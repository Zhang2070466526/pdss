# -*- coding: utf-8 -*-
# @Time    : 2023/6/6 11:29
# @Author  : zhuang
# @Site    : 
# @File    : sendClass.py
# @Software: PyCharm
import json
import os
import string
from ..models import *

from django.conf import settings
from django.core.mail import send_mail, send_mass_mail, EmailMessage
import pymssql
from datetime import date, timedelta, datetime

from openpyxl import Workbook
from openpyxl.styles import Side, Alignment, Font, Border
from openpyxl.utils import get_column_letter
from rest_framework.response import Response
from rest_framework import status
from pdss.settings import BASE_DIR
from ..serializers import *


def sql_connect():
    conn = pymssql.connect(host='172.16.6.197',
                           user='sa',
                           password='eHr123',
                           database='T9IMS_TG',
                           charset='utf8',
                           )
    return conn


def sql_connect2():
    conn = pymssql.connect(host='172.16.6.197',
                           user='sa',
                           password='eHr123',
                           database='T9IMS',
                           charset='utf8',
                           )
    return conn


def compute_week_start_end():
    today = date.today()
    start_of_last_week = today - timedelta(days=7)
    end_of_last_week = today - timedelta(days=1)
    return start_of_last_week, end_of_last_week


def get_alphabet(n):
    if 1 <= n <= 26:
        return string.ascii_uppercase[n - 1]
    elif n > 26:
        quotient, remainder = divmod(n - 1, 26)
        return get_alphabet(quotient) + string.ascii_uppercase[remainder]


# 保存xlsx文件
def save_xlsx(dir_name, title, row2, data_obj, save_filename):
    """
    :param dir_name: 保存文件目录，位于static/sendEmail/dir_name
    :param title:  excel的标题
    :param letter_list: ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
    :param row2: 第二行的数据
    :param data_obj: 保存的数据列表
    :param save_filename: 保存文件的名称
    :return: 返回文件地址
    """
    # letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
    letter_list = [get_alphabet(i) for i in range(1, len(row2) + 1)]
    side = Side(style="thick", color="000000")
    searchTime = datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(os.path.join(os.path.join(BASE_DIR, "static"), "sendEmail")):
        os.mkdir(os.path.join(os.path.join(BASE_DIR, "static"), "sendEmail"))
    if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "sendEmail"), dir_name)):
        os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "sendEmail"), dir_name))

    path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "sendEmail"), dir_name),
                        searchTime)
    if not os.path.exists(path):
        os.mkdir(path)
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(letter_list))
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
    ws.append(row2)
    for i in letter_list:
        ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
        ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
        ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
    for data in data_obj:
        ws.append(data)
    excel_columns = ws.columns

    len_list = count_excel_save(excel_columns)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
    wb.close()
    file_path = "static/sendEmail/" + dir_name + "/" + searchTime + '/' + save_filename + ".xlsx"
    wb.save(file_path)

    return file_path


# 用于设置xlsx的列宽
def count_excel_save(excel_columns):
    len_list = []
    prjTuple = tuple(excel_columns)
    for idx in range(0, len(prjTuple)):
        length2 = 0
        str_list = []
        for cell in prjTuple[idx]:
            if "1" not in str(cell):
                str_list.append(str(cell.value))
        for elem in str_list:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1.5
                else:
                    length += 2.5
            length2 = max(length, length2)
        len_list.append(length2)
    return len_list


# 发送邮件
def send_email(subject, message, recipient_list, file_list):
    """
    发送邮件
    :param subject: 主题
    :param message: 正文内容
    :param recipient_list: 接受者邮箱
    :param file_list: 附件列表
    :return:
    """
    email = EmailMessage(
        subject=subject,
        body=message,
        to=recipient_list,
    )
    for i in file_list:
        # print(i)
        email.attach_file(i)
    email.send()


# 保存发送邮件记录
def save_send_record(type_name, file_path, person_info, creator):
    person_list = [i["id"] for i in person_info]
    obj = SendRecord.objects.create(type_name=type_name, send_file_path=file_path, creator_id=creator)
    for i in person_list:
        obj.acceptor.add(i)


# 获取上下班数据
def get_sb_xb_data(start, end, deptID, ):
    sb_xb_data = []
    conn = sql_connect()
    if conn:
        cursor = conn.cursor()
        sql = """
        SELECT
            B.FullName,
            B.DepartmentName,
            A.Code,
            A.Name,
            D.PayTypeName,
            CONVERT ( DATE, C.WorkDate ) AS WorkDate,
            E.Name,
            C.b1sb,
            C.b1xb 
        FROM
            T_HR_Employee A
            left JOIN T_HR_Department B ON A.DeptID= B.id 
            left JOIN T_HR_WorkingTime C ON A.id= C.EmpID AND WorkDate BETWEEN %s AND %s
            left JOIN T_HR_PayType D ON A.PayTypeID= D.id 
            LEFT JOIN T_HR_ShiftsMst E ON E.MstID = C.ShiftID
        where A.EmployeeStatusID=1 AND A.DeptID in %s and A._gj='泰国'
        """
        cursor.execute(sql, (start, end, deptID))
        results = cursor.fetchall()
        sb_xb_data = []
        for i in results:
            sb_xb_data.append(i)
        conn.close()

    conn2 = sql_connect2()
    if conn2:
        cursor2 = conn2.cursor()
        sql = """
        SELECT
            B.FullName,
            B.DepartmentName,
            A.Code,
            A.Name,
            D.PayTypeName,
            CONVERT ( DATE, C.WorkDate ) AS WorkDate,
            E.Name,
            C.b1sb,
            C.b1xb 
        FROM
            T_HR_Employee A
            left JOIN T_HR_Department B ON A.DeptID= B.id 
            left JOIN T_HR_WorkingTime C ON A.id= C.EmpID AND WorkDate BETWEEN %s AND %s
            left JOIN T_HR_PayType D ON A.PayTypeID= D.id 
            LEFT JOIN T_HR_ShiftsMst E ON E.MstID = C.ShiftID
        where A.EmployeeStatusID=1 AND A.DeptID in %s and A._gj!='泰国'
        """
        cursor2.execute(sql, (start, end, deptID))
        results = cursor2.fetchall()
        for i in results:
            sb_xb_data.append(i)
        conn2.close()
    return sb_xb_data


# 获取离岗数据
def get_leave_data(start, end, deptID):
    leave_data = []
    conn = sql_connect2()
    if conn:
        cursor = conn.cursor()
        sql = """
        SELECT
            C.FUllName,
            C.DepartmentName,
            B.Code,
            B.Name,	
            A.WorkDate,
            A.OutTime,
            A.OutTimes,
            A.UseTime,
            D.Name 
        FROM
            Cust_MjDaily AS A
            LEFT JOIN T_HR_Employee B ON A.Code = B.Code
            LEFT JOIN T_HR_Department C ON A.DeptID = C.id
            LEFT JOIN T_HR_ShiftsMst D ON A.ShiftID= D.MstID 
        WHERE
             A.WorkDate BETWEEN %s AND %s AND A.DeptID in %s
        ORDER BY C.DepartmentName
        """
        cursor.execute(sql, (start, end, deptID))
        results = cursor.fetchall()
        leave_data = [list(i) for i in results]
        conn.close()
    return leave_data


# 获取所有部门
def get_dept():
    sql = """
        SELECT ID,FullName 
        FROM T_HR_Department
        WHERE IfUse=1
    """
    conn = sql_connect()
    cursor = conn.cursor()
    cursor.execute(sql)
    dept_list = []
    for i in cursor.fetchall():
        dept_list.append({
            "label": i[1],
            "key": i[0],
            'search': i[1],
        })
    conn.close()
    return dept_list


def get_person_email(name):
    person_email = []
    sql = """
        SELECT id,Name,email
        FROM T_HR_Employee
        WHERE Name LIKE '%{0}%' AND email is not null and email != ''
    """.format(name)
    conn = sql_connect()
    cursor = conn.cursor()
    cursor.execute(sql)
    for i in cursor.fetchall():
        person_email.append({
            "key": i[2],
            "label": i[1],
            "value": i[0]
        })
    cursor.close()
    conn.close()
    return person_email


def get_one_person_email(pk):
    sql = """
            SELECT id,Name,email
            FROM T_HR_Employee
            WHERE id={0}
        """.format(pk)
    conn = sql_connect()
    cursor = conn.cursor()
    cursor.execute(sql)
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    return result


def get_dept_fullname(name):
    dept_name = []
    sql = """
        SELECT ID,DepartmentName,FullName
        FROM T_HR_Department
        WHERE DepartmentName LIKE '%{0}%'
    """.format(name)
    conn = sql_connect()
    cursor = conn.cursor()
    cursor.execute(sql)
    for i in cursor.fetchall():
        dept_name.append({
            "key": i[2],
            "label": i[1],
            "value": i[0]
        })
    cursor.close()
    conn.close()
    return dept_name


# 获取单个部门
def get_one_dept(dept_id):
    sql = """
            SELECT ID,ParentID,DepartmentCode,DepartmentName 
            FROM T_HR_Department
            WHERE IfUse=1 AND ID=%s
        """
    conn = sql_connect()
    cursor = conn.cursor()
    cursor.execute(sql, dept_id)
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    return row


def judge_person_in_acceptorTable(pk, name, email, creator_id):
    obj = SendAcceptor.objects.filter(status=1, pk=pk)
    if obj.exists():
        return obj.values_list("id")[0][0]
    else:
        obj = SendAcceptor.objects.update_or_create(pk=pk, name=name, email_address=email, creator_id=creator_id)
        return obj[0].id


def judge_dept_in_deptTable(pk, ParentID, dept_code, dept_name, dept_type, creator_id):
    obj = SendDepartment.objects.filter(status=1, dept_id=pk, ParentID=ParentID, dept_code=dept_code,
                                        dept_name=dept_name, dept_type=dept_type)
    if obj.exists():
        return obj.values_list("id")[0][0]
    else:
        obj = SendDepartment.objects.update_or_create(dept_id=pk, ParentID=ParentID, dept_code=dept_code,
                                                      dept_name=dept_name,
                                                      dept_type=dept_type, creator_id=creator_id)
        return obj[0].id


class sendClass:
    def __init__(self, request, meth):
        self.request = request
        self.return_data = ''
        self.meth = meth
        self.token = ""
        self.methods = {
            "get_everyday_leave_data": self.get_everyday_leave_data,
            "get_everyWeek_attendance_data": self.get_everyWeek_attendance_data,
            "get_everyWeek_leave_data": self.get_everyWeek_leave_data,
            "get_attendance_list": self.get_attendance_list,
            "delete_one_dept_apt": self.delete_one_dept_apt,
            "get_dept": self.get_dept,
            "del_data": self.del_data,
            "patch_data": self.patch_data,
            "post_person": self.post_person,
            "test_send": self.test_send,
            "add_dept": self.add_dept,
        }

    def center_method(self):
        self.token = self.request.check_token
        self.methods[self.meth]()
        return Response(self.return_data)

    def token_center_method(self):
        self.token = self.request.check_token
        self.methods[self.meth]()
        return Response(self.return_data)

    # 新增部门
    def add_dept(self):
        info = json.loads(self.request.body)
        # try:
        for i in info["deptList"]:
            dept_info = get_one_dept(i)
            judge_dept_in_deptTable(dept_info[0], dept_info[1], dept_info[2], dept_info[3], info['type'],
                                    self.token)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "部门新增成功"
        }
        # except Exception as e:
        #     self.return_data = {
        #         "code": status.HTTP_400_BAD_REQUEST,
        #         "msg": "部门新增失败" + str(e)
        #     }

    def del_data(self):
        info = json.loads(self.request.body)
        for i in info['idList']:
            SendDepartment.objects.filter(pk=i).update(status=0)

    def patch_data(self):
        info = json.loads(self.request.body)
        for i in info['acceptorList']:
            accept_info = get_one_person_email(i)
            person_id = judge_person_in_acceptorTable(accept_info[0], accept_info[1], accept_info[2], self.token)
            obj = SendDepartment.objects.filter(pk=info['id'])
            if obj.exists():
                obj[0].dept_acceptor.add(person_id)
            else:
                pass
                # print("部门不存在")

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功"
        }

    def get_dept(self):
        dept_list = get_dept()
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "部门信息返回成功",
            "transfer_data": dept_list
        }

    def post_person(self):

        name = json.loads(self.request.body)['name']
        ac_type = json.loads(self.request.body)['type']
        self.return_data = {
            "code": status.HTTP_400_BAD_REQUEST,
        }
        if name != "":
            if ac_type == 1:
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "data": {
                        "personList": get_person_email(name)
                    }
                }
            else:
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "data": {
                        "deptList": get_dept_fullname(name)
                    }
                }

    # 获取每日离岗数据
    def get_everyday_leave_data(self):
        row2 = [
            '部门全称',
            '部门',
            '工号',
            '姓名',
            '日期',
            '离岗时长',
            '离岗次数',
            '在岗时长',
            '班制',
        ]
        today = date.today() - timedelta(1)
        dept_info = SendDepartment.objects.filter(status=1, dept_type=3)
        person_to_dept = {}
        person_info = {}
        for obj in dept_info:
            for k in obj.dept_acceptor.values("id", "name", "email_address"):
                email_address = k['email_address']
                if email_address in person_to_dept:
                    if obj.dept_id not in person_to_dept[email_address]:
                        person_to_dept[email_address].append(obj.dept_id)
                else:
                    person_to_dept[email_address] = [obj.dept_id]
                person_info[email_address] = k
        try:
            for email_address, dept_list in person_to_dept.items():
                title = person_info[email_address]['name'] + '-' + str(today) + "每日离岗"
                leave_data = get_leave_data(today, today, dept_list)

                file_path = save_xlsx("everyDayLeave", "每日离岗", row2, leave_data, title)
                send_email("离岗日报", "离岗日报", [email_address], [file_path])
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "邮件发送成功！"
                }
                save_send_record(title, file_path, [person_info[email_address]], self.token)
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_400_BAD_REQUEST,
                "msg": "邮件发送失败！" + str(e)
            }

    # 获取每周离岗数据
    def get_everyWeek_leave_data(self):
        row2 = [
            '部门全称',
            '部门',
            '工号',
            '姓名',
            '日期',
            '离岗时长',
            '离岗次数',
            '在岗时长',
            '班制',
        ]
        start, end = compute_week_start_end()
        dept_info = SendDepartment.objects.filter(status=1, dept_type=2)
        person_to_dept = {}
        person_info = {}
        for obj in dept_info:
            for k in obj.dept_acceptor.values("id", "name", "email_address"):
                email_address = k['email_address']
                if email_address in person_to_dept:
                    if obj.dept_id not in person_to_dept[email_address]:
                        person_to_dept[email_address].append(obj.dept_id)
                else:
                    person_to_dept[email_address] = [obj.dept_id]
                person_info[email_address] = k
        try:
            for email_address, dept_list in person_to_dept.items():
                title = person_info[email_address]['name'] + '-' + str(start) + "-" + str(end) + "每周离岗"
                leave_data = get_leave_data(start, end, dept_list)

                file_path = save_xlsx("everyWeekLeave", "每周离岗", row2, leave_data,
                                      title)
                send_email("离岗周报", "离岗周报", [email_address], [file_path])
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "邮件发送成功！"
                }
                save_send_record(title, file_path, [person_info[email_address]], self.token)
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_400_BAD_REQUEST,
                "msg": "邮件发送失败！" + str(e)
            }

    # 获取每周考勤数据
    def get_everyWeek_attendance_data(self):
        """
        这是发送每周考勤周报
        按照人来查找， 先找人，再找基地。通过sql语句去查询考勤记录。
        设想是 本周一，自动获取上周时间
        """
        row2 = [
            '部门全称',
            '部门名称',
            '工号',
            '姓名',
            '计薪方式',
            '日期',
            '班制',
            '上班时间',
            '下班时间',
        ]
        start, end = compute_week_start_end()
        dept_info = SendDepartment.objects.filter(status=1, dept_type=1)
        person_to_dept = {}
        person_info = {}
        for obj in dept_info:
            for k in obj.dept_acceptor.values("id", "name", "email_address"):
                email_address = k['email_address']
                if email_address in person_to_dept:
                    if obj.dept_id not in person_to_dept[email_address]:
                        person_to_dept[email_address].append(obj.dept_id)
                else:
                    person_to_dept[email_address] = [obj.dept_id]
                person_info[email_address] = k
        try:
            for email_address, dept_list in person_to_dept.items():
                title = person_info[email_address]['name'] + '-' + str(start) + "-" + str(end) + "每周考勤"
                sb_xb_data = get_sb_xb_data(start, end, dept_list)
                file_path = save_xlsx("everyWeekAttendance", "每周考勤", row2, sb_xb_data,
                                      title)
                send_email("考勤周报", "考勤周报", [email_address], [file_path])
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "邮件发送成功！"
                }
                save_send_record(title, file_path, [person_info[email_address]], self.token)
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_400_BAD_REQUEST,
                "msg": "邮件发送失败！" + str(e)
            }

    # 获取每日离岗数据
    def get_attendance_list(self):
        info = self.request.GET
        kwargs = {"status": 1, 'dept_type': info.get("type", 0)}
        if info['searchName'] != "":
            kwargs["dept_name__contains"] = info['searchName']
        pageSize = int(info['pageSize']) if info['pageSize'] != "" else 25
        currentPage = int(info['currentPage']) if info['currentPage'] != "" else 1
        columnList = [
            {
                "value": "序号",
                "label": "index",
                "width": 60
            }, {
                "value": "部门",
                "label": "dept_name",
                "width": 400
            }, {
                "value": "人数",
                "label": "dept_acceptor_num",
                "width": '60'
            }, {
                "value": "收件人",
                "label": "acceptor_name",
                "width": ''
            }
        ]
        obj = SendDepartment.objects.filter(**kwargs).order_by("ParentID")
        tableList = SendDepartmentSerializers(
            instance=obj[(currentPage - 1) * pageSize:currentPage * pageSize], many=True).data
        index = (currentPage - 1) * pageSize + 1
        for i in tableList:
            i['index'] = index
            index += 1
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                "columnList": columnList,
                "tableList": tableList,
                "totalNumber": obj.count(),
            }
        }

    def delete_one_dept_apt(self):
        info = json.loads(self.request.body)
        deptID = info['deptID']
        person_id = info['personID']
        obj = SendDepartment.objects.filter(pk=deptID)
        if obj.exists():
            obj[0].dept_acceptor.remove(person_id)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功！"
            }
        else:
            self.return_data = {
                "code": status.HTTP_400_BAD_REQUEST,
                "msg": "删除失败！"
            }

    def test_send(self):
        send_email("测试", "测试", ['1248493622@qq.com'], [])
