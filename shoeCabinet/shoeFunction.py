import datetime
import json
import os
import string
import time
import uuid

from openpyxl import Workbook
from openpyxl.styles import Side, Alignment, Font, Border

import arrow
import openpyxl
from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter

from pdss.settings import BASE_DIR
from utils.sqlServerConnect import EhrConnect

from django.http import JsonResponse
from django.db.models import Q
from django.db.models import ObjectDoesNotExist

from controller.controller import verify

from .models import *


class Shoe:
    def __init__(self, request):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = {
            'upload_info': self.upload_info,
            'get_list': self.get_list,
            'edit_person': self.edit_person,
            'delete_person': self.delete_person,
            'download_info': self.download_info,
            'create_single_info': self.create_single_info,
        }
        self.model_field = {
            'regin': 'regin',
            'wardrobeNumber': 'wardrobe_number',
            'deptName': 'dept_name',
            'code': 'person_code',
            'personName': 'person_name',
            'gender': 'gender',
            'employeeStatus': 'employee_status',
            'personPhone': 'person_phone',
            'productionProcesses': 'production_processes',
            'remark': 'remark',

        }

    def method_center(self, method):
        self.meth[method]()
        return JsonResponse(self.return_data)

    def upload_info(self):
        file = self.request.FILES.get('file')
        fileName = upload_file(file, "shoeCabinet", "upload", "")
        data = openpyxl.load_workbook(fileName)
        sheet = data.active
        value_list = []
        code_list = []
        for row in range(2, sheet.max_row + 1):
            if (sheet.cell(row, 5).value is None or sheet.cell(row, 5).value == '') and (
                    sheet.cell(row, 2).value is None or sheet.cell(row, 2).value == '') and \
                    sheet.cell(row, 3).value is None or sheet.cell(row, 3).value == '' and \
                    sheet.cell(row, 4).value is None or sheet.cell(row, 4).value == '':
                continue
            # for i in range(2, 9):
            #     if sheet.cell(row, i).value is None:
            #         sheet.cell(row, i).value = ''
            kwargs = {
                'regin': sheet.cell(row, 2).value if sheet.cell(row, 2).value is None else str(
                    sheet.cell(row, 2).value).strip(),
                'wardrobe_number': sheet.cell(row, 3).value if sheet.cell(row, 3).value is None else str(
                    sheet.cell(row, 3).value).strip(),
                'person_code': sheet.cell(row, 5).value if sheet.cell(row, 5).value is None else str(
                    sheet.cell(row, 5).value).strip(),
                'person_name': sheet.cell(row, 4).value if sheet.cell(row, 4).value is None else str(
                    sheet.cell(row, 4).value).strip(),
                'person_phone': sheet.cell(row, 6).value if sheet.cell(row, 6).value is None else str(
                    sheet.cell(row, 6).value).strip(),
                'production_processes': sheet.cell(row, 7).value if sheet.cell(row, 7).value is None else str(
                    sheet.cell(row, 7).value).strip(),
                'remark': sheet.cell(row, 8).value if sheet.cell(row, 8).value is None else str(
                    sheet.cell(row, 8).value).strip(),
            }
            if kwargs['person_code'] != "" and kwargs['person_code'] is not None and kwargs['person_code'] != 'None':
                code_list.append(kwargs['person_code'])
            value_list.append(kwargs)
        # 从数据库里获取详细的ehr人员信息
        result = get_personInfo_from_ehr(code_list)
        code_info = {}
        # 处理一下
        for i in result:
            code_info[i.pop('Code')] = i

        # flag = False
        for value in value_list:
            if type(value['person_code']) == str:
                value['person_code'] = value['person_code'].strip()
            for k, v in value.items():
                if type(v) == str:
                    value[k] = v.strip()
            # 如果有工号信息， 则根据 区域与衣柜号 去更新信息
            if value['person_code'] in code_info:
                value['dept_name'] = code_info[value['person_code']]['DepartmentName']
                if code_info[value['person_code']]['Sex'] != '':
                    value['gender'] = '男' if code_info[value['person_code']]['Sex'] == '1' else '女'
                else:
                    value['gender'] = '无信息'
                if code_info[value['person_code']]['EmployeeStatusID'] != '':
                    value['employee_status'] = '在职' if code_info[value['person_code']][
                                                           'EmployeeStatusID'] == '1' else '离职'
                else:
                    value['employee_status'] = '无信息'
                # try:
                #     # 如果人员已存在表中， 则报错，不新增
                #     ShoeRecord.objects.get(person_code=value['person_code'], status=True)
                #     flag = True
                #     self.return_data['msg'] += f"{value['person_code'],}"
                # except ObjectDoesNotExist:
                # 如果不存在，则根据区域与衣柜号进行创建或者更新
                # 如果柜子存在，怎更新，不存在则修改
            ShoeRecord.objects.update_or_create(defaults=value, regin=value['regin'],
                                                wardrobe_number=value['wardrobe_number'])
            # else:
            #     # 如果衣柜号存在，则更新
            #     try:
            #         ShoeRecord.objects.get(regin=value['regin'], wardrobe_number=value['wardrobe_number'], status=1)
            #         ShoeRecord.objects.update_or_create(defaults=value, regin=value['regin'],
            #                                             wardrobe_number=value['wardrobe_number'])
            # # 衣柜号不存在，则创建
            # except ObjectDoesNotExist:
            #     ShoeRecord.objects.create(**value)
        # if flag:
        #     self.return_data['msg'] += "该工号已存在衣柜号，请勿重复添加"
        #     self.return_data['code'] = 400
        # else:
        self.return_data["msg"] = "信息录入成功"
        self.return_data['code'] = 200

    # 获取信息列表
    def get_list(self):
        columnList = [
            {"value": "序号", "label": "index", "width": "60"},
        ]

        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        kwargs = {"status": True}
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        searchName = self.request.GET.get("searchName", None)
        regin = self.request.GET.get("regin", None)
        q_obj = Q()
        q_obj.connector = 'or'
        if searchName != '' and searchName is not None:
            q_obj.children.append(("person_code", searchName.strip()))
            q_obj.children.append(('person_name__contains', searchName.strip()))
            q_obj.children.append(('wardrobe_number', searchName.strip()))
        if regin != '' and regin is not None:
            kwargs['regin'] = regin.strip()

        obj = ShoeRecord.objects.filter(q_obj, **kwargs).values("id", "regin", "wardrobe_number", "dept_name",
                                                                "person_code",
                                                                "person_name", "gender", "employee_status",
                                                                "person_phone",
                                                                "production_processes", "remark").order_by(
            "-create_time")[
              (currentPage - 1) * pageSize: currentPage * pageSize]
        count = ShoeRecord.objects.filter(q_obj, **kwargs).count()
        tableList = []

        width_dict = {}
        # 输入初始的宽度
        for field in ShoeRecord._meta.get_fields():
            if field.name not in ['creator', 'modifier', 'status', 'create_time', 'modify_time', 'repair_person',
                                  "id"]:
                width_dict[field.name] = count_width(field.verbose_name)

        start = (int(currentPage) - 1) * int(pageSize) + 1
        for i in obj:
            for key, value in i.items():
                if key in width_dict:
                    width_dict[key] = max(count_width(str(value)), width_dict[key])
                else:
                    width_dict[key] = count_width(str(value))
            i['index'] = start
            start += 1
            tableList.append(i)
        for field in ShoeRecord._meta.get_fields():
            if field.name not in ['creator', 'modifier', 'status', 'create_time', 'modify_time', 'repair_person', "id"]:
                columnList.append(
                    {"value": field.verbose_name, "label": field.name, "width": width_dict[field.name]}
                )
        columnList[-1]['width'] = ''
        reginList = ShoeRecord.objects.all().values("regin").distinct()
        reginList = [{"value": i['regin'], "label": i['regin']} for i in reginList]
        self.return_data = {
            "code": 200,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': count,
                'reginList': reginList
            }
        }

    # 删除人员  仅仅支持idList
    def delete_person(self):
        info = json.loads(self.request.body)
        try:
            ShoeRecord.objects.filter(pk__in=info['personList']).update(status=False)
            self.return_data = {
                'code': 200,
                'msg': '删除成功'
            }
        except Exception as e:
            print(e)
            self.return_data = {
                'code': 400,
                'msg': '删除失败'
            }

    # 编辑人员
    def edit_person(self):
        info = json.loads(self.request.body)
        kwargs = {}
        person_id = info.pop("id")
        for field, value in info.items():
            if value != '' and type(value) == str:
                kwargs[self.model_field[field]] = value.strip()
        try:
            ShoeRecord.objects.filter(pk=person_id, status=1).update(**kwargs)
            self.return_data = {
                'code': 200,
                'msg': '修改成功'
            }
        except Exception as e:
            print(e)
            self.return_data = {
                'code': 400,
                'msg': '修改失败'
            }

    # 下载数据
    def download_info(self):
        info = json.loads(self.request.body)
        kwargs = {"status": True}
        searchName = info["searchName"]
        regin = info['regin']
        q_obj = Q()
        q_obj.connector = 'or'
        if searchName != '' and searchName is not None:
            q_obj.children.append(("person_code", searchName.strip()))
            q_obj.children.append(('person_name__contains', searchName.strip()))
        if regin != '' and regin is not None:
            kwargs['regin'] = regin.strip()

        if info['isAll'] == 0:
            q_obj = Q()
            kwargs = {
                "id__in": info['idList'],
                "status": True
            }
        try:
            obj = ShoeRecord.objects.filter(q_obj, **kwargs).values_list("regin", "wardrobe_number", "dept_name",
                                                                         "person_code",
                                                                         "person_name", "gender", "employee_status",
                                                                         "person_phone",
                                                                         "production_processes", "remark")
            if not obj.exists():
                self.return_data = {
                    'code': 200,
                    'msg': '数据为空，请重新选择'
                }
                return
        except Exception as e:
            print(e, "数据查询失败")
            self.return_data = {
                'code': 400,
                'msg': '下载失败'
            }
            return
        # try:
        url = save_xlsx('shoeCabinet', 'download', '衣柜使用情况信息表',
                        ['衣柜区域', '衣柜编号', '部门名称', '工号', '姓名', '性别', '在职状态', '手机号',
                         '工序', '备注'], obj,
                        arrow.now().format("YYYY_MM_DD") + '衣柜使用情况信息表')
        self.return_data = {
            'code': 200,
            'msg': '下载成功',
            "downloadUrl": url
        }
        # except Exception as e:
        #     print(e, "文件生成失败")
        #     self.return_data = {
        #         'code': 400,
        #         'msg': '下载失败'
        #     }

    # 新建单条信息
    def create_single_info(self):
        info = json.loads(self.request.body)
        kwargs = {}
        flag = False
        for key, value in info.items():
            kwargs[self.model_field[key]] = value
        regin_wardrobe_number = ShoeRecord.objects.filter(regin=kwargs['regin'],
                                                          wardrobe_number=kwargs['wardrobe_number'], status=True)

        if regin_wardrobe_number.exists():
            self.return_data = {
                'code': 200,
                'msg': '区域衣柜号已存在，无法新增'
            }
        # 如果衣柜号也不存在
        else:
            # 如果工号为空,则直接新增
            if kwargs['person_code'] == '' or kwargs['person_code'] is None:
                ShoeRecord.objects.create(**kwargs)
                self.return_data = {
                    'code': 200,
                    'msg': '区域衣柜号新增成功'
                }
            # 如果存在工号,则添加
            else:
                # 如果员工已存在
                # if ShoeRecord.objects.filter(person_code=kwargs['person_code'], status=True).exists():
                #     self.return_data = {
                #         'code': 200,
                #         'msg': '员工已有衣柜号，无法增加'
                #     }
                # # 如果员工不存在
                # else:
                result = get_personInfo_from_ehr([kwargs['person_code']])
                # 以下是找个人信息
                try:
                    kwargs['dept_name'] = result[0]['DepartmentName']
                    if result[0]['EmployeeStatusID'] != '':
                        kwargs['employee_status'] = "在职" if result[0]['EmployeeStatusID'] == '1' else "离职"
                    else:
                        kwargs['employee_status'] = "无信息"
                    kwargs['person_name'] = result[0]['Name']
                    if result[0]['Sex'] != '':
                        kwargs['gender'] = "男" if result[0]['Sex'] == '1' else "女"
                # 如果查找失败
                except Exception as e:
                    print(e)
                    flag = True
                try:
                    if flag:
                        self.return_data = {
                            'code': 200,
                            'msg': '新增失败,工号未找到个人信息，请检查工号'
                        }
                    else:
                        # 如果最终找到了个人信息,则新增成功
                        ShoeRecord.objects.create(**kwargs)
                        self.return_data = {
                            'code': 200,
                            'msg': '新增成功'
                        }
                # 如果保存数据库异常了，则报错
                except Exception as e:
                    print(e)
                    self.return_data = {
                        'code': 200,
                        'msg': '新增失败'
                    }

    # 获取报修列表
    def get_repair_list(self):
        kwargs = {
            "status": True
        }
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        searchName = self.request.GET.get("searchName", None)
        repairStatus = self.request.GET.get("repairStatus", None)


# 员工模块
class Employee:
    def __init__(self, request):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = {
            "get_my_info": self.get_my_info,
            "add_repair": self.add_repair,
        }
        self.model_field = {
            'code': "person_code",
            'description': "repair_content",
        }

    def method_center(self, method):
        self.meth[method]()
        return JsonResponse(self.return_data)

    def get_my_info(self):
        self.return_data = {
            'code': 200,
            'msg': '鞋衣柜查询成功'
        }
        code = self.request.GET.get("code").strip()
        if code == "" or code is None:
            self.return_data = {
                'code': 400,
                'msg': '工号不能为空！'
            }
            return

        try:
            shoe_obj = ShoeRecord.objects.filter(person_code=code, status=True)
            if not shoe_obj.exists():
                raise ObjectDoesNotExist
            repair_obj = RepairReport.objects.filter(person_code=code, status=True).order_by("-modify_time").values(
                "repair_content",
                "repair_status",
                "repair_suggestion",
                "repair_handle_person",
                "create_time",
                "modify_time",
                "pic_url",
            )
            repair_data = []
            for i in repair_obj:
                if i['repair_status'] == 1:
                    i['repair_status'] = '待处理'
                elif i['repair_status'] == 2:
                    i['repair_status'] = '处理中'
                else:
                    i['repair_status'] = '已处理'
                kwargs = {
                    'repair_content': i['repair_content'],
                    'repair_status': i['repair_status'],
                    'repair_suggestion': i['repair_suggestion'],
                    'repair_handle_person': i['repair_handle_person'],
                    'create_time': arrow.get(i['create_time']).format("YYYY-MM-DD HH:MM:SS"),
                    'modify_time': arrow.get(i['modify_time']).format("YYYY-MM-DD HH:MM:SS"),
                    'pic_url': i['pic_url'].split(','),
                }
                repair_data.append(kwargs)

            self.return_data = {
                'code': 200,
                'msg': '信息返回成功',
                'data': {
                    'regin_wardrobe_number': [
                        {'id': shoe.id, 'regin': shoe.regin, 'wardrobe_number': shoe.wardrobe_number} for shoe in
                        shoe_obj
                    ],
                    'repairList': [{"value": shoe.id,
                                    "text": "区域：" + shoe.regin + ' 衣柜号：' + shoe.wardrobe_number} for shoe in shoe_obj],
                    'repair': repair_data
                }
            }
        except ObjectDoesNotExist:
            self.return_data = {
                'code': 400,
                'msg': '未查询到您的鞋衣柜信息',
                'data': {
                    'regin_wardrobe_number': [{'id': 0, 'regin': "无柜号，请联系行政", 'wardrobe_number': "无柜号，请联系行政"}]
                    ,
                    'repairList': [],
                    'repair': [],
                }
            }

    def add_repair(self):
        # info = json.loads(self.request.body)
        code = self.request.POST.get("code")
        shoe_id = self.request.POST.get("shoe_id")
        if code == "" or code is None:
            self.return_data = {
                'code': 400,
                'msg': '请求错误'
            }
            return
        description = self.request.POST.get("description")
        file_list = self.request.FILES
        pic_path = []
        for key, value in file_list.items():
            pic_path.append(upload_file(value, 'shoeCabinet', 'upload', ''))
        pic_url = ''
        for pic in pic_path:
            pic_url += pic + ','
        try:
            RepairReport.objects.create(person_code=code, repair_content=description, pic_url=pic_url, shoe_id=shoe_id)
            self.return_data = {
                'code': 200,
                'msg': '报修成功'
            }
        except Exception as e:
            print(e)
            self.return_data = {
                'code': 400,
                'msg': '报修失败'
            }


# 计算宽度
def count_width(string):
    import re
    num_regex = re.compile(r'[0-9]')
    zimu_regex = re.compile(r'[a-zA-z]')
    hanzi_regex = re.compile(r'[\u4E00-\u9FA5]')
    num_list = num_regex.findall(string)
    zimu_list = zimu_regex.findall(string)
    hanzi_list = hanzi_regex.findall(string)
    return (len(num_list) + len(zimu_list)) * 15 + len(hanzi_list) * 25 + 10


# 从ehr表里获取详细信息
def get_personInfo_from_ehr(code_list):
    ehr = EhrConnect()
    if len(code_list) == 1:
        code_list = (code_list[0])
        sql = f"SELECT a.Code,a.Name, a.Sex, a.EmployeeStatusID, b.DepartmentName FROM T_HR_Employee AS a LEFT JOIN T_HR_Department AS b on a.DeptID=b.ID WHERE Code = '{code_list}'"
    else:
        code_list = tuple(code_list)
        sql = f"SELECT a.Code,a.Name, a.Sex, a.EmployeeStatusID, b.DepartmentName FROM T_HR_Employee AS a LEFT JOIN T_HR_Department AS b on a.DeptID=b.ID WHERE Code in {code_list}"
    result = ehr.select(sql)
    return result


def save_xlsx(root_dir, dir_name, title, row2, data_obj, save_filename):
    """
    :param root_dir static下面的一级目录
    :param dir_name: 保存文件目录，位于static/sendEmail/dir_name
    :param title:  excel的标题
    :param letter_list: ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
    :param row2: 第二行的数据
    :param data_obj: 保存的数据列表
    :param save_filename: 保存文件的名称
    :return: 返回文件地址
    """
    # letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
    letter_list = [get_alphabet(i) for i in range(1, len(row2) + 1)]
    side = Side(style="thick", color="000000")
    searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(os.path.join(os.path.join(BASE_DIR, "static"), root_dir)):
        os.mkdir(os.path.join(os.path.join(BASE_DIR, "static"), root_dir))
    if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name)):
        os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name))

    path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name),
                        searchTime)
    if not os.path.exists(path):
        os.mkdir(path)
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(letter_list))
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
    ws.append(row2)
    for i in letter_list:
        ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
        ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
        ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
    for data in data_obj:
        ws.append(data)
    excel_columns = ws.columns

    len_list = count_excel_save(excel_columns)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
    wb.close()
    file_path = "static/" + root_dir + '/' + dir_name + "/" + searchTime + '/' + save_filename + ".xlsx"
    wb.save(file_path)

    return file_path


def get_alphabet(n):
    if 1 <= n <= 26:
        return string.ascii_uppercase[n - 1]
    elif n > 26:
        quotient, remainder = divmod(n - 1, 26)
        return get_alphabet(quotient) + string.ascii_uppercase[remainder]


def count_excel_save(excel_columns):
    len_list = []
    prjTuple = tuple(excel_columns)
    for idx in range(0, len(prjTuple)):
        length2 = 0
        str_list = []
        for cell in prjTuple[idx]:
            if "1" not in str(cell):
                str_list.append(str(cell.value))
        for elem in str_list:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1.5
                else:
                    length += 2.5
            length2 = max(length, length2)
        len_list.append(length2)
    return len_list


def upload_file(file, model_file, method, fileName):
    searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
    flag = 1
    if fileName == "" or fileName is None:
        flag = 0
        fileName = searchTime + "_" + "".join(list(str(time.time()))[0:10]) + "_" + file.name
    try:
        if not os.path.exists(os.path.join(os.path.join(BASE_DIR, 'static'), model_file)):
            os.mkdir(os.path.join(os.path.join(BASE_DIR, 'static'), model_file))
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, 'static'), model_file), method)):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, 'static'), model_file), method))
        dirname = os.path.join(os.path.join(os.path.join(BASE_DIR, 'static'), model_file), method)
        if not os.path.exists(os.path.join(dirname, searchTime)):
            os.mkdir(os.path.join(dirname, searchTime))
        else:
            pass
        fp = open(os.path.join(os.path.join(dirname, searchTime), fileName), 'wb+')
        for chunk in file.chunks():
            fp.write(chunk)
        fp.close()
        if flag == 1:
            return fileName
        if flag == 0:
            return "static/" + model_file + '/' + method + '/' + fileName
    except:
        pass
    return 0
