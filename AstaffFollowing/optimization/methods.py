
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import status
import openpyxl,os,arrow,json
from pdss.settings import BASE_DIR
from employee.views import get_trees
from .models import *
from utils.genericMethods import methHeader, BasicClass,  FileClass
from staffFollowing.views import *
from utils.sqlServerConnect import EhrConnect
from employee.models import HrJobRank, HrDepartment
from datetime import datetime, timedelta
from collections import defaultdict
# from staffFollowing.optimization.views import generate_dates,get_template_file


def generate_dates(current_date):
    dates = []
    if current_date <= datetime(2024, 1, 31):
        dates.extend([date(2023, 11, 30), date(2023, 12, 31), date(2024, 1, 31), date(2024, 2, 29),date(2024, 3, 31)])
    elif datetime(2024, 2, 1) <= current_date <= datetime(2024, 3, 31):
        dates.extend([ date(2024, 1, 31), date(2024, 2, 29),date(2024, 3, 31),date(2024, 4, 30), date(2024, 5, 31)])
    elif datetime(2024, 4, 1) <= current_date <= datetime(2024, 5, 31):
        dates.extend([date(2024, 3, 31),date(2024, 4, 30), date(2024, 5, 31), date(2024, 6, 30), date(2024, 7, 31)])
    elif datetime(2024,6, 1) <= current_date <= datetime(2024, 7, 31):
        dates.extend([date(2024, 5, 31), date(2024, 6, 30), date(2024, 7, 31),date(2024,8, 31),date(2024, 9, 30),])
    elif datetime(2024,8, 1) <= current_date <= datetime(2024, 9, 30):
        dates.extend([date(2024, 7, 31),date(2024, 8, 31),date(2024, 9, 30),date(2024, 10, 31),date(2024, 11, 30),])
    return dates

def get_template_file(request):#获取上传模板
    current_date = datetime(2023, 12, 2)
    result_date = generate_dates(current_date)
    result_str_date = [dt.strftime('%Y/%m/%d') for dt in result_date]
    dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'optimize_template_file', "人员优化批量上传模板.xlsx")  # 创建文件
    # 加载工作簿和工作表
    wb = openpyxl.load_workbook(dummy_path)
    ws = wb.active
    ws['D2'].value,ws['E2'].value,ws['F2'].value,ws['G2'].value,ws['H2'].value= result_str_date
    ws['I2'].value, ws['J2'].value, ws['K2'].value, ws['L2'].value,ws['M2'].value,  = result_str_date
    wb.save(dummy_path)
    return JsonResponse({'url':dummy_path})



class BasicMobilize(methHeader, BasicClass):
    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete
        self.meth['options'] = self.options

        self.ehr = EhrConnect()


    def options(self):  #下拉框
        current_date = datetime.now()
        result = generate_dates(current_date)
        return_data = {
            'data': {
                'date_list': [
                    date.strftime('%Y/%m/%d')
                    # {'label': date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d')}
                    for date in result
                ],

            },
            "code": status.HTTP_200_OK,
            "msg": "下拉菜单返回成功",
            'hidden': True,
        }
        self.return_data=return_data


    def search(self):
        info = json.loads(self.request.body)
        print('info',info)
        current_page = int(info.get('currentPage',1))
        page_size = int(info.get('pageSize',25))

        # current_date = datetime.now()
        current_date = datetime(2023, 12, 2)
        result = generate_dates(current_date)
        print(current_date)
        print(result)
        column_list = [
            {'label': 'index', 'value': '序号', 'width': ''},
            {'label': 'optimize_dept__department_first_name', 'value': '一级部门', 'width': ''},
            {'label': 'optimize_dept__department_second_name', 'value': '二级部门', 'width': ''},
            {'value': '初始数据(2023/9/30)', 'children': [
                {'label': 'optimize_initial', 'value': '在职人数', 'width': 130},
            ]},
            # {'value': '在职人数预测', 'children': [
            #     {'label': 'month_forecast__'+date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d'), 'width': 130}
            #     for date in result
            #
            # ]},
            # {'value': '实际在职人数', 'children': [
            #     {'label': 'month_practical__'+date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d'), 'width': 130}
            #     for date in result
            # ]},
            {'value': '在职人数预测', 'children': [
                {'label': 'month_forecast__' + str(result.index(date)), 'value': date.strftime('%Y/%m/%d'),
                 'width': 130}
                for date in result

            ]},
            {'value': '实际在职人数', 'children': [
                {'label': 'month_practical__' + str(result.index(date)), 'value': date.strftime('%Y/%m/%d'),
                 'width': 130}
                for date in result
            ]},
        ]
        total_number=OptimizeTrace.objects.filter(optimize_status=True).count()
        trace_list = list(OptimizeTrace.objects.filter(optimize_status=True).values('id','optimize_dept','optimize_initial','optimize_dept__department_first_name','optimize_dept__department_second_name').order_by(
            '-optimize_create_time')[(current_page - 1) * page_size:current_page * page_size])
        print(trace_list)
        trace_all_id_list= [item['id'] for item in trace_list]
        month_list=list(OptimizeMonth.objects.filter(month_status=True,month_time__in=result,month_trace_id__in=trace_all_id_list).values('month_trace','month_time','month_forecast','month_practical'))
        transformed_month_list = []
        for entry in month_list:
            # transformed_entry = {
            #     'month_trace': entry['month_trace'],
            #     # 'month_time': entry['month_time'],
            #     f'month_forecast__{entry["month_time"].strftime("%Y/%m/%d")}': entry['month_forecast'],
            #     f'month_practical__{entry["month_time"].strftime("%Y/%m/%d")}': entry['month_practical']
            # }
            transformed_entry = {
                'month_trace': entry['month_trace'],
                # 'month_time': entry['month_time'],
                f'month_forecast__{str(result.index(entry["month_time"]))}': entry['month_forecast'],
                f'month_practical__{str(result.index(entry["month_time"]))}': entry['month_practical']
            }
            transformed_month_list.append(transformed_entry)
        """    
            from itertools import groupby
            from operator import itemgetter
            key_func = itemgetter('month_trace')
            sorted_data = sorted(transformed_month_list, key=key_func)
    
            merged_data = [
                {**{'month_trace': key}, **{k: v for d in group for k, v in d.items() if k != 'month_trace'}}
                for key, group in groupby(sorted_data, key=key_func)
            ]
            print('11111111111111')
            # Print the result
            for entry in merged_data:
                print(entry)
        """
        print('transformed_month_list',transformed_month_list)
        merged_data = defaultdict(dict)  #
        for entry in transformed_month_list:   #根据“month_trace”有效地对字典进行分组和合并
            trace_id = entry['month_trace']
            merged_data[trace_id] = {**merged_data[trace_id], **{k: v for k, v in entry.items() if k != 'month_trace'}}
        merged_list = [{'month_trace': key, **value} for key, value in merged_data.items()]
        print(merged_list)



        merged_data_dict = defaultdict(dict)
        for entry in trace_list + merged_list:
            key = entry.get('id') if 'id' in entry else entry.get('month_trace')
            merged_data_dict[key] = {**merged_data_dict[key], **entry}
        table_list = list(merged_data_dict.values())
        # print(table_list)
        for index, item in enumerate(table_list):
            item['index'] = (current_page - 1) * page_size + index + 1

        self.return_data['data']['columnList'] = column_list
        self.return_data['data']['tableList'] = table_list
        self.return_data['data']['totalNumber'] = total_number
    def create(self):
        info = json.loads(self.request.body)
        optimize_dept_id = info.pop('deptId', None)
        trace_obj,flag=OptimizeTrace.objects.update_or_create(defaults={'optimize_dept_id':optimize_dept_id,'optimize_status':True,'optimize_creator_id':self.operate_user_id},optimize_dept_id=optimize_dept_id,optimize_status=True)
        current_date = datetime(2023, 12, 2)
        result_date = generate_dates(current_date)
        print(result_date,type(result_date))
        print(info)

        '''
        {
            "deptId": "33",
            "optimize_forecast__1": "75",
            "optimize_forecast__2": "70",
            "optimize_forecast__3": "20",
            "optimize_forecast__4": "30",
            "optimize_forecast__5": "10",
            "optimize_practical__1": "14",
            "optimize_practical__2": "22",
            "optimize_practical__3": "71",
            "optimize_practical__4": "83",
            "optimize_practical__5": "69"
        }
        '''

        params_ls=[]
        for line in info.items():
            index=eval(line[0].split("__")[1])
            result={
                "month_time": result_date[index],
                line[0].split("__")[0]: eval(line[1])
            }
            params_ls.append(result)


        # for line in info.items():
        #     result = {"month_time": datetime.strptime(line[0].split("__")[1], '%Y/%m/%d').date(),
        #               line[0].split("__")[0]: eval(line[1])}
        #     params_ls.append(result)
        merged_dicts = defaultdict(dict)
        for d in params_ls:
            merged_dicts[d['month_time']].update(d)
            merged_dicts[d['month_time']].update({'month_trace_id':trace_obj.id,'month_status':True,'month_creator_id':self.operate_user_id})
        # 将defaultdict转换为普通字典
        merged_dicts = list(merged_dicts.values())
        # print(merged_dicts)
        for line in merged_dicts:
            month_obj=OptimizeMonth.objects.filter(month_trace_id=line['month_trace_id'], month_time=line['month_time'],month_status=1).first()
            if month_obj:  #对象存在  是否修改
                if month_obj.month_forecast is None and  month_obj.month_practical is None:
                    OptimizeMonth.objects.filter(pk=month_obj.id).update(month_forecast=line['month_forecast'],month_practical=line['month_practical'],month_modifier_id=self.operate_user_id)
                elif month_obj.month_forecast is None and  month_obj.month_practical is not None:
                    OptimizeMonth.objects.filter(pk=month_obj.id).update(month_forecast=line['month_forecast'],month_modifier_id=self.operate_user_id)
                elif month_obj.month_forecast is not None and month_obj.month_practical is None:
                    OptimizeMonth.objects.filter(pk=month_obj.id).update(month_practical=line['month_practical'],month_modifier_id=self.operate_user_id)
            else :#不存在 创建
                OptimizeMonth.objects.create(**line)



            # OptimizeMonth.objects.update_or_create(defaults=line,month_trace_id=line['month_trace_id'], month_time=line['month_time'],month_status=True,)
        self.return_data['msg'] = '新增成功'

    def update(self):
        info = json.loads(self.request.body)
        print(info)

    def delete(self):
        info = json.loads(self.request.body)
        OptimizeTrace.objects.filter(pk__in= info['idList']).update(
            optimize_status=False
        )
        self.return_data['msg'] = '删除成功'

    # @staticmethod
    # def generate_datess(current_date):
    #     dates = []
    #
    #     if current_date <= datetime(2024, 1, 31):
    #         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),
    #                       datetime(2024, 3, 31)])
    #     elif datetime(2024, 2, 1) <= current_date <= datetime(2024, 3, 31):
    #         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),

    #                       datetime(2024, 3, 31),
    #                       datetime(2024, 4, 30), datetime(2024, 5, 31)])
    #     elif datetime(2024, 4, 1) <= current_date <= datetime(2024, 5, 31):
    #         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),
    #                       datetime(2024, 3, 31),
    #                       datetime(2024, 4, 30), datetime(2024, 5, 31), datetime(2024, 6, 30), datetime(2024, 7, 31)])
    #     elif datetime(2024,6, 1) <= current_date <= datetime(2024, 7, 31):
    #         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),
    #                       datetime(2024, 3, 31),
    #                       datetime(2024, 4, 30), datetime(2024, 5, 31), datetime(2024, 6, 30), datetime(2024, 7, 31),datetime(2024,8, 31),datetime(2024, 9, 30),])
    #     elif datetime(2024,8, 1) <= current_date <= datetime(2024, 9, 30):
    #         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),
    #                       datetime(2024, 3, 31),
    #                       datetime(2024, 4, 30), datetime(2024, 5, 31), datetime(2024, 6, 30), datetime(2024, 7, 31),datetime(2024, 8, 31),datetime(2024, 9, 30),datetime(2024, 10, 31),datetime(2024, 11, 30),])
    #     return dates
    #     # return [date.strftime("%Y/%m/%d") for date in dates]
class MobilizeInfoFileClass(methHeader, FileClass):
    def __init__(self, request):
        super().__init__(request)
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.format('YYYY-MM-DD_HH_mm_ss')

    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download
    def upload(self):
        file = self.request.FILES.get('file',None)
        dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'upload_file', self.t1,'人员优化批量文件上传')  # 创建文件夹
        mkdir(dummy_path)
        file_url, file_name, file_suffix = createPath(file, '人员优化批量文件上传', 'talentDevelopFile','人员优化批量文件' + str(self.t2))
        saveFile(file_url, file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        current_date = datetime(2023, 12, 2)
        result_date = generate_dates(current_date)
        print(result_date)
        # # try:
        for i in range(2, sheet.max_row):  # 每行数据
            trace_params={}
            first_name = None if sheet.cell(i + 1, 2).value == '' else sheet.cell(i + 1, 2).value   #一级部门
            second_name = None if sheet.cell(i + 1, 3).value == '' else sheet.cell(i + 1, 3).value   #二级部门
            # print(first_name,second_name,third_name,forth_name)
            try:
                if first_name is not None and second_name is not None :  # 二级部门
                    trace_params['optimize_dept_id'] = HrDepartment.objects.filter(department_first_name=first_name,department_second_name=second_name).values_list('id', flat=True)[0]
                elif first_name is not None and second_name is None :  # 一级部门
                    trace_params['optimize_dept_id'] = HrDepartment.objects.filter(department_first_name=first_name).values_list('id', flat=True)[0]
                else:
                    trace_params['optimize_dept_id'] =None
            except:
                trace_params['optimize_dept_id'] = None
            # print(trace_params['optimize_dept_id'])
            if trace_params['optimize_dept_id'] is not None:
                trace_obj, flag = OptimizeTrace.objects.update_or_create(
                    defaults={'optimize_dept_id': trace_params['optimize_dept_id'], 'optimize_status': True,
                              'optimize_creator_id': self.operate_user_id}, optimize_dept_id=trace_params['optimize_dept_id'],optimize_status=True)
                print("trace_obj",trace_obj)
                month_params={}

                # for ceil in range(5):  #0,1,2,3,4
                #     value_ceil_forecast = sheet.cell(i + 1, ceil+4).value   #4,5,6,7,8 ,9 10  +4 +9   #预测在职
                #     value_ceil_practical= sheet.cell(i + 1, ceil+9).value     #实际在职
                #     if value_ceil_forecast:
                #


                for ceil in range(4,9):
                    value_ceil = sheet.cell(i + 1, ceil).value
                    if value_ceil is not None:
                        value_ceil_month_forecast = OptimizeMonth.objects.filter(month_trace=trace_obj.id,
                                                                               month_time=result_date[ceil-4],
                                                                               month_status=1).values('id',
                                                                                                      'month_forecast').first()  # .values_list('month_forecast',flat=True)
                        if value_ceil_month_forecast is None:  # 不存在创建
                            month_params = {
                                'month_forecast': value_ceil,
                                'month_time': result_date[ceil-4],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                            OptimizeMonth.objects.create(**month_params)
                        else:
                            if value_ceil_month_forecast['month_forecast'] is None:
                                OptimizeMonth.objects.filter(pk=value_ceil_month_forecast['id'], month_status=1).update(
                                    month_forecast=value_ceil)
                    else:
                        continue


                value_9 = sheet.cell(i + 1, 9).value
                if value_9 is not None:
                    value_9_month_practical=OptimizeMonth.objects.filter(month_trace=trace_obj.id,month_time=result_date[0],month_status=1).values('id','month_practical').first()#.values_list('month_forecast',flat=True)
                    if value_9_month_practical is None:  #不存在创建
                        month_params = {
                                'month_practical': value_9,
                                'month_time': result_date[0],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                        OptimizeMonth.objects.create(**month_params)
                    else:
                        if value_9_month_practical['month_practical'] is None:
                            print(value_9_month_practical['id'],value_9_month_practical['month_practical'])
                            OptimizeMonth.objects.filter(pk=value_9_month_practical['id'],month_status=1).update(month_practical=value_9)
                else:
                    continue
                value_10 = sheet.cell(i + 1, 10).value   #  只有在1,3,5,7,9,11才读取
                if self.now.month in [1,3,5,7,11]:#当前月份在列表中才读取
                    if value_10 is not None:
                        value_10_month_practical = OptimizeMonth.objects.filter(month_trace=trace_obj.id,
                                                                               month_time=result_date[1],
                                                                               month_status=1).values('id',
                                                                                                      'month_practical').first()  # .values_list('month_forecast',flat=True)
                        if value_10_month_practical is None:  # 不存在创建
                            month_params = {
                                'month_practical': value_10,
                                'month_time': result_date[0],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                            OptimizeMonth.objects.create(**month_params)
                        else:
                            if value_10_month_practical['month_practical'] is None:
                                OptimizeMonth.objects.filter(pk=value_10_month_practical['id'], month_status=1).update(
                                    month_practical=value_10)
                    else:
                        continue
                else:
                    pass
            else:
                continue

    def download(self):
        dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'download_file', self.t1,str(self.t2))  # 创建文件夹
        mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        download_all = json.loads(self.request.body).get('downloadAll')
        row_data = []
        current_date = datetime(2023, 12, 2)
        result_date = generate_dates(current_date)
        print(result_date)
        url=get_template_file()
        print(url)

        if download_all == True:  # 是下载全部   有条件
            pass
            # kwargs = {
            #
            # }
            # begin_date = to_date(json.loads(self.request.body).get('beginDate'))
            # end_date = to_date(json.loads(self.request.body).get('endDate'))
            # if begin_date is None or begin_date == '':
            #     begin_date = datetime(1901, 10, 29, 7, 17, 1, 177)
            # if end_date is None or end_date == '':
            #     end_date = datetime(3221, 10, 29, 7, 17, 1, 177)
            # record_list = list(ImmigrationRecords.objects.filter(records_status=1, records_create_time__gte=begin_date,
            #                                                      records_create_time__lte=end_date,
            #                                                      records_stationed_country_id=2).values('id',
            #                                                                                             'records_people__employee_code',
            #                                                                                             'records_people__employee_name',
            #                                                                                             'records_passport',
            #                                                                                             'records_stationed_country__country_name',
            #                                                                                             'records_stationed_base__base_name',
            #                                                                                             'records_people__employee_department__department_first_name',
            #                                                                                             'records_people__employee_department__department_second_name',
            #                                                                                             'records_people__employee_department__department_third_name',
            #                                                                                             'records_people__employee_join_date',
            #                                                                                             'records_begin_data',
            #                                                                                             'records_end_data',
            #                                                                                             'records_work_visa',
            #                                                                                             'records_local_bank',
            #                                                                                             'records_local_social_security',
            #                                                                                             'records_local_individual_taxes',
            #                                                                                             'records_leave_hour',
            #                                                                                             'records_absenteeism_hour'
            #                                                                                             ).order_by('-records_create_time'))
            #
            # record_id = [item['id'] for item in record_list]
            # index = 0
            # for line in record_list:
            #     line['index'] = index + 1
            #     if len(line) == 0:
            #         index = index
            #     index += 1
            #     line['records_work_visa']= '是' if line['records_work_visa'] == True else '否'
            #     line['records_local_bank'] = '是' if line['records_local_bank'] == True else '否'
            #     line['records_local_social_security'] = '是' if line['records_local_social_security'] == True else '否'
            #     line['records_local_individual_taxes'] = '是' if line['records_local_individual_taxes'] == True else '否'
            # print(record_list)
            # fill_list = list(ImmigrationFill.objects.filter(fill_status=True, fill_record_id__in=record_id,
            #                                                 # fill_create_time__lte=fill_end,
            #                                                 # fill_create_time__gte=fill_begin,
            #                                                 fill_approval_status=2).values('fill_record_id',
            #                                                                                'fill_inout_status__type_name',
            #                                                                                'fill_into_date',
            #                                                                                'fill_leave_date',
            #                                                                                'fill_trip_reason',
            #                                                                                'fill_leave_address',
            #                                                                                'fill_leave_days',
            #                                                                                'fill_remark'
            #                                                                                ))
            #
            # id_count = {item_a['id']: 1 for item_a in record_list}
            # table_list = []
            # for item_a in record_list:
            #     id_a = item_a['id']
            #     count = 0
            #     for item_b in fill_list:
            #         fill_record_id_b = item_b.get('fill_record_id')
            #         if id_a == fill_record_id_b:
            #             merged_item = item_a.copy()
            #             merged_item.update(item_b)
            #             table_list.append(merged_item)
            #             count += 1
            #     id_count[id_a] = max(count, 1)
            # for item_a in record_list:
            #     id_a = item_a['id']
            #     if not any('fill_record_id' in item and item['fill_record_id'] == id_a for item in table_list):
            #         table_list.append(item_a)  # Append unmatched items from 'record_list' to 'table_list'
            # count_list = [id_count[item['id']] for item in record_list]
            # table_list = self.sort_list_of_dicts(table_list)
            # table_list = sorted(table_list, key=lambda x: x['index'])
            # for line in table_list:
            #     # print(line)
            #     line_data = []
            #     for k, v in line.items():
            #         if k not in ('id', 'fill_record_id'):
            #             line_data.append(v)
            #     row_data.append(line_data)
            # exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            # sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            # for row in row_data:
            #     sheet.append(row)  # 在工作表中添加一行
            #
            # merge_rows_list = count_list
            # # 初始化起始行
            # start_row = 3
            #
            # # 循合并指定行数的A列到F列
            # for merge_rows in merge_rows_list:
            #     end_row = start_row + merge_rows - 1
            #     for col in range(1, 19):  # 合并A列到F列
            #         sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            #     start_row = end_row + 1
            #
            # # 保存修改后的Excel文件
            # exc.save(path)
            # # 关闭Excel文件
            # exc.close()
        else:
            pass
            # record_list = list(ImmigrationRecords.objects.filter(records_status=1,id__in=id_list,
            #                                                      records_stationed_country_id=2).values('id',
            #                                                                                             'records_people__employee_code',
            #                                                                                             'records_people__employee_name',
            #                                                                                             'records_passport',
            #                                                                                             'records_stationed_country__country_name',
            #                                                                                             'records_stationed_base__base_name',
            #                                                                                             'records_people__employee_department__department_first_name',
            #                                                                                             'records_people__employee_department__department_second_name',
            #                                                                                             'records_people__employee_department__department_third_name',
            #                                                                                             'records_people__employee_join_date',
            #                                                                                             'records_begin_data',
            #                                                                                             'records_end_data',
            #                                                                                             'records_work_visa',
            #                                                                                             'records_local_bank',
            #                                                                                             'records_local_social_security',
            #                                                                                             'records_local_individual_taxes',
            #                                                                                             'records_leave_hour',
            #                                                                                             'records_absenteeism_hour'
            #                                                                                             ).order_by('-records_create_time'))
            #
            # record_id = [item['id'] for item in record_list]
            # index = 0
            # for line in record_list:
            #     line['index'] = index + 1
            #     if len(line) == 0:
            #         index = index
            #     index += 1
            #     line['records_work_visa'] = '是' if line['records_work_visa'] == True else '否'
            #     line['records_local_bank'] = '是' if line['records_local_bank'] == True else '否'
            #     line['records_local_social_security'] = '是' if line['records_local_social_security'] == True else '否'
            #     line['records_local_individual_taxes'] = '是' if line['records_local_individual_taxes'] == True else '否'
            # print(record_list)
            # fill_list = list(ImmigrationFill.objects.filter(fill_status=True, fill_record_id__in=record_id,
            #                                                 # fill_create_time__lte=fill_end,
            #                                                 # fill_create_time__gte=fill_begin,
            #                                                 fill_approval_status=2).values('fill_record_id',
            #                                                                                'fill_inout_status__type_name',
            #                                                                                'fill_into_date',
            #                                                                                'fill_leave_date',
            #                                                                                'fill_trip_reason',
            #                                                                                'fill_leave_address',
            #                                                                                'fill_leave_days',
            #                                                                                'fill_remark'
            #                                                                                ))
            #
            # id_count = {item_a['id']: 1 for item_a in record_list}
            # table_list = []
            # for item_a in record_list:
            #     id_a = item_a['id']
            #     count = 0
            #     for item_b in fill_list:
            #         fill_record_id_b = item_b.get('fill_record_id')
            #         if id_a == fill_record_id_b:
            #             merged_item = item_a.copy()
            #             merged_item.update(item_b)
            #             table_list.append(merged_item)
            #             count += 1
            #     id_count[id_a] = max(count, 1)
            # for item_a in record_list:
            #     id_a = item_a['id']
            #     if not any('fill_record_id' in item and item['fill_record_id'] == id_a for item in table_list):
            #         table_list.append(item_a)  # Append unmatched items from 'record_list' to 'table_list'
            # count_list = [id_count[item['id']] for item in record_list]
            # table_list = self.sort_list_of_dicts(table_list)
            # table_list = sorted(table_list, key=lambda x: x['index'])
            # for line in table_list:
            #     # print(line)
            #     line_data = []
            #     for k, v in line.items():
            #         if k not in ('id', 'fill_record_id'):
            #             line_data.append(v)
            #     row_data.append(line_data)
            # exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            # sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            # for row in row_data:
            #     sheet.append(row)  # 在工作表中添加一行
            #
            # merge_rows_list = count_list
            # # 初始化起始行
            # start_row = 3
            #
            # # 循合并指定行数的A列到F列
            # for merge_rows in merge_rows_list:
            #     end_row = start_row + merge_rows - 1
            #     for col in range(1, 19):  # 合并A列到F列
            #         sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            #     start_row = end_row + 1
            #
            # # 保存修改后的Excel文件
            # exc.save(path)
            # # 关闭Excel文件
            # exc.close()

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            # "downloadUrl": path
        }






