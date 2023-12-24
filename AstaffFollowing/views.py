from django.db.models.query_utils import Q
from django.http.response import JsonResponse
from django.shortcuts import render

# Create your views here.
from datetime import datetime

from rest_framework import status
import os,arrow,openpyxl,time
from openpyxl.styles import Alignment

from datetime import datetime, date,timedelta
from employee.models import HrDepartment
from employee.views import get_trees


def get_department_option_two(request):
    return_data = {
        "code": 200,
        "msg": "下拉菜单返回成功",
        "data": [],
        'hidden': True
    }
    if request.check_token is not None:
        departments = HrDepartment.objects.filter(
            ~Q(id=999999),
            Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),
            department_status=1,department_level__lte=3
        ).values('id', 'department_name', 'department_parent_id')
        # from setup.userClass.userClass import get_trees
        hrbase_data = get_trees(departments, 'id', 'department_parent_id')
        base_ls = [1, ] + request.user_department_employee
        def filter_nodes(node):
            if 'children' in node:
                node['children'] = [child for child in node['children'] if child['id'] in base_ls]
                for child in node['children']:
                    filter_nodes(child)
        data = [node for node in hrbase_data if node['id'] in base_ls]
        for node in data:
            filter_nodes(node)
        def add_indexes(node_list, index=0):
            for node in node_list:
                node['index'] = index
                index += 1
                children = node.get('children', [])
                if children:
                    index = add_indexes(children, index)
            return index

        return_data['data'] = hrbase_data
        add_indexes(return_data['data'])
        base_ls = request.user_department_employee
    else:
        return_data = {
            "code": status.HTTP_403_FORBIDDEN,
            "msg": "没有权限访问",
            'hidden': False
        }
    return JsonResponse(return_data)




def mkdir(path):
    """
    判断path路径是否存在，不存再则创建
    :param path:
    :return:
    """
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)
    else:
        pass


def createPath(pic, path,fatherPath, fileName):
    """
    生成文件路径
    :param pic:    文件对象
    :param path:    生成的文件的父文件名称
    :param fileName:  生成的文件名称
    :param fatherPath:   static下面一级的文件名称
    :return:
    """
    now = arrow.now()
    t = now.format('YYYY-MM-DD')
    file_suffix = str(pic).split(".")[-1]  # 文件后缀

    file_name = f"{fileName}.{file_suffix}"  # 文件名称

    file_path = os.path.join('static',fatherPath , 'upload_file', t, path, file_name)  # 文件路径
    file_path = file_path.replace('\\', '/')
    return (file_path, file_name, file_suffix)  # 文件路径   文件名字  文件后缀



def saveFile(file_path, file_obj):
    """
    文件保存
    :param file_path:   文件路径
    :param file_obj:    文件对象
    :return:
    """
    with open(str(file_path), 'wb+') as f:
        for dot in file_obj.chunks():
            f.write(dot)

def createExcelPath(file_name,father_path, t2, title, num, interval, *file_ls):  # is not None
    """
    :param file_name: 要创建的excel的文件名称
    :param father_path:  static下面一级的路径
    :param t2:      excel的文件的上级目录
    :param num:  一个单元格的宽度
    :param title:    excel标题
    :param interval:   宽度范围
    :param file_ls:  表头字段

    :return:   excel文件路径
    """
    exc = openpyxl.Workbook()
    sheet = exc.active
    for column in sheet.iter_cols(min_col=0, max_col=num):
        for cell in column:
            sheet.column_dimensions[cell.column_letter].width = 30
    sheet.column_dimensions['A'].width = 10
    # sheet.column_dimensions['K'].width = 40

    sheet.title = file_name.split('.xlsx')[0]
    sheet.merge_cells(str(interval))  # 'A1:D1'

    sheet['A1'] = title
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.append(file_ls)
    t = time.strftime('%Y-%m-%d')   # 今天日期
    print(t)
    path = os.path.join('static', father_path, 'download_file', t, t2, file_name)
    path = path.replace(os.sep, '/')
    exc.save(path)
    return path

def is_date(value):
    """
    判断value是否为日期类型,或者字符串能转日期类型
    :param value:
    :return:
    """
    # 尝试将值转换为日期类型
    if isinstance(value, datetime):
        return True
    # 如果不是日期类型，检查是否是字符串
    if isinstance(value, str):
        try:
            datetime.strptime(value, '%Y-%m-%d')  # 根据日期格式修改这里
            return True
        except ValueError:
            return False
    return False



