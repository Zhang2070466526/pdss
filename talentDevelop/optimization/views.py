from utils.genericMethods import ViewBasicTemplate
from .methods import BasicMobilize, MobilizeInfoFileClass
from datetime import datetime, timedelta
import openpyxl,os
from pdss.settings import BASE_DIR
from django.http import JsonResponse
import pandas as pd
class basicDepartInfo(ViewBasicTemplate):
    def __init__(self):
        super().__init__(BasicMobilize)


class fileMobilizeInfo(ViewBasicTemplate):
    def __init__(self):
        super().__init__(MobilizeInfoFileClass)

# def get_template_file(request):#获取上传模板
#     current_date = datetime(2023, 12, 2)
#     result_date = generate_dates(current_date)
#     print(result_date)
#     result_str_date = [dt.strftime('%Y/%m/%d') for dt in result_date]
#     dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'optimize_template_file', "人员优化批量上传模板.xlsx")  # 创建文件
#     # 加载工作簿和工作表
#     wb = openpyxl.load_workbook(dummy_path)
#     ws = wb.active
#     ws['C2'].value,ws['D2'].value,ws['E2'].value,ws['F2'].value,ws['G2'].value, = result_str_date
#     ws['h2'].value, ws['i2'].value, ws['j2'].value, ws['k2'].value, ws['L2'].value, = result_str_date
#     wb.save(dummy_path)
#     return JsonResponse({'url':dummy_path})

# def generate_dates(current_date):
#     print(current_date)
#     current_date = datetime.now()
#     print(current_date)
#     dates = []
#     if current_date <= datetime(2024, 1, 31):
#         dates.extend([datetime(2023, 11, 30), datetime(2023, 12, 31), datetime(2024, 1, 31), datetime(2024, 2, 29),datetime(2024, 3, 31)])
#     elif datetime(2024, 2, 1) <= current_date <= datetime(2024, 3, 31):
#         dates.extend([ datetime(2024, 1, 31), datetime(2024, 2, 29),datetime(2024, 3, 31),datetime(2024, 4, 30), datetime(2024, 5, 31)])
#     elif datetime(2024, 4, 1) <= current_date <= datetime(2024, 5, 31):
#         dates.extend([datetime(2024, 3, 31),datetime(2024, 4, 30), datetime(2024, 5, 31), datetime(2024, 6, 30), datetime(2024, 7, 31)])
#     elif datetime(2024,6, 1) <= current_date <= datetime(2024, 7, 31):
#         dates.extend([datetime(2024, 5, 31), datetime(2024, 6, 30), datetime(2024, 7, 31),datetime(2024,8, 31),datetime(2024, 9, 30),])
#     elif datetime(2024,8, 1) <= current_date <= datetime(2024, 9, 30):
#         dates.extend([datetime(2024, 7, 31),datetime(2024, 8, 31),datetime(2024, 9, 30),datetime(2024, 10, 31),datetime(2024, 11, 30),])
#     return dates
#


