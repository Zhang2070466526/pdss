import shutil

from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import status
import openpyxl,os,arrow,json
from pdss.settings import BASE_DIR
from employee.views import get_trees
from .models import *
from utils.genericMethods import methHeader, BasicClass,  FileClass
from talentDevelop.views import *
from utils.sqlServerConnect import EhrConnect
from employee.models import HrJobRank, HrDepartment
from datetime import datetime, timedelta,date
from collections import defaultdict
# from staffFollowing.optimization.views import generate_dates,get_template_file
import calendar

def generate_dates(current_date):
    dates = []
    if current_date <= date(2024, 1, 31):
        dates.extend([date(2023, 11, 30), date(2023, 12, 31), date(2024, 1, 31), date(2024, 2, 29), date(2024, 3, 31)])
    elif date(2024, 2, 1) <= current_date <= date(2024, 3, 31):
        dates.extend([date(2024, 1, 31), date(2024, 2, 29), date(2024, 3, 31), date(2024, 4, 30), date(2024, 5, 31)])
    elif date(2024, 4, 1) <= current_date <= date(2024, 5, 31):
        dates.extend([date(2024, 3, 31), date(2024, 4, 30), date(2024, 5, 31), date(2024, 6, 30), date(2024, 7, 31)])
    elif date(2024, 6, 1) <= current_date <= date(2024, 7, 31):
        dates.extend([date(2024, 5, 31), date(2024, 6, 30), date(2024, 7, 31), date(2024, 8, 31), date(2024, 9, 30),])
    elif date(2024, 8, 1) <= current_date <= date(2024, 9, 30):
        dates.extend([date(2024, 7, 31), date(2024, 8, 31), date(2024, 9, 30), date(2024, 10, 31), date(2024, 11, 30),])
    elif date(2024, 10, 1) <= current_date <= date(2024, 11, 30):
        dates.extend([date(2024, 9, 30), date(2024, 10, 31), date(2024, 11, 30), date(2024, 12, 31), date(2025, 1, 31),])
    elif date(2024, 12, 1) <= current_date <= date(2025, 1, 31):
        dates.extend([date(2024, 11, 30), date(2024, 12, 31), date(2025, 1, 31), date(2025, 2, 28), date(2025, 3, 31),])
    elif date(2025, 2, 1) <= current_date <= date(2025, 3, 31):
        dates.extend([date(2025, 1, 31), date(2025, 2, 28), date(2025, 3, 31), date(2025, 4, 30), date(2025, 5, 31),])
    elif date(2025, 4, 1) <= current_date <= date(2025, 5, 31):
        dates.extend([date(2025, 3, 31), date(2025, 4, 30), date(2025, 5, 31), date(2025, 6, 30), date(2025, 7, 30),])
    elif date(2025, 6, 1) <= current_date <= date(2025, 7, 31):
        dates.extend([date(2025, 5, 31), date(2025, 6, 30), date(2025, 7, 30), date(2025, 8, 31), date(2025, 9, 30),])
    else:
        dates.extend([date(2025, 7, 30), date(2025, 8, 31), date(2025, 9, 30), date(2025,10, 31), date(2025, 11, 30), ])
    return dates

def get_template_file(request):#获取上传模板
    current_date = datetime.now().date()
    result_date = generate_dates(current_date)
    result_str_date = [dt.strftime('%Y/%m/%d') for dt in result_date]
    dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'optimize_template_file', "人员优化批量上传模板.xlsx")  # 创建文件
    # 加载工作簿和工作表
    wb = openpyxl.load_workbook(dummy_path)
    ws = wb.active
    ws['D2'].value,ws['E2'].value,ws['F2'].value,ws['G2'].value,ws['H2'].value= result_str_date
    ws['I2'].value, ws['J2'].value, ws['K2'].value, ws['L2'].value,ws['M2'].value,  = result_str_date
    wb.save(dummy_path)
    # return JsonResponse({'url':dummy_path})
    dummy_path = dummy_path.replace('\\', '/')
    dummy_path = 'static/' + dummy_path.split('static/')[1]
    response = JsonResponse({'url':dummy_path})
    return response


class BasicMobilize(methHeader, BasicClass):
    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete
        self.meth['options'] = self.options

        self.ehr = EhrConnect()


    def options(self):  #下拉框
        current_date = datetime.now()
        result = generate_dates(current_date)
        return_data = {
            'data': {
                'date_list': [
                    date.strftime('%Y/%m/%d')
                    # {'label': date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d')}
                    for date in result
                ],

            },
            "code": status.HTTP_200_OK,
            "msg": "下拉菜单返回成功",
            'hidden': True,
        }
        self.return_data=return_data


    def search(self):
        info = json.loads(self.request.body)
        current_page = int(info.get('currentPage',1))
        page_size = int(info.get('pageSize',25))
        employee_base = info.get('baseNameId','')
        month=info.get('month','')
        current_date = datetime.now().date() if month=='' or month is None or month == [] else datetime.strptime(month, '%Y-%m-%d').date()
        result = generate_dates(current_date)
        kwargs={}
        # # print(kwargs)
        if employee_base==[] or employee_base=='' or employee_base is None:
            kwargs['optimize_dept_id__in'] = self.request.user_department_employee
        else:
            kwargs['optimize_dept__in'] = employee_base

        column_list = [
            {'label': 'index', 'value': '序号', 'width': ''},
            {'label': 'optimize_dept__department_first_name', 'value': '一级部门', 'width': 200},
            {'label': 'optimize_dept__department_second_name', 'value': '二级部门', 'width': 200},
            {'value': '初始数据(2023/9/30)', 'children': [
                {'label': 'optimize_initial', 'value': '在职人数', 'width': 180},
            ]},
            # {'value': '在职人数预测', 'children': [
            #     {'label': 'month_forecast__'+date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d'), 'width': 130}
            #     for date in result
            #
            # ]},
            # {'value': '实际在职人数', 'children': [
            #     {'label': 'month_practical__'+date.strftime('%Y/%m/%d'), 'value': date.strftime('%Y/%m/%d'), 'width': 130}
            #     for date in result
            # ]},
            {'value': '在职人数预测', 'children': [
                {'label': 'month_forecast__' + str(result.index(date)), 'value': date.strftime('%Y/%m/%d'),
                 'width': 130}
                for date in result

            ]},
            {'value': '实际在职人数', 'children': [
                {'label': 'month_practical__' + str(result.index(date)), 'value': date.strftime('%Y/%m/%d'),
                 'width': 130}
                for date in result
            ]},
        ]

        total_number=OptimizeTrace.objects.filter(optimize_status=True,**kwargs).count()
        trace_list = list(OptimizeTrace.objects.filter(optimize_status=True,**kwargs).values('id','optimize_dept','optimize_initial','optimize_dept__department_first_name','optimize_dept__department_second_name').order_by(
            '-optimize_create_time')[(current_page - 1) * page_size:current_page * page_size])
        # print(trace_list)
        trace_all_id_list= [item['id'] for item in trace_list]
        month_list=list(OptimizeMonth.objects.filter(month_status=True,month_time__in=result,month_trace_id__in=trace_all_id_list).values('month_trace','month_time','month_forecast','month_practical'))
        transformed_month_list = []
        for entry in month_list:
            # transformed_entry = {
            #     'month_trace': entry['month_trace'],
            #     # 'month_time': entry['month_time'],
            #     f'month_forecast__{entry["month_time"].strftime("%Y/%m/%d")}': entry['month_forecast'],
            #     f'month_practical__{entry["month_time"].strftime("%Y/%m/%d")}': entry['month_practical']
            # }
            transformed_entry = {
                'month_trace': entry['month_trace'],
                # 'month_time': entry['month_time'],
                f'month_forecast__{str(result.index(entry["month_time"]))}': entry['month_forecast'],
                f'month_practical__{str(result.index(entry["month_time"]))}': entry['month_practical']
            }
            transformed_month_list.append(transformed_entry)
        """    
            from itertools import groupby
            from operator import itemgetter
            key_func = itemgetter('month_trace')
            sorted_data = sorted(transformed_month_list, key=key_func)
    
            merged_data = [
                {**{'month_trace': key}, **{k: v for d in group for k, v in d.items() if k != 'month_trace'}}
                for key, group in groupby(sorted_data, key=key_func)
            ]
            print('11111111111111')
            # Print the result
            for entry in merged_data:
                print(entry)
        """
        # print('transformed_month_list',transformed_month_list)
        merged_data = defaultdict(dict)  #
        for entry in transformed_month_list:   #根据“month_trace”有效地对字典进行分组和合并
            trace_id = entry['month_trace']
            merged_data[trace_id] = {**merged_data[trace_id], **{k: v for k, v in entry.items() if k != 'month_trace'}}
        merged_list = [{'month_trace': key, **value} for key, value in merged_data.items()]
        # print(merged_list)



        merged_data_dict = defaultdict(dict)
        for entry in trace_list + merged_list:
            key = entry.get('id') if 'id' in entry else entry.get('month_trace')
            merged_data_dict[key] = {**merged_data_dict[key], **entry}
        table_list = list(merged_data_dict.values())
        # print(table_list)
        for index, item in enumerate(table_list):
            item['index'] = (current_page - 1) * page_size + index + 1

        self.return_data['data']['columnList'] = column_list
        self.return_data['data']['tableList'] = table_list
        self.return_data['data']['totalNumber'] = total_number
    def create(self):
        info = json.loads(self.request.body)
        optimize_dept_id = info.pop('deptId', None)
        if optimize_dept_id is None or optimize_dept_id =='':
            self.return_data['msg'] = '部门未选择'
        else:
            trace_obj, flag = OptimizeTrace.objects.update_or_create(
                defaults={'optimize_dept_id': optimize_dept_id, 'optimize_status': True,
                          'optimize_creator_id': self.operate_user_id}, optimize_dept_id=optimize_dept_id,
                optimize_status=True)

            current_date = datetime.now().date()
            result_date = generate_dates(current_date)
            is_practical = info.get('type','1')   #是否是实际在职  0是预测在职 1是实际在职
            month = info.get('month','')
            num=info.get('num',None)
            month= datetime.now().date() if month=='' or month is None or month == [] else datetime.strptime(month, '%Y-%m-%d').date()
            if result_date[0]<=result_date[0]<=result_date[-1]:
                month_time=datetime(month.year, month.month, calendar.monthrange(month.year, month.month)[1]).date()
                month_obj = OptimizeMonth.objects.filter(month_trace_id=trace_obj.id,
                                                         month_time=month_time, month_status=1).first()

                if is_practical=='0':    #预测在职
                    if month_obj:
                        if month_obj.month_forecast is None:
                            OptimizeMonth.objects.filter(pk=month_obj.id).update(month_practical=num,month_modifier_id=self.operate_user_id)
                            self.return_data['msg'] = '新增成功'
                        else:
                            self.return_data['msg'] = '该部门该月份的预测在职人数已填写,无法再次填写'
                    else:
                        month_params={'month_trace_id': trace_obj.id, 'month_status': True,
                                      'month_creator_id': self.operate_user_id,
                                      'month_time':month_time,
                                      'month_forecast':num

                                      }

                        OptimizeMonth.objects.create(**month_params)
                        self.return_data['msg'] = '新增成功'
                elif is_practical=='1': #实际在职
                    if month_obj:
                        if month_obj.month_practical is None:
                            OptimizeMonth.objects.filter(pk=month_obj.id).update(month_practical=num,
                                                                                 month_modifier_id=self.operate_user_id)
                            self.return_data['msg'] = '新增成功'
                        else:
                            self.return_data['msg'] = '该部门该月份的实际在职人数已填写,无法再次填写'
                    else:
                        month_params = {'month_trace_id': trace_obj.id, 'month_status': True,
                                        'month_creator_id': self.operate_user_id,
                                        'month_time': month_time,
                                        'month_practical': num
                                        }
                        OptimizeMonth.objects.create(**month_params)
                        self.return_data['msg'] = '新增成功'

            else:
                self.return_data['msg'] = '所选日期不在时间范围内,请重新选择'

    def update(self):
        pass

    def delete(self):
        info = json.loads(self.request.body)
        OptimizeTrace.objects.filter(pk__in= info['idList']).update(
            optimize_status=False
        )
        self.return_data['msg'] = '删除成功'

class MobilizeInfoFileClass(methHeader, FileClass):
    def __init__(self, request):
        super().__init__(request)
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.format('YYYY-MM-DD_HH_mm_ss')

    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download
    def upload(self):
        file = self.request.FILES.get('file',None)
        dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'upload_file', self.t1,'人员优化批量文件上传')  # 创建文件夹
        mkdir(dummy_path)
        file_url, file_name, file_suffix = createPath(file, '人员优化批量文件上传', 'talentDevelopFile','人员优化批量文件' + str(self.t2))
        saveFile(file_url, file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        current_date =  datetime.now().date()
        result_date = generate_dates(current_date)

        # # try:
        for i in range(2, sheet.max_row):  # 每行数据
            trace_params={}
            first_name = None if sheet.cell(i + 1, 2).value == '' else sheet.cell(i + 1, 2).value   #一级部门
            second_name = None if sheet.cell(i + 1, 3).value == '' else sheet.cell(i + 1, 3).value   #二级部门
            try:
                if first_name is not None and second_name is not None :  # 二级部门

                    trace_params['optimize_dept_id'] = HrDepartment.objects.filter(department_first_name=first_name,department_second_name=second_name,department_name=second_name).values_list('id', flat=True)[0]
                elif first_name is not None and second_name is None :  # 一级部门

                    trace_params['optimize_dept_id'] = HrDepartment.objects.filter(department_first_name=first_name,department_name=first_name).values_list('id', flat=True)[0]
                else:
                    trace_params['optimize_dept_id'] =None
            except:
                trace_params['optimize_dept_id'] = None
            if trace_params['optimize_dept_id'] is not None:
                trace_obj, flag = OptimizeTrace.objects.update_or_create(
                    defaults={'optimize_dept_id': trace_params['optimize_dept_id'], 'optimize_status': True,
                              'optimize_creator_id': self.operate_user_id}, optimize_dept_id=trace_params['optimize_dept_id'],optimize_status=True)
                for ceil in range(4,9):
                    value_ceil = sheet.cell(i + 1, ceil).value
                    print(value_ceil)
                    if value_ceil is not None:
                        value_ceil_month_forecast = OptimizeMonth.objects.filter(month_trace=trace_obj.id,
                                                                               month_time=result_date[ceil-4],
                                                                               month_status=1).values('id',
                                                                                                      'month_forecast').first()  # .values_list('month_forecast',flat=True)
                        if value_ceil_month_forecast is None:  # 不存在创建
                            month_params = {
                                'month_forecast': value_ceil,
                                'month_time': result_date[ceil-4],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                            OptimizeMonth.objects.create(**month_params)
                        else:
                            if value_ceil_month_forecast['month_forecast'] is None:
                                OptimizeMonth.objects.filter(pk=value_ceil_month_forecast['id'], month_status=1).update(
                                    month_forecast=value_ceil)

                    else:
                        continue


                value_9 = sheet.cell(i + 1, 9).value
                if value_9 is not None:
                    value_9_month_practical=OptimizeMonth.objects.filter(month_trace=trace_obj.id,month_time=result_date[0],month_status=1).values('id','month_practical').first()#.values_list('month_forecast',flat=True)
                    if value_9_month_practical is None:  #不存在创建
                        month_params = {
                                'month_practical': value_9,
                                'month_time': result_date[0],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                        OptimizeMonth.objects.create(**month_params)
                    else:
                        if value_9_month_practical['month_practical'] is None:
                            print(value_9_month_practical['id'],value_9_month_practical['month_practical'])
                            OptimizeMonth.objects.filter(pk=value_9_month_practical['id'],month_status=1).update(month_practical=value_9)
                else:
                    continue
                value_10 = sheet.cell(i + 1, 10).value   #  只有在1,3,5,7,9,11才读取
                if self.now.month in [1,3,5,7,11]:#当前月份在列表中才读取
                    if value_10 is not None:
                        value_10_month_practical = OptimizeMonth.objects.filter(month_trace=trace_obj.id,
                                                                               month_time=result_date[1],
                                                                               month_status=1).values('id',
                                                                                                      'month_practical').first()  # .values_list('month_forecast',flat=True)
                        if value_10_month_practical is None:  # 不存在创建
                            month_params = {
                                'month_practical': value_10,
                                'month_time': result_date[0],
                                'month_trace_id': trace_obj.id,
                                'month_creator_id':self.operate_user_id
                            }
                            OptimizeMonth.objects.create(**month_params)
                        else:
                            if value_10_month_practical['month_practical'] is None:
                                OptimizeMonth.objects.filter(pk=value_10_month_practical['id'], month_status=1).update(
                                    month_practical=value_10)
                    else:
                        continue
                else:
                    pass
                self.return_data['msg'] = '上传成功'
            else:
                self.return_data['msg'] = '部门错误，新增失败'
                self.return_data['code'] = status.HTTP_401_UNAUTHORIZED
                continue

    def download(self):
        dummy_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'download_file', self.t1,str(self.t2))  # 创建文件夹
        mkdir(dummy_path)
        info = json.loads(self.request.body)
        id_list = json.loads(self.request.body).get('idList')
        download_all = json.loads(self.request.body).get('downloadAll')
        month = info.get('month', '')
        current_date = datetime.now().date() if month == '' or month is None or month == [] else datetime.strptime(
            month, '%Y-%m-%d').date()
        result_date = generate_dates(current_date)
        result_str_date = [dt.strftime('%Y/%m/%d') for dt in result_date]

        template_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'optimize_template_file','人员优化下载模板.xlsx')  # 创建文件
        # 加载工作簿和工作表
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws['E2'].value, ws['F2'].value, ws['G2'].value, ws['H2'].value,ws['I2'].value, = result_str_date
        ws['J2'].value, ws['K2'].value, ws['L2'].value, ws['M2'].value,ws['N2'].value, = result_str_date
        wb.save(template_path)


        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'talentDevelopFile', 'download_file', self.t1, str(self.t2),
                                        '编制及人员优化情况.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)


        if download_all == True:  # 是下载全部   有条件
            employee_base = info.get('baseNameId', '')


            kwargs = {}

            if employee_base == [] or employee_base == '' or employee_base is None:
                kwargs['optimize_dept_id__in'] = self.request.user_department_employee
            else:
                kwargs['optimize_dept__in'] = employee_base

            trace_list = list(OptimizeTrace.objects.filter(optimize_status=True, **kwargs).values('id', 'optimize_dept',
                                                                                                  'optimize_initial',
                                                                                                  'optimize_dept__department_first_name',
                                                                                                  'optimize_dept__department_second_name').order_by('-optimize_create_time'))

            trace_all_id_list = [item['id'] for item in trace_list]
            month_list = list(OptimizeMonth.objects.filter(month_status=True, month_time__in=result_date,
                                                           month_trace_id__in=trace_all_id_list).values('month_trace',
                                                                                                        'month_time',
                                                                                                        'month_forecast',
                                                                                                        'month_practical'))
            transformed_month_list = []
            for entry in month_list:

                transformed_entry = {
                    'month_trace': entry['month_trace'],
                    # 'month_time': entry['month_time'],
                    f'month_forecast__{str(result_date.index(entry["month_time"]))}': entry['month_forecast'],
                    f'month_practical__{str(result_date.index(entry["month_time"]))}': entry['month_practical']
                }
                transformed_month_list.append(transformed_entry)

            merged_data = defaultdict(dict)  #
            for entry in transformed_month_list:  # 根据“month_trace”有效地对字典进行分组和合并
                trace_id = entry['month_trace']
                merged_data[trace_id] = {**merged_data[trace_id],
                                         **{k: v for k, v in entry.items() if k != 'month_trace'}}
            merged_list = [{'month_trace': key, **value} for key, value in merged_data.items()]
            # print(merged_list)

            merged_data_dict = defaultdict(dict)
            for entry in trace_list + merged_list:
                key = entry.get('id') if 'id' in entry else entry.get('month_trace')
                merged_data_dict[key] = {**merged_data_dict[key], **entry}
            table_list = list(merged_data_dict.values())
            result_list = []
            count = 1
            for line in table_list:
                result_list.append([count,line.get('optimize_dept__department_first_name',None),
                                    line.get('optimize_dept__department_second_name',None),
                                    line.get('optimize_initial', None),
                                    line.get('month_forecast__0', None),
                                    line.get('month_forecast__1', None),
                                    line.get('month_forecast__2', None),
                                    line.get('month_forecast__3', None),
                                    line.get('month_forecast__4', None),
                                    line.get('month_practical__0', None),
                                    line.get('month_practical__1', None),
                                    line.get('month_practical__2', None),
                                    line.get('month_practical__3', None),
                                    line.get('month_practical__4', None),

                                   ])
                count += 1
            exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in result_list:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(destination_path)  # 指定路径,保存文件



        else:
            trace_list = list(OptimizeTrace.objects.filter(optimize_status=True, id__in=id_list).values('id', 'optimize_dept',
                                                                                                  'optimize_initial',
                                                                                                  'optimize_dept__department_first_name',
                                                                                                  'optimize_dept__department_second_name').order_by('-optimize_create_time'))
            trace_all_id_list = [item['id'] for item in trace_list]
            month_list = list(OptimizeMonth.objects.filter(month_status=True, month_time__in=result_date,
                                                           month_trace_id__in=trace_all_id_list).values('month_trace',
                                                                                                        'month_time',
                                                                                                        'month_forecast',
                                                                                                        'month_practical'))
            transformed_month_list = []
            for entry in month_list:
                transformed_entry = {
                    'month_trace': entry['month_trace'],
                    f'month_forecast__{str(result_date.index(entry["month_time"]))}': entry['month_forecast'],
                    f'month_practical__{str(result_date.index(entry["month_time"]))}': entry['month_practical']
                }
                transformed_month_list.append(transformed_entry)
            merged_data = defaultdict(dict)  #
            for entry in transformed_month_list:  # 根据“month_trace”有效地对字典进行分组和合并
                trace_id = entry['month_trace']
                merged_data[trace_id] = {**merged_data[trace_id],
                                         **{k: v for k, v in entry.items() if k != 'month_trace'}}
            merged_list = [{'month_trace': key, **value} for key, value in merged_data.items()]
            merged_data_dict = defaultdict(dict)
            for entry in trace_list + merged_list:
                key = entry.get('id') if 'id' in entry else entry.get('month_trace')
                merged_data_dict[key] = {**merged_data_dict[key], **entry}
            table_list = list(merged_data_dict.values())
            result_list = []
            count = 1
            for line in table_list:
                result_list.append([count,line.get('optimize_dept__department_first_name',None),
                                    line.get('optimize_dept__department_second_name',None),
                                    line.get('optimize_initial', None),
                                    line.get('month_forecast__0', None),
                                    line.get('month_forecast__1', None),
                                    line.get('month_forecast__2', None),
                                    line.get('month_forecast__3', None),
                                    line.get('month_forecast__4', None),
                                    line.get('month_practical__0', None),
                                    line.get('month_practical__1', None),
                                    line.get('month_practical__2', None),
                                    line.get('month_practical__3', None),
                                    line.get('month_practical__4', None),
                                   ])
                count += 1
            exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in result_list:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        self.return_data['msg'] = '下载成功'
        self.return_data['downloadUrl'] = destination_path




