import json, math, os, random, string, time, openpyxl

from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.generic.base import View
from rest_framework.status import HTTP_403_FORBIDDEN

from auther.models import AdminUser, UploadFiles
from django.core.paginator import Paginator
from django.db.models import Q
from memorabilia.models import *
from rest_framework.views import APIView
from rest_framework import serializers  # 序列化器
from rest_framework.generics import GenericAPIView
from datetime import datetime, date
from rest_framework.response import Response
from rest_framework import status
from memorabilia.serializers import * #序列化器
from general.models import center_base
from pdss.settings import BASE_DIR
class UploadFilesSerializers(serializers.ModelSerializer):
    class Meta:
        model = UploadFiles
        fields = '__all__'


class MemorabiliaListSerializers(serializers.ModelSerializer):
    # memorabilia_plans = serializers.SlugRelatedField(slug_field='file_url', read_only=True)
    #
    # memorabilia_photos=serializers.SerializerMethodField()
    # def get_memorabilia_photos(self, obj):
    #     

    class Meta:
        model = MemorabiliaList
        fields = '__all__'


class MemorabiliaListSerializersPart(serializers.ModelSerializer):  # 部分
    class Meta:
        model = MemorabiliaList
        # 指定这些字段，除了这些字段其他全部生成
        exclude = ['memorabilia_plans', 'memorabilia_photos']



class Reset:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "resetGet": self.resetGet,
            "resetPost": self.resetPost,
            'resetDelete': self.resetDelete,
            'resetPatch': self.resetPatch,
            'resetPut': self.resetPut,
        }

    def meth_center(self):
        # self.return_data = {'code': HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        # if self.request.check_token is None:
        #     return self.return_data
        self.methods[self.meth]()
        return Response(self.return_data)

    def resetGet(self):
        # print(self.request.GET)
        global serializer
        kwargs = {'memorabilia_status':True}
        memorabilia_base = self.request.GET.get('memorabilia_base_id', None)  # 中心/基地
        # memorabilia_base = self.request.GET.get('baseNameId', None)  # 中心/基地
        memorabilia_key_events_name = self.request.GET.get('memorabilia_key_events_name', None)  # 事件名称
        # memorabilia_date= self.request.GET.get('searchDate',None)  # 选择日期
        beginDate=self.request.GET.get('beginDate',None)
        endDate=self.request.GET.get('endDate',None)

        if memorabilia_key_events_name == ''  and memorabilia_base=='' and beginDate=="" and endDate=="":  # 全查
            kwargs['memorabilia_status'] = True
            kwargs['memorabilia_base__in']=self.request.user_base
        if memorabilia_base !='':
            kwargs['memorabilia_base'] = memorabilia_base
        if memorabilia_key_events_name != '':
            kwargs['memorabilia_key_events_name'] = memorabilia_key_events_name
        if beginDate != "" and endDate != "":
            kwargs['memorabilia_date__gte'] = datetime(2001, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
            kwargs['memorabilia_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate

        currentPage = eval(self.request.GET.get('currentPage'))  # 当前页码
        pageSize = eval(self.request.GET.get('pageSize'))  # 页面总条数
        record_count = MemorabiliaList.objects.all().count()
        if math.ceil(record_count / pageSize) < currentPage:
            currentPage = 1
        else:
            currentPage = currentPage

        print(memorabilia_base)
        kwargs['memorabilia_base__in']=self.request.user_base
        
        queryset = MemorabiliaList.objects.filter(**kwargs).order_by('-create_time')


        try:
            serializer = MemorabiliaListSerializersForge(instance=queryset[
                   (currentPage - 1) * pageSize:currentPage * pageSize], many=True)
        except:
            pass
            


        columnList = [{"value": "序号", "label": "index", "width": "60"}, ]   #表头数据的生成
        for field in MemorabiliaList._meta.get_fields():  # 获取该类内所有字段对象

            except_field = ["create_time", "modify_time", "creator", "modifier", "memorabilia_status"]
            if field.name not in except_field:
                if field.verbose_name == "公司":
                    columnList.append({
                        "value": "中心/事业部",
                        "label": "base_father",
                        "width": "180",
                    })
                field_label = {
                    "value": field.verbose_name,
                    "label": field.name,
                    "width": self.count_character(field.verbose_name),
                }
                columnList.append(field_label)

        tableList = []
        count_len = {}

        for i in serializer.data:
            # 
            for key, value in dict(i).items():
                count_len[key] = self.count_character(value)
            obj=MemorabiliaList.objects.filter(id=int(i['id'])).first()
            # 
            # i['memorabilia_photos_file'] = obj.memorabilia_photos.filter(status=True).values_list('file_url',flat=True)
            i['memorabilia_photos_file']=obj.memorabilia_photos.filter(status=True).values('id','file_url','file_name')
            i['memorabilia_plans_file'] = obj.memorabilia_plans.filter(status=True).values('id', 'file_url','file_name')
            for j in i['memorabilia_photos_file']:
                j['name'] = j['file_name']
                j['url']=j['file_url']
                del j['file_url']
                del j['file_name']
            for k in i['memorabilia_plans_file']:
                k['name'] = k['file_name']
                k['url']=k['file_url']
                del k['file_url']
                del k['file_name']
            # 
            k={
                'name':i['memorabilia_base'],
                'status':True,
            }
            # i['memorabilia_base_name']=list(center_base.objects.filter(**k).values_list('name',flat=True))[0]

            # i['memorabilia_base_id'] =center_base.objects.filter(**k).values_list('id',flat=True)
            try:
                id=list(center_base.objects.filter(**k).values_list('id',flat=True))[0]
            except:
                id=None
            i['memorabilia_base_id']=id

            tableList.append(dict(i))
        # 

        for i, item in enumerate(tableList):
            item["index"] = pageSize * (currentPage - 1) + i + 1
        else:
            pass
        del columnList[1]

        if len(count_len) > 0:
            count_len['index'] = "60"
            for i in columnList:
                i['width'] = max(int(i['width']), int(count_len[i['label'].lower()]))
        for i in columnList:
            i['width'] = 400 if int(i['width']) > 400 else int(i['width'])
            if i['label']=='memorabilia_remark':
                i['width']=''

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,  # 表头
                'tableList': tableList,
                'totalNumber': queryset.count()
            }
        }

    def resetPost(self):
        
        
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'memorabiliaFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)

        if self.request.POST.get('createData',None)!=None or self.request.FILES.get('memorabilia_photos', None)!=None or self.request.FILES.get('memorabilia_photos', None)!=None:    #新增数据
            
            memorabilia_photos = self.request.FILES.getlist('memorabilia_photos', None)  # 照片
            memorabilia_plans = self.request.FILES.getlist('memorabilia_plans', None)  # 方案
            try:
                createData=json.loads(self.request.POST.get('createData',None))
                
                data = {
                    "memorabilia_base_id": createData.get('memorabilia_base_id', None),
                    "memorabilia_date": createData.get('memorabilia_date', None),
                    'memorabilia_key_events_name': createData.get('memorabilia_key_events_name', None),
                    'memorabilia_main_attending_leaders': createData.get('memorabilia_main_attending_leaders', None),
                    'memorabilia_location': createData.get('memorabilia_location', None),
                    'memorabilia_remark': createData.get('memorabilia_remark', None),
                }
                if data['memorabilia_date']=='' or 'memorabilia_date'==None:
                    data['memorabilia_date']=date.today()
                
                try:
                    data['creator_id']=self.request.check_token
                except:
                    data['creator_id'] =None
                memo_obj = MemorabiliaList.objects.create(**data)
                
                for img_obj in memorabilia_photos:
                    img_path, img_name, img_suffix = self.createPath(img_obj,self.create_file_name(memo_obj.memorabilia_remark,
                                                                                             img_obj))

                    self.saveFile(img_name, img_obj, img_suffix)  # 保存文件
                    img_kwargs = {
                        "file_url": img_path,
                        "file_name": img_name,
                    }
                    up_obj = UploadFiles.objects.update_or_create(**img_kwargs)[0]  # <UploadFiles: UploadFiles object (8)>
                    memo_obj.memorabilia_photos.add(up_obj.id)

                for file_obj in memorabilia_plans:
                    file_path, file_name, file_suffix = self.createPath(file_obj,self.create_file_name(memo_obj.memorabilia_remark,file_obj))
                    self.saveFile(file_name, file_obj, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_path,
                        "file_name": file_name,
                    }
                    up_obj = UploadFiles.objects.update_or_create(**file_kwargs)[0]  # <UploadFiles: UploadFiles object (8)>
                    memo_obj.memorabilia_plans.add(up_obj.id)
                
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "信息添加成功",
                }
            except:
                self.return_data = {
                    "code": 400,
                    "msg": "您需要新增一条数据,不仅仅是图片或方案",
                }





        if self.request.FILES.getlist('file'):   #上传文件
            
            for file_obj in self.request.FILES.getlist('file'):  # 上传文件
                file_path, file_name, file_suffix = self.createPath(file_obj, self.create_file_name('up', file_obj))
                # self.saveFile(file_name, file_obj, file_suffix)  # 保存文件
                file=self.request.FILES.get("file", None)
                self.saveFile(file_name, file, file_suffix)  # 保存文件
                exc = openpyxl.load_workbook(file_path)  # 打开整个excel文件    static/memorabiliaFile/uploadFile/2023-05-06/润阳大事记模板.xlsx
                table = exc.active
                for i in range(2, table.max_row):
                    # try:
                    #     # memorabilia_base_name = table.cell(i + 1, 2).value  #基地名
                    #     k = {
                    #         'name': table.cell(i + 1, 2).value,
                    #         'status': True,
                    #     }
                    #     memorabilia_base_id=center_base.objects.filter(**k).values_list('id')[0][0]
                    # except:
                    #     memorabilia_base_id=None
                    try:
                        base_father = center_base.objects.get(status=1, name=table.cell(i + 1, 2).value)
                    except ObjectDoesNotExist:
                        self.return_data = {
                            "code": status.HTTP_400_BAD_REQUEST,
                            "msg": "某行基地名不存在，无法添加"
                        }
                        continue
                    obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                    # print(obj.exists())
                    if obj.exists():
                        try:
                            memorabilia_base_id = center_base.objects.get(name=table.cell(i + 1, 3).value,
                                                                               base_parent_id=base_father.id,
                                                                               status=True).id  # 二级基地id
                        except:
                            # talent_subsidies_base_id =None
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg": "基地与公司关系不符"
                            }
                            continue
                    else:
                        memorabilia_base_id = base_father.id
                    if memorabilia_base_id not in self.request.user_base:
                        self.return_data = {
                            "code": status.HTTP_403_FORBIDDEN,
                            "msg": "抱歉，您没有上传 " + table.cell(i + 1, 2).value + "-" + table.cell(i + 1,
                                                                                                      3).value if table.cell(
                                i + 1, 3).value else '' + " (基地/中心/公司)的权限"
                        }
                    else:
                        if table.cell(i + 1, 4).value != None:
                            try:
                                memorabilia_date= str(table.cell(i + 1, 4).value.date())
                            except:  # 可能不是日期
                                memorabilia_date= None
                        else:
                            memorabilia_date= None


                        data_kwargs = {
                            'memorabilia_base_id': memorabilia_base_id,
                            'memorabilia_date':memorabilia_date,
                            'memorabilia_key_events_name': table.cell(i + 1, 5).value,
                            'memorabilia_main_attending_leaders': table.cell(i + 1,6).value,
                            'memorabilia_location': table.cell(i + 1, 7).value,
                            'memorabilia_remark': table.cell(i + 1, 8).value
                        }
                        try:
                            data_kwargs['creator_id'] = self.request.check_token
                        except:
                            data_kwargs['creator_id'] = None
                        if data_kwargs['memorabilia_base_id']==None and data_kwargs['memorabilia_date']==None and data_kwargs['memorabilia_key_events_name']==None and data_kwargs['memorabilia_main_attending_leaders']==None and data_kwargs['memorabilia_location']==None and data_kwargs['memorabilia_remark']==None:
                            continue
                        else:
                            MemorabiliaList.objects.create(**data_kwargs)
            # for file_obj in self.request.FILES.getlist('file'):  # 上传文件
            #     file_path, file_name, file_suffix = self.createPath(file_obj)
            #     self.saveFile(file_name, file_obj, file_suffix)  # 保存文件
            #     exc = openpyxl.load_workbook(file_path)  # 打开整个excel文件    static/memorabiliaFile/uploadFile/2023-05-06/润阳大事记模板.xlsx
            #     sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            #     rowRange = sheet[3:sheet.max_row]  # 从第2行开始
            #     # 
            #     rows_fake=[]
            #
            #     for row in rowRange:  # 按行遍历
            #         
            #         rows=[]
            #         try:
            #             for cell in row:
            #                 rows.append(cell.value)
            #             if any(rows[1:]):
            #                 data_row = {
            #                     'memorabilia_base_id': rows[1:][0],
            #                     'memorabilia_date': str(rows[1:][1].date()) if rows[1:][1] != None else None,
            #                     'memorabilia_key_events_name': rows[1:][2],
            #                     'memorabilia_main_attending_leaders': rows[1:][3],
            #                     'memorabilia_location': rows[1:][4],
            #                     'memorabilia_remark': rows[1:][5]
            #                 }
            #
            #                 MemorabiliaList.objects.create(**data_row)
            #         except:
            #             rows_fake.append(row)
            #             # 
            #             
            #     
            #     try:
            #         
            #         data2 = {
            #             # 'memorabilia_base_id': rows_fake[1].value,
            #             # 'memorabilia_date': str(rows_fake[2].value.date())if rows_fake[2].value==None else None,
            #             # 'memorabilia_key_events_name': rows_fake[3].value,
            #             # 'memorabilia_main_attending_leaders': rows_fake[4].value,
            #             # 'memorabilia_location': rows_fake[5].value,
            #             'memorabilia_remark': rows_fake[6].value
            #         }
            #         
            #
            #         MemorabiliaList.objects.create(**data2)
            #     except:
            #         
            #
            #
            #
            #
            #

            
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息添加成功",
            }


    def resetPatch(self):
        
        # 
        # 
        # memorabilia_photos = self.request.FILES.getlist('memorabilia_photos')
        # memorabilia_plans = self.request.FILES.getlist('memorabilia_plans')

        # {'id': 92, 'memorabilia_base': 1, 'memorabilia_date': '2023-05-09', 'memorabilia_key_events_name': '6543',
        #  'memorabilia_main_attending_leaders': '', 'memorabilia_remark': '66', 'memorabilia_location': '666',
        #  'modifier': 2}

        data = {
            'id': self.request.data.get('id'),
            'memorabilia_base': self.request.data.get('memorabilia_base_id'),
            'memorabilia_date': self.request.data.get('memorabilia_date'),
            'memorabilia_key_events_name': self.request.data.get('memorabilia_key_events_name'),
            'memorabilia_main_attending_leaders': self.request.data.get('memorabilia_main_attending_leaders'),
            'memorabilia_remark': self.request.data.get('memorabilia_remark'),
            'memorabilia_location': self.request.data.get('memorabilia_location'),
            # 'memorabilia_plans': self.request.FILES.getlist('memorabilia_plans'),
            # 'memorabilia_photos': self.request.FILES.getlist('memorabilia_photos'),
            'modifier':self.request.check_token
        }
        

        # query_data = MemorabiliaList.objects.get(id=data['id'])
        MemorabiliaList.objects.filter(id=data['id']).update(**data)

        # for img_obj in memorabilia_photos:
        #     img_path, img_name, img_suffix = self.createPath(img_obj)
        #     self.saveFile(img_name, img_obj, img_suffix)  # 保存文件
        #     img_kwargs = {
        #         "file_url": img_path,
        #         "file_name": img_name,
        #     }
        #     up_obj = UploadFiles.objects.update_or_create(**img_kwargs)[0]  # <UploadFiles: UploadFiles object (8)>
        #     query_data.memorabilia_photos.add(up_obj.id)
        #
        # for file_obj in memorabilia_plans:
        #     file_path, file_name, file_suffix = self.createPath(file_obj)
        #     self.saveFile(file_name, file_obj, file_suffix)  # 保存文件
        #     file_kwargs = {
        #         "file_url": file_path,
        #         "file_name": file_name,
        #     }
        #     up_obj = UploadFiles.objects.update_or_create(**file_kwargs)[0]  # <UploadFiles: UploadFiles object (8)>
        #     query_data.memorabilia_plans.add(up_obj.id)

        # serializer = MemorabiliaListSerializersForge(instance=query_data, data=data)  # 构建序列化器对象
        # if serializer.is_valid():
        #     serializer.save()
        #     
        #     self.return_data = {
        #         "code": status.HTTP_200_OK,
        #         "msg": "信息修改成功",
        #     }
        # else:
        #     self.return_data = {
        #         "code": status.HTTP_401_UNAUTHORIZED,
        #         "msg": "信息修改失败！",
        #         'error': serializer.errors
        #     }
        #     
        self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息修改成功",
        }
    def resetDelete(self):
        
        
        # 

        idList=self.request.data.get('idList')
        # 
        for id in idList:
            # 
            MemorabiliaList.objects.filter(id=int(id)).update(memorabilia_status=False,modifier_id=self.request.check_token)



        # if 'True' in  self.request.data.get('idList'):  #True 删UploadFiles表的文件
        #     id= self.request.data.get('idList')[0]   #UploadFiles表的id
        #     UploadFiles.objects.filter(id=int(id)).update(status=False)
        # else:#删除这行数据
        #     idList = self.request.data.get('idList')
        #     for id in idList:
        #         
        #         obj=MemorabiliaList.objects.filter(id=int(id)).first()
        #         obj.memorabilia_photos.update(status=False)
        #         obj.memorabilia_plans.update(status=False)
        #         # 
        #         # # 
        #
        #         MemorabiliaList.objects.filter(id=int(id)).update(memorabilia_status=False)


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息删除成功",
        }

    def resetPut(self):
        
        
        id_list=self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        # 
        # 
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'memorabiliaFile', 'downloadFile', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path=self.createExcelPath('公司大事记收集表.xlsx')
        


        if downloadAll==True:#是下载全部   有条件
            
            kwargs={'memorabilia_status':True}
            memorabilia_base = self.request.GET.get('memorabilia_base_id', None)  # 中心/基地
            memorabilia_key_events_name = self.request.GET.get('memorabilia_key_events_name', None)  # 事件名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if memorabilia_key_events_name == '' and memorabilia_base == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['memorabilia_status'] = True
                kwargs['memorabilia_base__in'] = self.request.user_base
            if memorabilia_base != '':
                kwargs['memorabilia_base'] = memorabilia_base
            if memorabilia_key_events_name != '':
                kwargs['memorabilia_key_events_name'] = memorabilia_key_events_name
            if beginDate != "" and endDate != "":
                kwargs['memorabilia_date__gte'] = datetime(2001, 10, 29, 7, 17, 1,177) if beginDate == None else beginDate
                kwargs['memorabilia_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            index=1
            row_data = []
            all=list(MemorabiliaList.objects.filter(**kwargs).all().values_list('memorabilia_base__id','memorabilia_base__name','memorabilia_date','memorabilia_key_events_name','memorabilia_main_attending_leaders','memorabilia_location','memorabilia_remark'))

            row_data=[]
            index=1
            for i in all:
                i=list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                del i[1]
                i.insert(0,index)
                index+=1
                row_data.append(i)


            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data=[]
            index=1
            for id in id_list:
                data=MemorabiliaList.objects.filter(memorabilia_status=True,id=id).values_list('memorabilia_base__id','memorabilia_base__name','memorabilia_date','memorabilia_key_events_name','memorabilia_main_attending_leaders','memorabilia_location','memorabilia_remark')
                for i in data:
                    i=list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    del i[1]
                    i.insert(0,index)
                    row_data.append(list(i))
                index += 1


            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data ={
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        







    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def create_string_number(self, n):  # 随机名
        m = random.randint(1, n)
        a = "".join([str(random.randint(0, 9)) for _ in range(m)])
        b = "".join([random.choice(string.ascii_letters) for _ in range(n - m)])
        return ''.join(random.sample(list(a + b), n))

    # def createPath(self, pic):  # 生成路径
    #     pic_suffix = str(pic).split(".")[-1]
    #     pic_name = f"{self.create_string_number(20) + 'dsj'}.{pic_suffix}"
    #     pic_path = os.path.join('static', 'memorabiliaFile', pic_name)  # 图片路径
    #
    #     if pic_suffix in ['jpg', 'jpeg', 'png']:
    #         pic_path = os.path.join('static', 'memorabiliaFile', 'img', pic_name)  # 图片路径
    #         if not os.path.exists(pic_path):  # 创建目录
    #             os.mkdir(pic_path)
    #     elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
    #         t = time.strftime('%Y-%m-%d')
    #         pic_name = str(pic)
    #         pic_path = os.path.join('static', 'memorabiliaFile', 'uploadFile', t, str(pic))  # 文件路径
    #
    #         dummy_path = os.path.join('static', 'memorabiliaFile', 'uploadFile', t)
    #         if not os.path.exists(dummy_path):  # 创建目录
    #             os.mkdir(dummy_path)
    #
    #     pic_path = pic_path.replace('\\', '/')
    #     return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字




    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'memorabiliaFile','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'memorabiliaFile','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'memorabiliaFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字


    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        # t = time.strftime('%Y-%m-%d')
        # path1 = os.path.join('static', 'memorabiliaFile', 'img')
        # if not os.path.exists(path1):  # 创建目录
        #     
        #     os.mkdir(path1)
        # path2 = os.path.join('static', 'c', 'uploadFile',t)
        # if not os.path.exists(path2):  # 创建目录
        #     
        #     os.mkdir(path2)
        # if pic_suffix in ['jpg', 'jpeg', 'png']:
        #     with open('static/memorabiliaFile/img/' + pic_name, 'wb') as f:
        #         for dot in pic_obj.chunks():
        #             f.write(dot)
        # elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
        #
        #     with open('static/memorabiliaFile/uploadFile/' + t + '/' + pic_name, 'wb+') as f:
        #         for dot in pic_obj.chunks():
        #             f.write(dot)
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/memorabiliaFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/memorabiliaFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        

    def count_character(self, s):  # 计算宽度
        hanzi = 0
        num = 0
        date_str = ['日期', '时间', '照片', '方案','中心/基地']
        try:
            if any(sub in s for sub in date_str):
                return str(150)  # 日期字符串140宽度
        except:
            pass
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1
        len = 30 * hanzi + 20 * num  # 一个汉字30  一个非汉字20宽度
        return str(200) if len > 200 else str(len)

    def createExcelPath(selfself,file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 105):
            sheet.column_dimensions[chr(i).upper()].width = 20
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:H1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '公司大事记收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号', '中心/基地','公司', '时间', '关键事件名称', '主要出席领导（内外部）', '地点', '备注'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'memorabiliaFile', 'downloadFile', t,file_name)  # 文件路径
        
        path_fake=os.path.join('static', 'memorabiliaFile', 'downloadFile',t)
        path_fake = path_fake.replace('\\', '/')
        if not os.path.exists(path_fake):  # 创建目录
            os.mkdir(path_fake)
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def mkdir(self, path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            
        else:
            pass
            