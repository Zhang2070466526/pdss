import json, arrow, os,shutil, openpyxl,calendar
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import status
from pdss.settings import BASE_DIR
from openpyxl.utils.cell import get_column_letter
from utils.genericMethods import methHeader, BasicClass, FileClass
from talentDevelop.views import *
from django.db.models import Q
from datetime import datetime, date, timedelta
from rest_framework import status
from utils.sqlServerConnect import EhrConnect
from employee.models import *
from utils.save_data_to_redis import *
from employeePersonnel.publicMethods import *
import pandas as pd
from employeePersonnel.pretaxSalary.sql import *
from openpyxl.styles import Alignment


class BasicMobilize(methHeader, BasicClass):
    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete
        self.meth['options'] = self.options

        self.ehr = EhrConnect()

    def options(self):  # 下拉框
        hr_job_sequence_list = HrJobSequence.objects.filter().exclude(id=999999).values('id', 'sequence_name')# 序列
        hr_job_grade_list = list(
            HrJobGrade.objects.filter(job_grade_status=True).exclude(id=999999).values('id', 'job_grade_name'))  # 职级
        employee_pay_type_list = list(
            HrPayType.objects.filter(pay_type_status=True).exclude(id=999999).values('id', 'pay_type_name'))  # 计薪方式

        return_data = {
            'data': {
                'employee_pay_type_list': [
                    {"value": item["id"], "label": item["pay_type_name"]}
                    for item in employee_pay_type_list
                ],
                'employee_job_grade_list': [
                    {"value": item["id"], "label": item["job_grade_name"]}
                    for item in hr_job_grade_list
                ],
                'employee_job_sequence_list': [
                    {"value": item['id'], "label": item['sequence_name']}
                    for item in hr_job_sequence_list
                ],
                'employee_dl_list': [
                    {"value": 'DL', "label": 'DL'},
                    {"value": 'IDL', "label": 'IDL'},
                    {"value": 'SAL', "label": 'SAL'}
                ],


            },
            "code": status.HTTP_200_OK,
            "msg": "下拉菜单返回成功",
            'hidden': True,
        }
        self.return_data = return_data

    def search(self):
        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        info = json.loads(self.request.body)
        print(info)
        current_page = info.get('currentPage', 1)
        page_size = info.get('pageSize', 25)
        search_name = info.get('searchName', '')
        employee_department = info.get('department_id', [])  # 部门
        employee_dl = info.get('employeeDl', [])  # 成本类别
        employee_pay_type = info.get('employeePayType', [])  # 计薪方式
        employee_job_grade = info.get('employeeJobGrade', [])  # 职级
        employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列
        periodDate = info.get('PeriodDate', [])  # 绩效周期
        if periodDate is None or len(periodDate) == 0:
            period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
        else:
            period_begin_date, period_end_date = periodDate

        date_range = pd.date_range(
            start=formatted_dates[0] if len(period_begin_date) == 0 or period_begin_date is None else period_begin_date,
            end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
            freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        # print('2',date_range)
        # print(formatted_dates)
        # print(formatted_dates2)

        kwargs = {
            'employee_dl__in': employee_dl,
            'employee_department__in': self.request.user_department_employee if employee_department is None or len(
                employee_department) == 0 else employee_department,
            'employee_pay_type__in': employee_pay_type,
            'employee_job_grade__in': employee_job_grade,
            'employee_job_sequence__in': employee_job_sequence,
            'employee_status': '1',
        }
        kwargs = {key: value for key, value in kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项


        column_list = [
            {'label': 'index', 'value': '序号', 'width': 80},
            {'label': 'employee_code', 'value': '工号', 'width': 120},
            {'label': 'employee_name', 'value': '姓名', 'width': 150},
            {'label': 'employee_group_join_date', 'value': '集团入职日期', 'width': 160},
            {'label': 'employee_join_date', 'value': '入职日期', 'width': 160},
            {'label': 'employee_department__department_full_name', 'value': '部门全称', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'employee_job_grade__job_grade_name', 'value': '职级', 'width': 130},
            {'label': 'employee_pay_type__pay_type_name', 'value': '薪酬类型', 'width': 130},
            {'label': 'employee_dl', 'value': '成本类别', 'width': 130},
        ]
        column_list.extend([{
            "value": "发放月",
            "children": [
                {"label": 'pr_yfhj__'+month, "value": month, "width":130} for month in formatted_dates2
            ]},
            {'label': 'pr_yfhj__total', 'value': '合计', 'width': 130},
        ]
        )
        total_number = HrEmployee.objects.filter(
            Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name), **kwargs).count()
        employee_list = list(
            HrEmployee.objects.filter(Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name),
                                      **kwargs).values('id', 'employee_name',
                                                       'employee_code',
                                                       'employee_group_join_date',
                                                       'employee_join_date',
                                                       'employee_department__department_full_name',
                                                       'employee_department__department_first_name',
                                                       'employee_department__department_second_name',
                                                       'employee_department__department_third_name',
                                                       'employee_job_grade__job_grade_name',
                                                       "employee_pay_type__pay_type_name", "employee_dl"
                                                       ).order_by('-employee_group_join_date')[(current_page - 1) * page_size:current_page * page_size])
        employee_id_list= [item['id'] for item in employee_list]
        employee_id_tuple = tuple(employee_id_list)
        if len(employee_id_tuple) == 1:
            employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
        payroll_list = self.ehr.select(employee_payroll_sql(employee_id_tuple,formatted_dates[0]+'-01',formatted_dates[-1]+'-01'))  #薪资
        # print(payroll_list)
        decoded_payroll_list = [
            {
                'pr_emp_id': item['pr_emp_id'],
                'pr_pi_name': item['pr_pi_name'],
                'pr_yfhj':None if item['pr_yfhj'] is None else float(item['pr_yfhj']),
            }
            for item in payroll_list
        ]
        new_data = [
            {f'{k}__{v["pr_pi_name"]}': v[k] for k in ('pr_yfhj',)} | {'pr_emp_id': v['pr_emp_id'],'pr_pi_name': v['pr_pi_name']}
            for v in decoded_payroll_list
        ]  #json 数据格式化
        from collections import defaultdict
        merged_data = defaultdict(dict)
        for item in new_data:
            merged_data[item['pr_emp_id']].update(item)
        payroll_list = list(merged_data.values())
        header_date_list = []
        for header_date in formatted_dates2:
            header_date_list.extend(['pr_yfhj__' + header_date])

        result_list = []

        def get_payroll_entry(employee_id):
            return next((payroll for payroll in payroll_list if payroll['pr_emp_id'] == employee_id), None)

        result_list = [
            {
                **employee,
                "pr_emp_id": employee['id'],
                **{header_date: get_payroll_entry(employee['id']).get(header_date, None) if get_payroll_entry(
                    employee['id']) is not None else None
                   for header_date in header_date_list},
                "pr_yfhj__total": round(
                    sum(
                        value for key, value in (get_payroll_entry(employee['id']) or {}).items()
                        if key.startswith("pr_yfhj__") and key != "pr_yfhj__total"
                    ), 2
                )

            }
            for employee in employee_list
        ]
        for index, item in enumerate(result_list):
            item['index'] = (current_page - 1) * page_size + index + 1
            item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
            item['employee_join_date'] =  str(item['employee_join_date'])[:10]

        self.return_data['data']['columnList'] = column_list
        self.return_data['data']['tableList'] = result_list
        self.return_data['data']['totalNumber'] = total_number

    def create(self):
        pass
    def update(self):
        pass

    def delete(self):
        pass


class MobilizeInfoFileClass(methHeader, FileClass):
    def __init__(self, request):
        super().__init__(request)
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.format('YYYY-MM-DD_HH_mm_ss')

    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download
        self.ehr = EhrConnect()

    def upload(self):
        pass
    def download(self):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1,
                                  str(self.t2))  # 创建文件夹
        mkdir(dummy_path)
        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        info = json.loads(self.request.body)
        id_list = info.get('id_list',[])
        download_all = info.get('downloadAll',False)

        periodDate = info.get('PeriodDate', [])  # 绩效周期
        if periodDate is None or len(periodDate) == 0:
            period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
        else:
            period_begin_date, period_end_date = periodDate


        date_range = pd.date_range(
            start=formatted_dates[0] if len(period_begin_date) == 0 or period_begin_date is None else period_begin_date,
            end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
            freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        template_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'template_file',
                                     '员工历月税前工资总额下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1,
                                        str(self.t2),
                                        '员工历月税前工资总额.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        wb = openpyxl.load_workbook(destination_path)
        ws = wb.active
        # 合并单元格
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=13 + len(formatted_dates) - 1)
        ws['M1'].value = '发放月'
        ws['M1'].alignment = Alignment(horizontal='center')
        column = 'M'
        for i, value in enumerate(formatted_dates, start=2):
            ws.cell(row=2, column=ord(column) - ord('A') + 1).value = value
            column = chr(ord(column) + 1)
        start_column=ord(column) - ord('A') + 1
        ws.merge_cells(start_row=1, start_column=start_column, end_row=2, end_column=start_column)
        ws.cell(row=1, column=start_column).value = '合计'

        wb.save(destination_path)
        row_data = []
        if download_all:
            search_name = info.get('searchName', '')
            employee_department = info.get('department_id', [])  # 部门
            employee_dl = info.get('employeeDl', [])  # 成本类别
            employee_pay_type = info.get('employeePayType', [])  # 计薪方式
            employee_job_grade = info.get('employeeJobGrade', [])  # 职级
            employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列
            periodDate = info.get('PeriodDate', [])  # 绩效周期
            if periodDate is None or len(periodDate) == 0:
                period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
            else:
                period_begin_date, period_end_date = periodDate

            date_range = pd.date_range(
                start=formatted_dates[0] if len(
                    period_begin_date) == 0 or period_begin_date is None else period_begin_date,
                end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
                freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
            date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
            formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
            formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串


            kwargs = {
                'employee_dl__in': employee_dl,
                'employee_department__in': self.request.user_department_employee if employee_department is None or len(
                    employee_department) == 0 else employee_department,
                'employee_pay_type__in': employee_pay_type,
                'employee_job_grade__in': employee_job_grade,
                'employee_job_sequence__in': employee_job_sequence,
                'employee_status': '1',
            }
            kwargs = {key: value for key, value in kwargs.items() if
                      value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项
            employee_list = list(
                HrEmployee.objects.filter(
                    Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name),
                    **kwargs).values('id',
                                     'employee_code',
                                     'employee_name',
                                     'employee_group_join_date',
                                     'employee_join_date',
                                     'employee_department__department_full_name',
                                     'employee_department__department_first_name',
                                     'employee_department__department_second_name',
                                     'employee_department__department_third_name',
                                     'employee_job_grade__job_grade_name',
                                     "employee_pay_type__pay_type_name", "employee_dl"
                                     ).order_by('-employee_group_join_date'))
            employee_id_list = [item['id'] for item in employee_list]
            employee_id_tuple = tuple(employee_id_list)
            if len(employee_id_tuple) == 1:
                employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
            payroll_list = self.ehr.select(
                employee_payroll_sql(employee_id_tuple, formatted_dates[0] + '-01', formatted_dates[-1] + '-01'))  # 薪资
            # print(payroll_list)
            decoded_payroll_list = [
                {
                    'pr_emp_id': item['pr_emp_id'],
                    'pr_pi_name': item['pr_pi_name'],
                    'pr_yfhj': None if item['pr_yfhj'] is None else float(item['pr_yfhj']),
                }
                for item in payroll_list
            ]
            new_data = [
                {f'{k}__{v["pr_pi_name"]}': v[k] for k in ('pr_yfhj',)} | {'pr_emp_id': v['pr_emp_id'],
                                                                           'pr_pi_name': v['pr_pi_name']}
                for v in decoded_payroll_list
            ]  # json 数据格式化
            from collections import defaultdict
            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['pr_emp_id']].update(item)
            payroll_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['pr_yfhj__' + header_date])
            def get_payroll_entry(employee_id):
                return next((payroll for payroll in payroll_list if payroll['pr_emp_id'] == employee_id), None)

            result_list = [
                {
                    **employee,
                    "pr_emp_id": employee['id'],
                    **{header_date: get_payroll_entry(employee['id']).get(header_date, None) if get_payroll_entry(
                        employee['id']) is not None else None
                       for header_date in header_date_list},
                    "pr_yfhj__total": round(
                        sum(
                            value for key, value in (get_payroll_entry(employee['id']) or {}).items()
                            if key.startswith("pr_yfhj__") and key != "pr_yfhj__total"
                        ), 2
                    )

                }
                for employee in employee_list
            ]
            for index, item in enumerate(result_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]
            count = 1
            for line in result_list:
                del line['id']
                del line['pr_emp_id']
                # del line['pi_emp_id']
                # del line['ak_emp_id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1
        else:
            employee_list = list(
                HrEmployee.objects.filter(id__in=id_list).values('id',
                                     'employee_code',
                                     'employee_name',
                                     'employee_group_join_date',
                                     'employee_join_date',
                                     'employee_department__department_full_name',
                                     'employee_department__department_first_name',
                                     'employee_department__department_second_name',
                                     'employee_department__department_third_name',
                                     'employee_job_grade__job_grade_name',
                                     "employee_pay_type__pay_type_name", "employee_dl"
                                     ).order_by('-employee_group_join_date'))
            employee_id_list = [item['id'] for item in employee_list]
            employee_id_tuple = tuple(employee_id_list)
            if len(employee_id_tuple) == 1:
                employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
            payroll_list = self.ehr.select(
                employee_payroll_sql(employee_id_tuple, formatted_dates[0] + '-01', formatted_dates[-1] + '-01'))  # 薪资
            # print(payroll_list)
            decoded_payroll_list = [
                {
                    'pr_emp_id': item['pr_emp_id'],
                    'pr_pi_name': item['pr_pi_name'],
                    'pr_yfhj': None if item['pr_yfhj'] is None else float(item['pr_yfhj']),
                }
                for item in payroll_list
            ]
            new_data = [
                {f'{k}__{v["pr_pi_name"]}': v[k] for k in ('pr_yfhj',)} | {'pr_emp_id': v['pr_emp_id'],
                                                                           'pr_pi_name': v['pr_pi_name']}
                for v in decoded_payroll_list
            ]  # json 数据格式化
            from collections import defaultdict
            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['pr_emp_id']].update(item)
            payroll_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['pr_yfhj__' + header_date])

            def get_payroll_entry(employee_id):
                return next((payroll for payroll in payroll_list if payroll['pr_emp_id'] == employee_id), None)

            result_list = [
                {
                    **employee,
                    "pr_emp_id": employee['id'],
                    **{header_date: get_payroll_entry(employee['id']).get(header_date, None) if get_payroll_entry(
                        employee['id']) is not None else None
                       for header_date in header_date_list},
                    "pr_yfhj__total": round(
                        sum(
                            value for key, value in (get_payroll_entry(employee['id']) or {}).items()
                            if key.startswith("pr_yfhj__") and key != "pr_yfhj__total"
                        ), 2
                    )

                }
                for employee in employee_list
            ]
            for index, item in enumerate(result_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]
            count = 1
            for line in result_list:
                del line['id']
                del line['pr_emp_id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1

        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]

        self.return_data['msg'] = '下载成功'
        self.return_data['downloadUrl'] = destination_path

