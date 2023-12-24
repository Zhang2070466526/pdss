import json, arrow, os,shutil, openpyxl,calendar
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import status

from employeePersonnel.adjustmentSalary.sql import *
from pdss.settings import BASE_DIR
from openpyxl.utils.cell import get_column_letter
from utils.genericMethods import methHeader, BasicClass, FileClass
from talentDevelop.views import *
from django.db.models import Q
from datetime import datetime, date, timedelta
from rest_framework import status
from utils.sqlServerConnect import EhrConnect
from employee.models import *
from utils.save_data_to_redis import *
from employeePersonnel.publicMethods import *
import pandas as pd
from employeePersonnel.pretaxSalary.sql import *
from openpyxl.styles import Alignment
from employeePersonnel.models import *
from collections import defaultdict

class BasicMobilize(methHeader, BasicClass):
    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete
        self.meth['options'] = self.options

        self.ehr = EhrConnect()

    def options(self):  # 下拉框
        hr_job_sequence_list = HrJobSequence.objects.filter().exclude(id=999999).values('id', 'sequence_name')# 序列
        hr_job_grade_list = list(
            HrJobGrade.objects.filter(job_grade_status=True).exclude(id=999999).values('id', 'job_grade_name'))  # 职级
        employee_pay_type_list = list(
            HrPayType.objects.filter(pay_type_status=True).exclude(id=999999).values('id', 'pay_type_name'))  # 计薪方式

        return_data = {
            'data': {
                'employee_pay_type_list': [
                    {"value": item["id"], "label": item["pay_type_name"]}
                    for item in employee_pay_type_list
                ],
                'employee_job_grade_list': [
                    {"value": item["id"], "label": item["job_grade_name"]}
                    for item in hr_job_grade_list
                ],
                'employee_job_sequence_list': [
                    {"value": item['id'], "label": item['sequence_name']}
                    for item in hr_job_sequence_list
                ],
                'employee_dl_list': [
                    {"value": 'DL', "label": 'DL'},
                    {"value": 'IDL', "label": 'IDL'},
                    {"value": 'SAL', "label": 'SAL'}
                ],
                'employee_job_grade_history_list': [
                    {"value": item["id"], "label": item["job_grade_name"]}
                    for item in hr_job_grade_list
                ],


            },
            "code": status.HTTP_200_OK,
            "msg": "下拉菜单返回成功",
            'hidden': True,
        }
        self.return_data = return_data

    def search(self):
        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        info = json.loads(self.request.body)
        print(info)
        current_page = info.get('currentPage', 1)
        page_size = info.get('pageSize', 25)
        search_name = info.get('searchName', '')
        employee_department = info.get('department_id', [])  # 部门
        employee_dl = info.get('employeeDl', [])  # 成本类别
        employee_pay_type = info.get('employeePayType', [])  # 计薪方式
        employee_job_grade = info.get('employeeJobGrade', [])  # 职级
        employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列
        employee_job_grade_history_list = info.get('employeeJobGradeHistory', [])  # 历史职级
        periodDate = info.get('PeriodDate', [])  # 绩效周期

        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        if periodDate is None or len(periodDate) == 0:
            period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
        else:
            period_begin_date, period_end_date = periodDate

        date_range = pd.date_range(
            start=formatted_dates[0] if len(period_begin_date) == 0 or period_begin_date is None else period_begin_date,
            end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
            freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串

        kwargs = {
            'employee_dl__in': employee_dl,
            'employee_department__in': self.request.user_department_employee if employee_department is None or len(
                employee_department) == 0 else employee_department,
            'employee_pay_type__in': employee_pay_type,
            'employee_job_grade__in': employee_job_grade,
            'employee_job_sequence__in': employee_job_sequence,
            'employee_status': '1',
        }
        kwargs = {key: value for key, value in kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项


        column_list = [
            {'label': 'index', 'value': '序号', 'width': 80},
            {'label': 'employee_code', 'value': '工号', 'width': 120},
            {'label': 'employee_name', 'value': '姓名', 'width': 150},
            {'label': 'employee_group_join_date', 'value': '集团入职日期', 'width': 160},
            {'label': 'employee_join_date', 'value': '入职日期', 'width': 160},
            {'label': 'employee_department__department_full_name', 'value': '部门全称', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'employee_job_grade__job_grade_name', 'value': '职级', 'width': 130},
            {'label': 'employee_pay_type__pay_type_name', 'value': '薪酬类型', 'width': 130},
            {'label': 'employee_dl', 'value': '成本类别', 'width': 130},
        ]
        column_list.extend([{
            "value": "执行月",
            "children": [
                {"label": 'employee_job_grade__job_grade_name_history__'+month, "value": month, "width": 130} for month in formatted_dates2
            ]},

        ]
        )
        total_number = HrEmployee.objects.filter(
            Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name), **kwargs).count()
        employee_list = list(
            HrEmployee.objects.filter(Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name),
                                      **kwargs).values('id',
                                                       'employee_code',
                                                       'employee_name',
                                                       'employee_group_join_date',
                                                       'employee_join_date',
                                                       'employee_department__department_full_name',
                                                       'employee_department__department_first_name',
                                                       'employee_department__department_second_name',
                                                       'employee_department__department_third_name',
                                                       'employee_job_grade__job_grade_name',
                                                       "employee_pay_type__pay_type_name", "employee_dl"
                                                       ).order_by('-employee_group_join_date')[(current_page - 1) * page_size:current_page * page_size])
        employee_code_list= [item['employee_code'] for item in employee_list]
        formatted_dates_objects = [datetime.strptime(month, '%Y-%m') for month in formatted_dates]
        history_kwargs = {
            'employee_record_type': 2,
            'employee_code__in': employee_code_list,
            'employee_record_begin_time__in': formatted_dates_objects,
            'employee_job_grade__in':employee_job_grade_history_list,
        }
        history_kwargs = {key: value for key, value in history_kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项
        employee_history_list=list(HrEmployeeHistory.objects.filter(**history_kwargs).values('employee_code','employee_name','employee_record_begin_time','employee_job_grade__job_grade_name'))
        decoded_employee_history_list = [
            {
                'employee_code_history': item['employee_code'],
                'employee_record_begin_time_history': item['employee_record_begin_time'].strftime('%Y年%m月'),
                'employee_job_grade__job_grade_name_history': item['employee_job_grade__job_grade_name'] ,
            }
            for item in employee_history_list
        ]
        new_data = [
            {f'{k}__{v["employee_record_begin_time_history"]}': v[k] for k in ('employee_job_grade__job_grade_name_history',)} | {'employee_code_history': v['employee_code_history'],'employee_record_begin_time_history': v['employee_record_begin_time_history']}
            for v in decoded_employee_history_list
        ]  #json 数据格式化


        merged_data = defaultdict(dict)
        for item in new_data:
            merged_data[item['employee_code_history']].update(item)
        merged_employee_history_list = list(merged_data.values())
        header_date_list = []
        for header_date in formatted_dates2:
            header_date_list.extend(['employee_job_grade__job_grade_name_history__' + header_date])

        def get_employee_history_entry(employee_id):
            return next((employee_history for employee_history in merged_employee_history_list if employee_history['employee_code_history'] == employee_id), None)
        employee_history_entries = {
            employee['employee_code']: get_employee_history_entry(employee['employee_code'])
                    for employee in employee_list
        }

        result_list = [
            {
                **employee,
                **{
                    header_date: entry.get(header_date, None) if entry else None
                    for header_date in header_date_list
                },
            }
            for employee, entry in zip(employee_list, map(employee_history_entries.get, [e['employee_code'] for e in employee_list]))
        ]
        for index, item in enumerate(result_list):
            item['index'] = (current_page - 1) * page_size + index + 1
            item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
            item['employee_join_date'] = str(item['employee_join_date'])[:10]

        self.return_data['data']['columnList'] = column_list
        self.return_data['data']['tableList'] = result_list
        self.return_data['data']['totalNumber'] = total_number

    def create(self):
        pass
    def update(self):
        pass

    def delete(self):
        pass


class MobilizeInfoFileClass(methHeader, FileClass):
    def __init__(self, request):
        super().__init__(request)
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.format('YYYY-MM-DD_HH_mm_ss')

    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download
        self.ehr = EhrConnect()

    def upload(self):
        pass
    def download(self):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1,
                                  str(self.t2))  # 创建文件夹
        mkdir(dummy_path)
        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        info = json.loads(self.request.body)
        id_list = info.get('id_list',[])
        download_all = info.get('downloadAll',False)
        periodDate = info.get('PeriodDate', [])  # 绩效周期
        if periodDate is None or len(periodDate) == 0:
            period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
        else:
            period_begin_date, period_end_date = periodDate


        date_range = pd.date_range(
            start=formatted_dates[0] if len(period_begin_date) == 0 or period_begin_date is None else period_begin_date,
            end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
            freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        template_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'template_file',
                                     '员工职级历程下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1,
                                        str(self.t2),
                                        '员工职级历程.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        wb = openpyxl.load_workbook(destination_path)
        ws = wb.active
        # 合并单元格
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=13 + len(formatted_dates) - 1)
        ws['M1'].value = '执行月'
        ws['M1'].alignment = Alignment(horizontal='center')
        column = 'M'
        for i, value in enumerate(formatted_dates, start=2):
            ws.cell(row=2, column=ord(column) - ord('A') + 1).value = value
            column = chr(ord(column) + 1)


        wb.save(destination_path)
        row_data = []
        count = 1
        employee_job_grade_history_list = info.get('employeeJobGradeHistory', [])  # 历史职级
        if download_all:
            search_name = info.get('searchName', '')
            employee_department = info.get('department_id', [])  # 部门
            employee_dl = info.get('employeeDl', [])  # 成本类别
            employee_pay_type = info.get('employeePayType', [])  # 计薪方式
            employee_job_grade = info.get('employeeJobGrade', [])  # 职级
            employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列

            kwargs = {
                'employee_dl__in': employee_dl,
                'employee_department__in': self.request.user_department_employee if employee_department is None or len(
                    employee_department) == 0 else employee_department,
                'employee_pay_type__in': employee_pay_type,
                'employee_job_grade__in': employee_job_grade,
                'employee_job_sequence__in': employee_job_sequence,
                'employee_status': '1',
            }
            kwargs = {key: value for key, value in kwargs.items() if
                      value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项
            employee_list = list(
                HrEmployee.objects.filter(
                    Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name),
                    **kwargs).values('id',
                                     'employee_code',
                                     'employee_name',
                                     'employee_group_join_date',
                                     'employee_join_date',
                                     'employee_department__department_full_name',
                                     'employee_department__department_first_name',
                                     'employee_department__department_second_name',
                                     'employee_department__department_third_name',
                                     'employee_job_grade__job_grade_name',
                                     "employee_pay_type__pay_type_name", "employee_dl"
                                     ).order_by('-employee_group_join_date'))
            employee_code_list = [item['employee_code'] for item in employee_list]
            formatted_dates_objects = [datetime.strptime(month, '%Y-%m') for month in formatted_dates]
            history_kwargs = {
                'employee_record_type': 2,
                'employee_code__in': employee_code_list,
                'employee_record_begin_time__in': formatted_dates_objects,
                'employee_job_grade__in': employee_job_grade_history_list,
            }
            history_kwargs = {key: value for key, value in history_kwargs.items() if
                              value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项
            employee_history_list = list(
                HrEmployeeHistory.objects.filter(**history_kwargs).values('employee_code', 'employee_name',
                                                                          'employee_record_begin_time',
                                                                          'employee_job_grade__job_grade_name'))
            decoded_employee_history_list = [
                {
                    'employee_code_history': item['employee_code'],
                    'employee_record_begin_time_history': item['employee_record_begin_time'].strftime('%Y年%m月'),
                    'employee_job_grade__job_grade_name_history': item['employee_job_grade__job_grade_name'],
                }
                for item in employee_history_list
            ]
            new_data = [
                {f'{k}__{v["employee_record_begin_time_history"]}': v[k] for k in
                 ('employee_job_grade__job_grade_name_history',)} | {
                    'employee_code_history': v['employee_code_history'],
                    'employee_record_begin_time_history': v['employee_record_begin_time_history']}
                for v in decoded_employee_history_list
            ]  # json 数据格式化

            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['employee_code_history']].update(item)
            merged_employee_history_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['employee_job_grade__job_grade_name_history__' + header_date])

            def get_employee_history_entry(employee_id):
                return next((employee_history for employee_history in merged_employee_history_list if
                             employee_history['employee_code_history'] == employee_id), None)

            employee_history_entries = {
                employee['employee_code']: get_employee_history_entry(employee['employee_code'])
                for employee in employee_list
            }

            result_list = [
                {
                    **employee,
                    **{
                        header_date: entry.get(header_date, None) if entry else None
                        for header_date in header_date_list
                    },
                }
                for employee, entry in
                zip(employee_list, map(employee_history_entries.get, [e['employee_code'] for e in employee_list]))
            ]
            for index, item in enumerate(result_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]

            for line in result_list:
                del line['id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1


        else:
            employee_list = list(
                HrEmployee.objects.filter(id__in=id_list).values('id',
                                     'employee_code',
                                     'employee_name',
                                     'employee_group_join_date',
                                     'employee_join_date',
                                     'employee_department__department_full_name',
                                     'employee_department__department_first_name',
                                     'employee_department__department_second_name',
                                     'employee_department__department_third_name',
                                     'employee_job_grade__job_grade_name',
                                     "employee_pay_type__pay_type_name", "employee_dl"
                                     ).order_by('-employee_group_join_date'))
            employee_code_list = [item['employee_code'] for item in employee_list]
            formatted_dates_objects = [datetime.strptime(month, '%Y-%m') for month in formatted_dates]
            history_kwargs = {
                'employee_record_type': 2,
                'employee_code__in': employee_code_list,
                'employee_record_begin_time__in': formatted_dates_objects,
                'employee_job_grade__in': employee_job_grade_history_list,
            }
            history_kwargs = {key: value for key, value in history_kwargs.items() if
                              value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项
            employee_history_list = list(
                HrEmployeeHistory.objects.filter(**history_kwargs).values('employee_code', 'employee_name',
                                                                          'employee_record_begin_time',
                                                                          'employee_job_grade__job_grade_name'))
            decoded_employee_history_list = [
                {
                    'employee_code_history': item['employee_code'],
                    'employee_record_begin_time_history': item['employee_record_begin_time'].strftime('%Y年%m月'),
                    'employee_job_grade__job_grade_name_history': item['employee_job_grade__job_grade_name'],
                }
                for item in employee_history_list
            ]
            new_data = [
                {f'{k}__{v["employee_record_begin_time_history"]}': v[k] for k in
                 ('employee_job_grade__job_grade_name_history',)} | {
                    'employee_code_history': v['employee_code_history'],
                    'employee_record_begin_time_history': v['employee_record_begin_time_history']}
                for v in decoded_employee_history_list
            ]  # json 数据格式化

            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['employee_code_history']].update(item)
            merged_employee_history_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['employee_job_grade__job_grade_name_history__' + header_date])

            def get_employee_history_entry(employee_id):
                return next((employee_history for employee_history in merged_employee_history_list if
                             employee_history['employee_code_history'] == employee_id), None)

            employee_history_entries = {
                employee['employee_code']: get_employee_history_entry(employee['employee_code'])
                for employee in employee_list
            }

            result_list = [
                {
                    **employee,
                    **{
                        header_date: entry.get(header_date, None) if entry else None
                        for header_date in header_date_list
                    },
                }
                for employee, entry in
                zip(employee_list, map(employee_history_entries.get, [e['employee_code'] for e in employee_list]))
            ]
            for index, item in enumerate(result_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]

            for line in result_list:
                del line['id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1

        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]

        self.return_data['msg'] = '下载成功'
        self.return_data['downloadUrl'] = destination_path

