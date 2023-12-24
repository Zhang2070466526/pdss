import json, arrow, os,shutil, openpyxl
from django.db import connection
from django.db.models import Q
from datetime import datetime, date, timedelta
from rest_framework import status

from employee import views
from employee.models import *
from employeePersonnel.models import HrEmployeeHistory
from pdss.settings import BASE_DIR
# from IeProposal.proposal.proposalClass import *
from django.db.models import Count
from utils.sqlServerConnect import EhrConnect
from employee.models import *
from utils.save_data_to_redis import *
from employeePersonnel.publicMethods import *
# from employeePersonnel.views import get_week_begin_end,get_month_begin_end
import pandas as pd
from datetime import datetime, timedelta
from employeePersonnel.incumbencySalary.sql import *
from openpyxl.utils.cell import get_column_letter



def grade_assessment(value):
    if value is None:
        return None
    elif value < 80:
       return '不合格'
    elif value < 90:
       return '待改进'
    elif value < 110:
       return '合格'
    elif value < 120:
       return '良好'
    elif value < 140:
       return '优秀'


class Salary:   #员工档案切片
    def __init__(self,request,meth):
        self.request = request
        self.ehr = EhrConnect()
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.timestamp()
        self.now_last_year=(self.now - timedelta(days=365)).format('YYYY-MM-DD')
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = meth
        self.methods = {
            'analyse_salary_info': self.analyse_salary_info,
            'analyse_salary_options':self.analyse_salary_options,# 下拉框
            'analyse_salary_down':self.analyse_salary_down,#下载

        }


    def meth_center(self):
        self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return JsonResponse(self.return_data)
        self.methods[self.meth]()
        return JsonResponse(self.return_data)


    def analyse_salary_options(self):
        hr_job_sequence_list = HrJobSequence.objects.filter().exclude(id=999999).values('id','sequence_name')
        hr_job_grade_list = list(HrJobGrade.objects.filter(job_grade_status=True).exclude(id=999999).values('id', 'job_grade_name'))  # 职级
        employee_pay_type_list = list(HrPayType.objects.filter(pay_type_status=True).exclude(id=999999).values('id', 'pay_type_name'))  # 计薪方式

        return_data = {
            'data': {
                'employee_pay_type_list': [
                    {"value": item["id"], "label": item["pay_type_name"]}
                    for item in employee_pay_type_list
                ],
                'employee_job_grade_list': [
                    {"value": item["id"], "label": item["job_grade_name"]}
                    for item in hr_job_grade_list
                ],
                'employee_job_sequence_list': [
                    {"value": item['id'], "label": item['sequence_name']}
                    for item in hr_job_sequence_list
                ],
                'employee_dl_list': [
                    {"value": 'DL', "label": 'DL'},
                    {"value": 'IDL', "label": 'IDL'},
                    {"value": 'SAL', "label": 'SAL'}
                ],
                'jx_grade_list': [
                    {"value": '优秀', "label": '优秀'},
                    {"value": '良好', "label": '良好'},
                    {"value": '合格', "label": '合格'},
                    {"value": '待改进', "label": '待改进'},
                    {"value": '不合格', "label": '不合格'},
                ],

            },
            "code": status.HTTP_200_OK,
            "msg": "下拉菜单返回成功",
            'hidden': True,
        }
        self.return_data=return_data
    def analyse_salary_info(self):
        date_range = pd.date_range(end=datetime.now(),freq='M',periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串

        info = json.loads(self.request.body)
        current_page = info.get('currentPage', 1)
        page_size = info.get('pageSize', 25)
        search_name = info.get('searchName', '')
        employee_department = info.get('department_id', [])  # 部门
        employee_dl = info.get('employeeDl', [])  # 成本类别
        employee_pay_type = info.get('employeePayType', [])  # 计薪方式
        employee_job_grade = info.get('employeeJobGrade', [])  # 职级
        employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列
        periodDate = info.get('PeriodDate',[])  # 绩效周期
        if periodDate is None or len(periodDate)==0:
            period_begin_date,period_end_date= formatted_dates[0],formatted_dates[-1]
        else:
            period_begin_date,period_end_date=periodDate
        jx_grade = info.get('jxGrade', [])  # 绩效等级
        date_range = pd.date_range(start=formatted_dates[0] if len(period_begin_date)==0 or period_begin_date is None else period_begin_date , end=formatted_dates[-1] if len(period_end_date)==0 or period_end_date is None else period_end_date , freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        kwargs = {
            'employee_dl__in': employee_dl,
            'employee_department__in': self.request.user_department_employee if employee_department is None or  len(employee_department) == 0  else employee_department,
            'employee_pay_type__in': employee_pay_type,
            'employee_job_grade__in': employee_job_grade,
            'employee_job_sequence__in': employee_job_sequence,
            'employee_status':'1',
        }
        kwargs = {key: value for key, value in kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项

        salary_adjustment_type_dict={'01':"试用薪资","02":"转正调薪","03":"异动晋升","04":"异动降级","05":"年度调薪","06":"特别调薪","07":"其他调薪"}  #调薪类型
        transfer_type_dict = {'1': "平调", "2": "晋级", "3": "降级", "4": "工号变动","5": "跨基地平调", "6": "跨基地异动晋身", "7": "跨基地异动降级",'8':'基地内平调','9':'基地内晋升','10':'基地内降级','11':'用工性质调整','12':'降职','13':'基地内降职','14':'跨基地异动降职'}  # 调职类型
        ak_grjg_dict = {'ºÏ¸ñ': "合格", "´ý¸Ä½ø": "待改进", "ÓÅÐã": "优秀", "Á¼ºÃ": "良好", "²»ºÏ¸ñ": "不合格","None":None}

        column_list = [
            {'label': 'index', 'value': '序号', 'width': 80},
            {'label': 'employee_code', 'value': '工号', 'width': 120},
            {'label': 'employee_name', 'value': '姓名', 'width': 150},
            {'label': 'employee_group_join_date', 'value': '集团入职日期', 'width': 160},
            {'label': 'employee_join_date', 'value': '入职日期', 'width': 160},
            {'label': 'employee_department__department_full_name', 'value': '部门全称', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'employee_job_grade__job_grade_name', 'value': '职级', 'width': 130},
            {'label': 'employee_pay_type__pay_type_name', 'value': '薪酬类型', 'width': 130},
            {'label': 'employee_dl', 'value': '成本类别', 'width': 130},
            {'label': 'ts_date', 'value': '上次调职日期', 'width': 130},
            {'label': 'ts_type', 'value': '上次调职类型', 'width': 130},
            {'label': 'pi_salary_date', 'value': '上次调薪日期', 'width': 130},
            {'label': 'pi_salary_adjustment_type', 'value': '上次调薪类型', 'width': 130},
        ]
        column_list.extend(
            {'value': header_date, 'children': [
                {'label': 'ak_grdf__'+header_date, 'value': '绩效得分', 'width': 130},
                {'label': 'ak_grjg__'+header_date, 'value': '绩效等级', 'width': 130},
                # {'label': 'index', 'value': '绩效系数', 'width': 230},
                # {'label': 'index', 'value': '绩效薪资匹配度', 'width': 230},
            ]}
            for header_date in formatted_dates2
        )
        total_number = HrEmployee.objects.filter(Q(employee_name__contains=search_name) | Q(employee_code__contains= search_name ), **kwargs).count()
        employee_list=list(HrEmployee.objects.filter(Q(employee_name__contains=search_name) | Q(employee_code__contains= search_name ),**kwargs).values('id','employee_name',
                                                              'employee_code','employee_group_join_date','employee_join_date','employee_department__department_full_name','employee_department__department_first_name',
                                                           'employee_department__department_second_name','employee_department__department_third_name',
                                                           'employee_job_grade__job_grade_name',"employee_pay_type__pay_type_name","employee_dl"
                                                              ).order_by('-employee_group_join_date')[(current_page - 1) * page_size:current_page * page_size])

        employee_id_list= [item['id'] for item in employee_list]

        employee_id_tuple = tuple(employee_id_list)
        if len(employee_id_tuple) == 1:
            employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
        jx_grade_tuple = tuple(jx_grade)
        if len(jx_grade_tuple) == 1:
            jx_grade_tuple = "('" + str(jx_grade_tuple[0]) + "')"
        transfer_list = self.ehr.select(employee_newest_transfer_sql(employee_id_tuple))   #  调职
        payitem_list = self.ehr.select(employee_newest_payitem_sql(employee_id_tuple))    #  调薪
        asskpi_list = self.ehr.select(employee_asskpi_sql(employee_id_tuple,formatted_dates[0]+'-01',formatted_dates[-1]+'-01',jx_grade_tuple))  #绩效


        decoded_asskpi_list = [
            {
                'ak_emp_id': item['ak_emp_id'],
                'ak_pd_name': item['ak_pd_name'],
                'ak_grdf':None if item['ak_grdf'] is None else float(item['ak_grdf']),
                'ak_grjg': None if item['ak_grjg'] is None else ak_grjg_dict.get(str(item['ak_grjg']))
                # 'ak_grjg':grade_assessment(item['ak_grdf'])
            }
            for item in asskpi_list
        ]
        new_data = [
            {f'{k}__{v["ak_pd_name"]}': v[k] for k in ('ak_grdf', 'ak_grjg')} | {'ak_emp_id': v['ak_emp_id'],
                                                                                  'ak_pd_name': v['ak_pd_name']}
            for v in decoded_asskpi_list
        ]
        from collections import defaultdict

        merged_data = defaultdict(dict)
        for item in new_data:
            merged_data[item['ak_emp_id']].update(item)

        asskpi_list = list(merged_data.values())
        header_date_list=[]
        for header_date in formatted_dates2:
            header_date_list.extend(['ak_grdf__' + header_date,'ak_grjg__' + header_date])

        merged_dict = {}
        for employee in employee_list:
            matching_transfer = next((t for t in transfer_list if t['ts_emp_id'] == employee['id']), None)
            matching_payitem = next((p for p in payitem_list if p['pi_emp_id'] == employee['id']), None)
            matching_asskpi = next((k for k in asskpi_list if k['ak_emp_id'] == employee['id']), None)
            merged = {**employee,
                      'ts_emp_id': matching_transfer['ts_emp_id'] if matching_transfer else None,
                      'ts_date': matching_transfer['ts_date'] if matching_transfer else None,
                      'ts_type': matching_transfer['ts_type'] if matching_transfer else None,
                      'pi_emp_id': matching_payitem['pi_emp_id'] if matching_payitem else None,
                      'pi_salary_date': matching_payitem['pi_salary_date'] if matching_payitem else None,
                      'pi_salary_adjustment_type': matching_payitem['pi_salary_adjustment_type'] if matching_payitem else None,
                      'ak_emp_id':matching_asskpi['ak_emp_id'] if matching_asskpi else None,
                      # **{str(field): matching_asskpi[str(field)] if matching_asskpi else None for field in header_date_list},
                      **{str(field): matching_asskpi.get(str(field)) if matching_asskpi else None for field in header_date_list},
                      }
            merged_dict[employee['id']] = merged
        merged_list = list(merged_dict.values())
        for index, item in enumerate(merged_list):
            item['index'] = (current_page - 1) * page_size + index + 1
            item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
            item['employee_join_date'] =  str(item['employee_join_date'])[:10]
            try:
                if item['ts_date'] ==None:
                    item['ts_date']=None
                else:
                    item['ts_date'] = str(item['ts_date'])[:10]
            except:
                pass
            try:
                if item['pi_salary_date'] ==None:
                    item['pi_salary_date'] = None
                else:
                    item['pi_salary_date'] = str(item['pi_salary_date'])[:10]
            except:
                pass
            item['pi_salary_adjustment_type']=salary_adjustment_type_dict.get(item['pi_salary_adjustment_type'])
            item['ts_type']=transfer_type_dict.get(item['ts_type'])


        # print(merged_list)
        self.return_data={
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList':column_list,
                'tableList':merged_list,
                'totalNumber':total_number
            }
        }
    def analyse_salary_down(self):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1,str(self.t2))  # 创建文件夹
        mkdir(dummy_path)
        info = json.loads(self.request.body)
        id_list = json.loads(self.request.body).get('id_list')
        download_all = json.loads(self.request.body).get('downloadAll')
        date_range = pd.date_range(end=datetime.now(), freq='M', periods=12)
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        info = json.loads(self.request.body)
        periodDate = info.get('PeriodDate', [])  # 绩效周期
        if periodDate is None or len(periodDate) == 0:
            period_begin_date, period_end_date = formatted_dates[0], formatted_dates[-1]
        else:
            period_begin_date, period_end_date = periodDate

        jx_grade = info.get('jxGrade', [])  # 绩效等级
        date_range = pd.date_range(
            start=formatted_dates[0] if len(period_begin_date) == 0 or period_begin_date is None else period_begin_date,
            end=formatted_dates[-1] if len(period_end_date) == 0 or period_end_date is None else period_end_date,
            freq='M')  # 生成周期内的所有月初  MS：month start   默认是M即ME月末
        date_range = date_range.union([date_range[-1] + pd.offsets.MonthEnd()])
        formatted_dates = list(date_range.strftime('%Y-%m'))  # 将日期格式化为年月字符串
        formatted_dates2 = list(date_range.strftime('%Y年%m月'))  # 将日期格式化为年月字符串
        template_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'template_file','在职员工职级绩效调薪关联分析下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employeePersonnelFile', 'download_file', self.t1, str(self.t2),
                                        '在职员工职级绩效调薪关联分析.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)

        # 加载要下载的工作簿和工作表
        wb = openpyxl.load_workbook(destination_path)
        ws = wb.active

        # 循环遍历列表
        for i, item in enumerate(formatted_dates):
            # 计算列的索引
            col_index = i * 2 + 17  # Q列的索引是17
            # 获取列的字母表示
            start_col = get_column_letter(col_index)
            end_col = get_column_letter(col_index + 1)
            # 合并单元格
            ws.merge_cells(f'{start_col}1:{end_col}1')
            ws[f'{start_col}1'].alignment = Alignment(horizontal='center', vertical='center')  # 居中单元格

            # 在合并的单元格中写入数据
            ws[f'{start_col}1'].value = item
            ws[f'{start_col}2'].value = '绩效得分'
            ws[f'{end_col}2'].value = '绩效等级'

        wb.save(destination_path)
        row_data=[]
        salary_adjustment_type_dict = {'01': "试用薪资", "02": "转正调薪", "03": "异动晋升", "04": "异动降级",
                                       "05": "年度调薪", "06": "特别调薪", "07": "其他调薪"}  # 调薪类型
        transfer_type_dict = {'1': "平调", "2": "晋级", "3": "降级", "4": "工号变动", "5": "跨基地平调",
                              "6": "跨基地异动晋身", "7": "跨基地异动降级", '8': '基地内平调', '9': '基地内晋升',
                              '10': '基地内降级', '11': '用工性质调整', '12': '降职', '13': '基地内降职',
                              '14': '跨基地异动降职'}  # 调职类型
        ak_grjg_dict = {'ºÏ¸ñ': "合格", "´ý¸Ä½ø": "待改进", "ÓÅÐã": "优秀", "Á¼ºÃ": "良好", "²»ºÏ¸ñ": "不合格",
                        "None": None}
        if download_all:
            search_name = info.get('searchName', '')
            employee_department = info.get('department_id', [])  # 部门
            employee_dl = info.get('employeeDl', [])  # 成本类别
            employee_pay_type = info.get('employeePayType', [])  # 计薪方式
            employee_job_grade = info.get('employeeJobGrade', [])  # 职级
            employee_job_sequence = info.get('employeeJobSequence', [])  # 职级序列
            jx_grade = info.get('jxGrade', [])  # 绩效等级


            kwargs = {
                'employee_dl__in': employee_dl,
                'employee_department__in': self.request.user_department_employee if employee_department is None or len(employee_department) == 0 else employee_department,
                'employee_pay_type__in': employee_pay_type,
                'employee_job_grade__in': employee_job_grade,
                'employee_job_sequence__in': employee_job_sequence,
                'employee_status': '1',
            }
            kwargs = {key: value for key, value in kwargs.items() if
                      value is not None and value != '' and value != []}  # 过滤掉值为None或''或[]的项



            employee_list = list(HrEmployee.objects.filter(
                Q(employee_name__contains=search_name) | Q(employee_code__contains=search_name), **kwargs).values('id',
                                                                                                                  'employee_code',
                                                                                                                  'employee_name',
                                                                                                                  'employee_group_join_date',
                                                                                                                  'employee_join_date',
                                                                                                                  'employee_department__department_full_name',
                                                                                                                  'employee_department__department_first_name',
                                                                                                                  'employee_department__department_second_name',
                                                                                                                  'employee_department__department_third_name',
                                                                                                                  'employee_job_grade__job_grade_name',
                                                                                                                  "employee_pay_type__pay_type_name",
                                                                                                                  "employee_dl"
                                                                                                                  ).order_by(
                '-employee_group_join_date'))
            employee_id_list = [item['id'] for item in employee_list]

            employee_id_tuple = tuple(employee_id_list)
            if len(employee_id_tuple) == 1:
                employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
            jx_grade_tuple = tuple(jx_grade)
            if len(jx_grade_tuple) == 1:
                jx_grade_tuple = "('" + str(jx_grade_tuple[0]) + "')"

            transfer_list = self.ehr.select(employee_newest_transfer_sql(employee_id_tuple))  # 调职
            payitem_list = self.ehr.select(employee_newest_payitem_sql(employee_id_tuple))  # 调薪
            asskpi_list = self.ehr.select(employee_asskpi_sql(employee_id_tuple, formatted_dates[0] + '-01', formatted_dates[-1] + '-01', jx_grade_tuple))  # 绩效

            decoded_asskpi_list = [
                {
                    'ak_emp_id': item['ak_emp_id'],
                    'ak_pd_name': item['ak_pd_name'],
                    'ak_grdf': None if item['ak_grdf'] is None else float(item['ak_grdf']),
                    # 'ak_grjg':grade_assessment(item['ak_grdf'])
                    'ak_grjg': None if item['ak_grjg'] is None else ak_grjg_dict.get(str(item['ak_grjg']))
                }
                for item in asskpi_list
            ]
            new_data = [
                {f'{k}__{v["ak_pd_name"]}': v[k] for k in ('ak_grdf', 'ak_grjg')} | {'ak_emp_id': v['ak_emp_id'],'ak_pd_name': v['ak_pd_name']}
                for v in decoded_asskpi_list
            ]
            from collections import defaultdict

            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['ak_emp_id']].update(item)

            asskpi_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['ak_grdf__' + header_date, 'ak_grjg__' + header_date])

            merged_dict = {}
            for employee in employee_list:
                matching_transfer = next((t for t in transfer_list if t['ts_emp_id'] == employee['id']), None)
                matching_payitem = next((p for p in payitem_list if p['pi_emp_id'] == employee['id']), None)
                matching_asskpi = next((k for k in asskpi_list if k['ak_emp_id'] == employee['id']), None)
                merged = {**employee,
                          'ts_emp_id': matching_transfer['ts_emp_id'] if matching_transfer else None,
                          'ts_date': matching_transfer['ts_date'] if matching_transfer else None,
                          'ts_type': matching_transfer['ts_type'] if matching_transfer else None,
                          'pi_emp_id': matching_payitem['pi_emp_id'] if matching_payitem else None,
                          'pi_salary_date': matching_payitem['pi_salary_date'] if matching_payitem else None,
                          'pi_salary_adjustment_type': matching_payitem['pi_salary_adjustment_type'] if matching_payitem else None,

                          'ak_emp_id': matching_asskpi['ak_emp_id'] if matching_asskpi else None,
                          **{str(field): matching_asskpi.get(str(field)) if matching_asskpi else None for field in header_date_list},
                          }
                merged_dict[employee['id']] = merged
            merged_list = list(merged_dict.values())
            for index, item in enumerate(merged_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]
                try:
                    item['ts_date'] = str(item['ts_date'])[:10]
                except:
                    pass
                try:
                    item['pi_salary_date'] = str(item['pi_salary_date'])[:10]
                except:
                    pass
                item['pi_salary_adjustment_type'] = salary_adjustment_type_dict.get(item['pi_salary_adjustment_type'])
                item['ts_type'] = transfer_type_dict.get(item['ts_type'])
            count = 1
            for line in merged_list:
                if line['ts_date']=='None':
                    line['ts_date']= None
                if line['pi_salary_date']=='None':
                    line['pi_salary_date'] = None

                del line['id']
                del line['ts_emp_id']
                del line['pi_emp_id']
                del line['ak_emp_id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1
        else:
            employee_list = list(HrEmployee.objects.filter(id__in=id_list).values('id',
                                                                                  'employee_code',
                                                                                  'employee_name',
                                                                                  'employee_group_join_date',
                                                                                  'employee_join_date',
                                                                                  'employee_department__department_full_name',
                                                                                  'employee_department__department_first_name',
                                                                                  'employee_department__department_second_name',
                                                                                  'employee_department__department_third_name',
                                                                                  'employee_job_grade__job_grade_name',
                                                                                  "employee_pay_type__pay_type_name",
                                                                                  "employee_dl"
                                                                                  ).order_by(
                '-employee_group_join_date'))
            employee_id_list = [item['id'] for item in employee_list]

            employee_id_tuple = tuple(employee_id_list)
            if len(employee_id_tuple) == 1:
                employee_id_tuple = "(" + str(employee_id_tuple[0]) + ")"
            jx_grade_tuple = tuple(jx_grade)
            if len(jx_grade_tuple) == 1:
                jx_grade_tuple = "('" + str(jx_grade_tuple[0]) + "')"

            transfer_list = self.ehr.select(employee_newest_transfer_sql(employee_id_tuple))  # 调职
            payitem_list = self.ehr.select(employee_newest_payitem_sql(employee_id_tuple))  # 调薪
            asskpi_list = self.ehr.select(
                employee_asskpi_sql(employee_id_tuple, formatted_dates[0] + '-01', formatted_dates[-1] + '-01',
                                    jx_grade_tuple))  # 绩效

            decoded_asskpi_list = [
                {
                    'ak_emp_id': item['ak_emp_id'],
                    'ak_pd_name': item['ak_pd_name'],
                    'ak_grdf': None if item['ak_grdf'] is None else float(item['ak_grdf']),
                    # 'ak_grjg':grade_assessment(item['ak_grdf'])
                    'ak_grjg': None if item['ak_grjg'] is None else ak_grjg_dict.get(str(item['ak_grjg']))
                }
                for item in asskpi_list
            ]
            new_data = [
                {f'{k}__{v["ak_pd_name"]}': v[k] for k in ('ak_grdf', 'ak_grjg')} | {'ak_emp_id': v['ak_emp_id'],
                                                                                     'ak_pd_name': v['ak_pd_name']}
                for v in decoded_asskpi_list
            ]
            from collections import defaultdict

            merged_data = defaultdict(dict)
            for item in new_data:
                merged_data[item['ak_emp_id']].update(item)

            asskpi_list = list(merged_data.values())
            header_date_list = []
            for header_date in formatted_dates2:
                header_date_list.extend(['ak_grdf__' + header_date, 'ak_grjg__' + header_date])

            merged_dict = {}
            for employee in employee_list:
                matching_transfer = next((t for t in transfer_list if t['ts_emp_id'] == employee['id']), None)
                matching_payitem = next((p for p in payitem_list if p['pi_emp_id'] == employee['id']), None)
                matching_asskpi = next((k for k in asskpi_list if k['ak_emp_id'] == employee['id']), None)
                merged = {**employee,
                          'ts_emp_id': matching_transfer['ts_emp_id'] if matching_transfer else None,
                          'ts_date': matching_transfer['ts_date'] if matching_transfer else None,
                          'ts_type': matching_transfer['ts_type'] if matching_transfer else None,
                          'pi_emp_id': matching_payitem['pi_emp_id'] if matching_payitem else None,
                          'pi_salary_date': matching_payitem['pi_salary_date'] if matching_payitem else None,
                          'pi_salary_adjustment_type': matching_payitem[
                              'pi_salary_adjustment_type'] if matching_payitem else None,
                          'ak_emp_id': matching_asskpi['ak_emp_id'] if matching_asskpi else None,
                          **{str(field): matching_asskpi.get(str(field)) if matching_asskpi else None for field in header_date_list},
                          }
                merged_dict[employee['id']] = merged
            merged_list = list(merged_dict.values())
            for index, item in enumerate(merged_list):
                item['employee_group_join_date'] = str(item['employee_group_join_date'])[:10]
                item['employee_join_date'] = str(item['employee_join_date'])[:10]
                try:
                    item['ts_date'] = str(item['ts_date'])[:10]
                except:
                    pass
                try:
                    item['pi_salary_date'] = str(item['pi_salary_date'])[:10]
                except:
                    pass
                item['pi_salary_adjustment_type'] = salary_adjustment_type_dict.get(item['pi_salary_adjustment_type'])
                item['ts_type'] = transfer_type_dict.get(item['ts_type'])

            count = 1
            for line in merged_list:
                if line['ts_date'] == 'None':
                    line['ts_date'] = None
                if line['pi_salary_date'] == 'None':
                    line['pi_salary_date'] = None
                del line['id']
                del line['ts_emp_id']
                del line['pi_emp_id']
                del line['ak_emp_id']
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data.insert(0, count)
                row_data.append(line_data)
                if len(line_data) == 0:
                    count = count
                count += 1

        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }





    @staticmethod
    def days_until_given_date(calculate_time):
        """
        :param datetime:    要计算据今的日期   例datetime.datetime(2023, 7, 19, 0, 0)
        :return:   days_difference 相差的天数
        """
        given_date = calculate_time  # 给定的日期
        current_date = datetime.now()  # 获取当前日期
        time_difference = current_date - given_date
        days_difference = time_difference.days  # 天数
        return days_difference

    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic, path, fileName, father_file):  # 生成路径     文件对象  文件上一级目录名称 文件名称   static下面的目录名称（父目录）
        now = arrow.now().format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  # 文件后缀
        file_name = f"{fileName}.{file_suffix}"  # 文件名称
        file_path = os.path.join('static', father_file, 'report_upload_file', now, path, file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path, file_name, file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path, file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path), 'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)
