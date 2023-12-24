import os,datetime,json,openpyxl,time,random,string
from datetime import datetime, date

from django.db.models import Q
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from salarySurvey.models import SalarySurveyRecord
from auther.models import *
from ..serializers import *
from ..models import *
from django.core.exceptions import ObjectDoesNotExist
from controller.controller2 import Controller, upload_file
from general.models import *
from utils.check_token import *
class Reset:
    # def verify(self,request):
    #     return_data = {'code': '', "message": ''}
    #     new_token = CheckToken()
    #     try:
    #         check_token = new_token.check_token(request.headers['Authorization'])
    #     except Exception as e:
    #         print(e)
    #         return_data['code'] = 400
    #         return_data['message'] = '请求参数出错啦'
    #         return return_data
    #     if check_token is None:
    #         return_data['code'] = 403
    #         return_data['message'] = '没有权限查询'
    #         return return_data
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))
        self.methods[self.meth]()
        return Response(self.return_data)
    def get_list(self):
        columnList=[
                {
                    "value": "序号",
                    "label": "index",
                    "width": 60
                },
                {
                    "value": "奖项名称",
                    "label": "awards_name",
                    "width": 240
                },
                {
                    "value": "中心/事业部",
                    "label": "base_father",
                    "width": 180
                },
                {
                    "value": "公司",
                    "label": "evaluation_company",
                    "width": 120
                },
                {
                    "value": "内部评优时间",
                    "label": "awards_date",
                    "width": 200
                },
                {
                    "value": "获奖人员/团队",
                    "label": "awards_person_or_team",
                    "width": 200
                },
                {
                    "value": "岗位",
                    "label": "awards_position",
                    "width": 210
                },
                {
                    "value": "简要事迹",
                    "label": "brief_story",
                    "width": 300
                },
                {
                    "value": "备注",
                    "label": "awards_remark",
                    "width": ""
                },
                {
                    "value": "评优照片",
                    "label": "awards_photos",
                    "width": 120
                }
        ]
        # print(self.request.GET)
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'awards_status': True
        }
        searchName = self.request.GET.get('searchName', None)
        searchBase = self.request.GET.get('searchBase', None)
        beginDate = self.request.GET.get('beginDate', None)
        endDate = self.request.GET.get('endDate', None)
        if beginDate is not None and endDate is not None:
            kwargs['awards_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(endDate) == 0 else beginDate
            kwargs['awards_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate  #内部评优时间
        if searchName == '' and searchBase == '' and beginDate == "" and endDate == "":  # 全查
            kwargs['awards_status'] = True
            kwargs['evaluation_company__in'] = self.request.user_base
        if searchBase != '':
            kwargs['evaluation_company'] = searchBase
        totalNumber = InternalEvaluationList.objects.filter(Q(awards_name__contains=searchName) | Q(awards_person_or_team__contains=searchName),**kwargs).count()
        tableList=InternalEvaluationList.objects.filter(Q(awards_name__contains=searchName) | Q(awards_person_or_team__contains=searchName),**kwargs).values(
            'id',
            "awards_name",
            "evaluation_company__name",
            "evaluation_company__base_parent_id",
            "evaluation_company",
            "awards_date",
            "awards_person_or_team",
            "awards_position",
            "brief_story",
            "awards_remark",
        ).order_by('-create_time')[(currentPage - 1) * pageSize:currentPage * pageSize]
        base_list = center_base.objects.filter(status=True).values('id', 'name')
        all_inevid = [item['id'] for item in tableList]
        # print(InternalEvaluationList.objects.filter(id=1299,awards_status=True).values('id','awards_photos__id','awards_photos__file_name','awards_photos__file_url'))
        all_inev_file=InternalEvaluationList.objects.filter(id__in=all_inevid,awards_status=True,awards_photos__status=True).values('id','awards_photos__id','awards_photos__file_name','awards_photos__file_url')
        # print(all_inev_file)
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['evaluation_company_id']=item['evaluation_company']
            item['evaluation_company']=item['evaluation_company__name']
            item['awards_photos_files']=[]
            for base in list(base_list):
                if base['id']==item['evaluation_company__base_parent_id']:
                    item['base_father']=base['name']
            if item['evaluation_company__base_parent_id'] is None:
                item['base_father'] = item['evaluation_company__name']
            for inev_file_obj in all_inev_file:
                if inev_file_obj['id']==item['id']:
                    # item['awards_photos']=len(inev_file_obj['id'])
                    if inev_file_obj['awards_photos__id'] is not None and inev_file_obj['awards_photos__file_name'] is not None and inev_file_obj['awards_photos__file_url'] is not None:
                        item['awards_photos_files'].append(
                            {
                                "id": inev_file_obj['awards_photos__id'],
                                "name": inev_file_obj['awards_photos__file_name'],
                                "url": inev_file_obj['awards_photos__file_url'],
                            })
                    # print(inev_file_obj)
            item['awards_photos']=len(item['awards_photos_files'])
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }



        #
        # obj = Controller(InternalEvaluationList, "get_list", self.request)
        # self.return_data = obj.data_start()
    def get_upload(self):
        # print(self.request.POST)
        # print(self.request.FILES)
        file = self.request.FILES.get("file", None)
        createPhoto = self.request.FILES.getlist("createPhoto", None)   #照片
        createData = self.request.POST.get("createData", None)   #[{"awards_date":"2023-05-14","awards_name":"伟大母亲",'awards_remark':"备注一下","awards_person_or_team":"IT","awards_position":"IT工程师","brief_story":"光荣且伟大","evaluation_company_id":1}]
        # searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR, 'static', 'internalEvaluationFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        if file and file is not None:
            # print("文件上传")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name,file,file_suffix)#保存文件
                    exc = openpyxl.load_workbook(file_url)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        # print(sheet.cell(i + 1, 2).value)
                        # awards_date = sheet.cell(i + 1, 6).value
                        try:
                            base_father = center_base.objects.get(status=1, name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg": "某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                evaluation_company_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,
                                                                                   base_parent_id=base_father.id,
                                                                                   status=True).id  # 二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            evaluation_company_id = base_father.id
                        if evaluation_company_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,
                                                                                                          3).value if sheet.cell(
                                    i + 1, 3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:


                            if sheet.cell(i + 1, 8).value != None:
                                try:
                                    awards_date = str(sheet.cell(i + 1, 8).value.date())
                                except:  # 可能不是日期
                                    awards_date = None
                            else:
                                awards_date = None
                            if awards_date == None:
                                self.return_data = {
                                    "code": status.HTTP_401_UNAUTHORIZED,
                                    "msg": "请规范填写内部评优时间哦!"
                                }
                                continue




                            # print(awards_date)
                            # print(str(sheet.cell(i + 1, 6).value.date()) if sheet.cell(i + 1,6).value != None else None)
                            # print(awards_date,sheet.cell(i + 1, 2).value,sheet.cell(i + 1,3).value,sheet.cell(i + 1,4).value,sheet.cell(i + 1, 5).value,sheet.cell(i + 1, 7).value)
                            data_kwargs = {
                                'evaluation_company_id':evaluation_company_id,
                                'awards_date': awards_date,
                                'awards_name':sheet.cell(i + 1, 4).value if sheet.cell(i + 1, 4).value else None,
                                'awards_person_or_team':sheet.cell(i + 1, 5).value if sheet.cell(i + 1, 5).value else None,
                                'awards_position': sheet.cell(i + 1, 6).value,
                                'brief_story': sheet.cell(i + 1,7).value,
                                'awards_remark': sheet.cell(i + 1, 9).value,
                                'creator_id':self.request.check_token
                            }
                            # print(data_kwargs)
                            if data_kwargs['evaluation_company_id'] is None and data_kwargs['awards_date'] is None and data_kwargs['awards_name'] is None and data_kwargs['awards_person_or_team'] is None and data_kwargs['awards_position'] is None and data_kwargs['brief_story'] is None and data_kwargs['awards_remark'] is None:
                                continue
                            else:
                                InternalEvaluationList.objects.create(**data_kwargs)


            except Exception as e:
                # print("++++++++++++++++++")
                # print(e)
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            # print("新增数据")
            createData = eval(createData)
            try:
            # print(createData, type(createData))
            # if createData['evaluation_company_id'] ==None or len(createData['evaluation_company_id'])==0 or createData['evaluation_company_id'] =='':
            #     del createData['evaluation_company']
                createData['creator_id'] = self.request.check_token
                # print(createData)
                obj = InternalEvaluationList.objects.create(**createData)
                # print(obj)
                for file in createPhoto:
                    file_url, file_name, file_suffix = self.createPath(file,
                                                                       self.create_file_name(obj.awards_name, file))
                    # print(file_url,file_name,file_suffix)
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_url,
                        "file_name": file_name,
                    }
                    file_obj = UploadFiles.objects.create(**file_kwargs)
                    file_obj.awards_photos.add(obj.id)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            except:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据错误！"
                }
                # print(self.return_data)

            # print(obj)


    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(InternalEvaluationList, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            info=json.loads(self.request.body)
            InternalEvaluationList.objects.filter(pk=info['id']).update(**info)
            # obj = Controller(InternalEvaluationList, "patch", self.request)
            # obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def download_file(self):
        # print("putdata",self.request.data)
        # print("putdata", self.request.GET)
        id_list=self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')
        path=self.createExcelPath('内部评优收集表.xlsx')
        # print("path",path)

        if downloadAll==True:#是下载全部   有条件
            # print('是下载全部')
            kwargs={'awards_status':True}
            evaluation_company = self.request.GET.get('searchBase', None)  # 公司名称
            awards_name = self.request.GET.get('searchName', None)  # 奖项名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if evaluation_company == '' and awards_name == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['awards_status'] = True
                kwargs['evaluation_company_id__in']=self.request.user_base
            if evaluation_company != '':
                kwargs['evaluation_company'] = evaluation_company
            if awards_name != '':
                kwargs['awards_name'] = awards_name
            if beginDate != "" and endDate != "":
                kwargs['awards_date__gte'] = datetime(1901, 10, 29, 7, 17, 1,177) if beginDate == None else beginDate
                kwargs['awards_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            # print(kwargs)
            all=list(InternalEvaluationList.objects.filter(**kwargs).all().values_list('evaluation_company_id','evaluation_company__name','awards_name','awards_person_or_team','awards_position','brief_story','awards_date','awards_remark'))
            # print(all)
            row_data=[]
            index=1
            for i in all:
                i=list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                del i[1]
                i.insert(0,index)
                index+=1
                row_data.append(i)

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            # print('不是全部')
            row_data=[]
            index=1
            for id in id_list:
                data=InternalEvaluationList.objects.filter(awards_status=True,id=id,).values_list('evaluation_company_id','evaluation_company__name','awards_name','awards_person_or_team','awards_position','brief_story','awards_date','awards_remark')
                # print(data)
                for i in data:
                    i=list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    del i[1]

                    i.insert(0,index)
                    row_data.append(list(i))
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data ={
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # print(self.return_data)







    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # print(fileName)
        return fileName.split('.')[0]

    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'internalEvaluationFile','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'internalEvaluationFile','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'internalEvaluationFile', 'upload_file', t, pic_name)  # 文件路径

        dummy_path = os.path.join('static', 'internalEvaluationFile', 'upload_file', t)
        if not os.path.exists(dummy_path):  # 创建目录
            os.mkdir(dummy_path)
        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/internalEvaluationFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/internalEvaluationFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        # print('保存成功')

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 106):
            sheet.column_dimensions[chr(i).upper()].width = 25
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:I1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '内部评优收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号','中心/事业部','公司','奖项名称（可以是专项奖）','获奖人员/团队','岗位','简要事迹','内部评优时间','备注'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'internalEvaluationFile', 'download_file', t, file_name)  # 文件路径
        # print(path)
        path_fake = os.path.join('static', 'internalEvaluationFile', 'download_file', t)
        path_fake = path_fake.replace('\\', '/')
        if not os.path.exists(path_fake):  # 创建目录
            os.mkdir(path_fake)
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def mkdir(self, path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        else:
            pass