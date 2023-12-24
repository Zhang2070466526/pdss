import json, os,arrow,openpyxl
import time

from django.db.models import Q,F
from rest_framework import status
from django.http import JsonResponse

from auther.models import AdminUser
from employee.models import HrEmployee, HrEmployeeFiles, HrDepartment, HrJobRank
from rest_framework.response import Response
from datetime import datetime, date

from django.utils import timezone
from offlineTraining.models import TrainingLecturer, TrainingLecturerLevel, TrainingContentType, TrainingContent, \
    TrainingSessions
from pdss.settings import BASE_DIR
from employee import views
from offlineTraining.trainClass.contentClass import Content

class Analyse:
    def __init__(self, request,meth):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = meth
        self.methods = {
            'get_training_content_type': self.get_training_content_type, #培训类型 查询
            'post_training_content_type':self.post_training_content_type,#培训类型 新增
            'del_training_content_type': self.del_training_content_type,  # 培训类型 删除
            'edit_training_content_type': self.edit_training_content_type,  # 培训类型 修改

            'month_Training_hours_per_person':self.month_Training_hours_per_person,#本月人均课时  查询
            'download_month_Training_hours_per_person': self.download_month_Training_hours_per_person,  # 本月人均课时  下载
            'edit_month_Training_hours_per_person':self.edit_month_Training_hours_per_person,#修改本月人均课时

            'month_summary_analysis': self.month_summary_analysis,  # 本月汇总分析
            'download_month_summary_analysis': self.download_month_summary_analysis,  # 本月汇总分析 下载
        }

    def method_center(self):
        # self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        # if self.request.check_token is None:
        #     return JsonResponse(self.return_data)
        self.methods[self.meth]()
        return JsonResponse(self.return_data)

    # 获取信息列表
    def get_training_content_type(self):
        columnList = [{'label': 'index', 'value': '序号', 'width': 60},
                      {'label': 'type_name', 'value': '培训类型', 'width': ""},
                      {'label': 'type_first_name', 'value': '一级培训类型', 'width':""},
                      {'label': 'type_second_name', 'value': '二级培训类型', 'width':""},
                      {'label': 'type_third_name', 'value': '三级培训类型', 'width':""},]

        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        searchName = self.request.GET.get('searchName', None)
        totalNumber = TrainingContentType.objects.filter(type_status=True,type_name__contains=searchName).count()
        tableList = list(TrainingContentType.objects.filter(type_status=True,type_name__contains=searchName).values('id',
                  'type_parent_id',
                  'type_name',
                  'type_first_name',
                  'type_second_name',
                  'type_third_name',
                  )[(currentPage - 1) * pageSize:currentPage * pageSize])


        for index, item in enumerate(tableList):
            # ls=[item['id']]
            # if item['type_parent_id'] is not None:
            #     ls.append(item['type_parent_id'])
            ls=[]
            if item['type_third_name'] !=None:
                father_id=list(TrainingContentType.objects.filter(type_status=True,id=item['type_parent_id']).values_list('type_parent_id',flat=True))
                ls =father_id+[item['type_parent_id']]
            elif item['type_second_name'] !=None:
                ls=[item['type_parent_id']]
            else:
                pass

            item['type_parent_id'] =ls
            item['index'] = (currentPage - 1) * pageSize + index + 1


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }
    def post_training_content_type(self):
        info=json.loads(self.request.body)
        type_parent_id=info['type_parent_id']
        if type_parent_id !='':
            type_name_list=list(TrainingContentType.objects.filter(id__in=type_parent_id).values_list('type_name',flat=True))
        else:
            type_name_list=[]
        contentType_params = {
            'type_name': info['type_name'],
            'type_parent_id': type_parent_id[-1] if type_parent_id!='' else  None,
            'type_status':True
        }
        if len(type_name_list) == 2:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name']  = type_name_list[0],type_name_list[1], info['type_name']
        elif len(type_name_list) == 1:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name']  = type_name_list[0],info['type_name'], None
        elif len(type_name_list) ==0:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name'] =info['type_name'], None, None
        TrainingContentType.objects.update_or_create(defaults=contentType_params,type_status=contentType_params['type_status'],type_name=contentType_params['type_name'],type_parent_id=contentType_params['type_parent_id'])
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "新增成功",
        }
    def del_training_content_type(self):
        id_List = json.loads(self.request.body).get('idList')
        id_count=[]
        for id in id_List:
            new_id_List=list(TrainingContentType.objects.filter(Q(id=id)|Q(type_parent_id=id),type_status=True).values_list('id',flat=True))
            for new_id in new_id_List:
                id_count+=list(TrainingContentType.objects.filter(Q(id=new_id)|Q(type_parent_id=new_id),type_status=True).values_list('id',flat=True))
            id_count+=new_id_List
        id_count=list(set(id_count))
        TrainingContentType.objects.filter(id__in=id_count).update(type_status=False)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功",
        }
    def edit_training_content_type(self):
        info=json.loads(self.request.body)
        type_parent_id=info['type_parent_id']
        if len(type_parent_id)>0:
            type_name_list=list(TrainingContentType.objects.filter(id__in=type_parent_id).values_list('type_name',flat=True))
        else:
            type_name_list=[]
        contentType_params = {
            'type_name': info['type_name'],
            'type_parent_id': type_parent_id[-1] if type_parent_id !=[] else  None,
            'type_status':True
        }
        if len(type_name_list) == 2:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name']  = type_name_list[0],type_name_list[1], info['type_name']
        elif len(type_name_list) == 1:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name']  = type_name_list[0],info['type_name'], None
        elif len(type_name_list) ==0:
            contentType_params['type_first_name'],contentType_params['type_second_name'],contentType_params['type_third_name'] =info['type_name'], None, None


        TrainingContentType.objects.filter(pk=info['id']).update(**contentType_params)
        #找出所有父节点是id的
        id_List = list(TrainingContentType.objects.filter(type_parent_id=info['id'], type_status=True).values_list('id', flat=True))
        for id in id_List:
            id_List+=list(TrainingContentType.objects.filter(type_parent_id=id,type_status=True).values_list('id', flat=True))
        if len(type_name_list) == 0:
            TrainingContentType.objects.filter(id__in=id_List).update(type_first_name=info['type_name'])
        elif len(type_parent_id)==1:
            TrainingContentType.objects.filter(id__in=id_List).update(type_first_name=type_name_list[0],type_second_name=info['type_name'])
        elif len(type_parent_id)==2:
            TrainingContentType.objects.filter(id__in=id_List).update(type_first_name=type_name_list[0],type_second_name=type_name_list[1],type_third_name=info['type_name'])
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }



    # def month_Training_hours_per_person(self):   #计算本月人均课时     线下培训总时数=培训时长*参训人数
    #     columnList = [
    #         {'label': 'sessions_base', 'value': '培训基地', 'width':300},
    #         {'label': 'sessions_offline_total', 'value': '线下培训总时数', 'width':260},
    #         {'label': 'sessions_cloud_total', 'value': '线上(云学堂)培训总时数', 'width': 260},
    #         {'label': 'sessions_persons_register', 'value': '月平均在册人数', 'width': 260},
    #         {'label': 'sessions_per_people', 'value': '基地人均培训课时', 'width': ''}
    #     ]
    #
    #     month = self.request.GET.get('month', None)
    #
    #     if month == "":
    #         month = datetime.now().date()
    #     if type(month) == str:
    #         datetime_obj = datetime.strptime(month, "%Y-%m-%d")
    #     else:
    #         datetime_obj = datetime.combine(month, datetime.min.time())
    #     #
    #     current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    #     next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
    #
    #     training_hours_list = TrainingContent.objects.filter(content_begin_date__range=(current_month,next_month),content_end_date__range=(current_month,next_month),content_status=True).values('content_part__department_name', 'content_part__department_code',
    #         'content_part__department_first_name', 'content_part__department_second_name',
    #         'content_module','content_group','content_part','content_duration','content_people_number','content_title','content_begin_date','content_lecturer_id','content_object')
    #
    #     # 总对象
    #     training_hours_total_obj = {}
    #     # 部门分布对象
    #     training_hours_obj = {
    #         'index': '',
    #         'department_name': '',
    #         'content_part__department_first_name': '',
    #         'content_part__department_second_name': '',
    #         'content_part__department_third_name': '',
    #         'content_module':'',#模块==三级
    #         'content_group':'',#组 == 四级
    #
    #         'content_duration': 0,  #培训时长
    #         'content_people_number': 0,# 参训人数,
    #         'total_training_hours':0,#线下培训总时数
    #     }
    #     second_name_list_base = ['江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司',
    #                              '江苏润阳光伏科技有限公司',
    #                              '江苏海博瑞光伏科技有限公司', '润宝电力', '润阳光伏科技（泰国）有限公司',
    #                              '宁夏润阳硅材料科技有限公司', '江苏润阳光伏科技有限公司（二期）',
    #                              '润阳泰国四期组件', '润阳泰国四期电池', '云南润阳世纪光伏科技有限公司']  # 集团权限看的
    #     first_name_list_base = ['光伏研究院', '全球财务中心', '人力资源中心', '全球战略供应链管理中心']  # 集团权限看的
    #     base_ls = list(AdminUser.objects.filter(pk=self.request.check_token).values_list("user_department_employee",
    #                                                                                      flat=True).distinct())
    #     base_name_ls = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                     Q(department_expiry_date__isnull=True) | Q(
    #                                                         department_expiry_date__gt=datetime.now()),
    #                                                     department_second_name__isnull=False,
    #                                                     department_status=1, id__in=base_ls).values_list(
    #         'department_name', flat=True).distinct())
    #     if training_hours_list:
    #         for training_hours in list(training_hours_list):
    #             department_code = training_hours['content_part__department_code']
    #             content_title = training_hours['content_title']
    #             content_begin_date=training_hours['content_begin_date']
    #             code_title=str(department_code)+str(str(time.time())[:8]+str(content_title)+str(content_begin_date))
    #             if department_code not in training_hours_total_obj:
    #                 training_hours_total_obj[code_title] = training_hours_obj.copy()
    #                 training_hours_total_obj[code_title]['department_name'] = training_hours['content_part__department_name']
    #                 training_hours_total_obj[code_title]['content_part__department_first_name'] = training_hours['content_part__department_first_name']
    #                 training_hours_total_obj[code_title]['content_part__department_second_name'] = training_hours['content_part__department_second_name']
    #                 training_hours_total_obj[code_title]['content_module'] = training_hours['content_module']
    #                 training_hours_total_obj[code_title]['content_group'] = training_hours['content_group']
    #                 training_hours_total_obj[code_title]['content_duration'] = float(training_hours['content_duration'])
    #                 training_hours_total_obj[code_title]['content_people_number'] = float(training_hours['content_people_number'])
    #                 training_hours_total_obj[code_title]['total_training_hours'] = float(training_hours['content_duration'])*float(training_hours['content_people_number'])
    #
    #     dept_data_list=[]
    #     training_hours_total_obj = json.loads(json.dumps(training_hours_total_obj))
    #     table_list_true=[]
    #
    #     for code_title, line_data in training_hours_total_obj.items():
    #         line_data['code'] = code_title
    #         dept_data_list.append(line_data)
    #     # print(dept_data_list)
    #     tableList=[]
    #     if 1 in base_ls:#集团公司权限
    #
    #         for base in first_name_list_base + second_name_list_base:
    #             if base in first_name_list_base:
    #                 key = 'content_part__department_first_name'
    #             else:
    #                 key = 'content_part__department_second_name'
    #
    #             # 使用列表推导筛选出符合条件的键值对
    #             filtered_dict = [item['total_training_hours'] for item in dept_data_list if item[key] == base]
    #             training_hours_obj_true = {
    #                 'sessions_base': base,
    #                 'sessions_offline_total': sum(filtered_dict),  # 线下培训总时数
    #                 'sessions_record_time': str(current_month)[:10],  # 记录时间
    #             }
    #             TrainingSessions.objects.update_or_create(defaults=training_hours_obj_true,
    #                                                       sessions_base=training_hours_obj_true['sessions_base'],
    #                                                       sessions_record_time=training_hours_obj_true[
    #                                                           'sessions_record_time'])
    #         # base_name_ls_base=list(set(base_name_ls).intersection(
    #         #     set(second_name_list_base).union(set(first_name_list_base))))
    #         # print(base_name_ls_base)
    #         base_name_ls_base=set(second_name_list_base).union(set(first_name_list_base))
    #         tableList = list(TrainingSessions.objects.filter(sessions_record_time=str(current_month)[:10],sessions_base__in =base_name_ls_base).values('id',
    #                                                                                                 'sessions_base',
    #                                                                                                 'sessions_offline_total',
    #                                                                                                 "sessions_cloud_total",
    #                                                                                                 'sessions_persons_register',
    #                                                                                                 'sessions_per_people',
    #                                                                                                 'sessions_record_time'))
    #         for line in tableList:
    #             if line['sessions_cloud_total'] is None:
    #                 line['sessions_cloud_total'] = 0
    #             if line['sessions_persons_register'] is None:
    #                 line['sessions_persons_register'] = 0
    #             if line['sessions_per_people'] is None:
    #                 line['sessions_per_people'] = 0
    #             if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
    #                 'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
    #                 try:
    #                     line['sessions_per_people'] = round(
    #                         (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
    #                             'sessions_persons_register'], 2)
    #                 except:
    #                     line['sessions_per_people'] = 0
    #             # print(line)
    #             TrainingSessions.objects.update_or_create(defaults=line, sessions_base=line['sessions_base'],
    #                                                       sessions_record_time=line['sessions_record_time'])
    #
    #     else:   #部门权限
    #         first_name_list_base_dept = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                                      Q(department_expiry_date__isnull=True) | Q(
    #                                                                          department_expiry_date__gt=datetime.now()),
    #                                                                      department_first_name__isnull=False,
    #                                                                      department_status=1,
    #                                                                      department_first_name__in=first_name_list_base).values_list(
    #             'department_second_name', flat=True).distinct())
    #         first_name_list_base_dept = [item for item in first_name_list_base_dept if
    #                                      item is not None and item != '']  # 部门权限看的   排除None和''的
    #
    #         second_name_list_base_dept = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                                       Q(department_expiry_date__isnull=True) | Q(
    #                                                                           department_expiry_date__gt=datetime.now()),
    #                                                                       department_second_name__isnull=False,
    #                                                                       department_status=1,
    #                                                                       department_second_name__in=second_name_list_base).values_list(
    #             'department_third_name', flat=True).distinct())
    #         second_name_list_base_dept = [item for item in second_name_list_base_dept if
    #                                       item is not None and item != '']  # 部门权限看的
    #
    #
    #         # base_ls=self.request.user_department_employee
    #         # print(base_ls)
    #
    #
    #
    #         base_name_ls_dept = list(set(base_name_ls).intersection(
    #             set(second_name_list_base_dept).union(set(first_name_list_base_dept))))  # 两个部门权限的并集 与数据库权限的部门取交集
    #
    #
    #
    #         for base in first_name_list_base_dept + second_name_list_base_dept:
    #             if base in first_name_list_base_dept:
    #                 filter_condition = lambda item: item['content_part__department_second_name'] == base
    #             else:
    #                 filter_condition = lambda item: item['content_module'] == base
    #
    #             filtered_dict = [item['total_training_hours'] for item in dept_data_list if filter_condition(item)]
    #
    #             training_hours_obj_true = {
    #                 'sessions_base': base,
    #                 'sessions_offline_total': sum(filtered_dict),  # 线下培训总时数
    #                 'sessions_record_time': str(current_month)[:10],  # 记录时间
    #             }
    #
    #             TrainingSessions.objects.update_or_create(
    #                 defaults=training_hours_obj_true,
    #                 sessions_base=training_hours_obj_true['sessions_base'],
    #                 sessions_record_time=training_hours_obj_true['sessions_record_time']
    #             )
    #             tableList = list(TrainingSessions.objects.filter(sessions_record_time=str(current_month)[:10], sessions_base__in = base_name_ls_dept).values('id',
    #                                                                           'sessions_base',
    #                                                                           'sessions_offline_total',
    #                                                                           "sessions_cloud_total",
    #                                                                           'sessions_persons_register',
    #                                                                           'sessions_per_people',
    #                                                                           'sessions_record_time'))
    #             for line in tableList:
    #                 if line['sessions_cloud_total'] is None:
    #                     line['sessions_cloud_total'] = 0
    #                 if line['sessions_persons_register'] is None:
    #                     line['sessions_persons_register'] = 0
    #                 if line['sessions_per_people'] is None:
    #                     line['sessions_per_people'] = 0
    #                 if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
    #                     'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
    #                     try:
    #                         line['sessions_per_people'] = round(
    #                             (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
    #                                 'sessions_persons_register'], 2)
    #                     except:
    #                         line['sessions_per_people'] = 0
    #                 # print(line)
    #                 TrainingSessions.objects.update_or_create(defaults=line, sessions_base=line['sessions_base'],
    #                                                           sessions_record_time=line['sessions_record_time'])
    #
    #     self.return_data = {
    #         "code": status.HTTP_200_OK,
    #         "msg": "信息返回成功",
    #         "data": {
    #             'columnList': columnList,
    #             'tableList': tableList,
    #         }
    #     }





    def month_Training_hours_per_person(self):  # 计算本月人均课时     线下培训总时数=人数*课时 之和
        columnList = [
            # {'label': 'sessions_base', 'value': '培训基地', 'width': 600},
            # {'label': 'sessions_base_first', 'value': '培训基地(一级)', 'width': 200},
            # {'label': 'sessions_base_second', 'value': '培训基地(二级)', 'width':200},
            # {'label': 'sessions_base_third', 'value': '培训基地(三级)', 'width': 200},

            {'label': 'sessions_offline_total', 'value': '线下培训总时数', 'width': 220},
            {'label': 'sessions_cloud_total', 'value': '线上(云学堂)培训总时数', 'width': 220},
            {'label': 'sessions_persons_register', 'value': '月平均在册人数', 'width': 220},
            {'label': 'sessions_per_people', 'value': '基地人均培训课时', 'width': ''}
        ]
        month = self.request.GET.get('month', None)
        dept_level=int(self.request.GET.get('type', None))


        if month == "":
            month = datetime.now().date()
        if type(month) == str:
            datetime_obj = datetime.strptime(month, "%Y-%m-%d")
        else:
            datetime_obj = datetime.combine(month, datetime.min.time())
        #
        current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)

        training_hours_list = TrainingContent.objects.filter(content_begin_date__range=(current_month, next_month),
                                                             content_end_date__range=(current_month, next_month),
                                                             content_status=True).values(
            'content_part__department_full_name', 'content_part__department_code',
            'content_part__department_first_name', 'content_part__department_second_name',
            'content_part__department_third_name', 'content_part__department_forth_name', 'content_part', 'content_duration', 'content_people_number',
            'content_lecturer_id', 'content_object')

        # print(self.request.user_department_employee)
        user_department_list_employee =list(AdminUser.objects.filter(pk=self.request.check_token).values_list("user_department_employee",flat=True).distinct())
        all_dept=list(HrDepartment.objects.filter( ~Q(id=999999),Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_status=1,department_level=dept_level,id__in=user_department_list_employee).values_list('department_full_name',flat=True))   #一级部门
        #,id__in=self.request.user_department_employee
        for training_hours in  training_hours_list:
            training_hours['content_duration'] =0 if training_hours['content_duration']=='' or training_hours['content_duration'] is None else float(training_hours['content_duration'])
            training_hours['content_people_number'] = 0 if training_hours['content_people_number'] == '' or training_hours['content_people_number'] is None else float(training_hours['content_people_number'])
            training_hours['total_training_hours']=training_hours['content_duration']*training_hours['content_people_number']
        # print(training_hours_list)

        # key=''
        # if dept_level==2:
        #     key='content_part__department_first_name'
        # elif dept_level==3:
        #     key='content_part__department_second_name'
        # elif dept_level==4:
        #     key='content_part__department_third_name'


        if dept_level==2:
            columnList.insert(0,{'label': 'sessions_base_first', 'value': '培训基地(一级)', 'width': 200})
        elif dept_level==3:
            columnList.insert(0, {'label': 'sessions_base_first', 'value': '培训基地(一级)', 'width': 200})
            columnList.insert(1, {'label': 'sessions_base_second', 'value': '培训基地(二级)', 'width': 200})
        elif dept_level==4:
            columnList.insert(0, {'label': 'sessions_base_first', 'value': '培训基地(一级)', 'width': 200})
            columnList.insert(1, {'label': 'sessions_base_second', 'value': '培训基地(二级)', 'width': 200})
            columnList.insert(2, {'label': 'sessions_base_third', 'value': '培训基地(三级)', 'width': 200})


        for base in all_dept:
            # 使用列表推导筛选出符合条件的键值对
            dept, first_dept, second_dept, third_dept=None,None,None,None
            filtered_dict=0
            if dept_level==2:
                dept,first_dept=base.split()
                filtered_dict = [item['total_training_hours'] for item in training_hours_list if item['content_part__department_first_name'] == first_dept ]
            elif dept_level==3:
                dept,first_dept,second_dept=base.split()
                filtered_dict = [item['total_training_hours'] for item in training_hours_list if item['content_part__department_second_name'] == second_dept and item['content_part__department_first_name'] == first_dept ]
            elif dept_level==4:
                dept,first_dept,second_dept,third_dept=base.split()
                filtered_dict = [item['total_training_hours'] for item in training_hours_list if item['content_part__department_third_name'] == third_dept and item['content_part__department_second_name'] == second_dept and item['content_part__department_first_name'] == first_dept]

            # base_full_name=

            training_hours_obj_true = {
                'sessions_base': base,
                'sessions_base_first': first_dept,
                'sessions_base_second': second_dept,
                'sessions_base_third': third_dept,
                'sessions_offline_total': sum(filtered_dict),  # 线下培训总时数
                'sessions_record_time': str(current_month)[:10],  # 记录时间
            }
            # print(training_hours_obj_true)
            # new_training_hours_list.append(training_hours_obj_true)
            TrainingSessions.objects.update_or_create(defaults=training_hours_obj_true,
                                                      sessions_base=training_hours_obj_true['sessions_base'],
                                                      sessions_base_first=training_hours_obj_true['sessions_base_first'],
                                                      sessions_base_second=training_hours_obj_true['sessions_base_second'],
                                                      sessions_base_third=training_hours_obj_true['sessions_base_third'],
                                                      sessions_record_time=training_hours_obj_true[
                                                          'sessions_record_time'])

        tableList = list(TrainingSessions.objects.filter(sessions_record_time=str(current_month)[:10], sessions_base__in = all_dept).values('id',
                                                                    'sessions_base_first',
                                                                    'sessions_base_second',
                                                                    'sessions_base_third',

                                                                      'sessions_offline_total',
                                                                      "sessions_cloud_total",
                                                                      'sessions_persons_register',
                                                                      'sessions_per_people',
                                                                      'sessions_record_time'))
        for line in tableList:
            # line['sessions_offline_total']=[item['sessions_offline_total'] for item in new_training_hours_list if item['sessions_base'] == line['sessions_base'] ]
            if line['sessions_cloud_total'] is None:
                line['sessions_cloud_total'] = 0
            if line['sessions_persons_register'] is None:
                line['sessions_persons_register'] = 0
            if line['sessions_per_people'] is None:
                line['sessions_per_people'] = 0
            if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
                'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
                try:
                    line['sessions_per_people'] = round(
                        (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
                            'sessions_persons_register'], 2)
                except:
                    line['sessions_per_people'] = 0
            TrainingSessions.objects.update_or_create(defaults=line,
                                                      sessions_base_first=line['sessions_base_first'],
                                                      sessions_base_second=line['sessions_base_second'],
                                                      sessions_base_third=line['sessions_base_third'],
                                                      sessions_record_time=line['sessions_record_time'])
        # for line in tableList:
            if dept_level == 2:
                del line['sessions_base_second']
                del line['sessions_base_third']
            elif dept_level == 3:
                del line['sessions_base_third']
            elif dept_level == 4:
                pass

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
            }
        }


    # def month_Training_hours_per_person(self):  # 计算本月人均课时     线下培训总时数=人数*课时 之和
    #     columnList = [
    #         {'label': 'sessions_base', 'value': '培训基地', 'width': 300},
    #         {'label': 'sessions_offline_total', 'value': '线下培训总时数', 'width': 260},
    #         {'label': 'sessions_cloud_total', 'value': '线上(云学堂)培训总时数', 'width': 260},
    #         {'label': 'sessions_persons_register', 'value': '月平均在册人数', 'width': 260},
    #         {'label': 'sessions_per_people', 'value': '基地人均培训课时', 'width': ''}
    #     ]
    #
    #     month = self.request.GET.get('month', None)
    #
    #     if month == "":
    #         month = datetime.now().date()
    #     if type(month) == str:
    #         datetime_obj = datetime.strptime(month, "%Y-%m-%d")
    #     else:
    #         datetime_obj = datetime.combine(month, datetime.min.time())
    #     #
    #     current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    #     next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
    #
    #     training_hours_list = TrainingContent.objects.filter(content_begin_date__range=(current_month, next_month),
    #                                                          content_end_date__range=(current_month, next_month),
    #                                                          content_status=True).values(
    #         'content_part__department_name', 'content_part__department_code',
    #         'content_part__department_first_name', 'content_part__department_second_name',
    #         'content_part__department_third_name', 'content_part__department_forth_name','content_part__department_level', 'content_part', 'content_duration', 'content_people_number',
    #         'content_title', 'content_begin_date', 'content_lecturer_id', 'content_object')
    #
    #     # 总对象
    #     training_hours_total_obj = {}
    #     # 部门分布对象
    #     training_hours_obj = {
    #         'index': '',
    #         'department_name': '',
    #         'content_part__department_first_name': '',
    #         'content_part__department_second_name': '',
    #         'content_part__department_third_name': '',
    #         'content_part__department_forth_name':'',
    #
    #         'content_duration': 0,  # 培训时长
    #         'content_people_number': 0,  # 参训人数,
    #         'total_training_hours': 0,  # 线下培训总时数
    #     }
    #
    #
    #     all_dept_second=HrDepartment.objects.filter( ~Q(id=999999),Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_status=1,department_level=3).values_list('department_name',flat=True)   #二级部门
    #     # print(all_dept)
    #     if training_hours_list:
    #         for training_hours in list(training_hours_list):
    #             department_code = training_hours['content_part__department_code']
    #             content_title = training_hours['content_title']
    #             content_begin_date=training_hours['content_begin_date']
    #             content_people_number=training_hours['content_people_number']  #参训人数
    #
    #             code_title=str(department_code)+str(str(time.time())[:8]+str(content_title)+str(content_begin_date))+str(content_people_number)
    #             if department_code not in training_hours_total_obj:
    #                 training_hours_total_obj[code_title] = training_hours_obj.copy()
    #                 training_hours_total_obj[code_title]['department_name'] = training_hours['content_part__department_name']
    #                 training_hours_total_obj[code_title]['content_part__department_first_name'] = training_hours['content_part__department_first_name']
    #                 training_hours_total_obj[code_title]['content_part__department_second_name'] = training_hours['content_part__department_second_name']
    #                 training_hours_total_obj[code_title]['content_part__department_third_name'] = training_hours['content_part__department_third_name']
    #                 training_hours_total_obj[code_title]['content_part__department_forth_name'] = training_hours['content_part__department_forth_name']
    #                 training_hours_total_obj[code_title]['content_duration'] = float(training_hours['content_duration'])
    #                 training_hours_total_obj[code_title]['content_people_number'] = float(training_hours['content_people_number'])
    #                 training_hours_total_obj[code_title]['total_training_hours'] = float(training_hours['content_duration'])*float(training_hours['content_people_number'])
    #
    #     dept_data_list=[]
    #     training_hours_total_obj = json.loads(json.dumps(training_hours_total_obj))
    #
    #     tableList=[]
    #     for code_title, line_data in training_hours_total_obj.items():
    #         line_data['code'] = code_title
    #         dept_data_list.append(line_data)
    #     # print(dept_data_list)
    #     for base in all_dept_second:
    #         # 使用列表推导筛选出符合条件的键值对
    #         filtered_dict = [item['total_training_hours'] for item in dept_data_list if item['content_part__department_second_name'] == base]
    #
    #         training_hours_obj_true = {
    #             'sessions_base': base,
    #             'sessions_offline_total': sum(filtered_dict),  # 线下培训总时数
    #             'sessions_record_time': str(current_month)[:10],  # 记录时间
    #         }
    #         # print(training_hours_obj_true)
    #
    #         TrainingSessions.objects.update_or_create(defaults=training_hours_obj_true,
    #                                                   sessions_base=training_hours_obj_true['sessions_base'],
    #                                                   sessions_record_time=training_hours_obj_true[
    #                                                       'sessions_record_time'])
    #
    #         tableList = list(TrainingSessions.objects.filter(sessions_record_time=str(current_month)[:10], sessions_base__in = all_dept_second).values('id',
    #                                                                       'sessions_base',
    #                                                                       'sessions_offline_total',
    #                                                                       "sessions_cloud_total",
    #                                                                       'sessions_persons_register',
    #                                                                       'sessions_per_people',
    #                                                                       'sessions_record_time'))
    #         for line in tableList:
    #             if line['sessions_cloud_total'] is None:
    #                 line['sessions_cloud_total'] = 0
    #             if line['sessions_persons_register'] is None:
    #                 line['sessions_persons_register'] = 0
    #             if line['sessions_per_people'] is None:
    #                 line['sessions_per_people'] = 0
    #             if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
    #                 'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
    #                 try:
    #                     line['sessions_per_people'] = round(
    #                         (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
    #                             'sessions_persons_register'], 2)
    #                 except:
    #                     line['sessions_per_people'] = 0
    #             # print(line)
    #             TrainingSessions.objects.update_or_create(defaults=line, sessions_base=line['sessions_base'],
    #                                                       sessions_record_time=line['sessions_record_time'])
    #
    #     self.return_data = {
    #         "code": status.HTTP_200_OK,
    #         "msg": "信息返回成功",
    #         "data": {
    #             'columnList': columnList,
    #             'tableList': tableList,
    #         }
    #     }
    #



    def edit_month_Training_hours_per_person(self):
        info = json.loads(self.request.body)
        TrainingSessions.objects.filter(pk=info['id']).update(**info)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }
    def download_month_Training_hours_per_person(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        self.month_Training_hours_per_person()
        # file_ls = [
        #     "序号", "中心/基地", "线下培训总时数", "线上(云学堂)培训总时数", "月平均在册人数", "基地人均培训课时"
        # ]
        columnList = self.return_data['data']['columnList']

        file_ls = [line['value'] for line in columnList]

        file_ls.insert(0, '序号')
        path = self.createExcelPath('基地人均培训课时.xlsx', str(t2), '线下培训报表', 40, 'A1:F1', *file_ls)


        order_ls=[line['label'] for line in columnList]

        tableList = self.sort_list_of_dicts_per_person(self.return_data['data']['tableList'],order_ls=order_ls)
        row_data = []
        index = 1

        for line in tableList:
            line_data=[]
            for k,v in line.items():
                if k not in ('id', 'sessions_record_time'):
                    line_data.append(v)
            line_data.insert(0, index)
            row_data.append(line_data)
            if len(line_data) == 0:
                index = index
            index += 1
        exc = openpyxl.load_workbook(path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }


    # def month_summary_analysis(self):
    #     columnList = [
    #         {'value': '润阳集团', 'children': [
    #             {'label': 'department', 'value': '基地', 'width': 230}]
    #          },
    #         {'label': 'category', 'value': '类别', 'width': ''},
    #         {'value': '培训层级(场次)', 'children': [
    #             {'label': 'content_number_senior', 'value': '高层', 'width': 130},
    #             {'label': 'content_number_middle', 'value': '中层', 'width': 130},
    #             {'label': 'content_number_grass', 'value': '基层', 'width': 130},
    #             {'label': 'content_number_synthesis', 'value': '综合', 'width': 130},
    #         ]},
    #         {'value': '总分析', 'children': [
    #             {'label': 'content_number_Total', 'value': '场次', 'width': 130},
    #             {'label': 'content_people_number_sum', 'value': '人次', 'width': 130},
    #             {'label': 'content_duration_sum', 'value': '总时长(H)', 'width': 130},
    #             {'label': 'content_satisfaction_avg', 'value': '平均满意度', 'width': 130},
    #             # {'label': 'content_satisfaction_sum', 'value': '总满意度', 'width': 130},#省略
    #             {'label': 'content_satisfaction_noNone_avg', 'value': '已评分平均满意度', 'width': 130},
    #             {'label': 'content_satisfaction_None_count', 'value': '未评分场次', 'width': 130},
    #         ]},
    #     ]
    #
    #
    #
    #     month = self.request.GET.get('month', None)
    #
    #     if month == "":
    #         month = datetime.now().date()
    #     if type(month) == str:
    #         datetime_obj = datetime.strptime(month, "%Y-%m-%d")
    #     else:
    #         datetime_obj = datetime.combine(month, datetime.min.time())
    #     #
    #     current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    #     next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
    #
    #     second_name_list_base = ['江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司',
    #                              '江苏润阳光伏科技有限公司',
    #                              '江苏海博瑞光伏科技有限公司', '润宝电力', '润阳光伏科技（泰国）有限公司',
    #                              '宁夏润阳硅材料科技有限公司', '江苏润阳光伏科技有限公司（二期）',
    #                              '润阳泰国四期组件', '润阳泰国四期电池', '云南润阳世纪光伏科技有限公司']  # 集团权限看的
    #     first_name_list_base = ['光伏研究院', '全球财务中心', '人力资源中心', '全球战略供应链管理中心']  # 集团权限看的
    #
    #     month_summary_list = TrainingContent.objects.filter(content_begin_date__range=(current_month,next_month),content_end_date__range=(current_month,next_month),content_status=True).values('content_part__department_name', 'content_part__department_code',
    #         'content_part__department_first_name', 'content_part__department_second_name',
    #         'content_module','content_group','content_part','content_duration','content_people_number','content_title','content_begin_date','content_lecturer_id','content_object',
    #              'content_category__category_name','content_level__level_name','content_duration','content_people_number','content_satisfaction')
    #     # print(month_summary_list)
    #     # 总对象
    #     month_summary_total_obj = {}
    #     # 分布对象
    #     month_summary_obj = {
    #         'index': '',
    #         'department_name': '',
    #         'content_part__department_first_name': '',
    #         'content_part__department_second_name': '',
    #         'content_part__department_third_name': '',
    #         'content_module':'',#模块==三级
    #         'content_group':'',#组 == 四级
    #
    #         'content_category__category_name': '',  #培训类别
    #         'content_level__level_name': '',# 培训层级,
    #         'content_duration':'',#培训时长
    #
    #         # 'content_number_synthesis': 0,   #培训综合数量
    #         # 'content_number_senior': 0,  # 培训高层数量
    #         # 'content_number_middle': 0,  # 培训中层数量
    #         # 'content_number_grass': 0,  # 培训基层数量
    #     }
    #
    #     if month_summary_list:
    #         for training_hours in list(month_summary_list):
    #             department_code = training_hours['content_part__department_code']
    #             content_title = training_hours['content_title']
    #             content_begin_date=training_hours['content_begin_date']
    #             code_title=str(department_code)+str(str(time.time())[:8]+str(content_title)+str(content_begin_date))
    #             if department_code not in month_summary_total_obj:
    #                 month_summary_total_obj[code_title] = month_summary_obj.copy()
    #                 month_summary_total_obj[code_title]['department_name'] = training_hours['content_part__department_name']
    #                 month_summary_total_obj[code_title]['content_part__department_first_name'] = training_hours['content_part__department_first_name']
    #                 month_summary_total_obj[code_title]['content_part__department_second_name'] = training_hours['content_part__department_second_name']
    #                 month_summary_total_obj[code_title]['content_module'] = training_hours['content_module']
    #                 month_summary_total_obj[code_title]['content_group'] = training_hours['content_group']
    #                 month_summary_total_obj[code_title]['content_category__category_name'] = training_hours['content_category__category_name']
    #                 month_summary_total_obj[code_title]['content_level__level_name'] = training_hours['content_level__level_name']
    #                 month_summary_total_obj[code_title]['content_people_number'] = float(training_hours['content_people_number']) if training_hours['content_people_number'] is not None else None
    #                 month_summary_total_obj[code_title]['content_duration'] = float(training_hours['content_duration']) if training_hours['content_duration'] is not None else None
    #                 if training_hours['content_satisfaction'] is not None:
    #                     month_summary_total_obj[code_title]['content_satisfaction'] = float(training_hours['content_satisfaction'])   #评分
    #                 else:
    #                     month_summary_total_obj[code_title]['content_satisfaction'] =None    #未评分
    #
    #
    #     base_ls = list(AdminUser.objects.filter(pk=self.request.check_token).values_list("user_department_employee",flat=True).distinct())
    #
    #     base_name_ls = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                     Q(department_expiry_date__isnull=True) | Q(
    #                                                         department_expiry_date__gt=datetime.now()),
    #                                                     department_second_name__isnull=False,
    #                                                     department_status=1, id__in=base_ls).values_list(
    #         'department_name', flat=True).distinct())
    #
    #     # result_list=[]
    #     if 1 in base_ls:  # 集团公司权限
    #         department_category_counts_1 = {}
    #         first_name_list_base = ['光伏研究院', '全球财务中心', '人力资源中心', '全球战略供应链管理中心']
    #         categories_to_include = ['知识类', '技能类', '态度类']
    #         for department in first_name_list_base:
    #             for category in categories_to_include:
    #                 key = f"{department}_{category}"
    #                 department_category_counts_1[key] = {
    #                     "department": department,
    #                     "category": category,
    #                     "content_number_senior": 0,
    #                     "content_number_middle": 0,
    #                     "content_number_grass": 0,
    #                     "content_number_synthesis": 0,
    #                     "content_number_Total": 0,  # 总场次
    #                     "content_people_number_sum": 0,  # 总人数
    #                     "content_duration_sum": 0,  # 总时长
    #                     "content_satisfaction_sum": 0,  # 总满意度
    #                     "content_satisfaction_avg": 0,  # 平均满意度
    #                     "content_satisfaction_None_count": 0,  # 未评分场次
    #                     "content_satisfaction_noNone_count": 0,  # 已评分场次
    #                     "content_satisfaction_noNone_avg": 0,  # 已经评分平均满意度
    #                     # 已评分场次=总场次-未评分场次    平均满意度=总满意度/场次        已评分平均满意度=总满意度/已经评分场次
    #                 }
    #         for key, line in month_summary_total_obj.items():
    #             department_name = line['content_part__department_first_name']
    #             category_name = line['content_category__category_name']
    #             level_name = line['content_level__level_name']
    #             people_number = line['content_people_number']
    #             duration = line['content_duration']
    #             satisfaction = line['content_satisfaction']
    #             if department_name in first_name_list_base and category_name in categories_to_include:
    #                 department_category_key = f"{department_name}_{category_name}"
    #                 category_counts = department_category_counts_1[department_category_key]
    #                 category_counts["content_people_number_sum"] += people_number
    #                 if satisfaction is None:  # 评分未None:
    #                     category_counts["content_satisfaction_None_count"] += 1  # 未评分场次
    #                 category_counts["content_number_Total"] += 1
    #                 category_counts["content_duration_sum"] += duration
    #                 category_counts["content_satisfaction_sum"] += satisfaction if satisfaction != None else 0
    #                 if level_name == '高层':
    #                     category_counts["content_number_senior"] += 1
    #                 elif level_name == '中层':
    #                     category_counts["content_number_middle"] += 1
    #                 elif level_name == '基层':
    #                     category_counts["content_number_grass"] += 1
    #                 elif level_name == '综合':
    #                     category_counts["content_number_synthesis"] += 1
    #         result_list1 = list(department_category_counts_1.values())
    #         department_category_counts_2 = {}
    #         second_name_list_base = ['江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司',
    #                                  '江苏润阳光伏科技有限公司',
    #                                  '江苏海博瑞光伏科技有限公司', '润宝电力', '润阳光伏科技（泰国）有限公司',
    #                                  '宁夏润阳硅材料科技有限公司', '江苏润阳光伏科技有限公司（二期）',
    #                                  '润阳泰国四期组件', '润阳泰国四期电池', '云南润阳世纪光伏科技有限公司']
    #
    #         # Initialize the department_category_counts dictionary with the specified keys
    #         for department in second_name_list_base:
    #             for category in categories_to_include:
    #                 key = f"{department}_{category}"
    #                 department_category_counts_2[key] = {
    #                     "department": department,
    #                     "category": category,
    #                     "content_number_senior": 0,
    #                     "content_number_middle": 0,
    #                     "content_number_grass": 0,
    #                     "content_number_synthesis": 0,
    #                     "content_number_Total": 0,  # 总场次
    #                     "content_people_number_sum": 0,  # 总人数
    #                     "content_duration_sum": 0,  # 总时长
    #                     "content_satisfaction_sum": 0,  # 总满意度
    #                     "content_satisfaction_avg": 0,  # 平均满意度
    #                     "content_satisfaction_None_count": 0,  # 未评分场次
    #                     "content_satisfaction_noNone_avg": 0,  # 已经评分平均满意度
    #                     "content_satisfaction_noNone_count": 0,  # 已评分场次
    #
    #                     # 已评分场次=总场次-未评分场次    平均满意度=总满意度/场次        已评分平均满意度=总满意度/已经评分场次
    #                 }
    #         # Iterate through the data and calculate the counts
    #         for key, line in month_summary_total_obj.items():
    #             department_name = line['content_part__department_second_name']
    #             category_name = line['content_category__category_name']
    #             level_name = line['content_level__level_name']
    #             people_number = line['content_people_number']
    #             duration = line['content_duration']
    #             satisfaction = line['content_satisfaction']
    #             if department_name in second_name_list_base and category_name in categories_to_include:
    #                 department_category_key = f"{department_name}_{category_name}"
    #                 category_counts = department_category_counts_2[department_category_key]
    #                 category_counts["content_people_number_sum"] += people_number
    #                 if satisfaction is None:  # 评分未None:
    #                     category_counts["content_satisfaction_None_count"] += 1  # 未评分场次
    #                 category_counts["content_number_Total"] += 1
    #                 category_counts["content_duration_sum"] += duration
    #                 category_counts["content_satisfaction_sum"] += satisfaction if satisfaction != None else 0
    #                 if level_name == '高层':
    #                     category_counts["content_number_senior"] += 1
    #                 elif level_name == '中层':
    #                     category_counts["content_number_middle"] += 1
    #                 elif level_name == '基层':
    #                     category_counts["content_number_grass"] += 1
    #                 elif level_name == '综合':
    #                     category_counts["content_number_synthesis"] += 1
    #
    #         result_list2 = list(department_category_counts_2.values())
    #         result_list = result_list1 + result_list2
    #     else:      #部门权限
    #         first_name_list_base_dept = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                                      Q(department_expiry_date__isnull=True) | Q(
    #                                                                          department_expiry_date__gt=datetime.now()),
    #                                                                      department_first_name__isnull=False,
    #                                                                      department_status=1,
    #                                                                      department_first_name__in=first_name_list_base,
    #                                                                      department_second_name__in=base_name_ls,
    #                                                                      ).values_list('department_second_name', flat=True).distinct())
    #         first_name_list_base_dept = [item for item in first_name_list_base_dept if  item is not None and item != '']  # 部门权限看的   排除None和''的
    #
    #         second_name_list_base_dept = list(HrDepartment.objects.filter(~Q(id=999999),
    #                                                                       Q(department_expiry_date__isnull=True) | Q(
    #                                                                           department_expiry_date__gt=datetime.now()),
    #                                                                       department_second_name__isnull=False,
    #                                                                       department_status=1,
    #                                                                       department_second_name__in=second_name_list_base,
    #                                                                       department_third_name__in=base_name_ls).values_list(
    #             'department_third_name', flat=True).distinct())
    #         second_name_list_base_dept = [item for item in second_name_list_base_dept if  item is not None and item != '']  # 部门权限看的
    #
    #
    #
    #         department_category_counts_1 = {}
    #         first_name_list_base = first_name_list_base_dept
    #         categories_to_include = ['知识类', '技能类', '态度类']
    #         for department in first_name_list_base:
    #             for category in categories_to_include:
    #                 key = f"{department}_{category}"
    #                 department_category_counts_1[key] = {
    #                     "department": department,
    #                     "category": category,
    #                     "content_number_senior": 0,
    #                     "content_number_middle": 0,
    #                     "content_number_grass": 0,
    #                     "content_number_synthesis": 0,
    #                     "content_number_Total": 0,  # 总场次
    #                     "content_people_number_sum": 0,  # 总人数
    #                     "content_duration_sum": 0,  # 总时长
    #                     "content_satisfaction_sum": 0,  # 总满意度
    #                     "content_satisfaction_avg": 0,  # 平均满意度
    #                     "content_satisfaction_None_count": 0,  # 未评分场次
    #                     "content_satisfaction_noNone_count": 0,  # 已评分场次
    #                     "content_satisfaction_noNone_avg": 0,  # 已经评分平均满意度
    #                     # 已评分场次=总场次-未评分场次    平均满意度=总满意度/场次        已评分平均满意度=总满意度/已经评分场次
    #                 }
    #         for key, line in month_summary_total_obj.items():
    #             department_name = line['content_part__department_second_name']
    #             category_name = line['content_category__category_name']
    #             level_name = line['content_level__level_name']
    #             people_number = line['content_people_number']
    #             duration = line['content_duration']
    #             satisfaction = line['content_satisfaction']
    #             if department_name in first_name_list_base and category_name in categories_to_include:
    #                 department_category_key = f"{department_name}_{category_name}"
    #                 category_counts = department_category_counts_1[department_category_key]
    #                 category_counts["content_people_number_sum"] += people_number
    #                 if satisfaction is None:  # 评分未None:
    #                     category_counts["content_satisfaction_None_count"] += 1  # 未评分场次
    #                 category_counts["content_number_Total"] += 1
    #                 category_counts["content_duration_sum"] += duration
    #                 category_counts["content_satisfaction_sum"] += satisfaction if satisfaction != None else 0
    #                 if level_name == '高层':
    #                     category_counts["content_number_senior"] += 1
    #                 elif level_name == '中层':
    #                     category_counts["content_number_middle"] += 1
    #                 elif level_name == '基层':
    #                     category_counts["content_number_grass"] += 1
    #                 elif level_name == '综合':
    #                     category_counts["content_number_synthesis"] += 1
    #         result_list1 = list(department_category_counts_1.values())
    #
    #         department_category_counts_2 = {}
    #         second_name_list_base = second_name_list_base_dept
    #         for department in second_name_list_base:
    #             for category in categories_to_include:
    #                 key = f"{department}_{category}"
    #                 department_category_counts_2[key] = {
    #                     "department": department,
    #                     "category": category,
    #                     "content_number_senior": 0,
    #                     "content_number_middle": 0,
    #                     "content_number_grass": 0,
    #                     "content_number_synthesis": 0,
    #                     "content_number_Total": 0,  # 总场次
    #                     "content_people_number_sum": 0,  # 总人数
    #                     "content_duration_sum": 0,  # 总时长
    #                     "content_satisfaction_sum": 0,  # 总满意度
    #                     "content_satisfaction_avg": 0,  # 平均满意度
    #                     "content_satisfaction_None_count": 0,  # 未评分场次
    #                     "content_satisfaction_noNone_count": 0,  # 已评分场次
    #                     "content_satisfaction_noNone_avg": 0,  # 已经评分平均满意度
    #                     # 已评分场次=总场次-未评分场次    平均满意度=总满意度/场次        已评分平均满意度=总满意度/已经评分场次
    #                 }
    #         # Iterate through the data and calculate the counts
    #         for key, line in month_summary_total_obj.items():
    #             department_name = line['content_module']
    #             category_name = line['content_category__category_name']
    #             level_name = line['content_level__level_name']
    #             people_number = line['content_people_number']
    #             duration = line['content_duration']
    #             satisfaction = line['content_satisfaction']
    #             if department_name in second_name_list_base and category_name in categories_to_include:
    #                 department_category_key = f"{department_name}_{category_name}"
    #                 category_counts = department_category_counts_2[department_category_key]
    #                 category_counts["content_people_number_sum"] += people_number
    #                 if satisfaction is None:  # 评分未None:
    #                     category_counts["content_satisfaction_None_count"] += 1  # 未评分场次
    #                 category_counts["content_number_Total"] += 1
    #                 category_counts["content_duration_sum"] += duration
    #                 category_counts["content_satisfaction_sum"] += satisfaction if satisfaction != None else 0
    #                 if level_name == '高层':
    #                     category_counts["content_number_senior"] += 1
    #                 elif level_name == '中层':
    #                     category_counts["content_number_middle"] += 1
    #                 elif level_name == '基层':
    #                     category_counts["content_number_grass"] += 1
    #                 elif level_name == '综合':
    #                     category_counts["content_number_synthesis"] += 1
    #
    #         result_list2 = list(department_category_counts_2.values())
    #         result_list = result_list1 + result_list2
    #
    #
    #     department_totals = {}
    #     # 遍历数据列表，计算每个部门的合计值
    #     for item in result_list:   #计算合计
    #         # print(item)
    #         item['content_satisfaction_noNone_count']=item['content_number_Total']-item['content_satisfaction_None_count']
    #         item['content_satisfaction_avg']=item['content_satisfaction_sum']/item['content_number_Total'] if item['content_number_Total'] !=0 else 0
    #         item['content_satisfaction_noNone_avg']=item['content_satisfaction_sum']/item['content_satisfaction_noNone_count']  if item['content_satisfaction_noNone_count'] !=0 else 0   #已经评分平均满意度
    #         department = item['department']
    #         category = item['category']
    #         # print("item",item)
    #         # 检查部门是否已经在department_totals中
    #         if department in department_totals:
    #             department_totals[department]['content_number_senior'] += item['content_number_senior']
    #             department_totals[department]['content_number_middle'] += item['content_number_middle']
    #             department_totals[department]['content_number_grass'] += item['content_number_grass']
    #             department_totals[department]['content_number_synthesis'] +=  item['content_number_synthesis']
    #             department_totals[department]['content_number_Total'] += (item['content_number_senior'] + item['content_number_middle'] + item['content_number_grass'] + item['content_number_synthesis'])
    #             department_totals[department]['content_people_number_sum'] += item['content_people_number_sum']  #总人数
    #             department_totals[department]['content_duration_sum'] += item['content_duration_sum']  # 总时长
    #             department_totals[department]['content_satisfaction_sum'] += item['content_satisfaction_sum']  #  总满意度
    #             department_totals[department]['content_satisfaction_None_count'] += item['content_satisfaction_None_count']  # 未评分场次
    #             # department_totals[department]["content_satisfaction_noNone_count"] += item['content_satisfaction_noNone_count'],  # 已评分场次
    #             department_totals[department]["content_satisfaction_noNone_count"] =department_totals[department]['content_number_Total']-department_totals[department]['content_satisfaction_None_count'],  # 已评分场次
    #
    #             # department_totals[department]['content_satisfaction_avg'] = department_totals[department]['content_satisfaction_sum']/department_totals[department]['content_number_Total'] if department_totals[department]['content_number_Total'] !=0 else 0       # 平均满意度
    #             department_totals[department]['content_satisfaction_noNone_avg'] = department_totals[department]['content_satisfaction_sum'] / (department_totals[department]['content_number_Total']-department_totals[department]['content_satisfaction_None_count']) if (department_totals[department]['content_number_Total']-department_totals[department]['content_satisfaction_None_count'])!= 0 else 0  #已经评分平均满意度
    #             # print(department_totals[department]['content_satisfaction_noNone_count'],department_totals[department]['content_satisfaction_sum'])
    #         else:
    #             # 如果部门不在department_totals中，则创建一个新的条目
    #             department_totals[department] = {
    #                 'department': department,
    #                 'category': '合计',
    #                 'content_number_senior': item['content_number_senior'],
    #                 'content_number_middle': item['content_number_middle'],
    #                 'content_number_grass': item['content_number_grass'],
    #                 'content_number_synthesis': item['content_number_synthesis'],
    #                 'content_number_Total': item['content_number_senior'] + item['content_number_middle'] + item['content_number_grass'] + item['content_number_synthesis'],
    #                 'content_people_number_sum':item['content_people_number_sum'],
    #                 'content_duration_sum': item['content_duration_sum'],
    #                 'content_satisfaction_sum':item['content_satisfaction_sum'],
    #                 'content_satisfaction_avg':0,
    #                 'content_satisfaction_noNone_avg':0,
    #                 'content_satisfaction_None_count':0,
    #             }
    #
    #     # 将部门合计数据添加到原始数据列表
    #     result_list += list(department_totals.values())
    #     result_list=sorted(result_list,key=self.sorting_key)
    #
    #     #计算集团公司
    #     # 找出所有不同的部门和类别
    #     # departments = second_name_list_base+first_name_list_base
    #     categories = set(entry['category'] for entry in result_list)
    #     if 1 in base_ls:
    #         # 创建润阳集团的字典数据
    #         runyang_group_data = []
    #         for category in categories:
    #             senior_sum = sum(entry['content_number_senior'] for entry in result_list if entry['category'] == category)
    #             middle_sum = sum(entry['content_number_middle'] for entry in result_list if entry['category'] == category)
    #             grass_sum = sum(entry['content_number_grass'] for entry in result_list if entry['category'] == category)
    #             synthesis_sum = sum(entry['content_number_synthesis'] for entry in result_list if entry['category'] == category)
    #             total_sum = sum(entry['content_number_Total'] for entry in result_list if entry['category'] == category)
    #             people_number_sum=sum(entry['content_people_number_sum'] for entry in result_list if entry['category'] == category)
    #             duration_sum=sum(entry['content_duration_sum'] for entry in result_list if entry['category'] == category)
    #             satisfaction_sum=sum(entry['content_satisfaction_sum'] for entry in result_list if entry['category'] == category)
    #             satisfaction_avg=satisfaction_sum/total_sum if total_sum !=0 else 0
    #             satisfaction_None_count=sum(entry['content_satisfaction_None_count'] for entry in result_list if entry['category'] == category)#未评分场次
    #             satisfaction_noNone_count=total_sum-satisfaction_None_count
    #             satisfaction_noNone_avg=satisfaction_sum/satisfaction_noNone_count if  satisfaction_noNone_count !=0 else 0#已评分平均满意度
    #             runyang_group_data.append({
    #                 'department': '润阳集团',
    #                 'category': category,
    #                 'content_number_senior': senior_sum,
    #                 'content_number_middle': middle_sum,
    #                 'content_number_grass': grass_sum,
    #                 'content_number_synthesis': synthesis_sum,
    #                 'content_number_Total': total_sum,
    #                 'content_people_number_sum': people_number_sum,
    #                 'content_duration_sum': duration_sum,
    #                 'content_satisfaction_sum': satisfaction_sum,
    #                 'content_satisfaction_avg': satisfaction_avg,
    #                 'content_satisfaction_None_count': satisfaction_None_count, #未评分
    #                 'content_satisfaction_noNone_count':satisfaction_noNone_count,
    #                 'content_satisfaction_noNone_avg':satisfaction_noNone_avg
    #             })
    #         result_list[:0] = runyang_group_data
    #     else:
    #         pass
    #     result_list = sorted(result_list, key=lambda x: ('润阳集团' not in x['department'], x['department'],['知识类', '技能类', '态度类', '合计'].index(x['category'])))
    #     for line in result_list:
    #         line['content_duration_sum']=round(line['content_duration_sum'],4)
    #         line['content_satisfaction_avg'] = round(line['content_satisfaction_avg'], 4)
    #         line['content_satisfaction_sum'] = round(line['content_satisfaction_sum'], 4)
    #         line['content_satisfaction_noNone_avg'] = round(line['content_satisfaction_noNone_avg'], 4)
    #         if 'content_satisfaction_noNone_count' in line:
    #             del line['content_satisfaction_noNone_count']
    #         if 'content_satisfaction_sum' in line:
    #             del line['content_satisfaction_sum']
    #     # result_list = sorted(result_list, key=self.custom_sort)
    #     # print(result_list)
    #     self.return_data = {
    #         "code": status.HTTP_200_OK,
    #         "msg": "信息返回成功",
    #         "data": {
    #             'columnList': columnList,
    #             'tableList': result_list,
    #         }
    #     }



    def month_summary_analysis(self):    #本月汇总分析
        columnList = [

            # {'value': '润阳集团', 'children': [
            #     {'label': 'department', 'value': '基地', 'width': 230}
            # ]
            #  },
            {'value': '润阳集团', 'children': [

                # {'label': 'first_dept_name', 'value': '培训基地(一级部门)', 'width': 230},
                # {'label': 'second_dept_name', 'value': '培训基地(二级部门)', 'width': 230},
                # {'label': 'third_dept_name', 'value': '培训基地(三级部门)', 'width': 230},
                ]
             },
            {'label': 'category', 'value': '类别', 'width': ''},
            {'value': '培训层级(场次)', 'children': [
                {'label': 'content_number_senior', 'value': '高层', 'width': 100},
                {'label': 'content_number_middle', 'value': '中层', 'width': 100},
                {'label': 'content_number_grass', 'value': '基层', 'width': 100},
                {'label': 'content_number_synthesis', 'value': '综合', 'width': 100},
            ]},
            {'value': '总分析', 'children': [
                {'label': 'content_number_Total', 'value': '场次', 'width': 130},
                {'label': 'content_people_number_sum', 'value': '人次', 'width': 130},
                {'label': 'content_duration_sum', 'value': '总时长(H)', 'width': 130},
                {'label': 'content_satisfaction_avg', 'value': '平均满意度', 'width': 130},
                # {'label': 'content_satisfaction_sum', 'value': '总满意度', 'width': 130},#省略
                {'label': 'content_satisfaction_noNone_avg', 'value': '已评分平均满意度', 'width': 130},
                {'label': 'content_satisfaction_None_count', 'value': '未评分场次', 'width': 130},
                # {'label': 'content_satisfaction_noNone_count', 'value': '已评分场次', 'width': 130},#省略
            ]},
        ]

        dept_level = int(self.request.GET.get('type', 1))
        if dept_level==2:
            columnList[0]['children'].insert(0,{'label': 'first_dept_name', 'value': '培训基地(一级)', 'width': 200})
        elif dept_level==3:
            columnList[0]['children'].insert(0, {'label': 'first_dept_name', 'value': '培训基地(一级)', 'width': 200})
            columnList[0]['children'].insert(1, {'label': 'second_dept_name', 'value': '培训基地(二级)', 'width': 200})
        elif dept_level==4:
            columnList[0]['children'].insert(0, {'label': 'first_dept_name', 'value': '培训基地(一级)', 'width': 200})
            columnList[0]['children'].insert(1, {'label': 'second_dept_name', 'value': '培训基地(二级)', 'width': 200})
            columnList[0]['children'].insert(2, {'label': 'third_dept_name', 'value': '培训基地(三级)', 'width': 200})


        month = self.request.GET.get('month', None)

        if month == "":
            month = datetime.now().date()
        if type(month) == str:
            datetime_obj = datetime.strptime(month, "%Y-%m-%d")
        else:
            datetime_obj = datetime.combine(month, datetime.min.time())
        #
        current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)



        month_summary_list = TrainingContent.objects.filter(content_begin_date__range=(current_month,next_month),content_end_date__range=(current_month,next_month),content_status=True).values('content_part__department_name', 'content_part__department_code','content_part__department_full_name', 'content_part__department_full_code',
            'content_part__department_first_name', 'content_part__department_second_name', 'content_part__department_third_name', 'content_part__department_forth_name',
            'content_part_id','content_duration','content_people_number','content_lecturer_id','content_object',
                 'content_category__category_name','content_level__level_name','content_duration','content_people_number','content_satisfaction')



        user_department_list_employee =list(AdminUser.objects.filter(pk=self.request.check_token).values_list("user_department_employee",flat=True).distinct())
        all_dept = list(HrDepartment.objects.filter(~Q(id=999999), Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()), department_status=1, department_level=dept_level,id__in = user_department_list_employee).values_list('department_full_name',flat=True))  # 一级部门
        department_category_counts_1 = {}
        # first_name_list_base = all_dept
        # print(all_dept)
        dept, first_dept, second_dept, third_dept=None,None,None,None
        categories_to_include = ['知识类', '技能类', '态度类']
        for department in all_dept:
            if dept_level == 2:
                dept, first_dept = department.split()
            elif dept_level == 3:
                dept, first_dept, second_dept = department.split()
            elif dept_level == 4:
                dept, first_dept, second_dept, third_dept = department.split()

            for category in categories_to_include:
                key = f"{department}_{category}"
                department_category_counts_1[key] = {
                    "department": department,
                    "category": category,
                    'first_dept_name':first_dept,
                    'second_dept_name':second_dept,
                    'third_dept_name':third_dept,
                    "content_number_senior": 0,
                    "content_number_middle": 0,
                    "content_number_grass": 0,
                    "content_number_synthesis": 0,
                    "content_number_Total": 0,  # 总场次
                    "content_people_number_sum": 0,  # 总人数
                    "content_duration_sum": 0,  # 总时长
                    "content_satisfaction_sum": 0,  # 总满意度
                    "content_satisfaction_avg": 0,  # 平均满意度
                    "content_satisfaction_None_count": 0,  # 未评分场次
                    "content_satisfaction_noNone_count": 0,  # 已评分场次
                    "content_satisfaction_noNone_avg": 0,  # 已经评分平均满意度
                    # 已评分场次=总场次-未评分场次    平均满意度=总满意度/场次        已评分平均满意度=总满意度/已经评分场次
                }
        # key = ''
        # if dept_level == 2:
        #     key = 'content_part__department_first_name'
        # elif dept_level == 3:
        #     key = 'content_part__department_second_name'
        # elif dept_level == 4:
        #     key = 'content_part__department_third_name'
        # print(department_category_counts_1)
        for line in month_summary_list:
            # print(line)
            name=''
            if dept_level == 2:
                name='集团公司 {}'.format(line['content_part__department_first_name'])
            elif dept_level ==3:
                name='集团公司 {} {}'.format(line['content_part__department_first_name'],line['content_part__department_second_name'])
            elif dept_level ==4:
                name='集团公司 {} {} {}'.format(line['content_part__department_first_name'],line['content_part__department_second_name'],line['content_part__department_third_name'])
            # print(name)

            # department_name = line['content_part__department_full_name']
            department_name=name

            category_name = line['content_category__category_name']
            level_name = line['content_level__level_name']
            people_number = float(line['content_people_number']) if line['content_people_number'] and line['content_people_number'].strip() else 0
            # duration = line['content_duration']
            duration = float(line['content_duration']) if line['content_duration'] and line['content_duration'].strip() else 0
            # satisfaction = line['content_satisfaction']
            satisfaction = float(line['content_satisfaction']) if line['content_satisfaction'] and line['content_satisfaction'].strip() else None

            # month_summary_total_obj[code_title]['content_category__category_name'] = training_hours['content_category__category_name']
            # month_summary_total_obj[code_title]['content_level__level_name'] = training_hours['content_level__level_name']
            # month_summary_total_obj[code_title]['content_people_number'] = float(training_hours['content_people_number']) if training_hours['content_people_number'] is not None else None
            # month_summary_total_obj[code_title]['content_duration'] = float(training_hours['content_duration']) if training_hours['content_duration'] is not None else None


            if department_name in all_dept and category_name in categories_to_include:
                department_category_key = f"{department_name}_{category_name}"
                category_counts = department_category_counts_1[department_category_key]  #一个部门类别对象
                category_counts["content_people_number_sum"] += people_number
                if satisfaction is None:  # 评分未None:
                    category_counts["content_satisfaction_None_count"] += 1  # 未评分场次
                category_counts["content_number_Total"] += 1
                category_counts["content_duration_sum"] += duration
                category_counts["content_satisfaction_sum"] += satisfaction if satisfaction != None else 0
                if level_name == '高层':
                    category_counts["content_number_senior"] += 1
                elif level_name == '中层':
                    category_counts["content_number_middle"] += 1
                elif level_name == '基层':
                    category_counts["content_number_grass"] += 1
                elif level_name == '综合':
                    category_counts["content_number_synthesis"] += 1
        # print(department_category_counts_1.values())
        result_list= list(department_category_counts_1.values())
        # print(result_list)


        department_totals = {}
        # 遍历数据列表，计算每个部门的合计值
        for item in result_list:   #计算合计
            item['content_satisfaction_noNone_count']=item['content_number_Total']-item['content_satisfaction_None_count']  #已评分场次
            item['content_satisfaction_avg']=item['content_satisfaction_sum']/item['content_number_Total'] if item['content_number_Total'] !=0 else 0
            item['content_satisfaction_noNone_avg']=item['content_satisfaction_sum']/item['content_satisfaction_noNone_count']  if item['content_satisfaction_noNone_count'] !=0 else 0   #已经评分平均满意度
            department = item['department']
            category = item['category']

            # # 检查部门是否已经在department_totals中
            if department in department_totals:
                department_totals[department]['content_number_senior'] += item['content_number_senior']
                department_totals[department]['content_number_middle'] += item['content_number_middle']
                department_totals[department]['content_number_grass'] += item['content_number_grass']
                department_totals[department]['content_number_synthesis'] +=  item['content_number_synthesis']
                department_totals[department]['content_number_Total'] += item['content_number_Total']
                department_totals[department]['content_people_number_sum'] += item['content_people_number_sum']  #总人数
                department_totals[department]['content_duration_sum'] += item['content_duration_sum']  # 总时长
                department_totals[department]['content_satisfaction_sum'] += item['content_satisfaction_sum']  #  总满意度
                department_totals[department]['content_satisfaction_avg'] = department_totals[department]['content_satisfaction_sum'] / department_totals[department]['content_number_Total'] if department_totals[department]['content_number_Total'] != 0 else 0  # 平均满意度
                department_totals[department]['content_satisfaction_None_count'] += int(item['content_satisfaction_None_count'])  # 未评分场次
                department_totals[department]["content_satisfaction_noNone_count"]+=item['content_satisfaction_noNone_count']# 已评分场次
                department_totals[department]['content_satisfaction_noNone_avg']=float(department_totals[department]['content_satisfaction_sum'])/float(department_totals[department]["content_satisfaction_noNone_count"]) if float(department_totals[department]["content_satisfaction_noNone_count"]) != 0 else 0     #已经评分平均满意度=总满意度/已经评分场次
            else:
                # 如果部门不在department_totals中，则创建一个新的条目
                department_totals[department] = {
                    'department': department,
                    'category': '合计',
                    'content_number_senior': item['content_number_senior'],
                    'content_number_middle': item['content_number_middle'],
                    'content_number_grass': item['content_number_grass'],
                    'content_number_synthesis': item['content_number_synthesis'],
                    'content_number_Total': item['content_number_Total'],
                    'content_people_number_sum':item['content_people_number_sum'],
                    'content_duration_sum': item['content_duration_sum'],
                    'content_satisfaction_sum':item['content_satisfaction_sum'],
                    'content_satisfaction_avg':0,            #平均满意度
                    'content_satisfaction_noNone_avg':0,     #已评分平均满意度
                    'content_satisfaction_None_count':item['content_satisfaction_None_count'],     #未评分场次
                    'content_satisfaction_noNone_count': item['content_satisfaction_noNone_count'],  # 已评分场次

                }

        # 将部门合计数据添加到原始数据列表
        result_list += list(department_totals.values())
        result_list=sorted(result_list,key=self.sorting_key)

        for line in result_list:
            if dept_level == 2:
                dept, first_dept = line['department'].split()
            elif dept_level == 3:
                dept, first_dept, second_dept = line['department'].split()
            elif dept_level == 4:
                dept, first_dept, second_dept, third_dept = line['department'].split()
            line['first_dept_name']= first_dept
            line['second_dept_name'] = second_dept
            line['third_dept_name'] = third_dept

            line['content_duration_sum']=round(line['content_duration_sum'],4)
            line['content_satisfaction_avg'] = round(line['content_satisfaction_avg'], 4)
            line['content_satisfaction_sum'] = round(line['content_satisfaction_sum'], 4)
            line['content_satisfaction_noNone_avg'] = round(line['content_satisfaction_noNone_avg'], 4)
            if 'content_satisfaction_noNone_count' in line:
                del line['content_satisfaction_noNone_count']
            if 'content_satisfaction_sum' in line:
                del line['content_satisfaction_sum']
            if 'department' in line:
                del line['department']

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': result_list,
            }
        }

    def download_month_summary_analysis(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        import openpyxl
        template_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'template_file','本月汇总分析模板.xlsx')  # 创建文件夹

        import shutil
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2),
                                        '线下培训汇总分析.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)

        self.month_summary_analysis()
        # tableList = self.return_data['data']['tableList']
        row_data = []
        # tableList = sorted(self.return_data['data']['tableList'], key=self.custom_sort)
        # tableList = sorted(tableList, key=lambda x: ('润阳集团' not in x['department'], x['department'],
        #                                                  ['知识类', '技能类', '态度类', '合计'].index(x['category'])))


        tableList = self.sort_list_of_dicts(self.return_data['data']['tableList'])

        for line in tableList:
            line_data = []
            for k, v in line.items():
                line_data.append(v)
            row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件

        # 使用字符串替换将\替换为/
        destination_path = destination_path.replace('\\', '/')
        self.merge_cells(destination_path)
        destination_path = 'static/' + destination_path.split('static/')[1]


        # workbook = openpyxl.load_workbook(destination_path)
        # sheet = workbook['Sheet1']  # 替换为您的工作表名称
        # # 删除 B 列，包括数据和列标题
        # sheet.delete_cols(2)
        # workbook.save(destination_path)
        # workbook.close()


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
            # "downloadUrl":template_path
        }


    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10

        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'offlineTrainingFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path


    @staticmethod
    def merge_cells(file_path):
        from openpyxl import load_workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 从第3行开始，每四行合并column列 即1-3列
        # for row in range(3, sheet.max_row + 1, 4):
        #     start_cell = sheet.cell(row=row, column=1)
        #     end_cell = sheet.cell(row=row + 3, column=3)
        #     sheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
            # sheet.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row,end_column=end_cell.column)
        # 从第3行开始，每四行合并 A 列到 C 列的数据
        for row in range(3, sheet.max_row + 1, 4):
            for col in range(1, 3 + 1):
                start_cell = sheet.cell(row=row, column=col)
                end_cell = sheet.cell(row=row + 3, column=col)
                sheet.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row,
                                  end_column=end_cell.column)


        workbook.save(file_path)

    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        now = arrow.now()
        t = now.format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  #文件后缀

        file_name = f"{fileName}.{file_suffix}"    #文件名称

        file_path = os.path.join('static', 'offlineTrainingFile', 'upload_file', t,path,file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)

    @staticmethod
    def sorting_key(item):   #排序
        department_order = {
            '知识类': 1,
            '技能类': 2,
            '态度类': 3,
            '合计': 4,
        }
        return (item['department'], department_order[item['category']])

    # 定义排序键，按照多个字段依次排序
    @staticmethod
    def custom_sort(item):
        return (
            item['department'],
            ['知识类', '技能类', '态度类', '合计'].index(item['category']),
            item['content_number_senior'],
            item['content_number_middle'],
            item['content_number_grass'],
            item['content_number_synthesis'],
            item['content_number_Total'],
            item['content_people_number_sum'],
            item['content_duration_sum'],
            item['content_satisfaction_avg'],
            item['content_satisfaction_None_count'],
            item['content_satisfaction_noNone_avg']
        )

    @staticmethod
    def sort_list_of_dicts(list_of_dicts):
        """
        对列表中的字典按照指定的顺序排序。

        Args:
        list_of_dicts (list): 包含多个字典的列表。
        order (list): 指定排序顺序的键的列表。

        Returns:
        list: 排序后的列表中的字典。
        """
        # 对列表中的每个字典按照指定的顺序排序
        # 指定排序顺序
        order = [
            'first_dept_name',
            'second_dept_name',
            'third_dept_name',
            # 'department',
            'category',
            'content_number_senior',
            'content_number_middle',
            'content_number_grass',
            'content_number_synthesis',
            'content_number_Total',
            'content_people_number_sum',
            'content_duration_sum',
            'content_satisfaction_avg',
            'content_satisfaction_noNone_avg',
            'content_satisfaction_None_count',

            # 'content_satisfaction_sum',
            # 'content_satisfaction_noNone_count',
        ]
        sorted_list_of_dicts = [dict(sorted(d.items(), key=lambda x: order.index(x[0]))) for d in list_of_dicts]
        return sorted_list_of_dicts

    @staticmethod
    def sort_list_of_dicts_per_person(list_of_dicts,order_ls):  #人均课时排序
        """
        对列表中的字典按照指定的顺序排序。

        Args:
        list_of_dicts (list): 包含多个字典的列表。
        order (list): 指定排序顺序的键的列表。

        Returns:
        list: 排序后的列表中的字典。
        """
        # 对列表中的每个字典按照指定的顺序排序
        # 指定排序顺序
        # order = [
        #
        #     'sessions_base_first',
        #     'sessions_base_second',
        #     'sessions_base_third',
        #
        #     'sessions_offline_total',
        #     'sessions_cloud_total',
        #     'sessions_persons_register',
        #     'sessions_per_people',
        #     'id', 'sessions_record_time','sessions_base',
        #
        # ]
        public_order=['id','sessions_record_time']
        order=order_ls+public_order
        print('order',order)
        sorted_list_of_dicts = [dict(sorted(d.items(), key=lambda x: order.index(x[0]))) for d in list_of_dicts]
        return sorted_list_of_dicts