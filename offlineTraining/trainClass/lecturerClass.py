import json, os,arrow,openpyxl
import time

from django.db.models import Q,F,Sum
from rest_framework import status
from django.http import JsonResponse
from employee.models import HrEmployee, HrEmployeeFiles, HrDepartment, HrJobRank
from rest_framework.response import Response
from datetime import datetime, date

from offlineTraining.models import TrainingLecturer, TrainingLecturerLevel,TrainingContent
from pdss.settings import BASE_DIR
from employee import views
from offlineTraining.trainClass.contentClass import Content

class Lecturer:
    def __init__(self, request,meth):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = meth
        self.methods = {
            'get_lecturer_info': self.get_lecturer_info,
            'edit_lecturer_info': self.edit_lecturer_info,
            'delete_lecturer_info': self.delete_lecturer_info,
            'post_lecturer_info': self.post_lecturer_info,
            'down_lecturer_info':self.down_lecturer_info,
            'get_employee_info':self.get_employee_info,
            'summary_lecturer_info':self.summary_lecturer_info,
            'down_summary_lecturer_info':self.down_summary_lecturer_info,
            'batch_lecturer_info':self.batch_lecturer_info,
            'edit_retired_lecturer_info':self.edit_retired_lecturer_info,
            'get_retired_lecturer_info':self.get_retired_lecturer_info
        }

    def method_center(self):
        self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return JsonResponse(self.return_data)
        self.methods[self.meth]()
        return JsonResponse(self.return_data)


    # 获取信息列表
    def get_lecturer_info(self):
        columnList = [{'label': 'index', 'value': '序号', 'width': 60},
                      # {'label': 'lecturer_people__employee_job_rank__job_rank_name', 'value': '基地', 'width':220},
                      {'label': 'lecturer_people__employee_department__department_name', 'value': '部门', 'width':220},
                      {'label': 'lecturer_level__level_name', 'value': '讲师级别', 'width': 120},
                      {'label': 'lecturer_people__employee_code', 'value': '讲师工号', 'width': 220},
                      {'label': 'lecturer_people__employee_name', 'value': '讲师姓名', 'width': 220},
                      {'label': 'lecturer_people__employee_job_duty__job_duty_name', 'value': '职位', 'width':220},
                      {'label': 'total_duration', 'value': '共计授课时长(H)', 'width': 220},
                      {'label': 'avg_satisfaction', 'value': '平均满意度', 'width': 220},
                      {'label': 'lecturer_remark', 'value': '备注', 'width':''},
                      {'label': 'lecturer_content', 'value': '所授课程', 'width': 500},
                      {'label': 'lecturer_content_count', 'value': '所授课程次数', 'width': 500},
                      ]
        kwargs={
            'lecturer_status': True,
            # 'lecturer_name__isnull':True,   #排除有姓名的（外部讲师）
            'lecturer_people_id__isnull':False,   #关联hr表的
            'lecturer_type':'内部讲师',
            # # 'lecturer_people_id__employee_department__id__in' : self.request.user_department_employee,
            'lecturer_people_id__employee_status': '1',   #在职
            'lecturer_level_id__isnull':False,

        }
        info=json.loads(self.request.body)
        print(info)

        currentPage = info['currentPage'] if info['currentPage'] != "" else 1
        pageSize = info['pageSize'] if info['pageSize'] != "" else 25
        searchName = info['searchName']
        lecturer_level=info['lecturer_level_id']
        kwargs['lecturer_level__in'] = lecturer_level
        employee_base = info['baseNameId']
        begin_date = info.get('beginDate', None)
        end_date = info.get('endDate',None)
        train_kwargs={
            'content_begin_date__gte':datetime(1901, 10, 29, 7, 17, 1, 177) if begin_date is None or len(begin_date) == 0 else begin_date,
            'content_begin_date__lte': datetime(3221, 10, 29, 7, 17, 1, 177) if end_date is None or len(end_date) == 0 else end_date,
            'content_part_id__in':self.request.user_department_employee
        }

        print(train_kwargs)



        kwargs['lecturer_people__employee_department__id__in'] = employee_base
        kwargs = {key: value for key, value in kwargs.items() if value is not None and value != ''and value!=[] }  # 过滤掉值为None或''的项
        totalNumber = TrainingLecturer.objects.filter(Q(lecturer_people__employee_name__contains=searchName) | Q(lecturer_people__employee_code__contains=searchName),**kwargs).exclude(lecturer_level__level_name='无').count()
        tableList = list(TrainingLecturer.objects.filter(Q(lecturer_people__employee_name__contains=searchName) | Q(lecturer_people__employee_code__contains=searchName),**kwargs).exclude(lecturer_level__level_name='无').values('id',
                'lecturer_people__employee_department__department_name',
                'lecturer_level__level_name',
                'lecturer_people__employee_code',
                'lecturer_people__employee_name',
                'lecturer_people__employee_job_duty__job_duty_name',
              'lecturer_level',
                'lecturer_remark'
                           ).order_by('-lecturer_createTime')[(currentPage - 1) * pageSize:currentPage * pageSize])

        # print(tableList)
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            total_fraction= TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True).aggregate(Sum('content_duration'),Sum('content_satisfaction'))   #共计授课时长,共计满意度
            item['total_duration']=total_fraction['content_duration__sum'] if  total_fraction['content_duration__sum'] is not None else 0
            total_count = TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True).count()  # 场次
            if total_count ==0 or total_fraction['content_satisfaction__sum'] is None:
                item['avg_satisfaction']=0
            else:
                item['avg_satisfaction'] = round(total_fraction['content_satisfaction__sum']/total_count,4)

            content_ls=list(TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True,**train_kwargs).values_list('content_title',flat=True))
            sum_file=''
            for file in content_ls:
                sum_file+=file+','
            item['lecturer_content']=sum_file if len(content_ls) ==0 else sum_file[:-1]
            # if item['lecturer_level__level_name'] is None:
            #     item['lecturer_level__level_name']='无'
            item['lecturer_content_count']=len(content_ls)


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def edit_lecturer_info(self):
        info=json.loads(self.request.body)
        # info['expatriate_modifier_id'] = self.request.check_token
        TrainingLecturer.objects.filter(pk=info['id']).update(lecturer_level_id=info['lecturer_level'],lecturer_remark=info['lecturer_remark'],lecturer_modifier_id=self.request.check_token)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }

    def delete_lecturer_info(self):
        try:
            id_list = json.loads(self.request.body).get('idList')
            for id in id_list:
                TrainingLecturer.objects.filter(pk=id).update(lecturer_status=False,lecturer_modifier_id=self.request.check_token)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功！"
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def post_lecturer_info(self):
        info=self.request.POST.get('createData' ,None)
        info = json.loads(info)
        employee_obj=HrEmployee.objects.filter(employee_code=info['code'],employee_name=info['name'],employee_status=1).values('id','employee_code','employee_name')
        if employee_obj.exists():
            lecturer_info = {
                'lecturer_people_id': employee_obj[0]['id'],
                'lecturer_level_id': info['lecturer_level_id'],
                'lecturer_remark': info['lecturer_remark'],
                'lecturer_type': '内部讲师',
                'lecturer_status': True,
                'lecturer_creater_id':self.request.check_token
            }
            TrainingLecturer.objects.update_or_create(defaults=lecturer_info,lecturer_people_id=employee_obj[0]['id'],lecturer_type='内部讲师',lecturer_status=True,lecturer_creater_id=self.request.check_token)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "新增成功",
            }
        else:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "该人员不存在或已离职,无法创建",
            }

    def down_lecturer_info(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        file_ls = [
            "序号", "部门", "讲师级别", "讲师工号", "讲师姓名","职位","共计授课时长(H)","平均满意度","备注", "所授课程","所授课程次数"
        ]
        path = self.createExcelPath('讲师明细.xlsx', str(t2), '讲师明细表', 40, 'A1:G1', *file_ls)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        info = json.loads(self.request.body)
        begin_date = info.get('beginDate', None)
        end_date = info.get('endDate',None)
        train_kwargs={
            'content_begin_date__gte':datetime(1901, 10, 29, 7, 17, 1, 177) if begin_date is None or len(begin_date) == 0 else begin_date,
            'content_begin_date__lte': datetime(3221, 10, 29, 7, 17, 1, 177) if end_date is None or len(end_date) == 0 else end_date
        }


        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            kwargs = {
                'lecturer_status': True,
                'lecturer_name__isnull': True,  # 排除有姓名的（外部讲师）
                'lecturer_people_id__isnull': False,  # 关联hr表的
                'lecturer_type': '内部讲师',
                # 'lecturer_people_id__employee_department__id__in' : self.request.user_department_employee,
                'lecturer_people_id__employee_status': '1',  # 在职
                'lecturer_level_id__isnull': False
            }

            searchName = info['searchName']
            lecturer_level = info['lecturer_level_id']
            kwargs['lecturer_level__in'] = lecturer_level

            employee_base = info['baseNameId']
            kwargs['lecturer_people__employee_department__id__in'] = employee_base

            kwargs = {key: value for key, value in kwargs.items() if
                      value is not None and value != '' and value != []}  # 过滤掉值为None或''的项


            tableList = list(TrainingLecturer.objects.filter(Q(lecturer_people__employee_name__contains=searchName) | Q(
                lecturer_people__employee_code__contains=searchName), **kwargs).values('id',
                                                                                       'lecturer_people__employee_department__department_name',
                                                                                       'lecturer_level__level_name',
                                                                                       'lecturer_people__employee_code',
                                                                                       'lecturer_people__employee_name',
                                                                                       'lecturer_people__employee_job_duty__job_duty_name',
                                                                                       'lecturer_level',
                                                                                       'lecturer_remark'
                                                                                       ).exclude(lecturer_level__level_name='无').order_by( '-lecturer_createTime'))
            for index, item in enumerate(tableList):
                total_fraction = TrainingContent.objects.filter(content_lecturer_id=item['id'],
                                                                content_status=True).aggregate(Sum('content_duration'),
                                                                                               Sum('content_satisfaction'))  # 共计授课时长,共计满意度
                item['total_duration'] = total_fraction['content_duration__sum'] if total_fraction[
                                                                                        'content_duration__sum'] is not None else 0
                total_count = TrainingContent.objects.filter(content_lecturer_id=item['id'],
                                                             content_status=True).count()  # 场次
                if total_count == 0 or total_fraction['content_satisfaction__sum'] is None:
                    item['avg_satisfaction'] = 0
                else:
                    item['avg_satisfaction'] = round(total_fraction['content_satisfaction__sum'] / total_count, 4)
                content_ls = list(
                    TrainingContent.objects.filter(content_lecturer_id=item['id'], content_status=True,**train_kwargs).values_list(
                        'content_title', flat=True))
                sum_file = ''
                for file in content_ls:
                    sum_file += file + ','
                item['lecturer_content'] = sum_file if len(content_ls) == 0 else sum_file[:-1]
                item['lecturer_content_count'] = len(content_ls)
                # if item['lecturer_level__level_name'] is None:
                #     item['lecturer_level__level_name'] = '无'
            # print(tableList)
            tableList = self.sort_list_of_dicts(tableList)
            index=1
            for line in tableList:
                line_data = []
                for k, v in line.items():
                    if k not in ('id','lecturer_level'):
                        line_data.append(v)
                line_data.insert(0, index)
                row_data.append(line_data)
                if len(line_data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            tableList = list(TrainingLecturer.objects.filter(id__in=id_list,lecturer_status=True).values('id',
                                                                                       'lecturer_people__employee_department__department_name',
                                                                                       'lecturer_level__level_name',
                                                                                       'lecturer_people__employee_code',
                                                                                       'lecturer_people__employee_name',
                                                                                       'lecturer_people__employee_job_duty__job_duty_name',
                                                                                       'lecturer_level',
                                                                                       'lecturer_remark'
                                                                                       ).exclude(lecturer_level__level_name='无').order_by( '-lecturer_createTime'))
            for index, item in enumerate(tableList):
                total_fraction = TrainingContent.objects.filter(content_lecturer_id=item['id'],
                                                                content_status=True).aggregate(Sum('content_duration'),
                                                                                               Sum('content_satisfaction'))  # 共计授课时长,共计满意度
                item['total_duration'] = total_fraction['content_duration__sum'] if total_fraction[
                                                                                        'content_duration__sum'] is not None else 0
                total_count = TrainingContent.objects.filter(content_lecturer_id=item['id'],
                                                             content_status=True).count()  # 场次
                if total_count == 0 or total_fraction['content_satisfaction__sum'] is None:
                    item['avg_satisfaction'] = 0
                else:
                    item['avg_satisfaction'] = round(total_fraction['content_satisfaction__sum'] / total_count, 4)
                content_ls = list(
                    TrainingContent.objects.filter(content_lecturer_id=item['id'], content_status=True,**train_kwargs).values_list(
                        'content_title', flat=True))
                sum_file = ''
                for file in content_ls:
                    sum_file += file + ','
                item['lecturer_content'] = sum_file if len(content_ls) == 0 else sum_file[:-1]
                item['lecturer_content_count'] = len(content_ls)
                # if item['lecturer_level__level_name'] is None:
                #     item['lecturer_level__level_name'] = '无'
            # print(tableList)
            tableList = self.sort_list_of_dicts(tableList)
            index=1
            for line in tableList:
                line_data = []
                for k, v in line.items():
                    if k not in ('id','lecturer_level'):
                        line_data.append(v)
                line_data.insert(0, index)
                row_data.append(line_data)
                if len(line_data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }


    def get_employee_info(self):
        code=self.request.GET.get('code',None)
        type=self.request.GET.get('type',None)

        TrainingLecturer.objects.filter(Q(lecturer_people__employee_name=code)|Q(lecturer_people__employee_code=code),lecturer_status=1).values('id','lecturer_people__employee_name','lecturer_people__employee_code')

        tableList = list(
            HrEmployee.objects.filter(Q(employee_name__contains=code) | Q(employee_code__contains=code),employee_status=1).values('id',
                                                                                                                'employee_department__department_first_name',
                                                                                                                'employee_department__department_second_name',
                                                                                                                "employee_code",
                                                                                                                "employee_name",
                                                                                                                "employee_job_duty__job_duty_name"))
        # print(tableList)
        for line in tableList:
            # print(line['id'])
            if TrainingLecturer.objects.filter(lecturer_people_id=line['id'],lecturer_status=1).exists():
                # print(True)
                lecturer_level=list(TrainingLecturer.objects.filter(lecturer_people_id=line['id'],lecturer_status=1).values_list('lecturer_level_id',flat=True))
                if len(lecturer_level)>=1:
                    line['lecturer_level'] = lecturer_level[0]
                else:
                    line['lecturer_level'] = ''
            else:
                line['lecturer_level'] =''
            if line['employee_name']==None:
                line['employee_name']=''
            if line['employee_job_duty__job_duty_name']==None:
                line['employee_job_duty__job_duty_name']=''






        if type=='code':
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                'data': [
                    {
                        "value": item['employee_code']+"-"+item["employee_name"]+'-'+item['employee_job_duty__job_duty_name'],
                        "address": {
                            "employee_name": item["employee_name"],  # 姓名
                            'lecturer_code': item['employee_code'],
                            "post": item['employee_job_duty__job_duty_name'],  # 部门名称
                            'lecturer_level_id':item['lecturer_level']
                        }
                    }
                    for item in tableList
                ]

            }
        elif type=='name':
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                'data': [
                    {
                        "value": item['employee_name']+"-"+item["employee_code"]+'-'+item['employee_job_duty__job_duty_name'],
                        "address": {
                            "employee_name": item["employee_name"],  # 姓名
                            'lecturer_code':item['employee_code'],
                            "post": item['employee_job_duty__job_duty_name'],  # 部门名称
                            'lecturer_level_id': item['lecturer_level']
                        }
                    }
                    for item in tableList
                ]

            }

    def batch_lecturer_info(self):
        file = self.request.FILES.get('file')
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.format('YYYY-MM-DD_HH_mm_ss')
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', t1, '讲师库数据文件上传')  # 创建文件夹
        Content.mkdir(dummy_path)
        file_url, file_name, file_suffix = Content.createPath(file, '讲师库数据文件上传','讲师库数据台账' + str(t2))
        Content.saveFile(file_url, file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        for line in range(2, sheet.max_row):  # 每行数据
            lecturer = {}
            if sheet.cell(line + 1, 4).value==None and sheet.cell(line + 1, 3).value==None and sheet.cell(line + 1,2).value==None and sheet.cell(line + 1,5).value==None:  #该行没有数据
                continue
            else:
                if sheet.cell(line + 1, 4).value=='内部讲师':
                    employee_obj = HrEmployee.objects.filter(employee_code=sheet.cell(line + 1, 3).value,employee_status=1).values('id')
                    if employee_obj.exists():
                        lecturer_level_obj=TrainingLecturerLevel.objects.filter(level_name=sheet.cell(line + 1, 5).value).values('id')
                        lecturer = {
                            'lecturer_people_id': employee_obj[0]['id'],
                            'lecturer_level_id':lecturer_level_obj[0]['id'] if lecturer_level_obj.exists() else None,
                            'lecturer_remark':sheet.cell(line + 1, 6).value,
                            'lecturer_type': '内部讲师',
                            'lecturer_status': True,
                            'lecturer_creater_id': self.request.check_token
                        }
                        # print(lecturer)
                        TrainingLecturer.objects.update_or_create(defaults=lecturer, lecturer_people_id=lecturer['lecturer_people_id'],lecturer_type='内部讲师', lecturer_status=True)
                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "上传成功!"
                        }

                    else:
                        self.return_data = {
                            "code": status.HTTP_401_UNAUTHORIZED,
                            "msg": "该讲师已离职,或不是公司员工,无法上传!"
                        }
                        continue

                elif sheet.cell(line+1,4).value=='外部讲师':
                    lecturer = {
                        'lecturer_name': sheet.cell(line + 1, 2).value,
                        'lecturer_remark': sheet.cell(line + 1, 6).value,
                        'lecturer_type': sheet.cell(line+1,4).value,
                        'lecturer_status': True,
                        'lecturer_creater_id': self.request.check_token
                    }
                    # print(lecturer)
                    TrainingLecturer.objects.update_or_create(defaults=lecturer,lecturer_name=lecturer['lecturer_name'], lecturer_type='外部讲师', lecturer_status=True)
                    self.return_data = {
                        "code": status.HTTP_200_OK,
                        "msg": "上传成功!"
                    }
                else:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "讲师类型错误,无法上传!"
                    }
                    continue



    def get_retired_lecturer_info(self):
        columnList = [{'label': 'index', 'value': '序号', 'width': 60},
                      {'label': 'lecturer_people__employee_department__department_name', 'value': '部门', 'width':220},
                      {'label': 'lecturer_level__level_name', 'value': '讲师级别', 'width': 120},
                      {'label': 'lecturer_people__employee_code', 'value': '讲师工号', 'width': 200},
                      {'label': 'lecturer_people__employee_name', 'value': '讲师姓名', 'width': 200},
                      {'label': 'lecturer_people__employee_job_duty__job_duty_name', 'value': '职位', 'width':200},
                      {'label': 'total_duration', 'value': '共计授课时长(H)', 'width': 180},
                      {'label': 'avg_satisfaction', 'value': '平均满意度', 'width': 180},
                      {'label': 'lecturer_people__employee_status', 'value': '是否在职', 'width': 130},
                      {'label': 'lecturer_status', 'value': '是否已卸任', 'width': 130},
                      {'label': 'lecturer_remark', 'value': '备注', 'width':200},
                      {'label': 'lecturer_content', 'value': '所授课程', 'width': ''},]
        kwargs={
            'lecturer_level_id__isnull': False
        }

        info = json.loads(self.request.body)

        currentPage = info['currentPage'] if info['currentPage'] != "" else 1
        pageSize = info['pageSize'] if info['pageSize'] != "" else 25
        searchName = info['searchName']
        lecturer_level = info['lecturer_level_id']
        kwargs['lecturer_level__in'] = lecturer_level

        employee_base = info['baseNameId']
        kwargs['lecturer_people__employee_department__id__in'] = employee_base

        kwargs = {key: value for key, value in kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''的项


        employee_status_choices = {'1': '在职', '2': '离职', '99': '黑名单'}
        totalNumber = TrainingLecturer.objects.filter(Q(lecturer_people__employee_name=searchName) | Q(lecturer_people__employee_code__contains=searchName),Q(lecturer_status=False) | Q(lecturer_people__employee_status='2'),**kwargs).count()
        tableList = list(TrainingLecturer.objects.filter(Q(lecturer_people__employee_name=searchName) | Q(lecturer_people__employee_code__contains=searchName),Q(lecturer_status=False) | Q(lecturer_people__employee_status='2'),**kwargs).values('id',

                  'lecturer_people__employee_department__department_name',
                  'lecturer_level__level_name',
                  'lecturer_people__employee_code',
                  'lecturer_people__employee_name',
                  'lecturer_people__employee_job_duty__job_duty_name',
                    'lecturer_level',
                    'lecturer_remark',
                    "lecturer_status",
                    "lecturer_people__employee_status",
                  ).order_by('-lecturer_createTime')[(currentPage - 1) * pageSize:currentPage * pageSize])
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            total_fraction= TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True).aggregate(Sum('content_duration'),Sum('content_satisfaction'))   #共计授课时长,共计满意度
            item['total_duration']=total_fraction['content_duration__sum'] if  total_fraction['content_duration__sum'] is not None else 0
            total_count = TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True).count()  # 场次
            if total_count ==0 or total_fraction['content_satisfaction__sum'] is None:
                item['avg_satisfaction']=0
            else:
                item['avg_satisfaction'] = round(total_fraction['content_satisfaction__sum']/total_count,4)

            try:
                item['lecturer_people__employee_status']=employee_status_choices.get(item['lecturer_people__employee_status'])
            except:
                pass
            if item['lecturer_status']==False:
                item['lecturer_status']=True
            elif item['lecturer_status']==True:
                item['lecturer_status']=False
            content_ls=list(TrainingContent.objects.filter(content_lecturer_id=item['id'],content_status=True).values_list('content_title',flat=True))
            sum_file=''
            for file in content_ls:
                sum_file+=file+','
            item['lecturer_content']=sum_file if len(content_ls) ==0 else sum_file[:-1]

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def edit_retired_lecturer_info(self):
        try:
            id_list = json.loads(self.request.body).get('idList')
            for id in id_list:
                TrainingLecturer.objects.filter(pk=id).update(lecturer_status=True,lecturer_modifier_id=self.request.check_token)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "恢复任命成功！"
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "恢复任命失败！"
            }



    def summary_lecturer_info(self):
        from offlineTraining.models import TrainingLecturer, TrainingLecturerLevel
        from django.db.models import Count
        tableList = []
        columnList = [
                      {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width':300},
                        {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 300},
                      {'label': 'level_name_rongyu', 'value': '荣誉', 'width':''},
                      {'label': 'level_name_1', 'value': '高级', 'width': ''},
                      {'label': 'level_name_2', 'value': '中级', 'width':''},
                      {'label': 'level_name_3', 'value': '初级', 'width': ''}]


        table_list = []
        summary_lecturer_params = {
            'lecturer_people_id__isnull': False,  # 关联hr表的
            'lecturer_status':True,
            'lecturer_type': '内部讲师',
            'lecturer_level_id__isnull': False,
            'lecturer_people_id__employee_status': '1',  # 在职
        }

        summary_lecturer_params = {key: value for key, value in summary_lecturer_params.items() if
                                      value is not None and value != ''}  # 过滤掉值为None或''的项


        summary_lecturer_list = TrainingLecturer.objects.filter(**summary_lecturer_params).values('lecturer_people_id','lecturer_level__level_name',
                                                                                                  'lecturer_people__employee_department__department_name',
                                                                                                  'lecturer_people__employee_department__department_level',
                                                                                                  'lecturer_people__employee_department__department_code',
                                                                                                  'lecturer_people__employee_department__department_first_name',
                                                                                                  'lecturer_people__employee_department__department_second_name',
                                                                                                  )

        # 总对象
        summary_lecturer_total_obj = {}
        # 部门讲师级别对象
        summary_lecturer_obj = {
            'index': '',
            'department_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'level_name_rongyu': 0,   #荣誉
            'level_name_1': 0,        #高级
            'level_name_2': 0,        #中级
            'level_name_3': 0,        #初级
            # 'total_person': 0,
        }
        # print(summary_lecturer_list)
        if summary_lecturer_list:
            for summary_lecturer in list(summary_lecturer_list):
                department_code = summary_lecturer['lecturer_people__employee_department__department_code']
                department_level=summary_lecturer['lecturer_people__employee_department__department_level']
                code_title=str(department_code)+'_'+str(department_level)
                if department_code not in summary_lecturer_obj:
                    summary_lecturer_total_obj[code_title] = summary_lecturer_obj.copy()
                    summary_lecturer_total_obj[code_title]['department_name'] = summary_lecturer['lecturer_people__employee_department__department_name']
                    summary_lecturer_total_obj[code_title]['employee_department__department_first_name'] = summary_lecturer['lecturer_people__employee_department__department_first_name']
                    summary_lecturer_total_obj[code_title]['employee_department__department_second_name'] = summary_lecturer['lecturer_people__employee_department__department_second_name']

            for summary_lecturer in list(summary_lecturer_list):
                department_code = summary_lecturer['lecturer_people__employee_department__department_code']
                department_level = summary_lecturer['lecturer_people__employee_department__department_level']
                code_title = str(department_code) + '_' + str(department_level)
                lecturer_level__level_name = summary_lecturer['lecturer_level__level_name']
                if lecturer_level__level_name == '荣誉':  # 7天内
                    summary_lecturer_total_obj[code_title]['level_name_rongyu'] += 1
                elif lecturer_level__level_name == '高级':  # 7天内
                    summary_lecturer_total_obj[code_title]['level_name_1'] += 1
                elif lecturer_level__level_name == '中级':  # 7天内
                    summary_lecturer_total_obj[code_title]['level_name_2'] += 1
                elif lecturer_level__level_name == '初级':  # 7天内
                    summary_lecturer_total_obj[code_title]['level_name_3'] += 1
                # summary_lecturer_total_obj[code_title]['total_person'] += 1

        dept_data_list=[]
        summary_lecturer_total_obj = json.loads(json.dumps(summary_lecturer_total_obj))
        # print(summary_lecturer_total_obj)
        for code_title, line_data in summary_lecturer_total_obj.items():
            line_data['code'] = code_title
            dept_data_list.append(line_data)

        # all_dept_first=HrDepartment.objects.filter( ~Q(id=999999),Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_status=1,department_level=2).values_list('department_name',flat=True)   #一级部门
        # print(all_dept_first)
        # all_dept_second=HrDepartment.objects.filter( ~Q(id=999999),Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_status=1,department_level=3).values_list('department_name',flat=True)   #二级部门
        # print(all_dept_second)
        #
        # dept_data_list_new={}


        data = dept_data_list
        # 创建一个字典来跟踪每个组合的总和
        sum_dict = {}

        for row in data:
            department_first_name = row['employee_department__department_first_name']
            department_second_name = row['employee_department__department_second_name']
            level_name_rongyu = row['level_name_rongyu']
            level_name_1 = row['level_name_1']
            level_name_2 = row['level_name_2']
            level_name_3 = row['level_name_3']

            # 使用组合的名称作为键，将各级别的值累加到字典中
            key = (department_first_name, department_second_name)
            if key not in sum_dict:
                sum_dict[key] = {
                    'level_name_rongyu': 0,
                    'level_name_1': 0,
                    'level_name_2': 0,
                    'level_name_3': 0
                }
            sum_dict[key]['level_name_rongyu'] += level_name_rongyu
            sum_dict[key]['level_name_1'] += level_name_1
            sum_dict[key]['level_name_2'] += level_name_2
            sum_dict[key]['level_name_3'] += level_name_3

        # 将结果保存在一个新的列表中
        result = []
        for key, values in sum_dict.items():
            department_first_name, department_second_name = key
            result.append({
                'employee_department__department_first_name': department_first_name,
                'employee_department__department_second_name': department_second_name,
                'level_name_rongyu': values['level_name_rongyu'],
                'level_name_1': values['level_name_1'],
                'level_name_2': values['level_name_2'],
                'level_name_3': values['level_name_3']
            })


        # 按照指定顺序排列的参考列表
        order = [
            None,'全球战略供应链管理中心', '光伏研究院', '集团质量中心', '组件事业部',
            '全球组件营销中心', '战略与运营管理中心', '全球财务中心', '全球电池营销中心',
            '人力资源中心', '硅料事业部', '电池事业一部', '电池事业二部', '电站事业部',
            '润阳越南项目', '工程事业部'
        ]

        # 使用自定义排序函数来排序原始列表
        tableList =sorted(result, key=lambda x: order.index(x['employee_department__department_first_name']) if x['employee_department__department_first_name'] in order else len(order))
        for index, item in enumerate(tableList):
            if item['employee_department__department_first_name']== None and item['employee_department__department_second_name']==None:
                item['employee_department__department_first_name']='集团公司'



        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            'tableList':tableList,
            'columnList':columnList
        }

    def down_summary_lecturer_info(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)

        file_ls = [
            "一级部门", "二级部门", "荣誉", "高级", "中级", "初级",
        ]
        path = self.createExcelPath('讲师分析.xlsx', str(t2), '讲师分析报表', 25, 'A1:F1', *file_ls)
        self.summary_lecturer_info()
        tableList = self.return_data['tableList']
        row_data = []
        for line in tableList:
            line_data = []
            for k, v in line.items():
                line_data.append(v)
            row_data.append(line_data)

        exc = openpyxl.load_workbook(path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }




    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 30
        sheet.column_dimensions['A'].width = 10
        # sheet.column_dimensions['K'].width = 40

        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'offlineTrainingFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass


    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)
    @staticmethod
    def sort_list_of_dicts(list_of_dicts):
        """
        对列表中的字典按照指定的顺序排序。

        Args:
        list_of_dicts (list): 包含多个字典的列表。
        order (list): 指定排序顺序的键的列表。

        Returns:
        list: 排序后的列表中的字典。
        """
        # 对列表中的每个字典按照指定的顺序排序
        # 指定排序顺序
        order = [
            'lecturer_people__employee_department__department_name',
            'lecturer_level__level_name',
            'lecturer_people__employee_code',
            'lecturer_people__employee_name',
            'lecturer_people__employee_job_duty__job_duty_name',
            'total_duration',
            'avg_satisfaction',
            'lecturer_remark',
            'lecturer_content',
            'lecturer_content_count',
            'lecturer_level','id'
        ]
        sorted_list_of_dicts = [dict(sorted(d.items(), key=lambda x: order.index(x[0]))) for d in list_of_dicts]
        return sorted_list_of_dicts

# {
#     "job_rank_name": "内蒙古硅材料",
#     "level_name_rongyu": "仪表车间（内蒙古硅材料）",
#     "level_name_1": "三级",
#     "level_name_2": "7010000495",
#     "level_name_3": "聂宝升",
# },



# {
#     "job_rank_name": "内蒙古硅材料",
#     "level_name_rongyu": "仪表车间（内蒙古硅材料）",
#     "level_name_1": "三级",
#     "level_name_2": "7010000495",
#     "level_name_3": "聂宝升",
# },这是返回数据的格式 job_rank_name是合同归属名，level_name_rongyu是荣誉讲师的数量，level_name_1是一级讲师的数量，level_name_2是2级讲师的数量，level_name_3是3级讲师的数量