import json, os, arrow, openpyxl
import random

from django.db.models import Q, F
from rest_framework import status
from django.http import JsonResponse
from django.core import serializers
from rest_framework.status import HTTP_200_OK

from auther.models import AdminUser
# from employee.views import get_user_field
from django.db import models
from employee.models import HrEmployee, HrEmployeeFiles, HrDepartment
from rest_framework.response import Response
from datetime import datetime, date

from offlineTraining.models import *
from pdss.settings import BASE_DIR
from employee import views


def get_trees(data,
              key_column='id',
              parent_column='parent_id',
              child_column='children',

              current_column=None,
              current_path=None
              ):
    """
    :param data: 数据列表
    :param key_column: 主键字段，默认id
    :param parent_column: 父ID字段名，父ID默认从0开始
    :param child_column: 子列表字典名称
    :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
    :param current_path: 当前展开值
    :return: 树结构
    """
    data_dic = {}
    # data_dic.update({341: {'id': 341, 'department_name': '硅料事业部', 'department_parent_id': 1}})
    for d in data:
        data_dic[d.get(key_column)] = d  # 以自己的权限主键为键,以新构建的字典为值,构造新的字典

    # print(data_dic)
    # data_dic={3: {'id': 3, 'department_name': '全球战略供应链管理中心', 'department_parent_id': 1}, 187: {'id': 187, 'department_name': '物流部', 'department_parent_id': 3}, 540: {'id': 540, 'department_name': '物流部（欧美区域）', 'department_parent_id': 187}, 543: {'id': 543, 'department_name': '美国', 'department_parent_id': 540}}
    data_tree_list = []  # 整个数据大列表
    for d_id, d_dic in data_dic.items():
        d_dic['label'] = d_dic.pop('department_name')
        d_dic['value'] = d_id  # Change the key name 'id' to 'value'
        # print(d_id,d_dic)
        pid = d_dic.get(parent_column)  # 取每一个字典中的父id
        if pid == 1:  # 父id=0，就直接加入数据大列表
            data_tree_list.append(d_dic)
        else:  # 父id>0 就加入父id队对应的那个的节点列表
            try:  # 判断异常代表有子节点，增加子节点列表=[]
                data_dic[pid][child_column].append(d_dic)
            except KeyError:
                # if pid in data_dic:
                try:
                    data_dic[pid][child_column] = []
                    data_dic[pid][child_column].append(d_dic)
                except:
                    pass

                # else:
                #     pass
                # # 处理键不存在的情况
                # print(f"键 {pid} 不存在于 data_dic 中")
                # # print(request)
                # request.user_department_employee.append(pid)
                # print(request.user_department_employee)
    return data_tree_list


# def get_trees(data,
#               key_column='id',
#               parent_column='department_parent_id',
#               child_column='children',
#               current_column=None,
#               current_path=None):   #培训类型
#     """
#     :param data: 数据列表
#     :param key_column: 主键字段，默认'id'
#     :param parent_column: 父ID字段名，默认'department_parent_id'
#     :param child_column: 子列表字典名称，默认'children'
#     :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
#     :param current_path: 当前展开值
#     :return: 树结构
#     """
#     data_dic = {}
#     for d in data:
#         d[key_column] = str(d[key_column])  # Ensure the key is a string
#         data_dic[d[key_column]] = d  # Use the specified key_column as the dictionary key
#
#     data_tree_list = []  # 整个数据大列表
#     # print(data_dic)
#     # data_tree_list[str(341)] = {'id': '341', 'department_name': '硅料事业部', 'department_parent_id': 1}
#
#     # data_dic.update({'341': {'id': '341', 'department_name': '硅料事业部', 'department_parent_id': 1}})
#     # print(data_dic)
#     data_dic={'187': {'id': '187', 'department_name': '物流部', 'department_parent_id': 3}, '540': {'id': '540', 'department_name': '物流部（欧美区域）', 'department_parent_id': 187}, '543': {'id': '543', 'department_name': '美国', 'department_parent_id': 540}}
#     for d_id, d_dic in data_dic.items():
#         d_dic['label'] = d_dic.pop('department_name')
#         d_dic['value'] = int(d_id)
#         pid = d_dic.get(parent_column)
#         if pid is None or pid == 0:  # Check for root nodes (e.g., 0)
#             data_tree_list.append(d_dic)
#         else:
#             parent_node = data_dic.get(str(pid))  # Convert pid to a string for dictionary lookup
#             if parent_node:
#                 if child_column not in parent_node:
#                     parent_node[child_column] = []
#                 parent_node[child_column].append(d_dic)
#     print(data_tree_list)
#     return data_tree_list


def get_trees2(data,
               key_column='id',
               parent_column='department_parent_id',
               child_column='children',
               current_column=None,
               current_path=None):  # 培训类型
    """
    :param data: 数据列表
    :param key_column: 主键字段，默认'id'
    :param parent_column: 父ID字段名，默认'department_parent_id'
    :param child_column: 子列表字典名称，默认'children'
    :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
    :param current_path: 当前展开值
    :return: 树结构
    """
    data_dic = {}
    for d in data:
        d[key_column] = str(d[key_column])  # Ensure the key is a string
        data_dic[d[key_column]] = d  # Use the specified key_column as the dictionary key

    data_tree_list = []  # 整个数据大列表
    for d_id, d_dic in data_dic.items():
        d_dic['label'] = d_dic.pop('type_name')
        d_dic['value'] = int(d_id)
        pid = d_dic.get(parent_column)
        if pid is None or pid == 0:  # Check for root nodes (e.g., 0)
            data_tree_list.append(d_dic)
        else:
            parent_node = data_dic.get(str(pid))  # Convert pid to a string for dictionary lookup
            if parent_node:
                if child_column not in parent_node:
                    parent_node[child_column] = []
                parent_node[child_column].append(d_dic)
    return data_tree_list


class Content:
    def __init__(self, request, meth):
        self.request = request
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.timestamp()
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = meth
        self.methods = {
            'get_content_info': self.get_content_info,
            'post_content_info': self.post_content_info,
            'delete_content_info': self.delete_content_info,
            'edit_content_info': self.edit_content_info,
            'download_content_info': self.download_content_info,
            'batch_content_info': self.batch_content_info,  # 批量上传
            'offline_training_options': self.offline_training_options,  # 汇总下拉菜单
            'month_summary_analysis': self.month_summary_analysis,  # 本月汇总分析
            'month_Training_hours_per_person': self.month_Training_hours_per_person,  # 本月人均课时
            "edit_month_Training_hours_per_person": self.edit_month_Training_hours_per_person,
            'download_month_Training_hours_per_person': self.download_month_Training_hours_per_person,
            'download_month_summary_analysis': self.download_month_summary_analysis,
            'del_content_file': self.del_content_file,
            'post_content_file': self.post_content_file,
        }

    def method_center(self):
        if self.request.check_token is None:
            self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问!'}
            return JsonResponse(self.return_data)
        self.methods[self.meth]()
        return JsonResponse(self.return_data)

    # 获取信息列表
    def get_content_info(self):
        columnList = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'content_part__department_first_name', 'value': '一级部门', 'width': 160},
            {'label': 'content_part__department_second_name', 'value': '二级部门', 'width': 260},
            {'label': 'content_part__department_third_name', 'value': '三级部门', 'width': 160},
            {'label': 'content_part__department_forth_name', 'value': '四级部门', 'width': 260},
            # {'label': 'content_module', 'value': '模块', 'width': 160},
            # {'label': 'content_group', 'value': '组', 'width': 160},
            {'label': 'content_title', 'value': '培训主题/课题', 'width': 160},
            {'label': 'content_type__type_name', 'value': '培训类型', 'width': 160},
            {'label': 'content_category__category_name', 'value': '培训类别', 'width': 160},
            {'label': 'content_level__level_name', 'value': '培训层级', 'width': 160},
            {'label': 'content_manner', 'value': '培训方式', 'width': 160},
            {'label': 'content_begin_date', 'value': '开始培训日期', 'width': 160},
            {'label': 'content_end_date', 'value': '截至培训日期', 'width': 160},
            {'label': 'content_duration', 'value': '培训时长(H)', 'width': 160},
            {'label': 'content_object', 'value': '培训对象', 'width': 160},
            {'label': 'content_people_number', 'value': '参训人数', 'width': 160},
            {'label': 'content_lecturer__lecturer_type', 'value': '讲师类型', 'width': 160},
            {'label': 'content_lecturer__lecturer_people__employee_name', 'value': '培训讲师', 'width': 160},
            {'label': 'content_lecturer__lecturer_people__employee_code', 'value': '讲师工号', 'width': 160},
            {'label': 'content_lecturer__lecturer_people__employee_position__position_name', 'value': '讲师岗位',
             'width': 160},
            {'label': 'content_lecturer__lecturer_level__level_name', 'value': '讲师级别', 'width': 160},
            {'label': 'content_satisfaction', 'value': '培训满意度', 'width': 160},
            {'label': 'content_expenses', 'value': '培训费用', 'width': 160},
            {'label': 'content_plan', 'value': '计划内/计划外', 'width': 160},

            {'label': 'createPhoto_num', 'value': '培训照片', 'width': 160},
            {'label': 'createFile_num', 'value': '培训课件', 'width': 160},
            {'label': 'training_satisfaction_file_num', 'value': '培训满意度', 'width': 160},
            {'label': 'signin_sheet_file_num', 'value': '签到表', 'width': 160},
            {'label': 'content_creater__user', 'value': '用户昵称', 'width': 160},
        ]
        kwargs = {"content_status": True}

        info = json.loads(self.request.body)
        currentPage = info['currentPage'] if info['currentPage'] != "" else 1
        pageSize = info['pageSize'] if info['pageSize'] != "" else 25
        searchName = info['searchName']
        beginDate = info['beginDate']
        endDate = info['endDate']
        contentTypeList = info['contentTypeList']
        kwargs['content_type_id__in'] = contentTypeList
        employee_base = info['baseNameId']
        if beginDate != "" and endDate != "":
            kwargs['content_begin_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                beginDate) == 0 else beginDate
            kwargs['content_begin_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                endDate) == 0 else endDate
        print(endDate,type(endDate))

        kwargs['content_part_id__in'] = employee_base

        if employee_base == '':
            kwargs['content_part_id__in'] = self.request.user_department_employee

        kwargs = {key: value for key, value in kwargs.items() if
                  value is not None and value != '' and value != []}  # 过滤掉值为None或''的项
        print(kwargs)

        print(TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_name__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_code__contains=searchName), **kwargs).query)


        totalNumber = TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_name__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_code__contains=searchName), **kwargs).count()

        tableList = list(TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_name__contains=searchName) | Q(
            content_lecturer__lecturer_people__employee_code__contains=searchName), **kwargs).values('id',
                                                                                                     'content_part_id',
                                                                                                     'content_part__department_first_name',
                                                                                                     "content_part__department_second_name",
                                                                                                     "content_part__department_third_name",
                                                                                                     "content_part__department_forth_name",
                                                                                                     'content_module',
                                                                                                     'content_group',
                                                                                                     'content_title',
                                                                                                     'content_type__type_name',
                                                                                                     'content_category__category_name',
                                                                                                     'content_level__level_name',
                                                                                                     'content_manner',
                                                                                                     'content_begin_date',
                                                                                                     'content_end_date',
                                                                                                     'content_duration',
                                                                                                     'content_object',
                                                                                                     'content_people_number',
                                                                                                     'content_type_id',
                                                                                                     'content_category_id',
                                                                                                     'content_level_id',
                                                                                                     'content_lecturer__lecturer_type',
                                                                                                     'content_lecturer__lecturer_name',
                                                                                                     'content_lecturer__lecturer_people__employee_name',
                                                                                                     'content_lecturer__lecturer_people__employee_code',
                                                                                                     'content_lecturer__lecturer_people__employee_position__position_name',
                                                                                                     'content_lecturer__lecturer_level__level_name',
                                                                                                     "content_lecturer__lecturer_level_id",
                                                                                                     'content_lecturer',
                                                                                                     'content_satisfaction',
                                                                                                     'content_expenses',
                                                                                                     'content_plan',
                                                                                                     'content_creater__user').order_by(
            '-content_createTime')[
                         (currentPage - 1) * pageSize:currentPage * pageSize])
        all_id = [item['id'] for item in tableList]
        file_list = list(
            TrainingFiles.objects.filter(training_content_file_id__in=all_id, training_file_status=True).values('id',
                                                                                                                'training_file_name',
                                                                                                                'training_file_type',
                                                                                                                'training_file_url',
                                                                                                                'training_content_file_id'))

        createPhoto = {}  # 培训照片
        createFile = {}  # 培训附件
        signin_sheet_file = {}  # 簽到表
        training_satisfaction_file = {}  # 培訓滿意度

        for item in file_list:  # 查找每份提案对应的文件
            training_content_file_id = item.get('training_content_file_id')
            if item['training_file_type'] == 1:  # 活动照片
                if training_content_file_id not in createPhoto:
                    createPhoto[training_content_file_id] = []
                createPhoto[training_content_file_id].append(item)
            elif item['training_file_type'] == 2:  # 活动方案
                if training_content_file_id not in createFile:
                    createFile[training_content_file_id] = []
                createFile[training_content_file_id].append(item)
            elif item['training_file_type'] == 3:  # 簽到表
                if training_content_file_id not in signin_sheet_file:
                    signin_sheet_file[training_content_file_id] = []
                signin_sheet_file[training_content_file_id].append(item)
            elif item['training_file_type'] == 4:  # 培訓滿意度
                if training_content_file_id not in training_satisfaction_file:
                    training_satisfaction_file[training_content_file_id] = []
                training_satisfaction_file[training_content_file_id].append(item)

        for item in tableList:
            content_id = item.get('id')
            if content_id in createPhoto:
                item['createPhoto_ls'] = createPhoto[content_id]
                item['createPhoto_num'] = len(createPhoto[content_id])
            else:
                item['createPhoto_ls'] = []
                item['createPhoto_num'] = 0
            if content_id in createFile:
                item['createFile_ls'] = createFile[content_id]
                item['createFile_num'] = len(createFile[content_id])
            else:
                item['createFile_ls'] = []
                item['createFile_num'] = 0

            if content_id in signin_sheet_file:
                item['signin_sheet_file_ls'] = signin_sheet_file[content_id]
                item['signin_sheet_file_num'] = len(signin_sheet_file[content_id])
            else:
                item['signin_sheet_file_ls'] = []
                item['signin_sheet_file_num'] = 0
            if content_id in training_satisfaction_file:
                item['training_satisfaction_file_ls'] = training_satisfaction_file[content_id]
                item['training_satisfaction_file_num'] = len(training_satisfaction_file[content_id])
            else:
                item['training_satisfaction_file_ls'] = []
                item['training_satisfaction_file_num'] = 0

        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['content_begin_date'] = item['content_begin_date'] if item['content_begin_date'] is None else str(
                item['content_begin_date'])
            item['content_end_date'] = item['content_end_date'] if item['content_end_date'] is None else str(
                item['content_end_date'])
            if item['content_lecturer__lecturer_people__employee_name'] is None and item[
                'content_lecturer__lecturer_name'] is not None:
                item['content_lecturer__lecturer_people__employee_name'] = item['content_lecturer__lecturer_name']
            print(item['content_part__department_third_name'] is None,item['content_module'] is not None)
            if (item['content_part__department_third_name'] is None or  item['content_part__department_third_name'] =='') and item['content_module'] is not None:
                item['content_part__department_third_name']=item['content_module']
            if (item['content_part__department_forth_name'] is None or  item['content_part__department_forth_name'] =='') and item['content_group'] is not None:
                item['content_part__department_forth_name'] = item['content_group']

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def post_content_info(self):  # 输入内部讲师 讲师姓名  产生工号 岗位 讲师级别      外部讲师 工号 岗位 界别为空
        # try:
        info = self.request.POST.get('createData', None)
        info = json.loads(info)
        if len(info['nodePath']) > 0:
            info['content_part_id'] = info['nodePath'][-1]
        else:
            info['content_part_id'] = None

        if info['content_begin_date'] == '':
            info['content_begin_date'] = None
        if info['content_end_date'] == '':
            info['content_end_date'] = None
        if info['content_satisfaction'] == '':
            info['content_satisfaction'] = None
        file1_ls = self.request.FILES.getlist('createPhoto')  # 培训照片
        file2_ls = self.request.FILES.getlist('createFile')  # 附件
        file3_ls = self.request.FILES.getlist('signinFile')  # 签到表
        file4_ls = self.request.FILES.getlist('satisfactionFile')  # 培训满意度

        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', str(self.t1),
                                  str(info['content_title']))  # 创建文件夹
        self.mkdir(dummy_path)

        if info['lecturer_type'] == '内部讲师':
            employee_obj = HrEmployee.objects.filter(employee_code=info['lecturer_code'],
                                                     employee_name=info['content_lecturer'],
                                                     employee_status=1).values_list('id', flat=True)
            if employee_obj.exists():
                lecturer_info = {
                    'lecturer_people_id': employee_obj[0],
                    'lecturer_type': info['lecturer_type'],
                    'lecturer_level_id': info['lecturer_level_id'],
                    'lecturer_modifier_id': self.request.check_token
                }

                try:
                    lecturer_obj = TrainingLecturer.objects.filter(
                        lecturer_people_id=lecturer_info['lecturer_people_id'], lecturer_status=True)
                    if lecturer_obj.exists():  # 讲师库有该条数据
                        del info['nodePath']
                        del info['content_section']
                        del info['lecturer_type']
                        del info['content_lecturer']
                        del info['lecturer_code']
                        del info['post']
                        del info['lecturer_level_id']
                        del info['content_part']
                        info['content_lecturer_id'] = lecturer_obj[0].id
                        info['content_creater_id'] = self.request.check_token
                        content_obj = TrainingContent.objects.create(**info)
                        for file_obj in file1_ls:
                            file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                          str(random.random())[
                                                                          -5:] + '_培训照片')
                            self.saveFile(file_url, file_obj)  # 保存文件
                            file_kwargs = {
                                'training_file_name': file_name,
                                'training_file_url': file_url,
                                'training_file_type': 1,
                                'training_content_file_id': content_obj.id
                            }

                            file_dbobj = TrainingFiles.objects.create(**file_kwargs)
                        for file_obj in file2_ls:
                            file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                          str(random.random())[
                                                                          -5:] + '_培训课件')
                            self.saveFile(file_url, file_obj)  # 保存文件
                            file_kwargs = {
                                'training_file_name': file_name,
                                'training_file_url': file_url,
                                'training_file_type': 2,
                                'training_content_file_id': content_obj.id
                            }

                            file_dbobj = TrainingFiles.objects.create(**file_kwargs)
                        for file_obj in file3_ls:
                            file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                          str(random.random())[-5:] + '_签到表')
                            self.saveFile(file_url, file_obj)  # 保存文件
                            file_kwargs = {
                                'training_file_name': file_name,
                                'training_file_url': file_url,
                                'training_file_type': 3,
                                'training_content_file_id': content_obj.id
                            }
                            # print('file2', file_kwargs)
                            file_dbobj = TrainingFiles.objects.create(**file_kwargs)

                            try:
                                exc = openpyxl.load_workbook(file_url, data_only=True)
                                sheet = exc.active
                                for line in range(1, sheet.max_row):  # 每行数据
                                    checkin_info = {}
                                    name_code = sheet.cell(line + 1, 1).value
                                    name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")",
                                                                                                          "")
                                    employee_obj = HrEmployee.objects.filter(employee_code=code,
                                                                             employee_status=1).values_list(
                                        'id',
                                        flat=True)
                                    if employee_obj.exists():
                                        checkin_info['checkin_people_id'] = employee_obj[0]
                                        checkin_info['checkin_content_id'] = content_obj.id
                                        checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
                                        checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
                                        checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
                                        checkin_info['checkin_associated_files_id'] = file_dbobj.id
                                        TrainingCheckin.objects.update_or_create(defaults=checkin_info,
                                                                                 checkin_people_id=checkin_info[
                                                                                     'checkin_people_id'],
                                                                                 checkin_content_id=
                                                                                 checkin_info[
                                                                                     'checkin_content_id'],
                                                                                 checkin_time=checkin_info[
                                                                                     'checkin_time'],
                                                                                 checkin_address=checkin_info[
                                                                                     'checkin_address'],
                                                                                 checkin_status=True,

                                                                                 )
                                    else:  # 离职或不存在
                                        continue
                            except:
                                self.return_data = {
                                    "code": status.HTTP_401_UNAUTHORIZED,
                                    "msg": "该签到表不是excel或者其他错误，无法新增",
                                }
                        for file_obj in file4_ls:
                            file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                          str(random.random())[
                                                                          -5:] + '_培训满意度')
                            self.saveFile(file_url, file_obj)  # 保存文件
                            file_kwargs = {
                                'training_file_name': file_name,
                                'training_file_url': file_url,
                                'training_file_type': 4,
                                'training_content_file_id': content_obj.id
                            }
                            file_dbobj = TrainingFiles.objects.create(**file_kwargs)

                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "新增成功",
                        }
                    else:  # 讲师库没有该条数据 那么不可以新增   除非是无
                        if lecturer_info['lecturer_level_id'] == 5:  # 如果是无那么先创建讲师,在新增报表数据
                            lecturer_obj = TrainingLecturer.objects.create(**lecturer_info)
                            del info['nodePath']
                            del info['content_section']
                            del info['lecturer_type']
                            del info['content_lecturer']
                            del info['lecturer_code']
                            del info['post']
                            del info['lecturer_level_id']
                            del info['content_part']
                            info['content_lecturer_id'] = lecturer_obj.id
                            info['content_creater_id'] = self.request.check_token
                            content_obj = TrainingContent.objects.create(**info)
                            for file_obj in file1_ls:
                                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                              str(random.random())[
                                                                              -5:] + '_培训照片')
                                self.saveFile(file_url, file_obj)  # 保存文件
                                file_kwargs = {
                                    'training_file_name': file_name,
                                    'training_file_url': file_url,
                                    'training_file_type': 1,
                                    'training_content_file_id': content_obj.id
                                }

                                file_dbobj = TrainingFiles.objects.create(**file_kwargs)
                            for file_obj in file2_ls:
                                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                              str(random.random())[
                                                                              -5:] + '_培训课件')
                                self.saveFile(file_url, file_obj)  # 保存文件
                                file_kwargs = {
                                    'training_file_name': file_name,
                                    'training_file_url': file_url,
                                    'training_file_type': 2,
                                    'training_content_file_id': content_obj.id
                                }

                                file_dbobj = TrainingFiles.objects.create(**file_kwargs)
                            for file_obj in file3_ls:
                                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                              str(random.random())[-5:] + '_签到表')
                                self.saveFile(file_url, file_obj)  # 保存文件
                                file_kwargs = {
                                    'training_file_name': file_name,
                                    'training_file_url': file_url,
                                    'training_file_type': 3,
                                    'training_content_file_id': content_obj.id
                                }
                                # print('file2', file_kwargs)
                                file_dbobj = TrainingFiles.objects.create(**file_kwargs)

                                try:
                                    exc = openpyxl.load_workbook(file_url, data_only=True)
                                    sheet = exc.active
                                    for line in range(1, sheet.max_row):  # 每行数据
                                        checkin_info = {}
                                        name_code = sheet.cell(line + 1, 1).value
                                        # name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")",
                                        #                                                                       "")
                                        name, code = self.extract_info(name_code)
                                        employee_obj = HrEmployee.objects.filter(employee_code=code,
                                                                                 employee_status=1).values_list(
                                            'id',
                                            flat=True)
                                        if employee_obj.exists():
                                            checkin_info['checkin_people_id'] = employee_obj[0]
                                            checkin_info['checkin_content_id'] = content_obj.id
                                            checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
                                            checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
                                            checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
                                            checkin_info['checkin_associated_files_id'] = file_dbobj.id
                                            TrainingCheckin.objects.update_or_create(defaults=checkin_info,
                                                                                     checkin_people_id=checkin_info[
                                                                                         'checkin_people_id'],
                                                                                     checkin_content_id=
                                                                                     checkin_info[
                                                                                         'checkin_content_id'],
                                                                                     checkin_time=checkin_info[
                                                                                         'checkin_time'],
                                                                                     checkin_address=checkin_info[
                                                                                         'checkin_address'],
                                                                                     checkin_status=True,

                                                                                     )
                                        else:  # 离职或不存在
                                            continue
                                except:
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "该签到表不是excel或者其他错误，无法新增",
                                    }
                            for file_obj in file4_ls:
                                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                                              str(random.random())[
                                                                              -5:] + '_培训满意度')
                                self.saveFile(file_url, file_obj)  # 保存文件
                                file_kwargs = {
                                    'training_file_name': file_name,
                                    'training_file_url': file_url,
                                    'training_file_type': 4,
                                    'training_content_file_id': content_obj.id
                                }
                                file_dbobj = TrainingFiles.objects.create(**file_kwargs)

                            self.return_data = {
                                "code": status.HTTP_200_OK,
                                "msg": "新增成功",
                            }
                        else:
                            self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": "该内部讲师未加入讲师库,无法新增",
                            }
                except:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "该讲师相关信息错误,无法新增",
                    }

            else:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "不是内部员工或已离职,无法新增",
                }
        elif info['lecturer_type'] == '外部讲师':

            lecturer_info = {
                'lecturer_name': info['content_lecturer'],  # 讲师姓名
                'lecturer_type': info['lecturer_type']
            }
            lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer_info,
                                                                     lecturer_type=lecturer_info['lecturer_type'],
                                                                     lecturer_name=lecturer_info['lecturer_name'],
                                                                     lecturer_status=True)
            content_info = {
                "content_lecturer_id": lecturer_obj[0].id,
                'content_title': info['content_title'],
                'content_part_id': info['content_part_id'],
                'content_module': info['content_module'],
                'content_group': info['content_group'],
                'content_type_id': info['content_type_id'],
                'content_category_id': info['content_category_id'],
                'content_level_id': info['content_level_id'],
                'content_manner': info['content_manner'],
                'content_begin_date': info['content_begin_date'],
                'content_end_date': info['content_end_date'],
                'content_duration': info['content_duration'],
                'content_object': info['content_object'],
                'content_plan': info['content_plan'],
                'content_people_number': info['content_people_number'],
                'content_satisfaction': info['content_satisfaction'],
                'content_expenses': info['content_expenses'],
                'content_creater_id': self.request.check_token
            }
            content_obj = TrainingContent.objects.create(**content_info)
            for file_obj in file1_ls:
                TrainingFiles.objects.filter(training_content_file=info['id'], file_type=1).update(
                    is_valid=False)
                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                              str(random.random())[-5:] + '_培训照片')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 1,
                    'training_content_file_id': content_obj.id
                }
                # print('file1',file_kwargs)
                file_dbobj = TrainingFiles.objects.create(**file_kwargs)
            for file_obj in file2_ls:
                TrainingFiles.objects.filter(activity_file_id=info['id'], file_type=2).update(is_valid=False)
                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                              str(random.random())[-5:] + '_培训课件')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 2,
                    'training_content_file_id': content_obj.id
                }
                # print('file2',file_kwargs)
                file_dbobj = TrainingFiles.objects.create(**file_kwargs)
            for file_obj in file3_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                              str(random.random())[-5:] + '_签到表')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 3,
                    'training_content_file_id': content_obj.id
                }
                # print('file3', file_kwargs)
                file_dbobj = TrainingFiles.objects.create(**file_kwargs)
                try:
                    exc = openpyxl.load_workbook(file_url, data_only=True)
                    sheet = exc.active
                    for line in range(1, sheet.max_row):  # 每行数据
                        checkin_info = {}
                        name_code = sheet.cell(line + 1, 1).value
                        # name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")", "")
                        name, code = self.extract_info(name_code)
                        employee_obj = HrEmployee.objects.filter(employee_code=code,
                                                                 employee_status=1).values_list('id', flat=True)
                        if employee_obj.exists():
                            checkin_info['checkin_people_id'] = employee_obj[0]
                            checkin_info['checkin_content_id'] = content_obj.id
                            checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
                            checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
                            checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
                            checkin_info['checkin_associated_files_id'] = file_dbobj.id
                            TrainingCheckin.objects.update_or_create(defaults=checkin_info,
                                                                     checkin_people_id=checkin_info[
                                                                         'checkin_people_id'],
                                                                     checkin_content_id=checkin_info[
                                                                         'checkin_content_id'],
                                                                     checkin_time=checkin_info['checkin_time'],
                                                                     checkin_address=checkin_info[
                                                                         'checkin_address'],
                                                                     checkin_status=True
                                                                     )
                        else:  # 离职或不存在
                            continue
                except:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "该签到表不是excel或者其他错误，无法新增",
                    }
            for file_obj in file4_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
                                                              str(random.random())[-5:] + '_培训满意度')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 4,
                    'training_content_file_id': content_obj.id
                }
                file_dbobj = TrainingFiles.objects.create(**file_kwargs)

            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "新增成功",
            }

        else:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "请填写讲师类型",
            }

    # except:
    #     self.return_data = {
    #         "code": status.HTTP_401_UNAUTHORIZED,
    #         "msg": "数据不完整，新增失败",
    #     }
    def edit_content_info(self):
        info = json.loads(self.request.body)
        print(info)
        content_id = info['id']  # 培训表id
        lecturer_id = info['content_lecturer']  # 讲师表id
        # 内部--内部   内部--外部     外部--内部   外部 --外部
        if 'nodePath' in info:  # 修改中心了
            if len(info['nodePath']) >= 2:
                info['content_part_id'] = info['nodePath'][1]
            elif len(info['nodePath']) ==1:
                info['content_part_id'] = info['nodePath'][0]
            else:
                info['content_part_id'] = None
        else:  # 没有修改
            pass
        print(info)
        if info['content_lecturer__lecturer_type'] == '内部讲师':  # 新数据
            employee_obj = HrEmployee.objects.filter(
                employee_code=info['content_lecturer__lecturer_people__employee_code'],
                employee_name=info['content_lecturer__lecturer_people__employee_name']).values_list('id', flat=True)
            lecturer_info = {
                'lecturer_people_id': employee_obj[0],
                'lecturer_type': info['content_lecturer__lecturer_type'],
                'lecturer_level_id': info['content_lecturer__lecturer_level_id'],
                'lecturer_creater_id': self.request.check_token
            }

            lecturer_obj = TrainingLecturer.objects.filter(lecturer_people_id=lecturer_info['lecturer_people_id'],
                                                           lecturer_status=True)  # 讲师对象

            if lecturer_obj.exists():  # 讲师库有该条数据  级别是无 不修改   级别不是无 不修改
                content_info = {
                    "content_lecturer_id": lecturer_obj[0].id,
                    'content_title': info['content_title'],
                    'content_part_id': info['content_part_id'],
                    'content_module': info['content_module'],
                    'content_group': info['content_group'],
                    'content_type_id': info['content_type_id'],
                    'content_category_id': info['content_category_id'],
                    'content_level_id': info['content_level_id'],
                    'content_manner': info['content_manner'],
                    'content_begin_date': info['content_begin_date'],
                    'content_end_date': info['content_end_date'],
                    'content_duration': info['content_duration'],
                    'content_object': info['content_object'],
                    'content_plan': info['content_plan'],
                    'content_people_number': info['content_people_number'],
                    'content_satisfaction': info['content_satisfaction'],
                    'content_expenses': info['content_expenses'],
                    'content_modifier_id': self.request.check_token
                }
                TrainingContent.objects.filter(pk=content_id).update(**content_info)
            else:  # 讲师库没有该条数据
                if lecturer_info['lecturer_level_id'] == 5:  # 是无创建
                    lecturer_obj = TrainingLecturer.objects.create(**lecturer_info)
                    del info['nodePath']
                    del info['content_section']
                    del info['lecturer_type']
                    del info['content_lecturer']
                    del info['lecturer_code']
                    del info['post']
                    del info['lecturer_level_id']
                    del info['content_part']
                    info['content_lecturer_id'] = lecturer_obj.id
                    info['content_creater_id'] = self.request.check_token
                    content_obj = TrainingContent.objects.create(**info)

                    self.return_data = {
                        "code": status.HTTP_200_OK,
                        "msg": "新增成功",
                    }
                else:  # 不是提示无法新增
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "该内部讲师未加入讲师库,无法新增",
                    }
        elif info['content_lecturer__lecturer_type'] == '外部讲师':  # 新数据是外部讲师
            lecturer_info = {
                'lecturer_name': info['content_lecturer__lecturer_people__employee_name'],  # 讲师姓名
                'lecturer_type': info['content_lecturer__lecturer_type'],  # 讲师类型
                'lecturer_people_id': None,
                'lecturer_creater_id': self.request.check_token
            }
            # print(lecturer_info)
            lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer_info,
                                                                     lecturer_type=lecturer_info['lecturer_type'],
                                                                     lecturer_name=lecturer_info['lecturer_name'])
            # print(lecturer_obj)
            content_info = {
                "content_lecturer_id": lecturer_obj[0].id,
                'content_title': info['content_title'],
                'content_part_id': info['content_part_id'],
                'content_module': info['content_module'],
                'content_group': info['content_group'],
                'content_type_id': info['content_type_id'],
                'content_category_id': info['content_category_id'],
                'content_level_id': info['content_level_id'],
                'content_manner': info['content_manner'],
                'content_begin_date': info['content_begin_date'],
                'content_end_date': info['content_end_date'],
                'content_duration': info['content_duration'],
                'content_object': info['content_object'],
                'content_plan': info['content_plan'],
                'content_people_number': info['content_people_number'],
                'content_satisfaction': info['content_satisfaction'],
                'content_expenses': info['content_expenses'],
                'content_creater_id': self.request.check_token
            }

            TrainingContent.objects.filter(pk=content_id).update(**content_info)
        else:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "请选择讲师类型",
            }

    def delete_content_info(self):
        info = json.loads(self.request.body)
        for id in info['idList']:
            TrainingContent.objects.filter(pk=id).update(content_status=False)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功",
            }

    def download_content_info(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        file_ls = [
            "序号", "中心/基地", "部门", "模块", "组", "培训主题/课题", "培训类型", "培训类别", "培训层级", "培训方式",
            "开始培训日期", "截止培训日期", "培训时长(H)", "培训对象", "参训人数", "讲师类型", "培训讲师", "讲师工号",
            "讲师岗位", "讲师级别", "培训满意度", "培训费用", "计划内/计划外", "用户昵称"
        ]
        path = self.createExcelPath('线下培训报表.xlsx', str(t2), '线下培训报表', 25, 'A1:X1', *file_ls)
        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            index = 1
            kwargs = {"content_status": True}
            info = json.loads(self.request.body)
            searchName = info['searchName']
            beginDate = info['beginDate']
            endDate = info['endDate']
            contentTypeList = info['contentTypeList']
            kwargs['content_type_id__in'] = contentTypeList
            employee_base = info['baseNameId']
            if beginDate != "" and endDate != "":
                kwargs['content_begin_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                    beginDate) == 0 else beginDate
                kwargs['content_begin_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                    endDate) == 0 else endDate
            kwargs['content_part_id__in'] = employee_base
            if employee_base == '':
                kwargs['content_part_id__in'] = self.request.user_department_employee
            kwargs = {key: value for key, value in kwargs.items() if
                      value is not None and value != '' and value != []}  # 过滤掉值为None或''的项

            tableList = TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(
                content_lecturer__lecturer_people__employee_name__contains=searchName) | Q(
                content_lecturer__lecturer_people__employee_code__contains=searchName), **kwargs).values_list(
                'content_part__department_first_name',
                "content_part__department_second_name",
                "content_part__department_third_name",
                "content_part__department_forth_name",
                'content_title',
                'content_type__type_name',
                'content_category__category_name',
                'content_level__level_name',
                'content_manner',
                'content_begin_date',
                'content_end_date',
                'content_duration',
                'content_object',
                'content_people_number',  # 参训人数
                'content_lecturer__lecturer_type',  # 讲师类型
                'content_lecturer__lecturer_name',  # 培训讲师(外）
                'content_lecturer__lecturer_people__employee_name',  # 培训讲师（内）
                'content_lecturer__lecturer_people__employee_code',  # 讲师工号
                'content_lecturer__lecturer_people__employee_position__position_name',
                'content_lecturer__lecturer_level__level_name',
                'content_satisfaction',
                'content_expenses',
                'content_plan',
                'content_creater__user',
            ).order_by('-content_createTime')

            for line in tableList:
                line = list(line)
                line.insert(0, index)
                if line[16] is None:  # 内部讲师
                    line.pop(16)
                elif line[16] is not None:  # 外部讲师
                    line.pop(17)
                row_data.append(line)
                if len(line) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            index = 1
            for id in id_list:
                data = list(TrainingContent.objects.filter(pk=id, content_status=True).values_list(
                    'content_part__department_first_name',
                    "content_part__department_second_name",
                    "content_part__department_third_name",
                    "content_part__department_forth_name",
                    'content_title',
                    'content_type__type_name',
                    'content_category__category_name',
                    'content_level__level_name',
                    'content_manner',
                    'content_begin_date',
                    'content_end_date',
                    'content_duration',
                    'content_object',
                    'content_people_number',  # 参训人数
                    'content_lecturer__lecturer_type',  # 讲师类型
                    'content_lecturer__lecturer_name',  # 培训讲师(外）
                    'content_lecturer__lecturer_people__employee_name',  # 培训讲师（内）
                    'content_lecturer__lecturer_people__employee_code',  # 讲师工号
                    'content_lecturer__lecturer_people__employee_position__position_name',
                    'content_lecturer__lecturer_level__level_name',
                    'content_satisfaction',
                    'content_expenses',
                    'content_plan',
                    'content_creater__user',
                ))[0]
                data = (index,) + data
                data = list(data)
                if data[16] is None:  # 内部讲师
                    data.pop(16)
                elif data[16] is not None:  # 外部讲师
                    data.pop(17)
                row_data.append(data)
                if len(data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    def batch_content_info(self):
        # try:
        file = self.request.FILES.get("file", None)

        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.format('YYYY-MM-DD_HH_mm_ss')
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', t1,
                                  '培训报表文件上传')  # 创建文件夹
        self.mkdir(dummy_path)
        file_url, file_name, file_suffix = self.createPath(file, '培训报表文件上传', '培训报表' + str(t2))
        self.saveFile(file_url, file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        # try:
        for i in range(1, sheet.max_row):  # 每行数据

            content = {}
            lecturer = {}

            # try:
            first_name = None if sheet.cell(i + 1, 2).value == '' else sheet.cell(i + 1, 2).value
            second_name = None if sheet.cell(i + 1, 3).value == '' else sheet.cell(i + 1, 3).value
            third_name = None if sheet.cell(i + 1, 4).value == '' else sheet.cell(i + 1, 4).value
            forth_name = None if sheet.cell(i + 1, 5).value == '' else sheet.cell(i + 1, 5).value
            # print(first_name,second_name,third_name,forth_name)
            if first_name is not None and second_name is not None and third_name is not None and forth_name is not None:  # 四级部门
                content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
                                                                         department_second_name=second_name,
                                                                         department_third_name=third_name,
                                                                         department_forth_name=forth_name).values_list(
                    'id', flat=True)[0]
            elif first_name is not None and second_name is not None and third_name is not None and forth_name is None:  # 三级部门
                content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
                                                                         department_second_name=second_name,
                                                                         department_third_name=third_name, ).values_list(
                    'id', flat=True)[0]
            elif first_name is not None and second_name is not None and third_name is None and forth_name is None:  # 二级部门
                content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
                                                                         department_second_name=second_name).values_list(
                    'id', flat=True)[0]
            elif first_name is not None and second_name is None and third_name is None and forth_name is None:  # 一级部门
                content['content_part_id'] = \
                HrDepartment.objects.filter(department_first_name=first_name).values_list('id', flat=True)[0]
                # print(content)

            # except:
            #     content['content_part_id'] = None

            # content['content_module'] = sheet.cell(i + 1,4).value
            # content['content_group']= sheet.cell(i + 1, 5).value
            content['content_title'] = sheet.cell(i + 1, 6).value  # 课题
            try:
                content['content_type_id'] = \
                TrainingContentType.objects.filter(type_name=sheet.cell(i + 1, 7).value).values_list('id', flat=True)[0]
            except:
                content['content_type_id'] = None

            try:
                content['content_category_id'] = \
                TrainingContentCategory.objects.filter(category_name=sheet.cell(i + 1, 8).value).values_list('id',
                                                                                                             flat=True)[
                    0]
            except:
                content['content_category_id'] = None

            try:
                content['content_level_id'] = \
                TrainingContentLevel.objects.filter(level_name=sheet.cell(i + 1, 9).value).values_list('id', flat=True)[
                    0]
            except:
                content['content_level_id'] = None

            content['content_manner'] = sheet.cell(i + 1, 10).value
            content['content_begin_date'] = sheet.cell(i + 1, 11).value

            if content['content_begin_date'] is not None:
                if type(content['content_begin_date']) == datetime or type(content['content_begin_date']) == date:
                    content['content_begin_date'] = content['content_begin_date']
                elif type(content['content_begin_date']) == str:
                    try:
                        content['content_begin_date'] = datetime.strptime(content['content_begin_date'],
                                                                          "%Y-%m-%d %H:%M:%S")
                    except:
                        content['content_begin_date'] = None
                else:
                    content['content_begin_date'] = None

            content['content_end_date'] = sheet.cell(i + 1, 12).value
            if content['content_end_date'] is not None:
                if type(content['content_end_date']) == datetime or type(content['content_end_date']) == date:
                    content['content_end_date'] = content['content_end_date']
                elif type(content['content_end_date']) == str:
                    try:
                        content['content_end_date'] = datetime.strptime(content['content_end_date'],
                                                                        "%Y-%m-%d %H:%M:%S")
                    except:
                        content['content_end_date'] = None
                else:
                    content['content_end_date'] = None
                    # self.return_data = {
                    #     "code": status.HTTP_401_UNAUTHORIZED,
                    #     "msg": "培训日期格式错误，无法上传，可以将单元格改为日期格式,例如yyyy/m/d h:mm:ss 即 2023/7/17 17:30:59"
                    # }
            # print(type(content['content_end_date']),type(content['content_begin_date']))
            if content['content_end_date'] is not None and content['content_begin_date'] is not None:
                if content['content_end_date'] >= content['content_begin_date']:
                    # 计算时间差
                    time_difference = content['content_end_date'] - content['content_begin_date']
                    # 提取时间差的总秒数
                    total_seconds = time_difference.total_seconds()
                    # 将秒数转换为小时
                    hours_difference = total_seconds / 3600
                    content['content_duration'] = round(hours_difference, 1)
                else:
                    content['content_duration'] = None
                    content['content_end_date'] = None
                    content['content_begin_date'] = None

            else:
                content['content_duration'] = None
            # content['content_duration']=sheet.cell(i + 1, 13).value#培训时长
            content['content_object'] = sheet.cell(i + 1, 14).value
            content['content_people_number'] = sheet.cell(i + 1, 15).value  # 培训人数
            content['content_satisfaction'] = sheet.cell(i + 1, 21).value if sheet.cell(i + 1,
                                                                                        21).value != '' or sheet.cell(
                i + 1, 21).value is not None else None  # 培训满意度
            content['content_expenses'] = sheet.cell(i + 1, 22).value  # 培训费用
            content['content_plan'] = sheet.cell(i + 1, 23).value  # 计划内、计划外

            try:
                content['content_creater_id'] = \
                AdminUser.objects.filter(username=sheet.cell(i + 1, 24).value).values_list('id', flat=True)[0]  # 创建人
            except:
                try:
                    content['content_creater_id'] = self.request.check_token
                except:
                    content['content_creater_id'] = None
            lecturer['lecturer_type'] = sheet.cell(i + 1, 16).value  # 讲师类型
            try:
                lecturer['lecturer_level_id'] = \
                TrainingLecturerLevel.objects.filter(level_name=sheet.cell(i + 1, 20).value).values_list('id',
                                                                                                         flat=True)[
                    0] if sheet.cell(i + 1, 20).value != '' else None  # 讲师级别
            except:
                lecturer['lecturer_level_id'] = 5
                # self.return_data = {
                #     "code": status.HTTP_401_UNAUTHORIZED,
                #     "msg": "讲师级别不在一，二，三，荣誉讲师范围内"
                # }
            if lecturer['lecturer_type'] == '内部讲师':
                try:
                    lecturer['lecturer_people_id'] = \
                    HrEmployee.objects.filter(employee_code=sheet.cell(i + 1, 18).value).values_list('id', flat=True)[
                        0] if sheet.cell(i + 1, 18).value != '' else None
                    lecturer['lecturer_modifier_id'] = self.request.check_token

                    lecturer_obj = TrainingLecturer.objects.filter(lecturer_people_id=lecturer['lecturer_people_id'],
                                                                   lecturer_status=True).first()  # 讲师对象

                    if lecturer_obj:  # 讲师库有该条数据  级别是无 不修改   级别不是无 不修改
                        content['content_lecturer_id'] = lecturer_obj.id
                    else:  # 讲师库没有该条数据
                        if lecturer['lecturer_level_id'] == 5:  # 是无创建
                            print(lecturer)
                            lecturer_obj = TrainingLecturer.objects.create(**lecturer)
                            content['content_lecturer_id'] = lecturer_obj.id
                        else:  # 不是提示无法新增
                            pass
                            # lecturer_obj=None
                            # content['content_lecturer_id'] = None

                    # content['content_lecturer_id'] = lecturer_obj.id
                except:
                    content['content_lecturer_id'] = None
                    # self.return_data = {
                    #     "code": status.HTTP_401_UNAUTHORIZED,
                    #     "msg": "内部讲师中有讲师不是公司员工，无法新增"
                    # }


            elif lecturer['lecturer_type'] == '外部讲师':
                lecturer['lecturer_name'] = sheet.cell(i + 1, 17).value  # 讲师姓名
                lecturer['lecturer_modifier_id'] = self.request.check_token
                lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer,
                                                                         lecturer_type=lecturer['lecturer_type'],
                                                                         lecturer_name=lecturer['lecturer_name'],
                                                                         lecturer_status=True)
                content['content_lecturer_id'] = lecturer_obj[0].id
            else:
                content['content_lecturer_id'] = None

            # try:
            # TrainingContent.objects.update_or_create(defaults=content,content_part_id=content['content_part_id'],content_begin_date=content['content_begin_date'],content_people_number=content['content_people_number'],content_status=True,content_type_id=content['content_type_id'],content_category_id=content['content_category_id'],content_lecturer_id=content['content_lecturer_id'])
            # print("content",i,content)



            print(content)
            if all(value is None for key, value in content.items() if key != 'content_creater_id'):
                pass
            else:
                # try:
                a=TrainingContent.objects.update_or_create(defaults=content, content_part_id=content['content_part_id'],
                                                                content_title=content['content_title'],
                                                             content_begin_date=content['content_begin_date'],
                                                             content_people_number=content['content_people_number'],
                                                             content_status=True,
                                                             content_type_id=content['content_type_id'],
                                                             content_category_id=content['content_category_id'],
                                                             content_lecturer_id=content['content_lecturer_id'])
                print(i,a)
                # except:
                #     self.return_data = {
                #         "code": status.HTTP_200_OK,
                #         "msg": "参数错误,例如部门,开始培训时间,人数等",
                #     }

            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功",
            }
            # # except:
            # #     pass
            # #     self.return_data = {
            # #         "code": status.HTTP_401_UNAUTHORIZED,
            # #         "msg": "上传失败，数据错误",
            # #     }

            # print(self.return_data)

    # except:
    #     self.return_data = {
    #         "code": status.HTTP_401_UNAUTHORIZED,
    #         "msg": "上传异常",
    #     }

    def del_content_file(self):
        info = json.loads(self.request.body)
        file_obj = TrainingFiles.objects.filter(id=info['file_id'], training_file_status=True).first()
        if file_obj:
            TrainingFiles.objects.filter(pk=info['file_id'], training_file_status=True).update(
                training_file_status=False)
            if file_obj.training_file_type == 3:  # 删除文件的同时把文件数据也删除
                TrainingCheckin.objects.filter(checkin_associated_files_id=file_obj.id, checkin_status=True).update(
                    checkin_status=False, checkin_modifier_id=self.request.check_token)

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功",
        }

    def post_content_file(self):
        info = dict(self.request.POST)
        content_id = info['id'][0]
        content_obj = TrainingContent.objects.filter(pk=content_id, content_status=True).first()
        content_title = content_obj.content_title
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', str(self.t1),
                                  str(content_title))  # 创建文件夹
        self.mkdir(dummy_path)
        type = info['field'][0]
        if type == 'createFile_ls':  # 培训附件
            file_ls = self.request.FILES.getlist('file')
            for file_obj in file_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
                                                              str(random.random())[-5:] + '_培训课件')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 2,
                    'training_content_file_id': content_obj.id
                }
                file_dbobj2 = TrainingFiles.objects.create(**file_kwargs)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "附件新增成功",
                    'data': {
                        'id': file_dbobj2.id,
                        'name': file_dbobj2.training_file_name,
                        'url': file_dbobj2.training_file_url,
                    }
                }

        elif type == 'createPhoto_ls':  # 培训照片
            file_ls = self.request.FILES.getlist('file')
            for file_obj in file_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
                                                              str(random.random())[-5:] + '_培训照片')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 1,
                    'training_content_file_id': content_obj.id
                }
                file_dbobj1 = TrainingFiles.objects.create(**file_kwargs)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "照片新增成功",
                    'data': {
                        'id': file_dbobj1.id,
                        'name': file_dbobj1.training_file_name,
                        'url': file_dbobj1.training_file_url,
                    }
                }

        elif type == 'satisfactionFile':  # 满意度
            file_ls = self.request.FILES.getlist('file')
            for file_obj in file_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
                                                              str(random.random())[-5:] + '_培训满意度')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 4,
                    'training_content_file_id': content_obj.id
                }

                file_dbobj4 = TrainingFiles.objects.create(**file_kwargs)

                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "培训满意度新增成功",
                    'data': {
                        'id': file_dbobj4.id,
                        'name': file_dbobj4.training_file_name,
                        'url': file_dbobj4.training_file_url,
                    }
                }
        elif type == 'signinFile':  # 签到表
            file_ls = self.request.FILES.getlist('file')
            for file_obj in file_ls:
                file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
                                                              str(random.random())[-5:] + '_签到表')
                self.saveFile(file_url, file_obj)  # 保存文件
                file_kwargs = {
                    'training_file_name': file_name,
                    'training_file_url': file_url,
                    'training_file_type': 3,
                    'training_content_file_id': content_obj.id,

                }

                file_dbobj3 = TrainingFiles.objects.create(**file_kwargs)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "签到表新增成功",
                    'data': {
                        'id': file_dbobj3.id,
                        'name': file_dbobj3.training_file_name,
                        'url': file_dbobj3.training_file_url,
                    }
                }

                exc = openpyxl.load_workbook(file_url, data_only=True)
                sheet = exc.active
                for line in range(1, sheet.max_row):  # 每行数据
                    checkin_info = {}
                    name_code = sheet.cell(line + 1, 1).value
                    if name_code:
                        # try:
                        #     name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")", "")
                        # except:
                        #     name,code=None,None
                        name, code = self.extract_info(name_code)
                        print(line,name,code)

                        # employee_obj = HrEmployee.objects.filter(employee_code=code, employee_status=1).values_list('id', flat=True)
                        # print('employee_obj',employee_obj,employee_obj.name,employee_obj.code)
                        # if employee_obj.exists():
                        #     checkin_info['checkin_people_id'] = employee_obj[0]
                        #     checkin_info['checkin_content_id'] = content_id
                        #     checkin_info['checkin_time'] = self.to_date(sheet.cell(line + 1, 3).value)
                        #     checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
                        #     checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
                        #     checkin_info['checkin_associated_files_id'] = file_dbobj3.id
                        #     checkin_info['checkin_modifier_id'] = self.request.check_token
                        #     TrainingCheckin.objects.update_or_create(defaults=checkin_info,
                        #                                              checkin_people_id=checkin_info[
                        #                                                  'checkin_people_id'],
                        #                                              checkin_content_id=checkin_info[
                        #                                                  'checkin_content_id'],
                        #                                              checkin_time=checkin_info['checkin_time'],
                        #                                              checkin_address=checkin_info['checkin_address'],
                        #                                              checkin_status=True
                        #                                              )
                        # else:
                        #     continue
                    else:  # 离职或不存在
                        continue

        else:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "附件新增失败",
            }

    @staticmethod
    def to_date(value):
        """
        如果value是日期或者能被转换为日期 返回日期类型 不是返回None
        :param value:
        :return:
        """
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            try:
                date_obj = datetime.strptime(value, '%Y/%m/%d %H:%M:%S')  # 根据日期格式修改这里
                return date_obj
            except ValueError:
                try:
                    print('11111', value, str(value))
                    date_obj = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')  # 另一种日期格式
                    return date_obj
                except ValueError:
                    return None

        return None

    def month_summary_analysis(self):
        from django.utils import timezone
        # print(self.request.GET)
        month = self.request.GET.get('month', None)
        # print('month2',month)
        if month != "":
            month = month
        else:
            month = datetime.now().date()

        columnList = [
            {'value': '润阳集团', 'children': [
                {'label': 'department', 'value': '基地', 'width': 230}]
             },
            {'label': 'category', 'value': '类别', 'width': ''},
            {'value': '培训层级(场次)', 'children': [
                {'label': 'content_number_senior', 'value': '高层', 'width': 130},
                {'label': 'content_number_middle', 'value': '中层', 'width': 130},
                {'label': 'content_number_grass', 'value': '基层', 'width': 130},
                {'label': 'content_number_synthesis', 'value': '综合', 'width': 130},

            ]},
            {'value': '总分析', 'children': [
                {'label': 'content_number_Total', 'value': '场次', 'width': 130},
                {'label': 'content_people_number', 'value': '人次', 'width': 130},
                {'label': 'content_duration', 'value': '总时长(H)', 'width': 130},
                {'label': 'content_satisfaction', 'value': '平均满意度', 'width': 130},
                {'label': 'content_satisfaction_avg', 'value': '已评分平均满意度', 'width': 130},
                {'label': 'count_null_Satisfaction_Total', 'value': '未评分场次', 'width': 130},
            ]},
        ]

        from django.db import connection
        # print(month,type(month))
        # if type(month)==str:
        #     datetime_obj = datetime.strptime(month, "%Y-%m-%d")
        # else:
        #     datetime_obj = datetime.combine(month, datetime.min.time())
        # # print(datetime_obj,type(datetime_obj))
        #
        # # print(datetime_obj,type(datetime_obj))
        # current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
        # print(current_month,next_month)
        # # 构建 SQL 查询语句

        sql_query = """
            SELECT
                IFNULL( dc.department_first_name, 'Total' ) AS base_name,
                IFNULL( cat.category_name, 'Total' ) AS category_name,
                IFNULL( lvl.level_name, 'Total' ) AS level_name,
                COALESCE ( COUNT( tc.id ), 0 ) AS count,
                COALESCE ( SUM( tc.content_people_number ), 0 ) AS total_people,
                COALESCE ( SUM( tc.content_duration ), 0 ) AS total_duration,
                COALESCE ( SUM( tc.content_satisfaction ), 0 ) AS total_satisfaction,
            CASE

                    WHEN SUM(
                    CASE

				WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' ) 
				AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND 
				AND tc.content_status <> 0 
				AND tc.content_satisfaction IS NOT NULL THEN
					1 ELSE 0 
				END 
					) > 0 THEN
					COALESCE (
						SUM(
						CASE

								WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' ) 
								AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND 
								AND tc.content_status <> 0 
								AND tc.content_satisfaction IS NOT NULL THEN
									tc.content_satisfaction ELSE 0 
								END 
									) / SUM(
								CASE

										WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' ) 
										AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND 
										AND tc.content_status <> 0 
										AND tc.content_satisfaction IS NOT NULL THEN
											1 ELSE 0 
										END 
										),
										0 
									) ELSE 0 
								END AS avg_satisfaction,
								SUM(
								CASE

										WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' ) 
										AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND 
										AND tc.content_status <> 0 
										AND tc.content_satisfaction IS NULL THEN
											1 ELSE 0 
										END 
										) AS count_satisfaction_none 
									FROM
										(
										SELECT
											id,
											department_first_name 
										FROM
											hr_department 
										WHERE
											department_first_name IS NOT NULL 
											AND (
												department_expiry_date IS NULL 
											OR department_expiry_date >= NOW()) 
											AND department_first_name IN ('光伏研究院','全球财务中心', '人力资源中心','全球战略供应链管理中心' ) 
										) dc
										CROSS JOIN training_content_category cat
										CROSS JOIN training_content_level lvl
										LEFT JOIN (
										SELECT
											* 
										FROM
											training_content 
										WHERE
											content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' ) 
											AND content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND 
											AND content_status <> 0 
										) tc ON tc.content_part_id = dc.id 
										AND tc.content_category_id = cat.id 
										AND tc.content_level_id = lvl.id 
									GROUP BY
										dc.department_first_name,
									cat.category_name,
	                            lvl.level_name WITH ROLLUP;

        """.format(month, month, month, month, month, month, month, month, month, month)
        # print(sql_query)

        sql_query_second = """

           	            SELECT
            IFNULL(dc.department_second_name, 'Total') AS base_name,
            IFNULL(cat.category_name, 'Total') AS category_name,
            IFNULL(lvl.level_name, 'Total') AS level_name,
            COALESCE(COUNT(tc.id), 0) AS count,
            COALESCE(SUM(tc.content_people_number), 0) AS total_people,
            COALESCE(SUM(tc.content_duration), 0) AS total_duration,
            COALESCE(SUM(tc.content_satisfaction), 0) AS total_satisfaction,
            CASE
                WHEN SUM(
                    CASE
                        WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
                        AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
                        AND tc.content_status <> 0
                        AND tc.content_satisfaction IS NOT NULL
                        THEN 1
                        ELSE 0
                    END
                ) > 0 THEN 
                    COALESCE(SUM(
                        CASE
                            WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
                            AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
                            AND tc.content_status <> 0
                            AND tc.content_satisfaction IS NOT NULL
                            THEN tc.content_satisfaction
                            ELSE 0
                        END
                    ) / SUM(
                        CASE
                            WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
                            AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
                            AND tc.content_status <> 0
                            AND tc.content_satisfaction IS NOT NULL
                            THEN 1
                            ELSE 0
                        END
                    ), 0)
                ELSE
                    0
            END AS avg_satisfaction,
            SUM(
                    CASE
                        WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
                        AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY - INTERVAL 1 SECOND
                        AND tc.content_status <> 0
                        AND tc.content_satisfaction IS NULL
                        THEN 1
                        ELSE 0
                    END
                ) AS count_satisfaction_none

            FROM (
            SELECT id, department_second_name
            FROM hr_department
            WHERE department_second_name IN (
'江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司','江苏润阳光伏科技有限公司',
                '江苏海博瑞光伏科技有限公司', '润宝电力','润阳光伏科技（泰国）有限公司','宁夏润阳硅材料科技有限公司','江苏润阳光伏科技有限公司（二期）',
                '润阳泰国四期组件', '润阳泰国四期电池','云南润阳世纪光伏科技有限公司'
            )
            ) dc
            CROSS JOIN training_content_category cat
            CROSS JOIN training_content_level lvl
            LEFT JOIN (
            SELECT *
            FROM training_content
            WHERE content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00') 
            AND content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
            AND content_status <> 0
            ) tc ON tc.content_part_id = dc.id
            AND tc.content_category_id = cat.id
            AND tc.content_level_id = lvl.id
            GROUP BY
            dc.department_second_name, cat.category_name, lvl.level_name
            WITH ROLLUP;

        """.format(month, month, month, month, month, month, month, month, month, month)

        # 执行 SQL 查询
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            # print(cursor)
            result = cursor.fetchall()

        with connection.cursor() as cursor:
            cursor.execute(sql_query_second)
            # print(cursor)
            result2 = cursor.fetchall()

        result += result2

        tableData = []
        for entry in result:
            department, category, level, content_number, content_people_number, content_duration, content_satisfaction, content_satisfaction_avg, count_satisfaction_none = entry
            tableData.append({
                "department": department,
                "category": category,
                "level": level,
                "content_number": float(content_number),
                "content_people_number": float(content_people_number),
                "content_duration": float(content_duration),
                "content_satisfaction": float(content_satisfaction),
                "content_satisfaction_avg": float(content_satisfaction_avg),
                'count_null_Satisfaction_Total': float(count_satisfaction_none)
            })
        original_list = tableData
        transformed_list = []
        department_category_data = {}
        for item in original_list:
            department = item['department']
            category = item['category']
            level = item['level']
            content_number = item['content_number']

            if department not in department_category_data:
                department_category_data[department] = {}

            if category not in department_category_data[department]:
                department_category_data[department][category] = {
                    'department': department,
                    'category': category,
                    'content_number_中层': 0,
                    'content_number_基层': 0,
                    'content_number_综合': 0,
                    'content_number_高层': 0,
                    'content_people_number': 0.0,
                    'content_duration': 0.0,
                    'content_satisfaction': 0.0,
                    'content_satisfaction_avg': 0.0,
                    'count_null_Satisfaction_Total': 0
                }

            department_category_data[department][category]['content_number_' + level] = content_number
            department_category_data[department][category]['content_people_number'] = item['content_people_number']
            department_category_data[department][category]['content_duration'] = item['content_duration']
            department_category_data[department][category]['content_satisfaction'] = item['content_satisfaction']
            department_category_data[department][category]['content_satisfaction_avg'] = item[
                'content_satisfaction_avg']
            department_category_data[department][category]['count_null_Satisfaction_Total'] = item[
                'count_null_Satisfaction_Total']
        for department_data in department_category_data.values():
            transformed_list.extend(department_data.values())

        for item in transformed_list:
            item['content_number_middle'] = item.pop('content_number_中层', 0)
            item['content_number_grass'] = item.pop('content_number_基层', 0)
            item['content_number_synthesis'] = item.pop('content_number_综合', 0)
            item['content_number_senior'] = item.pop('content_number_高层', 0)
        transformed_list = [d for d in transformed_list if d['department'] != 'Total']
        # print('list',transformed_list)
        # tableList=transformed_list
        # from django.db.models import Count
        # from itertools import product
        #
        # # 获取所有可能的部门和类别组合
        # all_departments = HrDepartment.objects.filter(Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_first_name__isnull=False,
        #     department_status=1).exclude(id=999999).values_list('department_first_name', flat=True).distinct()
        # all_categories = TrainingContentCategory.objects.exclude(id=999999).values_list('category_name', flat=True)
        # #
        # # 查询并统计满意度为空的场次数量
        # if type(month)==str:
        #     datetime_obj = datetime.strptime(month, "%Y-%m-%d")
        # else:
        #     datetime_obj = datetime.combine(month, datetime.min.time())
        #
        # current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
        # # print(current_month,next_month)
        # result = TrainingContent.objects.filter(content_satisfaction__isnull=True,content_begin_date__range=(current_month, next_month),
        # content_end_date__range=(current_month, next_month),content_status=1).values(
        #     'content_part__department_first_name', 'content_category__category_name'
        # ).annotate(num_sessions=Count('id'))
        # print('result',result)
        #
        #
        #
        # # 生成所有可能的部门和类别组合的字典
        # department_category_combinations = list(product(all_departments, all_categories))
        # department_category_dict = {(department, category): 0 for department, category in
        #                             department_category_combinations}
        #
        # # 更新字典中的数量信息
        # for item in result:
        #     department_name = item['content_part__department_first_name']
        #     category = item['content_category__category_name']
        #     num_sessions = item['num_sessions']
        #     department_category_dict[(department_name, category)] = num_sessions
        #
        # # # 生成包含信息的列表
        # result_list = []
        # for department, category in department_category_combinations:
        #     num_sessions = department_category_dict[(department, category)]
        #     result_list.append(
        #         {'department': department, 'category': category, 'count_null_Satisfaction': num_sessions})
        # print(result_list)
        # department_category_sums = {}
        # for entry in result_list:
        #     department = entry['department']
        #     category = entry['category']
        #     # count = entry['count_null_Satisfaction']
        #
        #     if department not in department_category_sums:
        #         department_category_sums[department] = {}
        #
        #     # if category not in department_category_sums[department]:
        #     #     department_category_sums[department][category] = count
        #     # else:
        #     #     department_category_sums[department][category] += count
        # for department, categories in department_category_sums.items():
        #     total_count = sum(categories.values())
        #     categories['Total'] = total_count
        #
        # result = []
        # for department, categories in department_category_sums.items():
        #     for category, count in categories.items():
        #         result.append({'department': department, 'category': category, 'count_null_Satisfaction_Total': count})

        # print('result3333',result)
        # merged_dict = {}
        # for item in result:
        #     key = (item['department'], item['category'])
        #     merged_dict[key] = item

        # for item in transformed_list:
        #     key = (item['department'], item['category'])
        #     if key in merged_dict:
        #         merged_dict[key].update(item)
        #     else:
        #         merged_dict[key] = item
        #
        # tableList = list(merged_dict.values())[:-1]

        middle_sum_by_department = {}
        grass_sum_by_department = {}
        synthesis_sum_by_department = {}
        senior_sum_by_department = {}

        # tableList=transformed_list
        for d in transformed_list:
            if d['category'] != 'Total':
                if d['department'] not in middle_sum_by_department:
                    middle_sum_by_department[d['department']] = 0
                    grass_sum_by_department[d['department']] = 0
                    synthesis_sum_by_department[d['department']] = 0
                    senior_sum_by_department[d['department']] = 0
                middle_sum_by_department[d['department']] += int(d['content_number_middle'])
                grass_sum_by_department[d['department']] += int(d['content_number_grass'])
                synthesis_sum_by_department[d['department']] += int(d['content_number_synthesis'])
                senior_sum_by_department[d['department']] += int(d['content_number_senior'])
            else:
                d['content_number_middle'] = 0
                d['content_number_grass'] = 0
                d['content_number_synthesis'] = 0
                d['content_number_senior'] = 0

        for d in transformed_list:
            if d['category'] == 'Total':
                d['content_number_middle'] = float(middle_sum_by_department[d['department']])
                d['content_number_grass'] = float(grass_sum_by_department[d['department']])
                d['content_number_synthesis'] = float(synthesis_sum_by_department[d['department']])
                d['content_number_senior'] = float(senior_sum_by_department[d['department']])

        # tableList=transformed_list
        # for line in tableList:
        #     if line['category']=='Total':
        #         line['category']='合计'
        #
        #     line['content_satisfaction_avg']=round(line['content_satisfaction_avg'],2)
        #     line['content_satisfaction'] = round(line['content_satisfaction'], 2)
        #     line['content_duration']=round(line['content_duration'],2)
        #     if line['content_number_Total']==0:
        #
        #         line['content_satisfaction']=0
        #     else:
        #         print()
        #         line['content_satisfaction'] = round(float(line['content_satisfaction']) / float(line['content_number_Total']), 2)      #总满意度/总场次

        tableList = []
        # 创建一个字典用于存储不同部门和类别的总和
        department_category_totals = {}
        for entry in transformed_list:
            department = entry['department']
            category = entry['category']
            count_null_Satisfaction_Total = entry['count_null_Satisfaction_Total']
            content_duration = entry['content_duration']
            content_satisfaction = entry['content_satisfaction']
            content_satisfaction_avg = entry['content_satisfaction_avg']
            content_people_number = entry['content_people_number']
            content_number_Total = float(entry['content_number_Total'])
            content_number_middle = float(entry['content_number_middle'])
            content_number_grass = float(entry['content_number_grass'])
            content_number_synthesis = float(entry['content_number_synthesis'])
            content_number_senior = float(entry['content_number_senior'])
            count_null_Satisfaction_Total = float(entry['count_null_Satisfaction_Total'])

            # 初始化部门和类别的总和字典条目
            if (department, category) not in department_category_totals:
                department_category_totals[(department, category)] = {
                    'content_number_senior': 0,  # 高层
                    'content_number_middle': 0,  # 中层
                    'content_number_grass': 0,  # 基础
                    'content_number_synthesis': 0,  # 综合
                    'content_number_Total': 0,
                    'content_people_number': 0,
                    'content_duration': 0,
                    'content_satisfaction': 0,
                    'content_satisfaction_avg': 0,
                    'count_null_Satisfaction_Total': 0,
                }

            # 累加各个指标
            department_category_totals[(department, category)]['content_number_senior'] += content_number_senior
            department_category_totals[(department, category)]['content_number_middle'] += content_number_middle
            department_category_totals[(department, category)]['content_number_grass'] += content_number_grass
            department_category_totals[(department, category)]['content_number_synthesis'] += content_number_synthesis
            department_category_totals[(department, category)]['content_number_Total'] += content_number_Total
            department_category_totals[(department, category)]['content_people_number'] += content_people_number
            department_category_totals[(department, category)]['content_duration'] += content_duration
            department_category_totals[(department, category)]['content_satisfaction'] += content_satisfaction
            department_category_totals[(department, category)]['content_satisfaction_avg'] += content_satisfaction_avg
            department_category_totals[(department, category)][
                'count_null_Satisfaction_Total'] += count_null_Satisfaction_Total

        # 创建润阳集团的初始数据
        ruyang_data = []

        # 遍历字典中的项，根据不同的部门和类别动态创建数据
        for (department, category), totals in department_category_totals.items():
            ruyang_entry = {
                'department': '润阳集团',
                'category': category,
                'count_null_Satisfaction_Total': totals['count_null_Satisfaction_Total'],
                'content_duration': totals['content_duration'],
                'content_satisfaction': totals['content_satisfaction'],
                'content_satisfaction_avg': totals['content_satisfaction_avg'],
                'content_number_Total': totals['content_number_Total'],
                'content_number_middle': totals['content_number_middle'],
                'content_number_grass': totals['content_number_grass'],
                'content_number_synthesis': totals['content_number_synthesis'],
                'content_number_senior': totals['content_number_senior'],
                'content_people_number': totals['content_people_number']
            }
            ruyang_data.append(ruyang_entry)

        # 将润阳集团的数据添加到总和字典中
        for entry in ruyang_data:
            department = entry['department']
            category = entry['category']
            count_null_Satisfaction_Total = entry['count_null_Satisfaction_Total']
            content_duration = entry['content_duration']
            content_satisfaction = entry['content_satisfaction']
            content_satisfaction_avg = entry['content_satisfaction_avg']
            content_number_Total = entry['content_number_Total']
            content_number_middle = entry['content_number_middle']
            content_number_grass = entry['content_number_grass']
            content_number_synthesis = entry['content_number_synthesis']
            content_number_senior = entry['content_number_senior']
            content_people_number = entry['content_people_number']

            if (department, category) not in department_category_totals:
                department_category_totals[(department, category)] = {
                    'content_number_senior': 0,  # 高层
                    'content_number_middle': 0,  # 中层
                    'content_number_grass': 0,  # 基础
                    'content_number_synthesis': 0,  # 综合
                    'content_number_Total': 0,
                    'content_people_number': 0,
                    'content_duration': 0,
                    'content_satisfaction': 0,
                    'content_satisfaction_avg': 0,
                    'count_null_Satisfaction_Total': 0,
                }

            department_category_totals[(department, category)]['content_number_senior'] += content_number_senior
            department_category_totals[(department, category)]['content_number_middle'] += content_number_middle
            department_category_totals[(department, category)]['content_number_grass'] += content_number_grass
            department_category_totals[(department, category)]['content_number_synthesis'] += content_number_synthesis
            department_category_totals[(department, category)]['content_number_Total'] += content_number_Total
            department_category_totals[(department, category)]['content_people_number'] += content_people_number
            department_category_totals[(department, category)]['content_duration'] += content_duration
            department_category_totals[(department, category)]['content_satisfaction'] += content_satisfaction
            department_category_totals[(department, category)]['content_satisfaction_avg'] += content_satisfaction_avg
            department_category_totals[(department, category)][
                'count_null_Satisfaction_Total'] += count_null_Satisfaction_Total

        # 将结果转化为列表形式
        zz = [{'department': department, 'category': category, **totals}
              for (department, category), totals in department_category_totals.items()]

        # 打印 zz
        for entry in zz:
            tableList.append(entry)

        for line in tableList:
            if line['department'] == '润阳集团':

                if line['content_number_Total'] - line['count_null_Satisfaction_Total'] == 0:
                    line['content_satisfaction_avg'] = 0
                else:
                    line['content_satisfaction_avg'] = (line['content_satisfaction']) / (
                                line['content_number_Total'] - line['count_null_Satisfaction_Total'])

                if line['content_number_Total'] == 0:
                    line['content_satisfaction'] = 0
                else:

                    line['content_satisfaction'] = line['content_satisfaction'] / line[
                        'content_number_Total']  # 总满意度/总场次
            else:
                if line['content_number_Total'] == 0:
                    line['content_satisfaction'] = 0
                else:

                    line['content_satisfaction'] = line['content_satisfaction'] / line[
                        'content_number_Total']  # 总满意度/总场次
        for line in tableList:
            if line['category'] == 'Total':
                line['category'] = '合计'

            line['content_satisfaction_avg'] = round(line['content_satisfaction_avg'], 3)
            line['content_satisfaction'] = round(line['content_satisfaction'], 3)
            line['content_satisfaction'] = round(line['content_satisfaction'], 3)
            line['content_duration'] = round(line['content_duration'], 3)

        # 自定义排序规则
        def custom_sort(item):
            # 如果部门名称为'润阳集团'，将其排在第一个，否则按照原始顺序排列
            if item['department'] == '润阳集团':
                return 0
            return 1

        tableList = sorted(tableList, key=custom_sort)

        base_ls = list(HrDepartment.objects.filter(~Q(id=999999),
                                                   Q(department_expiry_date__isnull=True) | Q(
                                                       department_expiry_date__gt=datetime.now()),
                                                   department_first_name__isnull=False,
                                                   department_status=1,
                                                   id__in=self.request.user_department_employee).values_list(
            'department_name', flat=True))
        base_ls += ['润阳集团']
        tableList = [item for item in tableList if item['department'] in base_ls]  # 筛权限

        self.return_data = {
            'code': 200,
            'msg': '信息返回成功',
            'data': {
                'columnList': columnList,
                'tableList': tableList,
            }
        }

    def download_month_summary_analysis(self):  # 下载本月汇总分析
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)

        template_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'template_file',
                                     '本月汇总分析模板.xlsx')  # 创建文件夹
        import shutil

        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2),
                                        '线下培训汇总分析.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)

        self.month_summary_analysis()
        tableList = self.return_data['data']['tableList']
        row_data = []
        for line in tableList:
            line_data = []
            for k, v in line.items():
                line_data.append(v)
            row_data.append(line_data)
        # print(row_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        # print(destination_path)

        # 使用字符串替换将\替换为/
        destination_path = destination_path.replace('\\', '/')

        # print(destination_path)

        self.merge_cells(destination_path)
        destination_path = 'static/' + destination_path.split('static/')[1]
        # destination_path='static\\'+destination_path.split('static\\')[1]
        # print(destination_path)
        # print(destination_path)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

    def offline_training_options(self):
        #
        # def replace_id_with_value(data, start_index=1):
        #     for item in data:
        #         item["value"] = item.pop("id")
        #         item["index"] = start_index
        #         start_index += 1
        #         if "children" in item:
        #             start_index = replace_id_with_value(item["children"], start_index)
        #     return start_index
        #
        # replace_id_with_value(hrbase_data)

        # hrbase_boss_obj = HrDepartment.objects.filter(
        #     Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gte=datetime.now()), department_status=1,
        #     department_parent_id=0).values('id', 'department_name')
        # hrbase_data = [
        #     {"value": hrbase_boss_obj[0]['id'], "label": hrbase_boss_obj[0]['department_name'], "children": hrbase_data,"index": 0,}]

        # content_type_list=list(TrainingContentType.objects.filter(type_status=True).values('id','type_name'))
        content_category_list = list(
            TrainingContentCategory.objects.filter(category_status=True).values('id', 'category_name'))
        content_level_list = list(TrainingContentLevel.objects.filter(level_status=True).values('id', 'level_name'))
        lecturer_level_list = list(TrainingLecturerLevel.objects.filter(level_status=True).values('id', 'level_name'))

        base_ls = [value for value in self.request.user_department_employee if value != 1]
        departments = HrDepartment.objects.filter(
            ~Q(id=999999),
            Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),
            department_status=1,
            id__in=base_ls
        ).values('id', 'department_name', 'department_parent_id')
        hrbase_data = get_trees(departments, 'id', 'department_parent_id')

        def add_indexes(node_list, index=0):
            for node in node_list:
                node['index'] = index
                index += 1
                children = node.get('children', [])
                if children:
                    index = add_indexes(children, index)
            return index

        add_indexes(hrbase_data)

        # def flatten_tree(node, level, result):
        #     if len(result) <= level:
        #         result.append([])
        #     result[level].append({"label": node['label'], "value": node['value']})
        #     if 'children' in node:
        #         for child in node['children']:
        #             flatten_tree(child, level + 1, result)
        # flattened_result = [[] for _ in range(2)]
        # for node in hrbase_data:
        #     flatten_tree(node, 0, flattened_result)
        # print(flattened_result)
        # for level, nodes in enumerate(flattened_result):
        #     print(nodes)

        content_types = TrainingContentType.objects.filter(
            type_status=1
        ).values('id', 'type_name', 'type_parent_id')
        content_type_data = get_trees2(content_types, 'id', 'type_parent_id')

        def add_index(tree, start_index=1):
            for node in tree:
                node['index'] = start_index
                start_index += 1
                if 'children' in node:
                    start_index = add_index(node['children'], start_index)
            return start_index

        add_index(content_type_data)

        self.return_data = {
            'data': {
                'hrbase_data': hrbase_data,
                'content_type_list':
                # {"value": item["id"], "label": item["type_name"]}
                # for item in content_type_list
                    content_type_data
                ,
                'content_level_list': [
                    {"value": item["id"], "label": item["level_name"]}
                    for item in content_level_list
                ],
                'lecturer_level_list': [
                    {"value": item["id"], "label": item["level_name"]}
                    for item in lecturer_level_list
                ],
                "content_category_list": [
                    {"value": item["id"], "label": item["category_name"]}
                    for item in content_category_list
                ]

            },
            'code': HTTP_200_OK,
            'msg': '查询成功'
        }

    def month_Training_hours_per_person(self):
        from django.db.models import Sum, F, Q
        from datetime import datetime
        from django.utils import timezone
        month = self.request.GET.get('month', None)

        if month != "":
            month = month
        else:
            month = datetime.now().date()
        if type(month) == str:
            datetime_obj = datetime.strptime(month, "%Y-%m-%d")
        else:
            datetime_obj = datetime.combine(month, datetime.min.time())
        #
        current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)

        # from django.db.models import Sum, F, ExpressionWrapper, FloatField, Q
        # from datetime import datetime
        #
        # # Define an ExpressionWrapper to calculate the product of content_duration and content_people_number
        # product_expression = ExpressionWrapper(
        #     F('content_duration') * F('content_people_number'),
        #     output_field=FloatField()
        # )
        #
        # # Create a queryset that filters for content_manner '现场' and aggregates the sum of the product by department_first_name
        # result = (
        #     TrainingContent.objects.filter(
        #         content_manner='现场',
        #         content_part__department_expiry_date__isnull=True,
        #         content_part__department_status=True,
        #         content_begin_date__range=(current_month,next_month)
        #     )
        #     .annotate(product=product_expression)
        #     .values('content_part__department_first_name')
        #     .annotate(total=Sum('product'))
        #     .values('content_part__department_first_name', 'total')
        # )
        second_name_list = ['江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司', '江苏润阳光伏科技有限公司',
                            '江苏海博瑞光伏科技有限公司', '润宝电力', '润阳光伏科技（泰国）有限公司',
                            '宁夏润阳硅材料科技有限公司', '江苏润阳光伏科技有限公司（二期）',
                            '润阳泰国四期组件', '润阳泰国四期电池', '云南润阳世纪光伏科技有限公司']

        first_name_list = ['光伏研究院', '全球财务中心', '人力资源中心', '全球战略供应链管理中心']
        data = list(TrainingContent.objects.filter(Q(content_part__department_expiry_date__gt=datetime.now()) |
                                                   Q(content_part__department_first_name__isnull=False),
                                                   # Q(content_part__department_first_name__in=first_name_list)|Q(content_part__department_second_name__in=second_name_list),
                                                   content_manner='现场',
                                                   content_part__department_expiry_date__isnull=True,
                                                   content_part__department_status=True,
                                                   content_status=True,
                                                   content_begin_date__range=(current_month, next_month),
                                                   content_part__department_first_name__in=first_name_list
                                                   ).values('content_part__department_first_name', 'content_duration',
                                                            'content_people_number'))
        for line in data:
            duration = float(line['content_duration'])
            people_number = float(line['content_people_number'])
            line['total'] = duration * people_number
        # print(data)
        department_totals = {}
        for item in data:
            department_name = item['content_part__department_first_name']
            total = item['total']
            if department_name in department_totals:
                department_totals[department_name] += total
            else:
                department_totals[department_name] = total
        # 将结果存储到一个新的列表中
        result = [{'department_first_name': department_name, 'total': total} for department_name, total in
                  department_totals.items()]
        all_first_dp = list(HrDepartment.objects.filter(Q(department_expiry_date__gt=datetime.now()) |
                                                        Q(department_first_name__isnull=False),
                                                        department_expiry_date__isnull=True,
                                                        department_status=True,
                                                        department_first_name__in=first_name_list).values_list(
            'department_first_name',

            flat=True).distinct())
        for dp in all_first_dp:
            # 检查是否在列表b中的department_first_name对应值
            if not any(dp == department['department_first_name'] for department in result):
                # 如果不在，添加到列表b，并赋值total为0
                result.append({'department_first_name': dp, 'total': 0.0})

        data_second = list(TrainingContent.objects.filter(Q(content_part__department_expiry_date__gt=datetime.now()) |
                                                          Q(content_part__department_first_name__isnull=False),
                                                          # Q(content_part__department_first_name__in=first_name_list)|Q(content_part__department_second_name__in=second_name_list),
                                                          content_manner='现场',
                                                          content_part__department_expiry_date__isnull=True,
                                                          content_part__department_status=True,
                                                          content_status=True,
                                                          content_begin_date__range=(current_month, next_month),
                                                          content_part__department_second_name__in=second_name_list
                                                          ).values('content_part__department_second_name',
                                                                   'content_duration', 'content_people_number'))

        for line in data_second:
            duration = float(line['content_duration'])
            people_number = float(line['content_people_number'])
            line['total'] = duration * people_number
        # print(data)
        department_totals = {}
        for item in data_second:
            department_name = item['content_part__department_second_name']
            total = item['total']
            if department_name in department_totals:
                department_totals[department_name] += total
            else:
                department_totals[department_name] = total
        # 将结果存储到一个新的列表中
        result2 = [{'department_second_name': department_name, 'total': total} for department_name, total in
                   department_totals.items()]
        all_first_dp = list(HrDepartment.objects.filter(Q(department_expiry_date__gt=datetime.now()) |
                                                        Q(department_first_name__isnull=False),
                                                        department_expiry_date__isnull=True,
                                                        department_status=True,
                                                        department_second_name__in=second_name_list).values_list(
            'department_second_name', flat=True).distinct())

        for dp in all_first_dp:
            # 检查是否在列表b中的department_first_name对应值
            if not any(dp == department['department_second_name'] for department in result2):
                # 如果不在，添加到列表b，并赋值total为0
                result2.append({'department_second_name': dp, 'total': 0.0})

        update_result2 = [{'department_first_name': item['department_second_name'], 'total': item['total']} for item in
                          result2]
        result = result + update_result2

        for department in result:
            sessionInfo = {
                "sessions_base": department['department_first_name'],
                'sessions_offline_total': department['total'],
                'sessions_record_time': str(current_month)[:10],
            }
            TrainingSessions.objects.update_or_create(defaults=sessionInfo, sessions_base=sessionInfo['sessions_base'],
                                                      sessions_record_time=sessionInfo['sessions_record_time'])

        # TrainingContent.objects.filter()
        columnList = [
            {'label': 'sessions_base', 'value': '培训基地', 'width': 300},
            {'label': 'sessions_offline_total', 'value': '线下培训总时数', 'width': 260},
            {'label': 'sessions_cloud_total', 'value': '线上(云学堂)培训总时数', 'width': 260},
            {'label': 'sessions_persons_register', 'value': '月平均在册人数', 'width': 260},
            {'label': 'sessions_per_people', 'value': '基地人均培训课时', 'width': ''}
        ]

        kwargs = {}
        # print('month',month,type(month))
        if month != "":
            if type(month) == str:
                month = datetime.strptime(month, '%Y-%m-%d')
            modified_date = date(month.year, month.month, 1)
            kwargs['sessions_record_time'] = modified_date

        base_ls = list(HrDepartment.objects.filter(~Q(id=999999),
                                                   Q(department_expiry_date__isnull=True) | Q(
                                                       department_expiry_date__gt=datetime.now()),
                                                   department_first_name__isnull=False,
                                                   department_status=1,
                                                   id__in=self.request.user_department_employee).values_list(
            'department_name', flat=True))
        tableList = list(TrainingSessions.objects.filter(**kwargs).values('id',
                                                                          'sessions_base',
                                                                          'sessions_offline_total',
                                                                          "sessions_cloud_total",
                                                                          'sessions_persons_register',
                                                                          'sessions_per_people',
                                                                          'sessions_record_time'))
        for line in tableList:
            if line['sessions_cloud_total'] is None:
                line['sessions_cloud_total'] = 0
            if line['sessions_persons_register'] is None:
                line['sessions_persons_register'] = 0
            if line['sessions_per_people'] is None:
                line['sessions_per_people'] = 0
            if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
                'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
                try:
                    line['sessions_per_people'] = round(
                        (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
                            'sessions_persons_register'], 2)
                except:
                    line['sessions_per_people'] = 0
            # print(line)
            TrainingSessions.objects.update_or_create(defaults=line, sessions_base=line['sessions_base'],
                                                      sessions_record_time=line['sessions_record_time'])

        tableList = [item for item in tableList if item['sessions_base'] in base_ls]  # 筛权限

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
            }
        }

    def edit_month_Training_hours_per_person(self):
        info = json.loads(self.request.body)
        TrainingSessions.objects.filter(pk=info['id']).update(**info)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }

    def download_month_Training_hours_per_person(self):  # 下载本月人均课时
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        file_ls = [
            "序号", "中心/基地", "线下培训总时数", "线上(云学堂)培训总时数", "月平均在册人数", "基地人均培训课时"
        ]
        path = self.createExcelPath('基地人均培训课时.xlsx', str(t2), '线下培训报表', 40, 'A1:F1', *file_ls)

        month = self.request.GET.get('month', None)
        # print('month', month)
        kwargs = {

        }
        if month != "":
            kwargs['sessions_record_time'] = month
        else:
            today = datetime.today()
            month = datetime(today.year, today.month, 1)
            kwargs['sessions_record_time'] = month

        base_ls = list(HrDepartment.objects.filter(~Q(id=999999),
                                                   Q(department_expiry_date__isnull=True) | Q(
                                                       department_expiry_date__gt=datetime.now()),
                                                   department_first_name__isnull=False,
                                                   department_status=1,
                                                   id__in=self.request.user_department_employee).values_list(
            'department_name', flat=True))
        # print(base_ls)
        tableList = list(TrainingSessions.objects.filter(**kwargs).values(
            'sessions_base',
            'sessions_offline_total',
            "sessions_cloud_total",
            'sessions_persons_register',
            'sessions_per_people'))
        for line in tableList:
            if line['sessions_cloud_total'] is None:
                line['sessions_cloud_total'] = 0
            if line['sessions_persons_register'] is None:
                line['sessions_persons_register'] = 0
            if line['sessions_per_people'] is None:
                line['sessions_per_people'] = 0
            if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line[
                'sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
                try:
                    line['sessions_per_people'] = round(
                        (line['sessions_offline_total'] + line['sessions_cloud_total']) / line[
                            'sessions_persons_register'], 2)
                except:
                    line['sessions_per_people'] = 0

        tableList = [item for item in tableList if item['sessions_base'] in base_ls]
        row_data = []
        index = 1
        # print('ttttt',tableList)
        for line in tableList:
            # print(line)
            line_data = []
            for k, v in line.items():
                # print(k,v)
                line_data.append(v)
            #     line = list(line)
            line_data.insert(0, index)
            # print(line)
            row_data.append(line_data)
            if len(line_data) == 0:
                index = index
            index += 1

        exc = openpyxl.load_workbook(path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(path)  # 指定路径,保存文件
        # print('path',path)
        # self.merge_cells(path)

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10

        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'offlineTrainingFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path

    @staticmethod
    def merge_cells(file_path):
        from openpyxl import load_workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 从第3行开始，每四行合并A列
        for row in range(3, sheet.max_row + 1, 4):
            start_cell = sheet.cell(row=row, column=1)
            end_cell = sheet.cell(row=row + 3, column=1)
            sheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
        workbook.save(file_path)

    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def extract_info(input_string):
        import re
        # 使用正则表达式提取名字和学号
        match = re.match(r'^(.*?)[\(|（]([\w]+)[\)|）]$', input_string)

        if match:
            name = match.group(1)
            student_id = match.group(2)
            return name, student_id
        else:
            return None, None

    @staticmethod
    def createPath(pic, path, fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        now = arrow.now()
        t = now.format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  # 文件后缀

        file_name = f"{fileName}.{file_suffix}"  # 文件名称

        file_path = os.path.join('static', 'offlineTrainingFile', 'upload_file', t, path, file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path, file_name, file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path, file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path), 'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)





# import json, os,arrow,openpyxl
# import random
#
# from django.db.models import Q,F
# from rest_framework import status
# from django.http import JsonResponse
# from django.core import serializers
# from rest_framework.status import HTTP_200_OK
#
# from auther.models import AdminUser
# # from employee.views import get_user_field
# from django.db import models
# from employee.models import HrEmployee, HrEmployeeFiles, HrDepartment
# from rest_framework.response import Response
# from datetime import datetime, date
#
# from offlineTraining.models import *
# from pdss.settings import BASE_DIR
# from employee import views
#
#
# def get_trees(data,
#               key_column='id',
#               parent_column='parent_id',
#               child_column='children',
#
#               current_column=None,
#               current_path=None
#               ):
#     """
#     :param data: 数据列表
#     :param key_column: 主键字段，默认id
#     :param parent_column: 父ID字段名，父ID默认从0开始
#     :param child_column: 子列表字典名称
#     :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
#     :param current_path: 当前展开值
#     :return: 树结构
#     """
#     data_dic = {}
#     # data_dic.update({341: {'id': 341, 'department_name': '硅料事业部', 'department_parent_id': 1}})
#     for d in data:
#         data_dic[d.get(key_column)] = d  # 以自己的权限主键为键,以新构建的字典为值,构造新的字典
#
#     # print(data_dic)
#     # data_dic={3: {'id': 3, 'department_name': '全球战略供应链管理中心', 'department_parent_id': 1}, 187: {'id': 187, 'department_name': '物流部', 'department_parent_id': 3}, 540: {'id': 540, 'department_name': '物流部（欧美区域）', 'department_parent_id': 187}, 543: {'id': 543, 'department_name': '美国', 'department_parent_id': 540}}
#     data_tree_list = []  # 整个数据大列表
#     for d_id, d_dic in data_dic.items():
#         d_dic['label'] = d_dic.pop('department_name')
#         d_dic['value'] = d_id  # Change the key name 'id' to 'value'
#         # print(d_id,d_dic)
#         pid = d_dic.get(parent_column)  # 取每一个字典中的父id
#         if pid ==1:  # 父id=0，就直接加入数据大列表
#             data_tree_list.append(d_dic)
#         else:  # 父id>0 就加入父id队对应的那个的节点列表
#             try:  # 判断异常代表有子节点，增加子节点列表=[]
#                 data_dic[pid][child_column].append(d_dic)
#             except KeyError:
#                 # if pid in data_dic:
#                 try:
#                     data_dic[pid][child_column] = []
#                     data_dic[pid][child_column].append(d_dic)
#                 except:
#                     pass
#
#
#                 # else:
#                 #     pass
#                     # # 处理键不存在的情况
#                     # print(f"键 {pid} 不存在于 data_dic 中")
#                     # # print(request)
#                     # request.user_department_employee.append(pid)
#                     # print(request.user_department_employee)
#     return data_tree_list
#
#
# # def get_trees(data,
# #               key_column='id',
# #               parent_column='department_parent_id',
# #               child_column='children',
# #               current_column=None,
# #               current_path=None):   #培训类型
# #     """
# #     :param data: 数据列表
# #     :param key_column: 主键字段，默认'id'
# #     :param parent_column: 父ID字段名，默认'department_parent_id'
# #     :param child_column: 子列表字典名称，默认'children'
# #     :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
# #     :param current_path: 当前展开值
# #     :return: 树结构
# #     """
# #     data_dic = {}
# #     for d in data:
# #         d[key_column] = str(d[key_column])  # Ensure the key is a string
# #         data_dic[d[key_column]] = d  # Use the specified key_column as the dictionary key
# #
# #     data_tree_list = []  # 整个数据大列表
# #     # print(data_dic)
# #     # data_tree_list[str(341)] = {'id': '341', 'department_name': '硅料事业部', 'department_parent_id': 1}
# #
# #     # data_dic.update({'341': {'id': '341', 'department_name': '硅料事业部', 'department_parent_id': 1}})
# #     # print(data_dic)
# #     data_dic={'187': {'id': '187', 'department_name': '物流部', 'department_parent_id': 3}, '540': {'id': '540', 'department_name': '物流部（欧美区域）', 'department_parent_id': 187}, '543': {'id': '543', 'department_name': '美国', 'department_parent_id': 540}}
# #     for d_id, d_dic in data_dic.items():
# #         d_dic['label'] = d_dic.pop('department_name')
# #         d_dic['value'] = int(d_id)
# #         pid = d_dic.get(parent_column)
# #         if pid is None or pid == 0:  # Check for root nodes (e.g., 0)
# #             data_tree_list.append(d_dic)
# #         else:
# #             parent_node = data_dic.get(str(pid))  # Convert pid to a string for dictionary lookup
# #             if parent_node:
# #                 if child_column not in parent_node:
# #                     parent_node[child_column] = []
# #                 parent_node[child_column].append(d_dic)
# #     print(data_tree_list)
# #     return data_tree_list
#
#
# def get_trees2(data,
#               key_column='id',
#               parent_column='department_parent_id',
#               child_column='children',
#               current_column=None,
#               current_path=None):   #培训类型
#     """
#     :param data: 数据列表
#     :param key_column: 主键字段，默认'id'
#     :param parent_column: 父ID字段名，默认'department_parent_id'
#     :param child_column: 子列表字典名称，默认'children'
#     :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
#     :param current_path: 当前展开值
#     :return: 树结构
#     """
#     data_dic = {}
#     for d in data:
#         d[key_column] = str(d[key_column])  # Ensure the key is a string
#         data_dic[d[key_column]] = d  # Use the specified key_column as the dictionary key
#
#     data_tree_list = []  # 整个数据大列表
#     for d_id, d_dic in data_dic.items():
#         d_dic['label'] = d_dic.pop('type_name')
#         d_dic['value'] = int(d_id)
#         pid = d_dic.get(parent_column)
#         if pid is None or pid == 0:  # Check for root nodes (e.g., 0)
#             data_tree_list.append(d_dic)
#         else:
#             parent_node = data_dic.get(str(pid))  # Convert pid to a string for dictionary lookup
#             if parent_node:
#                 if child_column not in parent_node:
#                     parent_node[child_column] = []
#                 parent_node[child_column].append(d_dic)
#     return data_tree_list
#
# class Content:
#     def __init__(self, request,meth):
#         self.request = request
#         self.now = arrow.now()
#         self.t1 = self.now.format('YYYY-MM-DD')
#         self.t2 = self.now.timestamp()
#         self.return_data = {
#             'code': 200,
#             'msg': '信息返回成功'
#         }
#         self.meth = meth
#         self.methods = {
#             'get_content_info': self.get_content_info,
#             'post_content_info': self.post_content_info,
#             'delete_content_info': self.delete_content_info,
#             'edit_content_info': self.edit_content_info,
#             'download_content_info': self.download_content_info,
#             'batch_content_info':self.batch_content_info,#批量上传
#             'offline_training_options': self.offline_training_options,  # 汇总下拉菜单
#             'month_summary_analysis':self.month_summary_analysis,        #本月汇总分析
#             'month_Training_hours_per_person':self.month_Training_hours_per_person,   #本月人均课时
#             "edit_month_Training_hours_per_person":self.edit_month_Training_hours_per_person,
#             'download_month_Training_hours_per_person':self.download_month_Training_hours_per_person,
#             'download_month_summary_analysis':self.download_month_summary_analysis,
#             'del_content_file':self.del_content_file,
#             'post_content_file':self.post_content_file,
#         }
#
#     def method_center(self):
#         if self.request.check_token is None:
#             self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问!'}
#             return JsonResponse(self.return_data)
#         self.methods[self.meth]()
#         return JsonResponse(self.return_data)
#
#
#     # 获取信息列表
#     def get_content_info(self):
#         columnList = [
#             {'label': 'index', 'value': '序号', 'width': 60},
#             {'label': 'content_part__department_first_name', 'value': '一级部门', 'width':160},
#             {'label': 'content_part__department_second_name', 'value': '二级部门', 'width': 260},
#             {'label': 'content_part__department_third_name', 'value': '三级部门', 'width': 160},
#             {'label': 'content_part__department_forth_name', 'value': '四级部门', 'width': 260},
#             # {'label': 'content_module', 'value': '模块', 'width': 160},
#             # {'label': 'content_group', 'value': '组', 'width': 160},
#             {'label': 'content_title', 'value': '培训主题/课题', 'width': 160},
#             {'label': 'content_type__type_name', 'value': '培训类型', 'width': 160},
#             {'label': 'content_category__category_name', 'value': '培训类别', 'width': 160},
#             {'label': 'content_level__level_name', 'value': '培训层级', 'width': 160},
#             {'label': 'content_manner', 'value': '培训方式', 'width': 160},
#             {'label': 'content_begin_date', 'value': '开始培训日期', 'width': 160},
#             {'label': 'content_end_date', 'value': '截至培训日期', 'width': 160},
#             {'label': 'content_duration', 'value': '培训时长(H)', 'width': 160},
#             {'label': 'content_object', 'value': '培训对象', 'width': 160},
#             {'label': 'content_people_number', 'value': '参训人数', 'width': 160},
#             {'label': 'content_lecturer__lecturer_type', 'value': '讲师类型', 'width': 160},
#             {'label': 'content_lecturer__lecturer_people__employee_name', 'value': '培训讲师', 'width': 160},
#             {'label': 'content_lecturer__lecturer_people__employee_code', 'value': '讲师工号', 'width': 160},
#             {'label': 'content_lecturer__lecturer_people__employee_position__position_name', 'value': '讲师岗位', 'width': 160},
#             {'label': 'content_lecturer__lecturer_level__level_name', 'value': '讲师级别','width': 160},
#             {'label': 'content_satisfaction', 'value': '培训满意度', 'width': 160},
#             {'label': 'content_expenses', 'value': '培训费用', 'width': 160},
#             {'label': 'content_plan', 'value': '计划内/计划外', 'width': 160},
#
#             {'label': 'createPhoto_num', 'value': '培训照片', 'width': 160},
#             {'label': 'createFile_num', 'value': '培训课件', 'width': 160},
#             {'label': 'training_satisfaction_file_num', 'value': '培训满意度', 'width': 160},
#             {'label': 'signin_sheet_file_num', 'value': '签到表', 'width': 160},
#             {'label': 'content_creater__user', 'value': '用户昵称', 'width': 160},
#         ]
#         kwargs={"content_status":True}
#
#         info=json.loads(self.request.body)
#         currentPage = info['currentPage'] if info['currentPage'] != "" else 1
#         pageSize = info['pageSize'] if info['pageSize'] != "" else 25
#         searchName = info['searchName']
#         beginDate = info['beginDate']
#         endDate = info['endDate']
#         contentTypeList=info['contentTypeList']
#         kwargs['content_type_id__in'] = contentTypeList
#         employee_base = info['baseNameId']
#         if beginDate != "" and endDate != "":
#             kwargs['content_begin_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(beginDate) == 0 else beginDate
#             kwargs['content_begin_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate
#         kwargs['content_part_id__in'] = employee_base
#         print(kwargs)
#         if employee_base=='':
#             kwargs['content_part_id__in'] = self.request.user_department_employee
#
#         kwargs = {key: value for key, value in kwargs.items() if
#                   value is not None and value != '' and value != []}  # 过滤掉值为None或''的项
#         totalNumber = TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(content_lecturer__lecturer_people__employee_name__contains=searchName)| Q(content_lecturer__lecturer_people__employee_code__contains=searchName),**kwargs).count()
#         tableList = list(TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(content_lecturer__lecturer_people__employee_name__contains=searchName)| Q(content_lecturer__lecturer_people__employee_code__contains=searchName),**kwargs).values('id',
#                                                                'content_part_id',
#                                                                'content_part__department_first_name',
#                                                                "content_part__department_second_name",
#                                                                 "content_part__department_third_name",
#                                                                 "content_part__department_forth_name",
#                                                                'content_module',
#                                                                'content_group',
#                                                                'content_title',
#                                                                'content_type__type_name',
#                                                                'content_category__category_name',
#                                                                'content_level__level_name',
#                                                                'content_manner',
#                                                                'content_begin_date',
#                                                                'content_end_date',
#                                                                'content_duration',
#                                                                'content_object',
#                                                                'content_people_number',
#                                                                'content_type_id',
#                                                                'content_category_id',
#                                                                'content_level_id',
#                                                                'content_lecturer__lecturer_type',
#                                                                'content_lecturer__lecturer_name',
#                                                                'content_lecturer__lecturer_people__employee_name',
#                                                                'content_lecturer__lecturer_people__employee_code',
#                                                                'content_lecturer__lecturer_people__employee_position__position_name',
#                                                                'content_lecturer__lecturer_level__level_name',
#                                                                "content_lecturer__lecturer_level_id",
#                                                                 'content_lecturer',
#                                                                'content_satisfaction',
#                                                                'content_expenses',
#                                                                'content_plan',
#                                                                'content_creater__user').order_by('-content_createTime')[
#                          (currentPage - 1) * pageSize:currentPage * pageSize])
#         all_id = [item['id'] for item in tableList]
#         file_list=list(TrainingFiles.objects.filter(training_content_file_id__in=all_id,training_file_status=True).values('id','training_file_name','training_file_type','training_file_url','training_content_file_id'))
#
#         createPhoto = {}  # 培训照片
#         createFile = {}  # 培训附件
#         signin_sheet_file = {}  # 簽到表
#         training_satisfaction_file={}#   培訓滿意度
#
#         for item in file_list:  # 查找每份提案对应的文件
#             training_content_file_id = item.get('training_content_file_id')
#             if item['training_file_type'] == 1:  # 活动照片
#                 if training_content_file_id not in createPhoto:
#                     createPhoto[training_content_file_id] = []
#                 createPhoto[training_content_file_id].append(item)
#             elif item['training_file_type'] == 2:  # 活动方案
#                 if training_content_file_id not in createFile:
#                     createFile[training_content_file_id] = []
#                 createFile[training_content_file_id].append(item)
#             elif item['training_file_type'] == 3:  # 簽到表
#                 if training_content_file_id not in signin_sheet_file:
#                     signin_sheet_file[training_content_file_id] = []
#                 signin_sheet_file[training_content_file_id].append(item)
#             elif item['training_file_type'] ==4:  # 培訓滿意度
#                 if training_content_file_id not in training_satisfaction_file:
#                     training_satisfaction_file[training_content_file_id] = []
#                 training_satisfaction_file[training_content_file_id].append(item)
#
#
#
#         for item in tableList:
#             content_id = item.get('id')
#             if content_id in createPhoto:
#                 item['createPhoto_ls'] = createPhoto[content_id]
#                 item['createPhoto_num'] = len(createPhoto[content_id])
#             else:
#                 item['createPhoto_ls'] = []
#                 item['createPhoto_num'] = 0
#             if content_id in createFile:
#                 item['createFile_ls'] = createFile[content_id]
#                 item['createFile_num'] = len(createFile[content_id])
#             else:
#                 item['createFile_ls'] = []
#                 item['createFile_num'] = 0
#
#             if content_id in signin_sheet_file:
#                 item['signin_sheet_file_ls'] = signin_sheet_file[content_id]
#                 item['signin_sheet_file_num'] = len(signin_sheet_file[content_id])
#             else:
#                 item['signin_sheet_file_ls'] = []
#                 item['signin_sheet_file_num'] = 0
#             if content_id in training_satisfaction_file:
#                 item['training_satisfaction_file_ls'] = training_satisfaction_file[content_id]
#                 item['training_satisfaction_file_num'] = len(training_satisfaction_file[content_id])
#             else:
#                 item['training_satisfaction_file_ls'] = []
#                 item['training_satisfaction_file_num'] = 0
#
#         for index, item in enumerate(tableList):
#             item['index'] = (currentPage - 1) * pageSize + index + 1
#             item['content_begin_date'] = item['content_begin_date'] if item['content_begin_date'] is None else str(item['content_begin_date'])
#             item['content_end_date'] = item['content_end_date'] if item['content_end_date'] is None else str(item['content_end_date'])
#             if item['content_lecturer__lecturer_people__employee_name'] is None and item['content_lecturer__lecturer_name'] is not None :
#                 item['content_lecturer__lecturer_people__employee_name']=item['content_lecturer__lecturer_name']
#
#
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "信息返回成功",
#             "data": {
#                 'columnList': columnList,
#                 'tableList': tableList,
#                 'totalNumber': totalNumber,
#             }
#         }
#
#     def post_content_info(self):    #输入内部讲师 讲师姓名  产生工号 岗位 讲师级别      外部讲师 工号 岗位 界别为空
#         # try:
#             info=self.request.POST.get('createData',None)
#             info = json.loads(info)
#             if len(info['nodePath'])>0:
#                 info['content_part_id'] = info['nodePath'][-1]
#             else:
#                 info['content_part_id'] =None
#
#             if info['content_begin_date'] =='':
#                 info['content_begin_date']=None
#             if info['content_end_date'] =='':
#                 info['content_end_date']=None
#             if info['content_satisfaction'] =='':
#                 info['content_satisfaction']=None
#             file1_ls = self.request.FILES.getlist('createPhoto')  # 培训照片
#             file2_ls = self.request.FILES.getlist('createFile')  # 附件
#             file3_ls = self.request.FILES.getlist('signinFile')  # 签到表
#             file4_ls = self.request.FILES.getlist('satisfactionFile')  # 培训满意度
#
#
#             dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', str(self.t1),str(info['content_title']))  # 创建文件夹
#             self.mkdir(dummy_path)
#
#
#             if info['lecturer_type']=='内部讲师':
#                 employee_obj = HrEmployee.objects.filter(employee_code=info['lecturer_code'], employee_name=info['content_lecturer'],employee_status=1).values_list('id',flat=True)
#                 if employee_obj.exists():
#                     lecturer_info={
#                         'lecturer_people_id':employee_obj[0],
#                         'lecturer_type':info['lecturer_type'],
#                         'lecturer_level_id':info['lecturer_level_id'],
#                         'lecturer_modifier_id' : self.request.check_token
#                     }
#
#                     try:
#                         lecturer_obj = TrainingLecturer.objects.filter(lecturer_people_id=lecturer_info['lecturer_people_id'], lecturer_status=True)
#                         if lecturer_obj.exists():  #讲师库有该条数据
#                             del info['nodePath']
#                             del info['content_section']
#                             del info['lecturer_type']
#                             del info['content_lecturer']
#                             del info['lecturer_code']
#                             del info['post']
#                             del info['lecturer_level_id']
#                             del info['content_part']
#                             info['content_lecturer_id'] = lecturer_obj[0].id
#                             info['content_creater_id'] = self.request.check_token
#                             content_obj = TrainingContent.objects.create(**info)
#                             for file_obj in file1_ls:
#                                 file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                               str(random.random())[
#                                                                               -5:] + '_培训照片')
#                                 self.saveFile(file_url, file_obj)  # 保存文件
#                                 file_kwargs = {
#                                     'training_file_name': file_name,
#                                     'training_file_url': file_url,
#                                     'training_file_type': 1,
#                                     'training_content_file_id': content_obj.id
#                                 }
#
#                                 file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                             for file_obj in file2_ls:
#                                 file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                               str(random.random())[
#                                                                               -5:] + '_培训课件')
#                                 self.saveFile(file_url, file_obj)  # 保存文件
#                                 file_kwargs = {
#                                     'training_file_name': file_name,
#                                     'training_file_url': file_url,
#                                     'training_file_type': 2,
#                                     'training_content_file_id': content_obj.id
#                                 }
#
#                                 file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                             for file_obj in file3_ls:
#                                 file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                               str(random.random())[-5:] + '_签到表')
#                                 self.saveFile(file_url, file_obj)  # 保存文件
#                                 file_kwargs = {
#                                     'training_file_name': file_name,
#                                     'training_file_url': file_url,
#                                     'training_file_type': 3,
#                                     'training_content_file_id': content_obj.id
#                                 }
#                                 # print('file2', file_kwargs)
#                                 file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#
#                                 try:
#                                     exc = openpyxl.load_workbook(file_url, data_only=True)
#                                     sheet = exc.active
#                                     for line in range(1, sheet.max_row):  # 每行数据
#                                         checkin_info = {}
#                                         name_code = sheet.cell(line + 1, 1).value
#                                         name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")",
#                                                                                                               "")
#                                         employee_obj = HrEmployee.objects.filter(employee_code=code,
#                                                                                  employee_status=1).values_list(
#                                             'id',
#                                             flat=True)
#                                         if employee_obj.exists():
#                                             checkin_info['checkin_people_id'] = employee_obj[0]
#                                             checkin_info['checkin_content_id'] = content_obj.id
#                                             checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
#                                             checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
#                                             checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
#                                             checkin_info['checkin_associated_files_id'] = file_dbobj.id
#                                             TrainingCheckin.objects.update_or_create(defaults=checkin_info,
#                                                                                      checkin_people_id=checkin_info[
#                                                                                          'checkin_people_id'],
#                                                                                      checkin_content_id=
#                                                                                      checkin_info[
#                                                                                          'checkin_content_id'],
#                                                                                      checkin_time=checkin_info[
#                                                                                          'checkin_time'],
#                                                                                      checkin_address=checkin_info[
#                                                                                          'checkin_address'],
#                                                                                      checkin_status=True,
#
#                                                                                      )
#                                         else:  # 离职或不存在
#                                             continue
#                                 except:
#                                     self.return_data = {
#                                         "code": status.HTTP_401_UNAUTHORIZED,
#                                         "msg": "该签到表不是excel或者其他错误，无法新增",
#                                     }
#                             for file_obj in file4_ls:
#                                 file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                               str(random.random())[
#                                                                               -5:] + '_培训满意度')
#                                 self.saveFile(file_url, file_obj)  # 保存文件
#                                 file_kwargs = {
#                                     'training_file_name': file_name,
#                                     'training_file_url': file_url,
#                                     'training_file_type': 4,
#                                     'training_content_file_id': content_obj.id
#                                 }
#                                 file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#
#                             self.return_data = {
#                                 "code": status.HTTP_200_OK,
#                                 "msg": "新增成功",
#                             }
#                         else:    #讲师库没有该条数据 那么不可以新增   除非是无
#                             if lecturer_info['lecturer_level_id']==5:   #如果是无那么先创建讲师,在新增报表数据
#                                 lecturer_obj = TrainingLecturer.objects.create(**lecturer_info)
#                                 del info['nodePath']
#                                 del info['content_section']
#                                 del info['lecturer_type']
#                                 del info['content_lecturer']
#                                 del info['lecturer_code']
#                                 del info['post']
#                                 del info['lecturer_level_id']
#                                 del info['content_part']
#                                 info['content_lecturer_id'] = lecturer_obj.id
#                                 info['content_creater_id'] = self.request.check_token
#                                 content_obj = TrainingContent.objects.create(**info)
#                                 for file_obj in file1_ls:
#                                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                                   str(random.random())[
#                                                                                   -5:] + '_培训照片')
#                                     self.saveFile(file_url, file_obj)  # 保存文件
#                                     file_kwargs = {
#                                         'training_file_name': file_name,
#                                         'training_file_url': file_url,
#                                         'training_file_type': 1,
#                                         'training_content_file_id': content_obj.id
#                                     }
#
#                                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                                 for file_obj in file2_ls:
#                                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                                   str(random.random())[
#                                                                                   -5:] + '_培训课件')
#                                     self.saveFile(file_url, file_obj)  # 保存文件
#                                     file_kwargs = {
#                                         'training_file_name': file_name,
#                                         'training_file_url': file_url,
#                                         'training_file_type': 2,
#                                         'training_content_file_id': content_obj.id
#                                     }
#
#                                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                                 for file_obj in file3_ls:
#                                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                                   str(random.random())[-5:] + '_签到表')
#                                     self.saveFile(file_url, file_obj)  # 保存文件
#                                     file_kwargs = {
#                                         'training_file_name': file_name,
#                                         'training_file_url': file_url,
#                                         'training_file_type': 3,
#                                         'training_content_file_id': content_obj.id
#                                     }
#                                     # print('file2', file_kwargs)
#                                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#
#                                     try:
#                                         exc = openpyxl.load_workbook(file_url, data_only=True)
#                                         sheet = exc.active
#                                         for line in range(1, sheet.max_row):  # 每行数据
#                                             checkin_info = {}
#                                             name_code = sheet.cell(line + 1, 1).value
#                                             name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")",
#                                                                                                                   "")
#                                             employee_obj = HrEmployee.objects.filter(employee_code=code,
#                                                                                      employee_status=1).values_list(
#                                                 'id',
#                                                 flat=True)
#                                             if employee_obj.exists():
#                                                 checkin_info['checkin_people_id'] = employee_obj[0]
#                                                 checkin_info['checkin_content_id'] = content_obj.id
#                                                 checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
#                                                 checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
#                                                 checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
#                                                 checkin_info['checkin_associated_files_id'] = file_dbobj.id
#                                                 TrainingCheckin.objects.update_or_create(defaults=checkin_info,
#                                                                                          checkin_people_id=checkin_info[
#                                                                                              'checkin_people_id'],
#                                                                                          checkin_content_id=
#                                                                                          checkin_info[
#                                                                                              'checkin_content_id'],
#                                                                                          checkin_time=checkin_info[
#                                                                                              'checkin_time'],
#                                                                                          checkin_address=checkin_info[
#                                                                                              'checkin_address'],
#                                                                                          checkin_status=True,
#
#                                                                                          )
#                                             else:  # 离职或不存在
#                                                 continue
#                                     except:
#                                         self.return_data = {
#                                             "code": status.HTTP_401_UNAUTHORIZED,
#                                             "msg": "该签到表不是excel或者其他错误，无法新增",
#                                         }
#                                 for file_obj in file4_ls:
#                                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                                   str(random.random())[
#                                                                                   -5:] + '_培训满意度')
#                                     self.saveFile(file_url, file_obj)  # 保存文件
#                                     file_kwargs = {
#                                         'training_file_name': file_name,
#                                         'training_file_url': file_url,
#                                         'training_file_type': 4,
#                                         'training_content_file_id': content_obj.id
#                                     }
#                                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#
#                                 self.return_data = {
#                                     "code": status.HTTP_200_OK,
#                                     "msg": "新增成功",
#                                 }
#                             else:
#                                 self.return_data = {
#                                     "code": status.HTTP_401_UNAUTHORIZED,
#                                     "msg": "该内部讲师未加入讲师库,无法新增",
#                                 }
#                     except:
#                         self.return_data = {
#                             "code": status.HTTP_401_UNAUTHORIZED,
#                             "msg": "该讲师相关信息错误,无法新增",
#                         }
#
#                 else:
#                     self.return_data = {
#                         "code": status.HTTP_401_UNAUTHORIZED,
#                         "msg": "不是内部员工或已离职,无法新增",
#                     }
#             elif info['lecturer_type']=='外部讲师':
#
#                 lecturer_info = {
#                     'lecturer_name': info['content_lecturer'],   #讲师姓名
#                     'lecturer_type': info['lecturer_type']
#                 }
#                 lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer_info,lecturer_type=lecturer_info['lecturer_type'],lecturer_name=lecturer_info['lecturer_name'],lecturer_status=True)
#                 content_info={
#                     "content_lecturer_id":lecturer_obj[0].id,
#                     'content_title':info['content_title'],
#                     'content_part_id':info['content_part_id'],
#                     'content_module':info['content_module'],
#                     'content_group':info['content_group'],
#                     'content_type_id':info['content_type_id'],
#                     'content_category_id':info['content_category_id'],
#                     'content_level_id':info['content_level_id'],
#                     'content_manner':info['content_manner'],
#                     'content_begin_date':info['content_begin_date'],
#                     'content_end_date':info['content_end_date'],
#                     'content_duration':info['content_duration'],
#                     'content_object':info['content_object'],
#                     'content_plan':info['content_plan'],
#                     'content_people_number':info['content_people_number'],
#                     'content_satisfaction':info['content_satisfaction'],
#                     'content_expenses':info['content_expenses'],
#                     'content_creater_id' : self.request.check_token
#                 }
#                 content_obj=TrainingContent.objects.create(**content_info)
#                 for file_obj in file1_ls:
#                     TrainingFiles.objects.filter(training_content_file=info['id'], file_type=1).update(
#                         is_valid=False)
#                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                   str(random.random())[-5:] + '_培训照片')
#                     self.saveFile(file_url, file_obj)  # 保存文件
#                     file_kwargs = {
#                         'training_file_name': file_name,
#                         'training_file_url': file_url,
#                         'training_file_type': 1,
#                         'training_content_file_id': content_obj.id
#                     }
#                     # print('file1',file_kwargs)
#                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                 for file_obj in file2_ls:
#                     TrainingFiles.objects.filter(activity_file_id=info['id'], file_type=2).update(is_valid=False)
#                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                   str(random.random())[-5:] + '_培训课件')
#                     self.saveFile(file_url, file_obj)  # 保存文件
#                     file_kwargs = {
#                         'training_file_name': file_name,
#                         'training_file_url': file_url,
#                         'training_file_type': 2,
#                         'training_content_file_id':content_obj.id
#                     }
#                     # print('file2',file_kwargs)
#                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                 for file_obj in file3_ls:
#                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                   str(random.random())[-5:] + '_签到表')
#                     self.saveFile(file_url, file_obj)  # 保存文件
#                     file_kwargs = {
#                         'training_file_name': file_name,
#                         'training_file_url': file_url,
#                         'training_file_type': 3,
#                         'training_content_file_id': content_obj.id
#                     }
#                     # print('file3', file_kwargs)
#                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#                     try:
#                         exc = openpyxl.load_workbook(file_url, data_only=True)
#                         sheet = exc.active
#                         for line in range(1, sheet.max_row):  # 每行数据
#                             checkin_info = {}
#                             name_code = sheet.cell(line + 1, 1).value
#                             name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")", "")
#                             employee_obj = HrEmployee.objects.filter(employee_code=code,
#                                                                      employee_status=1).values_list('id', flat=True)
#                             if employee_obj.exists():
#                                 checkin_info['checkin_people_id'] = employee_obj[0]
#                                 checkin_info['checkin_content_id'] = content_obj.id
#                                 checkin_info['checkin_time'] = sheet.cell(line + 1, 3).value
#                                 checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
#                                 checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
#                                 checkin_info['checkin_associated_files_id'] = file_dbobj.id
#                                 TrainingCheckin.objects.update_or_create(defaults=checkin_info,
#                                                                          checkin_people_id=checkin_info[
#                                                                              'checkin_people_id'],
#                                                                          checkin_content_id=checkin_info[
#                                                                              'checkin_content_id'],
#                                                                          checkin_time=checkin_info['checkin_time'],
#                                                                          checkin_address=checkin_info[
#                                                                              'checkin_address'],
#                                                                          checkin_status=True
#                                                                          )
#                             else:  # 离职或不存在
#                                 continue
#                     except:
#                         self.return_data = {
#                             "code": status.HTTP_401_UNAUTHORIZED,
#                             "msg": "该签到表不是excel或者其他错误，无法新增",
#                         }
#                 for file_obj in file4_ls:
#                     file_url, file_name, suffix = self.createPath(file_obj, str(info['content_title']),
#                                                                   str(random.random())[-5:] + '_培训满意度')
#                     self.saveFile(file_url, file_obj)  # 保存文件
#                     file_kwargs = {
#                         'training_file_name': file_name,
#                         'training_file_url': file_url,
#                         'training_file_type': 4,
#                         'training_content_file_id': content_obj.id
#                     }
#                     file_dbobj = TrainingFiles.objects.create(**file_kwargs)
#
#
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "新增成功",
#                 }
#
#             else:
#                 self.return_data = {
#                     "code": status.HTTP_401_UNAUTHORIZED,
#                     "msg": "请填写讲师类型",
#                 }
#         # except:
#         #     self.return_data = {
#         #         "code": status.HTTP_401_UNAUTHORIZED,
#         #         "msg": "数据不完整，新增失败",
#         #     }
#     def edit_content_info(self):
#         info=json.loads(self.request.body)
#
#         content_id=info['id']     #培训表id
#         lecturer_id=info['content_lecturer'] #讲师表id
#         #内部--内部   内部--外部     外部--内部   外部 --外部
#         if 'nodePath' in info:  # 修改中心了
#             if len(info['nodePath']) >= 2:
#                 info['content_part_id'] = info['nodePath'][1]
#             else:
#                 info['content_part_id'] = None
#         else:  #没有修改
#             pass
#
#         if info['content_lecturer__lecturer_type'] == '内部讲师':  # 新数据
#             employee_obj = HrEmployee.objects.filter(employee_code=info['content_lecturer__lecturer_people__employee_code'],
#                                                      employee_name=info['content_lecturer__lecturer_people__employee_name']).values_list('id',flat=True)
#             lecturer_info = {
#                 'lecturer_people_id': employee_obj[0],
#                 'lecturer_type': info['content_lecturer__lecturer_type'],
#                 'lecturer_level_id': info['content_lecturer__lecturer_level_id'],
#                 'lecturer_creater_id': self.request.check_token
#             }
#
#
#             lecturer_obj = TrainingLecturer.objects.filter(lecturer_people_id=lecturer_info['lecturer_people_id'],lecturer_status=True)   #讲师对象
#
#             if lecturer_obj.exists():  # 讲师库有该条数据  级别是无 不修改   级别不是无 不修改
#                 content_info = {
#                     "content_lecturer_id": lecturer_obj[0].id,
#                     'content_title': info['content_title'],
#                     'content_part_id': info['content_part_id'],
#                     'content_module': info['content_module'],
#                     'content_group': info['content_group'],
#                     'content_type_id': info['content_type_id'],
#                     'content_category_id': info['content_category_id'],
#                     'content_level_id': info['content_level_id'],
#                     'content_manner': info['content_manner'],
#                     'content_begin_date': info['content_begin_date'],
#                     'content_end_date': info['content_end_date'],
#                     'content_duration': info['content_duration'],
#                     'content_object': info['content_object'],
#                     'content_plan': info['content_plan'],
#                     'content_people_number': info['content_people_number'],
#                     'content_satisfaction': info['content_satisfaction'],
#                     'content_expenses': info['content_expenses'],
#                     'content_modifier_id': self.request.check_token
#                 }
#                 TrainingContent.objects.filter(pk=content_id).update(**content_info)
#             else: #讲师库没有该条数据
#                 if lecturer_info['lecturer_level_id']==5:#是无创建
#                     lecturer_obj = TrainingLecturer.objects.create(**lecturer_info)
#                     del info['nodePath']
#                     del info['content_section']
#                     del info['lecturer_type']
#                     del info['content_lecturer']
#                     del info['lecturer_code']
#                     del info['post']
#                     del info['lecturer_level_id']
#                     del info['content_part']
#                     info['content_lecturer_id'] = lecturer_obj.id
#                     info['content_creater_id'] = self.request.check_token
#                     content_obj = TrainingContent.objects.create(**info)
#
#                     self.return_data = {
#                         "code": status.HTTP_200_OK,
#                         "msg": "新增成功",
#                     }
#                 else:#不是提示无法新增
#                     self.return_data = {
#                         "code": status.HTTP_401_UNAUTHORIZED,
#                         "msg": "该内部讲师未加入讲师库,无法新增",
#                     }
#
#
#
#
#         elif info['content_lecturer__lecturer_type'] == '外部讲师':#新数据是外部讲师
#             lecturer_info = {
#                 'lecturer_name': info['content_lecturer__lecturer_people__employee_name'],   #讲师姓名
#                 'lecturer_type': info['content_lecturer__lecturer_type'],    #讲师类型
#                 'lecturer_people_id':None,
#                 'lecturer_creater_id': self.request.check_token
#             }
#             # print(lecturer_info)
#             lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer_info,
#                                                                      lecturer_type=lecturer_info['lecturer_type'],
#                                                                      lecturer_name=lecturer_info['lecturer_name'])
#             # print(lecturer_obj)
#             content_info = {
#                 "content_lecturer_id": lecturer_obj[0].id,
#                 'content_title': info['content_title'],
#                 'content_part_id': info['content_part_id'],
#                 'content_module': info['content_module'],
#                 'content_group': info['content_group'],
#                 'content_type_id': info['content_type_id'],
#                 'content_category_id': info['content_category_id'],
#                 'content_level_id': info['content_level_id'],
#                 'content_manner': info['content_manner'],
#                 'content_begin_date': info['content_begin_date'],
#                 'content_end_date': info['content_end_date'],
#                 'content_duration': info['content_duration'],
#                 'content_object': info['content_object'],
#                 'content_plan': info['content_plan'],
#                 'content_people_number': info['content_people_number'],
#                 'content_satisfaction': info['content_satisfaction'],
#                 'content_expenses': info['content_expenses'],
#                 'content_creater_id': self.request.check_token
#             }
#
#             TrainingContent.objects.filter(pk=content_id).update(**content_info)
#         else:
#             self.return_data = {
#                 "code": status.HTTP_401_UNAUTHORIZED,
#                 "msg": "请选择讲师类型",
#             }
#
#
#
#
#     def delete_content_info(self):
#         info=json.loads(self.request.body)
#         for id in info['idList']:
#             TrainingContent.objects.filter(pk=id).update(content_status=False)
#             self.return_data = {
#                 "code": status.HTTP_200_OK,
#                 "msg": "删除成功",
#             }
#
#     def download_content_info(self):
#         now = arrow.now()
#         t1 = now.format('YYYY-MM-DD')
#         t2 = now.timestamp()
#         dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
#         self.mkdir(dummy_path)
#         id_list = json.loads(self.request.body).get('idList')
#         downloadAll = json.loads(self.request.body).get('downloadAll')
#         file_ls = [
#             "序号", "中心/基地", "部门", "模块", "组", "培训主题/课题", "培训类型", "培训类别", "培训层级", "培训方式",
#              "开始培训日期", "截止培训日期", "培训时长(H)", "培训对象", "参训人数", "讲师类型", "培训讲师", "讲师工号",
#              "讲师岗位", "讲师级别", "培训满意度", "培训费用", "计划内/计划外", "用户昵称"
#         ]
#         path = self.createExcelPath('线下培训报表.xlsx', str(t2), '线下培训报表',25, 'A1:X1', *file_ls)
#         if downloadAll == True:  # 是下载全部   有条件
#             row_data = []
#             index = 1
#             kwargs = {"content_status": True}
#             info = json.loads(self.request.body)
#             searchName = info['searchName']
#             beginDate = info['beginDate']
#             endDate = info['endDate']
#             contentTypeList = info['contentTypeList']
#             kwargs['content_type_id__in'] = contentTypeList
#             employee_base = info['baseNameId']
#             if beginDate != "" and endDate != "":
#                 kwargs['content_begin_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
#                     beginDate) == 0 else beginDate
#                 kwargs['content_begin_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
#                     endDate) == 0 else endDate
#             kwargs['content_part_id__in'] = employee_base
#             if employee_base == '':
#                 kwargs['content_part_id__in'] = self.request.user_department_employee
#             kwargs = {key: value for key, value in kwargs.items() if
#                       value is not None and value != '' and value != []}  # 过滤掉值为None或''的项
#
#             tableList = TrainingContent.objects.filter(Q(content_title__contains=searchName) | Q(content_lecturer__lecturer_people__employee_name__contains=searchName)| Q(content_lecturer__lecturer_people__employee_code__contains=searchName),**kwargs).values_list(
#                         'content_part__department_first_name',
#                     "content_part__department_second_name",
#                     "content_part__department_third_name",
#                     "content_part__department_forth_name",
#                     'content_title',
#                     'content_type__type_name',
#                     'content_category__category_name',
#                     'content_level__level_name',
#                     'content_manner',
#                     'content_begin_date',
#                     'content_end_date',
#                     'content_duration',
#                     'content_object',
#                     'content_people_number',#参训人数
#                     'content_lecturer__lecturer_type',#讲师类型
#                     'content_lecturer__lecturer_name',  # 培训讲师(外）
#                     'content_lecturer__lecturer_people__employee_name',    #培训讲师（内）
#                     'content_lecturer__lecturer_people__employee_code',  #讲师工号
#                     'content_lecturer__lecturer_people__employee_position__position_name',
#                     'content_lecturer__lecturer_level__level_name',
#                     'content_satisfaction',
#                     'content_expenses',
#                     'content_plan',
#                     'content_creater__user',
#             ).order_by('-content_createTime')
#
#             for line in tableList:
#                 line = list(line)
#                 line.insert(0, index)
#                 if line[16] is None:  #内部讲师
#                     line.pop(16)
#                 elif line[16] is not None:  #外部讲师
#                     line.pop(17)
#                 row_data.append(line)
#                 if len(line) == 0:
#                     index = index
#                 index += 1
#
#             exc = openpyxl.load_workbook(path)  # 打开整个excel文件
#             sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
#             for row in row_data:
#                 sheet.append(row)  # 在工作表中添加一行
#             exc.save(path)  # 指定路径,保存文件
#         else:
#             row_data = []
#             index = 1
#             for id in id_list:
#                 data = list(TrainingContent.objects.filter(pk=id, content_status=True).values_list(
#                     'content_part__department_first_name',
#                     "content_part__department_second_name",
#                     "content_part__department_third_name",
#                     "content_part__department_forth_name",
#                     'content_title',
#                     'content_type__type_name',
#                     'content_category__category_name',
#                     'content_level__level_name',
#                     'content_manner',
#                     'content_begin_date',
#                     'content_end_date',
#                     'content_duration',
#                     'content_object',
#                     'content_people_number',#参训人数
#                     'content_lecturer__lecturer_type',#讲师类型
#                     'content_lecturer__lecturer_name',  # 培训讲师(外）
#                     'content_lecturer__lecturer_people__employee_name',    #培训讲师（内）
#                     'content_lecturer__lecturer_people__employee_code',  #讲师工号
#                     'content_lecturer__lecturer_people__employee_position__position_name',
#                     'content_lecturer__lecturer_level__level_name',
#                     'content_satisfaction',
#                     'content_expenses',
#                     'content_plan',
#                     'content_creater__user',
#                 ))[0]
#                 data = (index,) + data
#                 data = list(data)
#                 if data[16] is None:  #内部讲师
#                     data.pop(16)
#                 elif data[16] is not None:  #外部讲师
#                     data.pop(17)
#                 row_data.append(data)
#                 if len(data) == 0:
#                     index = index
#                 index += 1
#             exc = openpyxl.load_workbook(path)  # 打开整个excel文件
#             sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
#             for row in row_data:
#                 sheet.append(row)  # 在工作表中添加一行
#             exc.save(path)  # 指定路径,保存文件
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "下载成功",
#             "downloadUrl": path
#         }
#
#     def batch_content_info(self):
#         # try:
#             file = self.request.FILES.get("file", None)
#
#             now = arrow.now()
#             t1 = now.format('YYYY-MM-DD')
#             t2 = now.format('YYYY-MM-DD_HH_mm_ss')
#             dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', t1,
#                                       '培训报表文件上传')  # 创建文件夹
#             self.mkdir(dummy_path)
#             file_url, file_name, file_suffix = self.createPath(file, '培训报表文件上传', '培训报表' + str(t2))
#             self.saveFile(file_url, file)
#             exc = openpyxl.load_workbook(file_url, data_only=True)
#             sheet = exc.active
#             # try:
#             for i in range(1, sheet.max_row):  # 每行数据
#
#                 content={}
#                 lecturer={}
#
#                 try:
#                     first_name=None if sheet.cell(i + 1, 2).value =='' else sheet.cell(i + 1, 2).value
#                     second_name=None if sheet.cell(i + 1, 3).value =='' else sheet.cell(i + 1,3).value
#                     third_name=None if sheet.cell(i + 1, 4).value =='' else sheet.cell(i + 1, 4).value
#                     forth_name=None if sheet.cell(i + 1, 5).value =='' else sheet.cell(i + 1,5).value
#                     # print(first_name,second_name,third_name,forth_name)
#                     if first_name is not None and second_name is not None and third_name is not None and forth_name is not None:  #四级部门
#                         content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
#                                                         department_second_name=second_name,
#                                                         department_third_name=third_name,
#                                                         department_forth_name=forth_name).values_list('id',flat=True)[0]
#                     elif first_name is not None and second_name is not None and third_name is not None and forth_name is None: #三级部门
#                         content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
#                                                                                  department_second_name=second_name,
#                                                                                  department_third_name=third_name,).values_list('id', flat=True)[0]
#                     elif first_name is not None and second_name is not None and third_name is  None and forth_name is None: #二级部门
#                         content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name,
#                                                                                  department_second_name=second_name).values_list('id', flat=True)[0]
#                     elif first_name is not None and second_name is  None and third_name is None and forth_name is None: #一级部门
#                         content['content_part_id'] = HrDepartment.objects.filter(department_first_name=first_name).values_list('id', flat=True)[0]
#                     # print(content)
#
#                 except:
#                     content['content_part_id']=None
#
#                 # content['content_module'] = sheet.cell(i + 1,4).value
#                 # content['content_group']= sheet.cell(i + 1, 5).value
#                 content['content_title']=sheet.cell(i + 1, 6).value   #课题
#                 try:
#                     content['content_type_id']=TrainingContentType.objects.filter(type_name=sheet.cell(i + 1,7).value ).values_list('id',flat=True)[0]
#                 except:
#                     content['content_type_id']=None
#
#                 try:
#                     content['content_category_id']=TrainingContentCategory.objects.filter(category_name=sheet.cell(i + 1,8).value).values_list('id',flat=True)[0]
#                 except:
#                     content['content_category_id']=None
#
#                 try:
#                     content['content_level_id']=TrainingContentLevel.objects.filter(level_name=sheet.cell(i + 1, 9).value ).values_list('id',flat=True)[0]
#                 except:
#                     content['content_level_id'] =None
#
#                 content['content_manner']=sheet.cell(i + 1, 10).value
#                 content['content_begin_date']=sheet.cell(i + 1, 11).value
#
#                 if content['content_begin_date'] is not None:
#                     if type(content['content_begin_date']) == datetime or type(content['content_begin_date']) == date:
#                         content['content_begin_date'] = content['content_begin_date']
#                     elif type(content['content_begin_date'] )==str:
#                         try:
#                             content['content_begin_date']=datetime.strptime(content['content_begin_date'] ,"%Y-%m-%d %H:%M:%S")
#                         except:
#                             content['content_begin_date'] = None
#                     else:
#                         content['content_begin_date'] = None
#
#
#                 content['content_end_date'] = sheet.cell(i + 1, 12).value
#                 if content['content_end_date'] is not None:
#                     if type(content['content_end_date']) == datetime or type(content['content_end_date']) == date:
#                         content['content_end_date'] = content['content_end_date']
#                     elif type(content['content_end_date'] )==str:
#                         try:
#                             content['content_end_date']=datetime.strptime(content['content_end_date'] ,"%Y-%m-%d %H:%M:%S")
#                         except:
#                             content['content_end_date'] = None
#                     else:
#                         content['content_end_date'] = None
#                         # self.return_data = {
#                         #     "code": status.HTTP_401_UNAUTHORIZED,
#                         #     "msg": "培训日期格式错误，无法上传，可以将单元格改为日期格式,例如yyyy/m/d h:mm:ss 即 2023/7/17 17:30:59"
#                         # }
#                 # print(type(content['content_end_date']),type(content['content_begin_date']))
#                 if content['content_end_date'] is not None and content['content_begin_date'] is not None:
#                     if  content['content_end_date']>=content['content_begin_date']:
#                         # 计算时间差
#                         time_difference = content['content_end_date'] - content['content_begin_date']
#                         # 提取时间差的总秒数
#                         total_seconds = time_difference.total_seconds()
#                         # 将秒数转换为小时
#                         hours_difference = total_seconds / 3600
#                         content['content_duration'] =round(hours_difference,1)
#                     else:
#                         content['content_duration']=None
#                         content['content_end_date'] = None
#                         content['content_begin_date'] = None
#
#                 else:
#                     content['content_duration'] = None
#                 # content['content_duration']=sheet.cell(i + 1, 13).value#培训时长
#                 content['content_object']=sheet.cell(i + 1, 14).value
#                 content['content_people_number']=sheet.cell(i + 1, 15).value #培训人数
#                 content['content_satisfaction'] = sheet.cell(i + 1, 21).value if sheet.cell(i + 1, 21).value != '' or sheet.cell(i + 1, 21).value is not None else None  # 培训满意度
#                 content['content_expenses']= sheet.cell(i + 1, 22).value#培训费用
#                 content['content_plan']= sheet.cell(i + 1, 23).value#  计划内、计划外
#
#                 try:
#                     content['content_creater_id']= AdminUser.objects.filter(username=sheet.cell(i + 1, 24).value).values_list('id',flat=True)[0]      #创建人
#                 except:
#                     content['content_creater_id']=None
#                 lecturer['lecturer_type']=sheet.cell(i + 1, 16).value#讲师类型
#                 try:
#                     lecturer['lecturer_level_id'] = TrainingLecturerLevel.objects.filter(level_name=sheet.cell(i + 1, 20).value).values_list('id', flat=True)[0] if sheet.cell(i + 1, 20).value!=''  else None  #讲师级别
#                 except:
#                     lecturer['lecturer_level_id'] =None
#                     # self.return_data = {
#                     #     "code": status.HTTP_401_UNAUTHORIZED,
#                     #     "msg": "讲师级别不在一，二，三，荣誉讲师范围内"
#                     # }
#                 print(lecturer['lecturer_level_id'])
#                 if lecturer['lecturer_type']=='内部讲师':
#                         print('内部讲师')
#                     # try:
#                         lecturer['lecturer_people_id']=HrEmployee.objects.filter(employee_code=sheet.cell(i + 1, 18).value).values_list('id', flat=True)[0] if sheet.cell(i + 1, 18).value!= ''else None
#                         lecturer['lecturer_modifier_id']=self.request.check_token
#
#                         lecturer_obj = TrainingLecturer.objects.filter(lecturer_people_id=lecturer['lecturer_people_id'], lecturer_status=True).first()  # 讲师对象
#                         print(lecturer_obj)
#                         if lecturer_obj:  # 讲师库有该条数据  级别是无 不修改   级别不是无 不修改
#                             content['content_lecturer_id'] = lecturer_obj.id
#                         else:  # 讲师库没有该条数据
#
#                             if lecturer['lecturer_level_id'] == 5:  # 是无创建
#                                 print('65656565656')
#                                 print(lecturer)
#                                 lecturer_obj = TrainingLecturer.objects.create(**lecturer)
#                                 print('lecturer_obj')
#                                 content['content_lecturer_id'] = lecturer_obj.id
#                             else:  # 不是提示无法新增
#                                 # lecturer_obj=None
#                                 content['content_lecturer_id'] = None
#
#                         # content['content_lecturer_id'] = lecturer_obj.id
#                     # except:
#                     #     content['content_lecturer_id'] =None
#                         # self.return_data = {
#                         #     "code": status.HTTP_401_UNAUTHORIZED,
#                         #     "msg": "内部讲师中有讲师不是公司员工，无法新增"
#                         # }
#
#
#                 elif lecturer['lecturer_type']=='外部讲师':
#                     lecturer['lecturer_name']=sheet.cell(i + 1, 17).value    #讲师姓名
#                     lecturer['lecturer_modifier_id'] = self.request.check_token
#                     lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer,
#                                                                              lecturer_type=lecturer['lecturer_type'],
#                                                                              lecturer_name=lecturer['lecturer_name'],
#                                                                              lecturer_status=True)
#                     content['content_lecturer_id'] = lecturer_obj[0].id
#                 else:
#                     content['content_lecturer_id']=None
#
#                 # try:
#
#                 if all(value is None for value in content.values()):   #全为None
#                     pass
#                 else:
#                     print('1111111111111111111')
#                     print('2', content)
#                     TrainingContent.objects.update_or_create(defaults=content,content_part_id=content['content_part_id'],content_begin_date=content['content_begin_date'],content_people_number=content['content_people_number'],content_status=True,content_type_id=content['content_type_id'],content_category_id=content['content_category_id'],content_lecturer_id=content['content_lecturer_id'])
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "上传成功",
#                 }
#                 # # except:
#                 # #     pass
#                 # #     self.return_data = {
#                 # #         "code": status.HTTP_401_UNAUTHORIZED,
#                 # #         "msg": "上传失败，数据错误",
#                 # #     }
#
#                 # print(self.return_data)
#         # except:
#         #     self.return_data = {
#         #         "code": status.HTTP_401_UNAUTHORIZED,
#         #         "msg": "上传异常",
#         #     }
#
#     def del_content_file(self):
#         info = json.loads(self.request.body)
#         file_obj = TrainingFiles.objects.filter(id=info['file_id'], training_file_status=True).first()
#         if file_obj:
#             TrainingFiles.objects.filter(pk=info['file_id'],training_file_status=True).update(training_file_status=False)
#             if file_obj.training_file_type==3:  #删除文件的同时把文件数据也删除
#                 TrainingCheckin.objects.filter(checkin_associated_files_id=file_obj.id,checkin_status=True).update(checkin_status=False,checkin_modifier_id=self.request.check_token)
#
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "删除成功",
#         }
#     def post_content_file(self):
#         info=dict(self.request.POST)
#         content_id=info['id'][0]
#         print(info)
#         content_obj=TrainingContent.objects.filter(pk=content_id,content_status=True).first()
#         print(content_obj)
#         content_title=content_obj.content_title
#         dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'upload_file', str(self.t1),str(content_title))  # 创建文件夹
#         self.mkdir(dummy_path)
#         type = info['field'][0]
#         if type=='createFile_ls':  #培训附件
#             file_ls=self.request.FILES.getlist('file')
#             for file_obj in file_ls:
#                 file_url, file_name, suffix = self.createPath(file_obj, str(content_title),str(random.random())[-5:] + '_培训课件')
#                 self.saveFile(file_url, file_obj)  # 保存文件
#                 file_kwargs = {
#                     'training_file_name': file_name,
#                     'training_file_url': file_url,
#                     'training_file_type': 2,
#                     'training_content_file_id': content_obj.id
#                 }
#                 file_dbobj2 = TrainingFiles.objects.create(**file_kwargs)
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "附件新增成功",
#                     'data': {
#                         'id': file_dbobj2.id,
#                         'name': file_dbobj2.training_file_name,
#                         'url': file_dbobj2.training_file_url,
#                     }
#                 }
#
#         elif type=='createPhoto_ls':#培训照片
#             file_ls=self.request.FILES.getlist('file')
#             for file_obj in file_ls:
#                 file_url, file_name, suffix = self.createPath(file_obj, str(content_title),str(random.random())[-5:] + '_培训照片')
#                 self.saveFile(file_url, file_obj)  # 保存文件
#                 file_kwargs = {
#                     'training_file_name': file_name,
#                     'training_file_url': file_url,
#                     'training_file_type': 1,
#                     'training_content_file_id': content_obj.id
#                 }
#                 file_dbobj1 = TrainingFiles.objects.create(**file_kwargs)
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "照片新增成功",
#                     'data':{
#                         'id':file_dbobj1.id,
#                         'name':file_dbobj1.training_file_name,
#                         'url': file_dbobj1.training_file_url,
#                     }
#                 }
#
#         elif type == 'satisfactionFile':  # 满意度
#             file_ls = self.request.FILES.getlist('file')
#             for file_obj in file_ls:
#                 file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
#                                                               str(random.random())[-5:] + '_培训满意度')
#                 self.saveFile(file_url, file_obj)  # 保存文件
#                 file_kwargs = {
#                     'training_file_name': file_name,
#                     'training_file_url': file_url,
#                     'training_file_type': 4,
#                     'training_content_file_id': content_obj.id
#                 }
#
#                 file_dbobj4 = TrainingFiles.objects.create(**file_kwargs)
#
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "培训满意度新增成功",
#                     'data': {
#                         'id': file_dbobj4.id,
#                         'name': file_dbobj4.training_file_name,
#                         'url': file_dbobj4.training_file_url,
#                     }
#                 }
#         elif type == 'signinFile':  # 签到表
#             file_ls = self.request.FILES.getlist('file')
#             for file_obj in file_ls:
#                 file_url, file_name, suffix = self.createPath(file_obj, str(content_title),
#                                                               str(random.random())[-5:] + '_签到表')
#                 self.saveFile(file_url, file_obj)  # 保存文件
#                 file_kwargs = {
#                     'training_file_name': file_name,
#                     'training_file_url': file_url,
#                     'training_file_type': 3,
#                     'training_content_file_id': content_obj.id,
#
#                 }
#
#                 file_dbobj3= TrainingFiles.objects.create(**file_kwargs)
#                 self.return_data = {
#                     "code": status.HTTP_200_OK,
#                     "msg": "签到表新增成功",
#                     'data': {
#                         'id': file_dbobj3.id,
#                         'name': file_dbobj3.training_file_name,
#                         'url': file_dbobj3.training_file_url,
#                     }
#                 }
#
#                 exc = openpyxl.load_workbook(file_url, data_only=True)
#                 sheet = exc.active
#                 for line in range(1, sheet.max_row):  # 每行数据
#                     checkin_info = {}
#                     name_code = sheet.cell(line + 1, 1).value
#                     if name_code:
#                         name,code=self.extract_info(name_code)
#
#                         # try:
#                         #     name, code = name_code.split("(")[0], name_code.split("(")[1].replace(")", "")
#                         # except:
#                         #     name,code=None,None
#
#                         employee_obj = HrEmployee.objects.filter(employee_code=code, employee_status=1).values_list('id',flat=True)
#                         print('1111',name, code,employee_obj)
#                         if employee_obj.exists():
#                             checkin_info['checkin_people_id'] = employee_obj[0]
#                             checkin_info['checkin_content_id'] = content_id
#                             checkin_info['checkin_time'] = self.to_date(sheet.cell(line + 1, 3).value)
#                             checkin_info['checkin_address'] = sheet.cell(line + 1, 5).value
#                             checkin_info['checkin_user_type'] = sheet.cell(line + 1, 4).value
#                             checkin_info['checkin_associated_files_id'] = file_dbobj3.id
#                             checkin_info['checkin_modifier_id']=self.request.check_token
#                             # print(checkin_info['checkin_time'],__builtins__.type(checkin_info['checkin_time']))
#
#                             # print('结果',checkin_info['checkin_time'])
#                             TrainingCheckin.objects.update_or_create(defaults=checkin_info,
#                                                                      checkin_people_id=checkin_info['checkin_people_id'],
#                                                                      checkin_content_id=checkin_info['checkin_content_id'],
#                                                                      checkin_time=checkin_info['checkin_time'],
#                                                                      checkin_address=checkin_info['checkin_address'],
#                                                                      checkin_status=True
#                                                                      )
#                         else:
#                             continue
#                     else:   #离职或不存在
#                         continue
#
#         else:
#             self.return_data = {
#                 "code": status.HTTP_401_UNAUTHORIZED,
#                 "msg": "附件新增失败",
#             }
#
#     @staticmethod
#     def extract_info(input_string):
#         import re
#         # 使用正则表达式提取名字和学号
#         match = re.match(r'^(.*?)[\(|（]([\w]+)[\)|）]$', input_string)
#
#         if match:
#             name = match.group(1)
#             student_id = match.group(2)
#             return name, student_id
#         else:
#             return None, None
#
#     @staticmethod
#     def to_date(value):
#         """
#         如果value是日期或者能被转换为日期 返回日期类型 不是返回None
#         :param value:
#         :return:
#         """
#         if isinstance(value, datetime):
#             return value
#         if isinstance(value, str):
#             try:
#                 date_obj = datetime.strptime(value,'%Y/%m/%d %H:%M:%S')  # 根据日期格式修改这里
#                 return date_obj
#             except ValueError:
#                 try:
#                     print('11111',value,str(value))
#                     date_obj = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')  # 另一种日期格式
#                     return date_obj
#                 except ValueError:
#                     return None
#
#         return None
#
#     def month_summary_analysis(self):
#         from django.utils import timezone
#         # print(self.request.GET)
#         month = self.request.GET.get('month', None)
#         # print('month2',month)
#         if month != "":
#             month=month
#         else:
#             month=datetime.now().date()
#
#
#         columnList = [
#             {'value': '润阳集团','children':[
#                             {'label': 'department', 'value': '基地', 'width': 230}]
#              },
#             {'label': 'category','value': '类别', 'width': ''},
#             {'value': '培训层级(场次)', 'children':[
#                             {'label': 'content_number_senior', 'value': '高层', 'width': 130},
#                             {'label': 'content_number_middle', 'value': '中层', 'width': 130},
#                             {'label': 'content_number_grass', 'value': '基层', 'width': 130},
#                             {'label': 'content_number_synthesis', 'value': '综合', 'width': 130},
#
#                                                                ]},
#             {'value': '总分析', 'children': [
#                 {'label': 'content_number_Total', 'value': '场次', 'width': 130},
#                 {'label': 'content_people_number', 'value': '人次', 'width': 130},
#                 {'label': 'content_duration', 'value': '总时长(H)', 'width': 130},
#                 {'label': 'content_satisfaction', 'value': '平均满意度', 'width': 130},
#                 {'label': 'content_satisfaction_avg', 'value': '已评分平均满意度', 'width': 130},
#                 {'label': 'count_null_Satisfaction_Total', 'value': '未评分场次', 'width': 130},
#             ]},
#         ]
#
#
#         from django.db import connection
#         # print(month,type(month))
#         # if type(month)==str:
#         #     datetime_obj = datetime.strptime(month, "%Y-%m-%d")
#         # else:
#         #     datetime_obj = datetime.combine(month, datetime.min.time())
#         # # print(datetime_obj,type(datetime_obj))
#         #
#         # # print(datetime_obj,type(datetime_obj))
#         # current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
#         # next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
#         # print(current_month,next_month)
#         # # 构建 SQL 查询语句
#
#         sql_query = """
#             SELECT
#                 IFNULL( dc.department_first_name, 'Total' ) AS base_name,
#                 IFNULL( cat.category_name, 'Total' ) AS category_name,
#                 IFNULL( lvl.level_name, 'Total' ) AS level_name,
#                 COALESCE ( COUNT( tc.id ), 0 ) AS count,
#                 COALESCE ( SUM( tc.content_people_number ), 0 ) AS total_people,
#                 COALESCE ( SUM( tc.content_duration ), 0 ) AS total_duration,
#                 COALESCE ( SUM( tc.content_satisfaction ), 0 ) AS total_satisfaction,
#             CASE
#
#                     WHEN SUM(
#                     CASE
#
# 				WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' )
# 				AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND
# 				AND tc.content_status <> 0
# 				AND tc.content_satisfaction IS NOT NULL THEN
# 					1 ELSE 0
# 				END
# 					) > 0 THEN
# 					COALESCE (
# 						SUM(
# 						CASE
#
# 								WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' )
# 								AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND
# 								AND tc.content_status <> 0
# 								AND tc.content_satisfaction IS NOT NULL THEN
# 									tc.content_satisfaction ELSE 0
# 								END
# 									) / SUM(
# 								CASE
#
# 										WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' )
# 										AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND
# 										AND tc.content_status <> 0
# 										AND tc.content_satisfaction IS NOT NULL THEN
# 											1 ELSE 0
# 										END
# 										),
# 										0
# 									) ELSE 0
# 								END AS avg_satisfaction,
# 								SUM(
# 								CASE
#
# 										WHEN tc.content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' )
# 										AND tc.content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND
# 										AND tc.content_status <> 0
# 										AND tc.content_satisfaction IS NULL THEN
# 											1 ELSE 0
# 										END
# 										) AS count_satisfaction_none
# 									FROM
# 										(
# 										SELECT
# 											id,
# 											department_first_name
# 										FROM
# 											hr_department
# 										WHERE
# 											department_first_name IS NOT NULL
# 											AND (
# 												department_expiry_date IS NULL
# 											OR department_expiry_date >= NOW())
# 											AND department_first_name IN ('光伏研究院','全球财务中心', '人力资源中心','全球战略供应链管理中心' )
# 										) dc
# 										CROSS JOIN training_content_category cat
# 										CROSS JOIN training_content_level lvl
# 										LEFT JOIN (
# 										SELECT
# 											*
# 										FROM
# 											training_content
# 										WHERE
# 											content_end_date >= DATE_FORMAT( '{}', '%Y-%m-01 00:00:00' )
# 											AND content_end_date <= LAST_DAY( '{}' ) + INTERVAL 1 DAY - INTERVAL 1 SECOND
# 											AND content_status <> 0
# 										) tc ON tc.content_part_id = dc.id
# 										AND tc.content_category_id = cat.id
# 										AND tc.content_level_id = lvl.id
# 									GROUP BY
# 										dc.department_first_name,
# 									cat.category_name,
# 	                            lvl.level_name WITH ROLLUP;
#
#         """.format(month,month,month,month,month,month,month,month,month,month)
#         # print(sql_query)
#
#
#         sql_query_second="""
#
#            	            SELECT
#             IFNULL(dc.department_second_name, 'Total') AS base_name,
#             IFNULL(cat.category_name, 'Total') AS category_name,
#             IFNULL(lvl.level_name, 'Total') AS level_name,
#             COALESCE(COUNT(tc.id), 0) AS count,
#             COALESCE(SUM(tc.content_people_number), 0) AS total_people,
#             COALESCE(SUM(tc.content_duration), 0) AS total_duration,
#             COALESCE(SUM(tc.content_satisfaction), 0) AS total_satisfaction,
#             CASE
#                 WHEN SUM(
#                     CASE
#                         WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
#                         AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
#                         AND tc.content_status <> 0
#                         AND tc.content_satisfaction IS NOT NULL
#                         THEN 1
#                         ELSE 0
#                     END
#                 ) > 0 THEN
#                     COALESCE(SUM(
#                         CASE
#                             WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
#                             AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
#                             AND tc.content_status <> 0
#                             AND tc.content_satisfaction IS NOT NULL
#                             THEN tc.content_satisfaction
#                             ELSE 0
#                         END
#                     ) / SUM(
#                         CASE
#                             WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
#                             AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
#                             AND tc.content_status <> 0
#                             AND tc.content_satisfaction IS NOT NULL
#                             THEN 1
#                             ELSE 0
#                         END
#                     ), 0)
#                 ELSE
#                     0
#             END AS avg_satisfaction,
#             SUM(
#                     CASE
#                         WHEN tc.content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
#                         AND tc.content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY - INTERVAL 1 SECOND
#                         AND tc.content_status <> 0
#                         AND tc.content_satisfaction IS NULL
#                         THEN 1
#                         ELSE 0
#                     END
#                 ) AS count_satisfaction_none
#
#             FROM (
#             SELECT id, department_second_name
#             FROM hr_department
#             WHERE department_second_name IN (
# '江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司','江苏润阳光伏科技有限公司',
#                 '江苏海博瑞光伏科技有限公司', '润宝电力','润阳光伏科技（泰国）有限公司','宁夏润阳硅材料科技有限公司','江苏润阳光伏科技有限公司（二期）',
#                 '润阳泰国四期组件', '润阳泰国四期电池','云南润阳世纪光伏科技有限公司'
#             )
#             ) dc
#             CROSS JOIN training_content_category cat
#             CROSS JOIN training_content_level lvl
#             LEFT JOIN (
#             SELECT *
#             FROM training_content
#             WHERE content_end_date >= DATE_FORMAT('{}', '%Y-%m-01 00:00:00')
#             AND content_end_date <= LAST_DAY('{}') + INTERVAL 1 DAY
#             AND content_status <> 0
#             ) tc ON tc.content_part_id = dc.id
#             AND tc.content_category_id = cat.id
#             AND tc.content_level_id = lvl.id
#             GROUP BY
#             dc.department_second_name, cat.category_name, lvl.level_name
#             WITH ROLLUP;
#
#         """.format(month,month,month,month,month,month,month,month,month,month)
#
#
#
#         # 执行 SQL 查询
#         with connection.cursor() as cursor:
#             cursor.execute(sql_query)
#             # print(cursor)
#             result = cursor.fetchall()
#
#
#         with connection.cursor() as cursor:
#             cursor.execute(sql_query_second)
#             # print(cursor)
#             result2 = cursor.fetchall()
#
#         result+=result2
#
#
#         tableData=[]
#         for entry in result:
#             department, category, level, content_number, content_people_number, content_duration, content_satisfaction, content_satisfaction_avg,count_satisfaction_none = entry
#             tableData.append({
#                 "department": department,
#                 "category": category,
#                 "level": level,
#                 "content_number": float(content_number),
#                 "content_people_number": float(content_people_number),
#                 "content_duration": float(content_duration),
#                 "content_satisfaction": float(content_satisfaction),
#                 "content_satisfaction_avg": float(content_satisfaction_avg),
#                 'count_null_Satisfaction_Total':float(count_satisfaction_none)
#             })
#         original_list = tableData
#         transformed_list = []
#         department_category_data = {}
#         for item in original_list:
#             department = item['department']
#             category = item['category']
#             level = item['level']
#             content_number = item['content_number']
#
#             if department not in department_category_data:
#                 department_category_data[department] = {}
#
#             if category not in department_category_data[department]:
#                 department_category_data[department][category] = {
#                     'department': department,
#                     'category': category,
#                     'content_number_中层': 0,
#                     'content_number_基层':0,
#                     'content_number_综合': 0,
#                     'content_number_高层': 0,
#                     'content_people_number': 0.0,
#                     'content_duration': 0.0,
#                     'content_satisfaction': 0.0,
#                     'content_satisfaction_avg': 0.0,
#                     'count_null_Satisfaction_Total':0
#                 }
#
#             department_category_data[department][category]['content_number_' + level] = content_number
#             department_category_data[department][category]['content_people_number'] = item['content_people_number']
#             department_category_data[department][category]['content_duration'] = item['content_duration']
#             department_category_data[department][category]['content_satisfaction'] = item['content_satisfaction']
#             department_category_data[department][category]['content_satisfaction_avg'] = item['content_satisfaction_avg']
#             department_category_data[department][category]['count_null_Satisfaction_Total'] = item['count_null_Satisfaction_Total']
#         for department_data in department_category_data.values():
#             transformed_list.extend(department_data.values())
#
#         for item in transformed_list:
#             item['content_number_middle'] = item.pop('content_number_中层', 0)
#             item['content_number_grass'] = item.pop('content_number_基层', 0)
#             item['content_number_synthesis'] = item.pop('content_number_综合', 0)
#             item['content_number_senior'] = item.pop('content_number_高层',0)
#         transformed_list = [d for d in transformed_list if d['department'] != 'Total']
#         # print('list',transformed_list)
#         # tableList=transformed_list
#         # from django.db.models import Count
#         # from itertools import product
#         #
#         # # 获取所有可能的部门和类别组合
#         # all_departments = HrDepartment.objects.filter(Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_first_name__isnull=False,
#         #     department_status=1).exclude(id=999999).values_list('department_first_name', flat=True).distinct()
#         # all_categories = TrainingContentCategory.objects.exclude(id=999999).values_list('category_name', flat=True)
#         # #
#         # # 查询并统计满意度为空的场次数量
#         # if type(month)==str:
#         #     datetime_obj = datetime.strptime(month, "%Y-%m-%d")
#         # else:
#         #     datetime_obj = datetime.combine(month, datetime.min.time())
#         #
#         # current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
#         # next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
#         # # print(current_month,next_month)
#         # result = TrainingContent.objects.filter(content_satisfaction__isnull=True,content_begin_date__range=(current_month, next_month),
#         # content_end_date__range=(current_month, next_month),content_status=1).values(
#         #     'content_part__department_first_name', 'content_category__category_name'
#         # ).annotate(num_sessions=Count('id'))
#         # print('result',result)
#         #
#         #
#         #
#         # # 生成所有可能的部门和类别组合的字典
#         # department_category_combinations = list(product(all_departments, all_categories))
#         # department_category_dict = {(department, category): 0 for department, category in
#         #                             department_category_combinations}
#         #
#         # # 更新字典中的数量信息
#         # for item in result:
#         #     department_name = item['content_part__department_first_name']
#         #     category = item['content_category__category_name']
#         #     num_sessions = item['num_sessions']
#         #     department_category_dict[(department_name, category)] = num_sessions
#         #
#         # # # 生成包含信息的列表
#         # result_list = []
#         # for department, category in department_category_combinations:
#         #     num_sessions = department_category_dict[(department, category)]
#         #     result_list.append(
#         #         {'department': department, 'category': category, 'count_null_Satisfaction': num_sessions})
#         # print(result_list)
#         # department_category_sums = {}
#         # for entry in result_list:
#         #     department = entry['department']
#         #     category = entry['category']
#         #     # count = entry['count_null_Satisfaction']
#         #
#         #     if department not in department_category_sums:
#         #         department_category_sums[department] = {}
#         #
#         #     # if category not in department_category_sums[department]:
#         #     #     department_category_sums[department][category] = count
#         #     # else:
#         #     #     department_category_sums[department][category] += count
#         # for department, categories in department_category_sums.items():
#         #     total_count = sum(categories.values())
#         #     categories['Total'] = total_count
#         #
#         # result = []
#         # for department, categories in department_category_sums.items():
#         #     for category, count in categories.items():
#         #         result.append({'department': department, 'category': category, 'count_null_Satisfaction_Total': count})
#
#         # print('result3333',result)
#         # merged_dict = {}
#         # for item in result:
#         #     key = (item['department'], item['category'])
#         #     merged_dict[key] = item
#
#         # for item in transformed_list:
#         #     key = (item['department'], item['category'])
#         #     if key in merged_dict:
#         #         merged_dict[key].update(item)
#         #     else:
#         #         merged_dict[key] = item
#         #
#         # tableList = list(merged_dict.values())[:-1]
#
#         middle_sum_by_department = {}
#         grass_sum_by_department = {}
#         synthesis_sum_by_department = {}
#         senior_sum_by_department = {}
#
#         # tableList=transformed_list
#         for d in transformed_list:
#             if d['category'] != 'Total':
#                 if d['department'] not in middle_sum_by_department:
#                     middle_sum_by_department[d['department']] = 0
#                     grass_sum_by_department[d['department']] = 0
#                     synthesis_sum_by_department[d['department']] = 0
#                     senior_sum_by_department[d['department']] = 0
#                 middle_sum_by_department[d['department']] += int(d['content_number_middle'])
#                 grass_sum_by_department[d['department']] += int(d['content_number_grass'])
#                 synthesis_sum_by_department[d['department']] += int(d['content_number_synthesis'])
#                 senior_sum_by_department[d['department']] += int(d['content_number_senior'])
#             else:
#                 d['content_number_middle'] = 0
#                 d['content_number_grass'] = 0
#                 d['content_number_synthesis'] = 0
#                 d['content_number_senior'] = 0
#
#         for d in transformed_list:
#             if d['category'] == 'Total':
#                 d['content_number_middle'] = float(middle_sum_by_department[d['department']])
#                 d['content_number_grass'] = float(grass_sum_by_department[d['department']])
#                 d['content_number_synthesis'] = float(synthesis_sum_by_department[d['department']])
#                 d['content_number_senior'] = float(senior_sum_by_department[d['department']])
#
#         # tableList=transformed_list
#         # for line in tableList:
#         #     if line['category']=='Total':
#         #         line['category']='合计'
#         #
#         #     line['content_satisfaction_avg']=round(line['content_satisfaction_avg'],2)
#         #     line['content_satisfaction'] = round(line['content_satisfaction'], 2)
#         #     line['content_duration']=round(line['content_duration'],2)
#         #     if line['content_number_Total']==0:
#         #
#         #         line['content_satisfaction']=0
#         #     else:
#         #         print()
#         #         line['content_satisfaction'] = round(float(line['content_satisfaction']) / float(line['content_number_Total']), 2)      #总满意度/总场次
#
#         tableList=[]
#         # 创建一个字典用于存储不同部门和类别的总和
#         department_category_totals = {}
#         for entry in transformed_list:
#             department = entry['department']
#             category = entry['category']
#             count_null_Satisfaction_Total = entry['count_null_Satisfaction_Total']
#             content_duration = entry['content_duration']
#             content_satisfaction = entry['content_satisfaction']
#             content_satisfaction_avg = entry['content_satisfaction_avg']
#             content_people_number = entry['content_people_number']
#             content_number_Total = float(entry['content_number_Total'])
#             content_number_middle = float(entry['content_number_middle'])
#             content_number_grass = float(entry['content_number_grass'])
#             content_number_synthesis = float(entry['content_number_synthesis'])
#             content_number_senior = float(entry['content_number_senior'])
#             count_null_Satisfaction_Total=float(entry['count_null_Satisfaction_Total'])
#
#
#             # 初始化部门和类别的总和字典条目
#             if (department, category) not in department_category_totals:
#                 department_category_totals[(department, category)] = {
#                     'content_number_senior': 0,  # 高层
#                     'content_number_middle': 0,  # 中层
#                     'content_number_grass': 0,  # 基础
#                     'content_number_synthesis': 0,  # 综合
#                     'content_number_Total': 0,
#                     'content_people_number': 0,
#                     'content_duration': 0,
#                     'content_satisfaction': 0,
#                     'content_satisfaction_avg': 0,
#                     'count_null_Satisfaction_Total': 0,
#                 }
#
#             # 累加各个指标
#             department_category_totals[(department, category)]['content_number_senior'] += content_number_senior
#             department_category_totals[(department, category)]['content_number_middle'] += content_number_middle
#             department_category_totals[(department, category)]['content_number_grass'] += content_number_grass
#             department_category_totals[(department, category)]['content_number_synthesis'] += content_number_synthesis
#             department_category_totals[(department, category)]['content_number_Total'] += content_number_Total
#             department_category_totals[(department, category)]['content_people_number'] += content_people_number
#             department_category_totals[(department, category)]['content_duration'] += content_duration
#             department_category_totals[(department, category)]['content_satisfaction'] += content_satisfaction
#             department_category_totals[(department, category)]['content_satisfaction_avg'] += content_satisfaction_avg
#             department_category_totals[(department, category)][
#                 'count_null_Satisfaction_Total'] += count_null_Satisfaction_Total
#
#         # 创建润阳集团的初始数据
#         ruyang_data = []
#
#         # 遍历字典中的项，根据不同的部门和类别动态创建数据
#         for (department, category), totals in department_category_totals.items():
#             ruyang_entry = {
#                 'department': '润阳集团',
#                 'category': category,
#                 'count_null_Satisfaction_Total': totals['count_null_Satisfaction_Total'],
#                 'content_duration': totals['content_duration'],
#                 'content_satisfaction': totals['content_satisfaction'],
#                 'content_satisfaction_avg': totals['content_satisfaction_avg'],
#                 'content_number_Total': totals['content_number_Total'],
#                 'content_number_middle': totals['content_number_middle'],
#                 'content_number_grass': totals['content_number_grass'],
#                 'content_number_synthesis': totals['content_number_synthesis'],
#                 'content_number_senior': totals['content_number_senior'],
#                 'content_people_number': totals['content_people_number']
#             }
#             ruyang_data.append(ruyang_entry)
#
#         # 将润阳集团的数据添加到总和字典中
#         for entry in ruyang_data:
#             department = entry['department']
#             category = entry['category']
#             count_null_Satisfaction_Total = entry['count_null_Satisfaction_Total']
#             content_duration = entry['content_duration']
#             content_satisfaction = entry['content_satisfaction']
#             content_satisfaction_avg = entry['content_satisfaction_avg']
#             content_number_Total = entry['content_number_Total']
#             content_number_middle = entry['content_number_middle']
#             content_number_grass = entry['content_number_grass']
#             content_number_synthesis = entry['content_number_synthesis']
#             content_number_senior = entry['content_number_senior']
#             content_people_number = entry['content_people_number']
#
#             if (department, category) not in department_category_totals:
#                 department_category_totals[(department, category)] = {
#                     'content_number_senior': 0,  # 高层
#                     'content_number_middle': 0,  # 中层
#                     'content_number_grass': 0,  # 基础
#                     'content_number_synthesis': 0,  # 综合
#                     'content_number_Total': 0,
#                     'content_people_number': 0,
#                     'content_duration': 0,
#                     'content_satisfaction': 0,
#                     'content_satisfaction_avg': 0,
#                     'count_null_Satisfaction_Total': 0,
#                 }
#
#             department_category_totals[(department, category)]['content_number_senior'] += content_number_senior
#             department_category_totals[(department, category)]['content_number_middle'] += content_number_middle
#             department_category_totals[(department, category)]['content_number_grass'] += content_number_grass
#             department_category_totals[(department, category)]['content_number_synthesis'] += content_number_synthesis
#             department_category_totals[(department, category)]['content_number_Total'] += content_number_Total
#             department_category_totals[(department, category)]['content_people_number'] += content_people_number
#             department_category_totals[(department, category)]['content_duration'] += content_duration
#             department_category_totals[(department, category)]['content_satisfaction'] += content_satisfaction
#             department_category_totals[(department, category)]['content_satisfaction_avg'] += content_satisfaction_avg
#             department_category_totals[(department, category)][
#                 'count_null_Satisfaction_Total'] += count_null_Satisfaction_Total
#
#         # 将结果转化为列表形式
#         zz = [{'department': department, 'category': category, **totals}
#               for (department, category), totals in department_category_totals.items()]
#
#         # 打印 zz
#         for entry in zz:
#             tableList.append(entry)
#
#         for line in tableList:
#             if line['department'] == '润阳集团':
#
#
#                 if line['content_number_Total']-line['count_null_Satisfaction_Total']==0:
#                     line['content_satisfaction_avg']=0
#                 else:
#                     line['content_satisfaction_avg']=(line['content_satisfaction']) / (line['content_number_Total'] - line['count_null_Satisfaction_Total'])
#
#                 if line['content_number_Total'] == 0:
#                     line['content_satisfaction'] = 0
#                 else:
#
#                     line['content_satisfaction'] =line['content_satisfaction']/ line['content_number_Total']# 总满意度/总场次
#             else:
#                 if line['content_number_Total'] == 0:
#                     line['content_satisfaction'] = 0
#                 else:
#
#                     line['content_satisfaction'] =line['content_satisfaction']/ line['content_number_Total']# 总满意度/总场次
#         for line in tableList:
#             if line['category']=='Total':
#                 line['category']='合计'
#
#             line['content_satisfaction_avg']=round(line['content_satisfaction_avg'],3)
#             line['content_satisfaction'] = round(line['content_satisfaction'], 3)
#             line['content_satisfaction'] = round(line['content_satisfaction'], 3)
#             line['content_duration']=round(line['content_duration'],3)
#
#
#     # 自定义排序规则
#         def custom_sort(item):
#             # 如果部门名称为'润阳集团'，将其排在第一个，否则按照原始顺序排列
#             if item['department'] == '润阳集团':
#                 return 0
#             return 1
#
#         tableList = sorted(tableList, key=custom_sort)
#
#         base_ls=list(HrDepartment.objects.filter(~Q(id=999999),
#             Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_first_name__isnull=False,
#             department_status=1,id__in=self.request.user_department_employee).values_list('department_name',flat=True))
#         base_ls+=['润阳集团']
#         tableList = [item for item in tableList if item['department'] in base_ls]  # 筛权限
#
#
#         self.return_data = {
#             'code': 200,
#             'msg': '信息返回成功',
#             'data':{
#                 'columnList': columnList,
#                 'tableList': tableList,
#             }
#         }
#
#
#     def download_month_summary_analysis(self):   #下载本月汇总分析
#         now = arrow.now()
#         t1 = now.format('YYYY-MM-DD')
#         t2 = now.timestamp()
#         dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
#         self.mkdir(dummy_path)
#
#         template_path=os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'template_file','本月汇总分析模板.xlsx')  # 创建文件夹
#         import shutil
#
#         # 指定原始文件路径和目标路径
#         source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
#         destination_path =os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file',t1,str(t2),'线下培训汇总分析.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
#         # 使用shutil库进行复制操作
#         shutil.copy(source_path, destination_path)
#
#         self.month_summary_analysis()
#         tableList=self.return_data['data']['tableList']
#         row_data = []
#         for line in tableList:
#             line_data = []
#             for k, v in line.items():
#
#                 line_data.append(v)
#             row_data.append(line_data)
#         # print(row_data)
#         exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
#         sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
#         for row in row_data:
#             sheet.append(row)  # 在工作表中添加一行
#         exc.save(destination_path)  # 指定路径,保存文件
#         # print(destination_path)
#
#         # 使用字符串替换将\替换为/
#         destination_path = destination_path.replace('\\', '/')
#
#         # print(destination_path)
#
#         self.merge_cells(destination_path)
#         destination_path = 'static/' + destination_path.split('static/')[1]
#         # destination_path='static\\'+destination_path.split('static\\')[1]
#         # print(destination_path)
#         # print(destination_path)
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "下载成功",
#             "downloadUrl": destination_path
#         }
#
#
#
#
#     def offline_training_options(self):
#         #
#         # def replace_id_with_value(data, start_index=1):
#         #     for item in data:
#         #         item["value"] = item.pop("id")
#         #         item["index"] = start_index
#         #         start_index += 1
#         #         if "children" in item:
#         #             start_index = replace_id_with_value(item["children"], start_index)
#         #     return start_index
#         #
#         # replace_id_with_value(hrbase_data)
#
#         # hrbase_boss_obj = HrDepartment.objects.filter(
#         #     Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gte=datetime.now()), department_status=1,
#         #     department_parent_id=0).values('id', 'department_name')
#         # hrbase_data = [
#         #     {"value": hrbase_boss_obj[0]['id'], "label": hrbase_boss_obj[0]['department_name'], "children": hrbase_data,"index": 0,}]
#
#         # content_type_list=list(TrainingContentType.objects.filter(type_status=True).values('id','type_name'))
#         content_category_list = list(TrainingContentCategory.objects.filter(category_status=True).values('id', 'category_name'))
#         content_level_list = list(TrainingContentLevel.objects.filter(level_status=True).values('id', 'level_name'))
#         lecturer_level_list = list(TrainingLecturerLevel.objects.filter(level_status=True).values('id', 'level_name'))
#
#
#
#         base_ls=[value for value in self.request.user_department_employee if value != 1]
#         departments = HrDepartment.objects.filter(
#             ~Q(id=999999),
#             Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),
#             department_status=1,
#             id__in=base_ls
#          ).values('id', 'department_name', 'department_parent_id')
#         hrbase_data = get_trees(departments,'id','department_parent_id')
#
#         def add_indexes(node_list, index=0):
#             for node in node_list:
#                 node['index'] = index
#                 index += 1
#                 children = node.get('children', [])
#                 if children:
#                     index = add_indexes(children, index)
#             return index
#         add_indexes(hrbase_data)
#
#
#         # def flatten_tree(node, level, result):
#         #     if len(result) <= level:
#         #         result.append([])
#         #     result[level].append({"label": node['label'], "value": node['value']})
#         #     if 'children' in node:
#         #         for child in node['children']:
#         #             flatten_tree(child, level + 1, result)
#         # flattened_result = [[] for _ in range(2)]
#         # for node in hrbase_data:
#         #     flatten_tree(node, 0, flattened_result)
#         # print(flattened_result)
#         # for level, nodes in enumerate(flattened_result):
#         #     print(nodes)
#
#         content_types = TrainingContentType.objects.filter(
#             type_status=1
#         ).values('id', 'type_name', 'type_parent_id')
#         content_type_data = get_trees2(content_types, 'id', 'type_parent_id')
#
#         def add_index(tree, start_index=1):
#             for node in tree:
#                 node['index'] = start_index
#                 start_index += 1
#                 if 'children' in node:
#                     start_index = add_index(node['children'], start_index)
#             return start_index
#
#         add_index(content_type_data)
#
#         self.return_data = {
#             'data': {
#                 'hrbase_data':hrbase_data,
#                 'content_type_list':
#                     # {"value": item["id"], "label": item["type_name"]}
#                     # for item in content_type_list
#                     content_type_data
#                 ,
#                 'content_level_list': [
#                     {"value": item["id"], "label": item["level_name"]}
#                     for item in content_level_list
#                 ],
#                 'lecturer_level_list': [
#                     {"value": item["id"], "label": item["level_name"]}
#                     for item in lecturer_level_list
#                 ],
#                 "content_category_list": [
#                     {"value": item["id"], "label": item["category_name"]}
#                     for item in content_category_list
#                 ]
#
#
#             },
#             'code': HTTP_200_OK,
#             'msg': '查询成功'
#         }
#
#     def month_Training_hours_per_person(self):
#         from django.db.models import Sum, F, Q
#         from datetime import datetime
#         from django.utils import timezone
#         month = self.request.GET.get('month', None)
#
#         if month != "":
#             month=month
#         else:
#             month=datetime.now().date()
#         if type(month)==str:
#             datetime_obj = datetime.strptime(month, "%Y-%m-%d")
#         else:
#             datetime_obj = datetime.combine(month, datetime.min.time())
#         #
#         current_month = datetime_obj.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
#         next_month = (current_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(seconds=1)
#
#         # from django.db.models import Sum, F, ExpressionWrapper, FloatField, Q
#         # from datetime import datetime
#         #
#         # # Define an ExpressionWrapper to calculate the product of content_duration and content_people_number
#         # product_expression = ExpressionWrapper(
#         #     F('content_duration') * F('content_people_number'),
#         #     output_field=FloatField()
#         # )
#         #
#         # # Create a queryset that filters for content_manner '现场' and aggregates the sum of the product by department_first_name
#         # result = (
#         #     TrainingContent.objects.filter(
#         #         content_manner='现场',
#         #         content_part__department_expiry_date__isnull=True,
#         #         content_part__department_status=True,
#         #         content_begin_date__range=(current_month,next_month)
#         #     )
#         #     .annotate(product=product_expression)
#         #     .values('content_part__department_first_name')
#         #     .annotate(total=Sum('product'))
#         #     .values('content_part__department_first_name', 'total')
#         # )
#         second_name_list = ['江苏润阳悦达光伏科技有限公司', '江苏润阳世纪光伏科技有限公司','江苏润阳光伏科技有限公司',
#                 '江苏海博瑞光伏科技有限公司', '润宝电力','润阳光伏科技（泰国）有限公司','宁夏润阳硅材料科技有限公司','江苏润阳光伏科技有限公司（二期）',
#                 '润阳泰国四期组件', '润阳泰国四期电池','云南润阳世纪光伏科技有限公司']
#
#         first_name_list = ['光伏研究院',  '全球财务中心', '人力资源中心','全球战略供应链管理中心']
#         data=list(TrainingContent.objects.filter(Q(content_part__department_expiry_date__gt=datetime.now())|
#                                                                        Q(content_part__department_first_name__isnull=False),
#                                                  # Q(content_part__department_first_name__in=first_name_list)|Q(content_part__department_second_name__in=second_name_list),
#                                                  content_manner='现场',
#                 content_part__department_expiry_date__isnull=True,
#                 content_part__department_status=True,
#                 content_status=True,
#                 content_begin_date__range=(current_month,next_month),
#                 content_part__department_first_name__in=first_name_list
#         ).values('content_part__department_first_name','content_duration','content_people_number'))
#         for line in data:
#             duration = float(line['content_duration'])
#             people_number = float(line['content_people_number'])
#             line['total'] = duration * people_number
#         # print(data)
#         department_totals = {}
#         for item in data:
#             department_name = item['content_part__department_first_name']
#             total = item['total']
#             if department_name in department_totals:
#                 department_totals[department_name] += total
#             else:
#                 department_totals[department_name] = total
#         # 将结果存储到一个新的列表中
#         result = [{'department_first_name': department_name, 'total': total} for department_name, total in
#                   department_totals.items()]
#         all_first_dp = list(HrDepartment.objects.filter(Q(department_expiry_date__gt=datetime.now()) |
#                                                         Q(department_first_name__isnull=False),
#                                                         department_expiry_date__isnull=True,
#                                                         department_status=True,department_first_name__in=first_name_list ).values_list('department_first_name',
#
#                                                                                               flat=True).distinct())
#         for dp in all_first_dp:
#             # 检查是否在列表b中的department_first_name对应值
#             if not any(dp == department['department_first_name'] for department in result):
#                 # 如果不在，添加到列表b，并赋值total为0
#                 result.append({'department_first_name': dp, 'total': 0.0})
#
#         data_second=list(TrainingContent.objects.filter(Q(content_part__department_expiry_date__gt=datetime.now())|
#                                                                        Q(content_part__department_first_name__isnull=False),
#                                                  # Q(content_part__department_first_name__in=first_name_list)|Q(content_part__department_second_name__in=second_name_list),
#                                                  content_manner='现场',
#                 content_part__department_expiry_date__isnull=True,
#                 content_part__department_status=True,
#                 content_status=True,
#                 content_begin_date__range=(current_month,next_month),
#                 content_part__department_second_name__in=second_name_list
#         ).values('content_part__department_second_name','content_duration','content_people_number'))
#
#         for line in data_second:
#             duration = float(line['content_duration'])
#             people_number = float(line['content_people_number'])
#             line['total'] = duration * people_number
#         # print(data)
#         department_totals = {}
#         for item in data_second:
#             department_name = item['content_part__department_second_name']
#             total = item['total']
#             if department_name in department_totals:
#                 department_totals[department_name] += total
#             else:
#                 department_totals[department_name] = total
#         # 将结果存储到一个新的列表中
#         result2 = [{'department_second_name': department_name, 'total': total} for department_name, total in
#                   department_totals.items()]
#         all_first_dp=list(HrDepartment.objects.filter(Q(department_expiry_date__gt=datetime.now())|
#                                             Q(department_first_name__isnull=False),
#                 department_expiry_date__isnull=True,
#                 department_status=True,department_second_name__in=second_name_list).values_list('department_second_name',flat=True).distinct())
#
#         for dp in all_first_dp:
#             # 检查是否在列表b中的department_first_name对应值
#             if not any(dp == department['department_second_name'] for department in result2):
#                 # 如果不在，添加到列表b，并赋值total为0
#                 result2.append({'department_second_name': dp, 'total': 0.0})
#
#         update_result2 = [{'department_first_name': item['department_second_name'], 'total': item['total']} for item in result2]
#         result=result+update_result2
#
#
#
#
#
#         for department in result:
#             sessionInfo={
#                 "sessions_base": department['department_first_name'],
#                 'sessions_offline_total':department['total'],
#                 'sessions_record_time':str(current_month)[:10],
#             }
#             TrainingSessions.objects.update_or_create(defaults=sessionInfo,sessions_base=sessionInfo['sessions_base'],sessions_record_time=sessionInfo['sessions_record_time'])
#
#         # TrainingContent.objects.filter()
#         columnList = [
#             {'label': 'sessions_base', 'value': '培训基地', 'width':300},
#             {'label': 'sessions_offline_total', 'value': '线下培训总时数', 'width':260},
#             {'label': 'sessions_cloud_total', 'value': '线上(云学堂)培训总时数', 'width': 260},
#             {'label': 'sessions_persons_register', 'value': '月平均在册人数', 'width': 260},
#             {'label': 'sessions_per_people', 'value': '基地人均培训课时', 'width': ''}
#         ]
#
#
#         kwargs={}
#         # print('month',month,type(month))
#         if month != "":
#             if type(month)==str:
#                 month = datetime.strptime(month, '%Y-%m-%d')
#             modified_date = date(month.year, month.month, 1)
#             kwargs['sessions_record_time']=modified_date
#
#         base_ls=list(HrDepartment.objects.filter(~Q(id=999999),
#             Q(department_expiry_date__isnull=True) | Q(department_expiry_date__gt=datetime.now()),department_first_name__isnull=False,
#             department_status=1,id__in=self.request.user_department_employee).values_list('department_name',flat=True))
#         tableList = list(TrainingSessions.objects.filter(**kwargs).values('id',
#                                                                'sessions_base',
#                                                                'sessions_offline_total',
#                                                                "sessions_cloud_total",
#                                                                'sessions_persons_register',
#                                                                'sessions_per_people',
#                                                                           'sessions_record_time'))
#         for line in tableList:
#             if line['sessions_cloud_total'] is None:
#                 line['sessions_cloud_total']=0
#             if line['sessions_persons_register'] is None:
#                 line['sessions_persons_register'] = 0
#             if line['sessions_per_people'] is None:
#                 line['sessions_per_people'] = 0
#             if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line['sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
#                 try:
#                     line['sessions_per_people']=round((line['sessions_offline_total']+line['sessions_cloud_total'])/line['sessions_persons_register'],2)
#                 except:
#                     line['sessions_per_people']=0
#             # print(line)
#             TrainingSessions.objects.update_or_create(defaults=line,sessions_base=line['sessions_base'],sessions_record_time=line['sessions_record_time'])
#
#
#         tableList = [item for item in tableList if item['sessions_base'] in base_ls]  #筛权限
#
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "信息返回成功",
#             "data": {
#                 'columnList': columnList,
#                 'tableList': tableList,
#             }
#         }
#
#
#
#     def edit_month_Training_hours_per_person(self):
#         info = json.loads(self.request.body)
#         TrainingSessions.objects.filter(pk=info['id']).update(**info)
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "修改成功",
#         }
#     def download_month_Training_hours_per_person(self):         #下载本月人均课时
#         now = arrow.now()
#         t1 = now.format('YYYY-MM-DD')
#         t2 = now.timestamp()
#         dummy_path = os.path.join(BASE_DIR, 'static', 'offlineTrainingFile', 'download_file', t1, str(t2))  # 创建文件夹
#         self.mkdir(dummy_path)
#         file_ls = [
#             "序号", "中心/基地", "线下培训总时数", "线上(云学堂)培训总时数", "月平均在册人数", "基地人均培训课时"
#         ]
#         path = self.createExcelPath('基地人均培训课时.xlsx', str(t2), '线下培训报表', 40, 'A1:F1', *file_ls)
#
#         month = self.request.GET.get('month', None)
#         # print('month', month)
#         kwargs = {
#
#         }
#         if month != "":
#             kwargs['sessions_record_time'] = month
#         else:
#             today = datetime.today()
#             month = datetime(today.year, today.month, 1)
#             kwargs['sessions_record_time']=month
#
#
#         base_ls = list(HrDepartment.objects.filter(~Q(id=999999),
#                                                    Q(department_expiry_date__isnull=True) | Q(
#                                                        department_expiry_date__gt=datetime.now()),
#                                                    department_first_name__isnull=False,
#                                                    department_status=1,
#                                                    id__in=self.request.user_department_employee).values_list(
#             'department_name', flat=True))
#         # print(base_ls)
#         tableList = list(TrainingSessions.objects.filter(**kwargs).values(
#                                                                           'sessions_base',
#                                                                           'sessions_offline_total',
#                                                                           "sessions_cloud_total",
#                                                                           'sessions_persons_register',
#                                                                           'sessions_per_people'))
#         for line in tableList:
#             if line['sessions_cloud_total'] is None:
#                 line['sessions_cloud_total']=0
#             if line['sessions_persons_register'] is None:
#                 line['sessions_persons_register'] = 0
#             if line['sessions_per_people'] is None:
#                 line['sessions_per_people'] = 0
#             if line['sessions_offline_total'] is not None and line['sessions_cloud_total'] is not None and line['sessions_persons_register'] is not None and line['sessions_per_people'] is not None:
#                 try:
#                     line['sessions_per_people']=round((line['sessions_offline_total']+line['sessions_cloud_total'])/line['sessions_persons_register'],2)
#                 except:
#                     line['sessions_per_people']=0
#
#
#         tableList = [item for item in tableList if item['sessions_base'] in base_ls]
#         row_data = []
#         index = 1
#         # print('ttttt',tableList)
#         for line in tableList:
#             # print(line)
#             line_data=[]
#             for k,v in line.items():
#                 # print(k,v)
#                 line_data.append(v)
#         #     line = list(line)
#             line_data.insert(0, index)
#             # print(line)
#             row_data.append(line_data)
#             if len(line_data) == 0:
#                 index = index
#             index += 1
#
#         exc = openpyxl.load_workbook(path)  # 打开整个excel文件
#         sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
#         for row in row_data:
#             sheet.append(row)  # 在工作表中添加一行
#         exc.save(path)  # 指定路径,保存文件
#         # print('path',path)
#         # self.merge_cells(path)
#
#         self.return_data = {
#             "code": status.HTTP_200_OK,
#             "msg": "下载成功",
#             "downloadUrl": path
#         }
#
#     @staticmethod
#     def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
#         import openpyxl
#         from openpyxl.styles import Alignment
#         import time
#         exc = openpyxl.Workbook()
#         sheet = exc.active
#         for column in sheet.iter_cols(min_col=0, max_col=num):
#             for cell in column:
#                 sheet.column_dimensions[cell.column_letter].width = 20
#         sheet.column_dimensions['A'].width = 10
#
#         sheet.title = file_name.split('.xlsx')[0]
#         sheet.merge_cells(str(interval))  # 'A1:D1'
#         sheet['A1'] = name
#         sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
#         sheet.append(args)
#         t = time.strftime('%Y-%m-%d')
#         path = os.path.join('static', 'offlineTrainingFile', 'download_file', t, t2, file_name)
#         path = path.replace(os.sep, '/')
#         exc.save(path)
#         return path
#
#
#     @staticmethod
#     def merge_cells(file_path):
#         from openpyxl import load_workbook
#         workbook = load_workbook(file_path)
#         sheet = workbook.active
#         # 从第3行开始，每四行合并A列
#         for row in range(3, sheet.max_row + 1, 4):
#             start_cell = sheet.cell(row=row, column=1)
#             end_cell = sheet.cell(row=row + 3, column=1)
#             sheet.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
#         workbook.save(file_path)
#
#     @staticmethod
#     def mkdir(path):
#         folder = os.path.exists(path)
#         if not folder:
#             os.makedirs(path)
#         else:
#             pass
#
#     @staticmethod
#     def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
#         now = arrow.now()
#         t = now.format('YYYY-MM-DD')
#         file_suffix = str(pic).split(".")[-1]  #文件后缀
#
#         file_name = f"{fileName}.{file_suffix}"    #文件名称
#
#         file_path = os.path.join('static', 'offlineTrainingFile', 'upload_file', t,path,file_name)  # 文件路径
#         file_path = file_path.replace('\\', '/')
#         return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀
#
#     @staticmethod
#     def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
#         with open(str(file_path),'wb+') as f:
#             for dot in file_obj.chunks():
#                 f.write(dot)