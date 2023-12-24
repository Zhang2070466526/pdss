from django.http import HttpResponse
from django.shortcuts import render
from employee.employeeClass.dataSync import DataSync
from django.http import JsonResponse
from rest_framework import status

from auther.models import AdminUser
from employee.models import HrEmployee, HrEmployeeFiles
from django.db import models

from offlineTraining.trainClass.analyseClass import Analyse   #分析  培训类型
from offlineTraining.trainClass.lecturerClass import *   #讲师
from offlineTraining.trainClass.contentClass import * #培训内容
from offlineTraining.trainClass.checkinClass import * #培训签到

from rest_framework.views import APIView
# Create your views here.


def get_lecturer_info(request):
    lecturer = Lecturer(request,'get_lecturer_info')
    res = lecturer.method_center()
    return res
def edit_lecturer_info(request):
    lecturer = Lecturer(request,'edit_lecturer_info')
    res = lecturer.method_center()
    return res
def delete_lecturer_info(request):
    lecturer = Lecturer(request,'delete_lecturer_info')
    res = lecturer.method_center()
    return res
def post_lecturer_info(request):
    lecturer = Lecturer(request,'post_lecturer_info')
    res = lecturer.method_center()
    return res
def batch_lecturer_info(request):
    lecturer = Lecturer(request,'batch_lecturer_info')
    res = lecturer.method_center()
    return res
def down_lecturer_info(request):
    lecturer = Lecturer(request,'down_lecturer_info')
    res = lecturer.method_center()
    return res

def get_retired_lecturer_info(request):
    lecturer = Lecturer(request,'get_retired_lecturer_info')
    res = lecturer.method_center()
    return res
def edit_retired_lecturer_info(request):
    lecturer = Lecturer(request,'edit_retired_lecturer_info')
    res = lecturer.method_center()
    return res





def get_employee_info(request):
    lecturer = Lecturer(request,'get_employee_info')
    res = lecturer.method_center()
    return res

def summary_lecturer_info(request):
    lecturer = Lecturer(request,'summary_lecturer_info')
    res = lecturer.method_center()
    return res
def down_summary_lecturer_info(request):
    lecturer = Lecturer(request,'down_summary_lecturer_info')
    res = lecturer.method_center()
    return res


def get_content_info(request):
    content = Content(request,'get_content_info')
    res = content.method_center()
    return res
def post_content_info(request):
    content = Content(request,'post_content_info')
    res = content.method_center()
    return res
def delete_content_info(request):
    content = Content(request,'delete_content_info')
    res = content.method_center()
    return res
def edit_content_info(request):
    content = Content(request,'edit_content_info')
    res = content.method_center()
    return res
def download_content_info(request):
    content = Content(request,'download_content_info')
    res = content.method_center()
    return res
def batch_content_info(request):
    content = Content(request,'batch_content_info')
    res = content.method_center()
    return res

def del_content_file(request):
    content = Content(request,'del_content_file')
    res = content.method_center()
    return res
def post_content_file(request):
    content = Content(request,'post_content_file')
    res = content.method_center()
    return res


def month_summary_analysis(request): #每月线下培训汇总分析
    # content = Content(request,'month_summary_analysis')
    # res = content.method_center()
    # return res
    analyse = Analyse(request, 'month_summary_analysis')
    res = analyse.method_center()
    return res

def download_month_summary_analysis(request): #每月线下培训汇总分析 下载
    # content = Content(request,'download_month_summary_analysis')
    # res = content.method_center()
    # return res
    analyse = Analyse(request, 'download_month_summary_analysis')
    res = analyse.method_center()
    return res

def offline_training_options(request):  #下拉接口
    content = Content(request,'offline_training_options')
    res = content.method_center()
    return res

def month_Training_hours_per_person(request):#基地人均培训课时
    # content = Content(request, 'month_Training_hours_per_person')
    # res = content.method_center()
    # return res

    analyse = Analyse(request, 'month_Training_hours_per_person')
    res = analyse.method_center()
    return res


def edit_month_Training_hours_per_person(request):#修改每月人均课时
    # content = Content(request, 'edit_month_Training_hours_per_person')
    # res = content.method_center()
    # return res
    analyse = Analyse(request, 'edit_month_Training_hours_per_person')
    res = analyse.method_center()
    return res
def download_month_Training_hours_per_person(request):#修改每月人均课时
    # content = Content(request, 'download_month_Training_hours_per_person')
    # res = content.method_center()
    # return res
    analyse = Analyse(request, 'download_month_Training_hours_per_person')
    res = analyse.method_center()
    return res



def test_data(request):
    import openpyxl

    # workbook = openpyxl.load_workbook('D:\Runergy_SourceCode\Python\pdss\static\9月培训报表.xlsx',data_only=True)
    workbook = openpyxl.load_workbook('D:\Runergy_SourceCode\Python\pdss\static\讲师库所有数据.xlsx')
    # 选择第一个工作表
    worksheet =  workbook['Sheet1']
    ls=[]
    # 遍历行并输出每行的内容


    #-----------------------    讲师库 --------------------------
    column='C'
    start_row = 3
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        # print(row)
        dicts={}
        try:
            dicts['lecturer_level_id']=TrainingLecturerLevel.objects.filter(level_name=row[4]).values_list('id',flat=True)[0]  #讲师级别
            dicts['lecturer_people_id']=HrEmployee.objects.filter(employee_code=row[2]).values_list('id', flat=True)[0]  #工号
            # dicts['lecturer_people_code'] = list(HrEmployee.objects.filter(employee_name=row[1],employee_department__department_first_name=row[6],employee_position__position_name=row[9]).values_list('employee_code', flat=True))  #工号
            # print(dicts['lecturer_people_code'])
            dicts['lecturer_type']='内部讲师'
            dicts['lecturer_remark']=row[5]
            lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=dicts,
                                                                     lecturer_type=dicts['lecturer_type'],
                                                                     lecturer_people_id=dicts['lecturer_people_id'],
                                                                     lecturer_status=True)


        except:

            print('11111111111111111111111111111111111',row)
        # ls.append(HrEmployee.objects.filter(employee_name=row[0], employee_job_duty__job_duty_name=row[1],
        #                                 employee_department__department_name=row[3]).values_list('id', flat=True))
        # print(dicts)
        # TrainingLecturer.objects.update_or_create(defaults=dicts,lecturer_people_id=dicts['lecturer_people_id'])
        # lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=dicts,
        #                                                      lecturer_type=dicts['lecturer_type'],
        #                                                      lecturer_people_id=dicts['lecturer_people_id'],
        #                                                      lecturer_status=True)
    # workbook.save('D:\Runergy_SourceCode\Python\pdss\static\讲师库所有数据.xlsx')
    # 关闭Excel文件

    workbook.close()
    # # -----------------------    讲师库 --------------------------




    #
    #
    # worksheet =  workbook['Sheet1']
    # # ls=[]
    # # 遍历行并输出每行的内容
    # # min_row = 3
    # column = 'F'
    # start_row = 2
    # for row in worksheet.iter_rows(min_row=2,values_only=True):#min_row=3, max_row=4, 3727
    #     try:
    #         dict = {}
    #         lecturer = {}
    #         dict['content_part_id']=HrDepartment.objects.filter(department_first_name=row[1],department_second_name=row[2],department_third_name=row[3],department_forth_name=row[4]).values_list('id',flat=True)[0]
    #         dict['content_title'] = row[5]
    #         # dict['content_title']=str(row[5]).encode('utf-8').decode('utf-8')   #.encode('utf-8') if row[5] is not None else None
    #         dict['content_type_id']=TrainingContentType.objects.filter(type_name=row[6]).values_list('id',flat=True)[0]
    #         dict['content_category_id']=TrainingContentCategory.objects.filter(category_name=row[7]).values_list('id',flat=True)[0]
    #         dict['content_level_id']=TrainingContentLevel.objects.filter(level_name=row[8]).values_list('id',flat=True)[0]
    #         dict['content_manner']=row[9]
    #         dict['content_begin_date']=row[10] #开始培训日期
    #         dict['content_end_date'] = row[11]
    #         dict['content_duration']=row[12]    #培训时长
    #         dict['content_object']=row[13]    #培训对象
    #         dict['content_people_number']=row[14]   #参训人数
    #         dict['content_satisfaction']=row[20] if row[20]!='' else None#培训满意度
    #         dict['content_expenses']=row[21]
    #         dict['content_plan']=row[22]   #计划内计划外
    #         try:
    #             dict['content_creater_id']= AdminUser.objects.filter(username=row[23]).values_list('id',flat=True)[0]      #创建人
    #         except:
    #             dict['content_creater_id']=None
    #
    #
    #         lecturer['lecturer_type']=row[15]#讲师类型
    #         try:
    #             lecturer['lecturer_level_id'] = TrainingLecturerLevel.objects.filter(level_name=row[19]).values_list('id', flat=True)[0] if row[19]!=''  else None
    #         except:
    #             pass
    #         if row[15]=='内部讲师':
    #
    #             lecturer['lecturer_people_id']=HrEmployee.objects.filter(employee_code=row[17]).values_list('id', flat=True)[0] if row[17]!=''  else None
    #             lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer,
    #                                                                  lecturer_type=lecturer['lecturer_type'],
    #                                                                  lecturer_people_id=lecturer['lecturer_people_id'],
    #                                                                  lecturer_status=True)
    #             dict["content_lecturer_id"]=lecturer_obj[0].id
    #         elif row[15]=='外部讲师':
    #             lecturer['lecturer_name']=row[16]
    #             lecturer_obj = TrainingLecturer.objects.update_or_create(defaults=lecturer,
    #                                                                      lecturer_type=lecturer['lecturer_type'],
    #                                                                      lecturer_name=lecturer['lecturer_name'],
    #                                                                      lecturer_status=True)
    #
    #             dict["content_lecturer_id"]=lecturer_obj[0].id
    #         # cell = "{}{}".format(column, start_row)
    #         #
    #         # dict['content_title'] = worksheet[cell].value
    #
    #         # print(dict)  #1586   4960
    #         TrainingContent.objects.update_or_create(defaults=dict,
    #                                                  content_part_id=dict['content_part_id'],
    #                                                  content_object=dict['content_object'],
    #                                                  content_begin_date=dict['content_begin_date'],
    #                                                  content_end_date=dict['content_end_date'],
    #                                                  content_people_number=dict['content_people_number'],
    #                                                  content_status=True
    #                                                  )
    #
    #     except:
    #         print(start_row)
    #     start_row+=1
    #
    #


        # print(row)

    # workbook.close()


    return JsonResponse({'1':1})


class Training_Content_Type_RecordView(APIView):  # 培训类型
    def get(self, request, **kwargs):
        new_query = Analyse(request, 'get_training_content_type')  # 根据工号获取部门等数据
        query = new_query.method_center()
        return query
    def post(self, request, **kwargs):
        new_query = Analyse(request, 'post_training_content_type')
        query = new_query.method_center()
        return query


class Edit_Training_Content_Type_RecordView(APIView):  # 培训类型   修改
    def post(self, request, **kwargs):
        new_query = Analyse(request, 'edit_training_content_type')
        query = new_query.method_center()
        return query

class Del_Training_Content_Type_RecordView(APIView):  # 培训类型  删除
    def post(self, request, **kwargs):
        new_query = Analyse(request, 'del_training_content_type')
        query = new_query.method_center()
        return query


class Training_Checkin_Get_RecordView(APIView): #培训签到
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'get_training_checkin')  # 培训签到数据 查询
        query = new_query.method_center()
        return query


class Training_Checkin_Post_RecordView(APIView): #培训签到
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'post_training_checkin')  # 培训签到数据 查询
        query = new_query.method_center()
        return query



class Training_Checkin_Down_RecordView(APIView): #培训签到 下载
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'down_training_checkin')  # 培训签到数据 下载
        query = new_query.method_center()
        return query
class Training_Checkin_Batch_RecordView(APIView):#培训签到 上传
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'batch_training_checkin')  # 培训签到数据 下载
        query = new_query.method_center()
        return query
class Training_Checkin_Del_RecordView(APIView): #培训签到 删除
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'del_training_checkin')  # 培训签到数据 删除
        query = new_query.method_center()
        return query
class Training_Checkin_Edit_RecordView(APIView):#培训签到 修改
    def post(self, request, **kwargs):
        new_query = Checkin(request, 'edit_training_checkin')  # 培训签到数据 修改
        query = new_query.method_center()
        return query



def get_trees(data,
              key_column='id',
              parent_column='department_parent_id',
              child_column='children',
              current_column=None,
              current_path=None):
    """
    :param data: 数据列表
    :param key_column: 主键字段，默认'id'
    :param parent_column: 父ID字段名，默认'department_parent_id'
    :param child_column: 子列表字典名称，默认'children'
    :param current_column: 当前展开值字段名，若找到展开值增加['open'] = '1'
    :param current_path: 当前展开值
    :return: 树结构
    """
    data_dic = {}
    for d in data:
        d[key_column] = str(d[key_column])  # Ensure the key is a string
        data_dic[d[key_column]] = d  # Use the specified key_column as the dictionary key

    data_tree_list = []  # 整个数据大列表
    for d_id, d_dic in data_dic.items():
        d_dic['label'] = d_dic.pop('type_name')
        d_dic['value'] = eval(d_id)
        pid = d_dic.get(parent_column)
        if pid is None or pid == 0:  # Check for root nodes (e.g., 0)
            data_tree_list.append(d_dic)
        else:
            parent_node = data_dic.get(str(pid))  # Convert pid to a string for dictionary lookup
            if parent_node:
                if child_column not in parent_node:
                    parent_node[child_column] = []
                parent_node[child_column].append(d_dic)

    # # Limit to two levels by removing children of children's children
    for item in data_tree_list:
        if child_column in item:
            for child_item in item[child_column]:
                if child_column in child_item:
                    child_item[child_column] = []

    return data_tree_list






def training_content_type_options(request):#培训类型下拉框   #只有两级
    content_types = TrainingContentType.objects.filter(
        type_status=1
    ).values('id', 'type_name', 'type_parent_id')
    content_type_data = get_trees(content_types, 'id', 'type_parent_id')
    for item in content_type_data:
        # item.pop("type_parent_id", None)
        if "children" in item :
            for chil in item['children']:
                # del chil['type_parent_id']
                if "children" in chil and chil['children']==[]:
                    del chil["children"]
    # for item in content_type_data:
    #     item.pop("type_parent_id", None)

    def add_index(tree, start_index=1):
        for node in tree:
            node['index'] = start_index
            start_index += 1
            if 'children' in node:
                start_index = add_index(node['children'], start_index)
        return start_index

    add_index(content_type_data)


    return_data={'code': status.HTTP_200_OK, "msg": '信息返回成功','data':content_type_data}
    return JsonResponse(return_data)
