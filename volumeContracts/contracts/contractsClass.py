import os, datetime, json, openpyxl, time, random, string, re, docx, zipfile, os,shutil,arrow
import pathlib
from datetime import datetime, date
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from rest_framework.status import *

from pdss.settings import BASE_DIR
from auther.models import *
# from ..serializers import *
from ..models import *
from django.core.exceptions import ObjectDoesNotExist
from controller.controller2 import Controller, upload_file
from general.models import *
from utils.check_token import *
from dateutil.relativedelta import relativedelta
from django.db.models import Q
from zipfile import ZipFile

from ..serializers import ContractsInfoSerializers


class infoMgmt:  # 批量合同
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
            'jobrank_drop':self.jobrank_drop,
            # 'download_data':self.download_data
        }

    def meth_center(self):
        self.methods[self.meth]()
        return Response(self.return_data)



    def get_list(self):
        new_token = CheckToken()
        check_token = new_token.check_token(self.request.headers['Authorization'])
        if check_token != None:
            columnList=[
                {
                    "value": "序号",
                    "label": "index",
                    "width": 60
                },
                {
                    "value": "是否已审核",
                    "label": "approval_status",
                    "width": 150
                },
                {
                    "value": "工号",
                    "label": "code",
                    "width": 200
                },
                {
                    "value": "姓名",
                    "label": "name",
                    "width": 60
                },
                {
                    "value": "性别",
                    "label": "gender",
                    "width": 60
                },
                {
                    "value": "籍贯",
                    "label": "nativePlaceId",
                    "width": 130
                },
                {
                    "value": "身份证号",
                    "label": "idCard",
                    "width": 200
                },
                {
                    "value": "出生日期",
                    "label": "birthday",
                    "width": 200
                },
                {
                    "value": "手机号",
                    "label": "phone",
                    "width": 220
                },
                {
                    "value": "入职日期",
                    "label": "entryData",
                    "width": 200
                },
                {
                    "value": "政治面貌",
                    "label": "politicsStatus",
                    "width": 120
                },
                {
                    "value": "户口性质",
                    "label": "accountNature",
                    "width": 120
                },
                {
                    "value": "户籍地址",
                    "label": "domicileAddress",
                    "width": 400
                },
                {
                    "value": "民族",
                    "label": "nationId",
                    "width": 60
                },
                {
                    "value": "法定送达地址",
                    "label": "legalAddress",
                    "width": 400
                },
                {
                    "value": "婚姻状况",
                    "label": "marriage",
                    "width": 120
                },
                {
                    "value": "紧急联系人姓名",
                    "label": "urgentPerson",
                    "width": 210
                },
                {
                    "value": "紧急联系人电话",
                    "label": "urgentPersonPhone",
                    "width": 220
                },
                {
                    "value": "与紧急联系人关系",
                    "label": "urgentRelation",
                    "width": 240
                },
                {
                    "value": "最高学历",
                    "label": "latestDegreeId",
                    "width": 120
                },
                {
                    "value": "最高学历毕业时间",
                    "label": "graduateTime",
                    "width": 240
                },
                {
                    "value": "毕业学校",
                    "label": "graduateInsituation",
                    "width": 150
                },
                {
                    "value": "毕业专业",
                    "label": "major",
                    "width": 270
                },
                {
                    "value": "教育方式",
                    "label": "educateMethod",
                    "width": 120
                },
                {
                    "value": "卡号",
                    "label": "bankIdCard",
                    "width": 380
                },
                {
                    "value": "开户银行",
                    "label": "openBank",
                    "width": 120
                },
                {
                    "value": "合同到期日",
                    "label": "contractExpirationDate",
                    "width": 200
                },
                {
                    "value": "试用期(月)",
                    "label": "probation",
                    "width": 160
                },
                {
                    "value": "夏装尺寸",
                    "label": "summerSize",
                    "width": 120
                },
                {
                    "value": "身高",
                    "label": "height",
                    "width": 60
                },
                {
                    "value": "体重",
                    "label": "weight",
                    "width": 60
                },
                {
                    "value": "胸围",
                    "label": "bust",
                    "width": 60
                },
                {
                    "value": "肚围",
                    "label": "bellyCircumference",
                    "width": 60
                },
                {
                    "value": "部门",
                    "label": "department",
                    "width": 90
                },
                {
                    "value": "岗位",
                    "label": "contracts_posts",
                    "width": 90
                },
                {
                    "value": "工作地",
                    "label": "contracts_placeWork",
                    "width": 90
                },
                {
                    "value": "基本工资",
                    "label": "base_salary",
                    "width": 120
                },
                {
                    "value": "合同归属",
                    "label": "jobRank",
                    "width": 400
                },
                {
                    "value": "备注",
                    "label": "contracts_remark",
                    "width": ""
                }
            ]
            # for field in ContractsInfo._meta.get_fields():
            #     if field.name not in ['contracts_status','modify_time','create_time','modifier','creator']:
            #         try:
            #             field_label = {
            #                 "value": field.verbose_name,
            #                 "label": field.name,
            #                 "width": self.count_character(field.verbose_name),
            #             }
            #             columnList.append(field_label)
            #         except AttributeError:
            #             pass
            class ContractsInfoClass():
                contracts_gender_choices = {
                    '0': '男',
                    '1': '女'
                }
                contracts_origin_choices ={'1': '北京市', '2': '天津市', '3': '河北省', '4': '山西省', '5': '内蒙古自治区', '6': '辽宁省', '7': '吉林省', '8': '黑龙江省', '9': '上海市', '10': '江苏省', '11': '浙江省', '12': '安徽省', '13': '福建省', '14': '江西省', '15': '山东省', '16': '河南省', '17': '湖北省', '18': '湖南省', '19': '广东省', '20': '广西壮族自治区', '21': '海南省', '22': '重庆市', '23': '四川省', '24': '贵州省', '25': '云南省', '26': '西藏自治区', '27': '陕西省', '28': '甘肃省', '29': '青海省', '30': '宁夏回族自治区', '31': '新疆维吾尔自治区', '32': '台湾省', '33': '香港特别行政区', '34': '澳门特别行政区'}
                contracts_politicalLandscape_choices = {'1': '党员', '2': '预备党员', '3': '群众'}
                contracts_natureOfAccount_choices = {'0': '农业户口', '1': '非农户口'}
                contracts_ethnicGroup_choices = {'1': '汉族', '2': '蒙古族', '3': '回族', '4': '藏族', '5': '维吾尔族', '6': '苗族', '7': '彝族', '8': '壮族', '9': '布依族', '10': '朝鲜族', '11': '满族', '12': '侗族', '13': '瑶族', '14': '白族', '15': '土家族', '16': '哈尼族', '17': '哈萨克族', '18': '傣族', '19': '黎族', '20': '傈僳族', '21': '佤族', '22': '畲族', '23': '高山族', '24': '拉祜族', '25': '水族', '26': '东乡族', '27': '纳西族', '28': '景颇族', '29': '柯尔克孜族', '30': '土族', '31': '达斡尔族', '32': '仫佬族', '33': '羌族', '34': '布朗族', '35': '撒拉族', '36': '毛难族', '37': '仡佬族', '38': '锡伯族', '39': '阿昌族', '40': '普米族', '41': '塔吉克族', '42': '怒族', '43': '乌孜别克族', '44': '俄罗斯族', '45': '鄂温克族', '46': '崩龙族', '47': '保安族', '48': '裕固族', '49': '京族', '50': '塔塔尔族', '51': '独龙族', '52': '鄂伦春族', '53': '赫哲族', '54': '门巴族', '55': '珞巴族', '56': '基诺族', '57': '其他'}
                contracts_highestEducation_choices = {'0': '高中及以下', '1': '中专', '2': '大专', '3': '本科', '4': '硕士', '5': '博士及以上'}
                contracts_educationalMethods_choices ={'0': '全日制', '1': '非全日制'}
                summerSize_choices = {'0': 'XS', '1': 'S', '2': 'M', '3': 'L', '4': 'XL', '5': 'XXL', '6': '3XL', '7': '4XL', '8': '5XL', '9': '其他'}
                contracts_maritalStatus_choices = {'1': '未婚', '0': '已婚'}
                contracts_emergencyContactRelationships_choices = {'0': '父母', '1': '配偶', '2': '兄弟姐妹', '3': '其他'}
                jobRank_choices = {'1': '江苏润阳新能源科技股份有限公司', '2': '润阳新能源（上海）有限公司', '3': '苏州润矽光伏科技有限公司', '4': '盐城润宝电力科技有限公司', '5': '江苏润阳悦达光伏科技有限公司', '6': '江苏润阳光伏科技有限公司', '7': '江苏润阳世纪光伏科技有限公司', '8': '云南润阳世纪光伏科技有限公司', '9': '江苏海博瑞光伏科技有限公司', '10': '宁夏润阳硅材料科技有限公司', '11': '内蒙古润阳悦达新能源科技有限公司'}



            currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
            pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
            kwargs = {
                'contracts_status': True
            }
            searchName = self.request.GET.get('searchName', None)
            baseNameId = self.request.GET.get('baseNameId', None)
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            # print(searchName)
            if len(baseNameId) == 0 or baseNameId is None  and  len(searchName) == 0 or searchName is None and len(beginDate) == 0 or beginDate is None and len(endDate) == 0 or endDate is None:
                kwargs['jobRank__in'] = self.request.user_jobRank

            if len(baseNameId) != 0  :  #有值
                kwargs['jobRank'] = baseNameId


            kwargs['entryData__gte'] = datetime(1901, 10, 29, 7, 17, 1,177) if beginDate is None or len(endDate)==0 else beginDate
            kwargs['entryData__lte'] = datetime(2521, 10, 29, 7, 17, 1,177) if endDate is None or len(endDate)==0 else endDate
            totalNumber=ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName), **kwargs).count()

            from django.db.models import Case, CharField, Value, When
            #    .annotate(
            # # gender_display=Case(When(gender=0, then=Value('男')),When(gender=1, then=Value('女')),output_field=CharField()),
            #                                                                                                         )
            # print(kwargs)
            tableList = ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs)\
                                                                                                                        .values('id',
                                                                                                                            'approval_status',
                                                                                                                            'code',
                                                                                                                            'name',
                                                                                                                            'gender',
                                                                                                                            # 'gender_display',
                                                                                                                            'nativePlaceId',
                                                                                                                            'idCard',
                                                                                                                            'birthday',
                                                                                                                            'phone',
                                                                                                                            'entryData',
                                                                                                                            'politicsStatus',
                                                                                                                            'accountNature',
                                                                                                                            'domicileAddress',
                                                                                                                            'nationId',
                                                                                                                            'legalAddress',
                                                                                                                            'marriage',
                                                                                                                            'urgentPerson',
                                                                                                                            'urgentPersonPhone',
                                                                                                                            'urgentRelation',
                                                                                                                            'latestDegreeId',
                                                                                                                            'graduateTime',
                                                                                                                            'graduateInsituation',
                                                                                                                            'major',
                                                                                                                            'educateMethod',
                                                                                                                            'bankIdCard',
                                                                                                                            'openBank',
                                                                                                                            'contractExpirationDate',
                                                                                                                            'probation',
                                                                                                                            'summerSize',
                                                                                                                            'height',
                                                                                                                            'weight',
                                                                                                                            'bust',
                                                                                                                            'bellyCircumference',
                                                                                                                            'department',
                                                                                                                            'contracts_posts',
                                                                                                                            'contracts_placeWork',
                                                                                                                            'base_salary',
                                                                                                                            'jobRank',
                                                                                                                            # 'contracts_infoFile',
                                                                                                                            'pic_file',
                                                                                                                            'contracts_remark'
                                                                                                                            ).order_by('-create_time')[(currentPage - 1) * pageSize:currentPage * pageSize]

            file_list= ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),contracts_infoFile__isnull=False,**kwargs).values('id','contracts_infoFile').order_by('-create_time')[(currentPage - 1) * pageSize:currentPage * pageSize*3]
            files={}
            for item in file_list:
                id_value = item['id']
                info_file = item['contracts_infoFile']
                if id_value in files:
                    files[id_value].append(info_file)
                else:
                    files[id_value] = [info_file]
            files = [{'id': key, 'contracts_infoFile': value} for key, value in files.items()]
            # print((currentPage - 1) * pageSize,currentPage * pageSize)
            # print(tableList.query)
            # tableList= []
            # for element in objList:
            #     id_value = element['id']
            #     contracts_infoFile_value = element['contracts_infoFile']
            #     matching_element = next((x for x in tableList if x['id'] == id_value), None)
            #
            #     if matching_element:
            #         if contracts_infoFile_value is not None:
            #             matching_element['contracts_infoFile'].append(contracts_infoFile_value)
            #     else:
            #         if contracts_infoFile_value is not None:
            #             element['contracts_infoFile'] = [contracts_infoFile_value]
            #         tableList.append(element)
            # print(tableList)
            # # print(len(tableList))

            # flag=(currentPage - 1) * pageSize
            # flag=(currentPage - 1) * pageSize+1
            for index, item in enumerate(tableList):
                item['index']=(currentPage - 1) * pageSize+index+1
                item['gender_id'] = item['gender']
                item['nativePlaceId_id'] = item['nativePlaceId']
                item['politicsStatus_id'] = item['politicsStatus']
                item['accountNature_id'] = item['accountNature']
                item['nationId_id'] = item['nationId']
                item['marriage_id'] = item['marriage']
                item['urgentRelation_id'] = item['urgentRelation']
                item['latestDegreeId_id'] = item['latestDegreeId']
                item['educateMethod_id'] = item['educateMethod']
                item['summerSize_id'] = item['summerSize']
                item['jobRank_id'] = item['jobRank']
                item['gender'] = ContractsInfoClass().contracts_gender_choices[str(item['gender'])]
                item['nativePlaceId'] = ContractsInfoClass().contracts_origin_choices[str(item['nativePlaceId'])]
                item['politicsStatus'] = ContractsInfoClass().contracts_politicalLandscape_choices[str(item['politicsStatus'])]
                item['accountNature'] = ContractsInfoClass().contracts_natureOfAccount_choices[str(item['accountNature'])]
                item['nationId'] = ContractsInfoClass().contracts_ethnicGroup_choices[str(item['nationId'])]
                item['marriage'] = ContractsInfoClass().contracts_maritalStatus_choices[str(item['marriage'])]
                item['urgentRelation'] = ContractsInfoClass().contracts_emergencyContactRelationships_choices[str(item['urgentRelation'])]
                item['latestDegreeId'] = ContractsInfoClass().contracts_highestEducation_choices[str(item['latestDegreeId'])]
                item['educateMethod'] = ContractsInfoClass().contracts_educationalMethods_choices[str(item['educateMethod'])]
                item['summerSize'] = ContractsInfoClass().summerSize_choices[str(item['summerSize'])]
                item['jobRank'] = ContractsInfoClass().jobRank_choices[str(item['jobRank'])]
                # print(item['id'])
                item['contracts_infoFile'] = None
                for file in files:
                    if file.get('id')==item['id']:
                        item['contracts_infoFile']=file['contracts_infoFile']
                # item['contracts_infoFile']=ContractsInfo.objects.filter(id=item['id'],contracts_status=True).values_list('contracts_infoFile__id',flat=True)

            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': columnList,
                    'tableList': tableList,
                    'totalNumber': totalNumber,
                }
            }
            # obj = Controller(ContractsInfo, "get_list", self.request)
            # self.return_data = obj.data_start()

        else:
            self.return_data = {
                "code": HTTP_403_FORBIDDEN,
                "msg": "没有权限访问",
                'hidden': False
            }

    def jobrank_drop(self):
        self.return_data = {
            "code": 200,
            "msg": "下拉菜单返回成功",
            "data": [],
            'hidden': True
        }
        new_token = CheckToken()
        check_token = new_token.check_token(self.request.headers['Authorization'])
        # print(check_token)
        if check_token!=None:
            # jobrankAll = AdminUser.objects.filter(user_jobrank__jobrank_status=True).values_list('user_jobrank__id',
            #                                                                                   'user_jobrank__jobrank_name').all()
            # jobrankAll = AdminUser.objects.filter(user_jobrank__jobrank_status=True).values_list('user_jobrank__id',
            #                                                                                   'user_jobrank__jobrank_name').all()
            jobrankAll =ContractsJobrank.objects.filter(jobrank_status=True).values_list('id','jobrank_name').all()
            '''
            id=check_token,
            '''
            for i in jobrankAll:
                self.return_data['data'].append({
                    "label": i[1],
                    "id": i[0]
                })

        else:
            self.return_data = {
                "code": HTTP_403_FORBIDDEN,
                "msg": "没有权限访问",
                'hidden': False
            }

    def get_upload(self):
        try:
            # info=eval(self.request.POST.get('creatData',None))

            info=self.request.POST.dict()
            # print(info)
            # # info = self.request.data
            # info=eval(self.request.POST.get('creatData',None))
            # print(self.request.FILES)
            pic_file=self.request.FILES.get("oneInchPic", None)
            # print(pic_file,type(pic_file))
            t = time.strftime('%Y-%m-%d')
            dummy_path = os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'upload_file', t,str(info['name'])+str(info['code']))  # 创建文件夹
            self.mkdir(dummy_path)
            file_url, file_name, file_suffix = self.createPath(pic_file,str(info['code']),str(info['name'])+str(info['code']))
            # print(file_url)
            if file_name:
                self.saveFile(file_name, pic_file, file_suffix,str(info['name'])+str(info['code']))  # 保存文件



            info['gender'] = int(info['gender'])
            info['accountNature'] = int(info['accountNature'])
            info['politicsStatus'] = int(info['politicsStatus'])
            info['nationId'] = int(info['nationId'])
            info['nativePlaceId'] = int(info['nativePlaceId'])
            info['marriage'] = int(info['marriage'])
            info['urgentRelation'] = int(info['urgentRelation'])
            info['latestDegreeId'] = int(info['latestDegreeId'])
            info['educateMethod'] = int(info['educateMethod'])
            info['summerSize'] = int(info['summerSize'])
            # print(type(info), info)  #{'entryData': '2023-06-28', 'idCard': '44444444444444', 'phone': '44444444', 'name': '444444', 'code': '34444444', 'department': '4444444', 'gender': '0', 'birthday': '2023-06-13', 'accountNature': '1', 'politicsStatus': '3', 'nationId': '2', 'domicileAddress': '4444444444', 'nativePlaceId': '2', 'legalAddress': '44444444444', 'marriage': '0', 'urgentPerson': '4444444444', 'urgentPersonPhone': '4444444444', 'urgentRelation': '0', 'latestDegreeId': '1', 'graduateTime': '2023-06-20', 'graduateInsituation': '44444444', 'major': '4444444', 'educateMethod': '1', 'bankIdCard': '44444444', 'openBank': '444444444', 'contractExpirationDate': '2023-06-06', 'probation': '44444444', 'summerSize': '1', 'height': '', 'weight': '', 'bust': '', 'bellyCircumference': '','department':''}
            # print(info['entryData'],type(info['entryData']))
            date = datetime.strptime(info['entryData'], "%Y-%m-%d")
            info['contractExpirationDate'] = str(date + relativedelta(years=3) - relativedelta(days=1))[:10]
            info['probation'] = str(3)
            info['pic_file']=file_url

            try:
                del info['oneInchPic']
            except:
                pass


            # print(info)
            obj = ContractsInfo.objects.update_or_create(defaults=info, idCard=info['idCard'])
            id = ContractsInfo.objects.filter(idCard=info['idCard'], contracts_status=True).values_list('id', flat=True)
            obj = ContractsInfo.objects.filter(id=id[0], contracts_status=True, idCard=info['idCard']).first()
            jobRank = obj.get_jobRank_display()
            # print(ContractsInfo.objects.filter(idCard=info['idCard'], contracts_status=True, id=id[0]))
            ContractsInfo.objects.filter(idCard=info['idCard'], contracts_status=True, id=id[0]).update(
                contracts_placeWork=self.findExcel(jobRank)[3])
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "填写成功！"
            }




        except ValueError as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "数据不符,填写失败！"
            }


    def patch_data(self):
        isPatch = self.request.GET.get('isPatch', 'None')
        info = self.request.data
        print(info)
        if self.request.check_token!=None:
            if isPatch=='true':  # 修改
                info['gender'] = int(info['gender_id'])
                info['accountNature'] = int(info['accountNature_id'])
                info['politicsStatus'] = int(info['politicsStatus_id'])
                info['nationId'] = int(info['nationId_id'])
                info['nativePlaceId'] = int(info['nativePlaceId_id'])
                info['marriage'] = int(info['marriage_id'])
                info['urgentRelation'] = int(info['urgentRelation_id'])
                info['latestDegreeId'] = int(info['latestDegreeId_id'])
                info['educateMethod'] = int(info['educateMethod_id'])
                info['summerSize'] = int(info['summerSize_id'])
                try:
                    info['base_salary']=int(info['base_salary'])
                except:
                    info['base_salary'] =0
                date = datetime.strptime(info['entryData'], "%Y-%m-%d")
                info['contractExpirationDate'] = str(date + relativedelta(years=3) - relativedelta(days=1))[:10]
                info['probation'] = str(3)
                id = info['id']
                ContractsInfo.objects.filter(id=id, contracts_status=True).update(jobRank=info['jobRank_id'])
                obj = ContractsInfo.objects.filter(id=id, contracts_status=True).first()
                # print(obj)
                jobRank = obj.get_jobRank_display()
                info['contracts_placeWork'] = self.findExcel(jobRank)[3]
                info2={
                    "code": info['code'],
                    "name": info['name'],
                    "phone": info['phone'],
                    "entryData": info['entryData'],
                    "politicsStatus":info['politicsStatus_id'],
                    "nationId":info['nationId_id'],
                    "marriage": info['marriage_id'],
                    "urgentPersonPhone": info['urgentPersonPhone'],
                    "urgentRelation": info['urgentRelation_id'],
                    "birthday": info['birthday'],
                    "idCard": info['idCard'],
                    "latestDegreeId":info['latestDegreeId_id'],
                    "graduateTime": info['graduateTime'],
                    "nativePlaceId": info['nativePlaceId_id'],
                    "gender": info['gender_id'],
                    "graduateInsituation":info['graduateInsituation'],
                    "domicileAddress": info['domicileAddress'],
                    "accountNature": info['accountNature_id'],
                    "major": info['major'],
                    "legalAddress": info['legalAddress'],
                    "educateMethod":info['educateMethod_id'],
                    "urgentPerson": info['urgentPerson'],
                    "bankIdCard":info['bankIdCard'],
                    "openBank": info['openBank'],
                    "contractExpirationDate":info['contractExpirationDate'],
                    # "contracts_posts": info['contracts_posts'],
                    "contracts_posts": '追光者',
                    "probation": info['probation'],
                    "contracts_placeWork": info['contracts_placeWork'],   #工作地
                    "jobRank": info['jobRank_id'],
                    "department": info['department'],
                    "summerSize":info['summerSize'],
                    "height": info['height'],
                    "weight": info['weight'],
                    "bust": info['bust'],
                    "bellyCircumference": info['bellyCircumference'],
                    "contracts_remark": info['contracts_remark'],
                    'base_salary':info['base_salary'],
                    'approval_status':False,
                    'modifier_id':self.request.check_token
                }
                if info2['summerSize']!=9:
                    info2["height"]= ''
                    info2["weight"]= ''
                    info2["bust"]=''
                    info2["bellyCircumference"]=''
                ContractsInfo.objects.filter(id=id, contracts_status=True).update(**info2)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": '修改成功！'
                }
            else:  # 审批
                # print('审核')
                for id in info['idList']:
                    ContractsInfo.objects.filter(id=id, contracts_status=True).update(approval_status=True,modifier_id=self.request.check_token)
                    self.write_docx(id)  #压缩目标文件夹,姓名
                    self.return_data = {
                        "code": status.HTTP_200_OK,
                        "msg": '用户审批通过！',
                    }
        else:
            self.return_data = {
                "code": HTTP_403_FORBIDDEN,
                "msg": "没有权限访问",
                'hidden': False
            }

    def count_character(self, s):
        hanzi = 0
        num = 0
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1

        return str(30 * hanzi + 20 * num)
    def write_docx(self, id):
        t = time.strftime('%Y-%m-%d')
        info = ContractsInfo.objects.filter(id=id, contracts_status=True, approval_status=True).values()[0]
        obj = ContractsInfo.objects.filter(id=id, contracts_status=True, approval_status=True).first()
        self.mkdir(os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'report_template_file'))  # 创建模板文件夹
        self.mkdir(os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'info_file', t, info['name']+info['code']))  # 用户个人文档文件夹
        path1 = 'static/volumeContractsFile/report_template_file/基本信息表、法律送达文书、重要事项、反腐败等.docx'
        doc1 = docx.Document(path1)  # 创建文档对象,获得word文档
        tables1 = doc1.tables  # 获取所有表格对象

        path2 = 'static/volumeContractsFile/report_template_file/入职资料-润阳新能源（合同、保密、竞业）.docx'
        doc2 = docx.Document(path2)  # 创建文档对象,获得word文档

        checkIn1 = tables1[0]  # 等级表  # 获取word中第一个表格对象
        checkIn1_name = checkIn1.cell(0, 2)  # 获取第1行第3列单元格的对象
        checkIn1_gender = checkIn1.cell(0, 4)
        checkIn1_origin = checkIn1.cell(1, 2)  # 籍贯
        checkIn1_idNumber = checkIn1.cell(3, 5)  # 身份证号码   获取第4行第6列单元格的对象
        checkIn1_birthday = checkIn1.cell(2, 2)  # 出生日期
        checkIn1_phoneNumber = checkIn1.cell(9, 2)  # 手机号
        # checkIn1_dateOfEntry = checkIn.cell(2, 2)  # 入职日期     --填写日期
        checkIn1_politicalLandscape = checkIn1.cell(2, 5)  # 政治面貌
        checkIn1_natureOfAccount = checkIn1.cell(3, 2)  # 户口性质
        checkIn1_registeredAddress = checkIn1.cell(4, 2)  # 户籍地址    获取第5行第3列单元格的对象
        checkIn1_ethnicGroup = checkIn1.cell(0, 6)  # 民族      获取第1行第7列单元格的对象
        checkIn1_deliveryAddress = checkIn1.cell(5, 2)  # 送达地址    -->  现居住住址  获取第6行第3列单元格的对象
        checkIn1_maritalStatus = checkIn1.cell(1, 5)  # 婚姻状况      获取第2行第5列单元格的对象
        checkIn1_emergencyContactName = checkIn1.cell(6, 2)  # 紧急联系人姓名  获取第7行第3列单元格的对象
        checkIn1_emergencyContactNumber = checkIn1.cell(6, 5)  # 紧急联系人电话
        checkIn1_emergencyContactRelationships = checkIn1.cell(6, 9)  # 紧急联系人关系
        checkIn1_highestEducation = checkIn1.cell(8, 5)  # 最高学历    获取第9行第6列单元格的对象
        checkIn1_graduationTime = checkIn1.cell(8, 2)  # 毕业时间
        checkIn1_graduationSchool = checkIn1.cell(7, 2)  # 毕业学校
        checkIn1_graduationMajor = checkIn1.cell(7, 5)  # 所学专业
        # checkIn_educationalMethods = checkIn.cell(2, 2)  # 教育方式
        checkIn1_cardNumber = checkIn1.cell(10, 4)  # 卡号   获取第11行第3列单元格的对象
        checkIn1_bankName = checkIn1.cell(10, 6)  # 开户银行   获取第11行第5列单元格的对象

        checkIn1_zip = checkIn1.cell(5, 4)  # 邮编 获取第6行第5列单元格的对象
        checkIn1_degree = checkIn1.cell(8, 6)  # 学位 获取第9行第7列单元格的对象
        checkIn1_mailbox = checkIn1.cell(9, 4)  # 邮箱 获取第10行第5列单元格的对象
        checkIn1_foreignLanguage = checkIn1.cell(7, 6)  # 外语语种 获取第8行第7列单元格的对象
        # # 家庭成员 部门  职位  微信號

        checkIn1_employeeHandbook = checkIn1.cell(14, 2)  # 员工手册
        checkIn1_department = checkIn1.cell(15, 2)  # 所属部门
        checkIn1_posts = checkIn1.cell(15, 5)  # 职位
        checkIn1_dateOfEntry = checkIn1.cell(15, 10)  # 入职时间

        checkIn2 = tables1[1]  # 确认书
        checkIn2_name = checkIn2.cell(1, 1)  # 姓名2
        checkIn2_dateOfEntry = checkIn2.cell(1, 4)  # 入职日期2
        checkIn2_idNumber = checkIn2.cell(1, 10)  # 身份证号2
        checkIn2_deliveryAddress = checkIn2.cell(3, 3)  # 送达地址2
        checkIn2_signatory = checkIn2.cell(4, 3)  # 签收人2    -->姓名2
        checkIn2_phoneNumber = checkIn2.cell(4, 10)  # 联系电话2
        checkIn2_weChatID = checkIn2.cell(5, 2)  # 微信號2
        checkIn2_wechatBoundMobile = checkIn2.cell(5, 4)  # 微信绑定手机号2
        checkIn2_emailAddress1 = checkIn2.cell(6, 2)  # 邮箱地址1  2
        checkIn2_emailAddress2 = checkIn2.cell(6, 4)  # 邮箱地址2 2
        checkIn2_emergencyContactName = checkIn2.cell(8, 3)  # 紧急联系人姓名1  2
        checkIn2_emergencyContactRelationships = checkIn2.cell(8, 5)  # 紧急联系人关系1  2
        checkIn2_emergencyContactNumber = checkIn2.cell(8, 11)  # 紧急联系人电话1  2
        checkIn2_emergencyContactWeChatID = checkIn2.cell(9, 2)  # 紧急联系人微信号1  2
        checkIn2_emergencyContactWechatBoundMobile = checkIn2.cell(9, 4)  # 紧急联系人微信绑定手机号1  2
        checkIn2_employeeConfirmation = checkIn2.cell(13, 0)  # 员工确认  2

        checkIn3 = tables1[2]  # 手续确认单
        checkIn3_workNumber = checkIn3.cell(0, 0)  # 工号 3
        checkIn3_name = checkIn3.cell(0, 2)  # 姓名     3
        checkIn3_dateOfEntry = checkIn3.cell(0, 4)  # 入职日期    3

        contracts_workNumber = info['code']
        contracts_name = info['name']
        contracts_gender = obj.get_gender_display()
        contracts_origin = obj.get_nativePlaceId_display()  # 籍贯
        contracts_idNumber = info['idCard']  # 身份证号码
        contracts_birthday = info['birthday']  # 出生日期
        contracts_phoneNumber = info['phone']  # 手机号
        contracts_dateOfEntry = info['entryData']  # 入职日期
        contracts_politicalLandscape = obj.get_politicsStatus_display()  # 政治面貌
        contracts_natureOfAccount = obj.get_accountNature_display()  # 户口性质
        contracts_registeredAddress = str(info['domicileAddress'])  # 户籍地址
        contracts_ethnicGroup = obj.get_nationId_display()  # 民族
        contracts_deliveryAddress = info['legalAddress']  # 送达地址
        contracts_maritalStatus = obj.get_marriage_display()  # 婚姻状况
        contracts_emergencyContactName = info['urgentPerson']  # 紧急联  系人姓名
        contracts_emergencyContactNumber = info['urgentPersonPhone']  # 紧急联系人电话
        contracts_emergencyContactRelationships = obj.get_urgentRelation_display()  # 紧急联系人关系
        contracts_highestEducation = obj.get_latestDegreeId_display()  # 最高学历
        contracts_graduationTime = info['graduateTime']  # 毕业时间
        contracts_graduationSchool = info['graduateInsituation']  # 毕业学校
        contracts_graduationMajor = info['major']  # 毕业专业
        contracts_educationalMethods = obj.get_educateMethod_display()  # 教育方式
        contracts_cardNumber = info['bankIdCard']  # 卡号
        contracts_bankName = info['openBank']  # 开户银行
        contracts_contractExpirationDate = info['contractExpirationDate']  # 合同到期日
        contracts_probationPeriod = info['probation']  # 试用期(月)
        contracts_posts = '追光者'  # 岗位
        contracts_placeWork=info['contracts_placeWork']#工作地
        # contracts_placeWork =  obj.get_contracts_placeWork_display()   # 工作地

        department = info['department']  # 部门
        base_salary = str(info['base_salary'])  # 部门
        jobRank = obj.get_jobRank_display()

        # 填充到world表格1中
        checkIn1_name.text = contracts_name
        checkIn1_gender.text = str(contracts_gender)
        checkIn1_origin.text = str(contracts_origin)  # 籍贯
        checkIn1_idNumber.text = str(contracts_idNumber)  # 身份证号码
        checkIn1_birthday.text = str(contracts_birthday)[:10]  # 出生日期
        checkIn1_phoneNumber.text = str(contracts_phoneNumber)  # 手机号
        # checkIn_dateOfEntry = checkIn.cell(2, 2)  # 入职日期     --填写日期
        checkIn1_politicalLandscape.text = str(contracts_politicalLandscape)  # 政治面貌
        checkIn1_natureOfAccount.text = contracts_natureOfAccount  # 户口性质
        checkIn1_registeredAddress.text = contracts_registeredAddress  # 户籍地址
        checkIn1_ethnicGroup.text = contracts_ethnicGroup  # 民族
        checkIn1_deliveryAddress.text = contracts_deliveryAddress  # 送达地址    -->  现居住住址
        checkIn1_maritalStatus.text = contracts_maritalStatus  # 婚姻状况
        checkIn1_emergencyContactName.text = contracts_emergencyContactName  # 紧急联系人姓名
        checkIn1_emergencyContactNumber.text = str(contracts_emergencyContactNumber)  # 紧急联系人电话
        checkIn1_emergencyContactRelationships.text = contracts_emergencyContactRelationships  # 紧急联系人关系
        checkIn1_highestEducation.text = contracts_highestEducation  # 最高学历
        checkIn1_graduationTime.text = str(contracts_graduationTime)[:10]  # 毕业时间
        checkIn1_graduationSchool.text = contracts_graduationSchool  # 毕业学校


        checkIn1_graduationMajor.text = contracts_graduationMajor  # 毕业专业
        # checkIn1_graduationMajor.width = 100
        # font=checkIn1_graduationMajor.paragraphs[0].runs[0].font
        # # font.name = '微软雅黑'
        # # 设置对齐方式
        # # 设置字号大小
        # font.size = docx.shared.Pt(8)


        # checkIn_educationalMethods.text = educationalMethods  # 教育方式
        checkIn1_cardNumber.text = str(contracts_cardNumber)  # 卡号
        checkIn1_bankName.text = contracts_bankName  # 开户银行
        for i in range(len(checkIn1_employeeHandbook.elements)):
            if i == 3:
                for j in range(len(checkIn1_employeeHandbook.elements[i].runs)):
                    if j == 1:
                        checkIn1_employeeHandbook.elements[i].runs[j].text = '已接受'
                        checkIn1_employeeHandbook.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 3:
                        checkIn1_employeeHandbook.elements[i].runs[j].text = '知悉'
                        checkIn1_employeeHandbook.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 5:
                        checkIn1_employeeHandbook.elements[i].runs[j].text = '理解'
                        checkIn1_employeeHandbook.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 7:
                        checkIn1_employeeHandbook.elements[i].runs[j].text = '自愿遵守'
                        checkIn1_employeeHandbook.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 10:
                        checkIn1_employeeHandbook.elements[i].runs[j].text = '愿意'
                        checkIn1_employeeHandbook.elements[i].runs[j].font.underline = True  # 添加下划线
            # try:
            #     print(i,checkIn1_employeeHandbook.elements[i].text)
            # except:
            #     pass
        checkIn1_department.text = department  # 所属部门
        checkIn1_posts.text = '追光者'  # 职位
        checkIn1_dateOfEntry.text = str(contracts_dateOfEntry)[:10]  # 入职时间
        for i in range(len(doc1.paragraphs)):
            if i == 2:
                doc1.paragraphs[i].text = '填写日期：{}年{}月{}日                                     编号：{}'.format(
                    str(contracts_dateOfEntry)[:4], str(contracts_dateOfEntry)[5:7], str(contracts_dateOfEntry)[8:10],
                    str(contracts_workNumber))
            if i == 13:  # 重要事项告知书
                for j in range(len(doc1.paragraphs[i].runs)):
                    if j == 1:
                        doc1.paragraphs[i].runs[j].text = contracts_name
                        doc1.paragraphs[i].runs[j].font.underline = True  # 添加下划线
            if i == 53:  # 部门
                for j in range(len(doc1.paragraphs[i].runs)):
                    # try:
                    #     print(j, doc1.paragraphs[i].runs[j].text)
                    # except:
                    #     pass
                    if j == 1:
                        doc1.paragraphs[i].runs[j].text = department.center(20, ' ')
            if i == 54:  # 工号
                for j in range(len(doc1.paragraphs[i].runs)):
                    # try:
                    #     print(j, doc1.paragraphs[i].runs[j].text)
                    # except:
                    #     pass
                    if j == 4:
                        doc1.paragraphs[i].runs[j].text = contracts_workNumber.center(22, ' ')
        checkIn2_name.text = contracts_name  # 姓名2
        checkIn2_dateOfEntry.text = str(contracts_dateOfEntry)[:10]  # 入职日期2
        checkIn2_idNumber.text = str(contracts_idNumber)  # 身份证号2
        checkIn2_deliveryAddress.text = contracts_deliveryAddress  # 送达地址2
        checkIn2_signatory.text = contracts_name  # 签收人2    -->姓名2
        checkIn2_phoneNumber.text = str(contracts_phoneNumber)  # 联系电话2
        checkIn2_emergencyContactName.text = contracts_emergencyContactName  # 紧急联系人姓名1  2
        checkIn2_emergencyContactRelationships.text = contracts_emergencyContactRelationships  # 紧急联系人关系1  2
        checkIn2_emergencyContactNumber.text = str(contracts_emergencyContactNumber)  # 紧急联系人电话1  2
        for i in range(len(checkIn2_employeeConfirmation.elements)):
            if i == 2:
                for j in range(len(checkIn2_employeeConfirmation.elements[i].runs)):
                    # print(j,checkIn2_employeeConfirmation.elements[i].runs[j].text)
                    if j == 7:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '注意事项'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 10:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '同意'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 12:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '接受'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 17:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '正确'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 19:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '有效'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划线
                    if j == 23:
                        checkIn2_employeeConfirmation.elements[i].runs[j].text = '本人自行承担'
                        checkIn2_employeeConfirmation.elements[i].runs[j].font.underline = True  # 添加下划
        for i in range(len(checkIn3_workNumber.elements)):
            if i == 1:
                checkIn3_workNumber.elements[i].add_run(str(contracts_workNumber))
        for i in range(len(checkIn3_name.elements)):
            if i == 1:
                checkIn3_name.elements[i].add_run(contracts_name)
        for i in range(len(checkIn3_dateOfEntry.elements)):
            if i == 1:
                checkIn3_dateOfEntry.elements[i].add_run(str(contracts_dateOfEntry)[:10])

        # name1 = '基本信息表、法律送达文书、重要事项、反腐败等{}.docx'.format(
        #     '_' + contracts_name + str(contracts_workNumber)[:4] + 'basic' + str(time.time()))
        name1 = '{}_基本信息表、法律送达文书、重要事项、反腐败等.docx'.format(
            contracts_name + str(contracts_workNumber))
        path1 = "static/volumeContractsFile/info_file/{}/{}/{}".format(t, info['name']+info['code'], name1)
        doc1.save(path1)
        file1_kwargs = {
            "url": path1,
            "name": name1,
            'filetype': 1
        }
        try:
            infoObj = ContractsInfo.objects.get(id=id)
            # print(infoObj.contracts_infoFile.all())
            infoId = infoObj.contracts_infoFile.filter(filetype=1, file_status=True).first().id
        except:
            infoId = None
        file_obj = ContractsFile.objects.update_or_create(defaults=file1_kwargs, id=infoId)[0]
        file_obj.contracts_infoFile.add(obj.id)

        for i in range(len(doc2.paragraphs)):

            if i == 15:  # 乙方
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 3:
                        if len(contracts_name)==2:
                            doc2.paragraphs[i].runs[j].text = contracts_name.center(26, ' ')
                        elif len(contracts_name)==3:
                            doc2.paragraphs[i].runs[j].text = contracts_name.center(25, ' ')
                        elif len(contracts_name)==4:
                            doc2.paragraphs[i].runs[j].text = contracts_name.center(24, ' ')
                        elif len(contracts_name) == 5:
                            doc2.paragraphs[i].runs[j].text = contracts_name.center(23, ' ')
                        else:
                            doc2.paragraphs[i].runs[j].text = contracts_name.center(22, ' ')
                        # print(contracts_name.center(24, '_'))
                        # doc2.paragraphs[i].runs[j].text = contracts_name.center(24, ' ')
                        # doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线

            if i == 16:  # 工號
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = str(contracts_workNumber).center(28, ' ')
                        # doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
            if i == 17:  # 簽訂日期
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[:10].center(29, ' ')
                        # doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线

            if i == 50:  # 乙方
                for j in range(len(doc2.paragraphs[i].runs)):
                    doc2.paragraphs[i].runs[j].add_text(contracts_name)
            if i == 51:  # 居民身份证
                for j in range(len(doc2.paragraphs[i].runs)):
                    doc2.paragraphs[i].runs[j].add_text(str(contracts_idNumber))
            try:
                if i == 52:  # 性别
                    # print("53",doc1.paragraphs[53].text)
                    for j in range(len(doc2.paragraphs[53].runs)):
                        doc2.paragraphs[i].runs[j].add_text(contracts_gender)
                if i == 53:  # 户籍地
                    # print(doc1.paragraphs[54].text)
                    for j in range(len(doc2.paragraphs[i].runs)):
                        doc2.paragraphs[i].runs[j].add_text(contracts_registeredAddress)
                if i == 54:  # 法定送达地址
                    for j in range(len(doc2.paragraphs[i].runs)):
                        if j == 1:
                            doc2.paragraphs[i].runs[j].text = '：' + contracts_deliveryAddress
            except:
                pass
            if i==163:#日期
                for j in range(len(doc2.paragraphs[i].runs)):
                    # try:
                    #     print(j,doc2.paragraphs[i].runs[j].text)
                    # except:
                    #     pass
                    if j==8:#年
                        doc2.paragraphs[i].runs[j].text=str(contracts_dateOfEntry)[:4].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j==12:#月
                        doc2.paragraphs[i].runs[j].text=str(contracts_dateOfEntry)[5:7].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j==14:#日
                        doc2.paragraphs[i].runs[j].text=str(contracts_dateOfEntry)[8:10].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j==16:#地点
                        doc2.paragraphs[i].runs[j].text="在中国{}市".format(self.findExcel(jobRank)[3])
                    if j in[9,10]:
                        doc2.paragraphs[i].runs[j].text =''

            if i==174:#日期
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:  # 年
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[:4].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 3:  # 月
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[5:7].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 5:  # 日
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[8:10].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线


            if i == 55:  # 联系方式
                for j in range(len(doc2.paragraphs[i].runs)):
                    doc2.paragraphs[i].runs[j].add_text(str(contracts_phoneNumber))
            if i == 59:  # 劳动合同期限
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[:4].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 5:
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[5:7].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 8:
                        doc2.paragraphs[i].runs[j].text = str(contracts_dateOfEntry)[8:10].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 12:
                        doc2.paragraphs[i].runs[j].text = str(contracts_contractExpirationDate)[:4].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 16:
                        doc2.paragraphs[i].runs[j].text = str(contracts_contractExpirationDate)[5:7].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 20:
                        doc2.paragraphs[i].runs[j].text = str(contracts_contractExpirationDate)[8:10].center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j in [2, 3, 6, 9, 10, 13, 14, 17, 18, 21, 22]:
                        doc2.paragraphs[i].runs[j].text = ''

            if i == 168 or i == 351:  # 姓名
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 5:
                        doc2.paragraphs[i].runs[j].text = contracts_name.center(18, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
            if i == 0 or i == 14 or i == 165:  # 江苏润阳新能源股份科技有限公司
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 3:
                        doc2.paragraphs[i].runs[j].text = jobRank

            # if i == 15:  # 甲方
            #     for j in range(len(doc2.paragraphs[i].runs)):
            #         print(j,doc2.paragraphs[i].runs[j].text)
            if i == 45:  # 甲方（用人单位）：  江苏润阳新能源股份科技有限公司
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = jobRank

            if i == 46:  # 注册地址：         江苏省盐城市湘江路58号
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[0]

            if i == 47:  # 注册类型：         有限责任公司
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 1:
                        doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[1]
            if i == 166 or i == 48 or i == 363:  # 法定代表人: 陶龙忠
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 2:
                        doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[2].center(14, ' ')
            # if i == 349:  # 甲方:	江苏润阳新能源科技股份有限公司
            #     for j in range(len(doc2.paragraphs[i].runs)):
            #         if j == 3:
            #             doc2.paragraphs[i].runs[j].text = jobRank
            #         if j == 4:
            #             doc2.paragraphs[i].runs[j].text = ''
            #         # print(j,doc2.paragraphs[i].runs[j].text)
            if i == 62:  # 试用期
                for j in range(len(doc2.paragraphs[i].runs)):
                    # print(j, doc2.paragraphs[i].runs[j].text)
                    if j == 2:  #还有个3
                        doc2.paragraphs[i].runs[j].text = '3'
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
            if i == 64:  # 从事工作
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 2:
                        # print(j,doc2.paragraphs[i].runs[j].text)
                        doc2.paragraphs[i].runs[j].text = '追光者'.center(6, ' ')
                        doc2.paragraphs[i].runs[j].font.underline = True  # 添加下划线
                    if j == 3:
                        doc2.paragraphs[i].runs[j].text = ''
            if i == 65 :  # 社会保险地
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 2:
                        doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[3]
                    if j in [3,4]:
                        doc2.paragraphs[i].runs[j].text =''
            if i == 90:  # 社会保险地
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 3:
                        doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[3]


            if i == 362:  # 甲方:	江苏润阳新能源科技股份有限公司
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 3:
                        doc2.paragraphs[i].runs[j].text = jobRank
            # if i == 422:  # 422甲方：润阳新能源（上海）有限公司
            #     for j in range(len(doc2.paragraphs[i].runs)):
            #         if j == 2:
            #             doc2.paragraphs[i].runs[j].text = jobRank
            # if i==66:  #(二)	双方对劳动合同履行地约定为
            #     for j in range(len(doc2.paragraphs[i].runs)):
            #         if j == 2:
            #             doc2.paragraphs[i].runs[j].text = self.findExcel(jobRank)[3]
            if i == 364:  # 乙方: 姓名:
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 5:
                        doc2.paragraphs[i].runs[j].text = contracts_name.center(14, ' ')
            # if i == 424:  # 乙方：                             身份证号：
            #     for j in range(len(doc2.paragraphs[i].runs)):
            #         doc2.paragraphs[i].runs[j].text = '乙方：{}                       身份证号：{}'.format(contracts_name,
            #                                                                                              str(contracts_idNumber))
            if i == 427:  # 422甲方：润阳新能源（上海）有限公司
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j == 2:
                        doc2.paragraphs[i].runs[j].text = jobRank
            if i == 429:  # 乙方：                             身份证号：
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j==0:
                        doc2.paragraphs[i].runs[j].text = '乙方：{}                       身份证号：{}'.format(contracts_name,str(contracts_idNumber))

            if i==83:
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j==3:
                        if base_salary=='None' or len(base_salary)==0:
                            doc2.paragraphs[i].runs[j].text =' '
                        else:
                            doc2.paragraphs[i].runs[j].text=base_salary
                    if j in [4,5]:
                        doc2.paragraphs[i].runs[j].text = ''
                    # try:
                    #     print(j,doc2.paragraphs[i].runs[j].text)
                    # except:
                    #     pass
            if i==85:
                for j in range(len(doc2.paragraphs[i].runs)):
                    if j==2:
                        if base_salary == 'None' or len(base_salary) == 0:
                            doc2.paragraphs[i].runs[j].text = ' '
                        else:
                            doc2.paragraphs[i].runs[j].text = base_salary
                    if j ==1:
                        doc2.paragraphs[i].runs[j].text = ''
                    # try:
                    #     print(j,doc2.paragraphs[i].runs[j].text)
                    # except:
                    #     pass
            # print(str(i), doc2.paragraphs[i].text)

        # name2 = '入职资料-{}（合同、保密、竞业）{}.docx'.format(jobRank, '_' + contracts_name + str(contracts_workNumber)[
        #                                                                                     :4] + 'Onboarding' + str(time.time()))
        name2 = '{}_入职资料-{}（合同、保密、竞业）.docx'.format(contracts_name + str(contracts_workNumber),jobRank)
        path2 = 'static/volumeContractsFile/info_file/{}/{}/{}'.format(t,info['name']+info['code'], name2)
        doc2.save(path2)
        file2_kwargs = {
            "url": path2,
            "name": name2,
            'filetype': 2
        }
        # ContractsInfo.objects.filter(id=id, contracts_status=True).update(contracts_placeWork=self.findExcel(jobRank)[3])
        try:
            infoObj = ContractsInfo.objects.get(id=id)
            infoId = infoObj.contracts_infoFile.filter(filetype=2, file_status=True).first().id
        except:
            infoId = None
        file_obj = ContractsFile.objects.update_or_create(defaults=file2_kwargs, id=infoId)[0]
        file_obj.contracts_infoFile.add(obj.id)

        # folderPath = 'static/volumeContractsFile/info_file/{}/{}'.format(t, info['name']+info['code'])
        folderPath = 'static/volumeContractsFile/info_file/{}'.format(t)
        return (folderPath, info['name'])

    def delete_data(self):  # 删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(ContractsInfo, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }


    def download_file(self):
        if self.request.check_token != None:
            # isContract= self.request.GET.get('isContract', 'None')
            downloadType= self.request.GET.get('downloadType', 'None')
            # print(downloadType,type(downloadType))
            # print(isContract,type(isContract))
            if downloadType=='1':   #合同下载
                # print('合同下载')
                id_list = self.request.data.get('idList')
                downloadAll = self.request.data.get('downloadAll')
                t = time.strftime('%Y-%m-%d')
                t1=str(time.time())
                dummy_path = os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'zip_file', t,t1,
                                                                                                      'copy')  # 创建文件夹
                self.mkdir(dummy_path)
                zip_path = os.path.join('static', 'volumeContractsFile', 'zip_file', t, t1,'合同下载.zip')  # 压缩后文件存放的路径
                zip_path = zip_path.replace('\\', '/')
                if downloadAll == True:  # 是下载全部   有条件
                    kwargs = {'contracts_status': True}
                    searchName = self.request.GET.get('searchName', None)  # 工号 /姓名
                    beginDate = self.request.GET.get('beginDate', None)
                    endDate = self.request.GET.get('endDate', None)
                    # jobRankId=self.request.GET.get('jobRankId', None)   #合同归属

                    baseNameId = self.request.GET.get('baseNameId', None)



                    if baseNameId == '' and beginDate == "" and endDate == "":  # 全查
                        kwargs['contracts_status'] = True
                    if len(baseNameId) == 0 or baseNameId == None:
                        kwargs['jobRank__in'] = self.request.user_jobRank
                    else:
                        kwargs['jobRank'] = baseNameId
                    if beginDate != "" and endDate != "":
                        kwargs['entryData__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                                       177) if beginDate == None else beginDate
                        kwargs['entryData__lte'] = datetime(2521, 10, 29, 7, 17, 1,
                                                                       177) if endDate == None else endDate

                    all =list(ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs).all().values_list('id',flat=True))
                    for id in all:
                        try:
                            flag=ContractsInfo.objects.filter(id=id,contracts_status=True).values_list('approval_status',flat=True)[0] #是否审批
                        except:
                            flag=False
                        if flag:

                            try:
                                infoObj = ContractsInfo.objects.get(id=id,contracts_status=True,approval_status=True)
                                infoId = infoObj.contracts_infoFile.filter(file_status=True).values_list('id',flat=True)
                            except:
                                infoId = None
                            for pk in infoId:

                                file_url = ContractsFile.objects.get(id=pk).url
                                # folderpath=file_url  #待压缩文件夹的路径
                                file_ls=file_url.split('/')[:-1]
                                folderpath="/".join(file_ls)
                                self.FindFile(folderpath,dummy_path)#要复制的文件所在目录     新路径

                                # if ContractsInfo.objects.get(id=id, contracts_status=True,approval_status=True).pic_file:
                                #     pic_url=ContractsInfo.objects.get(id=id, contracts_status=True,approval_status=True).pic_file
                                #     pic_ls=pic_url.split('/')[:-1]
                                #     folderpath2="/".join(pic_ls)
                                #     self.FindFile(folderpath2,dummy_path)#要复制的文件所在目录     新路径`
                                # else:
                                #     pass


                            self.folder_to_zip(dummy_path, zip_path)
                            self.return_data = {
                                "code": status.HTTP_200_OK,
                                "msg": '下载成功！',
                                'downloadUrl': zip_path
                            }
                            # print(self.return_data)
                        else:
                            self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": '未审批,无法下载！'
                            }
                            continue
                else:
                    for id in id_list:
                        try:
                            flag=ContractsInfo.objects.filter(id=id,contracts_status=True).values_list('approval_status',flat=True)[0] #是否审批
                        except:
                            flag=False
                        if flag:

                            try:
                                infoObj = ContractsInfo.objects.get(id=id,contracts_status=True,approval_status=True)
                                infoId = infoObj.contracts_infoFile.filter(file_status=True).values_list('id',flat=True)
                            except:
                                infoId = None
                            for pk in infoId:
                                file_url = ContractsFile.objects.get(id=pk).url
                                # folderpath=file_url  #待压缩文件夹的路径
                                file_ls=file_url.split('/')[:-1]
                                folderpath="/".join(file_ls)
                                self.FindFile(folderpath,dummy_path)#要复制的文件所在目录     新路径
                            self.folder_to_zip(dummy_path, zip_path)
                            self.return_data = {
                                "code": status.HTTP_200_OK,
                                "msg": '下载成功！',
                                'downloadUrl': zip_path
                            }

                        else:
                            self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": '未审批,无法下载！'
                            }
                            continue

                shutil.rmtree(dummy_path)
            elif downloadType == '2':   #照片下载
                # print('照片下载')
                id_list = self.request.data.get('idList')
                downloadAll = self.request.data.get('downloadAll')
                t = time.strftime('%Y-%m-%d')
                t1=str(time.time())
                # dummy_path = os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'pic_file', t,t1)  # 创建文件夹
                # self.mkdir(dummy_path)
                # zip_path = os.path.join('static', 'volumeContractsFile', 'zip_file', t, t1,'照片下载.zip')  # 压缩后文件存放的路径
                zip_path = os.path.join('static', 'volumeContractsFile', 'zip_file', t, t1,)  # 压缩后文件存放的路径
                zip_path = zip_path.replace('\\', '/')
                self.mkdir(zip_path)
                # zipf = zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED)
                if downloadAll == True:  # 是下载全部   有条件
                    kwargs = {'contracts_status': True}
                    searchName = self.request.GET.get('searchName', None)  # 工号 /姓名
                    beginDate = self.request.GET.get('beginDate', None)
                    endDate = self.request.GET.get('endDate', None)
                    jobRankId=self.request.GET.get('baseNameId', None)   #合同归属

                    if jobRankId == '' and beginDate == "" and endDate == "":  # 全查
                        kwargs['contracts_status'] = True
                    if jobRankId != '':
                        kwargs['jobRank'] = jobRankId
                    if beginDate != "" and endDate != "":
                        kwargs['entryData__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                                       177) if beginDate == None else beginDate
                        kwargs['entryData__lte'] = datetime(2521, 10, 29, 7, 17, 1,
                                                                       177) if endDate == None else endDate

                    all =list(ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs).all().values_list('id',flat=True))

                    pic_ls=[]
                    for id in all:
                        if ContractsInfo.objects.get(id=id, contracts_status=True).pic_file:
                            pic_url=ContractsInfo.objects.get(id=id, contracts_status=True).pic_file
                            # print(pic_url)
                            path = pathlib.Path(pic_url)

                            if path.is_file():
                                pic_ls.append(pic_url)
                            else:
                                pass
                        else:
                            pass
                    zip_path='static/volumeContractsFile/zip_file/{}/{}/照片下载.zip'.format(t,t1)
                    self.zip_files(pic_ls, file_o=zip_path)
                else:
                    pic_ls=[]
                    for id in id_list:
                        # print(id)
                        if ContractsInfo.objects.get(id=id, contracts_status=True).pic_file:
                            pic_url=ContractsInfo.objects.get(id=id, contracts_status=True).pic_file
                            # print(pic_url)

                            path = pathlib.Path(pic_url)

                            if path.is_file():
                                pic_ls.append(pic_url)
                            else:
                                pass
                        else:
                            pass
                    # print(pic_ls)
                    zip_path='static/volumeContractsFile/zip_file/{}/{}/照片下载.zip'.format(t,t1)
                    self.zip_files(pic_ls, file_o=zip_path)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": '下载成功！',
                    'downloadUrl': zip_path
                }
            elif downloadType=='3':
                id_list = self.request.data.get('idList')
                downloadAll = self.request.data.get('downloadAll')
                t = time.strftime('%Y-%m-%d')
                t1 = str(time.time())
                dummy_path = os.path.join(BASE_DIR, 'static', 'volumeContractsFile', 'download_data', t, t1)  # 创建文件夹
                self.mkdir(dummy_path)
                path = self.createExcelPath(t1, '追光者信息表.xlsx')
                if downloadAll == True:  # 是下载全部   有条件
                    kwargs = {'contracts_status': True}
                    searchName = self.request.GET.get('searchName', None)  # 工号 /姓名
                    beginDate = self.request.GET.get('beginDate', None)
                    endDate = self.request.GET.get('endDate', None)
                    # jobRankId=self.request.GET.get('jobRankId', None)   #合同归属

                    baseNameId = self.request.GET.get('baseNameId', None)

                    if baseNameId == '' and beginDate == "" and endDate == "":  # 全查
                        kwargs['contracts_status'] = True
                    if len(baseNameId) == 0 or baseNameId == None:
                        kwargs['jobRank__in'] = self.request.user_jobRank
                    else:
                        kwargs['jobRank'] = baseNameId
                    if beginDate != "" and endDate != "":
                        kwargs['entryData__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                            177) if beginDate == None else beginDate
                        kwargs['entryData__lte'] = datetime(2521, 10, 29, 7, 17, 1,
                                                            177) if endDate == None else endDate

                    all = list(ContractsInfo.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),
                                                            **kwargs).all().values_list('id', flat=True))
                    row_data = []
                    index = 1
                    for id in all:
                        table_data = ContractsInfo.objects.filter(contracts_status=True, id=id).values_list('code',
                                                                                                            'name',
                                                                                                            'gender',
                                                                                                            'nativePlaceId',
                                                                                                            'idCard',
                                                                                                            'birthday',
                                                                                                            'phone',
                                                                                                            'entryData',
                                                                                                            'politicsStatus',
                                                                                                            'accountNature',
                                                                                                            'domicileAddress',
                                                                                                            'nationId',
                                                                                                            'legalAddress',
                                                                                                            'marriage',
                                                                                                            'urgentPerson',
                                                                                                            'urgentPersonPhone',
                                                                                                            'urgentRelation',
                                                                                                            'latestDegreeId',
                                                                                                            'graduateTime',
                                                                                                            'graduateInsituation',
                                                                                                            'major',
                                                                                                            'educateMethod',
                                                                                                            'bankIdCard',
                                                                                                            'openBank',
                                                                                                            'contractExpirationDate',
                                                                                                            'probation',
                                                                                                            'summerSize',
                                                                                                            'height',
                                                                                                            'weight',
                                                                                                            'bust',
                                                                                                            'bellyCircumference',
                                                                                                            'department',
                                                                                                            'contracts_posts',
                                                                                                            'contracts_placeWork',
                                                                                                            'jobRank',
                                                                                                            'base_salary',
                                                                                                            'contracts_remark')
                        obj = ContractsInfo.objects.filter(id=id, contracts_status=True).first()

                        if table_data.exists():
                            for i in table_data:
                                line = list(i)
                                line.insert(0, index)
                                line[3] = obj.get_gender_display()
                                line[4] = obj.get_nativePlaceId_display()
                                line[9] = obj.get_politicsStatus_display()  # 政治面貌
                                line[10] = obj.get_accountNature_display()
                                line[12] = obj.get_nationId_display()  # 民族
                                line[14] = obj.get_marriage_display()  # 婚
                                line[17] = obj.get_urgentRelation_display()  # 关系
                                line[18] = obj.get_latestDegreeId_display()  # 最高学历
                                line[22] = obj.get_educateMethod_display()  # 教育方式
                                line[27] = obj.get_summerSize_display()  # 教育方式
                                line[-3] = obj.get_jobRank_display()  # 合同归属地

                                row_data.append(list(line))
                                if len(line) == 0:
                                    index = index
                                index += 1

                    exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                    sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                    for row in row_data:
                        sheet.append(row)  # 在工作表中添加一行
                    exc.save(path)  # 指定路径,保存文件
                else:
                    row_data = []
                    index = 1
                    for id in id_list:
                        table_data = ContractsInfo.objects.filter(contracts_status=True, id=id).values_list('code',
                                                                                                            'name',
                                                                                                            'gender',
                                                                                                            'nativePlaceId',
                                                                                                            'idCard',
                                                                                                            'birthday',
                                                                                                            'phone',
                                                                                                            'entryData',
                                                                                                            'politicsStatus',
                                                                                                            'accountNature',
                                                                                                            'domicileAddress',
                                                                                                            'nationId',
                                                                                                            'legalAddress',
                                                                                                            'marriage',
                                                                                                            'urgentPerson',
                                                                                                            'urgentPersonPhone',
                                                                                                            'urgentRelation',
                                                                                                            'latestDegreeId',
                                                                                                            'graduateTime',
                                                                                                            'graduateInsituation',
                                                                                                            'major',
                                                                                                            'educateMethod',
                                                                                                            'bankIdCard',
                                                                                                            'openBank',
                                                                                                            'contractExpirationDate',
                                                                                                            'probation',
                                                                                                            'summerSize',
                                                                                                            'height',
                                                                                                            'weight',
                                                                                                            'bust',
                                                                                                            'bellyCircumference',
                                                                                                            'department',
                                                                                                            'contracts_posts',
                                                                                                            'contracts_placeWork',
                                                                                                            'jobRank',
                                                                                                            'base_salary',
                                                                                                            'contracts_remark')
                        obj = ContractsInfo.objects.filter(id=id, contracts_status=True).first()

                        if table_data.exists():
                            for i in table_data:
                                line = list(i)
                                line.insert(0, index)
                                line[3] = obj.get_gender_display()
                                line[4] = obj.get_nativePlaceId_display()
                                line[9] = obj.get_politicsStatus_display()  # 政治面貌
                                line[10] = obj.get_accountNature_display()
                                line[12] = obj.get_nationId_display()  # 民族
                                line[14] = obj.get_marriage_display()  # 婚
                                line[17] = obj.get_urgentRelation_display()  # 关系
                                line[18] = obj.get_latestDegreeId_display()  #最高学历
                                line[22] = obj.get_educateMethod_display()  # 教育方式
                                line[27] = obj.get_summerSize_display()  # 教育方式
                                line[-3]=obj.get_jobRank_display()  # 合同归属地

                                row_data.append(list(line))
                                if len(line) == 0:
                                    index = index
                                index += 1


                    exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                    sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                    for row in row_data:
                        sheet.append(row)  # 在工作表中添加一行
                    exc.save(path)  # 指定路径,保存文件
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": '下载成功！',
                    'downloadUrl': path
                }
        else:
            self.return_data = {
                "code": HTTP_403_FORBIDDEN,
                "msg": "没有权限访问",
                'hidden': False
            }



    def createExcelPath(selfself,t1, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 123):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i >= 104:
                sheet.column_dimensions[chr(i).upper()].width = 18
        sheet.column_dimensions['AJ'].width=40
        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:AL1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '追光者信息表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号', '工号', '姓名','性别', '籍贯', '身份证号','出生日期', '手机号', '入职日期','政治面貌','户口性质','户籍地址','民族','法定送达地址','婚姻状况','紧急联系人姓名','紧急联系人电话','与紧急联系人关系','最高学历','最高学历毕业时间','毕业学校','毕业专业','教育方式','卡号','开户银行','合同到期日','试用期(月)','夏装尺寸','身高','体重','胸围','肚围','部门','岗位','工作地','合同归属','基本工资','备注'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'volumeContractsFile', 'download_data', t,t1, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def mkdir(self, path):
        folder = os.path.exists(path)
        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            # print('文件夹创建成功：', path)
        else:
            pass
            # print('文件夹已经存在：', path)

    def findExcel(self, name):
        import openpyxl
        exc = openpyxl.load_workbook("static/volumeContractsFile/report_template_file/追光者入职合同归属.xlsx")  # 打开整个excel文件
        # 获取工作表
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        registeredAddress = ""
        registeredType = ""
        registeredPeople = ""
        workPlace = ""
        for i in range(1, sheet.max_row):
            if sheet.cell(i + 1, 1).value == name:
                registeredAddress = sheet.cell(i + 1, 2).value
                registeredType = sheet.cell(i + 1, 3).value
                registeredPeople = sheet.cell(i + 1, 4).value
                workPlace = sheet.cell(i + 1, 5).value
            # else:
            #     registeredAddress=''
            #     registeredType = ''
            #     registeredPeople = ''
        return (registeredAddress, registeredType, registeredPeople, workPlace)

    # 定义压缩文件夹的方法
    def folder_to_zip(self, folderpath, zipath):
        """
        folderpath : 待压缩文件夹的路径
        zipath     ：压缩后文件存放的路径
        """

        zip = zipfile.ZipFile(zipath, "w", zipfile.ZIP_DEFLATED)
        for path, dirnames, filenames in os.walk(folderpath):
            # 将待压缩文件夹下的所有文件逐一添加到压缩文件
            fpath = path.replace(os.path.dirname(folderpath), '')
            for filename in filenames:
                zip.write(os.path.join(path, filename), os.path.join(fpath, filename))

        zip.close()



        # old_path = r'F:\1'  # 要复制的文件所在目录
        # new_path = r'F:\2'  # 新路径

    def FindFile(self,path,new_path):   #把一个文件夹内(包含子文件夹)的所有文件复制到另一个文件夹下
        for ipath in os.listdir(path):
            fulldir = os.path.join(path, ipath)  # 拼接成绝对路径
            # print(fulldir)  # 打印相关后缀的文件路径及名称
            if os.path.isfile(fulldir):  # 文件，匹配->打印
                shutil.copy(fulldir, new_path)
            if os.path.isdir(fulldir):  # 目录，递归
                self.FindFile(fulldir)

    def createPath(self, pic,func,path):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func}.{pic_suffix}"
        # print(pic_name)
        pic_path = os.path.join('static', 'volumeContractsFile', 'upload_file', t,path,pic_name)  # 图片路径
        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'volumeContractsFile','upload_file', t,path,pic_name)  # 图片路径
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'volumeContractsFile', 'upload_file', t,path, pic_name)  # 文件路径
        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix,path):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/volumeContractsFile/upload_file/' + t + '/' + path+'/' +pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/volumeContractsFile/upload_file/' + t + '/' + path+'/'+pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
    def zip_files(self,*files_i: list, file_o: str) -> None:
        with ZipFile(file_o, 'w') as z:
            for i in files_i:
                for f in i:
                    z.write(f, arcname=(n := os.path.basename(f)))
                    # print('zip_files:', n)