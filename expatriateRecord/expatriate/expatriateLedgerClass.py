
import pathlib
from zipfile import ZipFile
import os, datetime, json, openpyxl, time, random, string, re, docx, zipfile, os,shutil,arrow
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.generic.base import View
from auther.models import AdminUser, UploadFiles
from django.core.paginator import Paginator
from django.db.models import Q

from expatriateRecord.models import *
from memorabilia.models import *
from rest_framework.views import APIView
from rest_framework import serializers  # 序列化器
from rest_framework.generics import GenericAPIView
from datetime import datetime, date
from rest_framework.response import Response
from rest_framework import status
from memorabilia.serializers import *  # 序列化器
from general.models import center_base
from pdss.settings import BASE_DIR

from expatriateRecord.sql import *
from utils.check_token import CheckToken


class ResetExpatriateLedger:  # 外派台账
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "resetGet": self.resetGet,
            'resetPut': self.resetPut,
        }

    def meth_center(self):
        self.return_data = {'code': HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return Response(self.return_data)
        self.methods[self.meth]()
        return Response(self.return_data)
    def resetGet(self):
        columnList = [
            {'label': 'index','value':'序号','width':60},
            {'label': 'code', 'value': '工号', 'width': 100},
            {'label': 'name', 'value': '姓名', 'width': 90},
            {'label': 'idCard', 'value': '身份证号', 'width': 150},
            {'label': 'expatriate_jobRank', 'value': '合同归属', 'width': 140},
            {'label': 'expatriate_Before_Manage', 'value': '管理归属', 'width': 120},
            {'label': 'expatriate_Dept', 'value': '部门名称', 'width': 200},
            {'label': 'post', 'value': '岗位名称', 'width': 130},
            {'label': 'rank', 'value': '职级', 'width': 160},
            {'label': 'isSigned_Expatriate', 'value': '是否签订外派', 'width': 100},
            {'label': 'visaledger_info__arrival_Thailand', 'value': '落地时间', 'width': 100},  # 抵派驻地日期
            {'label': 'expatriate_Begin', 'value': '外派开始时间', 'width': 100},
            {'label': 'expatriate_End', 'value': '外派结束时间', 'width': 100},
            {'label': 'expatriate_Quality', 'value': '外派性质', 'width': 100},
            {'label': 'expatriate_Place', 'value': '外派地', 'width': 100},
            {'label': 'passport', 'value': '护照号', 'width': 160},
            {'label': 'visaledger_info__filing_Dead_90', 'value': '90天到期日', 'width': 100},
            {'label': 'visaledger_info__system_Post', 'value': 'BOI岗位', 'width': 150},
            {'label': 'visaledger_info__backSign', 'value': '回头签', 'width': 100},
            {'label': 'visaledger_info__visa_Validity_Period_End', 'value': '签证、工作证过期日', 'width': 100},
            {'label': 'visaledger_info__visa_Type', 'value': '签证类型', 'width': 100},
            {'label': 'visaledger_info__current_Progress', 'value': '签证办理状态', 'width': 100},
            {'label': 'departure_time', 'value': '回国机票', 'width': 170},
            {'label': 'arrival_time', 'value': '返回派驻地日期', 'width': 170},
            {'label': 'expatriate_passport__name', 'value': '护照首页', 'width': 200},
            {'label': 'expatriate_agreement__name', 'value': '外派协议', 'width': 200},
            {'label': 'expatriate_remark', 'value': '备注', 'width': ''},
            {'label': 'expatriate_creator__username', 'value': '创建人工号', 'width': 100},
            {'label': 'expatriate_creator__user', 'value': '创建人姓名', 'width': 100},
            {'label': 'expatriate_createTime', 'value': '创建时间', 'width': 100},
        ]
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'status': True
        }
        searchName = self.request.GET.get('searchName', None)
        if "jobRankid" in self.request.GET:
            jobRankId = self.request.GET.get('jobRankid', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank'] = jobRankId
        if "jobRankid[]" in self.request.GET:
            jobRankId = self.request.GET.getlist('jobRankid[]', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank__in'] = jobRankId

        beginDate = self.request.GET.get('beginDate', None)
        endDate = self.request.GET.get('endDate', None)
        expatriate_Place = self.request.GET.get('expatriate_Place', None)
        if len(expatriate_Place) != 0:
            kwargs['expatriate_Place'] = expatriate_Place
        if beginDate != "" and endDate != "":
            kwargs['expatriate_Begin__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                endDate) == 0 else beginDate
            kwargs['expatriate_Begin__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                endDate) == 0 else endDate
        totalNumber = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName)| Q(idCard__contains=searchName),**kwargs).count()
        tableList = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName)| Q(idCard__contains=searchName),**kwargs).values('id',
                                                                       'code',
                                                                       'name',
                                                                       'idCard',
                                                                       'expatriate_jobRank',
                                                                       'expatriate_Before_Manage',
                                                                       'expatriate_Dept',
                                                                       'post',
                                                                       'rank',
                                                                       'isSigned_Expatriate',
                                                                       'visaledger_info__arrival_Thailand',  # 抵泰日期
                                                                       'expatriate_Begin',
                                                                       'expatriate_End',
                                                                       'expatriate_Quality',
                                                                       'expatriate_Place',
                                                                       'passport',#护照号
                                                                       'visaledger_info__filing_Dead_90',
                                                                       'visaledger_info__system_Post',  #岗位名称（系统岗位名称）
                                                                       'visaledger_info__backSign',
                                                                       'visaledger_info__visa_Validity_Period_End',
                                                                       'visaledger_info__visa_Type',
                                                                       'visaledger_info__current_Progress',
                                                                       'expatriate_remark',
                                                                       'expatriate_Begin',#不显示
                                                                       'expatriate_End',#不显示
                                                                       # '回国机票',
                                                                       # '返回派驻地日期',

                                                                       # '外派协议',
                                                                       # 'expatriate_passport__name',# '护照首页',
                                                                       # 'expatriate_agreement__name',  # '护照首页',
                                                                       'expatriate_creator__username',
                                                                       'expatriate_creator__user',
                                                                       'expatriate_createTime'
                                                               ).order_by('-expatriate_createTime')[
                    (currentPage - 1) * pageSize:currentPage * pageSize]

        all_id = [item['id'] for item in tableList]
        # 计算每条信息在外派起始时间和外派结束时间范围内的最新的起飞时间和到达时间
        ticket_obj_time_return=TicketLedgerInfoList.objects.filter(people_id__in=all_id,is_assignment='回国').values('people__id','flight_date')
        ticket_obj_time_go = TicketLedgerInfoList.objects.filter(people_id__in=all_id, is_assignment='去往派驻地').values('people__id', 'flight_date')
        # print(ticket_obj_time_return)
        # print(ticket_obj_time_go)
        expat_file_obj_all=ExpatriateInfoList.objects.filter(id__in=all_id).values('id','expatriate_passport__id','expatriate_passport__name','expatriate_passport__url','expatriate_agreement__id','expatriate_agreement__name',"expatriate_agreement__url")
        # print(expat_file_obj_all)
        # print(tableList)
        for index, item in enumerate(tableList):
            # print(item['id'],item['expatriate_Begin'],item['expatriate_End'],type(item['expatriate_End']))
            max_return=None  #
            max_go=None   #
            for ticket_obj in ticket_obj_time_go:
                if ticket_obj['people__id']==item['id']:
                    departure_time=ticket_obj['flight_date']
                    if departure_time is None:
                        max_go= None
                    elif max_go is None or departure_time > max_go:
                        max_go = departure_time
            for ticket_obj in ticket_obj_time_return:
                if ticket_obj['people__id']==item['id']:
                    departure_time=ticket_obj['flight_date']
                    if departure_time is None:
                        max_return = None
                    elif max_return is None or departure_time > max_return:
                        max_return = departure_time



            for expat_file_obj in expat_file_obj_all:
                if expat_file_obj['id']==item['id']:
                    item['expatriate_passport'] = [{'id': expat_file_obj['expatriate_passport__id'],
                                                          'url': expat_file_obj['expatriate_passport__url'],
                                                          'name': expat_file_obj['expatriate_passport__name']}
                                                         ]
                    item['expatriate_agreement'] = [{'id': expat_file_obj['expatriate_agreement__id'],
                                                           'url': expat_file_obj['expatriate_agreement__url'],
                                                           'name': expat_file_obj['expatriate_agreement__name']}]
                    item['expatriate_passport__name']=expat_file_obj['expatriate_passport__name']
                    item['expatriate_agreement__name']=expat_file_obj['expatriate_agreement__name']
            item['departure_time']=max_return
            item['arrival_time'] = max_go
            item['expatriate_createTime'] = str(item['expatriate_createTime'])[:10]
            item['index'] = (currentPage - 1) * pageSize + index + 1
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }

   }

    def resetPut(self):
        # print(self.request.body)
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        file_ls = ['序号', '工号', '姓名', '身份证号码', '合同归属', '管理归属', '部门名称', '岗位名称', '职级', '是否签订外派', '落地时间', '外派开始时间', '外派结束时间', '外派性质', '外派地', '护照号', '90天到期日', 'BOI岗位', '回头签', '签证、工作证过期日', '签证类型', '签证办理状态', '备注', '回国机票', '返回派驻地日期',  '创建人工号', '创建人姓名', '创建时间']
        path = self.createExcelPath('外派台账表.xlsx', str(t2), '外派台账', 30, 'A1:AB1',*file_ls)
        # ticket_obj_time = TicketLedgerInfoList.objects.filter(people_id__in=id_list).values('people__id','departure_time','arrival_time')
        ticket_obj_time_return=TicketLedgerInfoList.objects.filter(people_id__in=id_list,is_assignment='回国').values('people__id','flight_date')
        ticket_obj_time_go = TicketLedgerInfoList.objects.filter(people_id__in=id_list, is_assignment='去往派驻地').values('people__id', 'flight_date')
        isAnnex =json.loads(self.request.body).get('isAnnex',None)


        if isAnnex ==True:#下载附件
            zip_path = os.path.join('static', 'expatriateRecordFile', 'zip_file', t1, str(t2),)  # 压缩后文件存放的路径
            zip_path = zip_path.replace('\\', '/')
            self.mkdir(zip_path)
            file_ls=[]
            if downloadAll==True:
                row_data = []
                index = 1
                kwargs = {
                    "status": 1
                }
                searchName = self.request.GET.get('searchName', None)
                if "jobRankid" in self.request.GET:
                    jobRankId = self.request.GET.get('jobRankid', None)
                    if len(jobRankId) != 0:  # 有值
                        kwargs['expatriate_jobRank'] = jobRankId
                if "jobRankid[]" in self.request.GET:
                    jobRankId = self.request.GET.getlist('jobRankid[]', None)
                    if len(jobRankId) != 0:  # 有值
                        kwargs['expatriate_jobRank__in'] = jobRankId

                beginDate = self.request.GET.get('beginDate', None)
                endDate = self.request.GET.get('endDate', None)
                expatriate_Place = self.request.GET.get('expatriate_Place', None)
                if len(expatriate_Place) != 0:
                    kwargs['expatriate_Place'] = expatriate_Place
                if beginDate != "" and endDate != "":
                    kwargs['expatriate_Begin__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                        endDate) == 0 else beginDate
                    kwargs['expatriate_Begin__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                        endDate) == 0 else endDate
                id_list = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName) | Q(idCard__contains=searchName),**kwargs).values_list('id',flat=True)
                for id in id_list:
                    file_all = ExpatriateInfoList.objects.filter(id=id, status=True).values_list(
                        'expatriate_passport__url', 'expatriate_agreement__url')  # 考虑一下信息id对应多个文件id
                    passport_url = file_all[0][0]
                    agreement_url = file_all[0][1]
                    if passport_url:
                        passport_path = pathlib.Path(passport_url)  # 校验本地是否存在该路径
                        if passport_path.is_file() or passport_path is not None:
                            file_ls.append(passport_path)
                        else:
                            pass
                    if agreement_url:
                        agreement_path = pathlib.Path(agreement_url)  # 校验本地是否存在该路径
                        if agreement_path.is_file() or agreement_path is not None:
                            file_ls.append(agreement_path)
                        else:
                            pass
                    if agreement_url and passport_url:
                        pass  # 数据库没有该url
                    zip_path = 'static/expatriateRecordFile/zip_file/{}/{}/附件下载.zip'.format(t1, t2)
                    self.zip_files(file_ls, file_o=zip_path)
            else:
                for id in id_list:
                    file_all=ExpatriateInfoList.objects.filter(id=id, status=True).values_list('expatriate_passport__url','expatriate_agreement__url')#考虑一下信息id对应多个文件id
                    passport_url=file_all[0][0]
                    agreement_url=file_all[0][1]
                    if passport_url :
                        passport_path = pathlib.Path(passport_url)  #校验本地是否存在该路径
                        if passport_path.is_file() :
                            file_ls.append(passport_path)
                        else:
                            pass
                    if agreement_url:
                        agreement_path = pathlib.Path(agreement_url)  # 校验本地是否存在该路径
                        if agreement_path.is_file()   :
                            file_ls.append(agreement_path)
                        else:
                            pass
                    if agreement_url and passport_url :
                        pass  #数据库没有该url

                    zip_path='static/expatriateRecordFile/zip_file/{}/{}/附件下载.zip'.format(t1,t2)
                    self.zip_files(file_ls, file_o=zip_path)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '下载成功！',
                'downloadUrl': zip_path
            }
        else:
            if downloadAll == True:  # 是下载全部   有条件
                row_data = []
                index = 1
                kwargs = {
                    "status": 1
                }
                searchName = self.request.GET.get('searchName', None)
                if "jobRankid" in self.request.GET:
                    jobRankId = self.request.GET.get('jobRankid', None)
                    if len(jobRankId) != 0:  # 有值
                        kwargs['expatriate_jobRank'] = jobRankId
                if "jobRankid[]" in self.request.GET:
                    jobRankId = self.request.GET.getlist('jobRankid[]', None)
                    if len(jobRankId) != 0:  # 有值
                        kwargs['expatriate_jobRank__in'] = jobRankId

                beginDate = self.request.GET.get('beginDate', None)
                endDate = self.request.GET.get('endDate', None)
                expatriate_Place = self.request.GET.get('expatriate_Place', None)
                if len(expatriate_Place) != 0:
                    kwargs['expatriate_Place'] = expatriate_Place
                if beginDate != "" and endDate != "":
                    kwargs['expatriate_Begin__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                        endDate) == 0 else beginDate
                    kwargs['expatriate_Begin__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                        endDate) == 0 else endDate
                tableList = ExpatriateInfoList.objects.filter(
                                Q(name__contains=searchName) | Q(code__contains=searchName) | Q(idCard__contains=searchName),
                                **kwargs).values_list('id',
                                                 'code',
                                                 'name',
                                                 'idCard',
                                                 'expatriate_jobRank',
                                                 'expatriate_Before_Manage',
                                                 'expatriate_Dept',
                                                 'post',
                                                 'rank',
                                                 'isSigned_Expatriate',
                                                 'visaledger_info__arrival_Thailand',  # 抵泰日期
                                                 'expatriate_Begin',
                                                 'expatriate_End',
                                                 'expatriate_Quality',
                                                 'expatriate_Place',
                                                 'passport',  # 护照号
                                                 'visaledger_info__filing_Dead_90',
                                                 'visaledger_info__system_Post',  # 岗位名称（系统岗位名称）
                                                 'visaledger_info__backSign',
                                                 'visaledger_info__visa_Validity_Period_End',
                                                 'visaledger_info__visa_Type',
                                                 'visaledger_info__current_Progress',
                                                 'expatriate_remark',
                                                 'expatriate_creator__username',
                                                 'expatriate_creator__user',
                                                 'expatriate_createTime'
                                                 ).order_by('-expatriate_createTime')
                for line in tableList:
                    max_return = None  #
                    max_go = None  #
                    for ticket_obj in ticket_obj_time_go:
                        if ticket_obj['people__id'] == line[0]:
                            departure_time = ticket_obj['flight_date']
                            if departure_time is None:
                                max_go = None
                            elif max_go is None or departure_time > max_go:
                                max_go = departure_time
                    for ticket_obj in ticket_obj_time_return:
                        if ticket_obj['people__id'] == line[0]:
                            departure_time = ticket_obj['flight_date']
                            if departure_time is None:
                                max_return = None
                            elif max_return is None or departure_time > max_return:
                                max_return = departure_time
                    # max_departure_time = None  # 最大起飞时间
                    # max_arrival_time = None  # 最大到达时间
                    # for ticket_obj in ticket_obj_time:
                    #     if ticket_obj['people__id'] == line[0]:
                    #         arrival_time = ticket_obj['arrival_time']
                    #         departure_time = ticket_obj['departure_time']
                    #         if departure_time is None:
                    #             max_departure_time = None
                    #         elif max_departure_time is None or departure_time > max_departure_time:
                    #             max_departure_time = departure_time
                    #         if arrival_time is None:
                    #             max_arrival_time = None
                    #         elif max_arrival_time is None or arrival_time > max_arrival_time:
                    #             max_arrival_time = arrival_time
                    line = list(line)
                    line.insert(0, index)
                    del line[1]

                    if line[9]==True or line[9]=='True':
                        line[9]='是'
                    else:
                        line[9]='否'
                    line.insert(-3,max_return)
                    line.insert(-3,max_go)
                    row_data.append(line)
                    if len(line) == 0:
                        index = index
                    index += 1

                exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                for row in row_data:
                    sheet.append(row)  # 在工作表中添加一行
                exc.save(path)  # 指定路径,保存文件
            else:
                row_data = []
                index = 1
                for id in id_list:
                    data = list(ExpatriateInfoList.objects.filter(id=id, status=True).values_list('code',
                                                                           'name',
                                                                           'idCard',
                                                                           'expatriate_jobRank',
                                                                           'expatriate_Before_Manage',
                                                                           'expatriate_Dept',
                                                                           'post',
                                                                           'rank',
                                                                           'isSigned_Expatriate',
                                                                           'visaledger_info__arrival_Thailand',  # 抵泰日期
                                                                           'expatriate_Begin',
                                                                           'expatriate_End',
                                                                           'expatriate_Quality',
                                                                           'expatriate_Place',
                                                                           'passport',#护照号
                                                                           'visaledger_info__filing_Dead_90',
                                                                           'visaledger_info__system_Post',  #岗位名称（系统岗位名称）
                                                                           'visaledger_info__backSign',
                                                                           'visaledger_info__visa_Validity_Period_End',
                                                                           'visaledger_info__visa_Type',
                                                                           'visaledger_info__current_Progress',

                                                                           # 'expatriate_Begin',#不显示   起飞时间
                                                                           # 'expatriate_End',#不显示  到达时间
                                                                           'expatriate_remark',
                                                                           'expatriate_creator__username',
                                                                           'expatriate_creator__user',
                                                                           'expatriate_createTime'))[0]

                    max_return = None  #
                    max_go = None  #
                    for ticket_obj in ticket_obj_time_go:
                        if ticket_obj['people__id'] == id:
                            departure_time = ticket_obj['flight_date']
                            if departure_time is None:
                                max_go = None
                            elif max_go is None or departure_time > max_go:
                                max_go = departure_time
                    for ticket_obj in ticket_obj_time_return:
                        if ticket_obj['people__id'] == id:
                            departure_time = ticket_obj['flight_date']
                            if departure_time is None:
                                max_return = None
                            elif max_return is None or departure_time > max_return:
                                max_return = departure_time
                    # max_departure_time = None  # 最大起飞时间
                    # max_arrival_time = None  # 最大到达时间
                    # for ticket_obj in ticket_obj_time:
                    #     if ticket_obj['people__id'] == id:
                    #         arrival_time = ticket_obj['arrival_time']
                    #         departure_time = ticket_obj['departure_time']
                    #         if departure_time is None:
                    #             max_departure_time = None
                    #         elif max_departure_time is None or departure_time > max_departure_time:
                    #             max_departure_time = departure_time
                    #         if arrival_time is None:
                    #             max_arrival_time = None
                    #         elif max_arrival_time is None or arrival_time > max_arrival_time:
                    #             max_arrival_time = arrival_time
                    data=list(data)
                    data.insert(0, index)
                    data.insert(-3,max_return)
                    data.insert(-3,max_go)
                    if data[9] ==True or data[9]=='True':
                        data[9]='是'
                    else:
                        data[9]='否'
                    row_data.append(data)
                    if len(data) == 0:
                        index = index
                    index += 1
                exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                for row in row_data:
                    sheet.append(row)  # 在工作表中添加一行
                exc.save(path)  # 指定路径,保存文件
            self.return_data = {
                "code":status.HTTP_200_OK,
                "msg": "下载成功",
                "downloadUrl": path
            }

    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['G'].width =70
        sheet.column_dimensions['I'].width = 30
        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'expatriateRecordFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass
    @staticmethod
    def FindFile(self,path,new_path):   #把一个文件夹内(包含子文件夹)的所有文件复制到另一个文件夹下
        for ipath in os.listdir(path):
            fulldir = os.path.join(path, ipath)  # 拼接成绝对路径
            if os.path.isfile(fulldir):  # 文件，匹配->打印
                shutil.copy(fulldir, new_path)
            if os.path.isdir(fulldir):  # 目录，递归
                self.FindFile(fulldir)

    @staticmethod  #压缩
    def zip_files(*files_i: list, file_o: str) -> None:
        with ZipFile(file_o, 'w') as z:
            for i in files_i:
                for f in i:
                    z.write(f, arcname=(n := os.path.basename(f)))
                    # print('zip_files:', n)
