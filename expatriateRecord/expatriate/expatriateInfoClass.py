from django.db.models import Q
from expatriateRecord.models import *
from datetime import datetime, date
from rest_framework.response import Response
from rest_framework import status
from pdss.settings import BASE_DIR
from expatriateRecord.sql import *

class ResetExpatriateInfo:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "resetGet": self.resetGet,
            "resetPost": self.resetPost,
            'resetDelete': self.resetDelete,
            'resetPatch': self.resetPatch,
            'resetPut':self.resetPut,
            'optionGet' :self.optionGet,   #所有下拉菜单总和
            'infoGet' :self.infoGet,  #根据工号获取ehr数据库的相关数据 例如部门等
            'infoSelete':self.infoSelete    #根据工号获取外派信息表的该人的数据，让他选择
        }

    def meth_center(self):
        self.return_data = {'code': HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return Response(self.return_data)
        self.methods[self.meth]()
        return Response(self.return_data)
    def resetGet(self):
        columnList = [
            {'label': 'index' ,'value' :'序号' ,'width' :60},
            {'label': 'code', 'value': '工号', 'width': 100},
            {'label': 'name', 'value': '姓名', 'width': 90},
            {'label': 'date_Of_Entry', 'value': '入职日期', 'width': 140},
            {'label': 'expatriate_Dept', 'value': '外派部门', 'width': 250},
            {'label': 'post', 'value': '岗位', 'width': 140},
            {'label': 'expatriate_jobRank', 'value': '合同归属', 'width': 140},
            {'label': 'expatriate_Before_Base', 'value': '外派前中心/事业部', 'width': 140},
            {'label': 'expatriate_Before_Manage', 'value': '管理归属', 'width': 100},
            {'label': 'expatriate_Before_Factory', 'value': '外派前厂区', 'width': 100},
            {'label': 'resident_Dept', 'value': '派驻部门', 'width':200},
            {'label': 'expatriate_Cycle', 'value': '外派周期', 'width': 100},
            {'label': 'expatriate_Begin', 'value': '外派起始时间', 'width': 100},
            {'label': 'expatriate_End', 'value': '外派结束时间', 'width': 100},
            {'label': 'last_Expatriate_Begin', 'value': '上次外派开始时间', 'width': 100},
            {'label': 'last_Expatriate_End', 'value': '上次外派结束时间', 'width': 100},
            {'label': 'isCross_Division', 'value': '是否跨事业部', 'width': 100},
            {'label': 'expatriate_After_Base', 'value': '外派后中心/事业部', 'width': 100},
            {'label': 'expatriate_After_Manage', 'value': '外派后管理归属', 'width': 100},
            {'label': 'expatriate_After_Factory', 'value': '外派后厂区', 'width': 100},
            {'label': 'expatriate_Reason', 'value': '外派缘由说明', 'width': 200},
            {'label': 'expatriate_Target', 'value': '外派工作目标', 'width': 200},
            {'label': 'expatriate_Allowance', 'value': '外派津贴', 'width': 100},
            {'label': 'description_Allowance', 'value': '津贴说明', 'width':''},
            {'label': 'expatriate_Type', 'value': '外派类型', 'width': 100},
            {'label': 'expatriate_After_Cost', 'value': '外派后成本归属', 'width': 100},
            {'label': 'first_Expatriate', 'value': '首次外派时间', 'width': 100},
            # {'label': 'number_Of_Expatriate', 'value': '外派次数', 'width': 100},
            {'label': 'rank', 'value': '职级', 'width': 150},
            {'label': 'isSigned_Expatriate', 'value': '是否签订外派', 'width': 100},
            {'label': 'expatriate_Quality', 'value': '外派性质', 'width': 100},
            {'label': 'expatriate_Place', 'value': '外派地', 'width': 100},

            {'label': 'expatriate_jobClass', 'value': '职等', 'width': 100},
            {'label': 'expatriate_country', 'value': '外派国家', 'width': 100},
            {'label': 'expatriate_several_frequency', 'value': '第几次外派', 'width': 150},
            {'label': 'expatriate_hardship', 'value': '艰苦补贴', 'width': 100},
            {'label': 'expatriate_life', 'value': '生活补贴', 'width': 100},
            {'label': 'expatriate_other', 'value': '其他', 'width': 100},
            {'label': 'expatriate_allowance_currency', 'value': '津贴币种', 'width': 100},


            {'label': 'expatriate_remark', 'value': '备注', 'width': ''},
            {'label': 'expatriate_creator__username', 'value': '创建人工号', 'width': 100},
            {'label': 'expatriate_creator__user', 'value': '创建人姓名', 'width': 100},
            {'label': 'expatriate_modifier__username', 'value': '修改人工号', 'width': 100},
            {'label': 'expatriate_modifier__user', 'value': '修改人姓名', 'width': 100},
            {'label': 'expatriate_createTime', 'value': '创建时间', 'width': 100},
            {'label': 'expatriate_modifyTime', 'value': '修改时间', 'width': 100},
        ]
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'status': True
        }
        searchName = self.request.GET.get('searchName',None)
        if "jobRankid" in self.request.GET:
            jobRankId = self.request.GET.get('jobRankid', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank'] = jobRankId
        if "jobRankid[]" in self.request.GET:
            jobRankId = self.request.GET.getlist('jobRankid[]', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank__in'] = jobRankId

        beginDate = self.request.GET.get('beginDate',None)
        endDate = self.request.GET.get('endDate',None)
        expatriate_Type = self.request.GET.get('expatriate_Type', None)
        # print(searchName)
        if len(expatriate_Type)!=0:
            kwargs['expatriate_Type'] = expatriate_Type
        if beginDate != "" and endDate != "":
            kwargs['expatriate_Begin__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(endDate) == 0 else beginDate
            kwargs['expatriate_Begin__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate

        totalNumber = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs).count()
        tableList = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs).values('id',
                                                                'code',
                                                                'name',
                                                                'date_Of_Entry',
                                                                'expatriate_Dept',
                                                                'post',
                                                                'idCard',
                                                                'expatriate_jobRank',
                                                                'expatriate_Before_Base',
                                                                'expatriate_Before_Manage',
                                                                'expatriate_Before_Factory',
                                                                'resident_Dept',
                                                                'expatriate_Cycle',
                                                                'expatriate_Begin',
                                                                'expatriate_End',
                                                                'isCross_Division',
                                                                'expatriate_After_Base',
                                                                'expatriate_After_Manage',
                                                                'expatriate_After_Factory',
                                                                'expatriate_Reason',
                                                                'expatriate_Target',
                                                                'expatriate_Allowance',
                                                                'description_Allowance',
                                                                'expatriate_Type',
                                                                'expatriate_After_Cost',
                                                                'first_Expatriate',
                                                                'last_Expatriate_Begin',
                                                                'last_Expatriate_End',
                                                                'number_Of_Expatriate',
                                                                'rank',
                                                                'isSigned_Expatriate',
                                                                'expatriate_Quality',
                                                                'expatriate_Place',

                                                                'expatriate_jobClass',
                                                               'expatriate_country',
                                                               'expatriate_several_frequency',
                                                               'expatriate_hardship',
                                                               'expatriate_life',
                                                               'expatriate_other',
                                                                'expatriate_allowance_currency',



                                                                'expatriate_remark',
                                                               'expatriate_creator__username',
                                                               'expatriate_creator__user',
                                                               'expatriate_modifier__username',
                                                               'expatriate_modifier__user',
                                                                'expatriate_createTime',
                                                                'expatriate_modifyTime',
                                                               'visaledger_info__id',
                                ).order_by('-expatriate_createTime')[(currentPage - 1) * pageSize:currentPage * pageSize]


        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['expatriate_createTime']=str(item['expatriate_createTime'])[:10]
            item['expatriate_modifyTime'] = str(item['expatriate_modifyTime'])[:10]
            item['first_Expatriate'] =str(item['first_Expatriate'])[:10] if item['first_Expatriate'] is not None else ''
            item['date_Of_Entry'] = str(item['date_Of_Entry'])[:10]

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }
        # print(self.return_data)

    def resetDelete(self):
        try:
            id_list = json.loads(self.request.body).get('idList')
            delete_all = json.loads(self.request.body).get('deleteAll')
            if delete_all: #删除全部
                ExpatriateInfoList.objects.update(status=False)
            else:
                for id in id_list:
                    ExpatriateInfoList.objects.filter(pk=id).update(status=False)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功！"
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def resetPost(self):#外派
        info=self.request.POST.get('createData' ,None)
        info = json.loads(info)
        # print(info)
        if len(info['date_Of_Entry'])>10:
            info['date_Of_Entry']=info['date_Of_Entry'][:10]
        if info['expatriate_several_frequency']=='':
            info['expatriate_several_frequency']=None
        all_expatriate_obj = ExpatriateInfoList.objects.filter(status=True,  name=info['name'])
        if len(all_expatriate_obj) == 0:  # 没有数据
            # print("信息用 信息没数据")
            info['expatriate_creator_id'] = self.request.check_token
            ExpatriateInfoList.objects.create(**info)   #这个人第一次在信息表就创建
        elif len(all_expatriate_obj) == 1:  # 只有1条数据 ，刚刚新增的
            # if all_expatriate_obj[0].abbreviation_Dept is None and all_expatriate_obj[0].passport is not None:
            #     print('部门简称为空 护照号不空，说明该条数据是机票台账新增的')
            #     info['expatriate_modifier_id'] = self.request.check_token
            #     ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)   #此时ssc新增更新该条数据
            # elif all_expatriate_obj[0].date_Of_Entry is None and all_expatriate_obj[0].expatriate_Dept is None and all_expatriate_obj[0].expatriate_Cycle is None and all_expatriate_obj[0].passport is not None:
            #     print('入职日期为空,外派部门为空,外派周期为空,护照号不空,说明该条数据是签证台账新增的')
            #     info['expatriate_modifier_id'] = self.request.check_token
            #     ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)   #此时ssc 新增应更新信息表数据
            # elif all_expatriate_obj[0].passport is  None and all_expatriate_obj[0].abbreviation_Dept is None:
            #     print('护照号空,部门简称也为空,说明该数据是信息ssc新增的')
            #     info['expatriate_creator_id'] = self.request.check_token
            #     ExpatriateInfoList.objects.create(**info)  #此时ssc 新增应创建数据
            # elif all_expatriate_obj[0].expatriate_Quality is None and all_expatriate_obj[0].expatriate_Place is None and all_expatriate_obj[0].passport is None and all_expatriate_obj[0].date_Of_Entry is not None and all_expatriate_obj[0].expatriate_Begin is not None and all_expatriate_obj[0].expatriate_creator is None:
            #     print('外派性质和外派地为空,护照号为空,入职日期不为空,外派开始时间不为空,创建人姓名为空,说明该条数据是oa新增的')
            #     info['expatriate_creator_id'] = self.request.check_token
            #     ExpatriateInfoList.objects.create(**info)     #此时ssc 新增应创建信息表数据




            if all_expatriate_obj[0].abbreviation_Dept  and all_expatriate_obj[0].expatriate_Begin: #这条数据签证台账和ssc都新增过
                # print("信息用，该条数据都签证和ssc都新增了,那么就创建")
                info['expatriate_creator_id'] = self.request.check_token
                ExpatriateInfoList.objects.create(**info)  # 这个人第一次在信息表就创建

            elif all_expatriate_obj[0].abbreviation_Dept is None:  #部门简称为空  这条数据是外派信息ssc新增的   或 机票新增的
                # print('部门简称为空')
                # if all_expatriate_obj[0].expatriate_Place is None:  #外派地为空 说明这条数据是oa流程的 但是要创建
                #     # print('1')
                #     info['expatriate_creator_id'] = self.request.check_token
                #     ExpatriateInfoList.objects.create(**info)
                if all_expatriate_obj[0].expatriate_creator is None:   #平台新增的
                    # print('2')
                    info['expatriate_creator_id'] = self.request.check_token
                    ExpatriateInfoList.objects.create(**info)
                else:
                    # print('3')
                    info['expatriate_modifier_id'] = self.request.check_token
                    ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)
            else:             #部门简称不为空，说明这条数据是签证台账新增的
                # print("签证用，部门简称不空，说明此次是签证台账先新增的")
                info['expatriate_modifier_id']=self.request.check_token
                ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)
            # if all_expatriate_obj[0].abbreviation_Dept is None:  #部门简称为空  这条数据是外派信息新增的
            #     print('部门简称为空')
            #     info['expatriate_modifier_id'] = self.request.check_token
            #     ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)
            # else:             #部门简称不为空，说明这条数据是签证台账新增的
            #     print("签证用，部门简称不空，说明此次是签证台账先新增的")
            #     info['expatriate_modifier_id']=self.request.check_token
            #     ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**info)

        else:
            # print('信息多')
            expat_obj=ExpatriateInfoList.objects.filter(status=True, code=info['code'], name=info['name'],date_Of_Entry__isnull=True).order_by('-expatriate_createTime')
            # print(expat_obj)
            if expat_obj.exists():
                if expat_obj[0].expatriate_Place is None and expat_obj[0].expatriate_Dept is None and expat_obj[0].date_Of_Entry is None and expat_obj[0].expatriate_creator is not None :  #外派地 和外派部门和date_Of_Entry为空  外派曾的
                    # print('1')
                    info['expatriate_modifier_id'] = self.request.check_token
                    ExpatriateInfoList.objects.filter(pk=expat_obj[0].id).update(**info)
                elif expat_obj[0].expatriate_Place is None:  #oa数据
                    # print('2')
                    info['expatriate_modifier_id'] = self.request.check_token
                    ExpatriateInfoList.objects.filter(pk=expat_obj[0].id).update(**info)
            else:
                # print('3')     #不是同工号同姓名的
                info['expatriate_creator_id'] = self.request.check_token
                ExpatriateInfoList.objects.create(**info)

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "新增成功",
        }

    def resetPatch(self):
        info = self.request.data
        info['expatriate_modifier_id'] = self.request.check_token
        ExpatriateInfoList.objects.filter(pk=info['id']).update(**info)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }

    def resetPut(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        file_ls = ['序号', '工号', '姓名', '入职日期', '外派部门', '岗位', '合同归属', '外派前中心/事业部',
                   '管理归属', '外派前厂区', '派驻部门', '外派周期', '外派起始时间', '外派结束时间',
                   '是否跨事业部', '外派后中心/事业部', '外派后管理归属', '外派后厂区', '外派缘由说明',
                   '外派工作目标', '外派津贴', '津贴说明', '外派类型', '外派后成本归属', '首次外派时间',
                   '上次外派开始时间','上次外派结束时间',  '职级', '是否签订外派', '外派性质', '外派地','职等','外派国家','第几次外派','艰苦补贴','生活补贴','其他','津贴币种','备注',
                   '创建人工号', '创建人姓名',  '修改人工号', '修改人姓名','创建时间', '修改时间']

        path = self.createExcelPath('信息SSC台账表.xlsx', str(t2), '信息SSC台账', 46, 'A1:AS1', *file_ls)
        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            index = 1
            kwargs = {
                'status': True
            }
            searchName = self.request.GET.get('searchName', None)
            if "jobRankid" in self.request.GET:
                jobRankId = self.request.GET.get('jobRankid', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['expatriate_jobRank'] = jobRankId
            if "jobRankid[]" in self.request.GET:
                jobRankId = self.request.GET.getlist('jobRankid[]', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['expatriate_jobRank__in'] = jobRankId

            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            expatriate_Type = self.request.GET.get('expatriate_Type', None)
            # print(searchName)
            if len(expatriate_Type) != 0:
                kwargs['expatriate_Type'] = expatriate_Type
            if beginDate != "" and endDate != "":
                kwargs['expatriate_Begin__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                    endDate) == 0 else beginDate
                kwargs['expatriate_Begin__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                    endDate) == 0 else endDate
            tableList = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName),**kwargs).values_list(
                                                                                                                                       'code',
                                                                                                                                       'name',
                                                                                                                                       'date_Of_Entry',
                                                                                                                                       'expatriate_Dept',
                                                                                                                                       'post',
                                                                                                                                       'expatriate_jobRank',
                                                                                                                                       'expatriate_Before_Base',
                                                                                                                                       'expatriate_Before_Manage',
                                                                                                                                       'expatriate_Before_Factory',
                                                                                                                                       'resident_Dept',
                                                                                                                                       'expatriate_Cycle',
                                                                                                                                       'expatriate_Begin',
                                                                                                                                       'expatriate_End',
                                                                                                                                       'isCross_Division',
                                                                                                                                       'expatriate_After_Base',
                                                                                                                                       'expatriate_After_Manage',
                                                                                                                                       'expatriate_After_Factory',
                                                                                                                                       'expatriate_Reason',
                                                                                                                                       'expatriate_Target',
                                                                                                                                       'expatriate_Allowance',
                                                                                                                                       'description_Allowance',
                                                                                                                                       'expatriate_Type',
                                                                                                                                       'expatriate_After_Cost',
                                                                                                                                       'first_Expatriate',
                                                                                                                                       'last_Expatriate_Begin',
                                                                                                                                       'last_Expatriate_End',
                                                                                                                                       'rank',
                                                                                                                                       'isSigned_Expatriate',
                                                                                                                                       'expatriate_Quality',
                                                                                                                                       'expatriate_Place',

                                                                                                                                        'expatriate_jobClass',
                                                                                                                                        'expatriate_country',
                                                                                                                                        'expatriate_several_frequency',
                                                                                                                                        'expatriate_hardship',
                                                                                                                                        'expatriate_life',
                                                                                                                                        'expatriate_other',
                                                                                                                                        'expatriate_allowance_currency',

                                                                                                                                        'expatriate_remark',
                                                                                                                                       'expatriate_creator__username',
                                                                                                                                       'expatriate_creator__user',
                                                                                                                                       'expatriate_modifier__username',
                                                                                                                                       'expatriate_modifier__user',
                                                                                                                                       'expatriate_createTime',
                                                                                                                                       'expatriate_modifyTime',
                                ).order_by('-expatriate_createTime')

            for line in tableList:
                line = list(line)
                line.insert(0, index)
                # print(line)
                if line[14]==True or line[14]=='True':
                    line[14]='是'
                elif line[14] == False or line[14] == 'False':
                    line[14]='否'
                if line[29]==True or line[29]=='True':
                    line[29]='是'
                elif line[29]==False or line[29]=='False':
                    line[29]='否'
                row_data.append(line)
                if len(line) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            index = 1
            for id in id_list:
                data = list(ExpatriateInfoList.objects.filter(pk=id, status=True).values_list(
                            'code',
                            'name',
                            'date_Of_Entry',
                            'expatriate_Dept',
                            'post',
                            'expatriate_jobRank',
                            'expatriate_Before_Base',
                            'expatriate_Before_Manage',
                            'expatriate_Before_Factory',
                            'resident_Dept',
                            'expatriate_Cycle',
                            'expatriate_Begin',
                            'expatriate_End',
                            'isCross_Division',
                            'expatriate_After_Base',
                            'expatriate_After_Manage',
                            'expatriate_After_Factory',
                            'expatriate_Reason',
                            'expatriate_Target',
                            'expatriate_Allowance',
                            'description_Allowance',
                            'expatriate_Type',
                            'expatriate_After_Cost',
                            'first_Expatriate',
                            'last_Expatriate_Begin',
                            'last_Expatriate_End',
                            'rank',
                            'isSigned_Expatriate',
                            'expatriate_Quality',
                            'expatriate_Place',

                    'expatriate_jobClass',
                    'expatriate_country',
                    'expatriate_several_frequency',
                    'expatriate_hardship',
                    'expatriate_life',
                    'expatriate_other',
                    'expatriate_allowance_currency',

                            'expatriate_remark',
                            'expatriate_creator__username',
                            'expatriate_creator__user',
                            'expatriate_modifier__username',
                            'expatriate_modifier__user',
                            'expatriate_createTime',
                            'expatriate_modifyTime',
                ))[0]
                data = (index,) + data
                data=list(data)
                if data[14]==True or data[14]=='True':
                    data[14]='是'
                elif data[14] == False or data[14] == 'False':
                    data[14]='否'
                if data[29]==True or data[29]=='True':
                    data[29]='是'
                elif data[29]==False or data[29]=='False':
                    data[29]='否'

                row_data.append(data)
                if len(data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    def infoGet(self):
        emp = Employee()
        resp, status = emp.select(self.request.GET.get('code'))
        self.return_data = {
            "code": status,
            "msg": resp['msg'],
            'data':[
                {
                    "value": item["code"],
                    "address": {
                        "name": item["name"],   #姓名
                        "idCard":item['idCard'],#身份证号
                        "date_Of_Entry":item['date_Of_Entry'],#入职日期
                        "expatriate_Before_Manage": item["expatriate_Before_Manage"],  #管理/基地归属
                        "expatriate_jobRank": item["expatriate_jobRank"],  #合同归属
                        "abbreviation_Dept": item["abbreviation_Dept"],   #部门简称
                        "post": item["post"],     #岗位
                        "rank":item["rank"],   #职级
                        "expatriate_Dept":item['expatriate_Dept'],   #部门名称
                        'expatriate_jobClass':item['expatriate_jobClass']#职等

                    }
                }
                for item in resp['data']
            ]
        }

    def infoSelete(self):  #获取外派信息表的数据
        code=self.request.POST.get('code',None)
        name=self.request.POST.get('name',None)
        info= {'code': code, 'name': name, 'status': True}
        obj_data_all=[]
        all_object=ExpatriateInfoList.objects.filter(**info).all()
        for obj in all_object:
            obj_data=ExpatriateInfoList.objects.filter(pk=obj.id).values('id','visaledger_info__visa_Validity_Period_Begin','visaledger_info__visa_Validity_Period_End','visaledger_info__arrival_Thailand','expatriate_Begin','expatriate_End')
            obj_data_all.append(obj_data[0])
        return_data=[]
        for dictionary in obj_data_all:
            arrival_Thailand_str=""
            visa_Validity_Period_Begin_str=""
            expatriate_Begin_str=""
            new_dict = {}
            if dictionary["expatriate_Begin"] is not None:
                expatriate_Begin_str = f"外派起始时间为{dictionary['expatriate_Begin']}"+";"
            if dictionary["visaledger_info__visa_Validity_Period_Begin"] is not None:
                visa_Validity_Period_Begin_str = f"签证有效期开始时间为{dictionary['visaledger_info__visa_Validity_Period_Begin']}"+";"
            if dictionary["visaledger_info__arrival_Thailand"] is not None:
                arrival_Thailand_str = f"抵派驻地日期为{dictionary['visaledger_info__arrival_Thailand']}"+";"
            new_dict["label"]=expatriate_Begin_str+visa_Validity_Period_Begin_str+arrival_Thailand_str
            if new_dict:
                new_dict["value"] = str(dictionary["id"])
                return_data.append(new_dict)

        self.return_data ={
            'data':return_data,
            # 'data': [
            #             {
            #                 "value": str(obj["id"]),
            #                 "label": {
            #                     k: v for k, v in obj.items() if k != "id"
            #                 }
            #             }
            #             for obj in obj_data_all
            #         ],
            'code': HTTP_200_OK,
            'msg': '查询成功'
        }

    def optionGet(self):
        expatriate_Quality = ExpatriateQuality()  # 外派性质
        expatriate_place = ExpatriatePlace()
        expatriate_re_enter_visa = ExpatriateReEnterVisa()
        expatriate_visa_type = ExpatriateVisaType()
        expatriate_visa_handle_status = ExpatriateVisaHandleStatus()
        expatriate_Dept = ExpatriateDept()  # 部门名称
        expatriate_Manage= ExpatriateManage()  # 管理归属
        expatriate_Post = ExpatriatePost()  # 岗位名称
        expatriate_Jobrank= ExpatriateJobrank()  # 合同归属
        expatriate_Type = ExpatriateType()  # 外派类型
        expatriate_Rank = ExpatriateRank()  # 职级
        abbreviation_Dept = AbbreviationDept()  # #部门简称
        expatriate_Jobclass=ExpatriateJobclass()  # 职等
        self.return_data ={
            'data': {
                'expatriate_Quality': expatriate_Quality.select(),
                'expatriate_place': expatriate_place.select(),
                'expatriate_re_enter_visa': expatriate_re_enter_visa.select(),
                'expatriate_visa_type': expatriate_visa_type.select(),
                'expatriate_visa_handle_status': expatriate_visa_handle_status.select(),
                'expatriate_Type' :expatriate_Type.select(),  # 外派类型
                'expatriate_Dept': expatriate_Dept.select(self.request.GET.get('code')),
                'expatriate_Manage': expatriate_Manage.select(self.request.GET.get('code')),
                'expatriate_Post': expatriate_Post.select(self.request.GET.get('code')),
                'expartiate_Jobrank':expatriate_Jobrank.select(),
                'expatriate_Rank' :expatriate_Rank.select(),
                'AbbreviationDept':abbreviation_Dept.select(self.request.GET.get('code')),
                'expatriate_Jobclass':expatriate_Jobclass.select()
            },
            'code': HTTP_200_OK,
            'msg': '查询成功'
        }

    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['K'].width = 40

        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'expatriateRecordFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        # t = time.strftime('%Y-%m-%d')
        now = arrow.now()
        t = now.format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  #文件后缀

        file_name = f"{fileName}.{file_suffix}"    #文件名称

        file_path = os.path.join('static', 'expatriateRecordFile', 'upload_file', t,path,file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)



class ResetExpatriateOA:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "oaPost": self.oaPost,
            "oaPostIncrement":self.oaPostIncrement,
        }
    def meth_center(self):
        self.methods[self.meth]()
        return Response(self.return_data)

    def oaPost(self):
        try:
            expOA =ExpatriateRecordOA()
            resp, status_oa = expOA.select()#oa数据      #黄国植全部重复，王朋部分数据重复
            resp_select,status_oa=expOA.select_count()  #外派次数 上次外派起止时间（禁掉） 首次外派时间
            empOA = Employee()
            resp_rank=empOA.select_rank_idCard_abbreviation_Dept()  # 职级 身份证号 部门简称
            for index, item in enumerate(resp['data']):  # 根据code找出所有的外派时间 然后遍历每一条 根据该条记录找出所有小于该条记录的外派起始时间中最大的外派起始时间就是该条的上次外派起始时间
                # print(index,item)
                last_begin = item['expatriate_Begin']
                last_end = item['expatriate_End']
                code = item['code']
                record = {'code': code, 'time_ls': []}
                for line in resp['data']:
                    if code == line['code']:
                        record['time_ls'].append(tuple([line['expatriate_Begin'], line['expatriate_End']]))
                record['time_ls'] = [t for t in record['time_ls'] if t[0] and last_begin and t[0] < last_begin]

                if len(record['time_ls']) >= 2:
                    max_item = max(record['time_ls'], key=lambda x: x[0])
                    record['time_ls'] = [max_item]
                if len(record['time_ls']) == 0:  # 第一次的
                    item['last_Expatriate_Begin'] = last_begin
                    item['last_Expatriate_End'] = last_end
                else:
                    item['last_Expatriate_Begin'] = record['time_ls'][0][0]
                    item['last_Expatriate_End'] = record['time_ls'][0][1]
            for line in resp['data']:
                del line['doc_status']
                del line['fd_ended_time']
                del line['fd_create_time']
                for exp in resp_select['data']:
                    if line['code'] == exp['code']:
                        line.update(exp)
                        break
                for emp in resp_rank:
                    if line['code'] == emp['code']:
                        line.update(emp)
                        break
                print(line)
                # ExpatriateInfoList.objects.update_or_create(defaults=line,code=line['code'],name=line['name'],expatriate_Begin=line['expatriate_Begin'],expatriate_After_Cost=line['expatriate_After_Cost'],resident_Dept=line['resident_Dept'],status=True)
                # ExpatriateInfoList.objects.create(**line)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "OA数据导入成功！"
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "OA数据导入失败！"
            }

    def oaPostIncrement(self):
        # print("+++++++++")
        try:
            expOA =ExpatriateRecordOA()
            resp, status_oa = expOA.selectIncrement()#oa数据      #黄国植全部重复，王朋部分数据重复
            resp_select,status_oa=expOA.select_count()  #外派次数 上次外派起止时间（禁掉） 首次外派时间
            empOA = Employee()
            resp_rank=empOA.select_rank_idCard_abbreviation_Dept()  # 职级 身份证号 部门简称
            for index, item in enumerate(resp['data']):  # 根据code找出所有的外派时间 然后遍历每一条 根据该条记录找出所有小于该条记录的外派起始时间中最大的外派起始时间就是该条的上次外派起始时间
                # print(index,item)
                last_begin = item['expatriate_Begin']
                last_end = item['expatriate_End']
                code = item['code']
                record = {'code': code, 'time_ls': []}
                for line in resp['data']:
                    if code == line['code']:
                        record['time_ls'].append(tuple([line['expatriate_Begin'], line['expatriate_End']]))
                record['time_ls'] = [t for t in record['time_ls'] if t[0] and last_begin and t[0] < last_begin]

                if len(record['time_ls']) >= 2:
                    max_item = max(record['time_ls'], key=lambda x: x[0])
                    record['time_ls'] = [max_item]
                if len(record['time_ls']) == 0:  # 第一次的
                    item['last_Expatriate_Begin'] = last_begin
                    item['last_Expatriate_End'] = last_end
                else:
                    item['last_Expatriate_Begin'] = record['time_ls'][0][0]
                    item['last_Expatriate_End'] = record['time_ls'][0][1]
            for line in resp['data']:
                del line['doc_status']
                del line['fd_ended_time']
                del line['fd_create_time']
                for exp in resp_select['data']:
                    if line['code'] == exp['code']:
                        line.update(exp)
                        break
                for emp in resp_rank:
                    if line['code'] == emp['code']:
                        line.update(emp)
                        break
                # print(line)
                #
                # ExpatriateInfoList.objects.filter(code=line['code'],name=line['name'],expatriate_Begin=line['expatriate_Begin'],expatriate_After_Cost=line['expatriate_After_Cost'],resident_Dept=line['resident_Dept']).update(expatriate_jobClass=line['expatriate_jobClass'],date_Of_Entry=line['date_Of_Entry'],expatriate_jobRank=line['expatriate_jobRank'],expatriate_Before_Manage=line['expatriate_Before_Manage'])
                # ExpatriateInfoList.objects.update_or_create(defaults=line,code=line['code'],name=line['name'],expatriate_Begin=line['expatriate_Begin'],expatriate_After_Cost=line['expatriate_After_Cost'],resident_Dept=line['resident_Dept'],status=True)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "OA数据导入成功！"
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "OA数据导入失败！"
            }

class ResetExpatriateAnnex:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "annexPost": self.annexPost,
        }

    def meth_center(self):
        self.methods[self.meth]()
        return Response(self.return_data)
    def annexPost(self):
        try:
            info=eval(self.request.POST.get('createData',None))
            expatriate_passport = self.request.FILES.get("expatriate_passport", None)  # 参保证明
            expatriate_agreement = self.request.FILES.get("expatriate_agreement", None)  # 外派协议
            exprate_obj=ExpatriateInfoList.objects.filter(code=info['code'],name=info['name'],status=True).order_by('-expatriate_createTime','-expatriate_Begin').first()
            if exprate_obj is None:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "您未申请外派,或信息缺失,暂时无法添加！"
                }
            else:
                now=arrow.now().format('YYYY-MM-DD')
                dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'upload_file', now,str(info['name']) + '_' + str(info['idCard']))  # 创建文件夹
                self.mkdir(dummy_path)
                if expatriate_passport is not None:
                    expatriate_passport_url, expatriate_passport_name,expatriate_passport_suffix= self.createPath(expatriate_passport, str(info['name']) + '_' + str(info['idCard']), str(info['name']) + str(info['idCard']) + '_护照首页')
                    self.saveFile(expatriate_passport_url, expatriate_passport)  # 保存文件
                    expatriate_passport_kwargs={
                        'name':expatriate_passport_name,
                        'url':expatriate_passport_url,
                        'filetype':1
                    }
                    expatriate_passport_file_obj = ExpatriateFile.objects.create(**expatriate_passport_kwargs)
                    expatriate_passport_file_obj.expatriate_passport.add(exprate_obj.id)
                if expatriate_agreement is not None:
                    expatriate_agreement_url, expatriate_agreement_name,expatriate_agreement_suffix= self.createPath(expatriate_agreement, str(info['name']) + '_' + str(info['idCard']), str(info['name']) + str(info['idCard']) + '_外派协议')
                    # print(expatriate_agreement_url,expatriate_agreement_name,expatriate_agreement_suffix)
                    self.saveFile(expatriate_agreement_url, expatriate_agreement)  # 保存文件
                    expatriate_agreement_kwargs={
                        'name':expatriate_agreement_name,
                        'url':expatriate_agreement_url,
                        'filetype':2
                    }
                    expatriate_agreement_file_obj = ExpatriateFile.objects.create(**expatriate_agreement_kwargs)
                    expatriate_agreement_file_obj.expatriate_agreement.add(exprate_obj.id)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "上传成功！"
                }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "上传失败！"
            }
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        now=arrow.now().format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  #文件后缀
        file_name = f"{fileName}.{file_suffix}"    #文件名称
        file_path = os.path.join('static', 'expatriateRecordFile', 'upload_file', now,path,file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)

class ResetExpatriateRemainingTime:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "remainingTime": self.remainingTime,
        }
    def meth_center(self):
        self.methods[self.meth]()
        return Response(self.return_data)
    def remainingTime(self):
        try:
            today = arrow.utcnow().date()
            visa_ls=VisaLedgerInfoList.objects.filter(status=True).values('id','visa_Validity_Period_End','filing_Dead_90')  #所有签证表数据   距签证过期日剩余  距报道过期日剩余
            for visa in visa_ls:
                info={}
                if visa['visa_Validity_Period_End']:
                    info['visa_Expiration_Date']=0 if (visa['visa_Validity_Period_End']-today).days<0 else (visa['visa_Validity_Period_End']-today).days   #计算距签证过期日剩余
                else:
                    info['visa_Expiration_Date'] =None
                if visa['filing_Dead_90']:
                    info['expiration_Article']=0 if (visa['filing_Dead_90']-today).days<0 else (visa['filing_Dead_90']-today).days                                #计算距报道过期日剩余
                else:
                    info['expiration_Article'] =None
                VisaLedgerInfoList.objects.filter(pk=visa['id']).update(**info)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "剩余时间计算成功！"
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "剩余时间计算失败！"
            }


class ResetTicketrelevancyssc:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "ticketRelevancySSC": self.ticketRelevancySSC,
        }

    def meth_center(self):
        self.methods[self.meth]()
        return Response(self.return_data)

    def ticketRelevancySSC(self):
        try:
            all_expatr=ExpatriateInfoList.objects.filter(status=True).values('id','code','name','passport','expatriate_Begin','expatriate_End').order_by('-expatriate_createTime')  #ssc
            all_ticket=TicketLedgerInfoList.objects.filter(status=True,people_id__isnull=True).values('id','ticket_code','ticket_name','ticket_passport','flight_date')  #机票
            # print(all_ticket)
            for tick_obj in all_ticket:
                for expatr_obj in all_expatr:
                    if tick_obj['ticket_name']==expatr_obj['name']  and tick_obj['ticket_passport'] ==expatr_obj['passport']and expatr_obj['expatriate_Begin']<=tick_obj['flight_date']<=expatr_obj['expatriate_End']:
                        # print(tick_obj['id'],tick_obj['ticket_name'],expatr_obj['id'],expatr_obj['name'])
                        TicketLedgerInfoList.objects.filter(pk=tick_obj['id']).update(ticket_code=expatr_obj['code'],people_id=expatr_obj['id'])
                        break
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "机票与信息关联成功！"
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "机票与信息关联失败！"
            }
