
from expatriateRecord.models import *
from rest_framework.response import Response
from rest_framework import status
from expatriateRecord.sql import *
from datetime import datetime, date
from django.db.models import Q

from pdss.settings import BASE_DIR


class ResetVisaLedger:    #签证
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "resetGet": self.resetGet,
            "resetPost": self.resetPost,
            'resetDelete': self.resetDelete,
            'resetPatch': self.resetPatch,
            'resetPut':self.resetPut,
            'resetPostTemplate': self.resetPostTemplate,   #上传模板文件数据
        }

    def meth_center(self):
        self.return_data = {'code': HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return Response(self.return_data)
        self.methods[self.meth]()
        return Response(self.return_data)
    def resetGet(self):
        columnList = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'code', 'value': '工号', 'width': 100},
            {'label': 'name', 'value': '姓名', 'width': 90},
            {'label': 'expatriate_jobRank', 'value': '合同归属', 'width': 140},
            {'label': 'expatriate_Before_Manage', 'value': '基地归属', 'width': 120},
            {'label': 'abbreviation_Dept', 'value': '部门简称', 'width': 120},
            {'label': 'post', 'value': '岗位名称', 'width': 120},
            {'label': 'passport', 'value': '护照号', 'width':150},
            {'label': 'first_Docking', 'value': '首次对接日期', 'width': 100},
            {'label': 'application_Approval', 'value': 'BOI申请批文日期', 'width': 90},
            {'label': 'invitation_Letter_Provided', 'value': '批文_邀请函提供日期', 'width': 90},
            {'label': 'collectInfo', 'value': '收齐资料日期', 'width': 100},
            {'label': 'submit_Embassy', 'value': '提交使馆日期', 'width': 100},
            {'label': 'last_SupplementaryInfo', 'value': '末次补资料日期', 'width':110},
            {'label': 'signed', 'value': '出签日期', 'width': 90},
            {'label': 'arrival_Thailand', 'value': '抵派驻地日期', 'width': 110},
            {'label': 'visa_Application_Remarks', 'value': '签证申请备注','width':200},
            {'label': 'current_Progress', 'value': '当前进度', 'width': 100},
            {'label': 'sending_Embassy', 'value': '送签使馆', 'width': 90},
            {'label': 'system_Post', 'value': '系统岗位名称', 'width': 100},
            {'label': 'visa_Type', 'value': '签证类型', 'width': 90},
            {'label': 'visa_Validity_Period_Begin', 'value': '签证有效期开始时间', 'width': 100},
            {'label': 'visa_Validity_Period_End', 'value': '签证有效期到期时间', 'width': 90},
            {'label': 'visa_Expiration_Date', 'value': '距签证过期日剩余', 'width': 90},
            {'label': 'backSign', 'value': '回头签', 'width': 100},
            {'label': 'arrival_Thailand_First', 'value': '首次抵派驻地日期', 'width': 120},
            {'label': 'report_90', 'value': '90天报道申报日', 'width':120},
            {'label': 'filing_Dead_90', 'value': '90天申报截止日', 'width': 120},
            {'label': 'expiration_Article', 'value': '距报道过期日剩余', 'width': 90},
            {'label': 'isAssignOver', 'value': '外派是否结束_签证是否注销', 'width': 100},
            {'label': 'visaLedger_remark', 'value': '备注', 'width': ''},
        ]
        # print(self.request.GET)
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'status': True
        }
        # print(self.request.GET)
        searchName = self.request.GET.get('searchName',None)
        if "jobRankid" in self.request.GET:
            jobRankId = self.request.GET.get('jobRankid', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank'] = jobRankId
        if "jobRankid[]" in self.request.GET:
            jobRankId = self.request.GET.getlist('jobRankid[]', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['expatriate_jobRank__in'] = jobRankId
        visaledger_info__visa_Type = self.request.GET.get('visa_Type', None)  #签证类型
        if len(visaledger_info__visa_Type)!=0:
            kwargs['visaledger_info__visa_Type'] = visaledger_info__visa_Type
        beginDate = self.request.GET.get('beginDate',None)
        endDate = self.request.GET.get('endDate',None)
        if beginDate != "" and endDate != "":
            kwargs['visaledger_info__signed__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(endDate) == 0 else beginDate
            kwargs['visaledger_info__signed__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate  # 出签日期
        # print(kwargs)
        totalNumber = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName)| Q(passport__contains=searchName),**kwargs).count()
        tableList=ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName)| Q(passport__contains=searchName),**kwargs).values('id',
                                                                    'code',
                                                                    'name',
                                                                    'expatriate_jobRank',
                                                                    'expatriate_Before_Manage',#基地归属(外派前管理归属)
                                                                    'abbreviation_Dept',  # 部门简称
                                                                    'post',  # 岗位名称
                                                                    'passport',  # 护照号
                                                                    "visaledger_info__first_Docking",
                                                                    "visaledger_info__application_Approval",
                                                                    "visaledger_info__invitation_Letter_Provided",
                                                                    "visaledger_info__collectInfo",
                                                                    "visaledger_info__submit_Embassy",
                                                                    "visaledger_info__last_SupplementaryInfo",
                                                                    "visaledger_info__signed",
                                                                    "visaledger_info__arrival_Thailand",
                                                                    "visaledger_info__visa_Application_Remarks",
                                                                    "visaledger_info__current_Progress",
                                                                    "visaledger_info__sending_Embassy",
                                                                    "visaledger_info__system_Post",
                                                                    "visaledger_info__visa_Type",
                                                                    "visaledger_info__visa_Validity_Period_Begin",
                                                                    "visaledger_info__visa_Validity_Period_End",
                                                                    "visaledger_info__backSign",
                                                                    "visaledger_info__visa_Expiration_Date",
                                                                    "visaledger_info__arrival_Thailand_First",
                                                                    "visaledger_info__report_90",
                                                                    "visaledger_info__filing_Dead_90",
                                                                    "visaledger_info__expiration_Article",
                                                                    "visaledger_info__isAssignOver",
                                                                    'visaledger_info__visaLedger_remark',
                                                                    'visaledger_info__id'
                                                                    ).order_by('-expatriate_createTime')[
                    (currentPage - 1) * pageSize:currentPage * pageSize]


        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['first_Docking']=item['visaledger_info__first_Docking']
            item['application_Approval'] = item['visaledger_info__application_Approval']
            item['invitation_Letter_Provided'] = item['visaledger_info__invitation_Letter_Provided']
            item['collectInfo'] = item['visaledger_info__collectInfo']
            item['submit_Embassy'] = item['visaledger_info__submit_Embassy']
            item['last_SupplementaryInfo'] = item['visaledger_info__last_SupplementaryInfo']
            item['signed'] = item['visaledger_info__signed']
            item['arrival_Thailand'] = item['visaledger_info__arrival_Thailand']
            item['visa_Application_Remarks'] = item['visaledger_info__visa_Application_Remarks']
            item['current_Progress'] = item['visaledger_info__current_Progress']
            item['sending_Embassy'] = item['visaledger_info__sending_Embassy']
            item['system_Post'] = item['visaledger_info__system_Post']
            item['visa_Type'] = item['visaledger_info__visa_Type']
            item['visa_Validity_Period_Begin'] = item['visaledger_info__visa_Validity_Period_Begin']
            item['visa_Validity_Period_End'] = item['visaledger_info__visa_Validity_Period_End']
            item['backSign'] = item['visaledger_info__backSign']
            item['visa_Expiration_Date'] = item['visaledger_info__visa_Expiration_Date']
            item['arrival_Thailand_First'] = item['visaledger_info__arrival_Thailand_First']
            item['report_90'] = item['visaledger_info__report_90']
            item['filing_Dead_90'] = item['visaledger_info__filing_Dead_90']
            item['expiration_Article'] = item['visaledger_info__expiration_Article']
            item['isAssignOver'] = item['visaledger_info__isAssignOver']
            item['visaLedger_remark'] = item['visaledger_info__visaLedger_remark']
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def resetDelete(self):
        try:
            id_list = json.loads(self.request.body).get('idList')
            delete_all = json.loads(self.request.body).get('deleteAll')
            if delete_all: #删除全部
                ExpatriateInfoList.objects.update(status=False)
            else:
                for id in id_list:
                    ExpatriateInfoList.objects.filter(pk=id).update(status=False)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功！"
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def resetPost(self):   #签证新增
        # print("签证新增")
        info=self.request.POST.get('createData',None)
        info = json.loads(info)
        expat_info = {
            'code': info['code'], 'name': info['name'], 'passport': info['passport'],'expatriate_jobRank':info['expatriate_jobRank'],'expatriate_Before_Manage':info['expatriate_Before_Manage'],
            'post':info['post'],'abbreviation_Dept':info['abbreviation_Dept']
        }

        all_expatriate_obj = ExpatriateInfoList.objects.filter(status=True, code=info['code'], name=info['name'])
        code = info['code']
        abbreviation_Dept=info['abbreviation_Dept']
        name=info['name']
        del info['code']
        del info['name']
        del info['passport']
        del info['expatriate_jobRank']
        del info['expatriate_Before_Manage']
        del info['post']
        del info['abbreviation_Dept']

        # print(info)
        # print(expat_info)
        if len(all_expatriate_obj) == 0:  # 没有数据
            # print("签证用,没有数据")
            expat_info['expatriate_creator_id'] = self.request.check_token
            info['visaLedger_creator_id'] = self.request.check_token
            # print(info)
            # print(VisaLedgerInfoList.objects.create(**info).query)
            # print(expat_info)
            # from django.db import connection
            # visa_obj = VisaLedgerInfoList.objects.create(**info)
            # sql = connection.queries[-1]['sql']
            # print(sql)
            # print("++++")
            exprate_obj=ExpatriateInfoList.objects.create(**expat_info)   #这个人第一次在信息表就创建
            visa_obj=VisaLedgerInfoList.objects.create(**info)
            visa_obj.visaledger_info.add(exprate_obj.id)
        elif len(all_expatriate_obj) == 1:  # 只有1条数据
            # if
            # print()  # 该条数据机票台账新增     签证更新
            # print()  # 该条数据oa新增   签证更新
            # print()  # 该条数据签证新增   签证创建
            # print()  # 该条数据ssc新增  签证创建
            if all_expatriate_obj[0].passport is None:    #护照号为空，说明数据库的该条数据是外派信息先增加的
                # print("信息用 护照号空，此次是信息ssc先开始新增的")
                # print("++++++++++++")
                try:
                    del info['id']
                except:
                    pass
                # print(all_expatriate_obj[0])
                expat_info['expatriate_modifier']= self.request.check_token
                ExpatriateInfoList.objects.filter(pk=all_expatriate_obj[0].id).update(**expat_info)
                visa_obj = VisaLedgerInfoList.objects.create(**info)
                visa_obj.visaledger_info.add(all_expatriate_obj[0].id)

            else:              #护照号不为空，说明数据库的该条数据是签证台账先增加的
                # print("签证用 护照号不空 签证信息新增成功")
                expat_info['expatriate_creator_id'] = self.request.check_token
                info['visaLedger_creator_id'] = self.request.check_token
                expat_obj = ExpatriateInfoList.objects.create(**expat_info)  # 在信息表就创建
                visa_obj = VisaLedgerInfoList.objects.create(**info)
                visa_obj.visaledger_info.add(expat_obj.id)
            # ExpatriateInfoList.objects.filter(id=all_expatriate_obj[0].id).update(**info) #这个人在信息表只有一条数据          根据出签日期来更新 如果出签日期为空则
        else:
            # print("签证多")
            expatriate_obj = ExpatriateInfoList.objects.filter(status=True, code=code, name=name,abbreviation_Dept__isnull=True).order_by('-expatriate_createTime') #多条中 找部门简称为空的
            # print(expatriate_obj)
            # print(expatriate_obj[-1],type(expatriate_obj[-1]))
            if expatriate_obj.exists():
                # print("信息用 1")
                info['visaLedger_creator_id'] = self.request.check_token
                expat_info['expatriate_modifier_id']=self.request.check_token
                visa_obj = VisaLedgerInfoList.objects.create(**info)
                visa_obj.visaledger_info.add(expatriate_obj[0].id)
                ExpatriateInfoList.objects.filter(pk=expatriate_obj[0].id).update(**expat_info)
            else:
                # print('签证用 2')
                expat_info['expatriate_creator_id'] = self.request.check_token
                info['visaLedger_creator_id'] = self.request.check_token
                expat_obj = ExpatriateInfoList.objects.create(**expat_info)
                visa_obj = VisaLedgerInfoList.objects.create(**info)
                visa_obj.visaledger_info.add(expat_obj.id)


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "新增成功",
        }

    def resetPatch(self):
        info = self.request.data
        # print(info['id'])
        expat_id=info['id']
        expat_info = {
            'code': info['code'], 'name': info['name'], 'passport': info['passport'],'expatriate_jobRank':info['expatriate_jobRank'],'expatriate_Before_Manage':info['expatriate_Before_Manage'],
            'post':info['post'],'abbreviation_Dept':info['abbreviation_Dept']
            # ,'expatriate_modifier':self.request.check_token
        }
        # print(expat_info)
        ExpatriateInfoList.objects.filter(pk=info['id']).update(**expat_info)
        # print("+++++++++++++++++++++++++++++")
        del info['code']
        del info['name']
        del info['passport']
        del info['expatriate_jobRank']
        del info['expatriate_Before_Manage']
        del info['post']
        del info['abbreviation_Dept']
        # info['visaLedger_modifier'] = self.request.check_token
        # print(info)
        # print(info['visaledger_info__id'],type(info['visaledger_info__id']))
        if info['visaledger_info__id'] is not None:   #这行数据有信息表的数据也有签证表的数据
            info['id']=info['visaledger_info__id']
            del info['visaledger_info__id']
            VisaLedgerInfoList.objects.filter(pk=info['id']).update(**info)
        else:   #只有信息表的数据，那么info其他信息要在签证表创建
            # print('创建签证表数据')
            del info['id']
            del info['visaledger_info__id']
            visa_obj=VisaLedgerInfoList.objects.create(**info)
            # print(visa_obj)
            visa_obj.visaledger_info.add(expat_id)



        # visa_id = ExpatriateInfoList.objects.filter(pk=info['id']).values_list('visaledger_info__id', flat=True)
        # print(visa_id)
        # print(info)
        # VisaLedgerInfoList.objects.filter(pk=info['id']).update(**info)
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }

    def resetPut(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        file_ls = [
            '序号', '工号', '姓名', '合同归属', '基地归属', '部门简称', '岗位名称', '护照号',
            '首次对接日期', 'BOI申请批文日期', '批文、邀请函提供日期',
            '收齐资料日期', '提交使馆日期', '末次补资料日期', '出签日期', '抵泰日期', '签证申请备注',
            '当前进度', '送签使馆', '系统岗位名称', '签证类型', '签证有效期开始时间',
            '签证有效期到期时间', '回头签', '距签证过期日剩余', '首次抵派驻地日期', '90天报道申报日',
            '90天申报截止日', '距报道过期日剩余',  '外派是否结束/签证是否注销','备注',
        ]

        path = self.createExcelPath('签证台账表.xlsx', str(t2), '签证台账', 30, 'A1:AE1', *file_ls)
        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            index = 1
            kwargs = {
                'status': True
            }

            searchName = self.request.GET.get('searchName', None)
            if "jobRankid" in self.request.GET:
                jobRankId = self.request.GET.get('jobRankid', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['expatriate_jobRank'] = jobRankId
            if "jobRankid[]" in self.request.GET:
                jobRankId = self.request.GET.getlist('jobRankid[]', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['expatriate_jobRank__in'] = jobRankId
            visaledger_info__visa_Type = self.request.GET.get('visa_Type', None)  # 签证类型
            if len(visaledger_info__visa_Type) != 0:
                kwargs['visaledger_info__visa_Type'] = visaledger_info__visa_Type
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            if beginDate != "" and endDate != "":
                kwargs['visaledger_info__signed__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                                  177) if beginDate is None or len(
                    endDate) == 0 else beginDate
                kwargs['visaledger_info__signed__lte'] = datetime(3221, 10, 29, 7, 17, 1,
                                                                  177) if endDate is None or len(
                    endDate) == 0 else endDate  # 出签日期
            # print(kwargs)
            tableList = ExpatriateInfoList.objects.filter(Q(name__contains=searchName) | Q(code__contains=searchName)| Q(passport__contains=searchName),**kwargs).values_list(
                                                                    'code',
                                                                    'name',
                                                                    'expatriate_jobRank',
                                                                    'expatriate_Before_Manage',#基地归属(外派前管理归属)
                                                                    'abbreviation_Dept',  # 部门简称
                                                                    'post',  # 岗位名称
                                                                    'passport',  # 护照号
                                                                    "visaledger_info__first_Docking",
                                                                    "visaledger_info__application_Approval",
                                                                    "visaledger_info__invitation_Letter_Provided",
                                                                    "visaledger_info__collectInfo",
                                                                    "visaledger_info__submit_Embassy",
                                                                    "visaledger_info__last_SupplementaryInfo",
                                                                    "visaledger_info__signed",
                                                                    "visaledger_info__arrival_Thailand",
                                                                    "visaledger_info__visa_Application_Remarks",
                                                                    "visaledger_info__current_Progress",
                                                                    "visaledger_info__sending_Embassy",
                                                                    "visaledger_info__system_Post",
                                                                    "visaledger_info__visa_Type",
                                                                    "visaledger_info__visa_Validity_Period_Begin",
                                                                    "visaledger_info__visa_Validity_Period_End",
                                                                    "visaledger_info__backSign",
                                                                    "visaledger_info__visa_Expiration_Date",
                                                                    "visaledger_info__arrival_Thailand_First",
                                                                    "visaledger_info__report_90",
                                                                    "visaledger_info__filing_Dead_90",
                                                                    "visaledger_info__expiration_Article",
                                                                    "visaledger_info__isAssignOver",
                                                                    'visaledger_info__visaLedger_remark'
                                                                    ).order_by('-expatriate_createTime')

            for line in tableList:
                line = list(line)
                line.insert(0, index)
                if line[-2] == True:
                    line[-2] = '是'
                elif line[-2] == False:
                    line[-2] = '否'
                row_data.append(line)
                if len(line) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
            pass
        else:
            row_data = []
            index = 1
            for id in id_list:
                data=list( ExpatriateInfoList.objects.filter(pk=id, status=True).values_list('code',
                                                                                                  'name',
                                                                                                  'expatriate_jobRank',
                                                                                                  'expatriate_Before_Manage',
                                                                                                  'abbreviation_Dept',
                                                                                                  'post',  # 岗位名称
                                                                                                  'passport',  # 护照号
                                                                                                  "visaledger_info__first_Docking",
                                                                                                  "visaledger_info__application_Approval",
                                                                                                  "visaledger_info__invitation_Letter_Provided",
                                                                                                  "visaledger_info__collectInfo",
                                                                                                  "visaledger_info__submit_Embassy",
                                                                                                  "visaledger_info__last_SupplementaryInfo",
                                                                                                  "visaledger_info__signed",
                                                                                                  "visaledger_info__arrival_Thailand",
                                                                                                  "visaledger_info__visa_Application_Remarks",
                                                                                                  "visaledger_info__current_Progress",
                                                                                                  "visaledger_info__sending_Embassy",
                                                                                                  "visaledger_info__system_Post",
                                                                                                  "visaledger_info__visa_Type",
                                                                                                  "visaledger_info__visa_Validity_Period_Begin",
                                                                                                  "visaledger_info__visa_Validity_Period_End",
                                                                                                  "visaledger_info__backSign",
                                                                                                  "visaledger_info__visa_Expiration_Date",
                                                                                                  "visaledger_info__arrival_Thailand_First",
                                                                                                  "visaledger_info__report_90",
                                                                                                  "visaledger_info__filing_Dead_90",
                                                                                                  "visaledger_info__expiration_Article",
                                                                                                  "visaledger_info__isAssignOver",
                                                                                                  'visaledger_info__visaLedger_remark',
                                                                                                  ))[0]
                # print(data)
                data = (index,) + data
                data=list(data)
                if data[-2] == True:
                    data[-2] = '是'
                elif data[-2] == False:
                    data[-2] = '否'
                row_data.append(data)
                if len(data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    def resetPostTemplate(self):
        try:
            file = self.request.FILES.get("file", None)
            now = arrow.now()
            t1 = now.format('YYYY-MM-DD')
            t2 = now.format('YYYY-MM-DD_HH_mm_ss')
            dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'upload_file', t1,
                                      '签证模板文件上传')  # 创建文件夹
            self.mkdir(dummy_path)
            file_url, file_name, file_suffix = self.createPath(file, '签证模板文件上传', '签证台账' + str(t2))
            self.saveFile(file_url, file)
            exc = openpyxl.load_workbook(file_url, data_only=True)
            sheet = exc.active
            for i in range(2, sheet.max_row):  # 每行数据
                try:
                    code = sheet.cell(i + 1, 2).value  # 工号
                    name = sheet.cell(i + 1, 3).value  #姓名
                    expatriate_jobRank = sheet.cell(i + 1, 4).value   #合同归属
                    expatriate_Before_Manage = sheet.cell(i + 1, 5).value  # 基地归属
                    abbreviation_Dept = sheet.cell(i + 1, 6).value  # 部门简称
                    post = sheet.cell(i + 1, 7).value  # 岗位名称
                    passport = sheet.cell(i + 1, 8).value  # 护照号
                    first_Docking = sheet.cell(i + 1, 9).value  # 首次对接日期
                    application_Approval = sheet.cell(i + 1, 10).value  # BOI申请批文日期
                    invitation_Letter_Provided = sheet.cell(i + 1, 11).value  # 批文、邀请函提供日期
                    collectInfo = sheet.cell(i + 1, 12).value  # 收齐资料日期
                    submit_Embassy = sheet.cell(i + 1, 13).value  # 提交使馆日期
                    last_SupplementaryInfo = sheet.cell(i + 1, 14).value  # 末次补资料日期
                    signed = sheet.cell(i + 1, 15).value  # 出签日期
                    arrival_Thailand = sheet.cell(i + 1, 16).value  # 抵派驻地日期
                    visa_Application_Remarks = sheet.cell(i + 1, 17).value  # 签证申请备注
                    current_Progress = sheet.cell(i + 1, 18).value  # 当前进度
                    sending_Embassy = sheet.cell(i + 1, 19).value  # 送签使馆
                    system_Post = sheet.cell(i + 1, 20).value  # 系统岗位名称
                    visa_Type = sheet.cell(i + 1, 21).value  # 签证类型
                    visa_Validity_Period_Begin = sheet.cell(i + 1,22).value  # 签证有效期开始时间
                    visa_Validity_Period_End = sheet.cell(i + 1,23).value  # 签证有效期结束时间
                    backSign = sheet.cell(i + 1, 24).value  # 回头签
                    visa_Expiration_Date = sheet.cell(i + 1, 25).value  #距签证过期日剩余
                    arrival_Thailand_First = sheet.cell(i + 1, 26).value  # 首次抵泰日期
                    report_90 = sheet.cell(i + 1, 27).value  # 90天报道申报日
                    filing_Dead_90 = sheet.cell(i + 1, 28).value  # 90天申报截止日
                    expiration_Article = sheet.cell(i + 1, 29).value  # 距报道过期日剩余
                    isAssignOver = sheet.cell(i + 1, 30).value  #外派是否结束_签证是否注销
                    visaLedger_remark = sheet.cell(i + 1, 31).value  # 备注
                    if passport is not None and name is not None:
                        expatriate_obj= ExpatriateInfoList.objects.filter(name=name, status=True,passport__isnull=True).order_by('-expatriate_createTime').all()  # 升序  从小到大 从近到远
                        expat_info = {'code': code, 'name': name, 'expatriate_jobRank': expatriate_jobRank,
                                      'expatriate_Before_Manage':expatriate_Before_Manage,"abbreviation_Dept":abbreviation_Dept,"post":post,"passport":passport}
                        visa_info = {
                            'first_Docking': first_Docking,
                            'application_Approval': application_Approval,
                            'invitation_Letter_Provided': invitation_Letter_Provided,
                            "collectInfo": collectInfo,
                            "submit_Embassy": submit_Embassy,
                            "last_SupplementaryInfo": last_SupplementaryInfo,
                            "signed": signed,
                            'arrival_Thailand': arrival_Thailand,
                            "visa_Application_Remarks": visa_Application_Remarks,
                            "current_Progress": current_Progress,
                            "sending_Embassy": sending_Embassy,
                            "system_Post": system_Post,
                            "visa_Type":visa_Type,
                            "visa_Validity_Period_Begin": visa_Validity_Period_Begin,
                            'visa_Validity_Period_End': visa_Validity_Period_End,
                            "backSign": backSign,
                            "visa_Expiration_Date": visa_Expiration_Date,
                            "arrival_Thailand_First": arrival_Thailand_First,
                            "report_90": report_90,
                            "filing_Dead_90": filing_Dead_90,
                            "expiration_Article":expiration_Article,
                            "isAssignOver": 1 if isAssignOver=='是' else 0,
                            "visaLedger_remark":visaLedger_remark
                        }
                        #     # print(ticket_name,ticket_passport,flight_date)
                        if expatriate_obj.exists():  # 在外派信息表至少有一条数据   签证创建后关联信息表
                            print('关联',expatriate_obj[0].id,expatriate_obj[0].name)
                            visa_info['visaLedger_creator_id'] = self.request.check_token
                            expat_info['expatriate_modifier_id'] = self.request.check_token
                            visa_obj = VisaLedgerInfoList.objects.create(**visa_info)
                            visa_obj.visaledger_info.add(expatriate_obj[0].id)
                            if len(str(expat_info['code']))==0 or expat_info['code']==None:
                                # print("删除")
                                del expat_info['code']
                            ExpatriateInfoList.objects.filter(pk=expatriate_obj[0].id).update(**expat_info)
                        else:   #签证创建后在创建信息后在关联
                            print("不关联",name)
                            expat_info['expatriate_creator_id'] = self.request.check_token
                            visa_info['visaLedger_creator_id'] = self.request.check_token
                            expat_obj = ExpatriateInfoList.objects.create(**expat_info)
                            visa_obj = VisaLedgerInfoList.objects.create(**visa_info)
                            visa_obj.visaledger_info.add(expat_obj.id)
                        # print(visa_info)
                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "数据新增成功",
                        }
                except:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "该行数据异常",
                    }
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功",
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "上传异常",
            }



    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['AD'].width = 25
        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'expatriateRecordFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        # t = time.strftime('%Y-%m-%d')
        now = arrow.now()
        t = now.format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  #文件后缀

        file_name = f"{fileName}.{file_suffix}"    #文件名称

        file_path = os.path.join('static', 'expatriateRecordFile', 'upload_file', t,path,file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)
