from expatriateRecord.models import *
from rest_framework.response import Response
from rest_framework import status
from expatriateRecord.sql import *
from datetime import datetime, date
from django.db.models import Q

from pdss.settings import BASE_DIR


class ResetTicketLedger:  # 机票台账
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "resetGet": self.resetGet,
            "resetPost": self.resetPost,
            'resetDelete': self.resetDelete,
            'resetPatch': self.resetPatch,
            'resetPut': self.resetPut,
            'resetPostTemplate':self.resetPostTemplate,   #上传模板文件数据
        }

    def meth_center(self):
        self.return_data = {'code': HTTP_403_FORBIDDEN, "msg": '没有权限访问!'}
        if self.request.check_token is None:
            return Response(self.return_data)
        self.methods[self.meth]()
        return Response(self.return_data)
    def resetGet(self):
        columnList = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'ticket_time', 'value': '订票时间', 'width': 120},
            {'label': 'ticket_code', 'value': '工号', 'width': 100},
            {'label': 'ticket_name', 'value': '姓名', 'width': 90},
            # {'label': 'people__expatriate_Dept', 'value': '部门名称', 'width':250},
            {'label': 'ticket_passport', 'value': '护照号', 'width': 90},
            {'label': 'is_assignment', 'value': '是否去往派驻地', 'width':130},

            {'label': 'flight_date', 'value': '飞行日期', 'width': 110},
            {'label': 'departure', 'value': '始发地', 'width': 100},
            {'label': 'departure_time', 'value': '起飞时间', 'width':160},
            {'label': 'destination', 'value': '目的地', 'width': 100},
            {'label': 'arrival_time', 'value': '到达时间', 'width': 160},
            {'label': 'price', 'value': '价格', 'width': 90},
            {'label': 'remark', 'value': '备注', 'width':500},
            {'label': 'ticketLedger_creator__username', 'value': '创建人工号', 'width': 100},
            {'label': 'ticketLedger_creator__user', 'value': '创建人姓名', 'width': 100},
            {'label': 'ticketLedger_modifier__username', 'value': '修改人工号', 'width': 100},
            {'label': 'ticketLedger_modifier__user', 'value': '修改人姓名', 'width': 100},
            {'label': 'ticketLedger_createTime', 'value': '创建时间', 'width': 100},
            {'label': 'ticketLedger_modifyTime', 'value': '修改时间', 'width': 100},
        ]
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'status': True
        }
        searchName = self.request.GET.get('searchName',None)
        beginDate = self.request.GET.get('beginDate',None)
        endDate = self.request.GET.get('endDate',None)
        expatriate_Dept= self.request.GET.get('expatriate_Dept',None)
        is_assignment = self.request.GET.get('is_assignment', None)  # 是否在职
        if is_assignment != '':
            kwargs['is_assignment'] = is_assignment

        # if len(expatriate_Dept) != 0:
        #     kwargs['people__expatriate_Dept'] = expatriate_Dept
        if beginDate != "" and endDate != "":
            kwargs['flight_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(endDate) == 0 else beginDate
            kwargs['flight_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate  #飞行日期

        totalNumber = TicketLedgerInfoList.objects.filter(Q(ticket_name__contains=searchName) | Q( ticket_code__contains = searchName)| Q(ticket_passport__contains = searchName),**kwargs).count()
        tableList = TicketLedgerInfoList.objects.filter(Q(ticket_name__contains=searchName) | Q(ticket_code__contains = searchName)| Q(ticket_passport__contains = searchName),**kwargs).values('id',
                                                               'ticket_code',
                                                               'ticket_name',
                                                               'ticket_passport',  # 护照号
                                                                'is_assignment',
                                                                "ticket_time",
                                                                "flight_date",
                                                                "departure",
                                                                "departure_time",
                                                                "destination",
                                                                "arrival_time",
                                                                "price",
                                                               'remark',
                                                               'ticketLedger_creator__username',
                                                               'ticketLedger_creator__user',
                                                                 'ticketLedger_modifier__username',
                                                                 'ticketLedger_modifier__user',
                                                                 'ticketLedger_createTime',
                                                                 'ticketLedger_modifyTime',
                                                                'people_id'
                                                                 ).order_by('-ticketLedger_createTime')[
                    (currentPage - 1) * pageSize:currentPage * pageSize]
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['ticket_time']=item['ticket_time'] if item['ticket_time'] is None else str(item['ticket_time'])[:10]
            item['departure_time']= item['departure_time'] if item['departure_time'] is None else str(item['departure_time'])
            item['arrival_time'] = item['arrival_time'] if item['arrival_time'] is None else str(item['arrival_time'])
            item['ticketLedger_createTime'] = item['ticketLedger_createTime'] if item['ticketLedger_createTime'] is None else str(item['ticketLedger_createTime'])[:10]
            item['ticketLedger_modifyTime'] = item['ticketLedger_modifyTime'] if item['ticketLedger_modifyTime'] is None else str(item['ticketLedger_createTime'])[:10]

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def resetDelete(self):
        try:
            id_list = json.loads(self.request.body).get('idList')
            delete_all = json.loads(self.request.body).get('deleteAll')
            if delete_all: #删除全部
                TicketLedgerInfoList.objects.update(status=False)
            else:
                for id in id_list:
                    TicketLedgerInfoList.objects.filter(pk=id).update(status=False)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "删除成功！"
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def resetPost(self):
        info=self.request.POST.get('createData',None)
        info = json.loads(info)
        # print(info)
        expat_info = {
            'code': info['ticket_code'], 'name': info['ticket_name'], 'passport': info['ticket_passport'],
        }
        expatriate_obj_all=ExpatriateInfoList.objects.filter(passport=expat_info['passport'],name=expat_info['name'],status=True,expatriate_Begin__lte=info['flight_date'], expatriate_End__gte=info['flight_date']).order_by('-expatriate_createTime').all()
        try:
            if expatriate_obj_all.exists():#在外派信息表至少有一条数据
                expatriate_obj=expatriate_obj_all[0]
                expat_info['expatriate_modifier_id'] = self.request.check_token
                if len(str(expat_info['code'])) == 0 or expat_info['code'] == None:
                    del expat_info['code']
                ExpatriateInfoList.objects.filter(pk=expatriate_obj.id).update(**expat_info)
                TicketLedgerInfoList.objects.update_or_create(defaults=info,ticket_time=info['ticket_time'],departure=info['departure'],flight_date=info['flight_date'],status=True,people_id=expatriate_obj.id)
            else:
                info['ticketLedger_creator_id']=self.request.check_token
                TicketLedgerInfoList.objects.create(**info)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "数据新增成功",
            }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "数据新增异常",
            }



    def resetPatch(self):
        try:
            info = self.request.data
            info['ticketLedger_modifier'] = self.request.check_token

            TicketLedgerInfoList.objects.filter(pk=info['id']).update(**info)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "修改成功",
            }
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def resetPut(self):  #下载
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        file_ls=['序号','订票时间','工号','姓名','护照号','是否去往派驻地','飞行日期','始发地','起飞时间','目的地','到达时间','价格','备注',"创建人工号","创建人姓名","修改人工号","修改人姓名","创建时间","修改时间"]


        path = self.createExcelPath('机票台账表.xlsx', str(t2), '机票台账', 30, 'A1:S1',*file_ls)
        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            index = 1
            kwargs = {
                'status': True
            }
            searchName = self.request.GET.get('searchName', None)
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            is_assignment = self.request.GET.get('is_assignment', None)  # 是否在职
            if is_assignment != '':
                kwargs['is_assignment'] = is_assignment
            if beginDate != "" and endDate != "":
                kwargs['flight_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                    endDate) == 0 else beginDate
                kwargs['flight_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                    endDate) == 0 else endDate  # 飞行日期
            # print(kwargs)
            tableList = TicketLedgerInfoList.objects.filter(
                Q(ticket_name__contains=searchName) | Q(ticket_code__contains=searchName) | Q(
                    ticket_passport__contains=searchName), **kwargs).values_list(
                                                                            "ticket_time",
                                                                             'ticket_code',
                                                                             'ticket_name',
                                                                             'ticket_passport',  # 护照号
                                                                             'is_assignment',
                                                                             "flight_date",
                                                                             "departure",
                                                                             "departure_time",
                                                                             "destination",
                                                                             "arrival_time",
                                                                             "price",
                                                                             'remark',
                                                                             'ticketLedger_creator__username',
                                                                             'ticketLedger_creator__user',
                                                                             'ticketLedger_modifier__username',
                                                                             'ticketLedger_modifier__user',
                                                                             'ticketLedger_createTime',
                                                                             'ticketLedger_modifyTime'
                                                                             ).order_by('-ticketLedger_createTime')

            for line in tableList:
                line = list(line)
                line.insert(0, index)
                # print(line)
                row_data.append(line)
                if len(line) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            index = 1
            for id in id_list:
                data = list(TicketLedgerInfoList.objects.filter(pk=id,status=True).values_list(
                    # 'id',
                                                               'ticket_code',
                                                               'ticket_name',
                                                               # 'people__expatriate_Dept',  # 部门名称
                                                               'ticket_passport',  # 护照号
                                                                'is_assignment',
                                                                "ticket_time",
                                                                "flight_date",
                                                                "departure",
                                                                "departure_time",
                                                                "destination",
                                                                "arrival_time",
                                                                "price",
                                                               'remark',
                                                               'ticketLedger_creator__username',
                                                               'ticketLedger_creator__user',
                                                                 'ticketLedger_modifier__username',
                                                                 'ticketLedger_modifier__user',
                                                                 'ticketLedger_createTime',
                                                                 'ticketLedger_modifyTime'
                                                                 ))[0]
                data = (index,) + data
                row_data.append(data)
                if len(data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    def resetPostTemplate(self):
        try:
            file = self.request.FILES.get("file", None)
            now = arrow.now()
            t1 = now.format('YYYY-MM-DD')
            t2 = now.format('YYYY-MM-DD_HH_mm_ss')
            dummy_path = os.path.join(BASE_DIR, 'static', 'expatriateRecordFile', 'upload_file', t1,'机票模板文件上传') # 创建文件夹
            self.mkdir(dummy_path)
            # print(dummy_path)
            file_url, file_name, file_suffix = self.createPath(file,'机票模板文件上传', '机票台账'+str(t2))
            self.saveFile(file_url, file)
            exc = openpyxl.load_workbook(file_url, data_only=True)
            sheet = exc.active
            for i in range(2, sheet.max_row):  #每行数据
                try:
                    ticket_time = sheet.cell(i + 1, 2).value#订票时间
                    ticket_code =sheet.cell(i + 1, 3).value  #
                    ticket_name =sheet.cell(i + 1, 4).value
                    ticket_passport = sheet.cell(i + 1, 5).value#护照号
                    is_assignment = sheet.cell(i + 1,6).value  #是否去往派驻地
                    flight_date =sheet.cell(i + 1, 7).value  #飞行日期
                    departure = sheet.cell(i + 1,8).value  #始发地
                    departure_time =sheet.cell(i + 1, 9).value  #起飞时间
                    destination =sheet.cell(i + 1, 10).value  #目的地
                    arrival_time = sheet.cell(i + 1,11).value  #到达时间
                    price =sheet.cell(i + 1, 12).value  #价格
                    remark =sheet.cell(i + 1, 13).value  #备注
                    # print(type(departure_time),type(arrival_time))


                    try:
                        if type(departure_time)==str:
                            departure_time = datetime.strptime(departure_time, "%Y/%m/%d %H:%M:%S")
                        if type(arrival_time)==str:
                            arrival_time = datetime.strptime(arrival_time, "%Y/%m/%d %H:%M:%S")
                    except:
                        pass
                    # print(type(departure_time), type(arrival_time))
                    if ticket_name is not None and flight_date is not None:
                        expatriate_obj_all = ExpatriateInfoList.objects.filter(name=ticket_name, status=True,passport=ticket_passport,
                                                                               expatriate_Begin__lte=flight_date,
                                                                               expatriate_End__gte=flight_date).order_by('-expatriate_createTime').all()   #升序  从小到大 从近到远

                        expat_info = {'code': ticket_code, 'name': ticket_name, 'passport':ticket_passport}
                        info={
                                'ticket_code':ticket_code,
                                'ticket_name': ticket_name,
                                'ticket_passport': ticket_passport,
                                "ticket_time": ticket_time,
                                "flight_date": flight_date,
                                "departure": departure,
                                "departure_time": departure_time,
                                'is_assignment':is_assignment,
                                "destination": destination,
                                "arrival_time": arrival_time,
                                "price": price,
                                "remark": remark
                            }
                        if expatriate_obj_all.exists():  # 在外派信息表至少有一条数据
                            # print('关联')
                            expatriate_obj = expatriate_obj_all[0]
                            if len(str(expat_info['code']))==0 or expat_info['code']==None:
                                del expat_info['code']
                            ExpatriateInfoList.objects.filter(pk=expatriate_obj.id).update(**expat_info)
                            TicketLedgerInfoList.objects.update_or_create(defaults=info, ticket_time=info['ticket_time'],departure=info['departure'],
                                                                          flight_date=info['flight_date'], status=True,
                                                                          people_id=expatriate_obj.id)   #有数据关联
                        else:
                            # print("不关联")
                            info['ticketLedger_creator_id'] = self.request.check_token
                            TicketLedgerInfoList.objects.update_or_create(defaults=info,ticket_time=info['ticket_time'],departure=info['departure'],
                                                                          flight_date=info['flight_date'], status=True)
                            # TicketLedgerInfoList.objects.create(**info)

                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "数据新增成功",
                        }
                # except:
                except Exception as e:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": e,
                    }
            # self.return_data = {
            #     "code": status.HTTP_200_OK,
            #     "msg": "上传成功",
            # }
        except:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "上传异常",
            }


    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 20
        sheet.column_dimensions['A'].width = 10
        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'expatriateRecordFile', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
        # t = time.strftime('%Y-%m-%d')
        now = arrow.now()
        t = now.format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  #文件后缀

        file_name = f"{fileName}.{file_suffix}"    #文件名称

        file_path = os.path.join('static', 'expatriateRecordFile', 'upload_file', t,path,file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    @staticmethod
    def saveFile(file_path,file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path),'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)

    # @staticmethod
    # def createPath(pic,path,fileName):  # 生成路径     文件对象  文件上一级目录名称 文件名称
    #     t = time.strftime('%Y-%m-%d')
    #     file_suffix = str(pic).split(".")[-1]  #文件后缀
    #
    #     file_name = f"{fileName}.{file_suffix}"    #文件名称
    #
    #     file_path = os.path.join('static', 'competeRestrictionsFile', 'upload_file', t,path,file_name)  # 文件路径
    #     file_path = file_path.replace('\\', '/')
    #     return (file_path,file_name,file_suffix)  # 文件路径   文件名字  文件后缀

    # def saveFile(self,file_path,file_obj):  # 文件名,图像对象   文件保存
    #     with open(str(file_path),'wb+') as f:
    #         for dot in file_obj.chunks():
    #             f.write(dot)

