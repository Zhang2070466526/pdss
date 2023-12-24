# -*- coding: utf-8 -*-
# @Time    : 2023/5/4 11:25
# @Author  : zhuang
# @Site    : 
# @File    : honorClass.py
# @Software: PyCharm
import os
import datetime
import json
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from salarySurvey.models import SalarySurveyRecord
from auther.models import *
from ..serializers import *
from ..models import *
from general.models import *
from django.core.exceptions import ObjectDoesNotExist

from controller.controller import Controller, upload_file
import time
from utils.check_token import CheckToken


def verify(request):
    return_data = {'code': '', "message": ''}
    new_token = CheckToken()
    try:
        check_token = new_token.check_token(request.headers['Authorization'])
    except Exception as e:
        # print(e)
        return_data['code'] = 400
        return_data['message'] = '请求参数出错啦'
        return return_data
    if check_token is None:
        return_data['code'] = 403
        return_data['message'] = '没有权限查询'
        return return_data


class honorClass:
    def __init__(self, request, meth):
        self.user_base = None
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.check_token = ""
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }

    def meth_center(self):
        if verify(self.request) is not None:
            return Response(verify(self.request))
        self.check_token = self.request.check_token
        self.user_base = self.request.user_base
        self.methods[self.meth]()
        return Response(self.return_data)

    def get_upload(self):
        file = self.request.FILES.get("file", None)
        createFile = self.request.FILES.getlist("createFile", None)
        createPhoto = self.request.FILES.getlist("createPhoto", None)
        createData = self.request.POST.get("createData", None)
        searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
        if file and file is not None:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                self.fileName = upload_file(file, "externalFile", "upload", None)
                if self.fileName:
                    self.save_Data()
            except ObjectDoesNotExist as e:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "中心/基地名不存在"
                }
            except Exception as e:
                # print(e)
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            createData = eval(createData)
            obj = ExternalHonorsList.objects.update_or_create(**createData)[0]

            for file in createFile:
                fileName = obj.honor_name + "_" + searchTime + "_" + str(
                    "".join(list(str(time.time()))[0:10])) + "_" + file.name
                fileName = upload_file(file, "externalFile", "other_file", fileName)
                file_url = "static/" + "externalFile" + "/other_file/" + searchTime + "/" + fileName
                file_kwargs = {
                    "file_url": file_url,
                    "file_name": fileName,
                }
                file_obj = UploadFiles.objects.update_or_create(**file_kwargs)[0]
                file_obj.honor_upload_declare_files.add(obj.id)

            for file in createPhoto:
                fileName = obj.honor_name + "_" + searchTime + "_" + str(
                    "".join(list(str(time.time()))[0:10])) + "_" + file.name
                fileName = upload_file(file, "externalFile", "other_file", fileName)
                file_url = "static/" + "externalFile" + "/other_file/" + searchTime + "/" + fileName
                file_kwargs = {
                    "file_url": file_url,
                    "file_name": fileName,
                }
                file_obj = UploadFiles.objects.update_or_create(**file_kwargs)[0]
                file_obj.honor_medal_photos.add(obj.id)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "用户数据新增成功"
            }

    def get_list(self):

        obj = Controller(ExternalHonorsList, "get_list", self.request)
        self.return_data = obj.data_start()

    # 将上传的文件数据保存到数据库中
    def save_Data(self):
        workbook = openpyxl.load_workbook(self.fileName)
        table = workbook.active
        for i in range(2, table.max_row):
            if table.cell(i + 1, 2).value != "" and table.cell(i + 1, 2).value is not None:
                try:
                    employee_base_father = center_base.objects.get(status=1,name=table.cell(i + 1, 1).value.replace("（","(").replace("）", ")"))
                except ObjectDoesNotExist:
                    self.return_data = {
                        "code": status.HTTP_400_BAD_REQUEST,
                        "msg": "基地名不存在，请重新选择"
                    }
                    return
                obj = center_base.objects.filter(base_parent_id=employee_base_father.id, status=1)
                if obj.exists():
                    employee_base_son = center_base.objects.get(status=1, name=table.cell(i + 1, 2).value.replace("（",
                                                                                                                  "(").replace(
                        "）", ")"))
                else:
                    employee_base_son = employee_base_father
                if employee_base_father.id not in self.user_base or employee_base_son.id not in self.user_base:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "抱歉，您没有上传 " + table.cell(i + 1, 1).value + "-" + table.cell(i + 1,
                                                                                           2).value + " (基地/中心/公司)的权限"
                    }
                else:
                    data_kwargs = {
                        "honor_base": employee_base_son,
                        "honor_name": table.cell(i + 1, 4).value,
                        "honor_level": table.cell(i + 1, 5).value,
                        "honor_issue_organize": table.cell(i + 1, 6).value,
                        "honor_date":  datetime.datetime.strptime(table.cell(i + 1, 7).value, "%Y-%m-%d"),
                        "honor_remark": table.cell(i + 1, 8).value,
                        "honor_status": 1,
                        "creator": AdminUser.objects.filter(pk=self.check_token)[0],
                        "modifier": AdminUser.objects.filter(pk=self.check_token)[0],
                    }
                    ExternalHonorsList.objects.update_or_create(**data_kwargs)

    # 下载文件
    def download_file(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功"
        }
        info = json.loads(self.request.body)
        get_obj = self.request.GET
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25

        kwargs = {
            'honor_status': 1,
            "honor_base__in": self.request.user_base

        }
        # print('kwargs:',kwargs)
        kwargs['honor_base__in'] = self.request.user_base
        searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
        filename = "外部荣誉清单"
        id_list = info['idList']
        is_all = info['downloadAll']
        fields = {}
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "externalFile"), "download")):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "externalFile"), "download"))
        path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "externalFile"), "download"),
                            searchTime)
        if not os.path.exists(path):
            os.mkdir(path)
        if is_all == 0 and len(id_list) <= 0:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "条件为空"
            }
            return

        if id_list is not None and len(id_list) > 0:
            kwargs["id__in"] = list(id_list)

        if is_all:
            kwargs = {
                'honor_status': 1,
                "honor_base__in": self.request.user_base
            }
            model_fields = {
                "searchName": "honor_name__contains",
                "beginDate": "honor_date__gte",
                "endDate": "honor_date__lte",
                "baseNameId": "honor_base_id",
                "user_base": "honor_base__in",
            }
            model_key = ["searchName", "beginDate", "endDate", "baseNameId"]
            for key, value in get_obj.items():
                if value != "" and key in model_key:
                    kwargs[model_fields[key]] = value
        # print(kwargs)
        obj = ExternalHonorsList.objects.filter(**kwargs).order_by("create_time")
        if len(obj) > 0:

            obj = HonorRecordPutSerializers(instance=obj, many=True).data
            for field in ExternalHonorsList._meta.get_fields():
                if field.name not in ["honor_status", "creator", "modifier", "modify_time", "create_time"]:
                    fields[field.verbose_name] = field.name

            letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            width = 3.0
            side = Side(style="thick", color="000000")
            font_header = Font(name="等线", size=14, bold=True)

            wb = Workbook()
            ws = wb.active
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
            font_header = Font(name="等线", size=14, bold=True)
            ws['A1'] = "外部荣誉"
            ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
            ws['A1'].font = font_header

            row2 = [key for key, value in fields.items()]
            row2[0] = "序号"
            row2.insert(2, "中心/事业部")
            fields["中心/事业部"] = "honor_base_father"
            ws.append(row2)
            for i in letter_list:
                ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
                ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
            index = 1
            for data in obj:
                row_data = []
                for k in row2:
                    if k != "序号":
                        row_data.append(dict(data)[fields[k]])
                    else:
                        row_data.append(index)
                index += 1
                ws.append(row_data)

            excel_columns = ws.columns
            len_list = self.count_excel_save(excel_columns)

            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
            wb.close()
            filename = filename + "".join(list(str(time.time()))[0:10])
            wb.save("static/externalFile/download/" + searchTime + '/' + filename + ".xlsx")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '下载成功',
                "downloadUrl": "static/externalFile/download/" + searchTime + '/' + filename + ".xlsx",
            }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '无该用户数据'
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(ExternalHonorsList, "patch", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def delete_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(ExternalHonorsList, "delete", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def count_character(self, s):
        hanzi = 0
        num = 0
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1

        return str(30 * hanzi + 20 * num)

    def count_excel_save(self, excel_columns):

        len_list = []
        prjTuple = tuple(excel_columns)
        for idx in range(0, len(prjTuple)):
            length2 = 0
            str_list = []
            for cell in prjTuple[idx]:
                if "1" not in str(cell):
                    str_list.append(str(cell.value))
            for elem in str_list:
                elem_split = list(elem)
                length = 0
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1.5
                    else:
                        length += 2.5
                length2 = max(length, length2)
            len_list.append(length2)
        return len_list
