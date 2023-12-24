import json
import os

from django.shortcuts import render

# Create your views here.
from rest_framework.viewsets import GenericViewSet

from pdss.settings import BASE_DIR
from rayongAttendance.attendance.raClass import *
import openpyxl
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect


# 获取数据列表
def attendance_data(request):
    res = get_person_data(request)
    return HttpResponse(json.dumps(res))


# 获取数据列表
def generate_data_method(request):
    res = generate_data(request)
    return HttpResponse(json.dumps(res))


# 下载
def get_rayong_department_data(request):
    from django.db.models import Q
    from django.db.models.functions import Lower
    from django.db.models.functions import TruncDate
    from django.db.models import Count
    import os
    new_token = CheckToken()
    check_token = new_token.check_token(request.headers['Authorization'])
    # print(check_token)
    if check_token is not None:
        isDp=json.loads(request.body).get('isDp')
        t = arrow.now()
        t1 = t.format('YYYY-MM-DD')
        t2 = t.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'rayongAttendance', 'download_file', t1, str(t2))  # 创建文件夹
        mkdir(dummy_path)
        print(isDp,type(isDp))
        if isDp==True:    #部门下载
            print('部门下载')
            start_time = request.GET.get('startTime', None)
            end_time = request.GET.get('endTime', None)
            if len(str(start_time)) == 0 or start_time is None and len(str(end_time)) == 0 or end_time is None:
                end_time = arrow.now().format('YYYY-MM-DD 23:59:59')  # 07-07
                start_time = arrow.now().shift(weeks=-1, days=-1).format('YYYY-MM-DD 00:00:00')  # 06-29
            file_ls=['ID', '部门', '日期', '次数']
            path = createExcelPath('罗勇部门考勤表.xlsx', str(t2),'罗勇部门考勤表',4,'A1:D1',*file_ls)
            # print(end_time,start_time)
            # result = AttendanceDetailRecord.objects.filter(
            #     dept_name__isnull=False,  # Exclude records with empty dept_name
            #     person__event_time__range=(start_time, end_time)
            # ).values('dept_name','day').annotate(count=Count('id')).annotate(day=TruncDate('person__event_time'))
            # print(result)

            records = AttendanceDetailRecord.objects.filter(dept_name__isnull=False,
                                                            person__event_time__range=(start_time, end_time)) \
                .annotate(day=TruncDate('person__event_time')).values_list('dept_name', 'day').annotate(
                count=Count('id')).order_by('dept_name', 'day')     #annotate分组    annotate 聚合函数
            index = 1
            row_data = []
            for line in records:
                line = list(line)
                line.insert(0, index)
                row_data.append(line)
                if len(line) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:  #详细数据下载

            id_list = json.loads(request.body).get('idList')
            downloadAll = json.loads(request.body).get('downloadAll')
            file_ls=['ID','姓名', '部门','工号','出门时间','班制']
            path = createExcelPath('罗勇部门个人出门时间表.xlsx', str(t2),'罗勇部门个人出门时间表',6,'A1:F1',*file_ls)

            if downloadAll == True:  # 是下载全部   有条件
                row_data = []
                index = 1
                kwargs = {
                    "status": 1
                }
                args = ()
                searchName = request.GET.get("searchName", None)
                startTime = request.GET.get("startTime", None)
                endTime = request.GET.get("endTime", None)
                # currentPage = request.GET.get("currentPage", None)
                # pageSize = request.GET.get("pageSize", None)
                deptName = request.GET.get("deptName", None)
                dept_list = AttendanceDetailRecord.objects.filter(status=1).values("dept_name").distinct()
                dept_list = [{"label": i['dept_name'], "value": i['dept_name']} for i in dept_list]
                # if currentPage == '' or currentPage is None:
                #     currentPage = 1
                # else:
                #     currentPage = int(currentPage)
                # if pageSize == '' or pageSize is None:
                #     pageSize = 25
                # else:
                #     pageSize = int(pageSize)
                if searchName is not None and searchName != "":
                    args += Q(person__name__contains=searchName) | Q(person__pin__contains=searchName),
                if startTime != "" and startTime is not None:
                    kwargs["event_time__gte"] = startTime + " 00:00:00"
                if endTime != "" and endTime is not None:
                    kwargs["event_time__lte"] = endTime + " 00:00:00"
                if deptName is not None and deptName != '':
                    kwargs['person__dept_name'] = deptName
                obj = AttendanceTimeRecord.objects.filter(*args, **kwargs).values_list("person__name", "person__dept_name", "person__pin", "event_time", "person__job_time")
                for line in obj:
                    line = list(line)
                    line.insert(0,index)
                    row_data.append(line)
                    if len(line) == 0:
                        index = index
                    index += 1

                exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                for row in row_data:
                    sheet.append(row)  # 在工作表中添加一行
                exc.save(path)  # 指定路径,保存文件
            else:
                row_data = []
                index = 1
                for id in id_list:
                    data =  list(AttendanceTimeRecord.objects.filter(id=id,status=True).values_list("person__name", "person__dept_name", "person__pin", "event_time", "person__job_time"))[0]
                    data=(index,)+data
                    row_data.append(data)
                    if len(data) == 0:
                        index = index
                    index += 1
                exc = openpyxl.load_workbook(path)  # 打开整个excel文件
                sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
                for row in row_data:
                    sheet.append(row)  # 在工作表中添加一行
                exc.save(path)  # 指定路径,保存文件
        return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        return HttpResponse(json.dumps(return_data))
    else:
        return_data = {
            "code": HTTP_403_FORBIDDEN,
            "msg": "没有权限访问",
            'hidden': False
        }
        return return_data



def createExcelPath(file_name, t2,name,num,interval, *args):  # is not None

    import openpyxl
    from openpyxl.styles import Alignment
    import time
    exc = openpyxl.Workbook()
    sheet = exc.active
    for column in sheet.iter_cols(min_col=0, max_col=num):
        for cell in column:
            sheet.column_dimensions[cell.column_letter].width = 30
    sheet.column_dimensions['A'].width = 10
    sheet.title = file_name.split('.xlsx')[0]
    sheet.merge_cells(str(interval)) #'A1:D1'
    sheet['A1'] = name
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.append(args)
    t = time.strftime('%Y-%m-%d')
    path = os.path.join('static', 'rayongAttendance', 'download_file', t, t2, file_name)
    path = path.replace(os.sep, '/')
    exc.save(path)
    return path


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)
    else:
        pass
