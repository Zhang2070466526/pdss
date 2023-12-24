import datetime
import os
import string
import time
import uuid

import arrow
import openpyxl
from django.http import JsonResponse
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Side, Alignment, Font, Border
from django.core.mail import EmailMessage

from employee.models import HrEmployee
from pdss.settings import BASE_DIR
from abc import ABC, abstractmethod
from rest_framework.views import APIView
from django.core.cache import cache
from auther.models import AdminUser


# 类名方法头
class methHeader(ABC):

    def __init__(self, request):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功',
            'data': {}
        }
        self.meth = {}
        self.now_time = arrow.now().format("YYYY-MM-DD HH:mm:ss")
        self.operate_user_id = ''
        self.operate_user_name = ''

    def method_center(self, method):
        self.operate_user_id = self.request.check_token
        self.operate_user_name = get_admin_user_by_id(self.operate_user_id)
        self.meth[method]()
        return JsonResponse(self.return_data)

    @abstractmethod
    def add_meth(self):
        pass


# 基本增删改查方法
class BasicClass(ABC):
    @abstractmethod
    def create(self):
        pass

    @abstractmethod
    def update(self):
        pass

    @abstractmethod
    def search(self):
        pass

    @abstractmethod
    def delete(self):
        pass



class FileClass(ABC):
    @abstractmethod
    def upload(self):
        pass

    def download(self):
        pass


class ViewBasicTemplate(APIView):

    def __init__(self, className):
        self.obj = className(None)
        self.obj.add_meth()

    def post(self, request):
        """
        查询
        :param request:
        :return:
        """
        self.obj.request = request
        return self.obj.method_center('search')

    def put(self, request):
        """
        新增
        检测创建重复放在前端，使用查询完成
        :param request:
        :return:
        """
        self.obj.request = request
        return self.obj.method_center('create')

    def patch(self, request):
        """
        修改
        :param request:
        :return:
        """
        self.obj.request = request
        return self.obj.method_center('update')

    def delete(self, request):
        """
        删除
        :param request:
        :return:
        """
        self.obj.request = request
        return self.obj.method_center('delete')

    def get(self, request):
        """
        下拉框
        :param request:
        :return:
        """
        self.obj.request = request
        return self.obj.method_center('options')


# 根据id来获取用户名
def get_admin_user_by_id(operate_user_id):
    user_name = cache.get(f'admin_user_{operate_user_id}')
    if user_name is None or user_name == '':
        operate_user_name = '匿名'
        if operate_user_id is not None and operate_user_id != '':
            admin = AdminUser.objects.filter(pk=operate_user_id)
            if admin.exists():
                cache.set(f'admin_user_{operate_user_id}', admin[0].username, 60 * 60 * 60)
                operate_user_name = admin[0].username
    else:
        operate_user_name = user_name
    return operate_user_name


def get_employee_id_by_code(employee_code):
    pk = cache.get(f'hr_employee_{employee_code}')
    if pk is not None:
        return pk
    else:
        emp_obj = HrEmployee.objects.filter(employee_code=employee_code)
        if emp_obj.exists():
            cache.set(f'hr_employee_{employee_code}', emp_obj[0].id)
            return emp_obj[0].id
        else:
            return 0
    pass


# 格式化时间
def arrow_format_time(format_time, format_type):
    if format_type == 1:
        return arrow.get(format_time).format("YYYY-MM-DD")
    elif format_type == 2:
        return arrow.get(format_time).format("YYYY-MM-DD HH:mm:ss")
    elif format_type == 3:
        return arrow.get(format_time).format("YYYY-MM-01")


def get_my_ip(request):
    client_ip = request.META['REMOTE_ADDR']
    return client_ip


# 按照员工个人来保存文件
def save_file_by_employee(dir_list, file, file_name, code, file_type, name):
    """
    :param dir_list:    目录列表 从static后面开始写
    :param file:        二进制文件
    :param file_name:   文件名
    :param code:        工号
    :param file_type:   文件类型
    :param name:        姓名
    :return:
    """
    path = 'static/'
    for file_ in dir_list:
        path += f'{file_}/'
        if not os.path.exists(path):
            os.mkdir(path)
    path2 = path + f"{file_type + '-' + name + code + '.' + file_name.split('.')[-1]}"
    index = 1
    while os.path.exists(path2):
        path2 = path + f"{file_type + '-' + name + code + f'_{str(index)}' + '.' + file_name.split('.')[-1]}"
        index += 1
    path = path2
    with open(f"{path}", 'wb+') as fp:
        for chunk in file.chunks():
            fp.write(chunk)
    return path


# 保存数据为excel
def save_xlsx(root_dir, dir_name, title, row2, data_obj, save_filename):
    """
    :param root_dir static下面的一级目录
    :param dir_name: 保存文件目录，位于static/sendEmail/dir_name
    :param title:  excel的标题
    :param letter_list: ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
    :param row2: 第二行的数据
    :param data_obj: 保存的数据列表
    :param save_filename: 保存文件的名称
    :return: 返回文件地址
    """
    if save_filename == '':
        save_filename = str(uuid.uuid4())
    letter_list = [get_alphabet(i) for i in range(1, len(row2) + 1)]
    side = Side(style="thick", color="000000")
    searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(os.path.join(os.path.join(BASE_DIR, "static"), root_dir)):
        os.mkdir(os.path.join(os.path.join(BASE_DIR, "static"), root_dir))
    if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name)):
        os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name))

    path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), root_dir), dir_name),
                        searchTime)
    if not os.path.exists(path):
        os.mkdir(path)
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(letter_list))
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
    ws.append(row2)
    for i in letter_list:
        ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
        ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
        ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
    for data in data_obj:
        ws.append(data)
    excel_columns = ws.columns

    len_list = count_excel_save(excel_columns)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
    wb.close()
    file_path = "static/" + root_dir + '/' + dir_name + "/" + searchTime + '/' + save_filename + ".xlsx"
    wb.save(file_path)

    return file_path


def save_xlsx_new(template_name, code, title, row2, data_obj, operate):
    """
    :param template_name: 那个模块下的操作
    :param code: 操作人工号
    :param title: excel的标题
    :param row2: excel的第二行数据
    :param data_obj:  保存的数据
    :param operate:  操作内容， 下载还是上传？
    :return:  返回地址
    """
    path = 'static/'
    dir_list = ['管理员文件操作', code, template_name, operate]
    for file_ in dir_list:
        path += f'{file_}/'
        if not os.path.exists(path):
            os.mkdir(path)
    path2 = path + f"{code + '_'  + template_name + '_' + arrow.now().format('YYYY-MM-DD') + '.xlsx'}"
    index = 1
    while os.path.exists(path2):
        path2 = path + f"{code + '_'  + template_name + '_' + arrow.now().format('YYYY-MM-DD') + '_' + str(index) + '.xlsx'}"
        index += 1
    path = path2
    letter_list = [get_alphabet(i) for i in range(1, len(row2) + 1)]
    side = Side(style="thick", color="000000")
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(letter_list))
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
    ws.append(row2)

    for i in letter_list:
        ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
        ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
        ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
    for data in data_obj:
        ws.append(data)
    excel_columns = ws.columns

    len_list = count_excel_save(excel_columns)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
    wb.close()
    wb.save(path)
    return path


# 保存上传的文件
def save_upload_file(dir_list, file, fileName):
    if fileName is None or fileName == '':
        try:
            fileName = str(uuid.uuid4()).replace('-', '') + '.' + file.name.split('.')[-1]
        except Exception as e:
            print(e)
            fileName = '保存文件报错.jpg'

    if fileName == '':
        pass
    path = 'static/'
    for file_ in dir_list:
        path += f'{file_}/'
        if not os.path.exists(path):
            os.mkdir(path)
    path = path + fileName
    with open(path, 'wb+') as fp:
        for chunk in file.chunks():
            fp.write(chunk)
    return path


# 获得字母对应的数字
def get_alphabet(n):
    if 1 <= n <= 26:
        return string.ascii_uppercase[n - 1]
    elif n > 26:
        quotient, remainder = divmod(n - 1, 26)
        return get_alphabet(quotient) + string.ascii_uppercase[remainder]


# 计算宽度
def count_excel_save(excel_columns):
    len_list = []
    prjTuple = tuple(excel_columns)
    for idx in range(0, len(prjTuple)):
        length2 = 0
        str_list = []
        for cell in prjTuple[idx]:
            if "1" not in str(cell):
                str_list.append(str(cell.value))
        for elem in str_list:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1.5
                else:
                    length += 2.5
            length2 = max(length, length2)
        len_list.append(length2)
    return len_list


def send_email(subject, message, recipient_list, file_list):
    """
    发送邮件
    :param subject: 主题
    :param message: 正文内容
    :param recipient_list: 接受者邮箱
    :param file_list: 附件列表
    :return:
    """
    email = EmailMessage(
        subject=subject,
        body=message,
        to=recipient_list,
    )
    for i in file_list:
        # print(i)
        email.attach_file(i)
    email.send()


# 计算宽度
def count_width(string):
    import re
    lang = 0
    string = str(string)
    num_regex = re.compile(r'[0-9]')
    zimu_regex = re.compile(r'[a-zA-z]')
    hanzi_regex = re.compile(r'[\u4E00-\u9FA5]')
    num_list = num_regex.findall(string)
    zimu_list = zimu_regex.findall(string)
    hanzi_list = hanzi_regex.findall(string)
    if len(num_list) >= 15:
        lang += len(num_list) * 10
    elif len(num_list) >= 10:
        lang += len(num_list) * 11
    else:
        lang += len(num_list) * 13
    if len(zimu_list) > 15:
        lang += len(zimu_list) * 9
    else:
        lang += len(zimu_list) * 13
    if len(hanzi_list) > 15:
        lang += len(hanzi_list) * 15
    elif len(hanzi_list) > 5:
        lang += len(hanzi_list) * 20
    else:
        lang += len(hanzi_list) * 25
    return lang
