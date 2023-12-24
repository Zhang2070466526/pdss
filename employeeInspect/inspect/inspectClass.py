import os,datetime,json,openpyxl,time,random,string
from datetime import datetime, date
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from salarySurvey.models import SalarySurveyRecord
from auther.models import *
from ..serializers import *
from ..models import *
from django.core.exceptions import ObjectDoesNotExist
from controller.controller2 import Controller, upload_file
from general.models import *
from utils.check_token import *
class Reset:
    # def verify(self,request):
    #     return_data = {'code': '', "message": ''}
    #     new_token = CheckToken()
    #     try:
    #         check_token = new_token.check_token(request.headers['Authorization'])
    #     except Exception as e:
    #         print(e)
    #         return_data['code'] = 400
    #         return_data['message'] = '请求参数出错啦'
    #         return return_data
    #     if check_token is None:
    #         return_data['code'] = 403
    #         return_data['message'] = '没有权限查询'
    #         return return_data

    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))
        self.methods[self.meth]()
        return Response(self.return_data)
    def get_list(self):
        obj = Controller(EmployeeInspect, "get_list", self.request)
        self.return_data = obj.data_start()
    def get_upload(self):
        # print(self.request.POST)
        # print(self.request.FILES)
        file = self.request.FILES.get("file", None)
        createPhoto = self.request.FILES.getlist("createPhoto", None)   #照片
        createPlan = self.request.FILES.getlist("createFile", None)  # 方案
        createData = self.request.POST.get("createData", None)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'employeeInspect', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)


        # print(createPlan,createPhoto,createData,file)
        if file and file is not None:
            # print("文件上传")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                # print(file_url,file_name,file_suffix)
                if file_name:
                    self.saveFile(file_name,file,file_suffix)#保存文件
                    exc = openpyxl.load_workbook(file_url)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1,name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg":"某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                employee_inspect_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,base_parent_id=base_father.id,status=True).id   #二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            employee_inspect_base_id = base_father.id
                        if employee_inspect_base_id not in self.request.user_base:
                            # print('no',employee_inspect_base_id,self.request.user_base)
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,3).value if sheet.cell(i + 1,3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:
                            data_kwargs = {
                                'employee_inspect_base_id':employee_inspect_base_id,
                                'employee_inspect_date': str(sheet.cell(i + 1, 4).value.date())if sheet.cell(i + 1, 4).value != None else None,
                                'employee_inspect_day_shift_no': sheet.cell(i + 1, 5).value,
                                'employee_inspect_night_shift_no': sheet.cell(i + 1, 6).value,
                                'employee_inspect_remark': sheet.cell(i + 1, 7).value,
                            }
                            data_kwargs['creator_id'] = self.request.check_token
                            if  data_kwargs['employee_inspect_base_id']==None and data_kwargs['employee_inspect_date'] ==None and data_kwargs['employee_inspect_day_shift_no']==None and data_kwargs['employee_inspect_night_shift_no']==None and data_kwargs['employee_inspect_remark']==None:
                                continue
                            else:
                                # print(data_kwargs)

                                EmployeeInspect.objects.create(**data_kwargs)

            except Exception as e:
                # print(e)
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            # print("新增数据")
            createData = eval(createData)  #{"employee_inspect_date": "2024-01-14", "employee_inspect_day_shift_no": 5, 'employee_inspect_night_shift_no': 2,"employee_inspect_remark": "说明一下","employee_inspect_bas_id": 1}
            # print(createData,type(createData))
            createData['employee_inspect_day_shift_no'] = self.decision_num(createData['employee_inspect_day_shift_no'],int)  # 关闭项
            createData['employee_inspect_night_shift_no'] = self.decision_num(createData['employee_inspect_night_shift_no'],int)  # 关闭项
            if len(createData['employee_inspect_date'])==0 or len(str(createData['employee_inspect_base_id']))==0:
                self.return_data = {
                    "code": status.HTTP_400_BAD_REQUEST,
                    "msg": "日期或中心是必填的哦!"
                }
            else:
                # print(createData)
                createData['creator_id'] = self.request.check_token
                obj = EmployeeInspect.objects.create(**createData)
                for file in createPhoto:
                    file_url,file_name,file_suffix=self.createPath(file,self.create_file_name(obj.employee_inspect_remark,file))
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_url,
                        "file_name": file_name,
                    }
                    file_obj = UploadFiles.objects.create(**file_kwargs)
                    file_obj.employee_inspect_photos.add(obj.id)
                for file in createPlan:
                    file_url,file_name,file_suffix=self.createPath(file,self.create_file_name(obj.employee_inspect_remark,file))
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_url,
                        "file_name": file_name,
                    }
                    file_obj = UploadFiles.objects.create(**file_kwargs)
                    file_obj.employee_inspect_plans.add(obj.id)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }

    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(EmployeeInspect, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(EmployeeInspect, "patch", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def download_file(self):
        # print("putdata", self.request.data)
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'employeeInspect', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('员工稽核收集表.xlsx')
        # print("path", path)

        if downloadAll == True:  # 是下载全部   有条件
            print('是下载全部')
            kwargs = {'employee_inspect_status': True}
            employee_inspect_base = self.request.GET.get('baseNameId', None)  # 公司名称
            # if len(str(employee_inspect_base))==0:
            #     print("+++++++++++++++++++++++++++++++++")
            #     employee_inspect_base =self.request.check_token
            # print("111", employee_inspect_base, len(str(employee_inspect_base)))
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            # print(beginDate,endDate,self.request.GET)
            if employee_inspect_base == ''  and beginDate == "" and endDate == "":  # 全查
                kwargs['employee_inspect_status'] = True
                kwargs['employee_inspect_base__in'] = self.request.user_base
            if employee_inspect_base != '':
                kwargs['employee_inspect_base'] = employee_inspect_base
            if beginDate != "" and endDate != "":
                kwargs['employee_inspect_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
                kwargs['employee_inspect_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            # print("kwargs:",kwargs)
            all = list(EmployeeInspect.objects.filter(**kwargs).all().values_list('employee_inspect_base_id',
                    'employee_inspect_base__name', 'employee_inspect_date', 'employee_inspect_day_shift_no', 'employee_inspect_night_shift_no',
                    'employee_inspect_remark'))
            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                if i[1] == None or i[3] == None:
                    # self.return_data = {
                    #     "code": status.HTTP_401_UNAUTHORIZED,
                    #     "msg": "该条数据日期或者公司为空，无法计算并下载！"
                    # }
                    pass
                else:
                    # print(i[2],type(i[2]),self.get_first_last(i[2]))
                    month_day_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                     employee_inspect_base_id=i[1],
                                                                     employee_inspect_date__gte=
                                                                     self.get_first_last(i[3])[0],
                                                                     employee_inspect_date__lte=
                                                                     self.get_first_last(i[3])[1]).values_list(
                        'employee_inspect_day_shift_no', flat=True)
                    month_day_count = list(filter(None, month_day_count))  # 列表去除None
                    month_night_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                       employee_inspect_base_id=i[1],
                                                                       employee_inspect_date__gte=
                                                                       self.get_first_last(i[3])[0],
                                                                       employee_inspect_date__lte=
                                                                       self.get_first_last(i[3])[1]).values_list(
                        'employee_inspect_night_shift_no', flat=True)
                    month_night_count = list(filter(None, month_night_count))  # 列表去除None
                    month_count = sum(month_day_count) + sum(month_night_count)  # 该月白班稽查次数+夜班稽查次数
                    # print(month_count)

                    year_day_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                    employee_inspect_base_id=i[1],
                                                                    employee_inspect_date__gte=
                                                                    self.get_first_last(i[3])[2],
                                                                    employee_inspect_date__lte=
                                                                    self.get_first_last(i[3])[3]).values_list(
                        'employee_inspect_day_shift_no', flat=True)
                    year_day_count = list(filter(None, year_day_count))  # 列表去除None
                    year_night_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                      employee_inspect_base_id=i[1],
                                                                      employee_inspect_date__gte=
                                                                      self.get_first_last(i[3])[2],
                                                                      employee_inspect_date__lte=
                                                                      self.get_first_last(i[3])[3]).values_list(
                        'employee_inspect_night_shift_no', flat=True)
                    year_night_count = list(filter(None, year_night_count))  # 列表去除None
                    year_count = sum(year_day_count) + sum(year_night_count)  # 该年白班稽查次数+夜班稽查次数
                    # print(year_count)
                    i.extend([month_count, year_count])

                i.insert(0, index)
                del i[2]
                index += 1
                # print(i)
                row_data.append(i)
            # print(row_data)

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            # print('不是全部')
            row_data = []
            index = 1
            for id in id_list:
                data = EmployeeInspect.objects.filter(employee_inspect_status=True, id=id).values_list('employee_inspect_base_id',
                    'employee_inspect_base__name', 'employee_inspect_date', 'employee_inspect_day_shift_no', 'employee_inspect_night_shift_no',
                    'employee_inspect_remark')
                for i in data:
                    i = list(i)

                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)


                    if i[1]==None or i[3]==None:
                        # self.return_data = {
                        #     "code": status.HTTP_401_UNAUTHORIZED,
                        #     "msg": "该条数据日期或者公司为空，无法计算并下载！"
                        # }
                        pass
                    else:
                    # print(i[2],type(i[2]),self.get_first_last(i[2]))
                        month_day_count= EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                     employee_inspect_base_id=i[1],
                                                                     employee_inspect_date__gte=self.get_first_last(i[3])[0],
                                                                     employee_inspect_date__lte=self.get_first_last(i[3])[1]).values_list('employee_inspect_day_shift_no',flat=True)
                        month_day_count = list(filter(None, month_day_count))  # 列表去除None
                        month_night_count =EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                     employee_inspect_base_id=i[1],
                                                                     employee_inspect_date__gte=self.get_first_last(i[3])[0],
                                                                     employee_inspect_date__lte=self.get_first_last(i[3])[1]).values_list('employee_inspect_night_shift_no',flat=True)
                        month_night_count = list(filter(None, month_night_count))  # 列表去除None
                        month_count=sum(month_day_count)+sum(month_night_count)   #该月白班稽查次数+夜班稽查次数
                        # print(month_count)

                        year_day_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                             employee_inspect_base_id=i[1],
                                                                             employee_inspect_date__gte=self.get_first_last(i[3])[2],
                                                                             employee_inspect_date__lte=self.get_first_last(i[3])[3]).values_list('employee_inspect_day_shift_no', flat=True)
                        year_day_count = list(filter(None, year_day_count))  # 列表去除None
                        year_night_count = EmployeeInspect.objects.filter(employee_inspect_status=True,
                                                                               employee_inspect_base_id=i[1],
                                                                               employee_inspect_date__gte=self.get_first_last(i[3])[2],
                                                                               employee_inspect_date__lte=self.get_first_last(i[3])[3]).values_list('employee_inspect_night_shift_no', flat=True)
                        year_night_count = list(filter(None, year_night_count))  # 列表去除None
                        year_count=sum(year_day_count)+sum(year_night_count)   #该年白班稽查次数+夜班稽查次数
                        # print(year_count)
                        i.extend([month_count, year_count])

                    i.insert(0, index)

                    del i[2]
                    row_data.append(list(i))
                    if len(i)==0:
                        index=index
                    index += 1
            # print(row_data)
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # print(self.return_data)


    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # print(fileName)
        return fileName.split('.')[0]

    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'employeeInspect','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'employeeInspect','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'employeeInspect', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/employeeInspect/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/employeeInspect/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        # print('保存成功')

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 107):
            sheet.column_dimensions[chr(i).upper()].width = 20
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:I1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '员工稽核收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号','中心/基地','公司','稽核日期',	'白班稽核次数','夜班稽核次数','说明','月度总稽核次数',"年度累计稽核次数"),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'employeeInspect', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path
    def mkdir(self,path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            # print('文件夹创建成功：', path)

        else:
            pass
            # print('文件夹已经存在：', path)


    def get_first_last(self,date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # print(weekDay,monthCountDay)
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month=datetime.date(year, month, day=1)
        last_month=datetime.date(year, month, day=monthCountDay)
        first_year=datetime.date(year,1,1)
        last_year=datetime.date(year,12,31)
        return first_month,last_month,first_year,last_year

    # print(get_first_last('2022', '08'))
    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return