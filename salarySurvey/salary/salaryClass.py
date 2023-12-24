# -*- coding: utf-8 -*-
# @Time    : 2023/4/25 16:28
# @Author  : zhuang
# @Site    : 
# @File    : salaryClass.py
# @Software: PyCharm
import os
import datetime
import json
import re
import time

from django.core.exceptions import ObjectDoesNotExist
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter

from general.models import center_base
from pdss.settings import BASE_DIR
from salarySurvey.models import SalarySurveyRecord
from auther.models import *
from ..serializers import *
from utils.check_token import *
from controller.controller import Controller, upload_file


def verify(request):
    return_data = {'code': '', "msg": ''}
    new_token = CheckToken()
    try:
        check_token = new_token.check_token(request.headers['Authorization'])
    except Exception as e:
        # print(e)
        return_data['code'] = 400
        return_data['msg'] = '请求参数出错啦'
        return return_data
    if check_token is None:
        return_data['code'] = 403
        return_data['msg'] = '没有权限查询'
        return return_data


class salaryClass:
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
            "delete_other_file": self.delete_other_file,
            "upload_other_file": self.upload_other_file,
        }
        self.auth = ''
        self.check_token = ''
        self.user_base = self.request.user_base

    def meth_center(self):
        if verify(self.request) is not None:
            return Response(verify(self.request))
        self.check_token = self.request.check_token
        self.methods[self.meth]()
        return Response(self.return_data)

    def get_upload(self):
        file = self.request.FILES.get("file", None)
        createFile = self.request.FILES.getlist("createFile", None)
        createData = self.request.POST.get("createData", None)
        searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
        print(createData)
        if file and file != "":
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "数据文件上传成功"
            }
            try:
                self.fileName = upload_file(file, "salaryFile", "upload", None)
                if self.fileName:
                    self.save_Data()
            except Exception as e:
                # print(e)
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": str(e)
                }
                if "cannot be null" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = SalarySurveyRecord._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 不能为空,请检查!"

                if "expected a number" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = SalarySurveyRecord._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 的数据格式有误，应为数字格式"

        if createData is not None and createData != "":
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "数据新增成功"
            }
            createData = eval(createData)
            new_data = {key: value for key, value in createData.items() if value != ''}
            obj = SalarySurveyRecord.objects.update_or_create(**new_data)[0]

            for file in createFile:
                fileName = obj.name + "_" + searchTime + "_" + str(
                    "".join(list(str(time.time()))[0:10])) + "_" + file.name
                fileName = upload_file(file, "salaryFile", "other_file", fileName)
                file_url = "static/" + "salaryFile" + "/other_file/" + searchTime + "/" + fileName
                file_kwargs = {
                    "file_url": file_url,
                    "file_name": fileName,
                }
                file_obj = UploadFiles.objects.update_or_create(**file_kwargs)[0]
                file_obj.historical_payslip.add(obj.id)

    def get_list(self):
        obj = Controller(SalarySurveyRecord, "get_list", self.request)
        self.return_data = obj.data_start()

    # 将上传的文件数据保存到数据库中
    def save_Data(self):
        # print("读取数据中")
        workbook = openpyxl.load_workbook(self.fileName)
        table = workbook.active
        for i in range(2, table.max_row):
            if table.cell(i + 1, 2).value != "" and table.cell(i + 1, 2).value is not None and \
                    table.cell(i + 1, 3).value != "" and table.cell(i + 1, 3).value is not None and \
                    table.cell(i + 1, 4).value != "" and table.cell(i + 1, 4).value is not None and \
                    table.cell(i + 1, 5).value != "" and table.cell(i + 1, 5).value is not None:
                try:
                    employee_base_father = center_base.objects.get(status=1,
                                                                   name=table.cell(i + 1, 1).value.replace("（",
                                                                                                           "(").replace(
                                                                       "）", ")"))
                except ObjectDoesNotExist:
                    self.return_data = {
                        "code": status.HTTP_400_BAD_REQUEST,
                        "msg": "基地名不存在，请重新选择"
                    }
                    return
                obj = center_base.objects.filter(base_parent_id=employee_base_father.id, status=1)
                if obj.exists():
                    employee_base_son = center_base.objects.get(status=1, name=table.cell(i + 1, 2).value.replace("（",
                                                                                                                  "(").replace(
                        "）", ")"))
                else:
                    employee_base_son = employee_base_father
                if employee_base_father.id not in self.user_base or employee_base_son.id not in self.user_base:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "抱歉，您没有上传 " + table.cell(i + 1, 1).value + "-" + table.cell(i + 1,
                                                                                           2).value + " (基地/中心/公司)的权限"
                    }
                else:
                    data_kwargs = {
                        "salary_base": employee_base_son,
                        "period": table.cell(i + 1, 4).value,
                        "name": table.cell(i + 1, 5).value,
                        "previous_work_country": table.cell(i + 1, 6).value,
                        "previous_work_city": table.cell(i + 1, 7).value,
                        "previous_company": table.cell(i + 1, 8).value,
                        "previous_department": table.cell(i + 1, 9).value,
                        "previous_post": table.cell(i + 1, 10).value,
                        "salary_calc_method": table.cell(i + 1, 11).value,
                        "monthly_income": table.cell(i + 1, 12).value,
                        "base_salary": table.cell(i + 1, 13).value,
                        "post_salary": table.cell(i + 1, 14).value,
                        "skill_salary": table.cell(i + 1, 15).value,
                        "performance_salary": table.cell(i + 1, 16).value,
                        "seniority_subsidy": table.cell(i + 1, 17).value,
                        "other_subsidy": table.cell(i + 1, 18).value,
                        "other_illustrate": table.cell(i + 1, 19).value,
                        "overtime_hours": table.cell(i + 1, 20).value,
                        "other_program": table.cell(i + 1, 21).value,
                        "overtime_base_salary": table.cell(i + 1, 22).value,
                        "description_of_leave": table.cell(i + 1, 23).value,
                        "bonus_standard": table.cell(i + 1, 24).value,
                        "bonus_total": table.cell(i + 1, 25).value,
                        "bonus_distribute_times": table.cell(i + 1, 26).value,
                        "bonus_distribute_way": table.cell(i + 1, 27).value,
                        "social_security_cardinality": table.cell(i + 1, 28).value,
                        "provident_fund_cardinality": table.cell(i + 1, 29).value,
                        "attach_business_insurance": table.cell(i + 1, 30).value,
                        "accommodation": table.cell(i + 1, 31).value,
                        "meals": table.cell(i + 1, 32).value,
                        "transportation": table.cell(i + 1, 33).value,
                        "spring_festival_benefits": table.cell(i + 1, 34).value,
                        "dragon_boat_festival_benefits": table.cell(i + 1, 35).value,
                        "mid_autumn_festival_benefits": table.cell(i + 1, 36).value,
                        "other_festival_benefits": table.cell(i + 1, 37).value,
                        "residency_country": table.cell(i + 1, 38).value,
                        "residency_start_date": table.cell(i + 1, 39).value,
                        "residency_end_date": table.cell(i + 1, 40).value,
                        "residency_allowance": table.cell(i + 1, 41).value,
                        "residency_welfare": table.cell(i + 1, 42).value,
                        "residency_allowance_distribute_way": table.cell(i + 1, 43).value,
                        "status": 1,
                        "fix_detail_creator": AdminUser.objects.filter(pk=self.check_token)[0],
                        "fix_detail_modifier": AdminUser.objects.filter(pk=self.check_token)[0],
                    }
                    SalarySurveyRecord.objects.create(**data_kwargs)

    # 下载文件
    def download_file(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功"
        }
        info = json.loads(self.request.body)
        get_obj = self.request.GET
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25

        kwargs = {'status': 1}

        searchTime = datetime.datetime.now().strftime("%Y-%m-%d")
        filename = '薪资调研'
        id_list = info['idList']
        is_all = info['downloadAll']
        fields = {}
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "salaryFile"), "download")):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "salaryFile"), "download"))
        path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "salaryFile"), "download"),
                            searchTime)
        if not os.path.exists(path):
            os.mkdir(path)
        if len(id_list) <= 0 and is_all == 0:
            return
        if is_all == 0 and len(id_list) <= 0:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "条件为空"
            }
            return

        if id_list is not None and len(id_list) > 0:
            kwargs["id__in"] = list(id_list)

        if is_all:
            kwargs = {'status': 1}
            model_fields = {"searchName": "name__contains",
                            "beginDate": "period__gte",
                            "endDate": "period__lte", }
            model_key = ["searchName", "beginDate", "endDate"]
            for key, value in get_obj.items():
                if value != "" and key in model_key:
                    kwargs[model_fields[key]] = value

        obj = SalarySurveyRecord.objects.filter(**kwargs).order_by("create_time")
        if len(obj) > 0:
            obj = SalarySurveyRecordPutSerializers(instance=obj, many=True).data
            for field in SalarySurveyRecord._meta.get_fields():
                if field.name not in ["fix_detail_creator", "fix_detail_modifier", "status", "create_time",
                                      "modify_time"]:
                    fields[field.verbose_name] = field.name

            letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                           'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI',
                           'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR']
            width = 3.0
            side = Side(style="thick", color="000000")
            font_header = Font(name="等线", size=14, bold=True)

            wb = Workbook()
            ws = wb.active
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=44)
            font_header = Font(name="等线", size=14, bold=True)
            ws['A1'] = "薪资统计"
            ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
            ws['A1'].font = font_header

            row2 = [key for key, value in fields.items()]
            row2[0] = "序号"
            row2.insert(2, "中心/事业部")
            fields["中心/事业部"] = "salary_base_father"
            ws.append(row2)
            for i in letter_list:
                ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
                ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
            index = 1
            for data in obj:
                row_data = []

                for k in row2:
                    if k != "序号":
                        row_data.append(dict(data)[fields[k]])
                    else:
                        row_data.append(0)
                row_data[0] = index
                index += 1
                ws.append(row_data)

            excel_columns = ws.columns
            len_list = self.count_excel_save(excel_columns)

            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
            wb.close()
            filename = filename + "".join(list(str(time.time()))[0:10])
            wb.save("static/salaryFile/download/" + searchTime + '/' + filename + ".xlsx")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '下载成功',
                "downloadUrl": "static/salaryFile/download/" + searchTime + '/' + filename + ".xlsx",
            }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '无该用户数据'
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(SalarySurveyRecord, "patch", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": str(e)
            }

    def delete_data(self):
        try:
            obj = Controller(SalarySurveyRecord, "delete", self.request)
            self.return_data = obj.data_start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def count_character(self, s):
        hanzi = 0
        num = 0
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1

        return str(30 * hanzi + 20 * num)

    def count_excel_save(self, excel_columns):

        len_list = []
        prjTuple = tuple(excel_columns)
        for idx in range(0, len(prjTuple)):
            length2 = 0
            str_list = []
            for cell in prjTuple[idx]:
                if "1" not in str(cell):
                    str_list.append(str(cell.value))
            for elem in str_list:
                elem_split = list(elem)
                length = 0
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1.5
                    else:
                        length += 2.5
                length2 = max(length, length2)
            len_list.append(length2)
        return len_list

    def delete_other_file(self):
        obj = Controller(SalarySurveyRecord, "delete_other_file", self.request)
        self.return_data = obj.data_start()

    def upload_other_file(self):
        obj = Controller(SalarySurveyRecord, "upload_other_file", self.request)
        self.return_data = obj.data_start()
