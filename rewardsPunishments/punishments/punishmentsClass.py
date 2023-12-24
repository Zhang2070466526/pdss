import os,datetime,json,openpyxl,time,random,string
from datetime import datetime, date
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from auther.models import *
from ..serializers import *
from ..models import *
from django.core.exceptions import ObjectDoesNotExist
from controller.controller2 import Controller, upload_file
from general.models import *
from utils.check_token import *
class Resetpb:  #项目奖金

    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))
        self.methods[self.meth]()
        return Response(self.return_data)
    def get_list(self):
        
        obj = Controller(ProjectBonus, "get_list", self.request)
        self.return_data = obj.data_start()
    def get_upload(self):
        
        
        file = self.request.FILES.get("file", None)
        createData = self.request.POST.get("createData", None)
        # print(createData)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'xprojectBonusFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        # 
        if file and file is not None:
            
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name,file,file_suffix)#保存文件
                    exc = openpyxl.load_workbook(file_url,data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1,name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg":"某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                project_bonus_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,base_parent_id=base_father.id,status=True).id   #二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            project_bonus_base_id = base_father.id
                        if project_bonus_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,3).value if sheet.cell(i + 1,3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:

                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    project_bonus_date=str(sheet.cell(i + 1, 4).value.date())
                                except:#可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                project_bonus_date=None
                            #
                            #
                            pb_avarge=sheet.cell(i + 1, 9).value#人均奖励
                            #

                            project_bonus_no = self.decision_num(sheet.cell(i + 1, 5).value, int)
                            project_bonus_reach_no=self.decision_num(sheet.cell(i + 1, 6).value, int)
                            project_bonus_total=self.decision_num(sheet.cell(i + 1, 7).value, float)
                            project_bonus_person_no=self.decision_num(sheet.cell(i + 1, 8).value, int)
                            if pb_avarge=='' or len(str(pb_avarge))==0 or pb_avarge==None:  #人均奖励为空
                                if sheet.cell(i + 1, 7).value!=None and sheet.cell(i + 1, 8).value!=None:  #奖金总额和享受人次都有  计算
                                    pb_avarge=sheet.cell(i + 1, 7).value/sheet.cell(i + 1, 8).value  #奖金总额/享受人次
                                elif sheet.cell(i + 1, 7).value!=None or sheet.cell(i + 1, 8).value!=None:#奖金总额和享受人次有一个没有  不计算，直接为None
                                    pb_avarge =None

                            data_kwargs = {
                                'project_bonus_base_id':project_bonus_base_id,
                                'project_bonus_date': project_bonus_date,
                                'project_bonus_no':project_bonus_no,#项目数量
                                'project_bonus_reach_no': project_bonus_reach_no,#达成数量
                                'project_bonus_total':project_bonus_total,  #奖金总额
                                'project_bonus_person_no':project_bonus_person_no,#享受人次
                                'project_bonus_average':pb_avarge,#人均奖励
                            }
                            data_kwargs['creator_id'] = self.request.check_token

                            if  data_kwargs['project_bonus_base_id']==None and data_kwargs['project_bonus_date'] ==None and data_kwargs['project_bonus_no']==None and data_kwargs['project_bonus_reach_no']==None and data_kwargs['project_bonus_total']==None and data_kwargs['project_bonus_person_no']==None :
                                continue
                            else:
                                ProjectBonus.objects.create(**data_kwargs)

            except Exception as e:
                
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            
            createData = eval(createData)#{"project_bonus_date": "2023-02-02", "project_bonus_no": 3,'project_bonus_reach_no': 21, "project_bonus_total": 2000,'project_bonus_person_no':4,"project_bonus_base_id": 3}
            # if len(createData['project_bonus_date']) ==10:#发送的年月日
            #     time_obj=datetime.strptime(createData['project_bonus_date'], "%Y-%m-%d")
            #     month=time_obj.month
            #     year=time_obj.year
            #     createData['project_bonus_date']='{}-{}'.format(year,month)
            # elif len(createData['project_bonus_date']) ==7:#发送的年月
            #     createData['project_bonus_date']=createData['project_bonus_date']
            # else:
            #     createData['project_bonus_date']=None
            createData['creator_id'] = self.request.check_token
            createData['project_bonus_no'] = self.decision_num(createData['project_bonus_no'], int)
            createData['project_bonus_reach_no'] = self.decision_num(createData['project_bonus_reach_no'], int)
            createData['project_bonus_total'] = self.decision_num(createData['project_bonus_total'], float)
            createData['project_bonus_person_no'] = self.decision_num(createData['project_bonus_person_no'], int)
            if createData['project_bonus_no']<createData['project_bonus_reach_no']:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据关系不符合,新增失败"
                }
            else:
                try:
                    createData['project_bonus_average']=createData['project_bonus_total']/createData['project_bonus_person_no']
                except:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "数据填写不完整，无法计算人均"
                    }
                
                ProjectBonus.objects.create(**createData)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(ProjectBonus, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }
    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(ProjectBonus, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }
    def download_file(self):
        
        
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'xprojectBonusFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('项目奖金收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'project_bonus_status': True}
            project_bonus_base = self.request.GET.get('baseNameId', None)  # 公司名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if project_bonus_base == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['project_bonus_status'] = True
                kwargs['project_bonus_base__in'] = self.request.user_base
            if project_bonus_base != '':
                kwargs['project_bonus_base'] = project_bonus_base
            if beginDate != "" and endDate != "":
                kwargs['project_bonus_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
                kwargs['project_bonus_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            
            all = list(ProjectBonus.objects.filter(**kwargs).all().values_list('project_bonus_base',
                    'project_bonus_base__name', 'project_bonus_date','project_bonus_no', 'project_bonus_reach_no', 'project_bonus_total',
                    'project_bonus_person_no', 'project_bonus_average'))

            
            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                if i[1] == None or i[3] == None or i[5] == None:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "该条数据日期或者公司或者金额为空，无法计算并下载！"
                    }
                    pass
                else:
                    year_project_bonus_total_count = ProjectBonus.objects.filter(project_bonus_status=True,
                                                                                 project_bonus_base_id=i[1],
                                                                                 project_bonus_date__gte=self.get_first_last(i[3])[2],
                                                                                 project_bonus_date__lte=self.get_first_last(i[3])[3]).values_list('project_bonus_total', flat=True)  # 年度奖金总额
                    year_project_bonus_total_count = list(filter(None, year_project_bonus_total_count))  # 列表去除None
                    
                    i.append(sum(year_project_bonus_total_count))
                i.insert(0, index)
                del i[2]
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data = []
            index = 1
            for id in id_list:
                data = ProjectBonus.objects.filter(project_bonus_status=True, id=id).values_list('project_bonus_base',
                    'project_bonus_base__name', 'project_bonus_date','project_bonus_no', 'project_bonus_reach_no', 'project_bonus_total',
                    'project_bonus_person_no', 'project_bonus_average')

                for i in data:
                    i=list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    if i[1] == None or i[3] == None or i[5] == None:
                        self.return_data = {
                            "code": status.HTTP_401_UNAUTHORIZED,
                            "msg": "该条数据日期或者公司或者金额为空，无法计算并下载！"
                        }
                        pass
                    else:
                        year_project_bonus_total_count = ProjectBonus.objects.filter(project_bonus_status=True,
                                                                                     project_bonus_base_id=i[1],
                                                                                     project_bonus_date__gte=
                                                                                     self.get_first_last(i[3])[2],
                                                                                     project_bonus_date__lte=
                                                                                     self.get_first_last(i[3])[
                                                                                         3]).values_list(
                            'project_bonus_total', flat=True)  # 年度奖金总额
                        year_project_bonus_total_count = list(filter(None, year_project_bonus_total_count))  # 列表去除None

                        i.append(sum(year_project_bonus_total_count))
                    i.insert(0, index)
                    del i[2]
                    row_data.append(list(i))
                    if len(i) == 0:
                        index = index
                    index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # 

    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'xprojectBonusFile','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'xprojectBonusFile','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'xprojectBonusFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/xprojectBonusFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/xprojectBonusFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 107):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i>=104:
                sheet.column_dimensions[chr(i).upper()].width = 18

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:J1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '项目奖金收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号','中心/基地','公司','日期','项目数量','达成数量','奖金总额','享受人次','人均奖励','年度累计金额'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'xprojectBonusFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path
    def mkdir(self,path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            
        else:
            pass
            
    def get_first_last(self,date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month=datetime.date(year, month, day=1)
        last_month=datetime.date(year, month, day=monthCountDay)
        first_year=datetime.date(year,1,1)
        last_year=datetime.date(year,12,31)
        return first_month,last_month,first_year,last_year
    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return
class Resetrp:   #奖惩数据
    # def verify(self,request):
    #     return_data = {'code': '', "message": ''}
    #     new_token = CheckToken()
    #     try:
    #         check_token = new_token.check_token(request.headers['Authorization'])
    #     except Exception as e:
    #         
    #         return_data['code'] = 400
    #         return_data['message'] = '请求参数出错啦'
    #         return return_data
    #     if check_token is None:
    #         return_data['code'] = 403
    #         return_data['message'] = '没有权限查询'
    #         return return_data
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))
        self.methods[self.meth]()
        return Response(self.return_data)

    def get_list(self):
        obj = Controller(RewardsAndPunishments, "get_list", self.request)
        self.return_data = obj.data_start()

    def get_upload(self):
        
        
        file = self.request.FILES.get("file", None)
        createData = self.request.POST.get("createData", None)

        # 
        if file and file is not None:
            
            t = time.strftime('%Y-%m-%d')
            dummy_path = os.path.join(BASE_DIR, 'static', 'rewardsAndPunishmentsFile', 'upload_file', t)  # 创建文件夹
            self.mkdir(dummy_path)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    exc = openpyxl.load_workbook(file_url, data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1, name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg": "某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                r_p_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,
                                                                                   base_parent_id=base_father.id,
                                                                                   status=True).id  # 二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            r_p_base_id = base_father.id
                        if r_p_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,
                                                                                                          3).value if sheet.cell(
                                    i + 1, 3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:


                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    r_p_date=str(sheet.cell(i + 1, 4).value.date())
                                except:#可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg":"有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                r_p_date=None
                #             #
                #             #

                            rewards_money = self.decision_num(sheet.cell(i + 1, 5).value, float)#奖励金额
                            rewards_person_no= self.decision_num(sheet.cell(i + 1, 6).value, int)#奖励人次
                            punishments_money= self.decision_num(sheet.cell(i + 1, 8).value, float)#惩处金额
                            punishments_person_no= self.decision_num(sheet.cell(i + 1, 9).value, int)#惩处人次


                            if rewards_money!=None and rewards_person_no!=None:
                                if rewards_person_no==0:
                                    rewards_average=0
                                else:
                                    rewards_average=rewards_money/rewards_person_no
                            else:
                                rewards_average =None
                            if punishments_money != None and punishments_person_no != None:
                                if punishments_person_no==0:
                                    punishments_average=0
                                else:
                                    punishments_average = punishments_money / punishments_person_no
                            else:
                                punishments_average = None



                            # rewards_average= sheet.cell(i + 1, 6).value  # 奖励人均
                            # if rewards_average == '' or len(str(rewards_average)) == 0 or rewards_average == None:  # 人均奖励为空
                            #     if sheet.cell(i + 1, 4).value != None and sheet.cell(i + 1,5).value != None:  # 奖金金额和獎勵人次都有  计算
                            #         rewards_average = sheet.cell(i + 1, 4).value / sheet.cell(i + 1, 5).value  # 奖金总额/獎勵人次
                            #     elif sheet.cell(i + 1, 4).value != None or sheet.cell(i + 1,5).value != None:  #奖金金额和獎勵人次一个都没有  不计算，直接为None
                            #         rewards_average = None
                            # else:#不为空
                            #     if sheet.cell(i + 1, 4).value != None or sheet.cell(i + 1,5).value != None:  # 奖金金额和獎勵人次有一个没有  不计算，直接为None
                            #         rewards_average = None
                            #     if sheet.cell(i + 1, 4).value != None and sheet.cell(i + 1,5).value != None:  # 奖金金额和獎勵人次一个都没有  不计算，直接为None
                            #         rewards_average = None


                            # punishments_average = sheet.cell(i + 1, 9).value  # 惩处人均
                            # if punishments_average == '' or len(str(punishments_average)) == 0 or punishments_average == None:  # 惩处人均为空
                            #     if sheet.cell(i + 1, 7).value != None and sheet.cell(i + 1,8).value != None:  # 惩处金额和惩处人次都有  计算
                            #         punishments_average = sheet.cell(i + 1, 7).value / sheet.cell(i + 1, 8).value  # 惩处金额/獎勵人次
                            #     elif sheet.cell(i + 1, 7).value != None or sheet.cell(i + 1,8).value != None:  #惩处金额和惩处人次有一个没有  不计算，直接为None
                            #         punishments_average= None
                            # else:#不为空
                            #     if sheet.cell(i + 1, 7).value != None or sheet.cell(i + 1,8).value != None:  # 惩处金额和惩处人次有一个没有  不计算，直接为None
                            #         punishments_average= None
                            #     if sheet.cell(i + 1, 7).value != None and sheet.cell(i + 1,8).value != None:  # 惩处金额和惩处人次有一个没有  不计算，直接为None
                            #         punishments_average= None





                            data_kwargs = {
                                'r_p_base_id': r_p_base_id,
                                'r_p_date': r_p_date,
                                'rewards_money': rewards_money,  # 奖励金额
                                'rewards_person_no': rewards_person_no,  # 奖励人次
                                "rewards_average":rewards_average,#奖励人均
                                'punishments_money': punishments_money,  # 惩处金额
                                'punishments_person_no':punishments_person_no,  # 惩处人次
                                'punishments_average': punishments_average,  # 惩处人均
                            }
                            data_kwargs['creator_id'] = self.request.check_token

                            if data_kwargs['r_p_base_id'] == None and data_kwargs['r_p_date'] == None and data_kwargs['rewards_money'] == None and data_kwargs['rewards_person_no'] == None and data_kwargs['punishments_money'] == None and data_kwargs['punishments_person_no'] == None:

                                continue
                            else:

                                RewardsAndPunishments.objects.create(**data_kwargs)

            except Exception as e:
                
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }

        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            
            createData = eval(createData)  # {"r_p_date": "2023-02-02", "rewards_money": 300.3,'rewards_person_no': 21, "punishments_money": 2000.434,'punishments_person_no':4,"r_p_base_id": 3}

            createData['rewards_money']  = self.decision_num(createData['rewards_money'],float)
            createData['punishments_money'] = self.decision_num(createData['punishments_money'], float)
            createData['rewards_person_no'] = self.decision_num(createData['rewards_person_no'], int)
            createData['punishments_person_no'] = self.decision_num(createData['punishments_person_no'], int)
            createData['creator_id'] = self.request.check_token
            if createData['rewards_person_no']==0:
                createData['rewards_average']=0
            else:
                createData['rewards_average'] = createData['rewards_money'] / createData['rewards_person_no']
            if createData['punishments_person_no']==0:
                createData['punishments_average'] = 0
            else:
                createData['punishments_average'] = createData['punishments_money'] / createData[
                    'punishments_person_no']
            # try:
            #     createData['rewards_average'] = createData['rewards_money'] / createData['rewards_person_no']
            #     createData['punishments_average'] = createData['punishments_money'] / createData['punishments_person_no']
            # except:
            #     self.return_data = {
            #         "code": status.HTTP_401_UNAUTHORIZED,
            #         "msg": "数据填写不完整，无法计算人均"
            #     }
            # print(createData)
            
            RewardsAndPunishments.objects.create(**createData)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "用户数据新增成功"
            }

    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(RewardsAndPunishments, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }
    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(RewardsAndPunishments, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }
    def download_file(self):
        
        
        
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'rewardsAndPunishmentsFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('奖惩数据收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'r_p_status': True}
            r_p_base = self.request.GET.get('baseNameId', None)  # 公司名称id
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)
            if r_p_base == ''  and beginDate == "" and endDate == "":  # 全查
                kwargs['r_p_status'] = True
                kwargs['r_p_base__in'] = self.request.user_base
            if r_p_base != '':
                kwargs['r_p_base'] = r_p_base
            if r_p_base != '':
                kwargs['r_p_base'] = r_p_base
            if beginDate != "" and endDate != "":
                kwargs['r_p_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
                kwargs['r_p_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            # 
            all = list(RewardsAndPunishments.objects.filter(**kwargs).all().values_list('r_p_base_id',
                    'r_p_base__name', 'r_p_date', 'rewards_money', 'rewards_person_no','rewards_average',
                    'punishments_money', 'punishments_person_no','punishments_average'))
            
            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                if i[1] == None or i[3] == None:
                    # self.return_data = {
                    #     "code": status.HTTP_401_UNAUTHORIZED,
                    #     "msg": "该条数据日期或者公司为空，无法计算并下载！"
                    # }
                    pass
                elif i[4] ==None or i[5] ==None:
                    pass
                elif i[7] ==None or i[8] ==None:
                    pass
                else:
                    year_rewards_person_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                         r_p_base_id=i[1],
                                                                                         r_p_date__gte=self.get_first_last(i[3])[2],
                                                                                         r_p_date__lte=self.get_first_last(i[3])[3]).values_list('rewards_person_no', flat=True)  # 年奖励累计人次
                    # 
                    year_rewards_money_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                        r_p_base_id=i[1],
                                                                                        r_p_date__gte=self.get_first_last(i[3])[2],
                                                                                        r_p_date__lte=self.get_first_last(i[3])[3]).values_list('rewards_money', flat=True)  # 年奖励累计金额

                    year_punishments_person_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                             r_p_base_id=i[1],
                                                                                             r_p_date__gte=self.get_first_last(i[3])[2],
                                                                                             r_p_date__lte=self.get_first_last(i[3])[3]).values_list('punishments_person_no', flat=True)  # 年惩处累计人次

                    year_punishments_money_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                            r_p_base_id=i[1],
                                                                                            r_p_date__gte=self.get_first_last(i[3])[2],
                                                                                            r_p_date__lte=self.get_first_last(i[3])[3]).values_list('punishments_person_no', flat=True)  # 年惩处累计金额

                    year_rewards_person_count = list(filter(None, year_rewards_person_count))
                    year_rewards_money_count = list(filter(None, year_rewards_money_count))
                    year_punishments_person_count = list(filter(None, year_punishments_person_count))
                    year_punishments_money_count = list(filter(None, year_punishments_money_count))

                    i.extend([sum(year_punishments_person_count), sum(year_punishments_money_count)])
                    i.insert(6, sum(year_rewards_person_count))
                    i.insert(7, sum(year_rewards_money_count))
                i.insert(0, index)
                del i[2]
                row_data.append(list(i))
                
                if len(i) == 0:
                    index = index
                index += 1
                # 

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data = []
            index = 1
            for id in id_list:
                data = RewardsAndPunishments.objects.filter(r_p_status=True, id=id).values_list('r_p_base_id',
                    'r_p_base__name', 'r_p_date', 'rewards_money', 'rewards_person_no','rewards_average',
                    'punishments_money', 'punishments_person_no','punishments_average')
                for i in data:
                    i = list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    if i[1] == None or i[3] == None:
                        # self.return_data = {
                        #     "code": status.HTTP_401_UNAUTHORIZED,
                        #     "msg": "该条数据日期或者公司为空，无法计算并下载！"
                        # }
                        pass
                    elif i[4] == None or i[5] == None:
                        pass
                    elif i[7] == None or i[8] == None:
                        pass
                    else:
                        year_rewards_person_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                         r_p_base_id=i[1],
                                                                                         r_p_date__gte=
                                                                                         self.get_first_last(i[3])[2],
                                                                                         r_p_date__lte=
                                                                                         self.get_first_last(i[3])[
                                                                                             3]).values_list(
                            'rewards_person_no', flat=True)  # 年奖励累计人次
                        #
                        year_rewards_money_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                        r_p_base_id=i[1],
                                                                                        r_p_date__gte=
                                                                                        self.get_first_last(i[3])[2],
                                                                                        r_p_date__lte=
                                                                                        self.get_first_last(i[3])[
                                                                                            3]).values_list(
                            'rewards_money', flat=True)  # 年奖励累计金额

                        year_punishments_person_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                             r_p_base_id=i[1],
                                                                                             r_p_date__gte=
                                                                                             self.get_first_last(i[3])[
                                                                                                 2],
                                                                                             r_p_date__lte=
                                                                                             self.get_first_last(i[3])[
                                                                                                 3]).values_list(
                            'punishments_person_no', flat=True)  # 年惩处累计人次

                        year_punishments_money_count = RewardsAndPunishments.objects.filter(r_p_status=True,
                                                                                            r_p_base_id=i[1],
                                                                                            r_p_date__gte=
                                                                                            self.get_first_last(i[3])[
                                                                                                2],
                                                                                            r_p_date__lte=
                                                                                            self.get_first_last(i[3])[
                                                                                                3]).values_list(
                            'punishments_person_no', flat=True)  # 年惩处累计金额

                        year_rewards_person_count = list(filter(None, year_rewards_person_count))
                        year_rewards_money_count = list(filter(None, year_rewards_money_count))
                        year_punishments_person_count = list(filter(None, year_punishments_person_count))
                        year_punishments_money_count = list(filter(None, year_punishments_money_count))

                        i.extend([sum(year_punishments_person_count), sum(year_punishments_money_count)])
                        i.insert(6, sum(year_rewards_person_count))
                        i.insert(7, sum(year_rewards_money_count))
                    i.insert(0, index)
                    del i[2]
                    row_data.append(list(i))

                    if len(i) == 0:
                        index = index
                    index += 1
                    # 

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        




    def mkdir(self,path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            

        else:
            pass
            

    def create_file_name(self, objName, file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t = time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def createPath(self, pic, func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func + 'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'rewardsAndPunishmentsFile', 'upload_file', t, pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png', 'icon']:
            pic_path = os.path.join('static', 'rewardsAndPunishmentsFile', 'upload_file', t, pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'rewardsAndPunishmentsFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/rewardsAndPunishmentsFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/rewardsAndPunishmentsFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 111):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i >=103:
                sheet.column_dimensions[chr(i).upper()].width = 20
            if i==98:
                sheet.column_dimensions[chr(i).upper()].width = 25

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:N1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '奖惩数据收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号', '中心/基地','公司', '日期', '奖励金额', '奖励人次', '奖励人均','年度奖励累计人次','年度奖励累计金额','惩处金额','惩处人次','惩处人均','年度惩处累计人次','年度惩处累计金额'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'rewardsAndPunishmentsFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def get_first_last(self,date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month=datetime.date(year, month, day=1)
        last_month=datetime.date(year, month, day=monthCountDay)
        first_year=datetime.date(year,1,1)
        last_year=datetime.date(year,12,31)
        return first_month,last_month,first_year,last_year

    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return