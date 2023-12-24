import json
from django.db.models import Q
from openpyxl import load_workbook

from monthReport.intern.basicInfo.methods import get_center_base
from monthReport.intern.models import internLeaveInfo
from utils.genericMethods import *


class internDepartInfoClass(methHeader, BasicClass):
    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete

    def create(self):
        info = json.loads(self.request.body)
        intern_center_base_id = info.pop('intern_leave_center_base_id')
        if type(intern_center_base_id) == list:
            intern_center_base_id = intern_center_base_id[0]
        update_field = ['intern_employee_id']
        for field in internLeaveInfo._meta.get_fields():
            update_field.append(field.name)
        kwargs = {'creator_id': self.operate_user_id, 'intern_employee_id': info.pop('intern_employee__employee_code')}
        for key, value in info.items():
            if key in update_field:
                if key == 'intern_leave_date' and value != '' and value is not None:
                    value = arrow_format_time(value, 1)
                kwargs[key] = value
        kwargs['intern_leave_center_base_id'] = intern_center_base_id
        if kwargs['intern_employee_id'] == '' or kwargs['intern_employee_id'] is None:
            self.return_data['msg'] = '请填写离职人信息'
            return
        internLeaveInfo.objects.create(**kwargs)
        self.return_data['msg'] = '新增成功'

    def search(self):
        info = json.loads(self.request.body)
        currentPage = int(info['currentPage'])
        pageSize = int(info['pageSize'])
        searchName = info['searchName']
        baseNameId = info['baseNameId']

        kwargs = {
            'intern_status': True
        }
        if baseNameId != '':
            kwargs['intern_leave_center_base_id'] = baseNameId
        if 'beginDate' in info:
            if info['beginDate'] != '' and info['beginDate'] is not None:
                kwargs['intern_leave_date__gte'] = arrow_format_time(info['beginDate'], 1)
        if 'endDate' in info:
            if info['endDate'] != '' and info['endDate'] is not None:
                kwargs['intern_leave_date__lte'] = arrow_format_time(info['endDate'], 1)

        q_obj = Q()
        q_obj.connector = 'or'
        q_obj.children.append(('creator_id', self.operate_user_id))
        q_obj.children.append(('intern_leave_center_base_id__in', self.request.user_base))
        if searchName is not None and searchName != '':
            q_obj.children.append(('intern_employee__employee_code__contains', searchName))
            q_obj.children.append(('intern_employee__employee_name__contains', searchName))
        tableList = []
        columnList = [
            {'label': 'index', "value": '序号', 'width': 60},
            {'label': 'center_base', "value": '中心/事业部', 'width': 10},
            {'label': 'intern_leave_center_base__name', "value": '公司名称', 'width': 10},
            {'label': 'intern_employee__employee_code', "value": '工号', 'width': 10},
            {'label': 'intern_employee__employee_name', "value": '姓名', 'width': 10},
            {'label': 'intern_employee__employee_group_join_date', "value": '入职日期', 'width': 10},
            {'label': 'intern_leave_date', "value": '离职日期', 'width': 10},
            {'label': 'intern_term', "value": '届别', 'width': 10},
            {'label': 'intern_type', "value": '类别', 'width': 10},
            {'label': 'intern_employee__employee_department__department_full_name', "value": '部门全称', 'width': 10},
            {'label': 'intern_employee__employee_position__position_name', "value": '岗位名称', 'width': 10},
            {'label': 'intern_employee__employee_job_class__job_class_name', "value": '职等', 'width': 10},
            {'label': 'performance_one', "value": '绩效1', 'width': 10},
            {'label': 'performance_two', "value": '绩效2', 'width': 10},
            {'label': 'performance_three', "value": '绩效3', 'width': 10},
            {'label': 'intern_leave_reason', "value": '离职原因', 'width': 10},
        ]
        field_width = {}
        for obj in columnList:
            field_width[obj['label']] = count_width(obj['value'])
        totalNumber = internLeaveInfo.objects.filter(q_obj, **kwargs).count()
        leave_obj = internLeaveInfo.objects.filter(q_obj, **kwargs)[
                    (currentPage - 1) * pageSize: currentPage * pageSize].values(
            'id',
            'intern_leave_center_base__name',
            'intern_leave_center_base_id',
            'intern_leave_center_base__base_parent_id',
            'intern_employee__employee_code',
            'intern_employee__employee_name',
            'intern_employee__employee_group_join_date',
            'intern_employee__employee_department__department_full_name',
            'intern_employee__employee_position__position_name',
            'intern_employee__employee_job_class__job_class_name',
            'intern_leave_date',
            'intern_term',
            'intern_type',
            'performance_one',
            'performance_two',
            'performance_three',
            'intern_leave_reason',
        )
        index = (currentPage - 1) * pageSize + 1
        for data in leave_obj:
            data['center_base'] = get_center_base(data.pop('intern_leave_center_base__base_parent_id'))
            # 计算宽度
            for field, value in data.items():
                if field in ['intern_employee__employee_group_join_date', 'intern_leave_date']:
                    if data[field] != '' and data[field] is not None:
                        data[field] = arrow_format_time(data[field], 1)
                if field in field_width:
                    field_width[field] = max(field_width[field], count_width(value))
            data['index'] = index
            index += 1
            tableList.append(data)
        for obj in columnList:
            if obj['label'] in field_width:
                obj['width'] = field_width[obj['label']]
        self.return_data['data']['columnList'] = columnList
        self.return_data['data']['tableList'] = tableList
        self.return_data['data']['totalNumber'] = totalNumber

    def update(self):
        info = json.loads(self.request.body)
        pk = info.pop('id')
        intern_center_base_id = info.pop('intern_leave_center_base_id')
        if type(intern_center_base_id) == list:
            intern_center_base_id = intern_center_base_id[0]
        update_field = [
            'intern_leave_date',
            'intern_term',
            'intern_type',
            'performance_one',
            'performance_two',
            'performance_three',
            'intern_leave_reason',
        ]
        kwargs = {
            'modify_time': self.now_time,
            'intern_leave_center_base_id': intern_center_base_id,
            'modifier_id': self.operate_user_id,
        }
        for key, value in info.items():
            if key in update_field:
                if 'performance' in key:
                    if value == '' or value is None:
                        value = 0
                    else:
                        value = float(value)
                kwargs[key] = value
        internLeaveInfo.objects.filter(pk=pk).update(**kwargs)
        self.return_data['msg'] = '修改成功'

    def delete(self):
        info = json.loads(self.request.body)
        idList = info['idList']
        internLeaveInfo.objects.filter(pk__in=idList).update(
            modify_time=self.now_time,
            intern_status=False
        )
        self.return_data['msg'] = '删除成功'


class internDepartFileInfoClass(methHeader, FileClass):
    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download

    def upload(self):
        file = self.request.FILES.get('file')
        baseNameId = self.request.POST['uploadBaseId']
        if baseNameId is None or baseNameId == '' or baseNameId == 'undefined':
            self.return_data['code'] = 400
            self.return_data['msg'] = "请选择中心/事业部"
            return
        if len(baseNameId) == list:
            baseNameId = baseNameId[0]
        url = save_file_by_employee(['intern', '离职信息', 'upload'], file, file.name, '测试', '追光者月度基本信息导入名单',
                                    self.operate_user_name)
        sheet = load_workbook(url, data_only=True)
        sheet = sheet.active
        field_to_index = {
            'intern_employee__employee_code': 2,
            'intern_leave_date': 3,
            'intern_term': 4,
            'intern_type': 5,
            'performance_one': 6,
            'performance_two': 7,
            'performance_three': 8,
            'intern_leave_reason': 9,
        }
        intern_term = ''
        msg = ''
        for row in sheet.iter_rows(min_row=3):
            kwargs = {
                'intern_leave_center_base_id': baseNameId,
                'creator_id': self.operate_user_id,
                'modifier_id': self.operate_user_id,
            }
            # 将元组转换为列表
            row = list(row)
            # 遍历循环获取字段对应的值
            value_is_none_count = 0
            for field, index in field_to_index.items():
                if row[index - 1].value is not None:
                    kwargs[field] = row[index - 1].value

                if row[index - 1].value is None or row[index - 1].value == '':
                    value_is_none_count += 1
            if value_is_none_count >= 5:
                continue
            # 如果届别为空，则默认为上一次的值
            if 'intern_term' not in kwargs:
                kwargs['intern_term'] = intern_term
            else:
                intern_term = kwargs['intern_term']
            if kwargs['intern_leave_date'] != '' and kwargs['intern_leave_date'] is not None:
                kwargs['intern_leave_date'] = arrow_format_time(kwargs['intern_leave_date'], 1)
            intern_employee_id = get_employee_id_by_code(kwargs['intern_employee__employee_code'])
            if intern_employee_id == 0:
                msg += f"{kwargs['intern_employee__employee_code']}该工号有误！；"
                continue
            kwargs.pop('intern_employee__employee_code')
            kwargs['intern_employee_id'] = intern_employee_id
            internLeaveInfo.objects.update_or_create(defaults=kwargs, intern_employee_id=intern_employee_id)
        if msg != '':
            self.return_data['msg'] = msg[:-1]
        else:
            self.return_data['msg'] = '上传成功'

    def download(self):
        info = json.loads(self.request.body)
        searchName = info['searchName']
        baseNameId = info['baseNameId']
        downloadAll = info['downloadAll']
        idList = info['idList']
        kwargs = {
            'intern_status': True
        }
        q_obj = Q()
        q_obj.connector = 'or'
        if downloadAll == 1:
            if baseNameId != '':
                kwargs['intern_leave_center_base_id'] = baseNameId
            if 'beginDate' in info:
                if info['beginDate'] != '' and info['beginDate'] is not None:
                    kwargs['intern_leave_date__gte'] = arrow_format_time(info['beginDate'], 1)
            if 'endDate' in info:
                if info['endDate'] != '' and info['endDate'] is not None:
                    kwargs['intern_leave_date__lte'] = arrow_format_time(info['endDate'], 1)

            if searchName is not None and searchName != '':
                q_obj.children.append(('intern_employee__employee_code__contains', searchName))
                q_obj.children.append(('intern_employee__employee_name__contains', searchName))
        else:
            kwargs['pk__in'] = idList
        leave_obj = internLeaveInfo.objects.filter(q_obj, **kwargs).values(
            'intern_leave_center_base__name',
            'intern_leave_center_base__base_parent_id',
            'intern_employee__employee_code',
            'intern_employee__employee_name',
            'intern_employee__employee_group_join_date',
            'intern_employee__employee_department__department_full_name',
            'intern_employee__employee_position__position_name',
            'intern_employee__employee_job_class__job_class_name',
            'intern_leave_date',
            'intern_term',
            'intern_type',
            'performance_one',
            'performance_two',
            'performance_three',
            'intern_leave_reason',
        )

        index = 1
        tableList = []
        for data in leave_obj:
            data['center_base'] = get_center_base(data.pop('intern_leave_center_base__base_parent_id'))
            # 计算宽度
            for field, value in data.items():
                if field in ['intern_employee__employee_group_join_date', 'intern_leave_date']:
                    if data[field] != '' and data[field] is not None:
                        data[field] = arrow_format_time(data[field], 1)
            data['index'] = index
            index += 1
            new_data = [
                data['index'],
                data['center_base'],
                data['intern_leave_center_base__name'],
                data['intern_employee__employee_code'],
                data['intern_employee__employee_name'],
                data['intern_employee__employee_group_join_date'],
                data['intern_employee__employee_department__department_full_name'],
                data['intern_employee__employee_position__position_name'],
                data['intern_employee__employee_job_class__job_class_name'],
                data['intern_leave_date'],
                data['intern_term'],
                data['intern_type'],
                data['performance_one'],
                data['performance_two'],
                data['performance_three'],
                data['intern_leave_reason'],
            ]
            tableList.append(new_data)
        row2 = [
            '序号',
            '中心/事业部',
            '公司',
            '工号',
            '姓名',
            '入职日期',
            '部门全称',
            '岗位名称',
            '职等',
            '离职日期',
            '界别',
            '类别',
            '绩效1',
            '绩效2',
            '绩效3',
            '离职原因',
        ]
        url = save_xlsx_new('追光者离职数据', self.operate_user_name, '追光者离职数据统计表', row2, tableList, 'download')
        self.return_data['downloadUrl'] = url
        self.return_data['msg'] = '下载成功'
