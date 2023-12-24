import json
from abc import ABC

import arrow

from general.models import center_base
from utils.genericMethods import *
from openpyxl import load_workbook
from monthReport.intern.models import *
from django.core.cache import cache
from django.db.models import Q


class internBasicInfoClass(methHeader, BasicClass):
    def add_meth(self):
        # self.meth['upload_basic_info'] = self.upload_basic_info
        # self.meth['get_basic_info'] = self.get_basic_info
        self.meth['create'] = self.create
        self.meth['update'] = self.update
        self.meth['search'] = self.search
        self.meth['delete'] = self.delete

    def create(self):
        info = json.loads(self.request.body)
        intern_center_base_id = info.pop('intern_center_base__id')
        if type(intern_center_base_id) == list:
            intern_center_base_id = intern_center_base_id[0]
        update_field = []
        for field in internBasicInfo._meta.get_fields():
            update_field.append(field.name)
        kwargs = {
            'creator_id': self.operate_user_id
        }
        for key, value in info.items():
            if key in update_field:
                if key in ['intern_num_at_begin', 'intern_num_join_in', 'intern_num_call_in', 'intern_num_call_out',
                           'intern_num_leave_out']:
                    if value == '' or value is None or value == 'None':
                        value = 0
                    else:
                        value = int(value)
                if key == 'intern_month' and value != '' and value is not None:
                    value = arrow.get(value).format("YYYY-MM-01")
                if value != '':
                    kwargs[key] = value
        kwargs['intern_num_at_end'] = kwargs['intern_num_at_begin'] + kwargs['intern_num_join_in'] + kwargs[
            'intern_num_call_in'] - kwargs['intern_num_call_out'] - kwargs['intern_num_leave_out']
        kwargs['modify_time'] = self.now_time
        kwargs['intern_center_base_id'] = intern_center_base_id

        internBasicInfo.objects.create(**kwargs)
        self.return_data['msg'] = '新增成功'

    def update(self):
        info = json.loads(self.request.body)
        pk = info.pop('id')
        intern_center_base_id = info.pop('intern_center_base__id')
        if type(intern_center_base_id) == list:
            intern_center_base_id = intern_center_base_id[0]
        update_field = []
        for field in internBasicInfo._meta.get_fields():
            update_field.append(field.name)
        kwargs = {
            'modifier_id': self.operate_user_id
        }
        for key, value in info.items():
            if key in update_field:
                if key in ['intern_num_at_begin', 'intern_num_join_in', 'intern_num_call_in', 'intern_num_call_out',
                           'intern_num_leave_out']:
                    if value == '' or value is None or value == 'None':
                        value = 0
                    else:
                        value = int(value)
                if key == 'intern_month' and value != '' and value is not None:
                    value = arrow.get(value).format("YYYY-MM-01")
                kwargs[key] = value
        kwargs['intern_num_at_end'] = kwargs['intern_num_at_begin'] + kwargs['intern_num_join_in'] + kwargs[
            'intern_num_call_in'] - kwargs['intern_num_call_out'] - kwargs['intern_num_leave_out']
        kwargs['modify_time'] = self.now_time
        kwargs['intern_center_base_id'] = intern_center_base_id
        internBasicInfo.objects.filter(pk=pk).update(**kwargs)
        self.return_data['msg'] = '修改成功'

    def search(self):
        info = json.loads(self.request.body)
        if 'currentPage' not in info or 'pageSize' not in info:
            self.return_data['msg'] = '缺少必要参数'
            self.return_data['code'] = 400
            return
        currentPage = int(info['currentPage'])
        pageSize = int(info['pageSize'])
        baseNameId = info['baseNameId']
        q_obj = Q()
        q_obj.connector = 'or'
        q_obj.children.append(('creator_id', self.operate_user_id))
        q_obj.children.append(('intern_center_base__in', self.request.user_base))
        kwargs = {
            'intern_status': True,
        }
        if info['beginDate'] != '' and info['beginDate'] is not None:
            kwargs['intern_month__gte'] = arrow.get(info['beginDate']).format("YYYY-MM-01")
        if info['endDate'] != '' and info['beginDate'] is not None:
            kwargs['intern_month__lte'] = arrow.get(info['endDate']).format("YYYY-MM-01")
        if baseNameId != '':
            kwargs['intern_center_base_id'] = baseNameId
        columnList = [
            {'label': 'index', 'value': '序号', 'width': '60'},
            {'label': 'center_base', 'value': '中心/事业部', 'width': count_width('中心/事业部')},
            {'label': 'intern_center_base__name', 'value': '公司', 'width': count_width('公司')},
        ]
        field_width = {'index': 60, 'center_base': count_width('中心/事业部'), 'intern_center_base__name': count_width('公司')}
        tableList = []
        no_show_list = ['id', 'create_time', 'modify_time', 'creator', 'modifier', 'intern_status',
                        'intern_center_base', 'intern_organization']

        for field in internBasicInfo._meta.get_fields():
            if field.name not in no_show_list:
                columnList.append(
                    {'label': field.name, 'value': field.verbose_name, 'width': count_width(field.verbose_name)}
                )
                field_width[field.name] = columnList[-1]['width']

        totalNumber = internBasicInfo.objects.filter(q_obj, **kwargs).count()
        intern_objs = internBasicInfo.objects.filter(q_obj, **kwargs).values(
            'id',
            'intern_center_base__name',
            'intern_center_base__id',
            'intern_center_base__base_parent_id',
            'intern_month',
            'intern_organization', 'intern_term', 'intern_type', 'intern_num_at_begin',
            'intern_num_join_in',
            'intern_num_call_in', 'intern_num_call_out', 'intern_num_leave_out', 'intern_num_at_end', 'intern_job_one',
            'intern_job_tow', 'intern_job_three', 'intern_job_four', 'intern_job_five', 'intern_yd_detail',
            'intern_yd_over_four_detail'
        )[(currentPage - 1) * pageSize: currentPage * pageSize]
        index = (currentPage - 1) * pageSize + 1
        for obj in intern_objs:

            obj['index'] = index
            index += 1
            obj['center_base'] = get_center_base(obj.pop('intern_center_base__base_parent_id'), obj['intern_center_base__name'])
            if obj['intern_month'] is not None and obj['intern_month'] != '':
                obj['intern_month'] = arrow.get(obj['intern_month']).format("YYYY-MM")
            for key, value in obj.items():
                if key in field_width:
                    field_width[key] = max(field_width[key], count_width(value))
            tableList.append(obj)
        for i in columnList:
            if i['label'] in field_width:
                i['width'] = field_width[i['label']]
        self.return_data['data']['tableList'] = tableList
        self.return_data['data']['columnList'] = columnList
        self.return_data['data']['totalNumber'] = totalNumber

    def delete(self):
        info = json.loads(self.request.body)
        idList = info['idList']
        internBasicInfo.objects.filter(pk__in=idList).update(intern_status=False, modify_time=self.now_time,
                                                             modifier_id=self.operate_user_id)
        self.return_data['msg'] = '删除成功'


class internBasicInfoFileClass(methHeader, FileClass):
    def add_meth(self):
        self.meth['upload'] = self.upload
        self.meth['download'] = self.download

    def upload(self):
        file = self.request.FILES.get('file')
        baseNameId = self.request.POST['uploadBaseId']
        if baseNameId is None or baseNameId == '' or baseNameId == 'undefined':
            self.return_data['code'] = 400
            self.return_data['msg'] = "请选择中心/事业部"
            return
        if len(baseNameId) == list:
            baseNameId = baseNameId[0]
        url = save_file_by_employee(['intern', '基本信息', 'upload'], file, file.name, '测试', '追光者月度基本信息导入名单',
                                    self.operate_user_name)
        sheet = load_workbook(url, data_only=True)
        sheet = sheet.active
        intern_month = sheet['D1'].value
        if intern_month == '' or intern_month is None:
            self.return_data['code'] = 400
            self.return_data['msg'] = '请输入月份'
            return
        intern_month = arrow.get(intern_month).format("YYYY-MM-01")
        field_to_index = {
            'intern_term': 1,
            'intern_type': 2,
            'intern_num_at_begin': 3,
            'intern_num_join_in': 4,
            'intern_num_call_in': 5,
            'intern_num_call_out': 6,
            'intern_num_leave_out': 7,
            'intern_num_at_end': 8,
            'intern_job_one': 9,
            'intern_job_tow': 10,
            'intern_job_three': 11,
            'intern_job_four': 12,
            'intern_job_five': 13,
            'intern_yd_detail': 14,
            'intern_yd_over_four_detail': 15,
        }
        intern_term = ''
        for row in sheet.iter_rows(min_row=4):
            kwargs = {
                'intern_month': intern_month,
                'intern_center_base_id': baseNameId,
            }
            # 将元组转换为列表
            row = list(row)
            # 遍历循环获取字段对应的值
            value_is_none_count = 0
            for field, index in field_to_index.items():
                if row[index - 1].value is not None:
                    if (row[index - 1].value is None or row[index - 1].value == '') and 3 <= index <= 13:
                        kwargs[field] = 0
                    else:
                        kwargs[field] = row[index - 1].value

                if row[index - 1].value is None or row[index - 1].value == '':
                    value_is_none_count += 1
            if value_is_none_count >= 14:
                print("存在13个为空哦")
                continue
            # 如果届别为空，则默认为上一次的值
            if 'intern_term' not in kwargs:
                kwargs['intern_term'] = intern_term
            else:
                intern_term = kwargs['intern_term']
            if 'intern_type' not in kwargs:
                kwargs['intern_type'] = ''
            kwargs['intern_status'] = True
            internBasicInfo.objects.update_or_create(defaults=kwargs, intern_month=intern_month,
                                                     intern_term=kwargs['intern_term'],
                                                     intern_type=kwargs['intern_type'])
        self.return_data['msg'] = '上传成功'

    def download(self):
        info = json.loads(self.request.body)
        kwargs = {'intern_status': True}
        if info['downloadAll'] == 1:
            baseNameId = info['baseNameId']
            kwargs = {'intern_status': True}
            if info['beginDate'] != '' and info['beginDate'] is not None:
                kwargs['intern_month__gte'] = arrow.get(info['beginDate']).format("YYYY-MM-01")
            if info['endDate'] != '' and info['beginDate'] is not None:
                kwargs['intern_month__lte'] = arrow.get(info['endDate']).format("YYYY-MM-01")
            if baseNameId != '':
                kwargs['intern_center_base_id'] = baseNameId
        else:
            kwargs['pk__in'] = info['idList']

        intern_objs = internBasicInfo.objects.filter(**kwargs).values_list(
            'intern_center_base__base_parent_id',
            'intern_center_base__name', 'intern_month', 'intern_term', 'intern_type', 'intern_num_at_begin',
            'intern_num_join_in',
            'intern_num_call_in', 'intern_num_call_out', 'intern_num_leave_out', 'intern_num_at_end', 'intern_job_one',
            'intern_job_tow', 'intern_job_three', 'intern_job_four', 'intern_job_five', 'intern_yd_detail',
            'intern_yd_over_four_detail'
        )
        data_obj = []
        for obj in intern_objs:
            obj = list(obj)
            if obj[1] != '' and obj[1] is not None:
                obj[2] = arrow.get(obj[2]).format('YYYY-MM')
            obj.insert(1, get_center_base(obj[0], obj[1]))
            del obj[0]
            data_obj.append(obj)

        row2 = [
            '中心/事业部',
            '公司',
            '月份',
            '届别',
            '类别',
            '期初人数',
            '入职 人数',
            '调入人数',
            '调出人数',
            '离职人数',
            '期末人数',
            '职等1',
            '职等2',
            '职等3',
            '职等4',
            '职等5',
            '本月新晋升/定岗说明',
            '职级4及以上说明(岗位、职级、姓名)',
        ]
        url = save_xlsx_new('追光者每月基本数据', self.operate_user_name, '追光者每月基本数据', row2, data_obj, 'download')
        self.return_data['downloadUrl'] = url
        self.return_data['msg'] = '下载成功'


def get_center_base(pk, base_name=None):
    name = cache.get(f"get_center_base_{pk}")
    if name is not None:
        return name

    try:
        name = center_base.objects.get(pk=pk).name
        cache.set(f"get_center_base_{pk}", name, 60 * 60)
    except:
        name = base_name
    return name
