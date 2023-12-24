import datetime
import json
import os
import time

import openpyxl
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from rest_framework import status
from rest_framework.response import Response
from rest_framework.status import HTTP_403_FORBIDDEN, HTTP_401_UNAUTHORIZED, HTTP_200_OK

from hikCanteen.hikMethods import HikvisonInterface
import requests
from rest_framework.request import Request
from hikCanteen.otherMethods import Rules
from hikCanteen.serializers import *
from pdss.settings import BASE_DIR
from utils.check_token import CheckToken
from .models import MonthmoneyToOAList
from django.apps import apps


"""
包含功能：
1、人员、部门自动同步；
2、新入职人员自动发起流程充钱；
3、人员查询消费记录、剩余金额；
4、人员自主挂失卡；
同步规则：
1、每日同步部门和当日新增人员以及对应人脸（创建日期大于等于今天）；
2、每天早上单独同步未录入人脸人员的人脸信息；
3、每天晚上5点半按合同归属自动发起首充金额审批流程，流程结束后充值到对应系统里。其中检测流程中金额改为0的人，在人员表中打上标签，以后月度充值不参与；
4、每月7日按部门架构自动发起上月餐补申请流程，餐补计算规则见对应文档；
"""


# Create your views here.
# 每日新入职人员金额首充
def daliy_entry_person_for_oa(request):
    hik = HikvisonInterface()
    hik.daliy_entry_person_to_oa()
    return HttpResponse('首充成功')


def receive_oa_daliy(request):
    print(request.GET)
    print(request.POST)
    print(request.body)
    return HttpResponse('首充成功')


# 总部平台同步部门
def daliy_add_zb_dept(request):
    hik = HikvisonInterface()
    hik.add_org()
    return HttpResponse('总部新增部门成功')


# 总部平台同步每日新增人员
def daliy_add_zb_person(request):
    hik = HikvisonInterface()
    hik.add_person('zongbu')
    return HttpResponse('总部新增人员成功')


# 总部平台同步没有人脸的人员的人脸同步
def zongbuSingleFace(request):
    hik = HikvisonInterface()
    hik.add_picture_by_code('zongbu')
    return HttpResponse('总部单独新增人脸成功')


# 建湖二期平台同步部门
def daliy_add_jheq_dept(request):
    hik = HikvisonInterface()
    hik.host = 'https://10.70.81.1'
    hik.port = '8443'
    hik.appKey = '23120519'
    hik.appSecret = 'cpXW0sDLKpvvvOpFFL5t'
    hik.add_org()
    return HttpResponse('建湖二期新增部门成功')


# 建湖二期平台同步每日新增人员
def daliy_add_jheq_person(request):
    hik = HikvisonInterface()
    hik.host = 'https://10.70.81.1'
    hik.port = '8443'
    hik.appKey = '23120519'
    hik.appSecret = 'cpXW0sDLKpvvvOpFFL5t'
    hik.add_person('jianhu')
    return HttpResponse('建湖二期新增人员成功')


# 建湖二期平台同步没有人脸的人员的人脸同步
def jianhuerqiSingleFace(request):
    hik = HikvisonInterface()
    hik.host = 'https://10.70.81.1'
    hik.port = '8443'
    hik.appKey = '23120519'
    hik.appSecret = 'cpXW0sDLKpvvvOpFFL5t'
    hik.add_picture_by_code('jianhuerqi')
    return HttpResponse('总部单独新增人脸成功')


# 每日新入职人员金额首充
def entry_add_money(request):
    hik = HikvisonInterface()
    hik.entry_add_money()
    return HttpResponse('首充成功')


# 查询消费记录 ok
def select_expenses_record(request):
    # 获取当前时间
    current_time = datetime.datetime.utcnow().replace(tzinfo=datetime.timezone.utc)
    previous_time = current_time - datetime.timedelta(7)
    startTime = request.GET.get("startTime", None)
    endTime = request.GET.get("endTime", None)
    if startTime is None or startTime == "" or endTime is None or endTime == "":
        startTime = str(previous_time.isoformat())
        endTime = str(current_time.isoformat())
    else:
        date1 = datetime.datetime.strptime(startTime, "%Y-%m-%d").date()
        date2 = datetime.datetime.strptime(endTime, "%Y-%m-%d").date()
        if (date2 - date1).days >= 31:
            return HttpResponse(json.dumps({
                "code": 200,
                "msg": "日期范围不能超过30天"
            }))
        if date1 > date2:
            return HttpResponse(json.dumps({
                "code": 200,
                "msg": "日期先后顺序有误"
            }))
        date_format = "%Y-%m-%d %H:%M:%S"
        startTime = str(datetime.datetime.strptime((startTime + " 23:59:59"), date_format).replace(
            tzinfo=datetime.timezone.utc).isoformat())
        endTime = str(datetime.datetime.strptime((endTime + " 23:59:59"), date_format).replace(
            tzinfo=datetime.timezone.utc).isoformat())
    person_code = request.GET.get("code", None)
    hik = HikvisonInterface()
    personID = hik.search_person_by_code(person_code)
    res = hik.select_expenses_record(startTime, endTime, personID)
    return HttpResponse(json.dumps(res))


# 查询人员 ok
def select_person(request):
    person_code = request.GET.get("code", None)
    if person_code and person_code != "":
        hik = HikvisonInterface()
        res = hik.search_person_by_code(person_code)
        res = {
            "code": 200,
            "personID": res
        }
    else:
        res = {
            "code": 200,
            "msg": "工号不能为空"
        }
    return HttpResponse(json.dumps(res))


# 查询卡号 ok
def select_person_cards(request):
    person_code = request.GET.get("code", None)
    if person_code and person_code != '':
        hik = HikvisonInterface()
        res = hik.select_person_cards(person_code)
    else:
        res = {
            "code": 200,
            "msg": "工号不能为空"
        }
    return HttpResponse(json.dumps(res))


# 挂失卡片 ok
def report_loss_card(request):
    card_code = request.GET.get("cardNo", None)
    if card_code is None or card_code == "":
        res = {
            "code": 200,
            "msg": "卡号不能为空"
        }
    else:
        hik = HikvisonInterface()
        res = hik.report_loss_card(card_code)
    return HttpResponse(json.dumps(res))


# 卡解挂 ok
def report_unloss_card(request):
    card_code = request.GET.get("cardNo", None)
    if card_code is None or card_code == "":
        res = {
            "code": 200,
            "msg": "卡号不能为空"
        }
    else:
        hik = HikvisonInterface()
        res = hik.report_unloss_card(card_code)
    return HttpResponse(json.dumps(res))


# 查询余额 ok
def select_balance(request):
    person_code = request.GET.get("code", None)
    if person_code and person_code != "":
        hik = HikvisonInterface()
        balance = hik.select_balance(person_code)
    else:
        balance = {
            "code": 200,
            "msg": "工号不能为空"
        }
    return HttpResponse(json.dumps(balance))


def wechat(request):
    print(request)
    print(request.GET)
    print(request.POST)
    access_token_res = requests.get(
        url='https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid=ww923f97f0a2f22221&corpsecret=eDWRBfFIV7IrXnB6bAeiugETotgCHphCbVvBWmNbTn0')
    access_token = json.loads(access_token_res.content.decode('utf-8'))['access_token']
    param = {
        'access_token': access_token,
        'code': request.GET['code']
    }
    employee_info_res = requests.get(url='https://qyapi.weixin.qq.com/cgi-bin/auth/getuserinfo', params=param)
    print(json.loads(employee_info_res.content.decode('utf-8'))['userid'])
    return HttpResponseRedirect(
        "http://10.60.19.50:8080/#/?code=" + json.loads(employee_info_res.content.decode('utf-8'))['userid'])


# 更新人脸 ok
def update_face(request):
    file = request.FILES.get("imgs", None)
    person_code = request.POST.get("code", None)
    if file is not None and file != "" and person_code is not None and person_code != "":
        hik = HikvisonInterface()
        res = hik.update_one_face(file, person_code)
    else:
        if file is None or file == "":
            res = {
                "code": 200,
                "msg": "请选择人脸照片"
            }
        else:
            res = {
                "code": 200,
                "msg": "工号不能为空"
            }

    # hik.test_photo()
    return HttpResponse(json.dumps(res))


def get_face_pic(request):
    person_code = request.GET.get("code", None)
    if person_code is not None and person_code != "":
        hik = HikvisonInterface()
        res = hik.get_face_pic(person_code)

    else:
        res = {
            "code": 200,
            "msg": "工号不能为空"
        }
    return HttpResponse(json.dumps(res))


def month_money_to_oa(request):
    hik = Rules()
    hik.month_money_to_oa()
    return HttpResponse('月度发起oa流程成功')


def get_card_top_up_system(request):   #获取数据
    from datetime import datetime
    new_token = CheckToken()
    check_token = new_token.check_token(request.headers['Authorization'])
    # print(check_token)
    if check_token is not None:
        # print(request.GET)
        # columnList = [{"value": "序号", "label": "index", "width": "60"}, ]
        columnList= [
            {
                "value": "序号",
                "label": "index",
                "width": "60"
            },
            {
                "value": "周期",
                "label": "month",
                "width": 100
            },
            {
                "value": "是否在职",
                "label": "status",
                "width": "120"
            },
            {
                "value": "合同归属",
                "label": "jobRankCode",
                "width": 180
            },
            {
                "value": "姓名",
                "label": "name",
                "width": "60"
            },
            {
                "value": "部门号",
                "label": "departmentCode",
                "width": "90"
            },
            {
                "value": "部门名称",
                "label": "departmentName",
                "width": 180
            },
            {
                "value": "工号",
                "label": "code",
                "width": 100
            },
            {
                "value": "岗位级别名称",
                "label": "jobLevelName",
                "width": "180"
            },
            {
                "value": "岗位名称",
                "label": "postName",
                "width": "120"
            },
            {
                "value": "最后结薪日期",
                "label": "SeparationDate",
                "width": "180"
            },
            {
                "value": "余额",
                "label": "balance",
                "width": "60"
            },
            {
                "value": "补贴总和",
                "label": "butie_sum",
                "width": "120"
            },
            {
                "value": "统一充值金额",
                "label": "entryCanteenAmount",
                "width": "180"
            },
            {
                "value": "应退/应补金额",
                "label": "amount_to_replenished",
                "width": "200"
            },
            {
                "value": "实际充值金额",
                "label": "realEntryCanteenAmount",
                "width": "180"
            },

            {
                "value": "季度返现金额",
                "label": "amount_to_quarter",
                "width": "180"
            },
            {
                "value": "白班出勤",
                "label": "baibanchuqin",
                "width": "120"
            },
            {
                "value": "白班补贴",
                "label": "baibanbutie",
                "width": "120"
            },
            {
                "value": "夜班出勤",
                "label": "yebanchuqin",
                "width": "120"
            },
            {
                "value": "夜班补贴",
                "label": "yebanbutie",
                "width": "120"
            },
            {
                "value": "强转班出勤",
                "label": "qiangzhuanbanchuqin",
                "width": "150"
            },
            {
                "value": "强转班补贴",
                "label": "qiangzhuanbanbutie",
                "width": "150"
            },
            {
                "value": "周末加班",
                "label": "zhoumojiaban",
                "width": "120"
            },
            {
                "value": "周末加班补贴",
                "label": "zhoumojiabanbutie",
                "width": "180"
            },
            {
                "value": "加班大于2",
                "label": "jiaban_gt_2",
                "width": "140"
            },
            {
                "value": "加班大于2补贴",
                "label": "jiaban_gt_2_butie",
                "width": "200"
            },
            {
                "value": "加班大于5",
                "label": "jiaban_gt_5",
                "width": "140"
            },
            {
                "value": "加班大于5补贴",
                "label": "jiaban_gt_5_butie",
                "width": "200"
            },
            {
                "value": "加班大于7.5",
                "label": "jiaban_gt_7_5",
                "width": "180"
            },
            {
                "value": "加班大于7.5补贴",
                "label": "jiaban_gt_7_5_butie",
                "width": "240"
            },
            {
                "value": "加班大于10",
                "label": "jiaban_gt_10",
                "width": "160"
            },
            {
                "value": "加班大于10补贴",
                "label": "jiaban_gt_10_butie",
                "width": "220"
            },
            {
                "value": "加班大于12.5",
                "label": "jiaban_gt_12_5",
                "width": "200"
            },
            {
                "value": "加班大于12.5补贴",
                "label": "jiaban_gt_12_5_butie",
                "width": "260"
            },
            {
                "value": "请假大于4,责任制",
                "label": "qingjia_gt_4",
                "width": "250"
            },
            {
                "value": "请假小于7.5",
                "label": "qingjia_lt_7_5",
                "width": "180"
            },
            {
                "value": "请假大于7.5",
                "label": "qingjia_gt_7_5",
                "width": "180"
            },
            {
                "value": "出勤总和",
                "label": "chuqin_sum",
                "width": "120"
            },


        ]





        tableList = []
        count_len = {}
        currentPage = eval(request.GET.get("currentPage", None)) if request.GET.get("currentPage", None) != "" else 1
        pageSize = eval(request.GET.get("pageSize", None)) if request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'month_money_status': True
        }
        # print(request.GET)
        code = request.GET.get('code', None)  # 工号
        beginDate = request.GET.get('beginDate', None)  #周期
        endDate = request.GET.get('endDate', None)
        # jobRankCode = request.GET.get('jobRankCode', None)  # 合同归属
        isStatus = request.GET.get('isStatus', None)  # 是否在职
        SeparationDate = request.GET.get('SeparationDate', None)  # 最后结薪日期
        # print(request.GET)
        if code == '' or code == None or len(str(code)) == 0 and beginDate == "" or beginDate == None or len(
                str(beginDate)) == 0  and SeparationDate == "" or SeparationDate == None or len(
                str(SeparationDate)) == 0 and endDate == "" or endDate == None or len(
                str(isStatus)) == 0 and isStatus == "" or isStatus == None or len(str(isStatus)) == 0:  # 全查
            kwargs['month_money_status']=True
            # kwargs['jobRankCode__in'] = request.user_jobRank_hik
            # kwargs['jobRankCode__in'] = request.user_base
        # if jobRankCode != '':
        #     kwargs['jobRankCode'] = jobRankCode
        if beginDate != "" and endDate != "":
            kwargs['month__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None else beginDate
            kwargs['month__lte'] = datetime(3421, 10, 29, 7, 17, 1, 177) if endDate is None else endDate
        if isStatus != '':
            kwargs['status'] = isStatus
        if SeparationDate != "":
            kwargs['SeparationDate__contains']=str(SeparationDate)[:7]

        obj = MonthmoneyToOAList.objects.filter(Q(name__contains=code) | Q(code__contains=code)| Q(departmentName__contains=code) , **kwargs).order_by('-create_time')
        # print(obj)

        serializer_obj = MonthmoneyToOAListSerializers(
            instance=obj[(currentPage - 1) * pageSize:currentPage * pageSize], many=True).data
        for i in serializer_obj:
            for key, value in dict(i).items():
                if key == 'jobRankCode':
                    k = {
                        'name': i[key],
                        'status': True,
                    }
                    try:
                        id = list(center_base.objects.filter(**k).values_list('id', flat=True))[0]
                    except:
                        id = None
                    i[key + '_id'] = id
            tableList.append(dict(i))
        # for field in MonthmoneyToOAList._meta.get_fields():  # 获取该类内所有字段对象
        #     except_field = ["month_money_status", "modify_time", "creator", "modifier", "create_time"]
        #     if field.name not in except_field:
        #         # if field.verbose_name == "公司":
        #         #     columnList.append({
        #         #         "value": "中心/事业部",
        #         #         "label": "base_father",
        #         #         "width": "180",
        #         #     })
        #         field_label = {
        #             "value": field.verbose_name,
        #             "label": field.name,
        #             "width": count_character(field.verbose_name),
        #         }
        #         columnList.append(field_label)
        for line in tableList:
            line['month'] = str(line['month'])[:7]
            if line['chuqin_sum'] is None:
                try:
                    line['chuqin_sum'] = eval(str(line['baibanchuqin'])) + eval(str(line['yebanchuqin'])) + eval(
                        str(line['qiangzhuanbanchuqin']))
                except:
                    line['chuqin_sum'] = None
            if line['butie_sum'] is None:
                try:
                    line['butie_sum'] = eval(str(line['yebanbutie'])) + eval(str(line['baibanbutie'])) + eval(
                        str(line['qiangzhuanbanbutie'])) + eval(str(line['zhoumojiabanbutie'])) + eval(
                        str(line['jiaban_gt_2_butie'])) + eval(str(line['jiaban_gt_5_butie'])) + eval(
                        str(line['jiaban_gt_7_5_butie'])) + eval(str(line['jiaban_gt_10_butie'])) + eval(
                        str(line['jiaban_gt_12_5_butie']))
                except:
                    line['butie_sum'] = None
        for i, item in enumerate(tableList):
            item["index"] = pageSize * (currentPage - 1) + i + 1
        # for i in columnList:
        #     if i['label'] in ['code', 'month']:
        #         i['width'] = 100
        #     if i['label'] in ['jobRankCode', 'departmentName']:
        #         i['width'] = 180

        # del columnList[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,  # 表头
                'tableList': tableList,
                'totalNumber': obj.count()
            }
        }
        return HttpResponse(json.dumps(return_data))
    else:
        return_data = {
            "code": HTTP_403_FORBIDDEN,
            "msg": "没有权限访问",
            'hidden': False
        }
        return HttpResponse(json.dumps(return_data))


def get_meal_card_top_up_download(request):
    from datetime import datetime
    id_list = json.loads(request.body).get('idList')
    downloadAll = json.loads(request.body).get('downloadAll')

    t = time.strftime('%Y-%m-%d')
    t2 = str(time.time())
    dummy_path = os.path.join(BASE_DIR, 'static', 'hikCanteen', 'download_file', t, t2)  # 创建文件夹
    mkdir(dummy_path)
    model = apps.get_model('hikCanteen', 'MonthmoneyToOAList')
    fields = model._meta.get_fields()
    verbose_names = [field.verbose_name for field in fields][:-6]
    # verbose_names.insert(3, '中心/基地')
    # print(verbose_names)

    path = createExcelPath('饭卡充值出勤收集表.xlsx', t2, *verbose_names)
    # print(path)
    if downloadAll == True:  # 是下载全部   有条件
        row_data = []
        index = 1
        kwargs = {
            'month_money_status': True
        }
        code = request.GET.get('code', None)  # 工号
        beginDate = request.GET.get('beginDate', None)
        endDate = request.GET.get('endDate', None)
        jobRankCode = request.GET.get('jobRankCode', None)  # 合同归属
        isStatus = request.GET.get('isStatus', None)  # 是否在职
        SeparationDate = request.GET.get('SeparationDate', None)  # 最后结薪日期
        if code == '' or code == None or len(str(code)) == 0 and SeparationDate == '' or SeparationDate == None or len(
                str(SeparationDate)) == 0 and beginDate == "" or beginDate == None or len(
                str(beginDate)) == 0 and endDate == "" or endDate == None or len(
                str(isStatus)) == 0 and isStatus == "" or isStatus == None or len(str(isStatus)) == 0:  # 全查
            kwargs['month_money_status'] = True

        if beginDate != "" and endDate != "":
            kwargs['month__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None else beginDate
            kwargs['month__lte'] = datetime(3421, 10, 29, 7, 17, 1, 177) if endDate is None else endDate
        if isStatus != '':
            kwargs['status'] = isStatus
        if SeparationDate != "":
            kwargs['SeparationDate__contains']=str(SeparationDate)[:7]

        all = MonthmoneyToOAList.objects.filter(Q(name__contains=code) | Q(code__contains=code)| Q(departmentName__contains=code),
                                                **kwargs).all().values_list()
        for i in all:
            i = list(i)
            try:
                i[3] = JobRank.objects.filter(status=True, id=i[3]).values_list('JobRankName', flat=True)[0]  # 合同归属
            except:
                i[3]=None
            try:
                i[1] = i[1].strftime("%Y-%m")
            except:
                pass
            if i[2] == True:
                i[2] = '在职'
            elif i[2] == False:
                i[2] = '离职'
            i.insert(0, index)
            del i[1]
            i = i[:-6]
            # try:
            #     if i[34] is None:
            #         if i[11] is None:
            #             i[34] = i[33] - i[32] - i[11]
            # except:
            #     pass
            row_data.append(list(i))
            if len(i) == 0:
                index = index
            index += 1
        exc = openpyxl.load_workbook(path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(path)  # 指定路径,保存文件

    else:
        row_data = []
        index = 1
        for id in id_list:
            data = MonthmoneyToOAList.objects.filter(month_money_status=True, id=id).values_list()
            # print(data)

            for i in data:
                i = list(i)
                try:
                    i[3] = JobRank.objects.filter(status=True, id=i[3]).values_list('JobRankName', flat=True)[0]  # 合同归属
                except:
                    i[3] = None
                try:
                    i[1] = i[1].strftime("%Y-%m")
                except:
                    pass
                if i[2] == True:
                    i[2] = '在职'
                elif i[2] == False:
                    i[2] = '离职'
                i.insert(0, index)
                del i[1]
                i = i[:-6]
                # try:
                #     if i[34] is None:
                #         if i[11] is None:
                #             i[34]=i[33]-i[32]-i[11]
                # except:
                #     pass
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1
        exc = openpyxl.load_workbook(path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(path)  # 指定路径,保存文件
    #
    return_data = {
        "code": 200,
        "msg": "下载成功",
        "downloadUrl": path
    }
    return HttpResponse(json.dumps(return_data))


def count_character(s):
    hanzi = 0
    num = 0
    for i in str(s):
        if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
            hanzi = hanzi + 1
        else:
            num += 1
    return str(30 * hanzi + 20 * num)


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)
    else:
        pass


def createExcelPath(file_name, t2, *args):  # is not None
    import openpyxl
    from openpyxl.styles import Alignment
    exc = openpyxl.Workbook()
    sheet = exc.active
    for column in sheet.iter_cols(min_col=0, max_col=50):
        for cell in column:
            sheet.column_dimensions[cell.column_letter].width = 20
    sheet.column_dimensions['A'].width = 10
    sheet.title = file_name.split('.xlsx')[0]
    sheet.merge_cells('A1:AM1')
    sheet['A1'] = '饭卡充值出勤收集表'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.append(args)
    t = time.strftime('%Y-%m-%d')
    path = os.path.join('static', 'hikCanteen', 'download_file', t, t2, file_name)
    path = path.replace(os.sep, '/')
    exc.save(path)
    return path


# def permission_authentication(func):
#     def wrapper(*args, **kw):
#         new_token = CheckToken()
#         check_token = new_token.check_token(request.headers['Authorization'])
#         if check_token != None:
#             print(func.__name__)
#         else:
#             return_data = {
#                 "code": HTTP_403_FORBIDDEN,
#                 "msg": "没有权限访问",
#                 'hidden': False
#             }
#     return wrapper
#
# @permission_authentication(request)


def get_jobrank_hik_Option(request):
    return_data = {
        "code": 200,
        "msg": "下拉菜单返回成功",
        "data": [],
        'hidden': True
    }
    new_token = CheckToken()
    check_token = new_token.check_token(request.headers['Authorization'])
    # print(check_token)
    if check_token != None:

        # jobrankAll = AdminUser.objects.filter(user_jobrank__jobrank_status=True).values_list('user_jobrank__id',
        #                                                                                   'user_jobrank__jobrank_name').all()
        # print(check_token)
        jobrankAll = AdminUser.objects.filter(id=check_token, user_jobrank_hik__status=True).values_list('user_jobrank_hik__id',
                                                                                          'user_jobrank_hik__JobRankName').all()

        # jobrankAll = JobRank.objects.filter(jobrank_status=True).values_list('id', 'JobRankName').all()
        '''
        id=check_token,
        '''
        for i in jobrankAll:
            return_data['data'].append({
                "label": i[1],
                "id": i[0]
            })


    else:
        return_data = {
            "code": HTTP_403_FORBIDDEN,
            "msg": "没有权限访问",
            'hidden': False
        }
    return HttpResponse(json.dumps(return_data))

# def get_amount_to_replenished(request):  #计算应退应补金额
#     try:
#         month='2023-06-01'
#         codeLs=MonthmoneyToOAList.objects.filter(month_money_status=True,balance__isnull=False,month=month).values_list('code',flat=True)
#         for code in codeLs:
#             entryCanteenAmount=MonthmoneyToOAList.objects.filter(month_money_status=True,balance__isnull=False,month=month,code=code).values_list('entryCanteenAmount',flat=True)[0]#统一充值金额
#             butie_sum = MonthmoneyToOAList.objects.filter(month_money_status=True, balance__isnull=False, month=month,code=code).values_list('butie_sum', flat=True)[0]  # 补贴总和
#             balance = MonthmoneyToOAList.objects.filter(month_money_status=True, balance__isnull=False, month=month,code=code).values_list('balance', flat=True)[0]  # 余额
#             if type(balance) is str:
#                 balance=eval(balance)
#             elif balance is None:
#                 balance=0
#             amount_to_replenished=entryCanteenAmount-butie_sum-balance
#             # print(code,entryCanteenAmount,butie_sum,balance,amount_to_replenished)
#             MonthmoneyToOAList.objects.filter(month_money_status=True, balance__isnull=False, month=month,code=code).update(amount_to_replenished=None)
#
#         # print(codeLs)
#         return_data = {
#             "code": HTTP_200_OK,
#             "msg": "计算应退应补金额成功",
#             'hidden': True
#         }
#     except:
#         return_data = {
#             "code": HTTP_401_UNAUTHORIZED,
#             "msg": "未知错误,计算失败！",
#             'hidden': False
#         }
#
#     return HttpResponse(json.dumps(return_data))