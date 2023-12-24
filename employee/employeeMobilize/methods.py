import json

import arrow
from django.db.models import Q
from openpyxl import load_workbook

from utils.genericMethods import methHeader, BasicClass, arrow_format_time, count_width, FileClass, \
    save_file_by_employee, save_xlsx_new
from .sql import *
from utils.sqlServerConnect import EhrConnect
from .models import HrEmployeeMobilize
from employee.models import HrJobRank


class BasicMobilize(methHeader, BasicClass):

    def add_meth(self):
        self.meth['create'] = self.create
        self.meth['search'] = self.search
        self.meth['update'] = self.update
        self.meth['delete'] = self.delete
        self.meth['data_sync'] = self.data_sync
        self.ehr = EhrConnect()

    def data_sync(self):
        sql = transfer()
        results = self.ehr.select(sql)
        bulk_objs = []
        for result in results:
            if result['new_join_date'] is not None and result['new_join_date'] != '':
                result['start_date_of_labor_contract'] = arrow_format_time(result['new_join_date'], 1)
                result['end_date_of_labor_contract'] = arrow.get(result['start_date_of_labor_contract']).shift(
                    years=3).shift(days=-1).format("YYYY-MM-DD")
            result['probation_period'] = judge_period(result['new_job_grade_id'])
            result['basic_salary'] = get_basic_salary(result['employee_id'])
            bulk_objs.append(HrEmployeeMobilize(**result))
        HrEmployeeMobilize.objects.bulk_create(bulk_objs)
        self.return_data['msg'] = '批量创建成功'

    def create(self):
        info = json.loads(self.request.body)
        info['basic_salary'] = get_basic_salary(info['employee_id'])
        HrEmployeeMobilize.objects.create(**info)
        self.return_data['msg'] = '新增成功'

    def search(self):
        info = json.loads(self.request.body)
        if 'currentPage' not in info or 'pageSize' not in info:
            self.return_data['msg'] = '缺少必要参数'
            self.return_data['code'] = 400
            return
        currentPage = int(info['currentPage'])
        pageSize = int(info['pageSize'])
        searchName = info['searchName']
        q_obj = Q()
        q_obj.connector = 'or'
        kwargs = {
            'mobilize_status': True
        }
        # if info['beginDate'] != '' and info['beginDate'] is not None:
        #     kwargs['intern_month__gte'] = arrow.get(info['beginDate']).format("YYYY-MM-01")
        # if info['endDate'] != '' and info['beginDate'] is not None:
        #     kwargs['intern_month__lte'] = arrow.get(info['endDate']).format("YYYY-MM-01")
        if searchName != '':
            q_obj.children.append(('employee__employee_name', searchName))
            q_obj.children.append(('employee__employee_code', searchName))

        columnList = [
            {'label': 'index', 'value': '序号', 'width': 0},
            {'label': 'employee__employee_code', 'value': '工号', 'width': 0},
            {'label': 'employee__employee_name', 'value': '姓名', 'width': 0},
            {'label': 'mobilize_date', 'value': '调职日期', 'width': 0},
            {'label': 'old_job_rank__job_rank_name', 'value': '原合同归属', 'width': 0},
            {'label': 'new_job_rank__job_rank_name', 'value': '新合同归属', 'width': 0},
            {'label': 'old_join_date', 'value': '原入职日期', 'width': 0},
            {'label': 'new_join_date', 'value': '新入职日期', 'width': 0},
            {'label': 'employee__employee_status', 'value': '在职状态', 'width': 0},
            {'label': 'employee__employee_group_join_date', 'value': '集团入职日期', 'width': 0},
            {'label': 'employee__employee_department__department_full_name', 'value': '部门全称', 'width': 0},
            {'label': 'employee__employee_position__position_name', 'value': '岗位', 'width': 0},
            {'label': 'employee__employee_work_place', 'value': '工作地', 'width': 0},
            {'label': 'employee__employee_identity_no', 'value': '身份证号码', 'width': 0},
            {'label': 'employee__employee_sex', 'value': '性别', 'width': 0},
            {'label': 'employee__employee_nation_place__nation_place_name', 'value': '户籍地', 'width': 0},
            {'label': 'employee__employee_nation_address', 'value': '法定送达地址', 'width': 0},
            {'label': 'employee__employee_phone', 'value': '联系方式', 'width': 0},
            {'label': 'term_of_labor_contract', 'value': '合同期限', 'width': 0},
            {'label': 'start_date_of_labor_contract', 'value': '合同开始日期', 'width': 0},
            {'label': 'end_date_of_labor_contract', 'value': '合同结束日期', 'width': 0},
            {'label': 'probation_period', 'value': '试用期(月)', 'width': 0},
            {'label': 'working_hours', 'value': '工时制', 'width': 0},
            {'label': 'basic_salary', 'value': '基本工资', 'width': 0},
            {'label': 'contract_status', 'value': '电子签状态', 'width': 0},
        ]
        field_width = {}
        tableList = []
        employee__employee_status = {
            '1': '在职',
            '2': '离职',
            '99': '黑名单',
            '': '',
            None: '',
        }
        employee__employee_sex = {
            '1': '男',
            '2': '女',
            '': '',
            None: '',
        }
        working_hours = {
            'A': '标准工时',
            'B': '不定时工时',
            'C': '综合工时',
            '': '',
            None: '',
        }
        contract_status = {
            0: '默认',
            1: '已创建、等待激活',
            2: '已激活',
            3: '合同已发起',
            4: '合同已签署',
            '': '默认',
            None: '默认',
        }

        for field in columnList:
            field_width[field['label']] = count_width(field['value'])

        totalNumber = HrEmployeeMobilize.objects.filter(q_obj, **kwargs).count()
        mobilize_objs = HrEmployeeMobilize.objects.filter(q_obj, **kwargs).values(
            'id',
            'employee__employee_code',
            'employee__employee_name',
            'mobilize_date',
            'old_job_rank__job_rank_name',
            'new_job_rank__job_rank_name',
            'old_join_date',
            'new_join_date',
            'employee__employee_status',
            'employee__employee_group_join_date',
            'employee__employee_department__department_full_name',
            'employee__employee_position__position_name',
            'employee__employee_work_place',
            'employee__employee_identity_no',
            'employee__employee_sex',
            'employee__employee_nation_place__nation_place_name',
            'employee__employee_nation_address',
            'employee__employee_phone',
            'term_of_labor_contract',
            'start_date_of_labor_contract',
            'end_date_of_labor_contract',
            'probation_period',
            'working_hours',
            'basic_salary',
            'contract_status',
        )[(currentPage - 1) * pageSize: currentPage * pageSize]
        index = (currentPage - 1) * pageSize + 1
        for obj in mobilize_objs:
            obj['index'] = index
            index += 1
            obj['employee__employee_status'] = employee__employee_status[obj['employee__employee_status']]
            obj['employee__employee_sex'] = employee__employee_sex[obj['employee__employee_sex']]
            obj['working_hours'] = working_hours[obj['working_hours']]
            obj['contract_status'] = contract_status[obj['contract_status']]
            for key, value in obj.items():
                if value is not None and value != '' and key in ['mobilize_date', 'old_join_date', 'old_join_date',
                                                                 'employee__employee_group_join_date']:
                    obj[key] = arrow.get(obj[key]).format("YYYY-MM-DD")
                if key in field_width:
                    field_width[key] = max(field_width[key], count_width(value))
            tableList.append(obj)
        for i in columnList:
            if i['label'] in field_width:
                i['width'] = field_width[i['label']]
        self.return_data['data']['tableList'] = tableList
        self.return_data['data']['columnList'] = columnList
        self.return_data['data']['totalNumber'] = totalNumber

    def update(self):
        info = json.loads(self.request.body)
        pk = info.pop('id')
        working_hours = {
            '不定时工时': 'B',
            '综合工时': 'C',
            '标准工时': 'A',
            None: '',
            '': '',
        }
        kwargs = {
            'term_of_labor_contract': info['term_of_labor_contract'],
            'start_date_of_labor_contract': info['start_date_of_labor_contract'],
            'end_date_of_labor_contract': info['end_date_of_labor_contract'],
            'probation_period': info['probation_period'],
            'working_hours': info['working_hours'],
        }
        if info['working_hours'] in working_hours:
            kwargs['working_hours'] = working_hours[info['working_hours']]
        HrEmployeeMobilize.objects.filter(pk=pk).update(**kwargs)
        self.return_data['msg'] = '修改成功'

    def delete(self):
        info = json.loads(self.request.body)
        idList = info['idList']
        HrEmployeeMobilize.objects.filter(pk__in=idList).update(mobilize_status=False, modify_time=self.now_time,
                                                                modifier_id=self.operate_user_id)
        self.return_data['msg'] = '删除成功'


# 职等5及以上试用期6个月，其他三个月
def judge_period(job_grade_id):
    if job_grade_id in [31, 32, 33, 34, 35, 36, 37, 52, 53, 54, 55, 56, 80, 81]:
        return 6
    else:
        return 3


class MobilizeInfoFileClass(methHeader, FileClass):
    def add_meth(self):
        self.meth['create'] = self.upload
        self.meth['search'] = self.download

    def upload(self):
        file = self.request.FILES.get('file')
        url = save_file_by_employee(['employee', 'employeeMobilize', 'upload'], file, file.name, '普通文件', '调职名单',
                                    self.operate_user_name)
        sheet = load_workbook(url, data_only=True)
        sheet = sheet.active
        field_to_index = {
            'employee__employee_code': 1,
            'mobilize_date': 2,
            'old_job_rank_id': 3,
            'new_job_rank_id': 4,
            'old_join_date': 5,
            'new_join_date': 6,
            'term_of_labor_contract': 7,
            'start_date_of_labor_contract': 8,
            'end_date_of_labor_contract': 9,
            'probation_period': 10,
            'working_hours': 11,
        }
        job_rank_dict = get_job_rank_list()
        job_rank_dict[''] = ''
        job_rank_dict[None] = ''
        job_rank_dict['None'] = ''
        workingHoursChoices = {
            '标准工时': 'A',
            '不定时工时': 'B',
            '综合工时': 'C',
            '': '',
            None: '',
        }
        # for row in sheet.iter_rows(min_row=2):
        for row in range(2, sheet.max_row + 1):

            # 将元组转换为列表
            # row = list(row)
            # 遍历循环获取字段对应的值
            pk = select_employee_id(sheet.cell(row, 1).value)
            if pk == 0:
                self.return_data['code'] = 400
                self.return_data['msg'] = f'{sheet.cell(row, 1).value}该工号人员不存在'
                return
            kwargs = {}
            # row[10] = workingHoursChoices[row[10].value]
            value_is_none_count = 0
            # print(row)
            for field, index in field_to_index.items():
                if sheet.cell(row, index).value is not None:
                    if sheet.cell(row, index).value is not None and sheet.cell(row, index).value != '':
                        kwargs[field] = sheet.cell(row, index).value

                if sheet.cell(row, index).value is None or sheet.cell(row, index).value == '':
                    value_is_none_count += 1
            kwargs.pop('employee__employee_code')
            kwargs['employee_id'] = pk
            kwargs['old_job_rank_id'] = job_rank_dict[kwargs['old_job_rank_id']]
            kwargs['new_job_rank_id'] = job_rank_dict[kwargs['new_job_rank_id']]
            kwargs['working_hours'] = workingHoursChoices[kwargs['working_hours']]
            if value_is_none_count >= 5:
                print("存在5个为空哦")
                continue
            print(kwargs)
            kwargs['basic_salary'] = get_basic_salary(kwargs['employee_id'])
            HrEmployeeMobilize.objects.create(**kwargs)
        self.return_data['msg'] = '上传成功'

    def download(self):
        info = json.loads(self.request.body)
        searchName = info['searchName']
        downloadAll = info['downloadAll']
        q_obj = Q()
        q_obj.connector = 'or'
        kwargs = {
            'mobilize_status': True
        }
        # if info['beginDate'] != '' and info['beginDate'] is not None:
        #     kwargs['intern_month__gte'] = arrow.get(info['beginDate']).format("YYYY-MM-01")
        # if info['endDate'] != '' and info['beginDate'] is not None:
        #     kwargs['intern_month__lte'] = arrow.get(info['endDate']).format("YYYY-MM-01")

        tableList = []
        employee__employee_status = {
            '1': '在职',
            '2': '离职',
            '99': '黑名单',
            '': '',
            None: '',
        }
        employee__employee_sex = {
            '1': '男',
            '2': '女',
            '': '',
            None: '',
        }
        working_hours = {
            'A': '标准工时',
            'B': '不定时工时',
            'C': '综合工时',
            '': '',
            None: '',
        }
        contract_status = {
            0: '默认',
            1: '已创建、等待激活',
            2: '已激活',
            3: '合同已发起',
            4: '合同已签署',
            '': '默认',
            None: '默认',
        }
        if downloadAll:
            kwargs = {
                'pk__in': info['idList']
            }
        else:
            if searchName != '':
                q_obj.children.append(('employee__employee_name', searchName))
                q_obj.children.append(('employee__employee_code', searchName))
        mobilize_objs = HrEmployeeMobilize.objects.filter(q_obj, **kwargs).values(
            # 'id',
            'employee__employee_name',
            'mobilize_date',
            'old_job_rank__job_rank_name',
            'new_job_rank__job_rank_name',
            'old_join_date',
            'new_join_date',
            'employee__employee_status',
            'employee__employee_group_join_date',
            'employee__employee_department__department_full_name',
            'employee__employee_position__position_name',
            'employee__employee_work_place',
            'employee__employee_identity_no',
            'employee__employee_sex',
            'employee__employee_nation_place__nation_place_name',
            'employee__employee_nation_address',
            'employee__employee_phone',
            'term_of_labor_contract',
            'start_date_of_labor_contract',
            'end_date_of_labor_contract',
            'probation_period',
            'working_hours',
            'basic_salary',
            'contract_status',
        )
        index = 1
        for obj in mobilize_objs:

            obj['employee__employee_status'] = employee__employee_status[obj['employee__employee_status']]
            obj['employee__employee_sex'] = employee__employee_sex[obj['employee__employee_sex']]
            obj['working_hours'] = working_hours[obj['working_hours']]
            obj['contract_status'] = contract_status[obj['contract_status']]
            for key, value in obj.items():
                if value is not None and value != '' and key in ['mobilize_date', 'old_join_date', 'old_join_date',
                                                                 'employee__employee_group_join_date']:
                    obj[key] = arrow.get(obj[key]).format("YYYY-MM-DD")
            data = [index, ]
            for key, value in obj.items():
                data.append(value)
            tableList.append(data)
            index += 1
        row2 = [
            '序号',
            '工号',
            '姓名',
            '调职日期',
            '原合同归属',
            '新合同归属',
            '原入职日期',
            '新入职日期',
            '员工在职状态',
            '集团入职日期',
            '部门全称',
            '岗位',
            '工作地',
            '身份证号码',
            '性别',
            '户籍地',
            '法定送达地址',
            '联系方式',
            '劳动合同期限',
            '劳动合同开始日期',
            '劳动合同结束日期',
            '试用期',
            '工时制',
            '基本工资',
        ]
        url = save_xlsx_new('调职数据', self.operate_user_name, '调职数据', row2, tableList, 'download')
        self.return_data['downloadUrl'] = url
        self.return_data['msg'] = '下载成功'


def get_job_rank_list():
    job_rank_obj = HrJobRank.objects.filter(job_rank_status=True).values("id", "job_rank_code")
    return {obj['job_rank_code']: obj['id'] for obj in job_rank_obj}


def select_employee_id(code):
    ehr = EhrConnect()
    sql = f"select ID from T_HR_Employee where Code='{code}'"
    result = ehr.select(sql)
    print(result)
    if len(result) < 0:
        return 0
    else:
        return result[0]['ID']


def get_basic_salary(empID):
    sql = f"""
    SELECT TOP 1 BasePay  
    FROM T_HR_Payroll
    WHERE EmpID = {empID}
    ORDER BY MonthID DESC
    """
    ehr = EhrConnect()
    result = ehr.select(sql)
    return result[0]['BasePay']
