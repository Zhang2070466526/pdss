import json, os,arrow,openpyxl

from django.db.models import Q,F
from rest_framework import status
from django.http import JsonResponse
from django.core import serializers
from auther.models import AdminUser
# from employee.views import get_user_field
from django.db import models
from employee.models import HrEmployee, HrEmployeeFiles
from rest_framework.response import Response
from datetime import datetime, date
from pdss.settings import BASE_DIR
from employee import views


class Roster:
    def __init__(self, request,meth):
        self.request = request
        self.return_data = {
            'code': 200,
            'msg': '信息返回成功'
        }
        self.meth = meth
        self.methods = {
            'get_roster_info': self.get_roster_info,
            'download_roster_info':self.download_roster_info,

        }

    def method_center(self):
        self.return_data = {'code': status.HTTP_403_FORBIDDEN, "msg": '没有权限访问'}
        if self.request.check_token is None:
            return JsonResponse(self.return_data)
        self.methods[self.meth]()
        return JsonResponse(self.return_data)


    # 获取信息列表
    def get_roster_info(self):   ##司龄=当前日期-集团入职日期
        user_field_all=views.get_user_field(self.request,97)
        user_field_all_keys= [list(d.keys())[0] for d in user_field_all]
        # print(user_field_all_keys)
        columnList = [{'label': 'index', 'value': '序号', 'width': 60}]
        for dictionary in user_field_all:
            key = list(dictionary.keys())[0]
            value = list(dictionary.values())[0]
            new_dict = {'label': key, 'value': value, 'width': 100}
            if new_dict['label'] in ['employee_department__department_full_name','employee_identity_no','employee_position__position_name','employee_group_join_date','employee_join_date','employee_department__department_first_name','employee_department__department_second_name','employee_department__department_third_name','employee_position__position_name','employee_job_duty__job_duty_name','employee_job_class__job_class_name']:
                new_dict['width']=210

            columnList.append(new_dict)
        columnList.insert(7, {'label': 'employee_seniority', 'value': '司龄(月)', 'width': ''})

        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'employee_job_rank__in':self.request.user_jobRank_employee
        }
        searchName = self.request.GET.get('searchName', None)

        if "employee_job_rank" in self.request.GET:
            jobRankId = self.request.GET.get('employee_job_rank', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['employee_job_rank'] = jobRankId
        if "employee_job_rank[]" in self.request.GET:
            jobRankId = self.request.GET.getlist('employee_job_rank[]', None)
            if len(jobRankId) != 0:  # 有值
                kwargs['employee_job_rank__in'] = jobRankId

        if "employee_work_status" in self.request.GET:
            employee_work_status = self.request.GET.get('employee_work_status', None)
            if len(employee_work_status) != 0:  # 有值
                kwargs['employee_work_status'] = employee_work_status
        if "employee_work_status[]" in self.request.GET:
            employee_work_status = self.request.GET.getlist('employee_work_status[]', None)
            if len(employee_work_status) != 0:  # 有值
                kwargs['employee_work_status__in'] = employee_work_status

        if "employee_status" in self.request.GET:
            employee_status = self.request.GET.get('employee_status', None)

            if len(employee_status) != 0:  # 有值
                kwargs['employee_status'] = employee_status
        if "employee_status[]" in self.request.GET:
            employee_status = self.request.GET.getlist('employee_status[]', None)
            if len(employee_status) != 0:  # 有值
                kwargs['employee_status__in'] = employee_status
        beginDate = self.request.GET.get('beginDate',None)
        endDate = self.request.GET.get('endDate',None)
        # print(self.request.GET)
        pay_type_name = self.request.GET.get('pay_type_name', None)   #计薪方式
        # print(pay_type_name)

        if pay_type_name is not None or len(pay_type_name)>0:
            if pay_type_name:
                kwargs['employee_pay_type_id'] = int(pay_type_name)
            else:
                pass
        if beginDate != "" and endDate != "":
            kwargs['employee_group_join_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(endDate) == 0 else beginDate
            kwargs['employee_group_join_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(endDate) == 0 else endDate
        # print(searchName)
        print(kwargs)
        totalNumber = HrEmployee.objects.filter(Q(employee_name__contains=searchName) | Q(employee_code__contains= searchName),**kwargs).count()
        tableList = list(HrEmployee.objects.filter(Q(employee_name__contains=searchName) | Q(employee_code__contains=searchName),**kwargs).values(*tuple(user_field_all_keys),'id').order_by('-employee_group_join_date')[(currentPage - 1) * pageSize:currentPage * pageSize])


        employee_status_choices= {'1': '在职','2':'离职','99':'黑名单'}
        employee_sex_choices={'1':'男','2':'女'}
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            try:
                item['employee_seniority']=self.calculate_month(item['employee_group_join_date'])
            except:
                pass
            try:
                item['employee_status']=employee_status_choices.get(item['employee_status'])
            except:
                pass
            try:
                item['employee_group_join_date']=str(item['employee_group_join_date'])[:10] if item['employee_group_join_date']!=None else None
            except:
                pass
            try:
                item['employee_join_date']=str(item['employee_join_date'])[:10] if item['employee_join_date']!=None else None
            except:
                pass
            try:
                item['employee_sex']=employee_sex_choices.get(item['employee_sex'])
            except:
                pass
            try:
                item['employee_birthday']=self.calculate_age(str(item['employee_birthday']))
            except:
                pass
            try:
                item['employee_departure_date']=str(item['employee_departure_date'])[:10] if item['employee_departure_date']!=None else None
            except:
                pass

        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

    def download_roster_info(self):
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.timestamp()
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'download_file', t1, str(t2))  # 创建文件夹
        self.mkdir(dummy_path)
        id_list = json.loads(self.request.body).get('idList')
        downloadAll = json.loads(self.request.body).get('downloadAll')
        user_field_all=views.get_user_field(self.request,97)
        user_field_all_keys = [list(d.keys())[0] for d in user_field_all]
        user_field_all_values= [list(d.values())[0] for d in user_field_all]
        file_ls=user_field_all_values
        file_ls.insert(0,"序号")
        file_ls.insert(7,'司龄(月)')
        path = self.createExcelPath('员工花名册表.xlsx', str(t2), '员工花名册表',len(file_ls),self.excel_range_from_numbers([1,len(file_ls)]), *file_ls)
        employee_status_choices = {'1': '在职', '2': '离职', '99': '黑名单'}
        employee_sex_choices = {'1': '男', '2': '女'}
        if downloadAll == True:  # 是下载全部   有条件
            row_data = []
            index = 1
            kwargs = { 'employee_job_rank__in':self.request.user_jobRank_employee}
            searchName = self.request.GET.get('searchName', None)
            if "employee_job_rank" in self.request.GET:
                jobRankId = self.request.GET.get('employee_job_rank', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['employee_job_rank'] = jobRankId
            if "employee_job_rank[]" in self.request.GET:
                jobRankId = self.request.GET.getlist('employee_job_rank[]', None)
                if len(jobRankId) != 0:  # 有值
                    kwargs['employee_job_rank__in'] = jobRankId

            if "employee_work_status" in self.request.GET:
                employee_work_status = self.request.GET.get('employee_work_status', None)
                if len(employee_work_status) != 0:  # 有值
                    kwargs['employee_work_status'] = employee_work_status
            if "employee_work_status[]" in self.request.GET:
                employee_work_status = self.request.GET.getlist('employee_work_status[]', None)
                if len(employee_work_status) != 0:  # 有值
                    kwargs['employee_work_status__in'] = employee_work_status

            if "employee_status" in self.request.GET:
                employee_status = self.request.GET.get('employee_status', None)
                if len(employee_status) != 0:  # 有值
                    kwargs['employee_status'] = employee_status
            if "employee_status[]" in self.request.GET:
                employee_status = self.request.GET.getlist('employee_status[]', None)
                if len(employee_status) != 0:  # 有值
                    kwargs['employee_status__in'] = employee_status

            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            pay_type_name = self.request.GET.get('pay_type_name', None)  # 计薪方式
            if pay_type_name is not None or len(pay_type_name) > 0:
                if pay_type_name:
                    kwargs['employee_pay_type_id'] = int(pay_type_name)
                else:
                    pass

            if beginDate != "" and endDate != "":
                kwargs['employee_group_join_date__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                                   177) if beginDate is None or len(
                    endDate) == 0 else beginDate
                kwargs['employee_group_join_date__lte'] = datetime(3221, 10, 29, 7, 17, 1,
                                                                   177) if endDate is None or len(
                    endDate) == 0 else endDate

            tableList = HrEmployee.objects.filter(Q(employee_name__contains=searchName) | Q(employee_code__contains=searchName),**kwargs).values(*tuple(user_field_all_keys)).order_by('-employee_group_join_date')

            for line in tableList:
                line_data = []
                try:
                    line['employee_sex']=employee_sex_choices.get(line['employee_sex'])
                except:
                    pass
                try:
                    line['employee_birthday']=self.calculate_age(str(line['employee_birthday']))
                except:
                    pass
                try:
                    line['employee_status']=employee_status_choices.get(line['employee_status'])
                except:
                    pass
                for k, v in line.items():
                    line_data.append(v)
                try:
                    line_data.insert(6, self.calculate_month(line_data[4]))
                except:
                    pass
                line_data.insert(0,index)
                row_data.append(line_data)
                if len(line) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            index = 1
            for id in id_list:
                line_data=[]
                line = list(HrEmployee.objects.filter(pk=id).values(*tuple(user_field_all_keys)))[0]
                try:
                    line['employee_sex']=employee_sex_choices.get(line['employee_sex'])
                except:
                    pass
                try:
                    line['employee_birthday']=self.calculate_age(str(line['employee_birthday']))
                except:
                    pass
                try:
                    line['employee_status']=employee_status_choices.get(line['employee_status'])
                except:
                    pass
                for k, v in line.items():
                    line_data.append(v)
                try:
                    line_data.insert(6, self.calculate_month(line_data[4]))
                except:
                    pass
                line_data.insert(0,index)
                row_data.append(line_data)
                if len(line_data) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": path
        }

    @staticmethod
    def calculate_month(date):
        """
        :param date:  集团入职日期      YYYY-MM-DD HH:mm:ss
        :return: int
        """
        employee_group_join_date = date
        if employee_group_join_date==None:
            return None
        now_data = datetime.now()
        try:
            if type(employee_group_join_date)==str:
                employee_group_join_date = datetime.strptime(employee_group_join_date, '%Y-%m-%d %H:%M:%S')
        except:
            pass
        # print(type(employee_group_join_date),employee_group_join_date)
        months_between = (now_data.year - employee_group_join_date.year) * 12 + (now_data.month - employee_group_join_date.month)
        return months_between

    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass

    @staticmethod
    def createExcelPath(file_name, t2, name, num, interval, *args):  # is not None
        import openpyxl
        from openpyxl.styles import Alignment
        import time
        exc = openpyxl.Workbook()
        sheet = exc.active
        for column in sheet.iter_cols(min_col=0, max_col=num):
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = 22
        sheet.column_dimensions['A'].width = 10
        # sheet.column_dimensions['L'].width = 70

        sheet.title = file_name.split('.xlsx')[0]
        sheet.merge_cells(str(interval))  # 'A1:D1'
        sheet['A1'] = name
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet.append(args)
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'employee', 'download_file', t, t2, file_name)
        path = path.replace(os.sep, '/')
        exc.save(path)
        return path
    @staticmethod
    def excel_range_from_numbers(numbers):
        def number_to_excel_column(n):
            result = ""
            while n > 0:
                n -= 1
                remainder = n % 26
                result = chr(65 + remainder) + result
                n //= 26
            return result

        columns = [number_to_excel_column(num) for num in numbers]
        if len(columns) == 1:
            return columns[0] + '1'
        else:
            return f"{columns[0]}1:{columns[-1]}1"

    # print(excel_range_from_numbers([1, 2]))  # 输出 "A1:B1"
    # print(excel_range_from_numbers([1, 28]))  # 输出 "A1:AB1"

    @staticmethod
    def calculate_age(birthdate):
        # print(birthdate)
        today = datetime.today()
        birthdate = datetime.strptime(birthdate, "%Y-%m-%d")
        age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
        return age






