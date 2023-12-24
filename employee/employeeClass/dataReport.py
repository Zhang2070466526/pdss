import json, arrow, os
from django.db import connection
from django.db.models import Q
from datetime import datetime, date, timedelta
from rest_framework import status
from employee.models import *
from employeePersonnel.models import HrEmployeeHistory
from pdss.settings import BASE_DIR
# from IeProposal.proposal.proposalClass import *
from django.db.models import Count
from utils.sqlServerConnect import EhrConnect
import shutil,openpyxl
import pandas as pd
from utils.save_data_to_redis import *
from threading import Thread

class EmployeeDataReport:
    def __init__(self):
        self.now = arrow.now()
        self.t1 = self.now.format('YYYY-MM-DD')
        self.t2 = self.now.timestamp()
    def limit_get_data(self, department_id, search_date, currentPage, pageSize):
        """
        基础数据查询，编制数据
        :param department_id:部门id列表
        :param search_date: 生效时间
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_peron_limit_department__department_full_name', 'value': '部门', 'width': ""},
            {'label': 'department_peron_limit_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'department_peron_limit_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'department_peron_limit_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'department_peron_limit_no', 'value': '总编制人数', 'width': ""},
            {'label': 'department_peron_limit_core_no', 'value': '关键核心岗位定编', 'width': ""},
            {'label': 'department_peron_limit_effect_date', 'value': '生效时间', 'width': ""}
        ]
        limit_params = {
            'department_peron_limit_effect_date__lte': search_date,  # 生效时间大于等于查询时间
            'department_peron_limit_department_id__in': department_id,
            'department_peron_limit_status': True,
            'department_peron_limit_expire_date__isnull': True
        }
        limit_params = {key: value for key, value in limit_params.items() if
                        value is not None and value != ''}  # 过滤掉值为None或''的项
        total_number = HrDepartmentPersonLimit.objects.filter(**limit_params).count()
        compile_data_list = list(HrDepartmentPersonLimit.objects.filter(**limit_params).values('id',
                                                                                               'department_peron_limit_department_id',
                                                                                               'department_peron_limit_department__department_first_name',
                                                                                               'department_peron_limit_department__department_second_name',
                                                                                               'department_peron_limit_department__department_third_name',
                                                                                               'department_peron_limit_department__department_full_name',
                                                                                               'department_peron_limit_no',
                                                                                               'department_peron_limit_core_no',
                                                                                               'department_peron_limit_effect_date').order_by(
            '-department_peron_limit_effect_date'))[(currentPage - 1) * pageSize:currentPage * pageSize]

        for index, item in enumerate(compile_data_list):
            item['index'] = (currentPage - 1) * pageSize + index + 1

        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': compile_data_list,
                'totalNumber': total_number,
                'sumLabelsList':['department_peron_limit_no','department_peron_limit_core_no']
            }
        }
        return return_data
    def limit_post_data(self, request_data):
        """
        基础数据新增，编制数据
        :param request_data:新增的数据
            {
                 department_peron_limit_department_id:部门
                 department_peron_limit_no:编制人数
                 department_peron_limit_core_no: 关键核心岗位定编
                 department_peron_limit_effect_date: 生效时间
            }
        :return:
        """
        #  失效时间是Null的 或者有日期大于今天的 无法新增 只有生效时间大于等于失效时间可以新增
        # 有失效时间为空的，但是又新增一条是更新还是不新增

        # 如果第一条是只有生效时间 第二条去找只有生效时间的，将第二条的生效时间赋值给第一条的失效时间，那第二次新增的要创建

        limit_obj = HrDepartmentPersonLimit.objects.filter(
            department_peron_limit_department_id=request_data['department_peron_limit_department_id'],
            department_peron_limit_expire_date__isnull=True).order_by(
            '-department_peron_limit_effect_date').first()  # 失效时间是空的
        if limit_obj:  # 存在失效时间是Null的
            date_obj = datetime.strptime(request_data['department_peron_limit_effect_date'], "%Y-%m-%d")
            previous_day = date_obj - timedelta(days=1)
            previous_day_str = previous_day.strftime("%Y-%m-%d")  # 日期的前一天
            HrDepartmentPersonLimit.objects.filter(id=limit_obj.id).update(
                department_peron_limit_expire_date=previous_day_str,department_peron_limit_status=False)
        HrDepartmentPersonLimit.objects.create(**request_data)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "新增成功",
        }
        return return_data
    def limit_batch_data(self, request_file):
        """
        :param request_file: 上传的文件对象
        :return:
        """

        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.format('YYYY-MM-DD_HH_mm_ss')
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_upload_file', t1, '编制数据文件上传')  # 创建文件夹
        self.mkdir(dummy_path)
        file_url, file_name, file_suffix = self.createPath(request_file, '编制数据文件上传',
                                                                    '编制报表' + str(t2), 'employee')
        self.saveFile(file_url, request_file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        for line in range(1, sheet.max_row):  # 每行数据
            limit = {}
            if sheet.cell(line + 1, 2).value == None or sheet.cell(line + 1, 3).value == None or sheet.cell(line + 1,
                                                                                                            4).value == None or sheet.cell(
                    line + 1, 5).value == None:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据缺失，无法上传"
                }
            if sheet.cell(line + 1, 2).value != '' or sheet.cell(line + 1, 2).value is not None:
                dept_obj = HrDepartment.objects.filter(~Q(id=999999), Q(department_expiry_date__isnull=True) | Q(
                    department_expiry_date__gt=datetime.now()),
                                                       department_name__isnull=False, department_status=1,
                                                       department_name=sheet.cell(line + 1, 2).value)
                if dept_obj.exists():
                    limit['department_peron_limit_department_id'] = dept_obj[0].id
                else:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "部门名称错误或已过期，无法上传"
                    }
            else:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "部门未填写，无法上传"
                }
            if type(sheet.cell(line + 1, 3).value) == int or type(sheet.cell(line + 1, 3).value) == float:
                limit['department_peron_limit_no'] = int(sheet.cell(line + 1, 3).value)
            else:
                try:
                    limit['department_peron_limit_no'] = int(sheet.cell(line + 1, 3).value)
                except:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "编制人数必须是整数，无法上传"
                    }
            if type(sheet.cell(line + 1, 4).value) == int or type(sheet.cell(line + 1, 4).value) == float:
                limit['department_peron_limit_core_no'] = int(sheet.cell(line + 1, 4).value)
            else:
                try:
                    limit['department_peron_limit_core_no'] = int(sheet.cell(line + 1, 4).value)
                except:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "关键核心岗位定编必须是整数,无法上传"
                    }
            if type(sheet.cell(line + 1, 5).value) == datetime or type(sheet.cell(line + 1, 5).value) == date:
                limit['department_peron_limit_effect_date'] = sheet.cell(line + 1, 5).value.date()
            elif type(sheet.cell(line + 1, 5).value) == str and len(sheet.cell(line + 1, 5).value) == 10:
                limit['department_peron_limit_effect_date'] = datetime.strptime(sheet.cell(line + 1, 5).value,
                                                                                "%Y-%m-%d")
            elif type(sheet.cell(line + 1, 5).value) == str and len(sheet.cell(line + 1, 5).value) == 19:
                limit['department_peron_limit_effect_date'] = datetime.strptime(sheet.cell(line + 1, 5).value,
                                                                                "%Y-%m-%d %H:%M:%S").date()
            else:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "生效时间必须是日期,例如2023-01-01,无法上传"
                }

            # 如果第一条是只有生效时间 第二条去找只有生效时间的，将第二条的生效时间赋值给第一条的失效时间，那第二次新增的要创建
            limit_obj = HrDepartmentPersonLimit.objects.filter(
                department_peron_limit_department_id=limit['department_peron_limit_department_id'],
                department_peron_limit_expire_date__isnull=True).order_by(
                '-department_peron_limit_effect_date').first()  # 失效时间是空的
            if limit_obj:  # 存在失效时间是Null的
                previous_day = limit['department_peron_limit_effect_date'] - timedelta(days=1)
                previous_day_str = previous_day.strftime("%Y-%m-%d")  # 日期的前一天
                HrDepartmentPersonLimit.objects.filter(id=limit_obj.id).update(
                    department_peron_limit_expire_date=previous_day_str,department_peron_limit_status=False)
            HrDepartmentPersonLimit.objects.create(**limit)

        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "上传成功!"
        }
        return return_data
    def limit_edit_data(self, request_data):
        """
        基础数据修改，编制数据
        :param request_data:修改的数据
            {
                 id:要修改的数据的id
                 department_peron_limit_department_id:部门
                 department_peron_limit_no:编制人数
                 department_peron_limit_core_no: 关键核心岗位定编
                 department_peron_limit_effect_date: 生效时间
            }
        :return:
        """
        HrDepartmentPersonLimit.objects.filter(pk=request_data['id']).update(**request_data)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }
        return return_data
    def limit_down_data(self,request):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file','编制数据下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),'编制数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            row_data = []
            index = 1

            info = json.loads(request.body)
            search_date=info['search_date']
            if search_date is None or search_date == '':
                search_date = '3000-01-01'
            department_id=info['department_id']
            if len(department_id) == 0:
                department_id = request.user_department_employee
            limit_params = {
                'department_peron_limit_effect_date__lte': search_date,  # 生效时间大于等于查询时间
                'department_peron_limit_department_id__in': department_id,
                'department_peron_limit_status': True,
                'department_peron_limit_expire_date__isnull': True
            }
            limit_params = {key: value for key, value in limit_params.items() if
                            value is not None and value != ''}  # 过滤掉值为None或''的项
            compile_data_list = list(HrDepartmentPersonLimit.objects.filter(**limit_params).values('id',
                                                                                                   'department_peron_limit_department__department_name',
                                                                                                   'department_peron_limit_department__department_first_name',
                                                                                                   'department_peron_limit_department__department_second_name',
                                                                                                   'department_peron_limit_department__department_third_name',
                                                                                                   'department_peron_limit_no',
                                                                                                   'department_peron_limit_core_no',
                                                                                                   'department_peron_limit_effect_date').order_by('-department_peron_limit_effect_date'))



            for line in compile_data_list:
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data[0] = index
                index += 1
                row_data.append(line_data)
        else:
            row_data = []
            index = 1
            for id in id_list:
                line_data = list(HrDepartmentPersonLimit.objects.filter(pk=id, department_peron_limit_status=True,department_peron_limit_expire_date__isnull=True).values_list(
                    'department_peron_limit_department__department_name',
                    'department_peron_limit_department__department_first_name',
                    'department_peron_limit_department__department_second_name',
                    'department_peron_limit_department__department_third_name',
                    'department_peron_limit_no',
                    'department_peron_limit_core_no',
                    'department_peron_limit_effect_date'
                ))[0]
                line_data = (index,) + line_data
                line_data = list(line_data)
                row_data.append(line_data)
                index+=1
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data

    def target_get_data(self, department_id, search_date, currentPage, pageSize):
        """
        基础数据查询，离职率目标
        :param department_id:部门id列表
        :param search_date:生效日期
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_turn_over_target_department__department_full_name', 'value': '部门', 'width': ""},
            {'label': 'department_turn_over_target_department__department_first_name', 'value': '一级部门',
             'width': 230},
            {'label': 'department_turn_over_target_department__department_second_name', 'value': '二级部门',
             'width': 230},
            {'label': 'department_turn_over_target_department__department_third_name', 'value': '三级部门',
             'width': 230},
            {'label': 'department_turn_over_target', 'value': '离职率目标', 'width': ""},
            {'label': 'department_turn_over_target_effect_date', 'value': '生效时间', 'width': ""}
        ]
        target_params = {
            'department_turn_over_target_effect_date__lte': search_date,
            'department_turn_over_target_department_id__in': department_id,
            'department_turn_over_target_status': True,
            'department_turn_over_target_expire_date__isnull': True
        }
        target_params = {key: value for key, value in target_params.items() if
                         value is not None and value != ''}  # 过滤掉值为None或''的项
        total_number = HrDepartmentTurnOverTarget.objects.filter(**target_params).count()
        compile_data_list = list(HrDepartmentTurnOverTarget.objects.filter(**target_params).values('id',
                                                                                                   'department_turn_over_target_department_id',
                                                                                                   'department_turn_over_target_department__department_full_name',
                                                                                                   'department_turn_over_target_department__department_first_name',
                                                                                                   'department_turn_over_target_department__department_second_name',
                                                                                                   'department_turn_over_target_department__department_third_name',
                                                                                                   'department_turn_over_target',
                                                                                                   'department_turn_over_target_effect_date').order_by(
            '-department_turn_over_target_effect_date'))[(currentPage - 1) * pageSize:currentPage * pageSize]
        for index, item in enumerate(compile_data_list):
            item['index'] = (currentPage - 1) * pageSize + index + 1
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': compile_data_list,
                'totalNumber': total_number,
                'sumLabelsList': ['department_turn_over_target']
            }
        }
        return return_data
    def target_post_data(self, request_data):
        """
        基础数据新增，部门离职率目标数据
        :param request_data:新增的数据
            {
                 department_turn_over_target_department_id:部门
                 department_turn_over_target:离职率目标
                 department_turn_over_target_effect_date: 生效时间
            }
        :return:
        """
        target_obj = HrDepartmentTurnOverTarget.objects.filter(
            department_turn_over_target_department_id=request_data['department_turn_over_target_department_id'],
            department_turn_over_target_expire_date__isnull=True).order_by(
            '-department_turn_over_target_effect_date').first()  # 失效时间是空的
        if target_obj:  # 存在失效时间是Null的
            date_obj = datetime.strptime(request_data['department_turn_over_target_effect_date'], "%Y-%m-%d")
            previous_day = date_obj - timedelta(days=1)
            previous_day_str = previous_day.strftime("%Y-%m-%d")  # 日期的前一天
            HrDepartmentTurnOverTarget.objects.filter(id=target_obj.id).update(
                department_turn_over_target_expire_date=previous_day_str,department_turn_over_target_status=False)
        HrDepartmentTurnOverTarget.objects.create(**request_data)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "新增成功",
        }
        return return_data
    def target_batch_data(self, request_file):
        """
        :param request_file: 上传的文件对象
        :return:
        """
        now = arrow.now()
        t1 = now.format('YYYY-MM-DD')
        t2 = now.format('YYYY-MM-DD_HH_mm_ss')
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_upload_file', t1,
                                  '离职率目标数据文件上传')  # 创建文件夹
        self.mkdir(dummy_path)
        file_url, file_name, file_suffix = self.createPath(request_file, '离职率目标数据文件上传',
                                                                    '离职率目标数据报表' + str(t2), 'employee')
        self.saveFile(file_url, request_file)
        exc = openpyxl.load_workbook(file_url, data_only=True)
        sheet = exc.active
        for line in range(1, sheet.max_row):  # 每行数据
            target = {}
            if sheet.cell(line + 1, 2).value == None or sheet.cell(line + 1, 3).value == None or sheet.cell(line + 1,
                                                                                                            4).value == None:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据缺失，无法上传"
                }
            if sheet.cell(line + 1, 2).value != '' or sheet.cell(line + 1, 2).value is not None:
                dept_obj = HrDepartment.objects.filter(~Q(id=999999), Q(department_expiry_date__isnull=True) | Q(
                    department_expiry_date__gt=datetime.now()),
                                                       department_name__isnull=False, department_status=1,
                                                       department_name=sheet.cell(line + 1, 2).value)
                if dept_obj.exists():
                    target['department_turn_over_target_department_id'] = dept_obj[0].id
                else:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "部门名称错误或已过期，无法上传"
                    }
            else:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "部门未填写，无法上传"
                }
            if type(sheet.cell(line + 1, 3).value) == int or type(sheet.cell(line + 1, 3).value) == float:
                if sheet.cell(line + 1, 3).value >= 0:
                    target['department_turn_over_target'] = sheet.cell(line + 1, 3).value
                else:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "离职率目标大于等于0，无法上传"
                    }

            else:
                try:
                    target['department_turn_over_target'] = float(sheet.cell(line + 1, 3).value)
                except:
                    return {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "离职率目标必须是数值，无法上传"
                    }

            if type(sheet.cell(line + 1, 4).value) == datetime or type(sheet.cell(line + 1, 4).value) == date:
                target['department_turn_over_target_effect_date'] = sheet.cell(line + 1, 4).value.date()
            elif type(sheet.cell(line + 1, 4).value) == str and len(sheet.cell(line + 1, 4).value) == 10:
                target['department_turn_over_target_effect_date'] = datetime.strptime(sheet.cell(line + 1, 5).value,
                                                                                      "%Y-%m-%d")
            elif type(sheet.cell(line + 1, 4).value) == str and len(sheet.cell(line + 1, 4).value) == 19:
                target['department_turn_over_target_effect_date'] = datetime.strptime(sheet.cell(line + 1, 5).value,
                                                                                      "%Y-%m-%d %H:%M:%S").date()
            else:
                return {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "生效时间必须是日期,例如2023-01-01,无法上传"
                }
            target_obj = HrDepartmentTurnOverTarget.objects.filter(
                department_turn_over_target_department_id=target['department_turn_over_target_department_id'],
                department_turn_over_target_expire_date__isnull=True).order_by(
                '-department_turn_over_target_effect_date').first()  # 失效时间是空的
            if target_obj:  # 存在失效时间是Null的
                previous_day = target['department_turn_over_target_effect_date'] - timedelta(days=1)
                previous_day_str = previous_day.strftime("%Y-%m-%d")  # 日期的前一天
                HrDepartmentTurnOverTarget.objects.filter(id=target_obj.id).update(
                    department_turn_over_target_expire_date=previous_day_str,department_turn_over_target_status=False)
            HrDepartmentTurnOverTarget.objects.create(**target)

        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "上传成功!"
        }
        return return_data
    def target_edit_data(self, request_data):
        """
        基础数据修改，部门离职率目标
        :param request_data:修改的数据
            {
                 id:要修改的数据的id
                 department_turn_over_target_department_id:部门
                 department_turn_over_target:离职率目标
                 department_turn_over_target_effect_date: 生效时间
            }
        :return:
        """
        HrDepartmentTurnOverTarget.objects.filter(pk=request_data['id']).update(**request_data)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功",
        }
        return return_data
    def target_down_data(self,request):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file','离职率目标数据下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),'离职率目标数据下载模板.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            row_data = []
            index = 1

            # department_id = []
            #
            # search_date=request.GET.get('search_date')
            # if "department_id" in request.GET:
            #     department_id = request.GET.get('department_id', None)
            #     if len(department_id) == 0:
            #         department_id = request.user_department_employee
            # if "department_id[]" in request.GET:
            #     department_id = request.GET.getlist('department_id[]', None)
            info = json.loads(request.body)
            search_date=info['search_date']
            if search_date is None or search_date == '':
                search_date = '3000-01-01'
            department_id=info['department_id']
            if len(department_id) == 0:
                department_id = request.user_department_employee
            target_params = {
                'department_turn_over_target_effect_date__lte': search_date,
                'department_turn_over_target_department_id__in': department_id,
                'department_turn_over_target_status': True,
                'department_turn_over_target_expire_date__isnull': True
            }
            target_params = {key: value for key, value in target_params.items() if
                             value is not None and value != ''}  # 过滤掉值为None或''的项
            compile_data_list = list(HrDepartmentTurnOverTarget.objects.filter(**target_params).values('id',
                                                                                                       'department_turn_over_target_department__department_full_name',
                                                                                                       'department_turn_over_target_department__department_first_name',
                                                                                                       'department_turn_over_target_department__department_second_name',
                                                                                                       'department_turn_over_target_department__department_third_name',
                                                                                                       'department_turn_over_target',
                                                                                                       'department_turn_over_target_effect_date').order_by('-department_turn_over_target_effect_date'))

            for line in compile_data_list:
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data[0] = index
                index += 1
                row_data.append(line_data)
        else:
            row_data = []
            index = 1
            for id in id_list:
                line_data = list(HrDepartmentTurnOverTarget.objects.filter(pk=id, department_turn_over_target_status=True,department_turn_over_target_expire_date__isnull=True).values_list(
                    'department_turn_over_target_department__department_name',
                    'department_turn_over_target_department__department_first_name',
                    'department_turn_over_target_department__department_second_name',
                    'department_turn_over_target_department__department_third_name',
                    'department_turn_over_target',
                    'department_turn_over_target_effect_date'
                ))[0]
                line_data = (index,) + line_data
                line_data = list(line_data)
                row_data.append(line_data)
                index+=1
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data

    def basic_data(self, department_id, effect_date, currentPage, pageSize):
        """
        基础数据查询，包含编制数和离职率目标
        :param department_id:部门id列表
        :param effect_date:生效日期
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_peron_limit_department__department_full_name', 'value': '部门', 'width': ""},
            {'label': 'department_peron_limit_department__department_first_name', 'value': '一级部门', 'width': ""},
            {'label': 'department_peron_limit_department__department_second_name', 'value': '二级部门', 'width': ""},
            {'label': 'department_peron_limit_department__department_third_name', 'value': '三级部门', 'width': ""},
            {'label': 'department_peron_limit_no', 'value': '编制人数', 'width': ""},
            {'label': 'department_peron_limit_core_no', 'value': '关键核心岗位定编', 'width': ""},
            {'label': 'department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target',
             'value': '离职率目标', 'width': ""}
        ]
        if effect_date is None or effect_date == '':
            effect_date = '3000-01-01'
        basic_params = {
            'department_peron_limit_effect_date__lte': effect_date,
            'department_peron_limit_department_id__in': department_id,
            'department_peron_limit_status': True
        }

        from employee.models import HrDepartmentPersonLimit
        result = list(HrDepartmentPersonLimit.objects.filter(
            Q(department_peron_limit_expire_date=None) &
            Q(department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target_expire_date=None),
            **basic_params
        ).values(
            'id',
            'department_peron_limit_department__department_full_name',
            'department_peron_limit_department__department_first_name',
            'department_peron_limit_department__department_second_name',
            'department_peron_limit_department__department_third_name',
            'department_peron_limit_no',
            'department_peron_limit_core_no',
            'department_peron_limit_effect_date',
            'department_peron_limit_expire_date',
            'department_peron_limit_status',
            'department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target'
        ))[(currentPage - 1) * pageSize:currentPage * pageSize]


        total_number=HrDepartmentPersonLimit.objects.filter(
            Q(department_peron_limit_expire_date=None) &
            Q(department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target_expire_date=None),
            **basic_params
        ).count()
        for index, item in enumerate(result):
            item['index'] = (currentPage - 1) * pageSize + index + 1

        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': result,
                'totalNumber': total_number,
                'sumLabelsList': ['department_peron_limit_no','department_peron_limit_core_no','department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target']
            }
        }
        return return_data

    def basic_down_data(self,request):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1,str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '基础数据下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '基础数据下载模板.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        from employee.models import HrDepartmentPersonLimit
        if download_all == True:  # 是下载全部   有条件
            row_data = []
            index = 1

            info = json.loads(request.body)
            effect_date = info['effect_date']
            if effect_date is None or effect_date == '':
                effect_date = '3000-01-01'
            department_id=info['department_id']
            if len(department_id) == 0:
                department_id = request.user_department_employee
            basic_params = {
                'department_peron_limit_effect_date__lte': effect_date,
                'department_peron_limit_department_id__in': department_id,
                'department_peron_limit_status': True
            }
            result = list(HrDepartmentPersonLimit.objects.filter(
                Q(department_peron_limit_expire_date=None) &
                Q(department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target_expire_date=None),
                **basic_params
            ).values(
                'id',
                'department_peron_limit_department__department_full_name',
                'department_peron_limit_department__department_first_name',
                'department_peron_limit_department__department_second_name',
                'department_peron_limit_department__department_third_name',
                'department_peron_limit_no',
                'department_peron_limit_core_no',
                'department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target'
            ))
            for line in result:
                line_data = []
                for k, v in line.items():
                    line_data.append(v)
                line_data[0] = index
                index += 1
                row_data.append(line_data)
        else:
            row_data = []
            index = 1
            for id in id_list:
                line_data = list(HrDepartmentPersonLimit.objects.filter(
                    Q(department_peron_limit_expire_date=None) &
                    Q(department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target_expire_date=None), pk=id,department_peron_limit_status=True,
                ).values_list(
                    'department_peron_limit_department__department_full_name',
                    'department_peron_limit_department__department_first_name',
                    'department_peron_limit_department__department_second_name',
                    'department_peron_limit_department__department_third_name',
                    'department_peron_limit_no',
                    'department_peron_limit_core_no',
                    'department_peron_limit_department__hrdepartmentturnovertarget__department_turn_over_target'
                ))[0]
                line_data = (index,) + line_data
                line_data = list(line_data)
                row_data.append(line_data)
                index+=1
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data

    def active_employee_total(self, department_id, begin_date, end_date, job_sequence, job_class):
        """
        在职人力情况-汇总
        :param department_id: 部门id列表
        :param begin_date: 开始日期
        :param end_date: 结束日期
        :param job_grade: 职级
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'begin_incumbency', 'value': '开始时间在职人数', 'width': 130},
            {'label': 'duration_onboarding', 'value': '期间入职人数', 'width': 130},
            {'label': 'duration_dimission', 'value': '期间离职人数', 'width': 130},
            {'label': 'turnover_rate', 'value': '离职率', 'width': 130},
            {'label': 'duration_tune', 'value': '期间调入人数', 'width': 130},
            {'label': 'duration_pull', 'value': '期间调出人数', 'width': 130},
            {'label': 'end_incumbency', 'value': '结束时间在职人数（含追光者）', 'width': 130},
            {'label': 'manage_trainees_incumbency', 'value': '管培生在职人数', 'width': 130},
            {'label': 'chaser_light_incumbency', 'value': '追光者在职人数', 'width': 130},
            {'label': 'Key_core_incumbency', 'value': '关键核心人员在岗人数', 'width': 130},
            {'label': 'SAL_incumbency', 'value': 'SAL在职人数', 'width': 130},
            {'label': 'IDL_incumbency', 'value': 'IDL在职人数', 'width': 130},
            {'label': 'DL_incumbency', 'value': 'DL在职人数', 'width': 130},
            {'label': 'total_rate', 'value': '总在岗率（含追光者）', 'width': 130},
            {'label': 'Key_core_rate', 'value': '关键核心人员在岗率', 'width': 130},
            # {'label': 'missing_weave', 'value': '缺编', 'width': 130}
        ]
        table_list = []

        if begin_date:
            begin_date = datetime.strptime(begin_date, "%Y-%m-%d %H:%M:%S")
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")

        if begin_date is None or begin_date == '':
            begin_date = datetime.strptime('2002-11-11 00:00:00', "%Y-%m-%d %H:%M:%S")
        if end_date is None or end_date == '':
            end_date = datetime.strptime('3000-09-01 00:00:00', "%Y-%m-%d %H:%M:%S")
        total_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
        }
        total_params = {key: value for key, value in total_params.items() if
                        value is not None and value != ''}  # 过滤掉值为None或''的项
        # print(total_params,begin_date,end_date)
        employee_total_list = HrEmployee.objects.filter(employee_group_join_date__range=(begin_date, end_date),
                                                        **total_params).values('employee_department__department_full_name',
                                                                               'employee_department__department_full_code',
                                                                               'employee_group_join_date',
                                                                               'employee_dl', 'employee_status',
                                                                               'employee_dim_reason',
                                                                               'employee_job_grade_id',
                                                                               'employee_departure_handle_date',
                                                                               'employee_department__department_first_name',
                                                                               'employee_department__department_second_name',
                                                                               'employee_department__department_third_name')

        employee_outside_begin_list = HrEmployee.objects.filter(
            employee_group_join_date__lte=begin_date, employee_status='1', **total_params).values(
            'employee_department__department_full_name').annotate(department_employee_count=Count('id'))  # 集团入职日期小于开始日期的在职人数
        employee_outside_end_list = HrEmployee.objects.filter(
            employee_group_join_date__lte=end_date, employee_status='1', **total_params).values(
            'employee_department__department_full_name').annotate(department_employee_count=Count('id'))  # 集团入职日期小于结束日期的在职人数
        limit_params = {
            'department_peron_limit_department_id__in': department_id,
        }
        limit_params = {key: value for key, value in limit_params.items() if
                        value is not None and value != ''}  # 过滤掉值为None或''的项
        person_limit_list = HrDepartmentPersonLimit.objects.filter(department_peron_limit_expire_date__isnull=True,
                                                                   department_peron_limit_effect_date__gte=begin_date,
                                                                   **limit_params).values(
            'department_peron_limit_department__department_full_name', 'department_peron_limit_no',
            'department_peron_limit_core_no')  # 编制

        department_id_tuple = tuple(department_id)
        if len(department_id_tuple) == 1:
            department_id_tuple = "(" + str(department_id_tuple[0]) + ")"

        job_class_tuple = tuple(int(x) for x in job_class)
        if len(job_class_tuple) == 1:
            job_class_tuple = "(" + str(job_class_tuple[0]) + ")"

        job_sequence_tuple = tuple(int(x) for x in job_sequence)
        if len(job_sequence_tuple) == 1:
            job_sequence_tuple = "(" + str(job_sequence_tuple[0]) + ")"

        # print(search_date)
        select_sql = """
            SELECT
                D.FullName AS department_full_name,
                SUM ( CASE WHEN T.NewDeptID = D.ID THEN 1 ELSE 0 END ) AS duration_tune,--调入
                SUM ( CASE WHEN T.OldDeptID = D.ID THEN 1 ELSE 0 END ) AS duration_pull --调出
            FROM
                T_HR_Transfer AS T
                JOIN T_HR_Department AS D ON T.NewDeptID = D.ID OR T.OldDeptID = D.ID
            WHERE
                T.TransferDate between '{}' and '{}'
                AND T.NewDeptID IN {}
                AND T.NewJobClassID IN {}
                AND T.NewJobSequenceID IN {}
            GROUP BY
                D.FullName;
        """.format(begin_date, end_date, department_id_tuple, job_class_tuple, job_sequence_tuple)
        if job_class is None or len(job_class_tuple) == 0:
            select_sql = select_sql.replace("AND T.NewJobClassID IN ()", "")
        if job_sequence is None or len(job_sequence_tuple) == 0:
            select_sql = select_sql.replace("AND T.NewJobSequenceID IN ()", "")
        if len(department_id_tuple) == 0:
            select_sql = select_sql.replace("AND T.NewDeptID IN ()", "")

        tune_pull_list = EhrConnect().select(select_sql)  # 期间调入调出

        # 总对象
        employee_total_obj = {}

        # 分布对象
        employee_obj = {
            'index': '',
            'department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            'begin_incumbency': 0,  # 开始时间在职人数
            'duration_onboarding': 0,  # 期间入职人数
            'duration_dimission': 0,  # 期间离职人数
            'turnover_rate': 0,  # 离职率
            'duration_tune': 0,  # 期间调入人数
            'duration_pull': 0,  # 期间调出人数
            'end_incumbency': 0,  # 结束时间在职人数（含追光者）
            'manage_trainees_incumbency': 0,  # 管培生在职人数
            'chaser_light_incumbency': 0,  # 追光者在职人数
            'Key_core_incumbency': 0,  # 关键核心人员在岗人数
            'SAL_incumbency': 0,  # SAL在职人数
            'IDL_incumbency': 0,  # IDL在职人数
            'DL_incumbency': 0,  # DL在职人数
            'total_rate': 0,  # 总在岗率（含追光者）
            'Key_core_rate': 0,  # 关键核心人员在岗率
            'missing_weave': 0,  # 缺编
            'fixed_weave': 0,  # 定编
            'key_core_weave': 0,  # 关键核心岗位定编
            'turnover_total_person': 0,  # 期间离职总人数
        }

        if employee_total_list:
            for employee_data in employee_total_list:
                department_code = employee_data['employee_department__department_full_code']
                if department_code not in employee_total_obj:
                    employee_total_obj[department_code] = employee_obj.copy()
                    employee_total_obj[department_code]['department_full_name'] = employee_data[
                        'employee_department__department_full_name']
                    employee_total_obj[department_code]['employee_department__department_first_name'] = employee_data[
                        'employee_department__department_first_name']
                    employee_total_obj[department_code]['employee_department__department_second_name'] = employee_data[
                        'employee_department__department_second_name']
                    employee_total_obj[department_code]['employee_department__department_third_name'] = employee_data[
                        'employee_department__department_third_name']

                if employee_data['employee_status'] == '1' and employee_data['employee_group_join_date'] >= begin_date and employee_data[
                    'employee_group_join_date'] <= end_date:  # 期间入职人数
                    employee_total_obj[employee_data['employee_department__department_full_code']]['duration_onboarding'] += 1
                if employee_data['employee_status'] == '2' and employee_data[
                    'employee_departure_handle_date'] is not None and employee_data[
                    'employee_departure_handle_date'] >= begin_date and employee_data[
                    'employee_departure_handle_date'] <= end_date:  # 期间离职人数
                    if employee_data['employee_dim_reason'] not in [12, 16, 17]:  # 辞退 体检 考试
                        employee_total_obj[employee_data['employee_department__department_full_code']][
                            'duration_dimission'] += 1
                    employee_total_obj[employee_data['employee_department__department_full_code']][
                        'turnover_total_person'] += 1  # 期间离职总人数
                if employee_data['employee_job_grade_id'] == 80 and employee_data['employee_status'] == '1':  # 管培生在职人数
                    employee_total_obj[employee_data['employee_department__department_full_code']][
                        'manage_trainees_incumbency'] += 1
                if employee_data['employee_job_grade_id'] == 79 and employee_data['employee_status'] == '1':  # 追光者在职人数
                    employee_total_obj[employee_data['employee_department__department_full_code']][
                        'chaser_light_incumbency'] += 1
                if (employee_data['employee_job_grade_id'] in range(6, 25) or employee_data[
                    'employee_job_grade_id'] in range(31, 40) or employee_data['employee_job_grade_id'] in range(52,
                                                                                                                 59)) and \
                        employee_data['employee_status'] == '1':  # 关键核心岗位在岗人数
                    employee_total_obj[employee_data['employee_department__department_full_code']][
                        'Key_core_incumbency'] += 1
                if employee_data['employee_dl'] == 'SAL' and employee_data['employee_status'] == '1':  # SAL在岗人数
                    employee_total_obj[employee_data['employee_department__department_full_code']]['SAL_incumbency'] += 1
                if employee_data['employee_dl'] == 'IDL' and employee_data['employee_status'] == '1':  # IDL在岗人数
                    employee_total_obj[employee_data['employee_department__department_full_code']]['IDL_incumbency'] += 1
                if employee_data['employee_dl'] == 'DL' and employee_data['employee_status'] == '1':  # DL在岗人数
                    employee_total_obj[employee_data['employee_department__department_full_code']]['DL_incumbency'] += 1
            for total_key, total_value in employee_total_obj.items():
                begin_result = [item for item in employee_outside_begin_list if
                                item['employee_department__department_full_name'] == total_value[
                                    'department_full_name']]  # 使用列表推导式筛选出符合条件的键值对
                if len(begin_result) == 0:
                    pass
                else:
                    total_value['begin_incumbency'] = begin_result[0]['department_employee_count']
                end_result = [item for item in employee_outside_end_list if
                              item['employee_department__department_full_name'] == total_value[
                                  'department_full_name']]  # 使用列表推导式筛选出符合条件的键值对
                if len(end_result) == 0:
                    pass
                else:
                    total_value['end_incumbency'] = end_result[0]['department_employee_count']

                waeve_result = [item for item in person_limit_list if
                                item['department_peron_limit_department__department_full_name'] == total_value[
                                    'department_full_name']]  # 使用列表推导式筛选出符合条件的键值对
                if len(waeve_result) == 0:
                    pass
                else:
                    total_value['key_core_weave'] = waeve_result[0]['department_peron_limit_core_no']
                    total_value['fixed_weave'] = waeve_result[0]['department_peron_limit_no']

                tune_pull_result = [item for item in tune_pull_list if
                                    item['department_full_name'] == total_value['department_full_name']]  # 计算调入调出
                if len(tune_pull_result) == 0:
                    pass
                else:
                    total_value['duration_tune'] = tune_pull_result[0]['duration_tune']
                    total_value['duration_pull'] = tune_pull_result[0]['duration_pull']

                if total_value['begin_incumbency'] + total_value['duration_onboarding'] != 0:  # 离职率
                    total_value['turnover_rate'] = total_value['duration_dimission'] / (
                                total_value['begin_incumbency'] + total_value[
                            'duration_onboarding'])  # 离职率= 期间离职人数/（起初人数+期间入职人数）

                if total_value['fixed_weave'] != 0:  # 总在岗率
                    total_value['total_rate'] = total_value['end_incumbency'] / total_value['fixed_weave']

                if total_value['fixed_weave'] != 0:  # 关键核心人员在岗率
                    total_value['Key_core_rate'] = total_value['Key_core_incumbency'] / total_value['key_core_weave']



            # employee_total_obj = {key: employee_total_obj[key] for key in list(employee_total_obj.keys())}
        employee_total_obj = {key: employee_total_obj[key] for key in list(employee_total_obj.keys())}
        employee_total_obj = json.loads(json.dumps(employee_total_obj))
        for code, line_data in employee_total_obj.items():
            line_data['code'] = code
            table_list.append(line_data)
        for index, item in enumerate(table_list):
            item['index'] = index + 1
            item['turnover_rate'] = "{:.2f}%".format(round(item['turnover_rate'], 4) * 100)
            item['total_rate'] = "{:.2f}%".format(round(item['total_rate'], 4) * 100)
            item['Key_core_rate'] = "{:.2f}%".format(round(item['Key_core_rate'], 4) * 100)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': table_list,
                'totalNumber': len(employee_total_obj),
            }
        }
        return return_data

    def active_employee_total_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file','汇总数据下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),'汇总数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('missing_weave', 'fixed_weave', 'key_core_weave', 'turnover_total_person', 'code'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in  ('missing_weave', 'fixed_weave', 'key_core_weave', 'turnover_total_person', 'code'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data

    def active_employee_seniority(self, department_id, search_date, dl_idl, job_sequence, job_class,job_grade):
        """
        在职人力情况-司龄分布
        取出所有满足查询条件的人员，以月为单位算出所有的司龄
        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员司龄人数分布', 'children': [
                {'label': '0-1', 'value': '0-1月', 'width': 100},
                {'label': '1-3', 'value': '1-3月', 'width': 100},
                {'label': '3-6', 'value': '3-6月', 'width': 100},
                {'label': '6-12', 'value': '6-12月', 'width': 100},
                {'label': '12-24', 'value': '1-2年', 'width': 100},
                {'label': '24-60', 'value': '2-5年', 'width': 100},
                {'label': '60', 'value': '5年以上', 'width': 100},
                {'label': 'average', 'value': '平均司龄(年)', 'width': 130},
            ]},
            {'value': '在职人员司龄占比分布', 'children': [
                {'label': '0-1-proportion', 'value': '0-1月', 'width': 130},
                {'label': '1-3-proportion', 'value': '1-3月', 'width': 130},
                {'label': '3-6-proportion', 'value': '3-6月', 'width': 130},
                {'label': '6-12-proportion', 'value': '6-12月', 'width': 130},
                {'label': '12-24-proportion', 'value': '1-2年', 'width': 130},
                {'label': '24-60-proportion', 'value': '2-5年', 'width': 130},
                {'label': '60-proportion', 'value': '5年以上', 'width': 130},
                # {'label': 'average', 'value': '平均司龄(年)', 'width': 130},
            ]},
        ]
        table_list = []
        seniority_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_dl__in': dl_idl,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in':job_grade,
        }
        seniority_params = {key: value for key, value in seniority_params.items() if
                            value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'
        join_data_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            employee_group_join_date__isnull=False,
            employee_status='1', **seniority_params).values('employee_name',
            'employee_group_join_date', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')
        now_data = datetime.now()
        # 总对象
        seniority_total_obj = {}
        # 司龄分布对象
        if join_data_list:
            for join_data in join_data_list:
                seniority_obj = {
                    'index': '',
                    'employee_department__department_full_name': '',
                    'employee_department__department_first_name': '',
                    'employee_department__department_second_name': '',
                    'employee_department__department_third_name': '',
                    '0-1': 0,
                    '1-3': 0,
                    '3-6': 0,
                    '6-12': 0,
                    '12-24': 0,
                    '24-60': 0,
                    '60': 0,
                    'average': 0,
                    'total_person': 0,
                    'total_seniority': 0,
                    '0-1-proportion': 0,
                    '1-3-proportion': 0,
                    '3-6-proportion': 0,
                    '6-12-proportion': 0,
                    '12-24-proportion': 0,
                    '24-60-proportion': 0,
                    '60-proportion': 0,
                    'employee_group_join_date': '',
                }
                # 根据部门编码判断是否在总对象里
                # 不在就新增键值对

                if not seniority_total_obj.get(join_data['employee_department__department_full_code']):
                    seniority_total_obj[join_data['employee_department__department_full_code']] = seniority_obj
                    seniority_total_obj[join_data['employee_department__department_full_code']]['employee_department__department_full_name'] = \
                    join_data['employee_department__department_full_name']

                    seniority_total_obj[join_data['employee_department__department_full_code']][
                        'employee_department__department_first_name'] = join_data[
                        'employee_department__department_first_name']
                    seniority_total_obj[join_data['employee_department__department_full_code']][
                        'employee_department__department_second_name'] = join_data[
                        'employee_department__department_second_name']
                    seniority_total_obj[join_data['employee_department__department_full_code']][
                        'employee_department__department_third_name'] = join_data[
                        'employee_department__department_third_name']

                    seniority_total_obj[join_data['employee_department__department_full_code']]['employee_group_join_date'] = join_data['employee_group_join_date'].strftime('%Y-%m-%d %H:%M:%S')
                if type(join_data['employee_group_join_date']) == datetime:
                    join_data['employee_group_join_date'] = join_data['employee_group_join_date'].strftime('%Y-%m-%d %H:%M:%S')
                # print(seniority_total_obj)
                employee_group_join_date = datetime.strptime(join_data['employee_group_join_date'], '%Y-%m-%d %H:%M:%S')
                # months_between = (now_data.year - employee_group_join_date.year) * 12 + (
                #         now_data.month - employee_group_join_date.month)
                # print(join_data['employee_name'],months_between,employee_group_join_date)
                timedelta=now_data-employee_group_join_date
                # print(timedelta.days)
                months_between=timedelta.days   #相差的天数
                if months_between <= 30:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['0-1'] += 1
                elif 1 < months_between <= 91:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['1-3'] += 1
                elif 3 < months_between <= 182:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['3-6'] += 1
                elif 6 < months_between <= 365:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['6-12'] += 1
                elif 12 < months_between <= 365*2:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['12-24'] += 1
                elif 24 < months_between <= 365*5:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['24-60'] += 1
                else:
                    seniority_total_obj[join_data['employee_department__department_full_code']]['60'] += 1
                seniority_total_obj[join_data['employee_department__department_full_code']]['total_person'] += 1
                seniority_total_obj[join_data['employee_department__department_full_code']]['total_seniority'] += months_between

            # 计算占比和平均司龄
            for total_key, total_value in seniority_total_obj.items():
                # 平均司龄
                seniority_total_obj[total_key]['average'] = total_value['total_seniority'] / total_value[
                    'total_person'] / 12  # 总人数/总人数
                seniority_total_obj[total_key]['0-1-proportion'] = total_value['0-1'] / total_value['total_person']
                seniority_total_obj[total_key]['1-3-proportion'] = total_value['1-3'] / total_value['total_person']
                seniority_total_obj[total_key]['3-6-proportion'] = total_value['3-6'] / total_value['total_person']
                seniority_total_obj[total_key]['6-12-proportion'] = total_value['6-12'] / total_value['total_person']
                seniority_total_obj[total_key]['12-24-proportion'] = total_value['12-24'] / total_value['total_person']
                seniority_total_obj[total_key]['24-60-proportion'] = total_value['24-60'] / total_value['total_person']
                seniority_total_obj[total_key]['60-proportion'] = total_value['60'] / total_value['total_person']

            seniority_total_obj = {key: seniority_total_obj[key] for key in list(seniority_total_obj.keys())}  # 切片
            # print(seniority_total_obj)
        total_number = len(seniority_total_obj)
        seniority_total_obj = json.loads(json.dumps(seniority_total_obj))
        for code, line_data in seniority_total_obj.items():
            line_data['code'] = code
            table_list.append(line_data)
        for index, item in enumerate(table_list):
            item['index'] =  index + 1
            item['average'] = round(item['average'], 2)
            item['0-1-proportion'] = "{:.2f}%".format(round(item['0-1-proportion'], 4) * 100)
            item['1-3-proportion'] = "{:.2f}%".format(round(item['1-3-proportion'], 4) * 100)
            item['3-6-proportion'] = "{:.2f}%".format(round(item['3-6-proportion'], 4) * 100)
            item['6-12-proportion'] = "{:.2f}%".format(round(item['6-12-proportion'], 4) * 100)
            item['12-24-proportion'] = "{:.2f}%".format(round(item['12-24-proportion'], 4) * 100)
            item['24-60-proportion'] = "{:.2f}%".format(round(item['24-60-proportion'], 4) * 100)
            item['60-proportion'] = "{:.2f}%".format(round(item['60-proportion'], 4) * 100)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': table_list,
                'totalNumber': total_number,
            }
        }
        return return_data

    def active_employee_seniority_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                    '司龄分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '司龄分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','total_person','total_seniority','employee_group_join_date'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person','total_seniority','employee_group_join_date'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }
        return return_data
    def active_employee_age(self, department_id, search_date, dl_idl, job_sequence, job_class,job_grade):
        """
        在职人力情况-年龄分布
        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员年龄人数分布', 'children': [
                {'label': '20', 'value': '<=20岁', 'width': 100},
                {'label': '21-25', 'value': '21-25岁', 'width': 100},
                {'label': '26-30', 'value': '26-30岁', 'width': 100},
                {'label': '31-35', 'value': '31-35岁', 'width': 100},
                {'label': '36-40', 'value': '36-40岁', 'width': 100},
                {'label': '40', 'value': '>40岁', 'width': 100},
                {'label': 'other', 'value': '其他', 'width': 100},
                {'label': 'average', 'value': '平均年龄', 'width': 100},
            ]},
            {'value': '在职人员年龄占比分布', 'children': [
                {'label': '20-proportion', 'value': '<=20岁', 'width': 100},
                {'label': '21-25-proportion', 'value': '21-25岁', 'width': 100},
                {'label': '26-30-proportion', 'value': '26-30岁', 'width': 100},
                {'label': '31-35-proportion', 'value': '31-35岁', 'width': 100},
                {'label': '36-40-proportion', 'value': '36-40岁', 'width': 100},
                {'label': '40-proportion', 'value': '>40岁', 'width': 100},
                {'label': 'other-proportion', 'value': '其他', 'width': 100},
            ]}
        ]
        table_list = []
        age_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_dl__in': dl_idl,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in': job_grade,
        }
        age_params = {key: value for key, value in age_params.items() if
                      value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'
        birthday_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            # employee_birthday__isnull=False,
            employee_status='1', **age_params).values(
            'employee_birthday', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')

        now_data = datetime.now()
        # 总对象
        age_total_obj = {}
        total_number = len(age_total_obj)
        # 年龄分布对象

        if birthday_list:
            for birthday in birthday_list:
                age_obj = {
                    'index': '',
                    'employee_department__department_full_name': '',
                    'employee_department__department_first_name': '',
                    'employee_department__department_second_name': '',
                    'employee_department__department_third_name': '',
                    '20': 0,
                    '21-25': 0,
                    '26-30': 0,
                    '31-35': 0,
                    '36-40': 0,
                    '40': 0,
                    'other':0,
                    'average': 0,
                    'total_person': 0,
                    'total_age': 0,
                    '20-proportion': 0,
                    '21-25-proportion': 0,
                    '26-30-proportion': 0,
                    '31-35-proportion': 0,
                    '36-40-proportion': 0,
                    '40-proportion': 0,
                    'other-proportion':0,

                }
                # 根据部门编码判断是否在总对象里
                # 不在就新增键值对
                if not age_total_obj.get(birthday['employee_department__department_full_code']):
                    age_total_obj[birthday['employee_department__department_full_code']] = age_obj
                    age_total_obj[birthday['employee_department__department_full_code']]['employee_department__department_full_name'] = birthday[
                        'employee_department__department_full_name']
                    age_total_obj[birthday['employee_department__department_full_code']][
                        'employee_department__department_first_name'] = birthday[
                        'employee_department__department_first_name']
                    age_total_obj[birthday['employee_department__department_full_code']][
                        'employee_department__department_second_name'] = birthday[
                        'employee_department__department_second_name']
                    age_total_obj[birthday['employee_department__department_full_code']][
                        'employee_department__department_third_name'] = birthday[
                        'employee_department__department_third_name']
                if type(birthday['employee_birthday']) == date:
                    birthday['employee_birthday'] = birthday['employee_birthday'].strftime('%Y-%m-%d %H:%M:%S')
                    employee_birthday = datetime.strptime(birthday['employee_birthday'], '%Y-%m-%d %H:%M:%S')
                    months_between = now_data.year - employee_birthday.year
                else:
                    months_between = 0

                if 0<months_between <= 20:
                    age_total_obj[birthday['employee_department__department_full_code']]['20'] += 1
                elif 21 <= months_between <= 25:
                    age_total_obj[birthday['employee_department__department_full_code']]['21-25'] += 1
                elif 26 <= months_between <= 30:
                    age_total_obj[birthday['employee_department__department_full_code']]['26-30'] += 1
                elif 31 <= months_between <= 35:
                    age_total_obj[birthday['employee_department__department_full_code']]['31-35'] += 1
                elif 36 <= months_between <= 40:
                    age_total_obj[birthday['employee_department__department_full_code']]['36-40'] += 1
                elif months_between>40:
                    age_total_obj[birthday['employee_department__department_full_code']]['40'] += 1
                elif months_between==0:
                    age_total_obj[birthday['employee_department__department_full_code']]['other'] += 1
                age_total_obj[birthday['employee_department__department_full_code']]['total_person'] += 1
                age_total_obj[birthday['employee_department__department_full_code']]['total_age'] += months_between
            # 计算占比和平均年龄
            for total_key, total_value in age_total_obj.items():
                # 平均年龄
                age_total_obj[total_key]['average'] = total_value['total_age'] / total_value['total_person']
                age_total_obj[total_key]['20-proportion'] = total_value['20'] / total_value['total_person']
                age_total_obj[total_key]['21-25-proportion'] = total_value['21-25'] / total_value['total_person']
                age_total_obj[total_key]['26-30-proportion'] = total_value['26-30'] / total_value['total_person']
                age_total_obj[total_key]['31-35-proportion'] = total_value['31-35'] / total_value['total_person']
                age_total_obj[total_key]['36-40-proportion'] = total_value['36-40'] / total_value['total_person']
                age_total_obj[total_key]['40-proportion'] = total_value['40'] / total_value['total_person']
                age_total_obj[total_key]['other-proportion'] = total_value['other'] / total_value['total_person']
            total_number = len(age_total_obj)
            age_total_obj = {key: age_total_obj[key] for key in list(age_total_obj.keys())}
        age_total_obj = json.loads(json.dumps(age_total_obj))
        for code, line_data in age_total_obj.items():
            line_data['code'] = code
            table_list.append(line_data)
        for index, item in enumerate(table_list):
            item['index'] =index + 1
            item['average'] = round(item['average'], 2)
            item['20-proportion'] = "{:.2f}%".format(round(item['20-proportion'], 4) * 100)
            item['21-25-proportion'] = "{:.2f}%".format(round(item['21-25-proportion'], 4) * 100)
            item['26-30-proportion'] = "{:.2f}%".format(round(item['26-30-proportion'], 4) * 100)
            item['31-35-proportion'] = "{:.2f}%".format(round(item['31-35-proportion'], 4) * 100)
            item['36-40-proportion'] = "{:.2f}%".format(round(item['36-40-proportion'], 4) * 100)
            item['40-proportion'] = "{:.2f}%".format(round(item['40-proportion'], 4) * 100)
            item['other-proportion'] = "{:.2f}%".format(round(item['other-proportion'], 4) * 100)
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': table_list,
                'totalNumber': total_number,
            }
        }
        return return_data
    def active_employee_age_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                    '年龄分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '年龄分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','total_person','total_age'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person','total_age'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def active_employee_education(self, department_id, search_date, dl_idl, job_sequence, job_class,job_grade):
        """
        在职人力情况-学历分布
        1	高中及以下
        2	中专
        3	大专
        4	本科
        5	硕士
        6	博士及以上
        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员学历人数分布', 'children': [
                {'label': '6', 'value': '博士级及以上', 'width': 130},
                {'label': '5', 'value': '硕士', 'width': 130},
                {'label': '4', 'value': '本科', 'width': 130},
                {'label': '3', 'value': '大专', 'width': 130},
                {'label': '2', 'value': '中专', 'width': 130},
                {'label': '1', 'value': '高中及以上', 'width': 130},
                {'label': 'other', 'value': '其他', 'width': 130},
            ]},
            {'value': '在职人员学历占比分布', 'children': [
                {'label': '6-proportion', 'value': '博士级及以上', 'width': 130},
                {'label': '5-proportion', 'value': '硕士', 'width': 130},
                {'label': '4-proportion', 'value': '本科', 'width': 130},
                {'label': '3-proportion', 'value': '大专', 'width': 130},
                {'label': '2-proportion', 'value': '中专', 'width': 130},
                {'label': '1-proportion', 'value': '高中及以上', 'width': 130},
                {'label':'other-proportion','value':"其他",'width':130}
            ]},
        ]
        table_list = []
        education_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_dl__in': dl_idl,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in': job_grade,
        }
        education_params = {key: value for key, value in education_params.items() if
                            value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'
        education_degree_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
             employee_status='1', **education_params).values(
            'employee_first_degree_id', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')
        # 总对象
        education_degree_total_obj = {}
        # 学历分布对象
        education_degree_obj = {
            'index': '',
            'employee_department__department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            '6': 0,
            '5': 0,
            '4': 0,
            '3': 0,
            '2': 0,
            '1': 0,
            'other':0,
            '6-proportion': 0,
            '5-proportion': 0,
            '4-proportion': 0,
            '3-proportion': 0,
            '2-proportion': 0,
            '1-proportion': 0,
            'other-proportion':0,
            'total_person': 0,

        }
        if education_degree_list:
            for education_degree in list(education_degree_list):
                department_code = education_degree['employee_department__department_full_code']
                if department_code not in education_degree_total_obj:
                    education_degree_total_obj[department_code] = education_degree_obj.copy()
                    education_degree_total_obj[department_code]['employee_department__department_full_name'] = education_degree[
                        'employee_department__department_full_name']
                    education_degree_total_obj[department_code]['employee_department__department_first_name'] = \
                    education_degree['employee_department__department_first_name']
                    education_degree_total_obj[department_code]['employee_department__department_second_name'] = \
                    education_degree['employee_department__department_second_name']
                    education_degree_total_obj[department_code]['employee_department__department_third_name'] = \
                    education_degree['employee_department__department_third_name']

                if education_degree['employee_first_degree_id'] == 1:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['1'] += 1
                elif education_degree['employee_first_degree_id'] == 2:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['2'] += 1
                elif education_degree['employee_first_degree_id'] == 3:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['3'] += 1
                elif education_degree['employee_first_degree_id'] == 4:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['4'] += 1
                elif education_degree['employee_first_degree_id'] == 5:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['5'] += 1
                elif education_degree['employee_first_degree_id'] == 6:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['6'] += 1
                else:
                    education_degree_total_obj[education_degree['employee_department__department_full_code']]['other'] += 1
                education_degree_total_obj[education_degree['employee_department__department_full_code']][
                    'total_person'] += 1
            # 计算占比
            for total_key, total_value in education_degree_total_obj.items():
                education_degree_total_obj[total_key]['1-proportion'] = total_value['1'] / total_value['total_person']
                education_degree_total_obj[total_key]['2-proportion'] = total_value['2'] / total_value['total_person']
                education_degree_total_obj[total_key]['3-proportion'] = total_value['3'] / total_value['total_person']
                education_degree_total_obj[total_key]['4-proportion'] = total_value['4'] / total_value['total_person']
                education_degree_total_obj[total_key]['5-proportion'] = total_value['5'] / total_value['total_person']
                education_degree_total_obj[total_key]['6-proportion'] = total_value['6'] / total_value['total_person']
                education_degree_total_obj[total_key]['other-proportion'] = total_value['other'] / total_value['total_person']
            total_number = len(education_degree_total_obj)
            education_degree_total_obj = {key: education_degree_total_obj[key] for key in
                                          list(education_degree_total_obj.keys())}
            education_degree_total_obj = json.loads(json.dumps(education_degree_total_obj))
            for code, line_data in education_degree_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] = index + 1
                item['1-proportion'] = "{:.2f}%".format(round(item['1-proportion'], 4) * 100)
                item['2-proportion'] = "{:.2f}%".format(round(item['2-proportion'], 4) * 100)
                item['3-proportion'] = "{:.2f}%".format(round(item['3-proportion'], 4) * 100)
                item['4-proportion'] = "{:.2f}%".format(round(item['4-proportion'], 4) * 100)
                item['5-proportion'] = "{:.2f}%".format(round(item['5-proportion'], 4) * 100)
                item['6-proportion'] = "{:.2f}%".format(round(item['6-proportion'], 4) * 100)
                item['other-proportion'] = "{:.2f}%".format(round(item['other-proportion'], 4) * 100)
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': total_number,
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': [],
                    'totalNumber': 0,

                }
            }
        return return_data
    def active_employee_education_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                    '学历分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '学历分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','total_person'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def active_employee_nationality(self, department_id, search_date, job_sequence, job_class,job_grade,check_token,flag):
        """
        在职人力情况-国籍分布
        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        # column_list = [
        #     {'label': 'index', 'value': '序号', 'width': 60},
        #     {'label': 'department_name', 'value': '部门', 'width': 230},
        #     {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
        #     {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
        #     {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
        #     {'value': '在职人员国籍人数分布', 'children': [
        #         {'label': '中国', 'value': '中国', 'width': 130},
        #         {'label': '泰国', 'value': '泰国', 'width': 130},
        #         {'label': '美国', 'value': '美国', 'width': 130},
        #         {'label': '越南', 'value': '越南', 'width': 130},
        #         {'label': '菲律宾', 'value': '菲律宾', 'width': 130},
        #         {'label': '西班牙', 'value': '西班牙', 'width': 130},
        #         {'label': '墨西哥', 'value': '墨西哥', 'width': 130},
        #         # {'label': '马来西亚', 'value': '马来西亚', 'width': 130},
        #         {'label': '哥伦比亚', 'value': '哥伦比亚', 'width': 130},
        #     ]},
        #     {'value': '在职人员国籍占比分布', 'children': [
        #         {'label': '中国proportion', 'value': '中国', 'width': 130},
        #         {'label': '泰国proportion', 'value': '泰国', 'width': 130},
        #         {'label': '美国proportion', 'value': '美国', 'width': 130},
        #         {'label': '越南proportion', 'value': '越南', 'width': 130},
        #         {'label': '菲律宾proportion', 'value': '菲律宾', 'width': 130},
        #         {'label': '西班牙proportion', 'value': '西班牙', 'width': 130},
        #         {'label': '墨西哥proportion', 'value': '墨西哥', 'width': 130},
        #         # {'label': '马来西亚proportion', 'value': '马来西亚', 'width': 130},
        #         {'label': '哥伦比亚proportion', 'value': '哥伦比亚', 'width': 130},
        #     ]},
        # ]
        table_list = []
        # print(job_sequence,job_class,job_grade)
        nationality_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in': job_grade,
        }
        nationality_params = {key: value for key, value in nationality_params.items() if
                              value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'

        # 找到当前查询条件下的所有国籍
        nationality_type_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            employee_status='1', **nationality_params).values(
            'employee_nationality').distinct()


        nationality_type_list = sorted(nationality_type_list, key=lambda x: (x['employee_nationality'] is None, x['employee_nationality']))

        # print(a)
        children = []
        children_pro = []

        for line in nationality_type_list:
            if line['employee_nationality'] is None:
                line['employee_nationality']='其他'
            new_dict = {'label': '', 'value': '', 'width': 130}
            new_dict['label'] = line['employee_nationality']
            new_dict['value'] = line['employee_nationality']
            children.append(new_dict)
            new_pro_dict = {'label': '', 'value': '', 'width': 130}
            new_pro_dict['label']=line['employee_nationality']+'proportion'
            new_pro_dict['value']=line['employee_nationality']
            children_pro.append(new_pro_dict)

        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员国籍人数分布', 'children': children},
            {'value': '在职人员国籍占比分布', 'children': children_pro},
        ]
        # print(column_list)

        if nationality_type_list:
            # 拼接查询语句
            sql_part = ""
            nationality_name = []  #所有国家的名字
            for nationality_type in nationality_type_list:
                if nationality_type['employee_nationality']=='其他':
                    sql_part += "sum(if(employee_nationality IS Null,1,0)) as '{}'," .format (nationality_type['employee_nationality'])
                else:
                    sql_part += "sum(if(employee_nationality = '{}',1,0)) as '{}'," .format (nationality_type['employee_nationality'], nationality_type['employee_nationality'])
                nationality_name.append(nationality_type['employee_nationality'])
            # print(nationality_name)
            if flag:
                save_list_data_to_redis('Active_Employee_Nationality_Name_POST_' + str(check_token)+'_default',
                                        nationality_name)  # 将国家名保存到redis中
            save_list_data_to_redis('Active_Employee_Nationality_Name_POST_'+str(check_token),nationality_name)  #将国家名保存到redis中
            # 初始化对象
            # 总对象
            nationality_total_obj = {}
            nationality_obj = {
                'index': '',
                'department_full_name': '',
                'employee_department__department_first_name': '',
                'employee_department__department_second_name': '',
                'employee_department__department_third_name': '',
                'total_person': 0,
            }


            for nationality_name_data in nationality_name:
                nationality_obj.setdefault(nationality_name_data, '')
                nationality_obj.setdefault(nationality_name_data + 'proportion', 0)
            # 完整sql   department_first_name,department_second_name,department_third_name,
            # sql_full = """select department_code,department_name,department_first_name,department_second_name,department_third_name,%s FROM
            #                 hr_employee
            #                 INNER JOIN hr_department on hr_employee.employee_department_id = hr_department.id
            #                 where employee_status = '1'
            #
            #                 GROUP BY employee_department_id""" % sql_part[:-1]
            department_id_tuple = tuple(department_id)
            if len(department_id_tuple) == 1:
                department_id_tuple = "('" + str(department_id_tuple[0]) + "')"
            job_class_tuple = tuple(int(x) for x in job_class)
            if len(job_class_tuple) == 1:
                job_class_tuple = "(" + str(job_class_tuple[0]) + ")"
            job_sequence_tuple = tuple(int(x) for x in job_sequence)
            if len(job_sequence_tuple) == 1:
                job_sequence_tuple = "(" + str(job_sequence_tuple[0]) + ")"
            sql_full = """SELECT department_code,department_full_name,department_first_name,department_second_name,department_third_name,{} 
                                     FROM hr_employee
                                 INNER JOIN hr_department on hr_employee.employee_department_id = hr_department.id
                                 where employee_status = '1' 
                                     AND employee_department_id IN {}
                                     AND employee_group_join_date  <= '{}'
                                     AND employee_job_class_id IN {}
                                     AND employee_job_sequence_id IN {}
                                 GROUP BY employee_department_id""".format(sql_part[:-1], department_id_tuple,
                                                                           search_date, job_class_tuple,
                                                                           job_sequence_tuple)

            if len(department_id_tuple) == 0 or department_id is None:
                sql_full = sql_full.replace("AND employee_department_id IN ()", "")
            if len(job_class_tuple) == 0 or job_class is None:
                sql_full = sql_full.replace("AND employee_job_class_id IN ()", "")
            if len(job_sequence_tuple) == 0 or job_sequence is None:
                sql_full = sql_full.replace("AND employee_job_sequence_id IN ()", "")


            with connection.cursor() as cursor:
                cursor.execute(sql_full)
                nationality_list = cursor.fetchall()
            # print(nationality_list)
            # 生成数据对象
            for nationality in nationality_list:
                if not nationality_total_obj.get(nationality[0]):
                    nationality_obj = {
                        'department_full_name': '',
                        'employee_department__department_first_name': '',
                        'employee_department__department_second_name': '',
                        'employee_department__department_third_name': '',
                        'total_person': 0,
                    }
                    for nationality_name_data in nationality_name:
                        nationality_obj.setdefault(nationality_name_data, '')
                        nationality_obj.setdefault(nationality_name_data + 'proportion', '')
                    nationality_total_obj.setdefault(nationality[0], nationality_obj)
                    nationality_total_obj[nationality[0]]['department_full_name'] = nationality[1]
                    nationality_total_obj[nationality[0]]['employee_department__department_first_name'] = nationality[2]
                    nationality_total_obj[nationality[0]]['employee_department__department_second_name'] = nationality[3]
                    nationality_total_obj[nationality[0]]['employee_department__department_third_name'] = nationality[4]
                for index, nationality_type in enumerate(nationality_name):
                    nationality_total_obj[nationality[0]][nationality_type] = int(nationality[index + 5])
                    nationality_total_obj[nationality[0]]['total_person'] += int(nationality[index + 5])
            for key, value in nationality_total_obj.items():
                for nationality_type in nationality_name:
                    if nationality_total_obj[key]['total_person'] > 0:
                        nationality_total_obj[key][nationality_type + 'proportion'] = nationality_total_obj[key][nationality_type] /nationality_total_obj[key]['total_person']
            total_number = len(nationality_total_obj)
            nationality_total_obj = {key: nationality_total_obj[key] for key in list(nationality_total_obj.keys())}
            nationality_total_obj = json.loads(json.dumps(nationality_total_obj))
            for code, line_data in nationality_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] =index + 1
                for country in nationality_name:
                    name=country+'proportion'
                    try:
                        item[name] = "{:.2f}%".format(round(item[name], 4) * 100)
                    except:
                        pass


            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': total_number,
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': 0,
                }
            }

        return return_data
    def active_employee_nationality_down(self, request, request_data,table_header_data):
        if len(request_data)==0:
            return  {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "没有数据，无法下载",
            }

        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1,
                                  str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)


        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')

        table_list = list(request_data)
        # print(table_list[1])
        row_data = []

        label_list=[]
        value_list=[]
        for line in table_header_data:
            if 'label' not in line and 'children' in line:
                for children_line in line['children']:
                    label_list.append(children_line['label'])
                    value_list.append(children_line['value'])
            else:
                label_list.append(line['label'])
                value_list.append(line['value'])
        self.createExcelPath(value_list)

        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '国籍分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '国籍分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)



        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line = {key: line.get(key, '') for key in label_list}#排序

                line_data = []
                for k, v in line.items():
                    if k not in ('code', 'total_person'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:

                if line['index'] in id_list:

                    line = {key: line.get(key, '') for key in label_list}  # 排序

                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    @staticmethod
    def createExcelPath(data):  # is not None

        def number_to_string(n):  # 跟据数字得到Excel的列 例如1是A,27是AA
            result = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                result = chr(65 + remainder) + result
            return result
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.styles import PatternFill

        file_len=len(data[5:])/2

        book = Workbook()
        sheet = book.active

        for i, value in enumerate(data):
            sheet.cell(column=i + 1, row=2, value=value)

        sheet.merge_cells('A1:A2')
        sheet.merge_cells('B1:B2')
        sheet.merge_cells('C1:C2')
        sheet.merge_cells('D1:D2')
        sheet.merge_cells('E1:E2')

        # Align the content of the merged cells to center
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')

        sheet['A1'] = data[0]
        sheet['B1'] = data[1]
        sheet['C1'] = data[2]
        sheet['D1'] = data[3]
        sheet['E1'] = data[4]

        num = 5
        letter_begin = chr(ord('A') + num)
        letter_end=chr(ord('A') + num + int(file_len)-1)


        sheet.merge_cells('{}1:{}1'.format(letter_begin,letter_end))
        sheet['F1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['F1'] = '在职人员国籍人数占比'

        # letter_pro_begin = str(chr(ord('A') + num + int(file_len)))+'1'
        letter_pro_begin  =str(number_to_string(num + int(file_len)+1))+'1'

        letter_pro_end=str(number_to_string( num + int(file_len)*2))+'1'


        sheet.merge_cells('{}:{}'.format(letter_pro_begin,letter_pro_end))
        sheet['{}'.format(letter_pro_begin)].alignment = Alignment(horizontal='center', vertical='center')
        sheet['{}'.format(letter_pro_begin)] = '在职人员国籍占比分布'
        # 设置填充颜色为浅蓝色
        fill = PatternFill(start_color="0070c0", end_color="0070c0", fill_type="solid")

        # 将A1到U2的单元格填充为浅蓝色
        for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(data)):
            for cell in row:
                cell.fill = fill
        # 将B列到E列的列宽调整为30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '国籍分布下载模板.xlsx')
        book.save(template_path)
    def active_employee_sex(self, department_id, search_date, job_sequence, job_class,job_grade):
        """
        在职人力情况-性别分布
        man	男
        woman 女
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员性别人数分布', 'children': [
                {'label': 'man', 'value': '男', 'width': ""},
                {'label': 'woman', 'value': '女', 'width': ""},
                {'label': 'other', 'value': '其他', 'width': ""},
            ]},
            {'value': '在职人员性别占比分布', 'children': [
                {'label': 'man-proportion', 'value': '男', 'width': ""},
                {'label': 'woman-proportion', 'value': '女', 'width': ""},
                {'label': 'other-proportion', 'value': '其他', 'width': ""},
            ]},
        ]
        table_list = []
        sex_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in': job_grade,
        }
        sex_params = {key: value for key, value in sex_params.items() if
                      value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'

        sex_type_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            # employee_sex__isnull=False,
            employee_status='1', **sex_params).values(
            'employee_sex', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')
        # 总对象
        sex_type_total_obj = {}
        # 性别分布对象
        sex_type_obj = {
            'index': '',
            'employee_department__department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            'man': 0,
            'woman': 0,
            'other':0,
            'total_person': 0,
            'man-proportion': 0,
            'woman-proportion': 0,
            'other-proportion':0,
        }
        if sex_type_list:
            for education_degree in list(sex_type_list):
                department_code = education_degree['employee_department__department_full_code']
                if department_code not in sex_type_total_obj:
                    sex_type_total_obj[department_code] = sex_type_obj.copy()
                    sex_type_total_obj[department_code]['employee_department__department_full_name'] = education_degree[
                        'employee_department__department_full_name']
                    sex_type_total_obj[department_code]['employee_department__department_first_name'] = \
                    education_degree['employee_department__department_first_name']
                    sex_type_total_obj[department_code]['employee_department__department_second_name'] = \
                    education_degree['employee_department__department_second_name']
                    sex_type_total_obj[department_code]['employee_department__department_third_name'] = \
                    education_degree['employee_department__department_third_name']

                if education_degree['employee_sex'] == '1':  # 男
                    sex_type_total_obj[education_degree['employee_department__department_full_code']]['man'] += 1
                elif education_degree['employee_sex'] == '2':  # 女
                    sex_type_total_obj[education_degree['employee_department__department_full_code']]['woman'] += 1
                else:
                    sex_type_total_obj[education_degree['employee_department__department_full_code']]['other'] += 1
                sex_type_total_obj[education_degree['employee_department__department_full_code']]['total_person'] += 1
            # 计算占比
            for total_key, total_value in sex_type_total_obj.items():
                sex_type_total_obj[total_key]['man-proportion'] = total_value['man'] / total_value['total_person']
                sex_type_total_obj[total_key]['woman-proportion'] = total_value['woman'] / total_value['total_person']
                sex_type_total_obj[total_key]['other-proportion'] = total_value['other'] / total_value['total_person']

            total_number = len(sex_type_total_obj)
            sex_type_total_obj = {key: sex_type_total_obj[key] for key in
                                  list(sex_type_total_obj.keys())}
            sex_type_total_obj = json.loads(json.dumps(sex_type_total_obj))
            for code, line_data in sex_type_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] = index + 1
                item['man-proportion'] = "{:.2f}%".format(round(item['man-proportion'], 4) * 100)
                item['woman-proportion'] = "{:.2f}%".format(round(item['woman-proportion'], 4) * 100)
                item['other-proportion'] = "{:.2f}%".format(round(item['other-proportion'], 4) * 100)
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': total_number,
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': [],
                    'totalNumber': 0,
                }
            }
        return return_data
    def active_employee_sex_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '性别分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '性别分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','total_person'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def active_employee_promotion(self, department_id, begin_date, end_date,job_sequence, job_class,new_job_grade,old_job_grade):
        """
        在职人力情况-晋升情况
            promote       晋升
            rise          晋级
            demotion      降职
            downgrade     降级

        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '晋升/晋级', 'children': [
                {'label': 'promote', 'value': '晋升', 'width': ""},
                {'label': 'rise', 'value': '晋级', 'width': ""}
            ]},
            {'value': '降职/降级', 'children': [
                {'label': 'demotion', 'value': '降职', 'width': ""},
                {'label': 'downgrade', 'value': '降级', 'width': ""}
            ]},
        ]
        # if search_date is None or search_date == '':
        #     search_date = '3000-01-01 00:00:00'

        if begin_date is None or begin_date == '':
            begin_date = '1949-10-01 00:00:00'
        if end_date is None or end_date == '':
            end_date = '3949-10-01 00:00:00'

        new_job_grade_tuple = tuple(new_job_grade)  #新职级
        if len(new_job_grade_tuple) == 1:
            new_job_grade_tuple = "(" + str(new_job_grade_tuple[0]) + ")"
        old_job_grade_tuple = tuple(old_job_grade)  #旧职级
        if len(old_job_grade_tuple) == 1:
            old_job_grade_tuple = "(" + str(old_job_grade_tuple[0]) + ")"

        department_id_tuple = tuple(department_id)
        if len(department_id_tuple) == 1:
            department_id_tuple = "(" + str(department_id_tuple[0]) + ")"

        job_class_tuple = tuple(job_class)
        if len(job_class_tuple) == 1:
            job_class_tuple = "(" + str(job_class_tuple[0]) + ")"

        job_sequence_tuple = tuple(job_sequence)
        if len(job_sequence_tuple) == 1:
            job_sequence_tuple = "(" + str(job_sequence_tuple[0]) + ")"

        from utils.sqlServerConnect import EhrConnect
        # print(search_date)
        select_sql = """
                SELECT
            D.FullName as department_full_name,
            D.Dept1 as employee_department__department_first_name,
            D.Dept2 as employee_department__department_second_name,
            D.Dept3 as employee_department__department_third_name,
            D.FullCode,
            COUNT ( CASE WHEN T.TransferType in (6,9) THEN 1 ELSE NULL END ) AS promote,
            COUNT ( CASE WHEN T.TransferType in (2) THEN 1 ELSE NULL END ) AS rise,
            COUNT ( CASE WHEN T.TransferType in (13,14,12) THEN 1 ELSE NULL END ) AS demotion,
            COUNT ( CASE WHEN T.TransferType in (7,10,3) THEN 1 ELSE NULL END ) AS downgrade 
        FROM
            T_HR_Employee AS E
            INNER JOIN T_HR_Department AS D ON E.DeptID = D.id
            LEFT JOIN T_HR_Transfer AS T ON E.id = T.EmpID 
            LEFT JOIN T_HR_JobSequence as S ON E.JobSequenceID=S.id
        WHERE
            E.EmployeeStatusID = '1' 
            AND E.DeptID IN {}
            AND E.JobClassID IN  {} 
            AND T.TransferDate BETWEEN '{}' AND '{}' 
            AND E.JobSequenceID IN {}
            AND T.NewJobGradeID IN {}
            AND T.OldJobGradeID IN {}
            
        GROUP BY
            D.FullName,
            D.FullCode,
            D.Dept1 ,
            D.Dept2 ,
            D.Dept3;
        """.format(department_id_tuple, job_class_tuple, begin_date,end_date,
                   job_sequence_tuple,new_job_grade_tuple,old_job_grade_tuple)  # ,job_class,search_date,job_sequence

        if job_class is None or len(job_class_tuple) == 0:
            select_sql = select_sql.replace("AND E.JobClassID IN  ()", "")
        if job_sequence is None or len(job_sequence_tuple) == 0:
            select_sql = select_sql.replace("AND E.JobSequenceID IN ()", "")
        if len(department_id_tuple) == 0:
            select_sql = select_sql.replace("AND E.DeptID IN ()", "")
        if new_job_grade is None or len(new_job_grade_tuple) == 0:
            select_sql = select_sql.replace("AND T.NewJobGradeID IN ()", "")
        if old_job_grade is None or len(old_job_grade_tuple) == 0:
            select_sql = select_sql.replace("AND T.OldJobGradeID IN ()", "")
        table_list = EhrConnect().select(select_sql)
        total_number = len(table_list)
        table_list = [{'index': index + 1, **item} for index, item in enumerate(table_list)]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': column_list,
                'tableList': table_list,
                'totalNumber': total_number,
            }
        }
        return return_data
    def active_employee_promotion_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '晋升情况下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '晋升情况数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('FullCode','total_person'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('FullCode','total_person'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def active_employee_job_grade(self, department_id, search_date, job_sequence, job_class,job_grade):
        """
        在职人力情况-职级分布
            1     T1-助攻级
            2     T2-工程师级
            2-2   T2.2-技术员级
            3     T3-中工级
            4     T4-高工级
            5     T5-资工级
            6     T6-专家级
            7     T7-总工级

        :param dl_idl:
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'value': '在职人员职等分布', 'children': [
                {'label': 't-1', 'value': 'T1-助工级', 'width': ""},
                {'label': 't-2', 'value': 'T2-工程师级', 'width': ""},
                {'label': 't-2-2', 'value': 'T2.2-技术员级', 'width': ""},
                {'label': 't-3', 'value': 'T3-中工级', 'width': ""},
                {'label': 't-4', 'value': 'T4-高工级', 'width': ""},
                {'label': 't-5', 'value': 'T5-资工级', 'width': ""},
                {'label': 't-6', 'value': 'T6-专家级', 'width': ""},
                {'label': 't-7', 'value': 'T7-总工级', 'width': ""},

                {'label': 'm-2', 'value': 'M2-班长级', 'width': ""},
                {'label': 'm-3', 'value': 'M3-倒班主管级', 'width': ""},
                {'label': 'm-4', 'value': 'M4-主管级', 'width': ""},
                {'label': 'm-5', 'value': 'M5-经理级', 'width': ""},
                {'label': 'm-6', 'value': 'M6-总监级', 'width': ""},
                {'label': 'm-7', 'value': 'M7-总经理级', 'width': ""},
                {'label': 'm-8', 'value': 'M8-总裁级', 'width': ""},

                {'label': 'p-1', 'value': 'P1-助理级', 'width': ""},
                {'label': 'p-2', 'value': 'P2-专员级', 'width': ""},
                {'label': 'p-3', 'value': 'P3-中专级', 'width': ""},
                {'label': 'p-4', 'value': 'P4-专业主管级', 'width': ""},
                {'label': 'p-5', 'value': 'P5-专业经理级', 'width': ""},
                {'label': 'p-6', 'value': 'P6-专家级', 'width': ""},

                {'label': 'o-1', 'value': 'O1-作业员级', 'width': ""},
                {'label': 'o-2-1', 'value': 'O2-技工级', 'width': ""},
                {'label': 'o-2-2', 'value': 'O2-技师级', 'width': ""},

                {'label': 'zhuiguang', 'value': '追光者', 'width': ""},
                {'label': 'guanpei', 'value': '管培生', 'width': ""},
                {'label': 'jingying', 'value': '精英人才', 'width': ""},
                {'label': 'shixi', 'value': '实习生', 'width': ""},

                # {'label': 'jiu-1', 'value': '操作员（旧）', 'width': ""},
                # {'label': 'jiu-2', 'value': '班长级（旧）', 'width': ""},
                # {'label': 'jiu-3', 'value': '员工级（旧）', 'width': ""},
                # {'label': 'jiu-4', 'value': '技术员（旧）', 'width': ""},
                # {'label': 'jiu-5', 'value': '助工级（旧）', 'width': ""},
                # {'label': 'jiu-6', 'value': '工程师级（旧）', 'width': ""},
                # {'label': 'jiu-7', 'value': '主管级（旧）', 'width': ""},
                # {'label': 'jiu-8', 'value': '经理级（旧）', 'width': ""},
                # {'label': 'jiu-9', 'value': '总监级及以上（旧）', 'width': ""},

            ]},
            # {'value': '在职人员职等占比分布', 'children': [
            #     {'label': '1-proportion', 'value': 'T1-助攻级', 'width': 130},
            #     {'label': '2-proportion', 'value': 'T2-工程师级', 'width': 130},
            #     {'label': '2-2-proportion', 'value': 'T2.2-技术员级', 'width': 130},
            #     {'label': '3-proportion', 'value': 'T3-中工级', 'width': 130},
            #     {'label': '4-proportion', 'value': 'T4-高工级', 'width': 130},
            #     {'label': '5-proportion', 'value': 'T5-资工级', 'width': 130},
            #     {'label': '6-proportion', 'value': 'T6-专家级', 'width': 130},
            #     {'label': '7-proportion', 'value': 'T7-总工级', 'width': 130}
            # ]},

        ]
        table_list = []
        job_grade_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
            'employee_job_grade_id__in': job_grade,
        }
        job_grade_params = {key: value for key, value in job_grade_params.items() if
                            value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'

        # job_grade_type_list = HrEmployee.objects.filter(
        #     (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__gt=search_date)),
        #     employee_job_grade_id__isnull=False, employee_status='1', **job_grade_params).values(
        #     'employee_job_grade_id', 'employee_department__department_name', 'employee_department__department_code','employee_department__department_first_name','employee_department__department_second_name','employee_department__department_third_name')


        all_job_grade_list = list(HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            employee_job_class_id__isnull=False, employee_status='1', **job_grade_params).values_list(
            'employee_job_class_id',flat=True).distinct())


        job_grade_type_list = HrEmployee.objects.filter(
            (Q(employee_group_join_date__isnull=True) | Q(employee_group_join_date__lte=search_date)),
            employee_job_class_id__isnull=False, employee_status='1', **job_grade_params).values(
            'employee_job_class_id', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')

        # 总对象
        job_grade_type_total_obj = {}
        # 职级分布对象
        job_grade_type_obj = {
            'index': '',
            'employee_department__department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            'total_person': 0,
            't-1': 0,
            't-2': 0,
            't-2-2': 0,
            't-3': 0,
            't-4': 0,
            't-5': 0,
            't-6': 0,
            't-7': 0,

            'm-2': 0,
            'm-3': 0,
            'm-4': 0,
            'm-5': 0,
            'm-6': 0,
            'm-7': 0,
            'm-8': 0,

            'p-1': 0,
            'p-2': 0,
            'p-3': 0,
            'p-4': 0,
            'p-5': 0,
            'p-6': 0,

            'o-1': 0,
            'o-2-1': 0,
            'o-2-2': 0,

            'zhuiguang': 0,
            'guanpei': 0,
            'jingying': 0,
            'shixi': 0,

            # 'jiu-1': 0,
            # 'jiu-2': 0,
            # 'jiu-3': 0,
            # 'jiu-4': 0,
            # 'jiu-5': 0,
            # 'jiu-6': 0,
            # 'jiu-7': 0,
            # 'jiu-8': 0,
            # 'jiu-9': 0,

            # '1-proportion': 0,
            # '2-proportion': 0,
            # '2-2-proportion': 0,
            # '3-proportion': 0,
            # '4-proportion': 0,
            # '5-proportion': 0,
            # '6-proportion': 0,
            # '7-proportion': 0,
        }
        if job_grade_type_list:
            for education_degree in list(job_grade_type_list):
                department_code = education_degree['employee_department__department_full_code']
                if department_code not in job_grade_type_total_obj:
                    job_grade_type_total_obj[department_code] = job_grade_type_obj.copy()
                    job_grade_type_total_obj[department_code]['employee_department__department_full_name'] = education_degree[
                        'employee_department__department_full_name']
                    job_grade_type_total_obj[department_code]['employee_department__department_first_name'] = \
                    education_degree['employee_department__department_first_name']
                    job_grade_type_total_obj[department_code]['employee_department__department_second_name'] = \
                    education_degree['employee_department__department_second_name']
                    job_grade_type_total_obj[department_code]['employee_department__department_third_name'] = \
                    education_degree['employee_department__department_third_name']
                # if education_degree['employee_job_grade_id'] in [49,50,51]:                    # T1-助工级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-1'] += 1
                # elif education_degree['employee_job_grade_id'] in [44,45,46,47,48]:            # T2-工程师级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-2'] += 1
                # elif education_degree['employee_job_grade_id'] in [101,102,103,104,105]:       # T2.2-技术员级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-2-2'] += 1
                # elif education_degree['employee_job_grade_id'] in [41,42,43]:                  # T3-中工级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-3'] += 1
                # elif education_degree['employee_job_grade_id'] in [38,39,40]:                  # T4-高工级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-4'] += 1
                # elif education_degree['employee_job_grade_id'] in [35,36,37]:                  # T5-资工级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-5'] += 1
                # elif education_degree['employee_job_grade_id'] in [33,34]:                     # T6-专家级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-6'] += 1
                # elif education_degree['employee_job_grade_id'] in [31,32]:                      # T7-总工级
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['t-7'] += 1

                # if education_degree['employee_job_class_id'] == 1:  # 旧 总监
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-9'] += 1
                # elif education_degree['employee_job_class_id'] == 2:  # 旧 经理
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-8'] += 1
                # elif education_degree['employee_job_class_id'] == 3:  # 旧 主管
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-7'] += 1
                # elif education_degree['employee_job_class_id'] == 4:  # 旧 工程师
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-6'] += 1
                # elif education_degree['employee_job_class_id'] == 5:  # 旧 助攻
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-5'] += 1
                # elif education_degree['employee_job_class_id'] == 6:  # 旧 技术员
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-4'] += 1
                # elif education_degree['employee_job_class_id'] == 7:  # 旧 员工
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-3'] += 1
                # elif education_degree['employee_job_class_id'] == 8:  # 旧 班长
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-2'] += 1
                # elif education_degree['employee_job_class_id'] == 9:  # 旧 操作员
                #     job_grade_type_total_obj[education_degree['employee_department__department_code']]['jiu-1'] += 1


                if education_degree['employee_job_class_id'] == 29:  # p1
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-1'] += 1
                elif education_degree['employee_job_class_id'] == 28:  # p2
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-2'] += 1
                elif education_degree['employee_job_class_id'] == 27:  # p3
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-3'] += 1
                elif education_degree['employee_job_class_id'] == 26:  # p4
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-4'] += 1
                elif education_degree['employee_job_class_id'] == 25:  # p5
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-5'] += 1
                elif education_degree['employee_job_class_id'] == 24:  # p6
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['p-6'] += 1


                elif education_degree['employee_job_class_id'] == 16:  # m2
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-2'] += 1
                elif education_degree['employee_job_class_id'] == 15:  # m3
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-3'] += 1
                elif education_degree['employee_job_class_id'] == 14:  # m4
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-4'] += 1
                elif education_degree['employee_job_class_id'] == 13:  # m5
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-5'] += 1
                elif education_degree['employee_job_class_id'] == 12:  # m6
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-6'] += 1
                elif education_degree['employee_job_class_id'] == 11:  # m7
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-7'] += 1
                elif education_degree['employee_job_class_id'] == 10:  # m8
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['m-8'] += 1

                elif education_degree['employee_job_class_id'] == 23:  # t1
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-1'] += 1
                elif education_degree['employee_job_class_id'] == 22:  # t2
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-2'] += 1
                elif education_degree['employee_job_class_id'] == 21:  # t3
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-3'] += 1
                elif education_degree['employee_job_class_id'] == 20:  # t4
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-4'] += 1
                elif education_degree['employee_job_class_id'] == 19:  # t5
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-5'] += 1
                elif education_degree['employee_job_class_id'] == 18:  # t6
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-6'] += 1
                elif education_degree['employee_job_class_id'] == 17:  # t7
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-7'] += 1
                elif education_degree['employee_job_class_id'] == 36:  # t2.2
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['t-2-2'] += 1

                elif education_degree['employee_job_class_id'] == 32:  # o1
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['o-1'] += 1
                elif education_degree['employee_job_class_id'] == 31:  # o2-技工
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['o-2-1'] += 1
                elif education_degree['employee_job_class_id'] == 30:  # o2-技师
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['o-2-2'] += 1

                elif education_degree['employee_job_class_id'] == 33:  # 追光者
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['zhuiguang'] += 1
                elif education_degree['employee_job_class_id'] == 34:  # 管培生
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['guanpei'] += 1
                elif education_degree['employee_job_class_id'] == 35:  # 精英
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['jingying'] += 1
                elif education_degree['employee_job_class_id'] == 37:  # 实习生
                    job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['shixi'] += 1

                job_grade_type_total_obj[education_degree['employee_department__department_full_code']]['total_person'] += 1
                # 计算比例
                # for total_key, total_value in job_grade_type_total_obj.items():
                #     job_grade_type_total_obj[total_key]['1-proportion'] = total_value['1'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['2-proportion'] = total_value['2'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['2-2-proportion'] = total_value['2-2'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['3-proportion'] = total_value['3'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['4-proportion'] = total_value['4'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['5-proportion'] = total_value['5'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['6-proportion'] = total_value['6'] / total_value['total_person']
                #     job_grade_type_total_obj[total_key]['7-proportion'] = total_value['7'] / total_value['total_person']

            total_number = len(job_grade_type_total_obj)
            job_grade_type_total_obj = {key: job_grade_type_total_obj[key] for key in
                                        list(job_grade_type_total_obj.keys())}

            job_grade_type_total_obj = json.loads(json.dumps(job_grade_type_total_obj))
            for code, line_data in job_grade_type_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] = index + 1

            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': total_number,
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': [],
                    'totalNumber': 0,
                }
            }

        return return_data
    def active_employee_job_grade_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '职级分布下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '职级分布数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)

        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','total_person'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','total_person'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def departure_employee_reason(self, departure_reason, department_id, search_date, job_sequence, job_class,departure_type,job_grade):#,departure_type
        """
        离职率分析-按离职原因
            department_name      部门
            departure_reason     离职原因
            SAL                  SAL离职人数
            IDL                  IDL离职人数
            DL                   DL离职人数
            total_person         总离职人数
            SAL-proportion       SAL离职人数占比
            IDL-proportion       IDL离职人数占比
            DL-proportion        DL离职人数占比
        :param departure_reason: 离职原因
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_sequence: 职级序列
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """

        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'employee_department__department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': 'departure_reason', 'value': '离职原因', 'width': 150},
            {'label': 'SAL', 'value': 'SAL离职人数', 'width': ""},
            {'label': 'IDL', 'value': 'IDL离职人数', 'width': ""},
            {'label': 'DL', 'value': 'DL离职人数', 'width': ""},
            {'label': 'total_person', 'value': '总离职人数', 'width': ""},
            {'label': 'SAL-proportion', 'value': 'SAL离职人数占比', 'width': ""},
            {'label': 'IDL-proportion', 'value': 'IDL离职人数占比', 'width': ""},
            {'label': 'DL-proportion', 'value': 'DL离职人数占比', 'width': ""}
        ]
        table_list = []
        employee_reason_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
            'employee_dim_reason_id__in': departure_reason,
            'employee_dim_type_id__in':departure_type,
            'employee_job_grade_id__in': job_grade,
        }
        employee_reason_params = {key: value for key, value in employee_reason_params.items() if
                            value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'
        departure_reason_list = HrEmployee.objects.filter(
            (Q(employee_departure_handle_date__isnull=True) | Q(employee_departure_handle_date__lte=search_date)),
            employee_dim_reason_id__isnull=False, employee_dl__isnull=False, employee_dl__gt='', employee_status='2',
            **employee_reason_params).values(
            'employee_dim_reason__dim_reason_name', 'employee_name', 'employee_dl',
            'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')

        # 总对象
        departure_reason_total_obj = {}
        # 离职原因对象
        departure_reason_obj = {
            'index': '',
            'employee_department__department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            'departure_reason': '',
            'SAL': 0,
            'IDL': 0,
            'DL': 0,
            'total_person': 0,
            'SAL-proportion': 0,
            'IDL-proportion': 0,
            'DL-proportion': 0,
        }
        # result = {}
        department_sums_reason = {

        }  # 部门离职原因合计
        if departure_reason_list:
            for reason in departure_reason_list:
                department_code = reason['employee_department__department_full_code']
                reason_name = reason['employee_dim_reason__dim_reason_name']
                dl = reason['employee_dl']
                key = f"{department_code}_{reason_name}"
                if key not in departure_reason_total_obj:
                    departure_reason_total_obj[key] = departure_reason_obj.copy()
                    departure_reason_total_obj[key]['employee_department__department_full_name'] = reason['employee_department__department_full_name']
                    departure_reason_total_obj[key]['employee_department__department_first_name'] = reason[
                        'employee_department__department_first_name']
                    departure_reason_total_obj[key]['employee_department__department_second_name'] = reason[
                        'employee_department__department_second_name']
                    departure_reason_total_obj[key]['employee_department__department_third_name'] = reason[
                        'employee_department__department_third_name']
                    departure_reason_total_obj[key]['departure_reason'] = reason_name
                try:
                    departure_reason_total_obj[key]['total_person'] += 1
                    departure_reason_total_obj[key][dl] += 1
                except:
                    print(reason)

            for key, values in departure_reason_total_obj.items():
                total_person = values['total_person']
                if total_person > 0:
                    # values['department_name'] = values['department_name']
                    values['employee_department__department_first_name']= values['employee_department__department_first_name']
                    values['employee_department__department_second_name'] = values['employee_department__department_second_name']
                    values['employee_department__department_third_name'] = values['employee_department__department_third_name']

                    values['SAL-proportion'] = values['SAL'] / total_person
                    values['IDL-proportion'] = values['IDL'] / total_person
                    values['DL-proportion'] = values['DL'] / total_person

            for key, values in departure_reason_total_obj.items():  # 计算总和
                department_name = values["employee_department__department_full_name"]
                if department_name in department_sums_reason:
                    values['employee_department__department_full_name'] = values["employee_department__department_full_name"]
                    values['employee_department__department_first_name'] = values['employee_department__department_first_name']
                    values['employee_department__department_second_name'] = values['employee_department__department_second_name']
                    values['employee_department__department_third_name'] = values['employee_department__department_third_name']
                    department_sums_reason[department_name]["SAL"] += values["SAL"]
                    department_sums_reason[department_name]["IDL"] += values["IDL"]
                    department_sums_reason[department_name]["DL"] += values["DL"]
                    department_sums_reason[department_name]["total_person"] += values["total_person"]

                else:
                    department_sums_reason[department_name] = {
                        'employee_department__department_full_name': values["employee_department__department_full_name"],
                        'employee_department__department_first_name': values['employee_department__department_first_name'],
                        'employee_department__department_second_name': values['employee_department__department_second_name'],
                        'employee_department__department_third_name': values['employee_department__department_third_name'],
                        'departure_reason':'合计',
                        "SAL": values["SAL"],
                        "IDL": values["IDL"],
                        "DL": values["DL"],
                        "total_person": values["total_person"],

                    }

            result_list = [{"employee_department__department_full_name": k, **v} for k, v in department_sums_reason.items()]
            # print(result_list)
            for entry in result_list:
                entry['departure_reason'] = '合计'
                entry['SAL-proportion'] = entry['SAL'] / entry['total_person']
                entry['IDL-proportion'] = entry['IDL'] / entry['total_person']
                entry['DL-proportion'] = entry['DL'] / entry['total_person']
                # entry['DL-departure_reason'] = '合计'


                departure_reason_total_obj[str(entry['employee_department__department_full_name']) + '合计'] = entry
            total_number = len(departure_reason_total_obj)

            departure_reason_total_obj = {key: departure_reason_total_obj[key] for key in
                                          list(departure_reason_total_obj.keys())}

            departure_reason_total_obj = json.loads(json.dumps(departure_reason_total_obj))
            for code, line_data in departure_reason_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] = index + 1
                item['SAL-proportion'] = "{:.2f}%".format(round(item['SAL-proportion'], 4) * 100)
                item['IDL-proportion'] = "{:.2f}%".format(round(item['IDL-proportion'], 4) * 100)
                item['DL-proportion'] = "{:.2f}%".format(round(item['DL-proportion'], 4) * 100)
            # table_list = [{'index': index + 1, **item} for index, item in enumerate(table_list)]
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': total_number,
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': [],
                    'totalNumber': 0,
                }
            }
        return return_data
    def departure_employee_reason_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '离职率分析-离职原因下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '离职率分析-离职原因数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)

        row_data = []
        if download_all == True:  # 是下载全部   有条件
            index = 1
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code','index'):
                        line_data.append(v)
                line_data.insert(0, index)
                index += 1
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code','index'):
                            line_data.append(v)
                    line_data.insert(0, index)
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def departure_employee_seniority(self, department_id, search_date, job_sequence, job_class,departure_type,departure_reason,job_grade):#,departure_type,departure_reason
        """
         离职率分析-按司龄
              department_name      部门
              7-d                 7天内
              1-m                 1个月内
              3-m                 1-3个月
              6-m                 3-6个月
              12-m                6-12个月
              2-y                 1-2年
              2-y-gt              2年以上
              total_person     total_person

         :param dl_idl:
         :param search_date: 查询日期
         :param department_id: 部门id列表
         :param job_sequence: 职级序列
         :param job_class: 职等
         :param currentPage: 当前页
         :param pageSize: 每页显示数量
         :return:
          """
        column_list = [
            {'label': 'index', 'value': '序号', 'width': 60},
            {'label': 'department_full_name', 'value': '部门', 'width': 230},
            {'label': 'employee_department__department_first_name', 'value': '一级部门', 'width': 230},
            {'label': 'employee_department__department_second_name', 'value': '二级部门', 'width': 230},
            {'label': 'employee_department__department_third_name', 'value': '三级部门', 'width': 230},
            {'label': '7-d', 'value': '7天内', 'width': ""},
            {'label': '1-m', 'value': '1个月内', 'width': ""},
            {'label': '3-m', 'value': '1-3个月', 'width': ""},
            {'label': '6-m', 'value': '3-6个月', 'width': ""},
            {'label': '12-m', 'value': '6-12个月', 'width': ""},
            {'label': '2-y', 'value': '1-2年', 'width': ""},
            {'label': '2-y-gt', 'value': '2年以上', 'width': ""},
            {'label': 'total_person', 'value': '总计', 'width': ""}
        ]
        table_list = []
        departure_seniority_params = {
            'employee_job_sequence_id__in': job_sequence,
            'employee_job_class_id__in': job_class,
            'employee_department_id__in': department_id,
            'employee_dim_reason_id__in': departure_reason,
            'employee_dim_type_id__in': departure_type,
            'employee_job_grade_id__in': job_grade,
        }


        departure_seniority_params = {key: value for key, value in departure_seniority_params.items() if
                                      value is not None and value != ''}  # 过滤掉值为None或''的项
        if search_date is None or search_date == '':
            search_date = '3000-01-01 00:00:00'

        departure_seniority_list = HrEmployee.objects.filter(
            (Q(employee_departure_handle_date__isnull=True) | Q(employee_departure_handle_date__lte=search_date)),
            employee_group_join_date__isnull=False, employee_status='2', **departure_seniority_params).values(
            'employee_group_join_date', 'employee_department__department_full_name', 'employee_department__department_full_code',
            'employee_department__department_first_name', 'employee_department__department_second_name',
            'employee_department__department_third_name')
        # 总对象
        departure_seniority_total_obj = {}
        # 离职司龄对象
        departure_seniority_obj = {
            'index': '',
            'department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            '7-d': 0,
            '1-m': 0,
            '3-m': 0,
            '6-m': 0,
            '12-m': 0,
            '2-y': 0,
            '2-y-gt': 0,
            'total_person': 0,
        }
        if departure_seniority_list:
            for education_degree in list(departure_seniority_list):
                department_code = education_degree['employee_department__department_full_code']
                if department_code not in departure_seniority_total_obj:
                    departure_seniority_total_obj[department_code] = departure_seniority_obj.copy()
                    departure_seniority_total_obj[department_code]['department_full_name'] = education_degree[
                        'employee_department__department_full_name']
                    departure_seniority_total_obj[department_code]['employee_department__department_first_name'] = education_degree['employee_department__department_first_name']
                    departure_seniority_total_obj[department_code]['employee_department__department_second_name'] = education_degree['employee_department__department_second_name']
                    departure_seniority_total_obj[department_code]['employee_department__department_third_name'] = education_degree['employee_department__department_third_name']
                employee_departure_seniority = self.days_until_given_date(education_degree['employee_group_join_date'])
                if employee_departure_seniority <= 7:  # 7天内
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['7-d'] += 1
                elif employee_departure_seniority <= 30:  # 1个月内
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['1-m'] += 1
                elif employee_departure_seniority <= 90:  # 1-3个月
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['3-m'] += 1
                elif employee_departure_seniority <= 180:  # 3-6个月
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['6-m'] += 1
                elif employee_departure_seniority <= 365:  # 6-12个月
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['12-m'] += 1
                elif employee_departure_seniority <= 730:  # 1-2年
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['2-y'] += 1
                else:  # 2年以上
                    departure_seniority_total_obj[education_degree['employee_department__department_full_code']]['2-y-gt'] += 1
                departure_seniority_total_obj[education_degree['employee_department__department_full_code']][
                    'total_person'] += 1

            departure_seniority_total_obj = {key: departure_seniority_total_obj[key] for key in
                                             list(departure_seniority_total_obj.keys())}
            departure_seniority_total_obj = json.loads(json.dumps(departure_seniority_total_obj))
            for code, line_data in departure_seniority_total_obj.items():
                line_data['code'] = code
                table_list.append(line_data)
            for index, item in enumerate(table_list):
                item['index'] = index + 1
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': table_list,
                    'totalNumber': len(table_list),
                }
            }
        else:
            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "data": {
                    'columnList': column_list,
                    'tableList': [],
                    'totalNumber': 0,
                }
            }
        return return_data
    def departure_employee_seniority_down(self, request, request_data):
        dummy_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2))  # 创建文件夹
        self.mkdir(dummy_path)
        template_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_template_file',
                                     '离职率分析-司龄原因下载模板.xlsx')  # 创建文件
        # 指定原始文件路径和目标路径
        source_path = template_path  # 例如：C:\\原始文件夹\\文件名.xlsx
        destination_path = os.path.join(BASE_DIR, 'static', 'employee', 'report_download_file', self.t1, str(self.t2),
                                        '离职率分析-司龄原因数据.xlsx')  # 例如：D:\\目标文件夹\\文件名.xlsx
        # 使用shutil库进行复制操作
        shutil.copy(source_path, destination_path)
        id_list = json.loads(request.body).get('id_list')
        download_all = json.loads(request.body).get('download_all')
        table_list = list(request_data)
        row_data = []
        if download_all == True:  # 是下载全部   有条件
            for line in table_list:
                line_data = []
                for k, v in line.items():
                    if k not in ('code'):
                        line_data.append(v)
                row_data.append(line_data)
        else:
            index = 1
            for line in table_list:
                if line['index'] in id_list:
                    line_data = []
                    for k, v in line.items():
                        if k not in ('code'):
                            line_data.append(v)
                    line_data[0]=index
                    index += 1
                    row_data.append(line_data)
        exc = openpyxl.load_workbook(destination_path)  # 打开整个excel文件
        sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
        for row in row_data:
            sheet.append(row)  # 在工作表中添加一行
        exc.save(destination_path)  # 指定路径,保存文件
        destination_path = destination_path.replace('\\', '/')
        destination_path = 'static/' + destination_path.split('static/')[1]
        return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功",
            "downloadUrl": destination_path
        }

        return return_data
    def departure_employee_area(self, department_id, search_date, job_grade, job_class, currentPage, pageSize):
        """
        离职率分析-按区域
        :param search_date: 查询日期
        :param department_id: 部门id列表
        :param job_grade: 职级
        :param job_class: 职等
        :param currentPage: 当前页
        :param pageSize: 每页显示数量
        :return:
        """

        pass
    def report_visualization(self,request):
        return_data ={
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "viewDataList": [],
                "departure_active_employee_reason":{}
            }
        # print('1',request.check_token)
        #
        # print(AdminUser.objects.filter(id=request.check_token,user_menu__nav_parent_id=127).values_list('user_menu__nav_name',flat=True))#报表平台的权限
        # authority_list = list(AdminUser.objects.filter(id=request.check_token, user_menu__nav_parent_id=127).values_list('user_menu__nav_name', flat=True))
        user_obj_list=list(AdminUser.objects.filter(Q(user_menu__id__in=list(range(127,140))) | Q(is_superuser=1),is_used=1,id=request.check_token).values('id','user_menu__id','is_superuser'))
        # print('1111111111111',user_obj_list)
        from collections import defaultdict
        merged_data = defaultdict(list)
        for d in user_obj_list:
            merged_data[str(d['id'])+'__'+str(d['is_superuser'])].append(d['user_menu__id'])
        authority_dict=dict(merged_data)
        if len(authority_dict)>=1:
            key, value = next(iter(authority_dict.items()))
            key_1, key_2 = key.split('__')
            user_id = int(key_1)
            is_super = eval(key_2)
            user_menu_list = value
            # retrieved_data = get_list_data_from_redis("Active_Employee_Seniority_RecordView_POST")    # 司龄分布
            # df = pd.DataFrame(retrieved_data)
            # sum_0_1 = df['0-1'].sum()
            # sum_1_3 = df['1-3'].sum()
            # sum_3_6 = df['3-6'].sum()
            # sum_6_12 = df['6-12'].sum()
            # sum_12_24 = df['12-24'].sum()
            # sum_24_60 = df['24-60'].sum()
            # sum_60 = df['60'].sum()
            # # sum_total_person = df['total_person'].sum()
            # # sum_total_seniority=df('total_seniority').sum()
            # # average=round(sum_total_seniority/sum_total_person,2)
            #
            # Active_Employee_Seniority_data = [int(sum_0_1), int(sum_1_3), int(sum_3_6), int(sum_6_12), int(sum_12_24), int(sum_24_60), int(sum_60)]
            # Active_Employee_Seniority_xAxis = ['0-1月', '1-3月', '3-6月', '6-12月', '1-2年', '2-5年', '5年以上']
            # print(Active_Employee_Seniority_data, Active_Employee_Seniority_xAxis)
            #
            # retrieved_data = get_list_data_from_redis("Active_Employee_Age_RecordView_POST")    # 年龄分布
            # df = pd.DataFrame(retrieved_data)
            # sum_20 = df['20'].sum()
            # sum_21_25 = df['21-25'].sum()
            # sum_26_30 = df['26-30'].sum()
            # sum_31_35 = df['31-35'].sum()
            # sum_36_40 = df['36-40'].sum()
            # sum_40 = df['40'].sum()
            # sum_other = df['other'].sum()
            #
            #
            # Active_Employee_Age_data = [int(sum_20), int(sum_21_25), int(sum_26_30), int(sum_31_35), int(sum_36_40), int(sum_40), int(sum_other)]
            # Active_Employee_Age_xAxis = ['<=20岁', '21-25岁', '26-30岁', '31-35岁', '36-40岁', '>40岁', '其他']
            # from datetime import datetime

            start_time = datetime.now()

            def retrieve_and_sum_redis_data(redis_key, column_names, labels):
                df = pd.DataFrame(get_list_data_from_redis(redis_key))
                data = [int(df[col].sum()) for col in column_names]
                return data, labels


            viewDataList=[]
            departure_active_employee_reason_dict={}
            if 131 in user_menu_list or is_super:
                try:
                    active_employee_seniority_data, active_employee_seniority_xaxis = retrieve_and_sum_redis_data('Active_Employee_Seniority_RecordView_POST_'+str(request.check_token)+"_default", ['0-1', '1-3', '3-6', '6-12', '12-24', '24-60', '60'], ['0-1月', '1-3月', '3-6月', '6-12月', '1-2年', '2-5年', '5年以上']) #司龄分布
                    data=[{"name": active_employee_seniority_xaxis[i], "value": active_employee_seniority_data[i]}  for i in range(len(active_employee_seniority_xaxis))]
                except:
                    data=[]
                viewDataList.append(
                    {
                        "type": "active_employee_seniority",
                        "data": {
                            'title': '司龄分布',
                            'data':data
                        }
                    }, )
            if 132 in user_menu_list or is_super:
                try:
                    active_employee_age_data, active_employee_age_xaxis = retrieve_and_sum_redis_data('Active_Employee_Age_RecordView_POST_'+str(request.check_token)+"_default", ['20', '21-25', '26-30', '31-35', '36-40', '40', 'other'],['<=20岁', '21-25岁', '26-30岁', '31-35岁', '36-40岁', '>40岁', '其他'])#年龄分布
                    data=[{"name": active_employee_age_xaxis[i], "value": active_employee_age_data[i]}  for i in range(len(active_employee_age_xaxis))]
                except:
                    data=[]
                viewDataList.append({
                        "type": "active_employee_age",
                        "data": {
                            'title': '年龄分布',
                            'data':data,
                            # 'xAxis': active_employee_age_xaxis
                        }
                    },)
            if 133 in user_menu_list or is_super:
                try:
                    active_employee_education_data, active_employee_education_xaxis = retrieve_and_sum_redis_data('Active_Employee_Education_RecordView_POST_'+str(request.check_token)+"_default", ['6', '5', '4', '3', '2', '1', 'other'],['博士级及以上', '硕士', '本科', '大专', '中专', '高中及以上', '其他'])  # 学历分布
                    data=[{"name": active_employee_education_xaxis[i], "value": active_employee_education_data[i]} for i in range(len(active_employee_education_xaxis))]
                except:
                    data=[]
                viewDataList.append(
                             {
                    "type": "active_employee_education",
                    "data": {
                        'title': '学历分布',
                        'data':data,
                    }
                },)
            if 134 in user_menu_list or is_super:
                try:
                    active_employee_nationality_name_data = get_list_data_from_redis("Active_Employee_Nationality_Name_POST_"+str(request.check_token)+"_default")  # 国家名
                    active_employee_nationality_data,active_employee_nationality_xaxis = retrieve_and_sum_redis_data('Active_Employee_Nationality_RecordView_POST_'+str(request.check_token)+"_default", active_employee_nationality_name_data, active_employee_nationality_name_data)  # 国籍分布
                    data=[{"name": active_employee_nationality_xaxis[i], "value": active_employee_nationality_data[i]} for i in range(len(active_employee_nationality_xaxis))]

                except:
                    data=[]
                viewDataList.append(
                    {
                        "type": "active_employee_nationality",
                        "data": {
                            'title': '国籍分布',
                            'data': data,
                        }
                    }, )
            if 135 in user_menu_list or is_super:
                try:
                    active_employee_sex_data, active_employee_sex_xaxis = retrieve_and_sum_redis_data('Active_Employee_Sex_RecordView_POST_'+str(request.check_token)+"_default", ['man', 'woman','other'], ['男', '女','其他'])  # 性别分布
                    data=[{"name": active_employee_sex_xaxis[i], "value": active_employee_sex_data[i]} for i in range(len(active_employee_sex_xaxis))]
                except:
                    data=[]
                viewDataList.append(
                  {
                    "type": "active_employee_sex",
                    "data":{
                          'title': '性别分布',
                          "data": data
                      }
                },)
            if 136 in user_menu_list or is_super:
                try:
                    active_employee_promotion_data, active_employee_promotion_xaxis = retrieve_and_sum_redis_data('Active_Employee_Promotion_RecordView_POST_'+str(request.check_token)+"_default", ['demotion','downgrade','promote','rise'], ['降职','降级','晋升', '晋级'])  # 晋升情况
                    active_employee_promotion_data[:2]=[-num for num in active_employee_promotion_data[:2]]
                except:
                    active_employee_promotion_data=[]
                    active_employee_promotion_xaxis=[]
                viewDataList.append( {
                                    "type": "active_employee_promotion",
                                    "data": {
                                        'title': '晋升情况',
                                        'data': active_employee_promotion_data,
                                        'xAxis': active_employee_promotion_xaxis
                                    }
                                },)
            if 137 in user_menu_list or is_super:
                try:
                    active_employee_job_grade_data_raw, active_employee_job_grade_xaxis_raw = retrieve_and_sum_redis_data(
                        'Active_Employee_Job_Grade_RecordView_POST_'+str(request.check_token)+"_default",
                        ['t-1', 't-2', 't-2-2', 't-3', 't-4', 't-5', 't-6', 't-7', 'm-2', 'm-3', 'm-4', 'm-5', 'm-6', 'm-7',
                         'm-8', 'p-1', 'p-2', 'p-3', 'p-4', 'p-5', 'p-6', 'o-1', 'o-2-1', 'o-2-2'],
                        ['T1-助工级', 'T2-工程师级', 'T2.2-技术员级', 'T3-中工级', 'T4-高工级', 'T5-资工级', 'T6-专家级',
                         'T7-总工级', 'M2-班长级', 'M3-倒班主管级', 'M4-主管级', 'M5-经理级', 'M6-总监级', 'M7-总经理级',
                         'M8-总裁级',
                         'P1-助理级', 'P2-专员级', 'P3-中专级', 'P4-专业主管级', 'P5-专业经理级', 'P6-专家级', 'O1-作业员级',
                         'O2-技工级', 'O2-技师级'])  # 职级分布
                    active_employee_job_grade_data, active_employee_job_grade_xaxis = [
                        sum(active_employee_job_grade_data_raw[:8]), sum(active_employee_job_grade_data_raw[8:15]),
                        sum(active_employee_job_grade_data_raw[15:21]), sum(active_employee_job_grade_data_raw[21:])], [
                        'T-技术序列', 'M-管理序列', 'P-专业序列', 'O-操作序列']  # 职级分布
                except:
                    active_employee_job_grade_data=[]
                    active_employee_job_grade_xaxis=[]
                viewDataList.append({
                                        "type": "active_employee_job_grade",
                                        "data": {
                                            'title': '职级分布',
                                            'data': active_employee_job_grade_data,
                                            'xAxis': active_employee_job_grade_xaxis
                                        }
                                    },)
            if 139 in user_menu_list or is_super:
                try:
                    departure_active_employee_seniority_data, departure_active_employee_seniority_xaxis = retrieve_and_sum_redis_data(
                        'Departure_Employee_Seniority_RecordView_POST_'+str(request.check_token)+"_default", ['7-d', '1-m', '3-m', '6-m', '12-m', '2-y', '2-y-gt'],
                        ['7天内', '1个月内', '1-3个月', '3-6个月', '6-12个月', '1-2年', '2年以上'])  # 离职率分析-司龄原因
                except:
                    departure_active_employee_seniority_data=[]
                    departure_active_employee_seniority_xaxis=[]
                viewDataList.append( {
                                        "type": "departure_active_employee_seniority",
                                        "data": {
                                            'title': '离职率分析-司龄原因',
                                            'data': departure_active_employee_seniority_data,
                                            'xAxis': departure_active_employee_seniority_xaxis
                                        }
                                    },)
            if 138 in user_menu_list or is_super:
                try:
                    departure_reason_retrieved_data = get_list_data_from_redis("Departure_Employee_Reason_RecordView_POST_"+str(request.check_token)+"_default")  # 离职原因
                    departure_reason_df = pd.DataFrame(departure_reason_retrieved_data)
                    departure_reason_df_filtered = departure_reason_df.loc[departure_reason_df['departure_reason'] != '合计']  # 排除合计
                    departure_reason_sums = departure_reason_df_filtered.groupby('departure_reason')[['SAL', 'IDL', 'DL']].sum()  # 按departure_reason分组并计算 SAL、IDL 和 DL 的总和
                    list_of_dicts = departure_reason_sums.reset_index().to_dict('records')  # 将 DataFrame 转换为字典列表
                except:
                    list_of_dicts=[]
                departure_active_employee_reason_dict={
                    'title': '离职率分析-离职原因',
                    'source': list_of_dicts,
                    'dimensions': ['departure_reason', 'SAL', 'IDL', 'DL'],
                },


            return_data = {
                "code": status.HTTP_200_OK,
                "msg": "信息返回成功",
                "viewDataList": viewDataList,
                "departure_active_employee_reason":{} if len(departure_active_employee_reason_dict)==0 else departure_active_employee_reason_dict[0]
            }


            """      active_employee_seniority_data, active_employee_seniority_xaxis = retrieve_and_sum_redis_data('Active_Employee_Seniority_RecordView_POST', ['0-1', '1-3', '3-6', '6-12', '12-24', '24-60', '60'], ['0-1月', '1-3月', '3-6月', '6-12月', '1-2年', '2-5年', '5年以上']) #司龄分布
    
    
            active_employee_age_data, active_employee_age_xaxis = retrieve_and_sum_redis_data('Active_Employee_Age_RecordView_POST', ['20', '21-25', '26-30', '31-35', '36-40', '40', 'other'],['<=20岁', '21-25岁', '26-30岁', '31-35岁', '36-40岁', '>40岁', '其他'])#年龄分布
    
            active_employee_education_data, active_employee_education_xaxis = retrieve_and_sum_redis_data('Active_Employee_Education_RecordView_POST', ['6', '5', '4', '3', '2', '1', 'other'],['博士级及以上', '硕士', '本科', '大专', '中专', '高中及以上', '其他'])  # 学历分布
    
            active_employee_sex_data, active_employee_sex_xaxis = retrieve_and_sum_redis_data('Active_Employee_Sex_RecordView_POST', ['man', 'woman'], ['男', '女'])  # 性别分布
    
            active_employee_promotion_data, active_employee_promotion_xaxis = retrieve_and_sum_redis_data('Active_Employee_Promotion_RecordView_POST', ['demotion','downgrade','promote','rise'], ['降职','降级','晋升', '晋级'])  # 晋升情况
            active_employee_promotion_data[:2]=[-num for num in active_employee_promotion_data[:2]]
    
            active_employee_nationality_name_data = get_list_data_from_redis("Active_Employee_Nationality_Name_POST")  # 国家名
            active_employee_nationality_data,active_employee_nationality_xaxis = retrieve_and_sum_redis_data('Active_Employee_Nationality_RecordView_POST', active_employee_nationality_name_data, active_employee_nationality_name_data)  # 国籍分布
    
            active_employee_job_grade_data_raw, active_employee_job_grade_xaxis_raw = retrieve_and_sum_redis_data(
                'Active_Employee_Job_Grade_RecordView_POST',
                ['t-1', 't-2','t-2-2','t-3','t-4','t-5','t-6','t-7','m-2','m-3','m-4','m-5','m-6','m-7','m-8','p-1','p-2','p-3','p-4','p-5','p-6','o-1','o-2-1','o-2-2'],
                ['T1-助工级', 'T2-工程师级','T2.2-技术员级','T3-中工级','T4-高工级','T5-资工级','T6-专家级','T7-总工级','M2-班长级','M3-倒班主管级','M4-主管级','M5-经理级','M6-总监级','M7-总经理级','M8-总裁级',
                 'P1-助理级','P2-专员级','P3-中专级','P4-专业主管级','P5-专业经理级','P6-专家级','O1-作业员级','O2-技工级','O2-技师级'])  # 职级分布
            active_employee_job_grade_data, active_employee_job_grade_xaxis=[sum(active_employee_job_grade_data_raw[:8]),sum(active_employee_job_grade_data_raw[8:15]),sum(active_employee_job_grade_data_raw[15:21]),sum(active_employee_job_grade_data_raw[21:])],['T-技术序列','M-管理序列','P-专业序列','O-操作序列'] #职级分布
    
            departure_active_employee_seniority_data, departure_active_employee_seniority_xaxis = retrieve_and_sum_redis_data('Departure_Employee_Seniority_RecordView_POST',['7-d', '1-m','3-m','6-m','12-m','2-y','2-y-gt'], ['7天内', '1个月内','1-3个月','3-6个月','6-12个月','1-2年','2年以上'])  # 离职率分析-司龄原因
    
            departure_reason_retrieved_data = get_list_data_from_redis("Departure_Employee_Reason_RecordView_POST")  # 离职原因
            departure_reason_df = pd.DataFrame(departure_reason_retrieved_data)
            departure_reason_df_filtered = departure_reason_df.loc[departure_reason_df['departure_reason'] != '合计']#排除合计
            departure_reason_sums = departure_reason_df_filtered.groupby('departure_reason')[['SAL', 'IDL', 'DL']].sum()#按departure_reason分组并计算 SAL、IDL 和 DL 的总和
            list_of_dicts = departure_reason_sums.reset_index().to_dict('records')  # 将 DataFrame 转换为字典列表"""





            """
                list_of_dicts = departure_reason_sums.reset_index().to_dict('records')  #将 DataFrame 转换为字典列表
                print(list_of_dicts)
    
                numpy_array = departure_reason_sums.reset_index().to_numpy()  #将 DataFrame 转换为 numpy 数组
                list_of_lists = numpy_array.tolist()   #将 numpy 数组转换为列表列表
                departure_reason = departure_reason_sums.index.tolist()
                SAL = departure_reason_sums['SAL'].tolist()
                IDL = departure_reason_sums['IDL'].tolist()
                DL = departure_reason_sums['DL'].tolist()
                print(departure_reason)
                print(SAL)
                print(IDL)
                print(DL)
            """
        return return_data



    # def report_visualization(self):
    #     start_time = datetime.now()
    #
    #     def retrieve_and_sum_redis_data(redis_key, column_names, labels, result_dict,name):
    #         """
    #         :param redis_key: redis的键
    #         :param column_names: 要计算的列名
    #         :param labels: 表头
    #         :param result_dict:   要返回的结果
    #         :param name: 唯一标识
    #         :return:
    #         """
    #         df = pd.DataFrame(get_list_data_from_redis(redis_key))
    #         data = [int(df[col].sum()) for col in column_names]
    #         result_dict[name]=(data,labels)
    #     results = {}
    #
    #     def retrieve_active_employee_seniority_data():  #司龄分布
    #         retrieve_and_sum_redis_data('Active_Employee_Seniority_RecordView_POST',['0-1', '1-3', '3-6', '6-12', '12-24', '24-60', '60'],['0-1月', '1-3月', '3-6月', '6-12月', '1-2年', '2-5年', '5年以上'], results,'employee_seniority_data')
    #     def retrieve_active_employee_age_data():#年龄分布
    #         retrieve_and_sum_redis_data('Active_Employee_Age_RecordView_POST',['20', '21-25', '26-30', '31-35', '36-40', '40', 'other'],['<=20岁', '21-25岁', '26-30岁', '31-35岁', '36-40岁', '>40岁', '其他'],results,'employee_age_data')
    #     def retrieve_active_employee_education_data():#学历分布
    #         retrieve_and_sum_redis_data('Active_Employee_Education_RecordView_POST',['6', '5', '4', '3', '2', '1', 'other'],['博士级及以上', '硕士', '本科', '大专', '中专', '高中及以上', '其他'], results,'employee_education_data')
    #     def retrieve_active_employee_sex_data():#性别分布
    #         retrieve_and_sum_redis_data('Active_Employee_Sex_RecordView_POST', ['man', 'woman'], ['男', '女'], results,'employee_sex_data')
    #     def retrieve_active_employee_promotion_data():#晋升情况
    #         retrieve_and_sum_redis_data('Active_Employee_Promotion_RecordView_POST',['demotion', 'downgrade', 'promote', 'rise'],['降职', '降级', '晋升', '晋级'],results,'employee_promotion_data')
    #     def retrieve_active_employee_nationality_data():  # 国籍分布
    #         active_employee_nationality_name_data = get_list_data_from_redis("Active_Employee_Nationality_Name_POST")  # 国家名
    #         retrieve_and_sum_redis_data('Active_Employee_Nationality_RecordView_POST', active_employee_nationality_name_data, active_employee_nationality_name_data,results,'employee_nationality_data')  # 国籍分布
    #     def retrieve_active_employee_job_grade_data():  # 职级分布
    #         retrieve_and_sum_redis_data(
    #             'Active_Employee_Job_Grade_RecordView_POST',
    #             ['t-1', 't-2','t-2-2','t-3','t-4','t-5','t-6','t-7','m-2','m-3','m-4','m-5','m-6','m-7','m-8','p-1','p-2','p-3','p-4','p-5','p-6','o-1','o-2-1','o-2-2'],
    #             ['T1-助工级', 'T2-工程师级','T2.2-技术员级','T3-中工级','T4-高工级','T5-资工级','T6-专家级','T7-总工级','M2-班长级','M3-倒班主管级','M4-主管级','M5-经理级','M6-总监级','M7-总经理级','M8-总裁级',
    #              'P1-助理级','P2-专员级','P3-中专级','P4-专业主管级','P5-专业经理级','P6-专家级','O1-作业员级','O2-技工级','O2-技师级'],results,'employee_job_grade_data')  # 职级分布
    #     def retrieve_departure_active_employee_seniority_data():  # 离职率分析-司龄原因
    #          retrieve_and_sum_redis_data('Departure_Employee_Seniority_RecordView_POST',['7-d', '1-m','3-m','6-m','12-m','2-y','2-y-gt'], ['7天内', '1个月内','1-3个月','3-6个月','6-12个月','1-2年','2年以上'],results,'departure_employee_seniority_data')  # 离职率分析-司龄原因
    #     def retrieve_departure_active_employee_reason_data():#离职率分析-离职原因
    #         departure_reason_retrieved_data = get_list_data_from_redis("Departure_Employee_Reason_RecordView_POST")  # 离职原因
    #         departure_reason_df = pd.DataFrame(departure_reason_retrieved_data)
    #         departure_reason_df_filtered = departure_reason_df.loc[departure_reason_df['departure_reason'] != '合计']#排除合计
    #         departure_reason_sums = departure_reason_df_filtered.groupby('departure_reason')[['SAL', 'IDL', 'DL']].sum()#按departure_reason分组并计算 SAL、IDL 和 DL 的总和
    #         list_of_dicts = departure_reason_sums.reset_index().to_dict('records')  # 将 DataFrame 转换为字典列表
    #         results['departure_employee_reason_data'] = (list_of_dicts, ['departure_reason', 'SAL', 'IDL', 'DL'])
    #
    #     threads = [#线程池
    #         Thread(target=retrieve_active_employee_seniority_data),
    #         Thread(target=retrieve_active_employee_age_data),
    #         Thread(target=retrieve_active_employee_education_data),
    #         Thread(target=retrieve_active_employee_sex_data),
    #         Thread(target=retrieve_active_employee_promotion_data),
    #         Thread(target=retrieve_active_employee_nationality_data),
    #         Thread(target=retrieve_active_employee_job_grade_data),
    #         Thread(target=retrieve_departure_active_employee_seniority_data),
    #         Thread(target=retrieve_departure_active_employee_reason_data)
    #     ]
    #
    #     for thread in threads:
    #         thread.start()
    #     for thread in threads:
    #         thread.join()
    #
    #     # print(results)
    #
    #     active_employee_seniority_data, active_employee_seniority_xaxis = results['employee_seniority_data']
    #     active_employee_age_data, active_employee_age_xaxis = results['employee_age_data']
    #     active_employee_education_data, active_employee_education_xaxis = results['employee_education_data']
    #     active_employee_sex_data, active_employee_sex_xaxis = results['employee_sex_data']
    #     active_employee_promotion_data, active_employee_promotion_xaxis = results['employee_promotion_data']
    #     active_employee_promotion_data[:2] = [-num for num in active_employee_promotion_data[:2]]
    #     active_employee_nationality_data, active_employee_nationality_xaxis = results['employee_nationality_data']
    #     active_employee_job_grade_data_raw, active_employee_job_grade_xaxis_raw = results['employee_job_grade_data']
    #     active_employee_job_grade_data, active_employee_job_grade_xaxis=[sum(active_employee_job_grade_data_raw[:8]),sum(active_employee_job_grade_data_raw[8:15]),sum(active_employee_job_grade_data_raw[15:21]),sum(active_employee_job_grade_data_raw[21:])],['T-技术序列','M-管理序列','P-专业序列','O-操作序列'] #职级分布
    #     departure_active_employee_seniority_data, departure_active_employee_seniority_xaxis = results['departure_employee_seniority_data']
    #     departure_active_employee_reason_data, departure_active_employee_reason_xaxis = results['departure_employee_reason_data']
    #     end_time = datetime.now()
    #     execution_time = end_time - start_time
    #     print(f"The code executed in: {execution_time} seconds")
    #     return {"code": status.HTTP_200_OK,
    #              "msg": "信息返回成功",
    #              'active_employee_seniority': {
    #                  'title':'司龄分布',
    #                  'data': active_employee_seniority_data,
    #                  'xAxis': active_employee_seniority_xaxis
    #              },
    #             'active_employee_age': {
    #                 'title': '年龄分布',
    #                 'data': active_employee_age_data,
    #                 'xAxis': active_employee_age_xaxis
    #             },
    #              'active_employee_education': {
    #                  'title': '学历分布',
    #                  'data': active_employee_education_data,
    #                  'xAxis': active_employee_education_xaxis
    #              },
    #             'active_employee_sex': {
    #                 'title': '性别分布',
    #                 'data': active_employee_sex_data,
    #                 'xAxis': active_employee_sex_xaxis
    #             },
    #             'active_employee_job_grade': {
    #                 'title': '职级分布',
    #                 'data': active_employee_job_grade_data,
    #                 'xAxis': active_employee_job_grade_xaxis
    #             },
    #             'departure_active_employee_seniority': {
    #                 'title': '离职率分析-司龄原因',
    #                 'data': departure_active_employee_seniority_data,
    #                 'xAxis': departure_active_employee_seniority_xaxis
    #             },
    #             'active_employee_nationality': {
    #                 'title': '国籍分布',
    #                 'data': active_employee_nationality_data,
    #                 'xAxis': active_employee_nationality_xaxis
    #             },
    #             'active_employee_promotion':{
    #                 'title':'晋升情况',
    #                 'data':active_employee_promotion_xaxis,
    #                 'value':active_employee_promotion_data
    #             },
    #
    #             'departure_active_employee_reason': {
    #                 'title':'离职率分析-离职原因',
    #                 'source':departure_active_employee_reason_data,
    #                 'dimensions':departure_active_employee_reason_xaxis,
    #             },
    #              }

    def employee_profile_slices(self,slice_begin_date,slice_end_date,slice_type):

        """

        :param slice_begin_date:   切片本周开始时间
        :param slice_end_date:   切片本周结束时间
        :return:
        """




        # list_key = "{}_{}_{}".format(slice_begin_date,slice_end_date,slice_type)    #本周时间
        # request_now_week_data = get_list_data_from_redis(list_key)

        slice_begin_date = datetime.strptime(slice_begin_date, "%Y-%m-%d %H:%M:%S")
        slice_end_date = datetime.strptime(slice_end_date, "%Y-%m-%d %H:%M:%S")


        # slice_last_week_begin_date = datetime.strptime(slice_begin_date, "%Y-%m-%d %H:%M:%S")- timedelta(days=7) # 向前推 7 天
        # slice_last_week_end_date = datetime.strptime(slice_end_date, "%Y-%m-%d %H:%M:%S")- timedelta(days=7) # 向前推 7 天
        # request_last_week_data = get_list_data_from_redis("{}_{}_{}".format(slice_last_week_begin_date,slice_last_week_end_date,slice_type) )

        employee_all_list=HrEmployeeHistory.objects.filter(employee_record_begin_time=slice_begin_date,employee_record_end_time=slice_end_date,employee_record_type='1').values('id','employee_status','employee_name','employee_group_join_date','employee_departure_notice_date','employee_department__department_full_name',
                                                                                                                                                                            'employee_department__department_full_code','employee_department__department_first_name',
                                                                                                                                                                            'employee_department__department_second_name','employee_department__department_third_name')





        # 总对象
        employee_profile_total_obj = {}
        # 分布对象
        employee_profile_obj = {
            'index': '',
            'employee_department__department_full_name': '',
            'employee_department__department_first_name': '',
            'employee_department__department_second_name': '',
            'employee_department__department_third_name': '',
            'onboarding': 0,  #入职
            'dimission': 0,# 离职


        }

        # all_second_dept = list(HrDepartment.objects.filter( ~Q(id=999999),
        #     Q(department_expiry_date__isnull=True) | Q(
        #         department_expiry_date__gt=datetime.now()), department_status=1, department_level=3, ).values_list('department_full_name', flat=True))  # 二级部门
        # for employee in employee_all_list:

        # for business_dept in employee_all_list:
        #     employee_profile_total_obj[business_dept] = {
        #                     }

        # for employee in employee_all_list:
        #     department_code = employee['employee_department__department_full_code']
        #     if department_code not in employee_profile_obj:
        #         employee_profile_total_obj[department_code] = employee_profile_obj.copy()
        #         employee_profile_total_obj[department_code]['department_full_name'] = employee_profile_obj['employee_department__department_full_name']
        #         employee_profile_total_obj[department_code]['employee_department__department_first_name'] = employee_profile_obj['employee_department__department_first_name']
        #         employee_profile_total_obj[department_code]['employee_department__department_second_name'] = employee_profile_obj['employee_department__department_second_name']
        #         employee_profile_total_obj[department_code]['employee_department__department_third_name'] = employee_profile_obj['employee_department__department_third_name']
        #     print(employee_profile_total_obj)


        all_first_dept = get_list_data_from_redis('all_first_dept') # 所有一级部门
        all_first_dept = get_list_data_from_redis('all_center_dept')  # 一级部门的中心
        all_first_dept = get_list_data_from_redis('all_business_dept')  # 一级部门的事业部
        all_first_dept = get_list_data_from_redis('all_business_dept_children')  # 一级部门事业部下面的所有二级部门







        for first_full_dept in all_first_dept:
            dept, first_dept = first_full_dept.split()    #集团公司 人力资源中心
            employee_profile_total_obj[first_full_dept] = {
                                    'department_name':first_full_dept,
                                    "employee_department__department_first_name": first_dept,
                                    'employee_department__department_second_name': '',
                                    'employee_department__department_third_name': '',
                                    'onboarding': 0,  # 入职
                                    'dimission': 0,  # 离职
                                    'final_incumbency':0,   #本期期末在职
                                }
        # print('+=============================')
        for employee in employee_all_list:
            dept_full_name_list=employee['employee_department__department_full_name'].split()   #这个人的部门
            if len(dept_full_name_list)>=2:
                dept,first_dept=dept_full_name_list[0],dept_full_name_list[1]
                first_full_dept='{} {}'.format(dept,first_dept)
                # print(employee['employee_group_join_date'],type(employee['employee_group_join_date']),)
                # print(employee['employee_group_join_date'])
                if slice_begin_date<=employee['employee_group_join_date']<=slice_end_date:
                    employee_profile_total_obj[first_full_dept]['onboarding']+=1
                if employee['employee_departure_notice_date'] is not None and  slice_begin_date<=employee['employee_departure_notice_date']<=slice_end_date:    #离职通报日期
                    employee_profile_total_obj[first_full_dept]['dimission']+=1
                if employee['employee_status']=='1':
                    employee_profile_total_obj[first_full_dept]['final_incumbency'] += 1

            # print(employee['id'],employee['employee_name'],dept_full_name_list)
            # pass
        print('employee_profile_total_obj',employee_profile_total_obj)








        # all_center_dept = list(HrDepartment.objects.filter(Q(department_full_name__endswith='中心') | Q(department_full_name__endswith='研究院'),~Q(id=999999), Q(department_expiry_date__isnull=True) | Q(
        #     department_expiry_date__gt=datetime.now()), department_status=1, department_level=2,).values_list(
        #     'department_full_name', flat=True))  # 二级部门    事业部

        # print(all_dept)
        return 'request_now_week_data'






    @staticmethod
    def days_until_given_date(calculate_time):
        """
        :param datetime:    要计算据今的日期   例datetime.datetime(2023, 7, 19, 0, 0)
        :return:   days_difference 相差的天数
        """
        given_date = calculate_time  # 给定的日期
        current_date = datetime.now()  # 获取当前日期
        time_difference = current_date - given_date
        days_difference = time_difference.days  # 天数
        return days_difference
    @staticmethod
    def mkdir(path):
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pass
    @staticmethod
    def createPath(pic, path, fileName, father_file):  # 生成路径     文件对象  文件上一级目录名称 文件名称   static下面的目录名称（父目录）
        now = arrow.now().format('YYYY-MM-DD')
        file_suffix = str(pic).split(".")[-1]  # 文件后缀
        file_name = f"{fileName}.{file_suffix}"  # 文件名称
        file_path = os.path.join('static', father_file, 'report_upload_file', now, path, file_name)  # 文件路径
        file_path = file_path.replace('\\', '/')
        return (file_path, file_name, file_suffix)  # 文件路径   文件名字  文件后缀
    @staticmethod
    def saveFile(file_path, file_obj):  # 文件名,图像对象   文件保存
        with open(str(file_path), 'wb+') as f:
            for dot in file_obj.chunks():
                f.write(dot)
