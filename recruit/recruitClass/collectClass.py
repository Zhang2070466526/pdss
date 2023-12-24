# -*- coding: utf-8 -*-
# @Time    : 2023/5/12 13:19
# @Author  : zhuang
# @Site    : 
# @File    : collectClass.py
# @Software: PyCharm
import os
import datetime
import json

from django.forms.models import model_to_dict
from openpyxl import Workbook
from PIL import Image as pilImage
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
from scipy.interpolate import make_interp_spline
import matplotlib.ticker as ticker
import matplotlib.font_manager as font_manager

from openpyxl.chart import ScatterChart, Reference, Series, BarChart

from decimal import Decimal

from pdss.settings import BASE_DIR
from auther.models import *
from ..serializers import *
from ..models import *
from django.db.models import Sum
from general.models import *
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl.drawing.image import Image

from django.core.exceptions import ObjectDoesNotExist
from utils.check_token import CheckToken
import time


def verify(request):
    return_data = {'code': '', "msg": ''}
    new_token = CheckToken()
    try:
        check_token = new_token.check_token(request.headers['Authorization'])
    except Exception as e:
        # print(e)
        return_data['code'] = 400
        return_data['msg'] = '请求参数出错啦'
        return return_data
    if check_token is None:
        return_data['code'] = 403
        return_data['msg'] = '没有权限查询'
        return return_data


class collectClass:
    def __init__(self, request, meth):
        self.border_style = None
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.user_base = []
        self.border_style = Border(
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin'))
        self.methods = {
            "get_list": self.get_list,
            "download_file": self.download_file,
        }

    def meth_center(self):
        if verify(self.request) is not None:
            return Response(verify(self.request))
        self.user_base = self.request.user_base
        self.methods[self.meth]()
        return Response(self.return_data)

    def get_list(self):
        # self.user_base = [i[0] for i in center_base.objects.filter(status=1).values_list("id")]
        # print(self.user_base)
        kwargsDL = {"recruit_dl_status": 1, "recruit_dl_base__in": self.user_base}
        kwargsIDL = {"recruit_idl_status": 1, "recruit_idl_base__in": self.user_base}
        kwargsSAL = {"recruit_sal_status": 1, "recruit_sal_base__in": self.user_base}
        # baseNameId = self.request.GET.get("baseNameId", None)
        beginDate = self.request.GET.get("beginDate", None)
        endDate = self.request.GET.get("endDate", None)
        # if baseNameId != "" and baseNameId is not None:
        #     kwargsDL['recruit_dl_base_id'] = baseNameId
        #     kwargsIDL['recruit_idl_base_id'] = baseNameId
        #     kwargsSAL['recruit_sal_base_id'] = baseNameId
        if beginDate is not None and beginDate != "" and endDate is not None and beginDate != "":
            kwargsIDL['recruit_idl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsIDL['recruit_idl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
            kwargsIDL['recruit_idl_base__in'] = self.user_base
            kwargsSAL['recruit_sal_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsSAL['recruit_sal_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
            kwargsSAL['recruit_sal_base__in'] = self.user_base
            kwargsDL['recruit_dl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsDL['recruit_dl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
            kwargsDL['recruit_dl_base__in'] = self.user_base

        columnList = [
            {"value": "类别", "label": "type", "width": "60"},
            {"value": "日期", "label": "date", "width": "90"},
            {"value": "需求人数", "label": "demand_no", "width": "120"},
            {"value": "面试人数", "label": "interview_no", "width": "120"},
            {"value": "面试通过", "label": "interview_pass_no", "width": "120"},
            {"value": "入职人数", "label": "entry_no", "width": "120"},
            {"value": "待入职人数", "label": "to_entry_no", "width": "150"},
            {"value": "达成率", "label": "completion_rate", "width": ""},
        ]

        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        dataDL = RecruitDl.objects.filter(**kwargsDL).values('recruit_dl_date').annotate(
            Sum('recruit_dl_demand_no'),
            Sum('recruit_dl_interview_no'),
            Sum("recruit_dl_interview_pass_no"),
            Sum("recruit_dl_entry_no"),
            Sum("recruit_dl_to_entry_no"),
        )
        dataIDL = RecruitIdl.objects.filter(**kwargsIDL).values('recruit_idl_date').annotate(
            Sum('recruit_idl_demand_no'),
            Sum('recruit_idl_interview_no'),
            Sum("recruit_idl_interview_pass_no"),
            Sum("recruit_idl_entry_no"),
            Sum("recruit_idl_to_entry_no"),
        )
        dataSAL = RecruitSal.objects.filter(**kwargsSAL).values('recruit_sal_date').annotate(
            Sum('recruit_sal_demand_no'),
            Sum('recruit_sal_interview_no'),
            Sum("recruit_sal_interview_pass_no"),
            Sum("recruit_sal_entry_no"),
            Sum("recruit_sal_to_entry_no"),
        )
        tableList = []
        for i in dataDL:
            for key, value in i.items():
                if value is None:
                    i[key] = 0
            tableList.append(
                {
                    "type": "DL",
                    "date": i["recruit_dl_date"].strftime('%Y-%m'),
                    "demand_no": i["recruit_dl_demand_no__sum"],
                    "interview_no": i["recruit_dl_interview_no__sum"],
                    "interview_pass_no": i["recruit_dl_interview_pass_no__sum"],
                    "entry_no": i["recruit_dl_entry_no__sum"],
                    "to_entry_no": i["recruit_dl_to_entry_no__sum"],
                    "completion_rate": "{:.2f}%".format(
                        (i["recruit_dl_to_entry_no__sum"] + i["recruit_dl_entry_no__sum"]) /
                        i["recruit_dl_demand_no__sum"] * 100),
                }
            )
        for i in dataIDL:
            tableList.append(
                {
                    "type": "IDL",
                    "date": i["recruit_idl_date"].strftime('%Y-%m'),
                    "demand_no": i["recruit_idl_demand_no__sum"],
                    "interview_no": i["recruit_idl_interview_no__sum"],
                    "interview_pass_no": i["recruit_idl_interview_pass_no__sum"],
                    "entry_no": i["recruit_idl_entry_no__sum"],
                    "to_entry_no": i["recruit_idl_to_entry_no__sum"],
                    "completion_rate": "{:.2f}%".format(
                        (i["recruit_idl_to_entry_no__sum"] + i["recruit_idl_entry_no__sum"]) /
                        i["recruit_idl_demand_no__sum"] * 100),
                }
            )
        for i in dataSAL:
            tableList.append(
                {
                    "type": "SAL",
                    "date": i["recruit_sal_date"].strftime('%Y-%m'),
                    "demand_no": i["recruit_sal_demand_no__sum"],
                    "interview_no": i["recruit_sal_interview_no__sum"],
                    "interview_pass_no": i["recruit_sal_interview_pass_no__sum"],
                    "entry_no": i["recruit_sal_entry_no__sum"],
                    "to_entry_no": i["recruit_sal_to_entry_no__sum"],
                    "completion_rate": "{:.2f}%".format(
                        (i["recruit_sal_to_entry_no__sum"] + i["recruit_sal_entry_no__sum"]) /
                        i["recruit_sal_demand_no__sum"] * 100),
                }
            )

        self.return_data = {
            "data": {
                "columnList": columnList,
                "tableList": tableList,
                "totalNumber": len(dataDL) + len(dataSAL) + len(dataIDL)
            },
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
        }

    def download_file(self):
        pattern_fill_name = PatternFill(fill_type='solid', fgColor='FFFF00')
        pattern_fill_field = PatternFill(fill_type='solid', fgColor='E7E6E6')
        font_path = r'static/font/simhei.ttf'
        #
        # 设置默认字体为中文字体
        rcParams['font.family'] = 'sans-serif'
        rcParams['font.sans-serif'] = ['SimHei']  # 替换为所需的中文字体名称


        # 创建一个新的工作簿
        searchTime = datetime.now().strftime("%Y-%m-%d")
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download")):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"))
        path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"),
                            searchTime)
        chart_path = os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "img"), "chart")
        if not os.path.exists(path):
            os.mkdir(path)
        if not os.path.exists(chart_path):
            os.mkdir(chart_path)

        workbook = Workbook()

        # 创建第一个工作表
        worksheet1 = workbook.active
        worksheet1.title = "招聘"
        worksheet1['A1'] = "汇总"
        worksheet1['A1'].alignment = Alignment(horizontal="center", vertical='center')
        worksheet1['A1'].fill = pattern_fill_name
        self.collect_get_list()
        row2 = []
        for i in self.return_data['data']['columnList']:
            row2.append(i['value'])
        worksheet1.append(row2)
        for i in self.return_data['data']['tableList']:
            row = []
            for key, values in i.items():
                row.append(values)
            worksheet1.append(row)
        # 加边框
        for row in worksheet1.iter_rows():
            for cell in row:
                cell.border = self.border_style

        img = self.collect_pic()
        worksheet1.add_image(img, "A10")

        # ======================worksheet2===============================================
        # 创建第二个工作表
        worksheet2 = workbook.create_sheet(title="DL")
        self.dl_get_list(worksheet2)
        # 创建第三个工作表
        worksheet3 = workbook.create_sheet(title="IDL")
        self.idl_get_list(worksheet3)
        # 创建第四个工作表
        worksheet4 = workbook.create_sheet(title="SAL")
        self.sal_get_list(worksheet4)
        filename = "汇总表" + "".join(list(str(time.time()))[0:10])
        workbook.save("static/recruitFile/download/" + searchTime + "/" + filename + ".xlsx")

        self.return_data = {
            "data": {
                "code": status.HTTP_200_OK,
                "msg": '下载成功',
                "downloadUrl": "static/recruitFile/download/" + searchTime + "/" + filename + ".xlsx",
            }
        }

    def percentage_to_decimal(self, percentage):
        # 去除百分号，并将字符串转换为浮点数
        percentage_float = float(percentage.strip('%'))

        # 将浮点数除以100，得到小数表示
        decimal_value = Decimal(percentage_float / 100)

        return decimal_value

    def dl_get_list(self, ws):
        x = []  # 公司
        y1 = []  # 劳务
        y2 = []  # 自招
        y3 = []  # 完成率
        y4 = []  # 自招率

        pattern_fill_field = PatternFill(fill_type='solid', fgColor='E7E6E6')
        fields = {}
        letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J"]
        kwargsDL = {"recruit_dl_status": 1}
        beginDate = self.request.GET.get("beginDate", None)
        endDate = self.request.GET.get("endDate", None)
        if beginDate is not None and beginDate != "" and endDate is not None and beginDate != "":
            kwargsDL['recruit_dl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsDL['recruit_dl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
        obj = RecruitDl.objects.filter(**kwargsDL).values("recruit_dl_base").annotate(
            recruit_dl_demand_no=Sum("recruit_dl_demand_no"),
            recruit_dl_interview_no=Sum('recruit_dl_interview_no'),
            recruit_dl_interview_pass_no=Sum('recruit_dl_interview_pass_no'),
            recruit_dl_entry_no=Sum('recruit_dl_entry_no'),
            recruit_dl_to_entry_no=Sum('recruit_dl_to_entry_no'),
            recruit_dl_labor_no=Sum('recruit_dl_labor_no'),
            recruit_dl_confess_no=Sum('recruit_dl_confess_no'), )

        if len(obj) > 0:
            obj = RecruitDlDownloadSerializers(instance=obj, many=True).data
            for field in RecruitDl._meta.get_fields():
                if field.name not in ["recruit_dl_status", "creator", "modifier", "modify_time",
                                      "create_time", 'id', "recruit_dl_date"]:
                    fields[field.name] = field.verbose_name
            row2 = ["公司", "需求人数", "面试人数", "面试通过人数", "入职人数", "待入职人数", "完成率", "劳务人数", "自招人数", "自招率"]
            ws.append(row2)
            for i in letter_list:
                ws[i + '1'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '1'].fill = pattern_fill_field

            for data in obj:
                row_data = []
                y1.append(dict(data)["recruit_dl_labor_no"])
                y2.append(dict(data)["recruit_dl_confess_no"])
                y3.append(int(dict(data)["recruit_dl_completion_rate"][:-1]))
                y4.append(int(dict(data)["recruit_dl_self_rate"][:-1]))
                x.append(dict(data)["recruit_dl_base"])
                for k in fields:
                    row_data.append(dict(data)[k])
                ws.append(row_data)
            if len(x) >= 2:
                img = self.dl_pic(x, y1, y2, y3, y4)
                ws.add_image(img, "A" + str(ws.max_row + 2))

            # len_list_dl = self.count_excel_save(ws.columns)
            # for i in range(1, ws.max_column + 1):
            #     ws.column_dimensions[get_column_letter(i)].width = len_list_dl[i - 1]

            for row in ws.iter_rows():
                row = row[:-1]
                for cell in row:
                    cell.border = self.border_style

    def idl_get_list(self, ws):
        x = []  # 公司
        y1 = []  # 劳务
        y2 = []  # 自招
        y3 = []  # 完成率

        pattern_fill_field = PatternFill(fill_type='solid', fgColor='E7E6E6')
        fields = {}
        letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        kwargsIDL = {"recruit_idl_status": 1}
        beginDate = self.request.GET.get("beginDate", None)
        endDate = self.request.GET.get("endDate", None)
        if beginDate is not None and beginDate != "" and endDate is not None and beginDate != "":
            kwargsIDL['recruit_idl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsIDL['recruit_idl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
        obj = RecruitIdl.objects.filter(**kwargsIDL).values("recruit_idl_base").annotate(
            recruit_idl_demand_no=Sum("recruit_idl_demand_no"),
            recruit_idl_interview_no=Sum('recruit_idl_interview_no'),
            recruit_idl_interview_pass_no=Sum('recruit_idl_interview_pass_no'),
            recruit_idl_offer_no=Sum('recruit_idl_offer_no'),
            recruit_idl_entry_no=Sum('recruit_idl_entry_no'),
            recruit_idl_to_entry_no=Sum('recruit_idl_to_entry_no'), )
        if len(obj) > 0:
            obj = RecruitIdlDownloadSerializers(instance=obj, many=True).data
            for field in RecruitIdl._meta.get_fields():
                if field.name not in ["recruit_idl_status", "creator", "modifier", "modify_time",
                                      "create_time", 'id', "recruit_idl_date"]:
                    fields[field.name] = field.verbose_name
            row2 = ["公司", "需求人数", "面试人数", "面试通过", "Offer人数", "入职人数", "待入职人数", "达成率"]
            ws.append(row2)
            for i in letter_list:
                ws[i + '1'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '1'].fill = pattern_fill_field

            for data in obj:
                row_data = []
                y1.append(dict(data)["recruit_idl_entry_no"])
                y2.append(dict(data)["recruit_idl_to_entry_no"])
                y3.append(int(dict(data)["recruit_idl_completion_rate"][:-1]))
                x.append(dict(data)["recruit_idl_base"])
                for k in fields:
                    row_data.append(dict(data)[k])
                ws.append(row_data)
            if len(x) >= 2:
                img = self.idl_pic(x, y1, y2, y3)
                ws.add_image(img, "A" + str(ws.max_row + 2))

            # len_list_dl = self.count_excel_save(ws.columns)
            # for i in range(1, ws.max_column + 1):
            #     ws.column_dimensions[get_column_letter(i)].width = len_list_dl[i - 1]
            for row in ws.iter_rows():
                row = row[:-1]
                for cell in row:
                    cell.border = self.border_style

    def sal_get_list(self, ws):
        x = []  # 公司
        y1 = []  # 劳务
        y2 = []  # 自招
        y3 = []  # 完成率

        pattern_fill_field = PatternFill(fill_type='solid', fgColor='E7E6E6')
        fields = {}
        letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        kwargsSAL = {"recruit_sal_status": 1}
        beginDate = self.request.GET.get("beginDate", None)
        endDate = self.request.GET.get("endDate", None)
        if beginDate is not None and beginDate != "" and endDate is not None and beginDate != "":
            kwargsSAL['recruit_sal_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsSAL['recruit_sal_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
        obj = RecruitSal.objects.filter(**kwargsSAL).values("recruit_sal_base").annotate(
            recruit_sal_demand_no=Sum("recruit_sal_demand_no"),
            recruit_sal_interview_no=Sum('recruit_sal_interview_no'),
            recruit_sal_interview_pass_no=Sum('recruit_sal_interview_pass_no'),
            recruit_sal_offer_no=Sum('recruit_sal_offer_no'),
            recruit_sal_entry_no=Sum('recruit_sal_entry_no'),
            recruit_sal_to_entry_no=Sum('recruit_sal_to_entry_no'), )
        if len(obj) > 0:
            obj = RecruitSalDownloadSerializers(instance=obj, many=True).data
            for field in RecruitSal._meta.get_fields():
                if field.name not in ["recruit_sal_status", "creator", "modifier", "modify_time",
                                      "create_time", 'id', "recruit_sal_date"]:
                    fields[field.name] = field.verbose_name
            row2 = ["公司", "需求人数", "面试人数", "面试通过", "Offer人数", "入职人数", "待入职人数", "达成率"]
            ws.append(row2)
            for i in letter_list:
                ws[i + '1'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '1'].fill = pattern_fill_field

            for data in obj:
                row_data = []
                y1.append(dict(data)["recruit_sal_entry_no"])
                y2.append(dict(data)["recruit_sal_to_entry_no"])
                y3.append(int(dict(data)["recruit_sal_completion_rate"][:-1]))
                x.append(dict(data)["recruit_sal_base"])
                for k in fields:
                    row_data.append(dict(data)[k])
                ws.append(row_data)
            if len(x) >= 2:
                img = self.sal_pic(x, y1, y2, y3)
                ws.add_image(img, "A" + str(ws.max_row + 2))

            # len_list_dl = self.count_excel_save(ws.columns)
            # for i in range(1, ws.max_column + 1):
            #     ws.column_dimensions[get_column_letter(i)].width = len_list_dl[i - 1]
            for row in ws.iter_rows():
                row = row[:-1]
                for cell in row:
                    cell.border = self.border_style

    def collect_get_list(self):

        tableList = []
        kwargsDL = {}
        kwargsIDL = {}
        kwargsSAL = {"recruit_sal_base__in": self.user_base}
        beginDate = self.request.GET.get("beginDate", None)
        endDate = self.request.GET.get("endDate", None)
        if beginDate is not None and beginDate != "" and endDate is not None and beginDate != "":
            kwargsIDL['recruit_idl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsIDL['recruit_idl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
            kwargsSAL['recruit_sal_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsSAL['recruit_sal_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
            kwargsDL['recruit_dl_date__gte'] = beginDate if len(beginDate) > 7 else beginDate + "-01"
            kwargsDL['recruit_dl_date__lte'] = endDate if len(endDate) > 7 else endDate + "-01"
        columnList = [
            {"value": "类别", "label": "type", "width": "60"},
            {"value": "需求人数", "label": "demand_no", "width": "120"},
            {"value": "面试人数", "label": "interview_no", "width": "120"},
            {"value": "面试通过", "label": "interview_pass_no", "width": "120"},
            {"value": "入职人数", "label": "entry_no", "width": "120"},
            {"value": "待入职人数", "label": "to_entry_no", "width": "150"},
            {"value": "达成率", "label": "completion_rate", "width": ""},
        ]

        # currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
        #                                                                                       None) != "" else 1
        # pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        try:
            dataDL = RecruitDl.objects.filter(recruit_dl_status=1, **kwargsDL).aggregate(
                Sum('recruit_dl_demand_no'),
                Sum('recruit_dl_interview_no'),
                Sum("recruit_dl_interview_pass_no"),
                Sum("recruit_dl_entry_no"),
                Sum("recruit_dl_to_entry_no"),
            )
            tableList.append({
                "type": "DL",
                "demand_no": dataDL["recruit_dl_demand_no__sum"],
                "interview_no": dataDL["recruit_dl_interview_no__sum"],
                "interview_pass_no": dataDL["recruit_dl_interview_pass_no__sum"],
                "entry_no": dataDL["recruit_dl_entry_no__sum"],
                "to_entry_no": dataDL["recruit_dl_to_entry_no__sum"],
                "completion_rate": "{:.2f}%".format(
                    (dataDL["recruit_dl_to_entry_no__sum"] + dataDL["recruit_dl_entry_no__sum"]) /
                    dataDL["recruit_dl_demand_no__sum"] * 100),
            }, )
        except TypeError as e:
            tableList.append({
                "type": "DL",
                "demand_no": 0,
                "interview_no": 0,
                "interview_pass_no": 0,
                "entry_no": 0,
                "to_entry_no": 0,
                "completion_rate": "0%",
            }, )
            # print("DL数据为空")
        try:
            dataIDL = RecruitIdl.objects.filter(recruit_idl_status=1, **kwargsIDL).aggregate(
                Sum('recruit_idl_demand_no'),
                Sum('recruit_idl_interview_no'),
                Sum("recruit_idl_interview_pass_no"),
                Sum("recruit_idl_entry_no"),
                Sum("recruit_idl_to_entry_no"),
            )
            tableList.append({
                "type": "IDL",
                "demand_no": dataIDL["recruit_idl_demand_no__sum"],
                "interview_no": dataIDL["recruit_idl_interview_no__sum"],
                "interview_pass_no": dataIDL["recruit_idl_interview_pass_no__sum"],
                "entry_no": dataIDL["recruit_idl_entry_no__sum"],
                "to_entry_no": dataIDL["recruit_idl_to_entry_no__sum"],
                "completion_rate": "{:.2f}%".format(
                    (dataIDL["recruit_idl_to_entry_no__sum"] + dataIDL["recruit_idl_entry_no__sum"]) /
                    dataIDL["recruit_idl_demand_no__sum"] * 100),
            }, )
        except:
            tableList.append({
                "type": "IDL",
                "demand_no": 0,
                "interview_no": 0,
                "interview_pass_no": 0,
                "entry_no": 0,
                "to_entry_no": 0,
                "completion_rate": "0%",
            }, )
            # print("IDL数据为空")
        try:
            dataSAL = RecruitSal.objects.filter(recruit_sal_status=1, **kwargsSAL).aggregate(
                Sum('recruit_sal_demand_no'),
                Sum('recruit_sal_interview_no'),
                Sum("recruit_sal_interview_pass_no"),
                Sum("recruit_sal_entry_no"),
                Sum("recruit_sal_to_entry_no"),
            )
            tableList.append({
                "type": "SAL",
                "demand_no": dataSAL["recruit_sal_demand_no__sum"],
                "interview_no": dataSAL["recruit_sal_interview_no__sum"],
                "interview_pass_no": dataSAL["recruit_sal_interview_pass_no__sum"],
                "entry_no": dataSAL["recruit_sal_entry_no__sum"],
                "to_entry_no": dataSAL["recruit_sal_to_entry_no__sum"],
                "completion_rate": "{:.2f}%".format(
                    (dataSAL["recruit_sal_to_entry_no__sum"] + dataSAL["recruit_sal_entry_no__sum"]) /
                    dataSAL["recruit_sal_demand_no__sum"] * 100),
            })
        except:
            tableList.append({
                "type": "SAL",
                "demand_no": 0,
                "interview_no": 0,
                "interview_pass_no": 0,
                "entry_no": 0,
                "to_entry_no": 0,
                "completion_rate": "0%",
            }, )
            # print("SAL数据为空")
        self.return_data = {
            "data": {
                "columnList": columnList,
                "tableList": tableList,
            }
        }

    def count_excel_save(self, excel_columns):

        len_list = []
        prjTuple = tuple(excel_columns)
        # print(prjTuple)
        for idx in range(0, len(prjTuple)):
            length2 = 0
            str_list = []
            for cell in prjTuple[idx]:
                if "1" not in str(cell):
                    str_list.append(str(cell.value))
            for elem in str_list:
                elem_split = list(elem)
                length = 0
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1.5
                    else:
                        length += 2.5
                length2 = max(length, length2)
            len_list.append(length2)
        return len_list

    def count_character(self, s):
        hanzi = 0
        num = 0
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1

        return str(30 * hanzi + 20 * num)

    def dl_pic(self, x, y1, y2, y3, y4):
        # 创建Figure和Axes对象
        fig, ax1 = plt.subplots(figsize=(16, 6), dpi=300)
        # 绘制柱状图
        bar_width = 0.6
        bar1 = np.arange(len(x))

        ax1.bar(bar1, y1, width=bar_width, label='劳务', alpha=0.7, color="#ED7D31")
        ax1.bar(bar1, y2, width=bar_width, label='自招', alpha=0.7, color="#9DC3E6", bottom=y1)
        ax1.set_xticks(bar1)
        ax1.set_xticklabels(x)
        ax1.tick_params(axis='y')

        ax2 = ax1.twinx()
        x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        y_smooth = make_interp_spline(bar1, y3, bc_type='natural')(x_smooth)
        ax2.plot(x_smooth, y_smooth, color='#92D050', linestyle='-', label='完成率')
        x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        y_smooth = make_interp_spline(bar1, y4, bc_type='natural')(x_smooth)
        ax2.plot(x_smooth, y_smooth, color='#0070C0', linestyle='-', label='自招率')
        ax2.tick_params(axis='y')

        ax2.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=100))
        ax2.set_ylim(0, max(max(y3), max(y4)) + 20 if max(max(y3), max(y4)) >= 100 else 100)

        for i, v in enumerate(y1):
            ax1.text(bar1[i] - 0.05, v - 10, str(v))
        for i, v in enumerate(y2):
            ax1.text(bar1[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(y3):
            ax2.text(bar1[i], v + 1, f'{v:.0f}%')
        for i, v in enumerate(y4):
            ax2.text(bar1[i], v + 1, f'{v:.0f}%')

        # 获取所有图例句柄和标签
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        # 合并所有图例句柄和标签
        all_lines = lines1 + lines2
        all_labels = labels1 + labels2
        # 设置图例位置在底部中心，并排显示
        ax1.legend(all_lines, all_labels, loc='upper center', bbox_to_anchor=(0.5, -0.2), ncol=4)
        # 设置标题

        fig.tight_layout()
        tmp_filename = "static/img/chart/chart2" + "".join(list(str(time.time()))[0:10]) + ".png"
        plt.savefig(tmp_filename, dpi=300, )
        # 显示图表
        plt.show()
        plt.close()

        # img = self.img_shape(tmp_filename)
        img = Image(tmp_filename)
        return img

    def idl_pic(self, x, y1, y2, y3):
        # 数据
        # x = ['润阳新能源', '宁夏硅材料', '内蒙古硅材料', '润阳建湖', '润阳世纪', '润阳泰国', '泰国四期', '云南曲靖', '润阳悦达&建湖二期', '海博瑞', '泰国WHA']
        # y1 = [10, 15, 12, 8, 20, 18, 14, 16, 10, 12, 9]  # 劳务
        # y2 = [8, 10, 6, 4, 12, 10, 8, 12, 8, 6, 5]  # 自招
        # y3 = [80, 90, 70, 60, 85, 78, 75, 82, 76, 68, 72]  # 完成率
        # y4 = [60, 70, 50, 40, 65, 55, 60, 68, 62, 58, 55]  # 自招率

        # 创建Figure和Axes对象
        fig, ax1 = plt.subplots(figsize=(16, 6), dpi=300)
        # 绘制柱状图
        bar_width = 0.6
        bar1 = np.arange(len(x))

        ax1.bar(bar1, y1, width=bar_width, label='入职人数', alpha=0.7, color="#2F5597")
        ax1.bar(bar1, y2, width=bar_width, label='待入职人数', alpha=0.7, color="#70AD47", bottom=y1)
        ax1.set_xticks(bar1)
        ax1.set_xticklabels(x)
        ax1.tick_params(axis='y')

        ax2 = ax1.twinx()
        x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        y_smooth = make_interp_spline(bar1, y3, bc_type='natural')(x_smooth)
        ax2.plot(x_smooth, y_smooth, color='#92D050', linestyle='-', label='达成率')
        # x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        # y_smooth = make_interp_spline(bar1, y4, bc_type='natural')(x_smooth)
        # ax2.plot(x_smooth, y_smooth, color='#0070C0', linestyle='-', label='自招率')
        ax2.tick_params(axis='y')

        ax2.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=100))
        ax2.set_ylim(0, max(y3) + 20 if max(y3) >= 100 else 100)

        for i, v in enumerate(y1):
            ax1.text(bar1[i] - 0.05, v - 10, str(v))
        for i, v in enumerate(y2):
            ax1.text(bar1[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(y3):
            ax2.text(bar1[i], v + 1, f'{v:.0f}%')
        # for i, v in enumerate(y4):
        #     ax2.text(bar1[i], v + 1, f'{v:.0f}%')

        # 获取所有图例句柄和标签
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        # 合并所有图例句柄和标签
        all_lines = lines1 + lines2
        all_labels = labels1 + labels2
        # 设置图例位置在底部中心，并排显示
        ax1.legend(all_lines, all_labels, loc='upper center', bbox_to_anchor=(0.5, -0.2), ncol=4)
        # 设置标题

        fig.tight_layout()
        tmp_filename = "static/img/chart/chart2" + "".join(list(str(time.time()))[0:10]) + ".png"
        plt.savefig(tmp_filename, dpi=300, )
        # 显示图表
        plt.show()
        plt.close()

        # img = self.img_shape(tmp_filename)
        img = Image(tmp_filename)
        return img

    def sal_pic(self, x, y1, y2, y3):
        # 创建Figure和Axes对象
        fig, ax1 = plt.subplots(figsize=(16, 6), dpi=300)
        # 绘制柱状图
        bar_width = 0.6
        bar1 = np.arange(len(x))

        ax1.bar(bar1, y1, width=bar_width, label='入职人数', alpha=0.7, color="#2F5597")
        ax1.bar(bar1, y2, width=bar_width, label='待入职人数', alpha=0.7, color="#70AD47", bottom=y1)
        ax1.set_xticks(bar1)
        ax1.set_xticklabels(x)
        ax1.tick_params(axis='y')

        ax2 = ax1.twinx()
        x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        y_smooth = make_interp_spline(bar1, y3, bc_type='natural')(x_smooth)
        ax2.plot(x_smooth, y_smooth, color='#92D050', linestyle='-', label='达成率')
        # x_smooth = np.linspace(bar1[0], bar1[-1], 300)
        # y_smooth = make_interp_spline(bar1, y4, bc_type='natural')(x_smooth)
        # ax2.plot(x_smooth, y_smooth, color='#0070C0', linestyle='-', label='自招率')
        ax2.tick_params(axis='y')

        ax2.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=100))
        ax2.set_ylim(0, max(y3) + 20 if max(y3) >= 100 else 100)

        for i, v in enumerate(y1):
            ax1.text(bar1[i] - 0.05, v - 10, str(v))
        for i, v in enumerate(y2):
            ax1.text(bar1[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(y3):
            ax2.text(bar1[i], v + 1, f'{v:.0f}%')
        # for i, v in enumerate(y4):
        #     ax2.text(bar1[i], v + 1, f'{v:.0f}%')

        # 获取所有图例句柄和标签
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        # 合并所有图例句柄和标签
        all_lines = lines1 + lines2
        all_labels = labels1 + labels2
        # 设置图例位置在底部中心，并排显示
        ax1.legend(all_lines, all_labels, loc='upper center', bbox_to_anchor=(0.5, -0.2), ncol=4)
        # 设置标题

        fig.tight_layout()
        tmp_filename = "static/img/chart/chart2" + "".join(list(str(time.time()))[0:10]) + ".png"
        plt.savefig(tmp_filename, dpi=300)
        # 显示图表
        plt.show()
        plt.close()

        # img = self.img_shape(tmp_filename)
        img = Image(tmp_filename)
        return img

    def collect_pic(self):
        # print(self.return_data['data']['tableList'])
        # 示例数据
        categories = ['DL', 'IDL', 'SAL']
        # 需求人数
        demand_values = [i["demand_no"] for i in self.return_data['data']['tableList']]
        # 入职人数
        join_values = [i["entry_no"] for i in self.return_data['data']['tableList']]
        # 待入职人数
        pending_values = [i["to_entry_no"] for i in self.return_data['data']['tableList']]
        #
        achievement_values = [float(i["completion_rate"][0:-1]) for i in self.return_data['data']['tableList']]
        # achievement_values = [80, 85, 78]
        # 创建Figure和Axes对象
        fig, ax1 = plt.subplots(dpi=300)
        # 绘制柱状图
        bar_width = 0.2
        bar1 = np.arange(len(categories))
        bar2 = [x + bar_width for x in bar1]
        bar3 = [x + bar_width for x in bar2]
        ax1.bar(bar1, demand_values, width=bar_width, label='需求人数', alpha=0.7)
        ax1.bar(bar2, join_values, width=bar_width, label='入职人数', alpha=0.7)
        ax1.bar(bar3, pending_values, width=bar_width, label='待入职人数', alpha=0.7)
        ax1.set_xticks(bar2)
        ax1.set_xticklabels(categories)
        ax1.tick_params(axis='y')
        # 创建第二个y轴并绘制光滑曲线图,共享x轴
        ax2 = ax1.twinx()
        # 使用插值方法生成光滑曲线
        x_smooth = np.linspace(bar2[0], bar2[-1], 300)
        y_smooth = make_interp_spline(bar2, achievement_values, bc_type='natural')(x_smooth)
        ax2.plot(x_smooth, y_smooth, color='red', linestyle='-', label='达成率')
        ax2.tick_params(axis='y')
        # 设置左侧y轴刻度为百分比，从0%开始，每次增加20%
        # ax2.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=max(demand_values), decimals=0, symbol='%'))
        # ax2.set_yticks(np.arange(0, max(demand_values) + 1, 20))
        # 设置右侧y轴刻度为百分比
        ax2.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=100))
        ax2.set_ylim(0, max(achievement_values) + 20 if max(achievement_values) >= 100 else 100)
        # 显示柱状图和曲线图数据标签
        for i, v in enumerate(demand_values):
            ax1.text(bar1[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(join_values):
            ax1.text(bar2[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(pending_values):
            ax1.text(bar3[i] - 0.05, v + 1, str(v))
        for i, v in enumerate(achievement_values):
            ax2.text(bar2[i], v + 1, f'{v:.0f}%')
        # 获取所有图例句柄和标签
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        # 合并所有图例句柄和标签
        all_lines = lines1 + lines2
        all_labels = labels1 + labels2
        # 设置图例位置在底部中心，并排显示
        ax1.legend(all_lines, all_labels, loc='lower center', bbox_to_anchor=(0.5, -0.2), ncol=4)
        # 设置标题
        ax1.set_title('集团招聘数据', fontname='SimHei')
        # 调整图表布局
        fig.tight_layout()
        tmp_filename = "static/img/chart/chart" + "".join(list(str(time.time()))[0:10]) + ".png"
        plt.savefig(tmp_filename, dpi=300, )
        # 显示图表
        plt.show()
        plt.close()
        # img = self.img_shape(tmp_filename)
        img = Image(tmp_filename)

        return img

    def img_shape(self, image_path):
        path = "static/img/chart/change_chart" + "".join(list(str(time.time()))[0:10]) + ".png"
        # 打开原始图片
        image = pilImage.open(image_path)

        # 设置目标尺寸
        target_width = 1200
        target_height = 600

        # 调整图像尺寸
        resized_image = image.resize((target_width, target_height))
        # 将缩小后的图像插入到工作表中
        resized_image.save(path)
        return path
