# -*- coding: utf-8 -*-
# @Time    : 2023/5/11 14:16
# @Author  : zhuang
# @Site    : 
# @File    : idlClass.py
# @Software: PyCharm
import os
import datetime
import json
import re
import traceback

from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, Side, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from auther.models import *
from ..serializers import *
from ..models import *
from general.models import *
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist

from controller.controller import Controller, upload_file
import time


class idlClass(Controller):
    def __init__(self, model, meth, request):
        super().__init__(model, meth, request)

    def get_upload(self):
        file = self.request.FILES.get("file", None)

        createData = self.request.POST.get("createData", None)
        # print(createData)
        searchTime = datetime.now().strftime("%Y-%m-%d")
        if file and file is not None:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                self.fileName = upload_file(file, "recruitFile", "upload", None)
                if self.fileName:
                    self.save_Data()
            except ObjectDoesNotExist as e:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "中心/基地名不存在"
                }
            except Exception as e:
                # print(e)
                # traceback.print_exc()
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
                if "cannot be null" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = RecruitIdl._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 不能为空,请检查!"

                if "expected a number" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = RecruitIdl._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 的数据格式有误，应为数字格式"
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            createData = eval(createData)
            try:
                createData['recruit_idl_date'] = datetime.strptime(createData['recruit_idl_date'][0:7], "%Y-%m")
            except Exception as e:
                createData['recruit_idl_date'] = datetime.strptime(createData['recruit_idl_date'], "%Y-%m")
                # print(e)
            createData = {key: value for key, value in createData.items() if value != ""}
            for i in createData:
                if "rate" in i:
                    percentage = float(createData[i].strip("%"))
                    createData[i] = percentage / 100
            # print(createData)
            obj = RecruitIdl.objects.create(**createData)

            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "用户数据新增成功"
            }

    def save_Data(self):
        workbook = openpyxl.load_workbook(self.fileName, data_only=True)
        table = workbook.active
        # upload_time = self.request.POST.get("uploadTime", None)
        for i in range(2, table.max_row):
            if table.cell(i + 1, 1).value != "" or table.cell(i + 1, 1).value is None:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "请在A1单元格填写日期，如2023-05"
                }
                return
            if table.cell(i + 1, 2).value != "" and table.cell(i + 1, 2).value is not None:
                try:
                    employee_base_father = center_base.objects.get(status=1,
                                                                   name=table.cell(i + 1, 1).value.replace("（",
                                                                                                           "(").replace(
                                                                       "）", ")"))
                except ObjectDoesNotExist:
                    self.return_data = {
                        "code": status.HTTP_400_BAD_REQUEST,
                        "msg": "基地名不存在，请重新选择"
                    }
                    return
                obj = center_base.objects.filter(base_parent_id=employee_base_father.id, status=1)
                if obj.exists():
                    # print("2", table.cell(i + 1, 2).value)
                    employee_base_son = center_base.objects.get(status=1, name=table.cell(i + 1, 2).value.replace("（",
                                                                                                                  "(").replace(
                        "）", ")"))
                else:
                    employee_base_son = employee_base_father
                if employee_base_father.id not in self.user_base or employee_base_son.id not in self.user_base:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "抱歉，您没有上传 " + table.cell(i + 1, 1).value + "-" + table.cell(i + 1,
                                                                                           2).value + " (基地/中心/公司)的权限"
                    }
                else:
                    data_kwargs = {
                        "recruit_idl_base": employee_base_son,
                        "recruit_idl_date": datetime.strptime(table.cell(1, 1).value, "%Y-%m"),
                        "recruit_idl_demand_no": table.cell(i + 1, 3).value,
                        "recruit_idl_interview_no": table.cell(i + 1, 4).value,
                        "recruit_idl_interview_pass_no": table.cell(i + 1, 5).value,
                        "recruit_idl_offer_no": table.cell(i + 1, 6).value,
                        "recruit_idl_entry_no": table.cell(i + 1, 7).value,
                        "recruit_idl_to_entry_no": table.cell(i + 1, 8).value,
                        "recruit_idl_completion_rate": round((int(table.cell(i + 1, 8).value)+int(table.cell(i + 1, 7).value))/int(table.cell(i + 1, 3).value), 2),
                        "recruit_idl_remark": table.cell(i + 1, 10).value,
                        "recruit_idl_status": 1,
                        "creator": AdminUser.objects.filter(pk=self.check_token)[0],
                        "modifier": AdminUser.objects.filter(pk=self.check_token)[0],
                    }
                    RecruitIdl.objects.update_or_create(**data_kwargs)

    def download_file(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功"
        }
        info = json.loads(self.request.body)
        get_obj = self.request.GET
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25

        kwargs = {
            'recruit_idl_status': 1,
            "recruit_idl_base__in": self.user_base
        }

        searchTime = datetime.now().strftime("%Y-%m-%d")
        filename = "IDL招聘"
        id_list = info['idList']
        is_all = info['downloadAll']
        fields = {}
        if not os.path.exists(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile")):
            os.mkdir(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"))
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download")):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"))
        path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"),
                            searchTime)
        if not os.path.exists(path):
            os.mkdir(path)
        if is_all == 0 and len(id_list) <= 0:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "条件为空"
            }
            return

        if id_list is not None and len(id_list) > 0:
            kwargs["id__in"] = list(id_list)

        if is_all:
            kwargs = {'recruit_idl_status': 1}
            model_fields = {"beginDate": "recruit_idl_date__gte",
                            "endDate": "recruit_idl_date__lte",
                            "baseNameId": "recruit_idl_base_id",
                            }
            model_key = ["beginDate", "baseNameId", "endDate"]
            for key, value in get_obj.items():
                if value != "" and key in model_key:
                    kwargs[model_fields[key]] = value
        obj = RecruitIdl.objects.filter(**kwargs).order_by("create_time")
        if len(obj) > 0:
            obj = RecruitIdlPutSerializers(instance=obj, many=True).data
            for field in RecruitIdl._meta.get_fields():
                if field.name not in ["recruit_idl_status", "creator", "modifier", "modify_time",
                                      "create_time", 'id']:
                    fields[field.verbose_name] = field.name

            letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J', 'K']
            side = Side(style="thick", color="000000")
            percentage_format = '0.00%'
            wb = Workbook()
            ws = wb.active
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
            font_header = Font(name="等线", size=14, bold=True)

            pattern_fill = PatternFill(fill_type='solid', fgColor='FFFF00')

            ws['A1'] = "IDL招聘"
            ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
            ws['A1'].fill = pattern_fill

            row2 = [key for key, value in fields.items()]
            # row2[1] = "公司"
            row2.insert(1, "中心/事业部")
            fields["中心/事业部"] = "recruit_idl_base_father"
            ws.append(row2)
            for i in letter_list:
                ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
                ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
            index = 1
            for data in obj:
                row_data = []
                for k in row2:
                    row_data.append(dict(data)[fields[k]])
                index += 1
                ws.append(row_data)
            max_row = ws.max_row
            # ws.merge_cells(start_row=max_row + 1, start_column=1, end_row=max_row + 1, end_column=2)
            # ws["A" + str(max_row + 1)] = "合计"
            # ws["A" + str(max_row + 1)].alignment = Alignment(horizontal="center", vertical='center')

            excel_columns = ws.columns
            # font = Font(color="FF0000", italic=True)
            # for i in range(3, 10):
            #     column_letter = get_column_letter(i)
            #     formula_cell = ws.cell(row=max_row + 1, column=i)
            #     if i == 9:
            #         formula_cell.number_format = percentage_format
            #         formula = f"=(G{max_row + 1}+H{max_row + 1})/C{max_row + 1}"
            #     else:
            #         formula = f"=SUM({column_letter}3:{column_letter}{max_row})"
            #     formula_cell.value = formula

            len_list = self.count_excel_save(excel_columns)

            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
            # ws.column_dimensions["A"].width = 7
            wb.close()
            filename = filename + "".join(list(str(time.time()))[0:10])

            wb.save("static/recruitFile/download/" + searchTime + '/' + filename + ".xlsx")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '下载成功',
                "downloadUrl": "static/recruitFile/download/" + searchTime + '/' + filename + ".xlsx",
            }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '无该用户数据'
            }
