# -*- coding: utf-8 -*-
# @Time    : 2023/5/10 15:44
# @Author  : zhuang
# @Site    : 
# @File    : dlClass.py
# @Software: PyCharm
import os
import datetime
import json
import re

from openpyxl.styles import numbers
from django.forms.models import model_to_dict
import traceback
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, Side, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from auther.models import *
from ..serializers import *
from ..models import *
from general.models import *
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist

from controller.controller import Controller, upload_file
import time
from utils.check_token import CheckToken


def verify(request):
    return_data = {'code': '', "msg": ''}
    new_token = CheckToken()
    try:
        check_token = new_token.check_token(request.headers['Authorization'])
    except Exception as e:
        # print(e)
        return_data['code'] = 400
        return_data['msg'] = '请求参数出错啦'
        return return_data
    if check_token is None:
        return_data['code'] = 403
        return_data['msg'] = '没有权限查询'
        return return_data


class dlClass:
    def __init__(self, request, meth):
        self.user_base = None
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.check_token = request.check_token
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }

    def meth_center(self):
        return_data = {'code': '', "msg": ''}
        if self.check_token is None:
            return_data['code'] = 400
            return_data['msg'] = '请求参数出错啦'
            return Response(return_data)
        self.user_base = self.request.user_base
        self.methods[self.meth]()
        return Response(self.return_data)

    def get_upload(self):
        file = self.request.FILES.get("file", None)

        createData = self.request.POST.get("createData", None)
        searchTime = datetime.now().strftime("%Y-%m-%d")
        if file and file is not None:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                self.fileName = upload_file(file, "recruitFile", "upload", None)
                if self.fileName:
                    self.save_Data()
            except ObjectDoesNotExist as e:
                # print(e)
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "中心/基地名不存在"
                }
            except Exception as e:
                # traceback.print_exc()
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
                if "cannot be null" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = RecruitDl._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 不能为空,请检查!"

                if "expected a number" in str(e):
                    field = re.search(r"'(\w+)'", str(e)).group(1)
                    verbose_name = RecruitDl._meta.get_field(field).verbose_name
                    self.return_data["msg"] = verbose_name + " 的数据格式有误，应为数字格式"

        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            createData = eval(createData)
            try:
                createData['recruit_dl_date'] = datetime.strptime(createData['recruit_dl_date'][0:7], "%Y-%m")

            except Exception as e:
                createData['recruit_dl_date'] = datetime.strptime(createData['recruit_dl_date'], "%Y-%m")
                # print(e)
            createData = {key: value for key, value in createData.items() if value != ""}
            for i in createData:
                if "rate" in i:
                    percentage = float(createData[i].strip("%"))
                    createData[i] = percentage / 100
            obj = RecruitDl.objects.create(**createData)
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "用户数据新增成功"
            }

    def get_list(self):
        obj = Controller(RecruitDl, "get_list", self.request)
        self.return_data = obj.data_start()

    # 将上传的文件数据保存到数据库中
    def save_Data(self):
        workbook = openpyxl.load_workbook(self.fileName, data_only=True)
        table = workbook.active
        # print(table.cell(1, 1).value)
        if table.cell(1, 1).value == "" or table.cell(1, 1).value is None:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "请在A1单元格填写日期，如2023-05"
            }
            return
        f = r"[（）()]"
        # upload_time = self.request.POST.get("uploadTime", None)
        for i in range(2, table.max_row):
            if table.cell(i + 1, 1).value != "" and table.cell(i + 1, 1).value is not None:
                # print("1", table.cell(i + 1, 1).value)
                try:
                    employee_base_father = center_base.objects.get(status=1,
                                                                   name=table.cell(i + 1, 1).value.replace("（",
                                                                                                           "(").replace(
                                                                       "）", ")"))
                except ObjectDoesNotExist:
                    self.return_data = {
                        "code": status.HTTP_400_BAD_REQUEST,
                        "msg": "第%s行基地名不存在，请重新选择" % str(i + 1)
                    }
                    return
                obj = center_base.objects.filter(base_parent_id=employee_base_father.id, status=1)
                if obj.exists():
                    employee_base_son = center_base.objects.get(status=1, name=table.cell(i + 1, 2).value.replace("（",
                                                                                                                  "(").replace(
                        "）", ")"))
                else:
                    employee_base_son = employee_base_father
                if employee_base_father.id not in self.user_base or employee_base_son.id not in self.user_base:
                    self.return_data = {
                        "code": status.HTTP_403_FORBIDDEN,
                        "msg": "抱歉，您没有上传 " + table.cell(i + 1, 1).value + "-" + table.cell(i + 1,
                                                                                           2).value + " (基地/中心/公司)的权限"
                    }
                else:
                    data_kwargs = {
                        "recruit_dl_base": employee_base_son,
                        "recruit_dl_date": datetime.strptime(table.cell(1, 1).value, "%Y-%m"),
                        "recruit_dl_demand_no": table.cell(i + 1, 3).value,
                        "recruit_dl_interview_no": table.cell(i + 1, 4).value,
                        "recruit_dl_interview_pass_no": table.cell(i + 1, 5).value,
                        "recruit_dl_entry_no": table.cell(i + 1, 6).value,
                        "recruit_dl_to_entry_no": table.cell(i + 1, 7).value,
                        "recruit_dl_completion_rate": round((int(table.cell(i + 1, 6).value)+int(table.cell(i + 1, 7).value))/int(table.cell(i + 1, 3).value), 2),
                        "recruit_dl_labor_no": table.cell(i + 1, 9).value,
                        "recruit_dl_confess_no": table.cell(i + 1, 10).value,
                        "recruit_dl_self_rate": round((int(table.cell(i + 1, 6).value)/int(table.cell(i + 1, 10).value)), 2),
                        "recruit_dl_remark": table.cell(i + 1, 12).value,
                        "creator": AdminUser.objects.filter(pk=self.check_token)[0],
                        "modifier": AdminUser.objects.filter(pk=self.check_token)[0],
                    }
                    RecruitDl.objects.create(**data_kwargs)
                    self.return_data = {
                        "code": status.HTTP_200_OK,
                        "msg": "上传成功"
                    }

    # 下载文件
    def download_file(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "下载成功"
        }
        info = json.loads(self.request.body)
        get_obj = self.request.GET
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25

        kwargs = {
            'recruit_dl_status': 1,
            "recruit_dl_base__in": self.user_base,
        }

        searchTime = datetime.now().strftime("%Y-%m-%d")
        filename = "DL招聘"
        id_list = info['idList']
        is_all = info['downloadAll']
        fields = {}
        if not os.path.exists(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile")):
            os.mkdir(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"))
        if not os.path.exists(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download")):
            os.mkdir(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"))
        path = os.path.join(os.path.join(os.path.join(os.path.join(BASE_DIR, "static"), "recruitFile"), "download"),
                            searchTime)
        if not os.path.exists(path):
            os.mkdir(path)
        if is_all == 0 and len(id_list) <= 0:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "条件为空"
            }
            return

        if id_list is not None and len(id_list) > 0:
            kwargs["id__in"] = list(id_list)

        if is_all:
            kwargs = {'recruit_dl_status': 1}
            model_fields = {
                "beginDate": "recruit_dl_date__gte",
                "endDate": "recruit_dl_date__lte",
                "baseNameId": "recruit_dl_base_id",
            }
            model_key = ["beginDate", "baseNameId", "endDate"]
            for key, value in get_obj.items():
                if value != "" and key in model_key:
                    kwargs[model_fields[key]] = value
        for key, value in kwargs.items():
            if "recruit_dl_date" in key and len(value) <= 7:
                # print(1)
                kwargs[key] = value + "-01"
        obj = RecruitDl.objects.filter(**kwargs).order_by("create_time")
        if len(obj) > 0:
            obj = RecruitDlPutSerializers(instance=obj, many=True).data
            for field in RecruitDl._meta.get_fields():
                if field.name not in ["recruit_dl_status", "creator", "modifier", "modify_time",
                                      "create_time", 'id']:
                    fields[field.verbose_name] = field.name

            letter_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', "L"]
            side = Side(style="thick", color="000000")
            percentage_format = '0.00%'
            wb = Workbook()
            ws = wb.active
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=7)
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
            font_header = Font(name="等线", size=14, bold=True)

            pattern_fill = PatternFill(fill_type='solid', fgColor='FFFF00')

            ws['A1'] = "DL招聘"
            ws['A1'].alignment = Alignment(horizontal="center", vertical='center')
            ws['A1'].fill = pattern_fill

            ws['C1'] = "达成情况"
            ws['C1'].alignment = Alignment(horizontal="center", vertical='center')

            ws['H1'] = '来源'
            ws['H1'].alignment = Alignment(horizontal="center", vertical='center')

            # row2 = [value for key, value in fields.items()]
            row2 = ["日期", "中心/事业部", "公司", "需求人数", "面试人数", "面试通过人数", "入职人数", "待入职人数", "完成率", "劳务人数", "自招人数", "自招率"]
            # row2.append("check")
            ws.append(row2)
            for i in letter_list:
                ws[i + '2'].alignment = Alignment(horizontal="center", vertical='center')
                ws[i + '2'].font = Font(name="黑体", size=10, bold=True)
                ws[i + '2'].border = Border(top=side, bottom=side, left=side, right=side)
            index = 1
            fields["中心/事业部"] = "recruit_dl_base_father"
            for data in obj:
                row_data = []
                for k in row2:
                    row_data.append(dict(data)[fields[k]])
                index += 1
                ws.append(row_data)
                # formula_cell = ws.cell(row=index + 1, column=12)
                # formula = f"=(G{index + 1}-I{index + 1})-J{index + 1}"
                # formula_cell.value = formula
            # max_row = ws.max_row
            # ws.merge_cells(start_row=max_row + 1, start_column=1, end_row=max_row + 1, end_column=2)
            # ws["A" + str(max_row + 1)] = "合计"
            # ws["A" + str(max_row + 1)].alignment = Alignment(horizontal="center", vertical='center')

            excel_columns = ws.columns
            # font = Font(color="FF0000", italic=True)
            # for i in range(3, 12):
            #     column_letter = get_column_letter(i)
            #     formula_cell = ws.cell(row=max_row + 1, column=i)
            #     if i == 8:
            #         formula_cell.number_format = percentage_format
            #         formula = f"=(G{max_row + 1}+F{max_row + 1})/D{max_row + 1}"
            #     elif i == 11:
            #         formula_cell.number_format = percentage_format
            #         formula = f"=(J{max_row + 1}/F{max_row + 1})"
            #     elif i == 12:
            #         formula = f"=(G{max_row + 1}-I{max_row + 1}-J{max_row + 1})"
            #     else:
            #         formula = f"=SUM({column_letter}3:{column_letter}{max_row})"
            #     formula_cell.value = formula

            len_list = self.count_excel_save(excel_columns)
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = len_list[i - 1]
            wb.close()
            filename = filename + "".join(list(str(time.time()))[0:10])

            wb.save("static/recruitFile/download/" + searchTime + '/' + filename + ".xlsx")
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '下载成功',
                "downloadUrl": "static/recruitFile/download/" + searchTime + '/' + filename + ".xlsx",
            }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": '无该用户数据'
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(RecruitDl, "patch", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def delete_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(RecruitDl, "delete", self.request)
            obj.start()
        except Exception as e:
            # print(e)
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def count_character(self, s):
        hanzi = 0
        num = 0
        for i in str(s):
            if u'\u4e00' <= i <= u'\u9fa5':  # \u4E00 ~ \u9FFF  中文字符
                hanzi = hanzi + 1
            else:
                num += 1

        return str(30 * hanzi + 20 * num)

    def count_excel_save(self, excel_columns):

        len_list = []
        prjTuple = tuple(excel_columns)
        for idx in range(0, len(prjTuple)):
            length2 = 0
            str_list = []
            for cell in prjTuple[idx]:
                if "1" not in str(cell):
                    str_list.append(str(cell.value))
            for elem in str_list:
                elem_split = list(elem)
                length = 0
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1.5
                    else:
                        length += 2.5
                length2 = max(length, length2)
            len_list.append(length2)
        return len_list
