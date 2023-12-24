import os,datetime,json,openpyxl,time,random,string
from datetime import datetime, date
from django.forms.models import model_to_dict
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.utils import get_column_letter
from pdss.settings import BASE_DIR
from auther.models import *
from ..serializers import *
from ..models import *
from django.core.exceptions import ObjectDoesNotExist
from controller.controller2 import Controller, upload_file
from general.models import *
from utils.check_token import *
from django.http import HttpResponse





class Resetts:  #人才补贴
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))

        self.methods[self.meth]()
        return Response(self.return_data)
    def get_list(self):
        obj = Controller(TalentSubsidies, "get_list", self.request)
        # print(obj)
        self.return_data = obj.data_start()
    def get_upload(self):
        file = self.request.FILES.get("file", None)
        createData = self.request.POST.get("createData", None)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'talentSubsidiesFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        # 
        if file and file is not None:
            
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name,file,file_suffix)#保存文件
                    exc = openpyxl.load_workbook(file_url,data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        # try:
                        #     k = {
                        #         'name': sheet.cell(i + 1, 2).value,
                        #         'status': True,
                        #     }
                        #     talent_subsidies_base_id=center_base.objects.filter(**k).values_list('id')[0][0]
                        # except:
                        #     talent_subsidies_base_id=None

                        # try:
                        #     # print("1",center_base.objects.filter(status=True,base_parent_id=None).values_list('name',flat=True)[0])
                        #
                        #     if sheet.cell(i + 1, 2).value!='润阳新能源':      #base_parent_id==None
                        #         father= {
                        #             'name': sheet.cell(i + 1, 2).value,  #
                        #             'status': True,
                        #         }
                        #         son={
                        #             'name': sheet.cell(i + 1,3).value,  #
                        #             'status': True,
                        #         }
                        #         k={
                        #             'name': sheet.cell(i + 1, 3).value,  #
                        #             'status': True,
                        #             'base_parent_id':center_base.objects.filter(**son).values_list('base_parent_id',flat=True)[0]
                        #         }
                        #         talent_subsidies_base_id= center_base.objects.filter(**k).values_list('id', flat=True)[0]
                        #     else:
                        #         father= {
                        #             'name': sheet.cell(i + 1, 2).value,  #
                        #             'status': True,
                        #         }
                        #         talent_subsidies_base_id= center_base.objects.filter(**father).values_list('id', flat=True)[0]
                        # except:
                        #     talent_subsidies_base_id = None

                        try:
                            base_father = center_base.objects.get(status=1,name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg":"某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                talent_subsidies_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,base_parent_id=base_father.id,status=True).id   #二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            talent_subsidies_base_id = base_father.id
                        if talent_subsidies_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,3).value if sheet.cell(i + 1,3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:
                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    talent_subsidies_date=str(sheet.cell(i + 1, 4).value.date())
                                    # print(talent_subsidies_date)
                                except:#可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                talent_subsidies_date=None

                            talent_subsidies_conditions = self.decision_num(sheet.cell(i + 1,5).value, int)
                            talent_subsidies_applied=self.decision_num(sheet.cell(i + 1, 6).value, int)
                            talent_subsidies_claimed = self.decision_num(sheet.cell(i + 1, 7).value, int)
                            data_kwargs = {
                                'talent_subsidies_base_id':talent_subsidies_base_id,
                                'talent_subsidies_date':talent_subsidies_date,
                                'talent_subsidies_conditions': talent_subsidies_conditions,#满足条件HC
                                'talent_subsidies_applied': talent_subsidies_applied,#已申请HC
                                'talent_subsidies_claimed':talent_subsidies_claimed,  #已领取HC
                                'talent_subsidies_remark':sheet.cell(i + 1, 8).value,#备注
                            }
                            data_kwargs['creator_id'] = self.request.check_token
                            # try:
                            #     if type(data_kwargs['talent_subsidies_claimed'])==str:
                            #         data_kwargs['talent_subsidies_claimed'] = eval(data_kwargs['talent_subsidies_claimed'])
                            # except:
                            #     data_kwargs['talent_subsidies_claimed'] =None
                            #
                            # try:
                            #     if type(data_kwargs['talent_subsidies_applied']) == str:
                            #         data_kwargs['talent_subsidies_applied'] = eval(data_kwargs['talent_subsidies_applied'])
                            # except:
                            #     data_kwargs['talent_subsidies_applied'] =None
                            #
                            # try:
                            #     if type(data_kwargs['talent_subsidies_conditions']) == str:
                            #         data_kwargs['talent_subsidies_conditions'] = eval(data_kwargs['talent_subsidies_conditions'])
                            # except:
                            #     data_kwargs['talent_subsidies_conditions'] =None




                            if  data_kwargs['talent_subsidies_base_id']==None and data_kwargs['talent_subsidies_date'] ==None and data_kwargs['talent_subsidies_conditions']==None and data_kwargs['talent_subsidies_applied']==None and data_kwargs['talent_subsidies_claimed']==None and data_kwargs['talent_subsidies_remark']==None :
                                continue
                            else:
                                if data_kwargs['talent_subsidies_conditions']==None or data_kwargs['talent_subsidies_applied']==None or  data_kwargs['talent_subsidies_claimed']==None   :
                                    if data_kwargs['talent_subsidies_conditions']==None:
                                        if data_kwargs['talent_subsidies_applied']!=None and data_kwargs['talent_subsidies_claimed']!=None:
                                            if data_kwargs['talent_subsidies_applied']>=data_kwargs['talent_subsidies_claimed']:
                                                TalentSubsidies.objects.create(**data_kwargs)
                                        elif data_kwargs['talent_subsidies_applied']==None or data_kwargs['talent_subsidies_claimed']==None:
                                            TalentSubsidies.objects.create(**data_kwargs)

                                    elif data_kwargs['talent_subsidies_applied']==None:
                                        if data_kwargs['talent_subsidies_conditions'] != None and data_kwargs['talent_subsidies_claimed'] != None:
                                            if data_kwargs['talent_subsidies_conditions']>=data_kwargs['talent_subsidies_claimed']:
                                                TalentSubsidies.objects.create(**data_kwargs)
                                        elif data_kwargs['talent_subsidies_conditions'] == None or data_kwargs['talent_subsidies_claimed'] == None:
                                            TalentSubsidies.objects.create(**data_kwargs)

                                    elif data_kwargs['talent_subsidies_claimed']==None:
                                        if data_kwargs['talent_subsidies_conditions'] != None and data_kwargs['talent_subsidies_applied'] != None:
                                            if data_kwargs['talent_subsidies_conditions']>=data_kwargs['talent_subsidies_applied']:
                                                TalentSubsidies.objects.create(**data_kwargs)
                                        elif data_kwargs['talent_subsidies_conditions'] == None or data_kwargs['talent_subsidies_applied']== None:
                                            TalentSubsidies.objects.create(**data_kwargs)

                                    #
                                    # TalentSubsidies.objects.create(**data_kwargs)
                                    # if data_kwargs['talent_subsidies_conditions']>=data_kwargs['talent_subsidies_applied'] or data_kwargs['talent_subsidies_conditions']>=data_kwargs['talent_subsidies_claimed'] or data_kwargs['talent_subsidies_applied']>=data_kwargs['talent_subsidies_claimed']:
                                    #
                                else:
                                    if data_kwargs['talent_subsidies_conditions']>=data_kwargs['talent_subsidies_applied'] and data_kwargs['talent_subsidies_applied']>=data_kwargs['talent_subsidies_claimed']: #满足数据存在的同时 数据关系也要成立
                                        #
                                        TalentSubsidies.objects.create(**data_kwargs)
                                    else:
                                        continue



            except Exception as e:
                
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            
            createData = eval(createData)#     {"talent_subsidies_base_id": 11, "talent_subsidies_date": "2023-06-25", "talent_subsidies_conditions": '2e',"talent_subsidies_applied": '1332', "talent_subsidies_claimed": 21,"talent_subsidies_remark": "备注一下吧"}
            # 
            createData['talent_subsidies_conditions'] = self.decision_num(createData['talent_subsidies_conditions'], int)
            createData['talent_subsidies_applied'] = self.decision_num(createData['talent_subsidies_applied'],int)
            createData['talent_subsidies_claimed'] = self.decision_num(createData['talent_subsidies_claimed'],int)
            createData['creator_id'] = self.request.check_token
            # if createData['talent_subsidies_conditions']>=createData['talent_subsidies_applied']>=createData['talent_subsidies_claimed']:
            #     TalentSubsidies.objects.create(**createData)
            #     self.return_data = {
            #         "code": status.HTTP_200_OK,
            #         "msg": "用户数据新增成功"
            #     }
            # else:
            #     # 
            #     self.return_data = {
            #         "code": status.HTTP_401_UNAUTHORIZED,
            #         "msg": "新增失败,数据关系不符合"
            #     }

            if createData['talent_subsidies_conditions'] != None and createData['talent_subsidies_applied']  != None and createData['talent_subsidies_claimed'] != None:  #三条数据都存在
                if createData['talent_subsidies_conditions']>=createData['talent_subsidies_applied']>=createData['talent_subsidies_claimed']:
                    TalentSubsidies.objects.create(**createData)
                    self.return_data = {
                        "code": status.HTTP_200_OK,
                        "msg": "用户数据新增成功"
                    }
                else:
                    self.return_data = {
                        "code": status.HTTP_401_UNAUTHORIZED,
                        "msg": "新增失败,数据关系不符合"
                    }
            elif (createData['talent_subsidies_conditions'] != None and createData['talent_subsidies_applied'] != None)  or (createData['talent_subsidies_conditions'] != None and createData['talent_subsidies_claimed'] != None)  or (createData['talent_subsidies_applied'] != None and createData['talent_subsidies_claimed'] != None):  # 存在两条数据
                if createData['talent_subsidies_conditions'] ==None:
                    if createData['talent_subsidies_applied'] >= createData['talent_subsidies_claimed']:
                        TalentSubsidies.objects.create(**createData)
                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "用户数据新增成功"
                        }
                    else:
                        self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": "新增失败,数据关系不符合"
                            }
                elif createData['talent_subsidies_applied'] ==None:
                    if createData['talent_subsidies_conditions'] >= createData['talent_subsidies_claimed']:
                        TalentSubsidies.objects.create(**createData)
                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "用户数据新增成功"
                        }
                    else:
                        self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": "新增失败,数据关系不符合"
                            }
                elif createData['talent_subsidies_claimed'] ==None:
                    if createData['talent_subsidies_conditions'] >=createData['talent_subsidies_applied']:
                        TalentSubsidies.objects.create(**createData)
                        self.return_data = {
                            "code": status.HTTP_200_OK,
                            "msg": "用户数据新增成功"
                        }
                    else:
                        self.return_data = {
                                "code": status.HTTP_401_UNAUTHORIZED,
                                "msg": "新增失败,数据关系不符合"
                            }
            elif createData['talent_subsidies_conditions'] ==None or  createData['talent_subsidies_applied'] ==None or   createData['talent_subsidies_claimed'] ==None:#只有一个数据的
                TalentSubsidies.objects.create(**createData)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(TalentSubsidies, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }
    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(TalentSubsidies, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }
    def download_file(self):
        
        
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'talentSubsidiesFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('人才补贴收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'talent_subsidies_status': True}
            project_bonus_base = self.request.GET.get('baseNameId', None)  # 公司名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if project_bonus_base == '' and beginDate == "" and endDate == "":  # 全查

                kwargs['talent_subsidies_status'] = True
                kwargs['talent_subsidies_base__in'] = self.request.user_base
            if project_bonus_base != '':
                kwargs['talent_subsidies_base'] = project_bonus_base
            if beginDate != "" and endDate != "":
                kwargs['talent_subsidies_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
                kwargs['talent_subsidies_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            
            all = list(TalentSubsidies.objects.filter(**kwargs).all().values_list('talent_subsidies_base',
                    'talent_subsidies_base__name', 'talent_subsidies_date','talent_subsidies_conditions', 'talent_subsidies_applied', 'talent_subsidies_claimed','talent_subsidies_remark'))
            
            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                i.insert(0, index)
                del i[2]
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            row_data = []
            index = 1
            for id in id_list:
                data = TalentSubsidies.objects.filter(talent_subsidies_status=True, id=id).values_list('talent_subsidies_base',
                    'talent_subsidies_base__name', 'talent_subsidies_date','talent_subsidies_conditions', 'talent_subsidies_applied', 'talent_subsidies_claimed','talent_subsidies_remark')
                for i in data:
                    i=list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',flat=True).exists():
                        # print(i[1],'1级基地')  润阳新能源
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                            father_name = center_base.objects.filter(
                                id=
                                center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                    0],
                                status=True).values_list('name', flat=True)[0]
                            i.insert(0, father_name)
            #         if i[1] != '润阳新能源':
            #             # print(center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[0])
            #             father_name = center_base.objects.filter(
            #                 id=
            #                 center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
            #                     0],
            #                 status=True).values_list('name', flat=True)[0]
            #             i.insert(0, father_name)
            #         else:
            #             i.insert(0, '润阳新能源')
            #             i[2] = ''
                    i.insert(0, index)
                    del i[2]
                    row_data.append(list(i))
                    if len(i) == 0:
                        index = index
                    index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        #
    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'talentSubsidiesFile','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'talentSubsidiesFile','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'talentSubsidiesFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/talentSubsidiesFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/talentSubsidiesFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 107):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i>=104:
                sheet.column_dimensions[chr(i).upper()].width = 18

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:H1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '人才补贴收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号','中心/基地','公司','日期','满足条件HC','已申请HC','已领取HC','备注'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'talentSubsidiesFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path
    def mkdir(self,path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        else:
            pass
            
    def get_first_last(self,date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month=datetime.date(year, month, day=1)
        last_month=datetime.date(year, month, day=monthCountDay)
        first_year=datetime.date(year,1,1)
        last_year=datetime.date(year,12,31)
        return first_month,last_month,first_year,last_year

    def decision_num(self, num, typp):
        if num != None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) == int:
                    if num >= 0:
                        return num
                    else:
                        return
                elif type(num) == typp:
                    if num >= 0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return

class Resetei:  #离职访谈
    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }
    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))

        self.methods[self.meth]()
        return Response(self.return_data)
    def get_list(self):
        obj = Controller(ExitInterviews, "get_list", self.request)
        self.return_data = obj.data_start()
    def get_upload(self):
        file = self.request.FILES.get("file", None)
        createData = self.request.POST.get("createData", None)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'exitInterviewsFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        # 
        if file and file is not None:
            
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name,file,file_suffix)#保存文件
                    exc = openpyxl.load_workbook(file_url,data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1,name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg":"某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                exit_interviews_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,base_parent_id=base_father.id,status=True).id   #二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            exit_interviews_base_id = base_father.id
                        if exit_interviews_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,3).value if sheet.cell(i + 1,3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:


                            # try:
                            #     if sheet.cell(i + 1, 2).value != '润阳新能源':
                            #         father = {
                            #             'name': sheet.cell(i + 1, 2).value,  #
                            #             'status': True,
                            #         }
                            #         son = {
                            #             'name': sheet.cell(i + 1, 3).value,  #
                            #             'status': True,
                            #         }
                            #         k = {
                            #             'name': sheet.cell(i + 1, 3).value,  #
                            #             'status': True,
                            #             'base_parent_id':
                            #                 center_base.objects.filter(**son).values_list('base_parent_id', flat=True)[0]
                            #         }
                            #         exit_interviews_base_id = center_base.objects.filter(**k).values_list('id', flat=True)[
                            #             0]
                            #     else:
                            #         father = {
                            #             'name': sheet.cell(i + 1, 2).value,  #
                            #             'status': True,
                            #         }
                            #         exit_interviews_base_id = center_base.objects.filter(**father).values_list('id', flat=True)[0]
                            #
                            # except:
                            #     exit_interviews_base_id = None

                            #
                            #
                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    exit_interviews_date=str(sheet.cell(i + 1, 4).value.date())
                                except:#可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                exit_interviews_date=None


                            # if sheet.cell(i + 1,4).value != None:   #访谈人次
                            #     try:
                            #         if type(sheet.cell(i + 1,4).value)==int and sheet.cell(i + 1,4).value>=0:
                            #             exit_interviews_numberInterviews=sheet.cell(i + 1, 4).value
                            #         else:
                            #             exit_interviews_numberInterviews = None
                            #     except:#可能不是日期
                            #         exit_interviews_numberInterviews=None
                            # else:
                            #     exit_interviews_numberInterviews=None
                            exit_interviews_numberInterviews = self.decision_num(sheet.cell(i + 1,5).value, int)  # 挽留成功人次
                            exit_interviews_retentionSuccess=self.decision_num(sheet.cell(i + 1, 6).value,int)#挽留成功人次

                            # if sheet.cell(i + 1, 5).value != None:  #挽留成功人次
                            #     try:
                            #         if type(sheet.cell(i + 1, 5).value) == int and sheet.cell(i + 1, 5).value>=0:
                            #             exit_interviews_retentionSuccess = sheet.cell(i + 1, 5).value
                            #         else:
                            #             exit_interviews_retentionSuccess = None
                            #     except:  # 可能不是日期
                            #         exit_interviews_retentionSuccess = None
                            # else:
                            #     exit_interviews_retentionSuccess = None


                            if exit_interviews_numberInterviews!=None and exit_interviews_retentionSuccess !=None:
                                if exit_interviews_numberInterviews==0:
                                    exit_interviews_retentionSuccessRate=0
                                else:
                                    exit_interviews_retentionSuccessRate=exit_interviews_retentionSuccess/exit_interviews_numberInterviews
                            else:
                                exit_interviews_retentionSuccessRate=None
                            data_kwargs = {
                                'exit_interviews_base_id':exit_interviews_base_id,
                                'exit_interviews_date':exit_interviews_date ,
                                'exit_interviews_numberInterviews': exit_interviews_numberInterviews,#访谈人次
                                'exit_interviews_retentionSuccess':exit_interviews_retentionSuccess,#挽留成功人次
                                'exit_interviews_retentionSuccessRate': exit_interviews_retentionSuccessRate,  #挽留成功率
                                'exit_interviews_typicalCase':sheet.cell(i+1,7).value ,#典型案例
                                'exit_interviews_remark': sheet.cell(i + 1, 8).value,  # 备注
                            }
                            data_kwargs['creator_id'] = self.request.check_token

                            #
                            if  data_kwargs['exit_interviews_base_id']==None and data_kwargs['exit_interviews_date'] ==None and data_kwargs['exit_interviews_numberInterviews']==None and data_kwargs['exit_interviews_retentionSuccess']==None and data_kwargs['exit_interviews_typicalCase']==None and data_kwargs['exit_interviews_remark']==None :
                                continue
                            else:
                                if exit_interviews_retentionSuccessRate!=None:
                                    if 0<=data_kwargs['exit_interviews_retentionSuccessRate']<=1:
                                        ExitInterviews.objects.create(**data_kwargs)
                                else:
                                    ExitInterviews.objects.create(**data_kwargs)

            except Exception as e:
                
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            
            createData = eval(createData)# {"exit_interviews_date": "2023-03-02", "exit_interviews_base_id": 1,'exit_interviews_numberInterviews': 1, "exit_interviews_retentionSuccess":1212,'exit_interviews_typicalCase':222,"exit_interviews_remark": '备注一下吧'}
            # print(createData)
            # try:
            #     if type(createData['exit_interviews_numberInterviews']) == str:
            #         createData['exit_interviews_numberInterviews'] = eval(createData['exit_interviews_numberInterviews'])
            # except:
            #     createData['exit_interviews_numberInterviews'] = None
            # try:
            #     if type(createData['exit_interviews_retentionSuccess']) == str:
            #         createData['exit_interviews_retentionSuccess'] = eval(createData['exit_interviews_retentionSuccess'])
            # except:
            #     createData['exit_interviews_retentionSuccess'] = None

            createData['exit_interviews_numberInterviews'] = self.decision_num(createData['exit_interviews_numberInterviews'], int)
            createData['exit_interviews_retentionSuccess']=self.decision_num(createData['exit_interviews_retentionSuccess'],int)
            createData['creator_id'] = self.request.check_token
            # print(createData)
            if createData['exit_interviews_numberInterviews']==None or createData['exit_interviews_retentionSuccess']==None:
                ExitInterviews.objects.create(**createData)
            elif createData['exit_interviews_numberInterviews']>=createData['exit_interviews_retentionSuccess']:
                if createData['exit_interviews_numberInterviews']==0:
                    createData['exit_interviews_retentionSuccessRate'] =0
                else:
                    createData['exit_interviews_retentionSuccessRate']=createData['exit_interviews_retentionSuccess']/createData['exit_interviews_numberInterviews']
                # print(createData)
                ExitInterviews.objects.create(**createData)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            else:
                
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "新增失败,数据关系不符合"
                }
    def delete_data(self):  #删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(ExitInterviews, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }
    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(ExitInterviews, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }
    def download_file(self):
        
        
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR,'static', 'exitInterviewsFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('离职访谈收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'exit_interviews_status': True}
            project_bonus_base = self.request.GET.get('baseNameId', None)  # 公司名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if project_bonus_base == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['exit_interviews_status'] = True
                kwargs['exit_interviews_base__in'] = self.request.user_base
            if project_bonus_base != '':
                kwargs['exit_interviews_base'] = project_bonus_base
            if beginDate != "" and endDate != "":
                kwargs['exit_interviews_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate == None else beginDate
                kwargs['exit_interviews_date__lte'] = datetime(2521, 10, 29, 7, 17, 1, 177) if endDate == None else endDate
            
            all = list(ExitInterviews.objects.filter(**kwargs).all().values_list('exit_interviews_base',
                    'exit_interviews_base__name', 'exit_interviews_date','exit_interviews_numberInterviews', 'exit_interviews_retentionSuccess', 'exit_interviews_retentionSuccessRate','exit_interviews_typicalCase','exit_interviews_remark'))

            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')  润阳新能源
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                i.insert(0, index)
                del i[2]
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data = []
            index = 1
            for id in id_list:
                data = ExitInterviews.objects.filter(exit_interviews_status=True, id=id).values_list('exit_interviews_base',
                    'exit_interviews_base__name', 'exit_interviews_date','exit_interviews_numberInterviews', 'exit_interviews_retentionSuccess', 'exit_interviews_retentionSuccessRate','exit_interviews_typicalCase','exit_interviews_remark')
                for i in data:
                    i=list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',flat=True).exists():
                        # print(i[1],'1级基地')  润阳新能源
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    i.insert(0, index)
                    del i[2]
                    row_data.append(list(i))
                    if len(i) == 0:
                        index = index
                    index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # 
    def create_file_name(self, objName,file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t=time.strftime('%Y-%m-%d')
        fileName = str(objName) + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]
    def createPath(self, pic,func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func+'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'exitInterviewsFile','upload_file', t,pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png','icon']:
            pic_path = os.path.join('static', 'exitInterviewsFile','upload_file', t,pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'exitInterviewsFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字
    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/exitInterviewsFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/exitInterviewsFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
    def createExcelPath(self, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 105):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i>=104:
                sheet.column_dimensions[chr(i).upper()].width = 18

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:I1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '离职访谈收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号','中心/基地','公司','日期','访谈人次','挽留成功人次','挽留成功率','典型案例','备注'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'exitInterviewsFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path
    def mkdir(self,path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            
        else:
            pass
            
    def get_first_last(self,date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month=datetime.date(year, month, day=1)
        last_month=datetime.date(year, month, day=monthCountDay)
        first_year=datetime.date(year,1,1)
        last_year=datetime.date(year,12,31)
        return first_month,last_month,first_year,last_year

    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return

                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return

class Resetcq:  # 座谈会

    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }

    def meth_center(self):
        # if self.verify(self.request):
        #     return Response(self.verify(self.request))

        self.methods[self.meth]()
        return Response(self.return_data)

    def get_list(self):
        columnList = [
        {
            "value": "序号",
            "label": "index",
            "width": 60
        },
        {
            "value": "中心/事业部",
            "label": "base_father",
            "width": 180
        },
        {
            "value": "公司",
            "label": "colloquium_base",
            "width": 160
        },
        {
            "value": "日期",
            "label": "colloquium_date",
            "width": 200
        },
        {
            "value": "场次",
            "label": "colloquium_numberSessions",
            "width": 60
        },
        {
            "value": "参与人次",
            "label": "colloquium_numberParticipants",
            "width": 120
        },
        {
            "value": "占比",
            "label": "colloquium_percentage",
            "width": 80
        },
        {
            "value": "产出行动项",
            "label": "colloquium_outputItems",
            "width": 150
        },
        {
            "value": "关闭项",
            "label": "colloquium_closeItem",
            "width": 90
        },
        {
            "value": "完成率",
            "label": "colloquium_completionRate",
            "width": 90
        },
        {
            "value": "备注",
            "label": "colloquium_remark",
            "width": 500
        },
        {
            "value": "典型事项",
            "label": "colloquium_typical",
            "width": 800
        },
        {
            "value": "座谈会图片",
            "label": "colloquium_photos",
            "width": 150
        },
        {
            "value": "年度累计场次",
            "label": "year_colloquium_numberSessions",
            "width": 70
        },
        {
            "value": "年度累计人次",
            "label": "year_colloquium_numberParticipants",
            "width": 70
        }
        ]
        # print(self.request.GET)
        currentPage = eval(self.request.GET.get("currentPage", None)) if self.request.GET.get("currentPage",
                                                                                              None) != "" else 1
        pageSize = eval(self.request.GET.get("pageSize", None)) if self.request.GET.get("pageSize", None) != "" else 25
        kwargs = {
            'coll_interviews_status': True
        }
        searchName = self.request.GET.get('searchName', None)
        searchBase = self.request.GET.get('searchBase', None)
        beginDate = self.request.GET.get('beginDate', None)
        endDate = self.request.GET.get('endDate', None)
        if beginDate is not None and endDate is not None:
            kwargs['colloquium_date__gte'] = datetime(1901, 10, 29, 7, 17, 1, 177) if beginDate is None or len(
                endDate) == 0 else beginDate
            kwargs['colloquium_date__lte'] = datetime(3221, 10, 29, 7, 17, 1, 177) if endDate is None or len(
                endDate) == 0 else endDate  # 内部评优时间
        if searchName == '' and searchBase == '' and beginDate == "" and endDate == "":  # 全查
            kwargs['coll_interviews_status'] = True
            kwargs['colloquium_base__in'] = self.request.user_base
        if searchBase != '':
            kwargs['colloquium_base'] = searchBase
        totalNumber = Colloquium.objects.filter(
             **kwargs).count()
        tableList = Colloquium.objects.filter(
            **kwargs).values(
            'id',
            "colloquium_base__name",
            "colloquium_base__base_parent_id",
            "colloquium_base",
            "colloquium_date",
            "colloquium_numberSessions",
            "colloquium_numberParticipants",
            "colloquium_percentage",
            "colloquium_outputItems",
            "colloquium_closeItem",
            "colloquium_completionRate",
            "colloquium_remark",
            "colloquium_typical",

        ).order_by('-create_time')[(currentPage - 1) * pageSize:currentPage * pageSize]
        base_list = center_base.objects.filter(status=True).values('id', 'name')
        all_inevid = [item['id'] for item in tableList]
        all_inev_file = Colloquium.objects.filter(id__in=all_inevid, coll_interviews_status=True,
                                                              colloquium_photos__status=True).values('id',
                                                                                                 'colloquium_photos__id',
                                                                                                 'colloquium_photos__file_name',
                                                                                                 'colloquium_photos__file_url')
        for index, item in enumerate(tableList):
            item['index'] = (currentPage - 1) * pageSize + index + 1
            item['colloquium_base_id'] = item['colloquium_base']
            item['colloquium_base'] = item['colloquium_base__name']

            item['colloquium_photos_files'] = []
            for base in list(base_list):
                if base['id'] == item['colloquium_base__base_parent_id']:
                    item['base_father'] = base['name']
            if item['colloquium_base__base_parent_id'] is None:
                item['base_father'] = item['colloquium_base__name']
            for inev_file_obj in all_inev_file:
                if inev_file_obj['id'] == item['id']:
                    # item['awards_photos']=len(inev_file_obj['id'])
                    if inev_file_obj['colloquium_photos__id'] is not None and inev_file_obj[
                        'colloquium_photos__file_name'] is not None and inev_file_obj[
                        'colloquium_photos__file_url'] is not None:
                        item['colloquium_photos_files'].append(
                            {
                                "id": inev_file_obj['colloquium_photos__id'],
                                "name": inev_file_obj['colloquium_photos__file_name'],
                                "url": inev_file_obj['colloquium_photos__file_url'],
                            })
            item['colloquium_photos'] = len(item['colloquium_photos_files'])

            #计算年度累计场次 年度累计人次
            if item['colloquium_base_id'] == None or item['colloquium_date'] == None:  # 无法统计
                # #print(i['employee_inspect_date'],type(i['employee_inspect_date']))
                pass
            else:  # 两个都不为None
                year_colloquium_numberSessions = Colloquium.objects.filter(coll_interviews_status=True,
                                                                           colloquium_base_id=item['colloquium_base_id'],
                                                                           colloquium_date__gte=self.get_first_last(item['colloquium_date'])[2],
                                                                           colloquium_date__lte=self.get_first_last(item['colloquium_date'])[3]).values_list(
                    'colloquium_numberSessions', flat=True)  # 年度累计场次
                year_colloquium_numberParticipants = Colloquium.objects.filter(coll_interviews_status=True,
                                                                               colloquium_base_id=item['colloquium_base_id'],
                                                                               colloquium_date__gte=self.get_first_last(item['colloquium_date'])[2],
                                                                               colloquium_date__lte=self.get_first_last(item['colloquium_date'])[3]).values_list(
                    'colloquium_numberParticipants', flat=True)  # 年度累计人次
            #
                year_colloquium_numberSessions = list(filter(None, year_colloquium_numberSessions))  # 列表去除None
                year_colloquium_numberParticipants = list(filter(None, year_colloquium_numberParticipants))  # 列表去除None
                item['year_colloquium_numberSessions'] = sum(year_colloquium_numberSessions)
                item['year_colloquium_numberParticipants'] = sum(year_colloquium_numberParticipants)


            try:
                item['colloquium_percentage'] = "{:.2f}%".format(round(item['colloquium_percentage'], 4) * 100)
            except:
                pass
            try:
                item['colloquium_completionRate'] = "{:.2f}%".format(round(item['colloquium_completionRate'], 4) * 100)
            except:
                pass


        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "信息返回成功",
            "data": {
                'columnList': columnList,
                'tableList': tableList,
                'totalNumber': totalNumber,
            }
        }

        # obj = Controller(Colloquium, "get_list", self.request)
        # self.return_data = obj.data_start()

    def get_upload(self):

        file = self.request.FILES.get("file", None)
        createPhoto = self.request.FILES.getlist("createPhoto", None)  # 照片
        createData = self.request.POST.get("createData", None)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR, 'static', 'colloquiumFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        #
        if file and file is not None:
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    exc = openpyxl.load_workbook(file_url, data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1, name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg": "某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                colloquium_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,
                                                                             base_parent_id=base_father.id,
                                                                             status=True).id  # 二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            colloquium_base_id = base_father.id
                        if colloquium_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,
                                                                                                          3).value if sheet.cell(
                                    i + 1, 3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:
                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    colloquium_date = str(sheet.cell(i + 1, 4).value.date())
                                    colloquium_numberSessions = self.decision_num(sheet.cell(i + 1, 5).value, int)
                                    colloquium_numberParticipants = self.decision_num(sheet.cell(i + 1, 6).value, int)
                                    colloquium_percentage = self.decision_num(sheet.cell(i + 1, 7).value, float)  # 占比
                                    colloquium_outputItems = self.decision_num(sheet.cell(i + 1, 8).value, int)
                                    colloquium_closeItem = self.decision_num(sheet.cell(i + 1, 9).value, int)

                                    colloquium_completionRate = None
                                    if colloquium_percentage != None:
                                        if colloquium_percentage > 1 or colloquium_percentage < 0:
                                            colloquium_percentage = None

                                    if colloquium_outputItems != None and colloquium_closeItem != None:
                                        if colloquium_outputItems >= colloquium_closeItem:
                                            if colloquium_outputItems == 0:
                                                colloquium_completionRate = 0
                                            else:
                                                colloquium_completionRate = colloquium_closeItem / colloquium_outputItems
                                        else:
                                            colloquium_closeItem = None
                                            colloquium_outputItems = None

                                    data_kwargs = {
                                        'colloquium_base_id': colloquium_base_id,
                                        'colloquium_date': colloquium_date,
                                        'colloquium_numberSessions': colloquium_numberSessions,  # 场次
                                        'colloquium_numberParticipants': colloquium_numberParticipants,  # 参与人次
                                        'colloquium_percentage': colloquium_percentage,  # 占比
                                        'colloquium_outputItems': colloquium_outputItems,  # 产出行动项
                                        'colloquium_closeItem': colloquium_closeItem,  # 关闭项
                                        'colloquium_completionRate': colloquium_completionRate,  # 完成率
                                        'colloquium_typical': sheet.cell(i + 1, 10).value,  # 典型事项
                                        'colloquium_remark': sheet.cell(i + 1, 11).value,  # 备注
                                        'creator_id': self.request.check_token
                                    }

                                    if data_kwargs['colloquium_base_id'] == None and data_kwargs[
                                        'colloquium_date'] == None and \
                                            data_kwargs['colloquium_numberSessions'] == None and data_kwargs[
                                        'colloquium_numberParticipants'] == None and data_kwargs[
                                        'colloquium_outputItems'] == None and data_kwargs[
                                        'colloquium_closeItem'] == None and \
                                            data_kwargs['colloquium_typical'] == None and data_kwargs[
                                        'colloquium_remark'] == None:
                                        continue
                                    else:
                                        Colloquium.objects.create(**data_kwargs)
                                        self.return_data = {
                                            "code": status.HTTP_200_OK,
                                            "msg": "上传成功"
                                        }
                                except:  # 可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                self.return_data = {
                                    "code": status.HTTP_401_UNAUTHORIZED,
                                    "msg": "日期为空，无法上传"
                                }


            except Exception as e:

                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":

            createData = eval(
                createData)  # {"colloquium_base_id":11,"colloquium_date":"2023-06-25","colloquium_numberSessions":'2e',"colloquium_numberParticipants":'1332',"colloquium_outputItems":21,"colloquium_closeItem":"","colloquium_typical":"23333","colloquium_remark":"备注一下吧"}
            # print(createData)
            createData['creator_id'] = self.request.check_token
            createData['colloquium_numberSessions'] = self.decision_num(createData['colloquium_numberSessions'], int)
            createData['colloquium_numberParticipants'] = self.decision_num(createData['colloquium_numberParticipants'],
                                                                            int)
            createData['colloquium_percentage'] = self.decision_num(createData['colloquium_percentage'], float)  # 占比
            createData['colloquium_outputItems'] = self.decision_num(createData['colloquium_outputItems'], int)
            createData['colloquium_closeItem'] = self.decision_num(createData['colloquium_closeItem'], int)

            #

            if createData['colloquium_percentage'] != None:
                if createData['colloquium_percentage'] > 1 or createData['colloquium_percentage'] < 0:
                    createData['colloquium_percentage'] = None

            if createData['colloquium_outputItems'] == None or createData['colloquium_closeItem'] == None:
                createData['colloquium_completionRate'] = None
                obj = Colloquium.objects.create(**createData)
                for file in createPhoto:
                    file_url, file_name, file_suffix = self.createPath(file,
                                                                       self.create_file_name(obj.colloquium_remark,
                                                                                             file))
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_url,
                        "file_name": file_name,
                    }
                    file_obj = UploadFiles.objects.create(**file_kwargs)
                    file_obj.colloquium_photos.add(obj.id)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            elif createData['colloquium_outputItems'] != None and createData['colloquium_closeItem'] != None and \
                    createData['colloquium_outputItems'] >= createData['colloquium_closeItem']:
                if createData['colloquium_outputItems'] == 0:
                    createData['colloquium_completionRate'] = 0
                else:
                    createData['colloquium_completionRate'] = createData['colloquium_closeItem'] / createData[
                        'colloquium_outputItems']
                obj = Colloquium.objects.create(**createData)
                for file in createPhoto:
                    file_url, file_name, file_suffix = self.createPath(file,
                                                                       self.create_file_name(obj.colloquium_remark,
                                                                                             file))
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    file_kwargs = {
                        "file_url": file_url,
                        "file_name": file_name,
                    }
                    file_obj = UploadFiles.objects.create(**file_kwargs)
                    file_obj.colloquium_photos.add(obj.id)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            else:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据关系不符合,用户数据新增失败"
                }

    def delete_data(self):  # 删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(Colloquium, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(Colloquium, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def download_file(self):
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR, 'static', 'colloquiumFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('座谈会收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'coll_interviews_status': True}
            colloquium_base = self.request.GET.get('searchBase', None)  # 公司名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if colloquium_base == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['coll_interviews_status'] = True
                kwargs['colloquium_base__in'] = self.request.user_base
            if colloquium_base != '':
                kwargs['colloquium_base'] = colloquium_base
            if beginDate != "" and endDate != "":
                kwargs['colloquium_date__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                               177) if beginDate == None else beginDate
                kwargs['colloquium_date__lte'] = datetime(2521, 10, 29, 7, 17, 1,
                                                               177) if endDate == None else endDate
            
            all = list(Colloquium.objects.filter(**kwargs).all().values_list('colloquium_base',
                                                                            'colloquium_base__name', 'colloquium_date', 'colloquium_numberSessions',
                                                                            'colloquium_numberParticipants', 'colloquium_percentage',
                                                                            'colloquium_outputItems', 'colloquium_closeItem','colloquium_completionRate',
                                                                             'colloquium_typical','colloquium_remark'))

            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')  润阳新能源
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                if i[1] == None or i[3] == None:
                    pass
                else:
                    year_colloquium_numberSessions= Colloquium.objects.filter(coll_interviews_status=True,
                                                                              colloquium_base_id=i[1],
                                                                              colloquium_date__gte=self.get_first_last(i[3])[2],
                                                                              colloquium_date__lte=self.get_first_last(i[3])[3]).values_list('colloquium_numberSessions', flat=True)  #年度累计场次

                    year_colloquium_numberSessions= list(filter(None,year_colloquium_numberSessions))
                    year_colloquium_numberParticipants= Colloquium.objects.filter(coll_interviews_status=True,
                                                                              colloquium_base_id=i[1],
                                                                              colloquium_date__gte=self.get_first_last(i[3])[2],
                                                                              colloquium_date__lte=self.get_first_last(i[3])[3]).values_list('colloquium_numberParticipants', flat=True)  #年度累计人次

                    year_colloquium_numberParticipants= list(filter(None,year_colloquium_numberParticipants))

                    i.extend([sum(year_colloquium_numberSessions),sum(year_colloquium_numberParticipants)])
                i.insert(0, index)
                del i[2]
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data = []
            index = 1
            for id in id_list:
                data = Colloquium.objects.filter(coll_interviews_status=True, id=id).values_list(
                    'colloquium_base',
                    'colloquium_base__name', 'colloquium_date', 'colloquium_numberSessions',
                    'colloquium_numberParticipants', 'colloquium_percentage',
                    'colloquium_outputItems', 'colloquium_closeItem','colloquium_completionRate','colloquium_typical','colloquium_remark')
                for i in data:
                    i = list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                           flat=True).exists():
                        # print(i[1],'1级基地')  润阳新能源
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    if i[1] == None or i[3] == None:
                        pass
                    else:
                        year_colloquium_numberSessions = Colloquium.objects.filter(coll_interviews_status=True,
                                                                                   colloquium_base_id=i[1],
                                                                                   colloquium_date__gte=
                                                                                   self.get_first_last(i[3])[2],
                                                                                   colloquium_date__lte=
                                                                                   self.get_first_last(i[3])[
                                                                                       3]).values_list(
                            'colloquium_numberSessions', flat=True)  # 年度累计场次

                        year_colloquium_numberSessions = list(filter(None, year_colloquium_numberSessions))
                        year_colloquium_numberParticipants = Colloquium.objects.filter(coll_interviews_status=True,
                                                                                       colloquium_base_id=i[1],
                                                                                       colloquium_date__gte=
                                                                                       self.get_first_last(i[3])[2],
                                                                                       colloquium_date__lte=
                                                                                       self.get_first_last(i[3])[
                                                                                           3]).values_list(
                            'colloquium_numberParticipants', flat=True)  # 年度累计人次

                        year_colloquium_numberParticipants = list(filter(None, year_colloquium_numberParticipants))

                        i.extend([sum(year_colloquium_numberSessions), sum(year_colloquium_numberParticipants)])
                    i.insert(0, index)
                    del i[2]
                    row_data.append(list(i))
                    if len(i) == 0:
                        index = index
                    index += 1

            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # 

    #
    def create_file_name(self, objName, file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t = time.strftime('%Y-%m-%d')
        fileName = str(objName)[:5] + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def createPath(self, pic, func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func + 'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'colloquiumFile', 'upload_file', t, pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png', 'icon']:
            pic_path = os.path.join('static', 'colloquiumFile', 'upload_file', t, pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'colloquiumFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/colloquiumFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/colloquiumFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 110):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i >= 104:
                sheet.column_dimensions[chr(i).upper()].width = 18

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:L1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '座谈会收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号', '中心/基地', '公司','日期', '场次', '参与人次','占比','产出行动项', '关闭项', '完成率','典型事项','备注','年度累计场次','年度累计人次'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'colloquiumFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def mkdir(self, path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            
        else:
            pass
            

    def get_first_last(self, date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month = datetime.date(year, month, day=1)
        last_month = datetime.date(year, month, day=monthCountDay)
        first_year = datetime.date(year, 1, 1)
        last_year = datetime.date(year, 12, 31)
        return first_month, last_month, first_year, last_year
    def round_down(self, num):  # 小数向下取整2位
        import math
        return math.floor(num * 100) / 100

    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return

class Resetji:  # 在职访谈
    def verify(self,request):
        return_data = {'code': '', "msg": ''}
        new_token = CheckToken()
        try:
            check_token = new_token.check_token(request.headers['Authorization'])
        except Exception as e:
            
            return_data['code'] = 400
            return_data['msg'] = '请求参数出错啦'
            return return_data
        if check_token is None:
            return_data['code'] = 403
            return_data['msg'] = '没有权限查询'
            return return_data

    def __init__(self, request, meth):
        self.request = request
        self.fileName = ""
        self.return_data = {}
        self.meth = meth
        self.methods = {
            "get_upload": self.get_upload,
            "get_list": self.get_list,
            "patch_data": self.patch_data,
            "delete_data": self.delete_data,
            "download_file": self.download_file,
        }

    def meth_center(self):
        if self.verify(self.request):
            return Response(self.verify(self.request))

        self.methods[self.meth]()
        return Response(self.return_data)

    def get_list(self):
        
        obj = Controller(JobInterviews, "get_list", self.request)
        self.return_data = obj.data_start()

    def get_upload(self):
        
        
        file = self.request.FILES.get("file", None)
        createPhoto = self.request.FILES.getlist("createPhoto", None)  # 照片
        createData = self.request.POST.get("createData", None)
        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR, 'static', 'jobInterviewsFile', 'upload_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        # 
        if file and file is not None:
            
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "上传成功"
            }
            try:
                file_url, file_name, file_suffix = self.createPath(file, self.create_file_name('up', file))
                if file_name:
                    self.saveFile(file_name, file, file_suffix)  # 保存文件
                    exc = openpyxl.load_workbook(file_url, data_only=True)
                    sheet = exc.active
                    for i in range(2, sheet.max_row):
                        try:
                            base_father = center_base.objects.get(status=1, name=sheet.cell(i + 1, 2).value)
                        except ObjectDoesNotExist:
                            self.return_data = {
                                "code": status.HTTP_400_BAD_REQUEST,
                                "msg": "某行基地名不存在，无法添加"
                            }
                            continue
                        obj = center_base.objects.filter(base_parent_id=base_father.id, status=True)
                        # print(obj.exists())
                        if obj.exists():
                            try:
                                job_interviews_base_id = center_base.objects.get(name=sheet.cell(i + 1, 3).value,
                                                                             base_parent_id=base_father.id,
                                                                             status=True).id  # 二级基地id
                            except:
                                # talent_subsidies_base_id =None
                                self.return_data = {
                                    "code": status.HTTP_400_BAD_REQUEST,
                                    "msg": "基地与公司关系不符"
                                }
                                continue
                        else:
                            job_interviews_base_id = base_father.id
                        if job_interviews_base_id not in self.request.user_base:
                            self.return_data = {
                                "code": status.HTTP_403_FORBIDDEN,
                                "msg": "抱歉，您没有上传 " + sheet.cell(i + 1, 2).value + "-" + sheet.cell(i + 1,
                                                                                                          3).value if sheet.cell(
                                    i + 1, 3).value else '' + " (基地/中心/公司)的权限"
                            }
                        else:

                            if sheet.cell(i + 1, 4).value != None:
                                try:
                                    job_interviews_date=str(sheet.cell(i + 1, 4).value.date())
                                except:#可能不是日期
                                    self.return_data = {
                                        "code": status.HTTP_401_UNAUTHORIZED,
                                        "msg": "有数据日期格式错误，无法上传，可以将单元格改为日期格式,例如2023/07/01"
                                    }
                                    continue
                            else:
                                job_interviews_date=None

                            job_interviews_number= self.decision_num(sheet.cell(i + 1, 5).value,int)
                            job_interviews_percentage = self.decision_num(sheet.cell(i + 1, 6).value, float)
                            job_interviews_outputItem = self.decision_num(sheet.cell(i + 1, 7).value,int)
                            job_interviews_closeItem = self.decision_num(sheet.cell(i + 1, 8).value,int)
                            job_interviews_completionRate=None
                            if job_interviews_percentage!=None:
                                if job_interviews_percentage>1 or job_interviews_percentage<0:
                                    job_interviews_percentage=None
                            if job_interviews_outputItem!=None and job_interviews_closeItem !=None:

                                if job_interviews_outputItem>=job_interviews_closeItem:
                                    if job_interviews_closeItem == 0:
                                        job_interviews_completionRate=0
                                    else:
                                        job_interviews_completionRate=job_interviews_closeItem/job_interviews_outputItem
                                else:
                                    job_interviews_outputItem=None
                                    job_interviews_closeItem=None
                            # if job_interviews_outputItem>=job_interviews_closeItem

                            data_kwargs = {
                                'job_interviews_base_id': job_interviews_base_id,
                                'job_interviews_date': job_interviews_date,
                                'job_interviews_number': job_interviews_number,  # 访谈人次
                                'job_interviews_percentage':job_interviews_percentage,#占比
                                'job_interviews_outputItem': job_interviews_outputItem,  # 产出行动项
                                'job_interviews_closeItem':job_interviews_closeItem,  # 关闭项
                                'job_interviews_completionRate':job_interviews_completionRate,#完成率
                                'job_interviews_typical':sheet.cell(i + 1, 9).value,#典型事项
                                'job_interviews_remark':sheet.cell(i + 1,10).value#备注
                            }


                            data_kwargs['creator_id'] = self.request.check_token

                            if data_kwargs['job_interviews_base_id'] == None and data_kwargs['job_interviews_date'] == None and data_kwargs['job_interviews_number'] == None and data_kwargs['job_interviews_outputItem'] == None and data_kwargs['job_interviews_closeItem'] == None and data_kwargs['job_interviews_typical'] == None and data_kwargs['job_interviews_remark'] == None:
                                continue
                            else:
                                # if data_kwargs['job_interviews_percentage'] != None:
                                #     if 0 <= data_kwargs['job_interviews_percentage'] <= 1:
                                #         JobInterviews.objects.create(**data_kwargs)
                                # elif data_kwargs['job_interviews_completionRate']!=None:
                                #     if 0 <= data_kwargs['job_interviews_completionRate'] <= 1:
                                #         JobInterviews.objects.create(**data_kwargs)
                                # else:
                                JobInterviews.objects.create(**data_kwargs)
            except Exception as e:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "上传失败" + str(e)
                }
        else:
            self.return_data = {
                "code": status.HTTP_200_OK,
                "msg": "空文件"
            }

        if createData is not None and createData != "":
            
            createData = eval(createData)  # {"job_interviews_base_id":11,"job_interviews_date":"2023-06-25","job_interviews_number":'2e',"job_interviews_outputItem":'1332',"job_interviews_closeItem":21,"job_interviews_typical":None,"job_interviews_remark":"备注一下吧"}
            createData['job_interviews_number']=self.decision_num(createData['job_interviews_number'],int)#访谈人次
            createData['job_interviews_percentage'] = self.decision_num(createData['job_interviews_percentage'],float)  # 占比
            createData['job_interviews_outputItem']=self.decision_num(createData['job_interviews_outputItem'],int)#产出行动项
            createData['job_interviews_closeItem'] = self.decision_num(createData['job_interviews_closeItem'],int)#关闭项
            # if len(str(createData['job_interviews_typical']))==0:
            #     createData['job_interviews_typical']=None
            if createData['job_interviews_percentage']!=None:
                if createData['job_interviews_percentage']>1 or createData['job_interviews_percentage']<0:
                    createData['job_interviews_percentage']=None

            if createData['job_interviews_outputItem']== None or createData['job_interviews_closeItem'] == None:
                createData['job_interviews_completionRate']=None
                createData['creator_id'] = self.request.check_token
                JobInterviews.objects.create(**createData)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            elif createData['job_interviews_outputItem']!=None and createData['job_interviews_closeItem']!=None and createData['job_interviews_outputItem']>=createData['job_interviews_closeItem']:
                if createData['job_interviews_outputItem']==0:
                    createData['job_interviews_completionRate'] =0
                else:
                    createData['job_interviews_completionRate']=createData['job_interviews_closeItem']/createData['job_interviews_outputItem']
                createData['creator_id'] = self.request.check_token
                JobInterviews.objects.create(**createData)
                self.return_data = {
                    "code": status.HTTP_200_OK,
                    "msg": "用户数据新增成功"
                }
            else:
                self.return_data = {
                    "code": status.HTTP_401_UNAUTHORIZED,
                    "msg": "数据关系不符合,用户数据新增失败"
                }


    def delete_data(self):  # 删除数据
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "删除成功！"
        }
        try:
            obj = Controller(JobInterviews, "delete", self.request)
            obj.start()
        except Exception as e:
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "删除失败！"
            }

    def patch_data(self):
        self.return_data = {
            "code": status.HTTP_200_OK,
            "msg": "修改成功！"
        }
        try:
            obj = Controller(JobInterviews, "patch", self.request)
            obj.start()
        except Exception as e:
            
            self.return_data = {
                "code": status.HTTP_401_UNAUTHORIZED,
                "msg": "修改失败！"
            }

    def download_file(self):
        
        
        id_list = self.request.data.get('idList')
        downloadAll = self.request.data.get('downloadAll')

        t = time.strftime('%Y-%m-%d')
        dummy_path = os.path.join(BASE_DIR, 'static', 'jobInterviewsFile', 'download_file', t)  # 创建文件夹
        self.mkdir(dummy_path)
        path = self.createExcelPath('在职访谈收集表.xlsx')
        

        if downloadAll == True:  # 是下载全部   有条件
            
            kwargs = {'job_interviews_status': True}
            colloquium_base = self.request.GET.get('baseNameId', None)  # 公司名称
            beginDate = self.request.GET.get('beginDate', None)
            endDate = self.request.GET.get('endDate', None)

            if colloquium_base == '' and beginDate == "" and endDate == "":  # 全查
                kwargs['job_interviews_status'] = True
                kwargs['job_interviews_base__in'] = self.request.user_base
            if colloquium_base != '':
                kwargs['job_interviews_base'] = colloquium_base
            if beginDate != "" and endDate != "":
                kwargs['job_interviews_date__gte'] = datetime(1901, 10, 29, 7, 17, 1,
                                                               177) if beginDate == None else beginDate
                kwargs['job_interviews_date__lte'] = datetime(2521, 10, 29, 7, 17, 1,
                                                               177) if endDate == None else endDate
            
            all = list(JobInterviews.objects.filter(**kwargs).all().values_list(
                    'job_interviews_base_id',
                    'job_interviews_base__name', 'job_interviews_date', 'job_interviews_number',
                    'job_interviews_percentage', 'job_interviews_outputItem',
                    'job_interviews_closeItem', 'job_interviews_completionRate','job_interviews_typical','job_interviews_remark'))

            row_data = []
            index = 1
            for i in all:
                i = list(i)
                if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',
                                                                                                       flat=True).exists():
                    # print(i[1],'1级基地')  润阳新能源
                    i.insert(0, i[1])
                    i[2] = ''
                else:
                    # print(i[2],'2级基地')
                    father_name = center_base.objects.filter(
                        id=
                        center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                            0],
                        status=True).values_list('name', flat=True)[0]
                    i.insert(0, father_name)
                if i[1] == None or i[3] == None:
                    pass
                else:
                    year_job_interviews_number = JobInterviews.objects.filter(job_interviews_status=True,
                                                                              job_interviews_base_id=i[1],
                                                                              job_interviews_date__gte=
                                                                              self.get_first_last(i[3])[2],
                                                                              job_interviews_date__lte=
                                                                              self.get_first_last(i[3])[3]).values_list(
                        'job_interviews_number', flat=True)

                    year_job_interviews_number = list(filter(None, year_job_interviews_number))
                    i.append(sum(year_job_interviews_number))
                i.insert(0, index)
                del i[2]
                if i[4] != None:
                    i[4] = self.round_down(i[4])
                if i[7] != None:
                    i[7] = self.round_down(i[7])
                row_data.append(list(i))
                if len(i) == 0:
                    index = index
                index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件
        else:
            
            row_data = []
            index = 1
            for id in id_list:
                data = JobInterviews.objects.filter(job_interviews_status=True, id=id).values_list(
                    'job_interviews_base_id',
                    'job_interviews_base__name', 'job_interviews_date', 'job_interviews_number',
                    'job_interviews_percentage', 'job_interviews_outputItem',
                    'job_interviews_closeItem', 'job_interviews_completionRate','job_interviews_typical','job_interviews_remark')
                for i in data:
                    i = list(i)
                    if center_base.objects.filter(status=True, base_parent_id=None, name=i[1]).values_list('name',flat=True).exists():
                        # print(i[1],'1级基地')  润阳新能源
                        i.insert(0, i[1])
                        i[2] = ''
                    else:
                        # print(i[2],'2级基地')
                        father_name = center_base.objects.filter(
                            id=
                            center_base.objects.filter(id=i[0], status=True).values_list('base_parent_id', flat=True)[
                                0],
                            status=True).values_list('name', flat=True)[0]
                        i.insert(0, father_name)
                    if i[1] == None or i[3] == None:
                        pass
                    else:
                        year_job_interviews_number = JobInterviews.objects.filter(job_interviews_status=True,
                                                                            job_interviews_base_id=i[1],
                                                                            job_interviews_date__gte=self.get_first_last(i[3])[2],
                                                                            job_interviews_date__lte=self.get_first_last(i[3])[3]).values_list('job_interviews_number', flat=True)

                        year_job_interviews_number = list(filter(None, year_job_interviews_number))
                        i.append(sum(year_job_interviews_number))
                    i.insert(0, index)
                    del i[2]
                    if i[4] != None:
                        i[4] = self.round_down(i[4])
                    if i[7] != None:
                        i[7] = self.round_down(i[7])
                    row_data.append(list(i))
                    if len(i) == 0:
                        index = index
                    index += 1
            exc = openpyxl.load_workbook(path)  # 打开整个excel文件
            sheet = exc.worksheets[0]  # 打开第0张工作表(也就是第一张)
            for row in row_data:
                sheet.append(row)  # 在工作表中添加一行
            exc.save(path)  # 指定路径,保存文件

        self.return_data = {
            "code": 200,
            "msg": "下载成功",
            "downloadUrl": path
        }
        # 

    #
    def create_file_name(self, objName, file_obj):  # 随机名  模块字段名(奖项名称 awards_name)  原文件
        t = time.strftime('%Y-%m-%d')
        fileName = str(objName)[:5] + "_" + t + "_" + str(
            "".join(list(str(time.time()))[0:10])) + "_" + str(file_obj)
        # 
        return fileName.split('.')[0]

    def createPath(self, pic, func):  # 生成路径
        t = time.strftime('%Y-%m-%d')
        pic_suffix = str(pic).split(".")[-1]
        pic_name = f"{func + 'nbpy'}.{pic_suffix}"
        pic_path = os.path.join('static', 'jobInterviewsFile', 'upload_file', t, pic_name)  # 图片路径

        if pic_suffix in ['jpg', 'jpeg', 'png', 'icon']:
            pic_path = os.path.join('static', 'jobInterviewsFile', 'upload_file', t, pic_name)  # 图片路径

        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            t = time.strftime('%Y-%m-%d')
            # pic_name = str(pic)
            pic_path = os.path.join('static', 'jobInterviewsFile', 'upload_file', t, pic_name)  # 文件路径

        pic_path = pic_path.replace('\\', '/')
        return (pic_path, pic_name, pic_suffix)  # 图片路径    图像名字

    def saveFile(self, pic_name, pic_obj, pic_suffix):  # 文件名,图像对象   文件保存
        t = time.strftime('%Y-%m-%d')
        if pic_suffix in ['jpg', 'jpeg', 'png']:
            with open('static/jobInterviewsFile/upload_file/' + t + '/' + pic_name, 'wb') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        elif pic_suffix in ['doc', 'docx', 'xls', 'xlsx', 'txt']:
            with open('static/jobInterviewsFile/upload_file/' + t + '/' + pic_name, 'wb+') as f:
                for dot in pic_obj.chunks():
                    f.write(dot)
        

    def createExcelPath(selfself, file_name):
        import openpyxl
        from openpyxl.styles import Alignment
        exc = openpyxl.Workbook()  # 创建一个excel文档,W是大写
        sheet = exc.active  # 取第0个工作表
        for i in range(97, 109):
            sheet.column_dimensions[chr(i).upper()].width = 15
            if i == 97:
                sheet.column_dimensions[chr(i).upper()].width = 10
            if i == 98:
                sheet.column_dimensions[chr(i).upper()].width = 25
            if i >= 104:
                sheet.column_dimensions[chr(i).upper()].width = 18

        sheet.title = file_name.split('.xlsx')[0]  # 工作表命名为表名
        sheet.merge_cells('A1:L1')  # 从C2到D3合并为一个单元格,但此后名为A1
        sheet['A1'] = '在职访谈收集表'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        dataRows = (('序号', '中心/基地', '公司','日期', '访谈人次', '占比','产出行动项', '关闭项', '完成率','典型事项','备注','年度累计人次'),)  # 可以是列表和元组
        for row in dataRows:
            sheet.append(row)  # 在工作表中添加一行
        t = time.strftime('%Y-%m-%d')
        path = os.path.join('static', 'jobInterviewsFile', 'download_file', t, file_name)  # 文件路径
        path = path.replace('\\', '/')
        exc.save(path)  # 指定路径,保存文件
        return path

    def mkdir(self, path):
        # os.path.exists 函数判断文件夹是否存在
        folder = os.path.exists(path)

        # 判断是否存在文件夹如果不存在则创建为文件夹
        if not folder:
            # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
            os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
            
        else:
            pass
            

    def get_first_last(self, date):
        import calendar
        import datetime
        year, month = int(date.year), int(date.month)
        weekDay, monthCountDay = calendar.monthrange(year, month)
        # 
        # range_day = str(datetime.date(year, month, day=1)) + "至" + str(datetime.date(year, month, day=monthCountDay))
        first_month = datetime.date(year, month, day=1)
        last_month = datetime.date(year, month, day=monthCountDay)
        first_year = datetime.date(year, 1, 1)
        last_year = datetime.date(year, 12, 31)
        return first_month, last_month, first_year, last_year

    def round_down(self, num):  # 小数向下取整2位
        import math
        return math.floor(num * 100) / 100

    def decision_num(self, num,typp):
        if num!=None:
            try:
                if type(num) == str:
                    if eval(num) >= 0:
                        return eval(num)
                    else:
                        return
                elif type(num) ==int:
                    if num>=0:
                        return num
                    else:
                        return
                elif type(num) ==typp:
                    if num>=0:
                        return num
                    else:
                        return
            except:
                return
        else:
            return

        # try:
        #     if type(createData['colloquium_numberSessions']) == str:
        #         createData['colloquium_numberSessions'] = eval(createData['colloquium_numberSessions'])
        #     elif type(createData['colloquium_numberSessions']) == int:
        #         createData['colloquium_numberSessions'] = createData['colloquium_numberSessions']
        # except:
        #     createData['colloquium_numberSessions'] = None


        # if sheet.cell(i + 1, 5).value != None:  # 挽留成功人次
        #     try:
        #         if type(sheet.cell(i + 1, 5).value) == int and sheet.cell(i + 1, 5).value >= 0:
        #             exit_interviews_retentionSuccess = sheet.cell(i + 1, 5).value
        #         else:
        #             exit_interviews_retentionSuccess = None
        #     except:  # 可能不是日期
        #         exit_interviews_retentionSuccess = None
        # else:
        #     exit_interviews_retentionSuccess = None