

from openpyxl import load_workbook

wb = load_workbook('D:\Runergy_SourceCode\Python\pdss\\testApp\时间序列.xlsx')
ws = wb.active
from datetime import datetime, time, date

for i in range(2, ws.max_row + 1):
    time_str = ws.cell(row=i, column=2).value
    date_obj = ws.cell(row=i, column=3).value


    time_obj = datetime.strptime(time_str, "%H:%M:%S").time()
    datetime_obj = datetime.combine(date_obj, time_obj)

    ws.cell(row=i, column=5, value=datetime_obj)
wb.save('D:\Runergy_SourceCode\Python\pdss\\testApp\时间序列.xlsx')

"""
# 导入readxl包
library(readxl)
# 读取Excel文件
excel_data <- read_excel("your_excel_file.xlsx") # 替换为你的文件路径

# 创建DataFrame
df <- data.frame(Eta = 时间序列$Eta, TestTime = 时间序列$TestTime)

# 导入所需的包
library(ggplot2)
# library(dplyr)

# 创建时间回归模型
model <- lm(Eta ~ TestTime, data = df)

# 绘制时间回归分析的散点图和拟合线
ggplot(df, aes(x = TestTime, y = Eta)) +
  geom_point() +
  geom_smooth(method = "lm", formula = y ~ x, se = FALSE) +
  labs(title = "时间回归分析", x = "时间", y = "Eta") +
  theme_minimal()

# 打印回归分析摘要
summary(model)




lm(formula = Eta ~ TestTime, data = df)

Residuals:
     Min       1Q   Median       3Q      Max 
-22.5752  -0.0635   0.0533   0.1319   0.3469 

Coefficients:
              Estimate Std. Error t value Pr(>|t|)
(Intercept)  7.084e+01  3.784e+02   0.187    0.852
TestTime    -2.807e-08  2.233e-07  -0.126    0.900

Residual standard error: 0.4273 on 24052 degrees of freedom
Multiple R-squared:  6.572e-07,	Adjusted R-squared:  -4.092e-05 
F-statistic: 0.01581 on 1 and 24052 DF,  p-value: 0.9




这是线性回归模型的摘要信息，它提供了有关回归分析的各种统计指标和参数的信息。

Residuals (残差):

Min: 残差的最小值
1Q: 第一四分位数（Q1），表示残差的25th百分位数
Median: 残差的中位数，即第二四分位数（Q2）
3Q: 第三四分位数（Q3），表示残差的75th百分位数
Max: 残差的最大值
这些统计量描述了模型中预测值和观测值之间的差异。

Coefficients (系数):

Intercept (截距): 模型的截距项的估计值
TestTime: 自变量 TestTime 的估计系数
Estimate: 估计值，表示截距和 TestTime 系数的估计值
Std. Error: 估计值的标准误差
t value: t统计量，用于检验系数的显著性
Pr(>|t|): p-值，表示 t统计量的显著性水平
截距和 TestTime 系数的估计值、标准误差、t值和p值用于判断它们对因变量（Eta）的影响是否显著。

Residual standard error (残差标准差): 衡量了残差的标准差，表示模型对数据的拟合程度。

Multiple R-squared (多重R方): 衡量了模型中自变量的解释方差的程度。在这种情况下，R方非常接近零，说明模型中的自变量无法解释因变量的方差。

Adjusted R-squared (调整后的R方): 考虑了自变量的数量和模型的复杂性，以更准确地反映模型的解释能力。也非常接近零，表示模型的解释能力较差。

F-statistic (F统计量): 用于检验整个模型是否显著，即自变量是否总体上对因变量有显著影响。

在这种情况下，F统计量为0.01581，p值为0.9，说明模型的整体显著性不足。
综合来看，根据这个回归模型的结果，截距和 TestTime 系数对因变量 Eta 的影响都不显著，而且模型的解释能力非常差，因为多重R方和调整后的R方都接近零。这可能意味着时间 TestTime 对 Eta 的影响非常微弱，或者模型可能需要更多的自变量来解释因变量的变化。



['5月', '6月', '7月', '8月', '8月', '10月']
[11829, 21929, 146818, 227985, 233320, 229170]

"""